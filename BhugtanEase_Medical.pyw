# -*- coding: utf-8 -*-
"""
BHUGTANEASE MEDICAL BILLING SOFTWARE — MEDICAL STORE VERSION
- Login protected
- Sale Bill, Purchase
- Sale History, Purchase History
- All Bills (Sale + Purchase tabs)
- Stock, Products, Parties, Ledger
- P&L Report, Sale Report
- Exact invoice format
- Back date bill support
- Scrollable tables with mousewheel
"""

import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
import datetime
import os
import sys
import hashlib

# ── Grocery Add-ons: Barcode + Expiry ──────────────────────────────────────
try:
    import barcode
    from barcode.writer import ImageWriter
    BARCODE_OK = True
except ImportError:
    BARCODE_OK = False

try:
    from PIL import Image, ImageTk, ImageDraw, ImageFont
    PIL_OK = True
except ImportError:
    PIL_OK = False

import io, base64, re, json

# ─── DATABASE PATH — Pehle define karo ───────────────────────────────────────
# PyInstaller onefile build me __file__ ek temp extraction folder ko point
# karta hai jo har restart pe delete ho jata hai -> data loss. Isliye frozen
# build me sys.executable (actual .exe ki location) use karo.
if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(BASE_DIR, "bhugtanease.db")

# ─── LICENSE SYSTEM v2 ────────────────────────────────────────────────────────
# ⚠️  SECRET KEY — Sirf aap jaante hain — kabhi share mat karo!
# ─── LICENSE ENGINE v3 — Self-contained serial (no JSON file needed) ─────────
# Secret key — kabhi share mat karo
_LIC_SECRET  = "BhugtanEase@2025#VNS"
_LIC_BASE_DT = datetime.date(2020, 1, 1)   # epoch for date encoding

# Serial format: BE-DDDD-HHHH-HHHH-HHHH
#   BE   = prefix
#   DDDD = 4-hex expiry offset from 2020-01-01 (up to year ~2199)
#   HHHH = SHA-256 hash chunks (first 12 hex chars)
# Customer sirf serial type kare — koi naam ya date nahi chahiye

def _make_serial_v2(expiry_date: datetime.date, customer_name: str = "") -> str:
    """Customer naam + expiry date se serial banao. BE-XXXX-XXXX-XXXX-XXXX format.
    Customer naam hash mein shamil — alag naam = alag serial!
    """
    days = (expiry_date - _LIC_BASE_DT).days
    date_block = f"{days:04X}"
    cname = customer_name.strip().upper()
    raw = f"{_LIC_SECRET}|{expiry_date.isoformat()}|{cname}"
    h = hashlib.sha256(raw.encode()).hexdigest().upper()
    return f"BE-{date_block}-{h[0:4]}-{h[4:8]}-{h[8:12]}"

def _decode_serial(serial: str, customer_name: str = ""):
    """
    Serial validate karo — customer naam + expiry date se verify hoga.
    Returns: datetime.date agar valid, else None
    """
    serial = serial.strip().upper().replace(" ", "")
    parts = serial.split("-")
    if len(parts) != 5 or parts[0] != "BE" or len(parts[1]) != 4:
        return None
    try:
        days   = int(parts[1], 16)
        expiry = _LIC_BASE_DT + datetime.timedelta(days=days)
    except Exception:
        return None
    # Verify hash with customer name
    cname = customer_name.strip().upper()
    raw = f"{_LIC_SECRET}|{expiry.isoformat()}|{cname}"
    h   = hashlib.sha256(raw.encode()).hexdigest().upper()
    expected = f"BE-{days:04X}-{h[0:4]}-{h[4:8]}-{h[8:12]}"
    return expiry if expected == serial else None

def _get_license_info():
    """
    Returns: (status, days_left, install_date_str, customer_name)
    status: 'ok' | 'expired' | 'new'
    """
    try:
        conn = sqlite3.connect(DB_FILE)
        rows = {r[0]: r[1] for r in conn.execute(
            "SELECT key,value FROM settings WHERE key IN "
            "('install_date','license_expiry','license_customer','license_serial')"
        ).fetchall()}
        conn.close()

        install_date_str = rows.get('install_date', '').strip()
        expiry_str       = rows.get('license_expiry', '').strip()
        customer         = rows.get('license_customer', '').strip()

        if not install_date_str or not expiry_str:
            return ('new', 0, '', customer)

        expiry_date = datetime.date.fromisoformat(expiry_str)
        today       = datetime.date.today()
        days_left   = (expiry_date - today).days

        if today > expiry_date:
            return ('expired', days_left, install_date_str, customer)
        return ('ok', days_left, install_date_str, customer)
    except Exception:
        return ('new', 0, '', '')

def _activate_license(serial: str, customer_name: str = "") -> bool:
    """
    Serial validate karo customer naam ke saath. Agar valid ho to DB mein save karo.
    Serial + customer naam dono chahiye — naam galat ho to serial nahi chalega.
    """
    expiry_date = _decode_serial(serial, customer_name)
    if expiry_date is None:
        return False
    if expiry_date <= datetime.date.today():
        return False   # already expired serial
    try:
        today = datetime.date.today()
        conn  = sqlite3.connect(DB_FILE)
        for k, v in [
            ('install_date',     today.isoformat()),
            ('license_expiry',   expiry_date.isoformat()),
            ('license_customer', customer_name.strip()),
            ('license_serial',   serial.strip().upper()),
        ]:
            conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", (k, v))
        conn.commit()
        conn.close()
        return True
    except Exception:
        return False

def _record_first_install():
    """Deprecated — kept for compatibility."""
    pass

def _get_logo_image(size=(60,60)):
    """Logo PIL Image banao — cached bytes use karta hai, baar-baar decode nahi."""
    try:
        from PIL import Image, ImageTk
        import io
        img = Image.open(io.BytesIO(_LOGO_BYTES)).convert("RGBA").resize(size, Image.LANCZOS)
        return ImageTk.PhotoImage(img)
    except Exception:
        return None

def _apply_logo(win):
    """Window pe logo icon set karo (taskbar + titlebar).
    FAST: disk write nahi hota — PIL se seedha memory mein load.
    """
    try:
        from PIL import Image, ImageTk
        import io
        img = Image.open(io.BytesIO(_LOGO_BYTES))
        ico = ImageTk.PhotoImage(img)
        win.iconphoto(True, ico)
        win._logo_ref = ico          # garbage collection se bachao
    except Exception:
        pass



def _show_first_install_window():
    root = tk.Tk()
    root.title("BhugtanEase — Activate")
    root.resizable(False, False)
    root.configure(bg="#F0F0F0")
    root.protocol("WM_DELETE_WINDOW", root.destroy)

    W, H = 560, 480
    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()
    root.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")
    _apply_logo(root)

    # ── MAROON HEADER ──────────────────────────────────────────────────────
    hdr = tk.Frame(root, bg="#8B1A1A", height=82)
    hdr.pack(fill="x")
    hdr.pack_propagate(False)

    logo_img = _get_logo_image((224, 58))
    if logo_img:
        lbl = tk.Label(hdr, image=logo_img, bg="#8B1A1A")
        lbl.image = logo_img
        lbl.place(relx=0.5, rely=0.5, anchor="center")
    else:
        tk.Label(hdr,
                 text="BhugtanEase",
                 font=("Segoe UI", 22, "bold"),
                 bg="#8B1A1A", fg="#FFFFFF").place(relx=0.5, rely=0.5, anchor="center")

    # ── DARK SUB-BAR ───────────────────────────────────────────────────────
    sub = tk.Frame(root, bg="#5C1010", height=28)
    sub.pack(fill="x")
    sub.pack_propagate(False)
    tk.Label(sub,
             text="Medical Store Billing & Inventory",
             font=("Segoe UI",9),
             bg="#5C1010", fg="#E8C0C0").place(relx=0.5, rely=0.5, anchor="center")

    # ── WHITE CARD ─────────────────────────────────────────────────────────
    gap = tk.Frame(root, bg="#F0F0F0", height=20); gap.pack(fill="x")

    card_outer = tk.Frame(root, bg="#CCCCCC", padx=1, pady=1)
    card_outer.pack(fill="x", padx=24)

    card = tk.Frame(card_outer, bg="#FFFFFF", padx=30, pady=26)
    card.pack(fill="both")

    # Heading
    tk.Label(card,
             text="Serial Number Enter Karein",
             font=("Segoe UI", 13, "bold"),
             bg="#FFFFFF", fg="#111111",
             anchor="w").pack(fill="x")

    # Format hint
    tk.Label(card,
             text="Format: BE-XXXX-XXXX-XXXX-XXXX",
             font=("Segoe UI", 9),
             bg="#FFFFFF", fg="#666666",
             anchor="w").pack(fill="x", pady=(3, 14))

    # ── Customer Name Field ─────────────────────────────────────────────
    tk.Label(card,
             text="Aapka Naam (Shop/Owner ka naam — serial ke saath milna chahiye):",
             font=("Segoe UI", 9),
             bg="#FFFFFF", fg="#333333",
             anchor="w").pack(fill="x", pady=(0, 3))

    name_border = tk.Frame(card, bg="#AAAAAA", padx=1, pady=1)
    name_border.pack(fill="x")
    v_customer = tk.StringVar()
    e_name = tk.Entry(name_border,
                      textvariable=v_customer,
                      font=("Segoe UI", 11),
                      bg="#FFFFFF", fg="#111111",
                      insertbackground="#8B1A1A",
                      relief="flat", bd=7)
    e_name.pack(fill="x")
    e_name.focus_set()

    tk.Label(card,
             text="(Bilkul wahi naam likhen jo dealer ko diya tha — chhote bade akshar matter nahi karte)",
             font=("Segoe UI",7),
             bg="#FFFFFF", fg="#AAAAAA",
             anchor="w").pack(fill="x", pady=(3, 12))

    # Input border frame
    inp_border = tk.Frame(card, bg="#AAAAAA", padx=1, pady=1)
    inp_border.pack(fill="x")

    v_serial = tk.StringVar()
    e = tk.Entry(inp_border,
                 textvariable=v_serial,
                 font=("Segoe UI", 12),
                 bg="#FFFFFF", fg="#111111",
                 insertbackground="#8B1A1A",
                 relief="flat", bd=7)
    e.pack(fill="x")

    # Hint below input
    tk.Label(card,
             text="(Poora serial yahan paste karein)",
             font=("Segoe UI",7),
             bg="#FFFFFF", fg="#AAAAAA",
             anchor="w").pack(fill="x", pady=(5, 0))

    # Error label
    msg_var = tk.StringVar()
    tk.Label(card,
             textvariable=msg_var,
             font=("Segoe UI", 9, "bold"),
             bg="#FFFFFF", fg="#C0392B",
             wraplength=480,
             anchor="w").pack(fill="x", pady=(6, 0))

    # Spacer
    tk.Frame(card, bg="#FFFFFF", height=12).pack()

    # ACTIVATE button
    def activate(ev=None):
        customer = v_customer.get().strip()
        serial = v_serial.get().strip().upper()
        if not customer:
            msg_var.set("Pehle apna naam daalna zaroori hai!"); return
        if not serial:
            msg_var.set("Serial number daalna zaroori hai!"); return
        if _activate_license(serial, customer):
            root.destroy()
        else:
            msg_var.set("Serial galat ya expire ho chuka hai! Naam aur serial dono dealer se milana chahiye.")

    e_name.bind("<Return>", lambda ev: e.focus_set())
    e.bind("<Return>", activate)

    btn = tk.Button(card,
                    text="✔  ACTIVATE KARO",
                    font=("Segoe UI", 12, "bold"),
                    bg="#C0392B", fg="#FFFFFF",
                    activebackground="#922B21",
                    activeforeground="#FFFFFF",
                    relief="flat", bd=0,
                    cursor="hand2",
                    pady=14,
                    command=activate)
    btn.pack(fill="x")

    root.mainloop()

# LICENSE EXPIRED SCREEN
class LicenseExpiredWin:
    """Jab license expire ho jaye — taskbar mein tab NAHI dikhega."""
    def __init__(self, days_left, customer):
        self._root = tk.Tk()
        self._root.title("BhugtanEase — License Expired")
        self._root.configure(bg="#1A1A2E")
        self._root.resizable(False, False)
        w, h = 600, 640
        sw, sh = self._root.winfo_screenwidth(), self._root.winfo_screenheight()
        self._root.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        self._root.protocol("WM_DELETE_WINDOW", self._root.destroy)
        self.dlg = self._root   # sabhi references self.dlg use karte hain
        _apply_logo(self._root)

        # Logo + Header
        top = tk.Frame(self.dlg, bg="#1A1A2E"); top.pack(pady=(16, 0))
        logo_img = _get_logo_image((232, 56))
        if logo_img:
            lbl_logo = tk.Label(top, image=logo_img, bg="#1A1A2E")
            lbl_logo.image = logo_img
            lbl_logo.pack(anchor="center")

        tk.Frame(self.dlg, bg="#C53030", height=3).pack(fill="x", pady=(10, 0))
        tk.Label(self.dlg, text="\u23f0  License Expire Ho Gayi!",
                 font=("Segoe UI", 15, "bold"), bg="#1A1A2E", fg="#FC8181").pack(pady=(10, 2))
        tk.Label(self.dlg,
                 text=f"Aapki license {abs(days_left)} din pehle expire hui.",
                 font=("Segoe UI",9), bg="#1A1A2E", fg="#A0AEC0").pack()
        tk.Label(self.dlg,
                 text="Software jaari rakhne ke liye naya Serial Number daalo.",
                 font=("Segoe UI",9), bg="#1A1A2E", fg="#A0AEC0").pack(pady=(2, 10))

        card = tk.Frame(self.dlg, bg="#2D3748", padx=24, pady=14)
        card.pack(fill="x", padx=28)

        tk.Label(card, text="👤 Aapka Naam (Serial ke saath register hua naam):",
                 font=("Segoe UI", 9), bg="#2D3748", fg="#A0AEC0", anchor="w").pack(fill="x", pady=(8, 2))
        self._customer_var = tk.StringVar()
        name_entry = tk.Entry(card, textvariable=self._customer_var,
                              font=("Segoe UI", 11),
                              bg="#1A202C", fg="#FBD38D",
                              insertbackground="#FBD38D", relief="flat")
        name_entry.pack(fill="x", ipady=6)
        name_entry.focus_set()

        tk.Label(card, text="🔑 Naya Serial Number (BE-XXXX-XXXX-XXXX-XXXX):",
                 font=("Segoe UI", 9), bg="#2D3748", fg="#A0AEC0", anchor="w").pack(fill="x", pady=(10, 2))
        self._serial_var = tk.StringVar()
        serial_entry = tk.Entry(card, textvariable=self._serial_var,
                                font=("Courier New", 13, "bold"),
                                bg="#1A202C", fg="#68D391",
                                insertbackground="#68D391", relief="flat")
        serial_entry.pack(fill="x", ipady=2)

        self._msg_var = tk.StringVar()
        tk.Label(self.dlg, textvariable=self._msg_var,
                 font=("Segoe UI", 9, "bold"), bg="#1A1A2E",
                 fg="#FC8181", wraplength=510).pack(pady=(10, 4))

        name_entry.bind("<Return>", lambda e: serial_entry.focus_set())
        serial_entry.bind("<Return>", lambda e: self._do_activate())
        tk.Button(self.dlg, text="🔓  Activate Karo",
                  font=("Segoe UI", 11, "bold"), bg="#276749", fg="white",
                  relief="flat", cursor="hand2", padx=20, pady=9, bd=0,
                  command=self._do_activate).pack(pady=(4, 0))
        tk.Label(self.dlg, text="Naya serial lene ke liye apne dealer se sampark karein.",
                 font=("Segoe UI",7), bg="#1A1A2E", fg="#4A5568").pack(pady=(8, 0))
        tk.Frame(self.dlg, bg="#C53030", height=3).pack(fill="x", side="bottom")

        self._root.mainloop()

    def _do_activate(self):
        customer = self._customer_var.get().strip()
        serial = self._serial_var.get().strip().upper()
        if not customer:
            self._msg_var.set("❌ Pehle apna naam daalna zaroori hai!"); return
        if not serial:
            self._msg_var.set("❌ Serial number daalna zaroori hai!"); return
        if _activate_license(serial, customer):
            self.dlg.destroy(); self._root.destroy()
        else:
            self._msg_var.set("❌ Serial ya naam galat hai! Dealer se milaya hua naam aur serial dono sahi honay chahiye.")

# ─── DATABASE ────────────────────────────────────────────────────────────────
LOGO_B64 = """iVBORw0KGgoAAAANSUhEUgAAARgAAABYCAIAAAChw38DAAABCGlDQ1BJQ0MgUHJvZmlsZQAAeJxjYGA8wQAELAYMDLl5JUVB7k4KEZFRCuwPGBiBEAwSk4sLGHADoKpv1yBqL+viUYcLcKakFicD6Q9ArFIEtBxopAiQLZIOYWuA2EkQtg2IXV5SUAJkB4DYRSFBzkB2CpCtkY7ETkJiJxcUgdT3ANk2uTmlyQh3M/Ck5oUGA2kOIJZhKGYIYnBncAL5H6IkfxEDg8VXBgbmCQixpJkMDNtbGRgkbiHEVBYwMPC3MDBsO48QQ4RJQWJRIliIBYiZ0tIYGD4tZ2DgjWRgEL7AwMAVDQsIHG5TALvNnSEfCNMZchhSgSKeDHkMyQx6QJYRgwGDIYMZAKbWPz9HbOBQAAAwy0lEQVR42u19d3xUVdr/85xzy7TMpIdUIAm9Sq8KKLCAqKusq676UyyoWBB7eW27VkRWUFDX7tpYRQVBpHeQFnoPSUgndZJpt5xzfn/cZAgBdS2vu+/u/X4Qw8ydO/ee+3yf8j3POUEhBNiwYeOXgdhDYMOGTSQbNmwi2bBhE8mGDRs2kWzYsIlkw4ZNJBs2bCLZsGHDJpINGzaRbNiwiWTDhk0kGzZs2ESyYcMmkg0bNpFs2LCJZMOGDZtINmzYRLJhwyaSDRs2kWzYsGETyYaNfwWkn/oBIcT/0sZDiNjifzZs/F8C2ttx2bDxL4hIBw8e8Pv9lNKzBisrogghoCmuWD8iIgIIgYCAhBBE5Jxzk3POGOemYQrOPC5XrNfr8/o8Pp/T7QY7Mtn4DybSizNeWLdundvj4VwAiGgu1pTxISBBASC4AAQAIBaNBDDOOCIzTcMwOOeqLLudLqfqzMrIzMnNTUtPzWnbPs7r9aYky5JsnfCXZ3m/1klOSz5t2PglRLLCDGM8HI5IsoMxBoAATZFDNMUbMIEjIkVEQM7MiB7hJpMIKpLikJSstIz0jLTMdlnt27dv3yG3XW4Hb2y8z+W0voLrul5SWldaLKW08eV2+Nk8EAAIEAwEJFlWVfXnkScaYKOX8KvQ0oYdkQBAgECJKpRSi0VRmyOIFqUURJOZuh4BELFeX6fsdmnpadm5HXp07dY2JzchuU18cpLSfDpeU2ueKKqpKNNKiuFoPi85gbW1oq7BSM1SnnrM0SEXuACCP5UEiOivqdENLTEltUVE+UGR44wXo68HAv7GQCAxIVGWVdtibPxyIln2KLgQnJtCsGb1XLT4D4Qp4uPihg4b0q9f306dO2dltUtKSmo6ga7p1dWRXXn1RYVacSErLHCUVshVtWaDnwpNNUwHEEEJVVTt4P662a8mPPmIGp8IpyqufzZuNtbV1lSUZXXq/M9kZJzzVswBgFAodODA3i1bNh88dKC6urqmpmbgwEH3TL8/Pj7Bjks/nk5/j8s6zZAQQTT/8F9FpKZaCAXg9/p0gshNJlHpqj9d3a9fPwAIVZRXrlpuHD3GysrkygosLcMav2xG3FxHzlQkinBJAIwilV2cIXITNQMc1L9tY+Xb72ROm4ZEAvjnuCQEIIbq60t272zbt78kydbLjDHOOWNMCC6E4JwDIICglCqKGhVOOOeEEOvvrVu33HTz5IYGv6KoBCWJqkcOf+T1+h64/+F/dyP+V9slIopmp9uSQOLUzwgA3HJepzPsv0hsaAY5fayavbvgRBXllSf+8vTjr7/2Vnxs/MF33vItWZoY1MAME0miVFI5EgSBMpEkEAJNAYITIBwRhQRcCGpSU4uRldqFX1V3yEm6+FLBOeKPDbgQgKgHgsdXLEvt29cV42tsqH/3jTeKykoMJjRN03XNZCbnnHNmHa4ois8bl56e1rdvv8GDh7rdrqgh5uTkXD95cllpWd7OXVVVtYri5MFA1cmaf/PH+a9kkeXFKipO/mOBGozIEuVgCMElQI6EI0iCUyY4GIxS4JwIIkA2UxITrrxcVp3Nhe1/L5FaeR8QwBwOdVfeztlzXnrqqWdzJk4MbtvhChRSh4KcAEckCByRCwAOQoDQBQgBFJAKAigICiRMOImZYATr3n67sVOXmM5dQPAfDEoCAJiuH/1iflzb7IT2OQBgGGzD5s3bd+4wBQASAoAguOCGYZimaeVykiQjosvl6tWz19SpU889b4RVUWVmtn34wcfC4dDNN99UWlqhKA7OmSTRf5PHZqWjhJzWmMIYM01dUVRE8q/gkUBEIxio+W5TYn6pK9KAZlgBREI0wcOcA1c5CkojgFzIqmLKIkzCHbPYReNl1fl/nEdAn3jiiZ/0gS+//KKwsEhR1FYlvKU1EIIEKUVVkV279+5La5PSd+QFEZfbn7dDNXWZCxCMIxcIBAgAAhGCMCACkQAQIdDK4gQIKpgKBOv99fW1rmHDqaz8QD4thEAkRQvmG3q4/cRLrKxBdThGjBp1ye8vO3T4UHV1tcvhlojkdsd06dL5nHP6jBgxqnfvPlVVJxljBKUTRSUrVi5Pz0zr3KlLVLITAr7+elFhYaGiqKZpdOvedfQFY35Lx3+m8m7Z65laYn193UMP3T/v9VcTExNyczpyzn/0IqOy5K8YDCWPO/bcYd4BA/DoQae/VlGclAhIiDd69hYp6SwpDnxORXFKKCRJooCYnKhccpHkcP5fL5Z+XkRCAILAueCENI0g44wxwzS54FRwjRJZkaW5f53brVPXrhdOKD60LzR/gSwJEEwwjggCCBACwIEgF1ziDIAiASYAACSUgBMAEkNR37C54etv4i+fBIIhnG0imHMk5OSyJTV7dnSb/hAhFAQHQIIkOTklOTnF5/XqWsShqGFNa5fT/uWXX27fPsf67MiRIx586IEGf4Mnxh0IBma9NLNP7z4ZmVmWdXLexCjSrPG3tMJoWGhli80kFM1GD9EQ0eqtM6NKSyG05butXsnL21lTUz1ixEhJkq1T7dixbfGSRcFgcM/uPWNGj28p35/tzFaMJt8X3KwrbXmSlgeIZrR6HQAkWfWkpDGfL+BzMiJUpDozRHaHuD8/LRwOEQyKcKMUNuo/f58tW+ymDuJ2UFVulVqg9d1N/CSAIDi3Buv7PEKTwoGI0TEGbCUgNanOgqMQAhEQsGnykyDB35pI1qUigECQKDEMzTA1zrnX601MSM7O7tC+XW679u0OHzj0xaeflBYXvTjj2dnzXku67sbqI8fVHbtUKkmEAwPBGVIUgMCRIgJwBF0ICZEAIBfEGgOCxKPpjZ9+HBnQ39Gu7VnUcM6BkMYdW8s++XvO1LuccQkgOCBp0kUEcM4FNwkRiFyAoShSbGyclQgh4gUXjL5o04Z33nmbUOpwOE4UFu3ZvdsiUrOURwCACwEIQvCWjjyqUrQq8a2fW74SPeAH3mqZtkXJyZhpmiYhVJZl65W6utpZs2Z+9vn85OSkzp07Z2Rkcc4RybJlywQHVXVUV1e3vLaWpz0jlHHOhXXkGZdh8b/1dTZNdZzuO07/rAABPKwDMATTamXRFdkd65NUB/h8TQfl9+eLlxFOWWIioUrLq7S+WCBaRLLKWUIIwveUUVwgaZp64QAASBDOrmFY3Iz6ptN0RPEL46H08wISEh4OhIRgPl9M/579+vbt27dP365de8TGxsqybIUJAP2jTz9ZtXXD3HlzHrj/Uc8tN9c/8nh8TTURnIAAIumcUVkiAoALACKEQOAEAQRlQAQwiii4kCSqFuYH53+i3HcfPT37F0IgIVrR8cLZsxJGjIjrOzDKoqg5EGKVc1wABwQOnDFm+WArqvTr1++DDz4QHJACIPgb/Gc6ckTgjCE2aXoAcKKoaOOmjQg4aNCgdu3btzzYNI3j+cePHDlSUVnhb/APHTJk0KAhUe9eXl524MDB8vKyqqqq/v37nXvuiJaGaJlpKBTavHnDzrydxSeKw+GQoqrJyW169ew9aNDAufPmvPPOOzExPr8/cKL4REZGliRJx4/nb9iwSVacZjhYVVUdCDRIkiIEJ0hkRbEu2Po7FAoWFB4/fPhIYUFBWVmZrhvZ2e1/97txnTt3Oe0yOCspLT1y5EhxSXFtTU3nzl3Gj58QZXhtbc3GjRurq6v79evbo0cvRDzNDhGRIgGgnMkoOGcRZM31sLDSBzktK5SYWlPnl7Lbo6ycUmUJEQA8GDBOlPHjxwKV5Xp1naLpJCnOOWqkq2NnhFN8suIKEDTqakO7d0WKi3mjHwwTHU41KUXu2sWZk0tluSVVBED4WH5wz25+sopFItTjcbbNdvXvI8X6fiGXfjKRuBAIyDnr3bvXoEEDR40a1b17D5fL1SrnoZQ++Mhj/mDks4WLPvr7J/179B01bpz+xz/Uv/56vKAq6hy4LskiElYEKFTmRCacCW6AyQE5QSIIARCIlAA40GQrv9FHjXb063PqoQkOSMz62vJZM6nH3eayK1rNYGDTgCMIAIEARIAMAq1cwEppKKX+Bj/nXJKIECjLcmpqWoviRJzupAkhpKKi7KMPP1y8eHFR0QnBeVpG2jXXXHv99TdQSgkh27ZtfeH5Z4qLS/11DcFw2GT6Z+2ynnvuhfPOHaHr2qOPPrx27dpwWGtsaNT1SJvUlL88/ez4cROi/h4Rt2/f/vLLf92+fVtDg18AUEJNk1GJul2uzKzMyopyh9OtaWZNbU1+/rEhg4dt2LDhb397vaqmGglVVVde3q4rr7xSkiTDMFRFeXHGzPY5OcFgYPeevHXr1m7durWgoLCysjIS1t1ut0JVwzT+8ek/7rnn3ksnXWZdQFlpyTPP/Hn7jrxAMBRobNANPdYX29AQuOKKPzY2Nn711ZfzP51/+PBhTQunJCddeNHEO++82+eLbclDK0AgpwBIhaDATiXFlAKAq3NXeOYvoBtq23bReGCGI42H9rIt26TtO6TjhUZtDdPDHtkpcW4S079kcfjGm+LHTyCUQlN4Ai5E/fJl5oefKkcPQkM9CoMKCSXFcFF/bKLnmskJf7oSm2erjPq6kx9/iEu+dZSU8JCfqJIquxCcDT37SLde6x0w8JdwSfoZ0QhQIJJBgwbed9/9VmJgpUnRvIVSKoRwu92PPvJoTWXdprXrXp45q2P37mlX/PHEgX3B1SsdiExBacgQ0+MNHjnYcKJIDppuArJE0eBcEAISMCaIQBAMgSgUK8vr1qxM7denRbGBwjBOvjY3vG9PmxdelL0+wbkgZ2a7AgBBEAACAhCJLCvR5KequuqLL74QICilDQ0NQ4cO7NOnT4uIhE1TzUJQSkzTXPz1opmzXjiwb5/TGeNyugFIWVn5s8897XK5rr76WgBoaPQXF58IhSKAEBMTg4SUlVasW7v2vHNHIEIwHNQ0jXPu8cQQ4qupqV2xYtn4cROiGd369evvu+/+iopKWZJ79ep7+eWXu93OtevWrlq5WghRVFgMAAkJieedd25KSkqH3E6vvz73pVkv6RpTVIflNEKh0L59+xAgEolkZ+dIsgwAS5cufvyJJxr8jVSiKcltxowe53K5tm/f3uBvdLldFZUVTz71eJvU5CFDhwOAaZqFRUWhUIQzcDjcXm9cbW3tgs8/z85u/+KM5zds2ECp7HF73a7Y+vrGN15/XYB49JHHT5MKOQgGhAMKIBwpx6geZQ2rHBPj6923+dimga7duqXh6afiqytVk0FKanjkUIfkkHftVBtOCkmVSkrqX54XSkjxDB0EXACCQKz6ckHkr7NT6hqFosjnn4/tcoySMnXbTqLX0qpSfvyY9fCQEL2qqu7Z55V1q1Ru0IQk+bxBkaqTRkGh22TSzu9qnywRjz7oGzz4Z3PpJ+mkVm0gAAQH8c5771199dWrV6+2/HpUTYp6b855cnLSE0892qVH57z9+1+Z9RJT1LibbwpktNUMYAYL5B+mQwbHvjpP+fMzxpiR9T53IzMMFFxWdJQEp4ID40xwqDdMo0evuLFjMZqGC4GIdf+Y3/jFV/EXXhzbuz9wga1YFE1UsIkNBE3T0MrLyysrK44XFiz9dukdd96xfWeeojoC4UCHTjn33nuf9zTnKqK8lWV11apVDz/8gFNx/GHSH3Nzsg1mcOBOp0dwsmTJYsMwAGDY0OHz//HF7DmvZrbNDGthJjih1Or3k2X1+WdmvP/+Bx07dghFwlwIRKLIjmhGV1NTPWvmzMqKKo8nhlBx1113XHfd//vDHy5/aeZLF198USSiqaoTAFwux5133jl9+j0DBgzct38vpTQ2NhYBgHOChBDJ6431en3Z2e2efOrPmVlZABAMhsLhsNfrlag0ffrd77777ty5c2+5ZQoTEVPoDqez3l+/bv0660oyMrNee+3Njz/+qF//XlokwJjhcbsL8gtuufmWstKySy65pG/fXoCMcU5l1eH0rlq1qqS0GBGtGhIAODBAIIAAnAuCOgOTcUNnDX6jpkYrLQ0czzf8dYIzIYQl1wIAr6+PqQ/EqB7d5eWTb0qc+XLsjBeNG64PUYoGU6nLW1mj7dghmpIRjJQXBz/+MDZSD1SC3n3Vp/6i3n235+knwz176AYRTklO8hEARGCaVvXaG8rKlT4iG7Hxxu1T3TNf9b4wh48aH+Kcuoiv9ETwrbf0ulrA0ySl/6WIZCWmp6rtjRs37t+/f9KkSVOmTElJSWklE1ldAh06dnzsicem3XnHwgWf9+vbZ9KfrtGvv7H+uRe8LELKy+tmzfDdcXfsuAli4NDInrzAt1+b69dCfY1DVWmYCAImlUxN4336xDz0sLNjl6ZMlzMktGHdusibb6s5HXxXXEUIsZS6H557kWW5sKjw1ltvJpTWNzT66/2cC0VRXS7XpZdOuvWWKZkZWYxxSskZ4j4KwRMSEqZNm37eeed6PDGHDx+aOvW2wsIThFBEDAZDgUAgLi5OUdSsrLYJCfFeb4yVxLcQ00WM19uzZ6/YWB9nzBonK8O0DtizO2//gb1OpxqONGRkpHfr1o1zbpqmoqjXXPOnVatW1dX5JUkqLCxcu3btFVdcKYSYfvd911930z8++3T+p5+qisMwjfT0jOnT70lISPB6PT169LScQmpqqsOhGobh8/k6d7Y6p2Ds2LEfffxBYUGRqroAsN5fH31wGRkZGRkZyUnJhmkiIhecSjj1ltsnTJyYlJRUW1Nz1123b9y41eFyAUA4FK6vr8/KbOHNBXAgJiESApVkeqLI/8yzjJmk0S+HQrwxFAhHPJMnx026rKXVqkkphqSSiK55Ykn3XkglAHBdMLb+Hwuc+QVEBSQaC/qjxxtHj3srymIY10TYgLDCTQQgDmegT/fQ5tWSKSVkZAkARFK3ZgUuXex2gxkK0PPGeS65jACo6WkJ06aF84+yIwdkhSh794S35Sljzv+xGctfISKdKhWsUigmJkbTtLfffvOGG65fv36dFZEsy2jJpUFDhky/715BcN4rc/KPHk28cLw5ZrRpCh+nvqqa+hdm1n+9GJyqY+CghIcfdz3/kjZhYlCRBI9wKjUKFhxwju/Rx5wduwjOAVFwQQgNHz4YePVFCNTHTJrkbJMGQsA/NwtpmmZ9fUNtrR8YdygKASIYaGH98MGjS79ZXl1dfSaLrNim61rfvv0mTJjg8cRwzjt16jxgwEDD0K2ExTAMXdeiJWIoFGYmO7P5QwgRiYTNU29hS+9z8mSVYZhIuADmcrkoIYQQSZKEEGlp6enpqYZhEkKQQGVluVX3t2+f3adPn8yMLCszFIK53Y5hw4YOHTqkR4+eUTPt1LmT1+sNBBvi4+MSEpIsbvt8sXGx8dbcNABwxlvO7XLODWZYdWEkEs7tkHPd5MlJSUmc8/iEhLFjxkmSxBlHJCYzmclOL6SZADSpxCggolxVJb74wlzyJV2/yvXdZvXIIbWsDMNhcapxCAHAlZlutImvYZqWmaLExVo+m/jiIK4Nmki4CcTkLHzKupiJpgGMKFSS9+zQZr1knCgSgL5xv/Pcfbf71jsc/QYgANO1wIplqh7gwE0HUTt35LU1oZLC4LEjRkmRmpwkaQKJpGiacfQY/7nTWT9DtYsWb8RkJiJxudwHDx6cevttV1155W233e71+qLSFgAQgpyLP1x+RUFBwdxZs+a9NPPZ2XPibrjev3+f6/BBl+ogwXDD7Neox+MZcR4S4ujTP7nnOdqItf73Pwzv382GDEy55yFHZqYlcwshkKBRfbLh5RmOw/vDw0Z6Lzj/n3YAxDCMLl27PP/cjLi4eM55VdXJr75auGDB57qu5+Vt37p109q1K59/7oX0jIzmCc3TxAvOedRNCCGSEhMlSbI696yVi1H3QSkBtAaqVbKJUbm5eREKRiOSz+ejlAIIiUr19fXhcFgIwRijlDLGdV0npEmQSElJbtniYJoGAHLBGWcCRDgc4jxWCEEptb4+tU36Aw88VHWyqm3btklJSafmuEiLhTAtaG/xhyJFBM4tqpOW9x4XnwAgCLF0GyFOdxkIBAQRQBARBMOEeNpvACHCDIVDRsRgRkBDX3LSKYkaEQDkNqneO+7iJcVKdnspyWpWBgQCFEzkFAEMTgUVzYNGO+aGUjMchQWyQoAb2jeL+IHjjvETHBeO9l59XfSOtPJS+cARN1cdGlKCkbc+NP7+aVBiwgh7QlwN6kSVdBEJOmQ5MZ78ZqodAEfkiMAYI0hACABUVYeh63974/UDB/Y/8uj/dOrYpYWjRUSBiLfdfsfRY8c+XfhVt359r7/hptB119Q/+2RcWFdRij1Z2fjXF4gvxnVOH2GaRJJcI86XuvcW27Z4evVypGVYLLJGloXDta++Lm/LY67YmIt+L3t9Te9+bwdnU5exlWUpipKVleX1+gAgMzOzT5++GRlpM2fOdLmcQjhXrV45Y+bzLzz/omw1UpyyDwHAASyv32SFVCZCWCoLRMuDqNYqwPImlpFhq+4eAGz5Ecsyevbq3T673dGjx9wez8nK6uUrVkyZkmtNJxw4sL+4uMShOiJaMCUlcdCgodEPtpwdIoQKLggh1nWeesySfMnFl7YaH7fbTYkkBAqBTUZ7lrJYIHIEEIITgpY3QURZlpCA1QQsWt+9dT+EcI4oBDPNdpm+hx8WHq/QNTS5MM0Y0yBOJzQxsKkaIJIUM3JUa/cng4NFgDCGIIGEPDodIRzpbR1/urp+9ixPYwM6JImoypFS8+ic2uWfOS6+wjPxEup2AoBeVkEaA5KgTCBDyoMBCDAnABegERJUZZQ9BjWUyybFTbzQWpv6G80jCSGE4IAiKjPous5MzelQN2xYf+utU+6Zft+ECRNb9rNwzmM8MQ8++PDRo0dnvzp74ICBXcaPP7F7B13wpY8Ziip8Bfn+556nzzyj5uQg54CoJCYkjpvQZJWnJvgh8M7brq+/QUMKDhnsHD7yp8os1tU2eVEhCKHXXnvtmjVrNmzcGOPxxni8a9as2b17V//+A09P8LD5T+sOjzPbeVpqM03VEec/mi5zztPS0u+4/a6HH3koEAjJsvrGG28apjF0yNDa2tp58+YGg2Fumi634667pmVn51j6RIsvRUR61qVX1lOoq6tbv35tKBS21k8ahqFpWnl5uZU6fl8TthACgLSifTPFOAD93s4XARTQ0qXCDkVWFIlQcLhamh0D65gWPccNDYHtO8DfIAsOpk50gwSDauUJIQugBAQVzQMuEFCIhIsvbfB4tXc+FEf3ukFXJaqCYPn5oZdnGAcO+m6fqiSnYCSgomESXQiiJyfId91GUlKFaSIQA8EEATpzqrK7axficOLPVe1+MpEIsSQWpFTiBo/oEcMw2rbN7N69+769u0vLSouLSx948KHjx/NvuWWqLMtWmmcVS7m5uY8++tjUKTe+8sILL7/9btzV19Xu2ufOP0FMg6ou18EjjS/Ppk89IcfGCcGbumGxqS0fOEdCGr5Zwj56z2vyOleM46JLJKfLmuD7UbGxldVaZR4hhDGmqs6uXbtuWL+RMSZJUm1tbUFBQQsiiRacIa0TGLA0p7MaIlq2TpD8M/1s1jzSxIsuRsQ333orP78wEAg899wzsiwzxgzdjInxdu7e9bZbbxs/4cIzGhGsCyAI0Rh4GotOnCh69NFHNm3aTCgaRkTXdVV1EkIQiCzLZ23kP91ZICI9M1iddXibiSQEcgYMkDAhWY0hTV5PCECs3LA+snoDbdMm/pILXcnJgBguK2mYMUds3qQCA8Ns4BHhIBRovBmRJCoEMxFF87NGBC4AkfhGjzW69679dklgyZfK4QIJVJU6CZqBbxbUOpWUBx8hbq9GBUWdMKpHQlJ6htK9p5UkOM7WwP4bpXaMMyQEkUQiEYqYm5s7duyYiRMvzMnJ3bU775VXZq9dtzkc0ea8Mqe8vOKhhx6OifFGucQYGz169O13Tnv9uRcX/P2DP15/feiKK6tnzkw0dM5AcUraxnW1732QcOcdtHn7h2aT5EBI4/794blzk4JBAyjr2ck9YAD+YGnYwtTw9HjKo0qaNXfMTC44IhBEgQiEktOndZsmlBB/gKVnxByBiGCt2mCctRDoIDrl1mQQnFv1jBCCMX7hxIviExKm3zM9Uh0cMnio1xdDkKSmpp9zTu9hw4YnJiadZdGROHWbiCBAtJR8dF2fNWvmylUrY30JhqG1aZN2ySW/79SxMyLOnfvq0aPHXC45KpNYV3L6vQsAgkBbDuMZ1Z9o3Y4kuAkmlWQpjBKjyEUTkQQAAtM0bf4n6spVRlIa9O8DycmmYdS89ZZ71eIYp8wZjySkuMefjx06yQLIW29i/nF0oqAmgImcgwBB0epTQwAlNSXluuu1888Lf/WNWPit52QZ5VQmom79eu2aYjk7u9GXwBqCKCP668PLV8jdu6MprKXUTWk/IUDIb9rZQIklNbN27bKm3nr7mDFjYmJirMfQ55x+r74y7933//7Wm2/W1tZ8+unH1TU1jz/2RHp6erS7RAhx08235G3b8fLs2f3PO7fdhROPb1nvWLnaRxyCRXxg1s7/uLFDduz4C6NyttXZHakob3zpJV9pGSqSxoVr2FAa4/nRhehNHQx4yklH+9YsW5RluaqqauvWbZIkEyIZhh4fl9i+XftmgxAtqIotKhywagOrC9JqfATgrRRLITgSjggFxwsAQJIkAJAk2rT2BAUiUGtAT0kRuHzF0hdnzqitPXndddffffe9qnrawv7T22Gbr4xYpSAgoqYbFGnLXtKSkuJNmzd6PG5AoevGTTfeMnnyZOut+fM/PXTosKV3SJLUskeWc84ZR2tKGgUHHu0ob3JDLfa0QKSc8yYBxfJQIBRrShaFLAjKSss61igpdRccczrlgNcBMgKAVlpK8na5ZCQmi4AkLr889sYbLYJGvljkOFLEOSFCEIJWAmI5ovIFX5DGQPxFF8pxcY7MbPn2qcEO7YNPv+Bt1IhgDn+Al1U4BgyAHj3M4/mKqqpc8i9e1tjrHM+okaTFmHLODX9A9Xl+OyJp4YgeibhdbkPTDx06kJ3dvnv3HpZpmqbpdLpvnTKlf5++L82a+d3WzUuXLq6prpoxY2Z2dk7T+lMhnC7nw08+ce1VV7384oxZr76WdN3kuoOHHeV1KicEuTfir319XrBdtqdr12jZwbRI3dxXvdt2ORwUBaNCAklu2Xb1fRNHhBBKqaVOCSEIkSMRvaSkBJE4HE7GzOLi4tmz/3ro8EGn28GFGQgELrnkql69zrE+63SqAMibIoCwskFLHSaE0CbLQIISZ0KWlWjR4nK5PB63EAy5UFXlu63fzZnz8vDh59XU1CxZ8vXOnXkOp8KYhgiFhQW7d+/inHXt2l2W5a8XffXUn5+sqa2OifHu3rPrzjvv8Hhcqalp6enpWVlZqalpaWlpbnfr580YB0ABXJKUivKTc+bM6dq168GDB+Lj4++++x7D0HU9gigQuNXlK4SoqKjYf2BvSWmJosjW7RQXF+/Zs5sx1rVrN0Wxmj9ki7BCMEqbuvUsDquqIhA5cIHIOVLSRF0r00ZZAgqKIGgSJFwOhfiRQ7rqZOGI0DQRDosNaz0nawWXDaRccQIA6AY1dYHMJCpqRDGA6ZpRV6sfPUwqq4FQQRjK6CwoDu/MM10Od26u2eiHf3xGDh1r3LBZ/v1E55CBcmy8Z8DAYFKK0VgoC8G9ThHjJAAxF18W3rzBcbKeyj5nTZX53IyG/QfU4cNErJfoullWVrdmdYUR7vHgo44Y389bZfwT1iNZX1BSUlxcfMLvb6itrd22feuK5csOHjzgcDjS09MtpcswjMzMzAsuuIALfvTo4WPHju7bv6/POX0TExMt62ScJyYkJCUmv/m3v7Vrm9Fj+IhgoDGyd6cDCRGESbKo8Ycb/I6hw4gsW0pDzScfyh/83UeIAEZNwXSjMSnRc96IJk/8/amd31+3cOFXy5cv1zRGkFIqBwKBzZs3Ll++fPnyZQsXLnzvvXfz8nZRSgxDN4zIxIkXPvLIox6PBxFN09iyZdPSpUsbGwKUypSSQKCRc5aZmeF2ewzDXLr02717D8iySinV9UhxSZHb7UpLTSeEyLKy/8D+LZu/czrdlFLTNDdv3vzFFws+/vij6uqabt26lZWXE6SUKpWVlZ/O/3Dhws8njJ8oy8r9999fWFTocXtNg5WWlh87dvzw4UM7d+5cvWbV4iWLly3/du261fv27g2FwklJSQ6H03oo+fn5K1YsV1UZgAsh9uzZ8+2yJdu2bfPGxPzud+NiYmLWr19fUHDC4XAj0mPHjn733Zb33nvno48/JEg1zSCEyLJcUVH+8ScffrP0m7FjxyUkJBw4cPDrrxeePFktSaokSeFwKBQKpaamxsbGAsDGjRvXrVsvSw6CkhC8tOQEIZiWlqo6nJHKitA3S3DjBiUUlEHmlBrhoL55s7Z4kbloiVi0GJcudRzeJwsODBuTkzy/v1j2xIDD0Zi3QxQcR1VREMWxguCu3Y1ffNbw1VcgDGQmJYaQhVlTGvx6ccWW7zzDh0c2fKd+9qXPKbHSfGPjpvDO3SzoZ4cOk43rEQONhhYaOSb24t8jJUpqRkRxBvbuEmajKoEcbiDbt7FVq7VVq4wlSyNffR3IP6S3zfANHOhwxeDPmkr6CUSyaDpo4KBR549u166tELzB31hXV3fgwKE1a9bs37/f6XRlZmZaAoPT6Rw2bHhObu7x/PztO3bs3LGzW7cuaWnpnHOCKITo3LlT4YnChV8uGH/RxXGdOwd27pAqTipCBsGpQnhhUTg+3tmjJyIGN20Ov/hCrB4xJWsFFDEpRjjKw4bKMd6zFoiWeeXl5d1ww/977/13amrqTNMIh4OaFgmHg6Wlpfn5+QcPHjpy5HAwGHA6VZfL3a1bt1tvnTpt2t0+n08IUVtbe+edU59//rny8nLD0CORSCSiVVdXffb5p+3b52RmZt508/ULF36laUYkEtH1UGPAv3XrptWrV1xwwZjk5BQhRMcOnYqKig4dPsSaoSjK6NEX/OUvz1xzzTV5O/MOHzmEAJyZWVmZk6+/ccCAwV6vd8eOnQUFRYqiGoYZiRiUEkJkSmVFURAwGAwVFhbt2rVr8ZJFAGLYsHOtO23Tps2ePbuPHj1MKOHcBMJ79uhx11133z71jpiYGFVVU1LabNmytbKyklLa2NCwa/cuxszJ199w3333HTp0MD//GABomp7dPufmm27u06fvk08+/tBDDxQUHDdNMxKJGIZeV1f39deLVFUdPHjwAw/cO3fuKxFNC4WDuqaFQ6ED+3cv+2ZR7379fJqx5/4HcclSpz9gMjMizKAwNV0362p5Y7UU8ktaiIER5jqaTAvojVkZvkl/kBSFKoqIiQ/vOyhqqxC4HAxq+cf8BNy/vyxu8o3B0jI8VkwFNEoQyWrru/Qyd+8+jXv36YeOQLCekAgFXSkpZ5u36jvySDgUBr1hwOCEadMdCYkAHJE4u/bQEhMChQVqTZ0GJiqSpBnc7w9quujRyzvl5sxrr3fHJf7s9YU/f8viSCSyffu2lStXrF23pqCgQNd0rzd2xIjzrr32usGDB1v5D6W0qKhg5syZCxYsyM7Ofm3e6127deOcIRJEPFl18tprrhp57rkPPPxY3TdL9WeejY8EKZhMRi1ihFLbeme9JDyu2runxR87pFCJCcEAJSFMSWqkLsfjj3vHjD6rameZ1/79+z9fMF+RZbBWewmMrmZr6p9EdLlcbVJScnM75OZ2dLvd0YSwpqbm7bffjGgRWZaF4AQlIZAQCIcDEydenJaW9s67bzITJEkRAgBMJCiEcDrVP111TVJSm6ji/OGH769YsTISifTt23f06DHDhg23KqWioqIXX3zh2LGj48eNv+yySWnpGdZll5aWbdy4fs6c2Zqmjxs3LhKJlJaWFhcX19bWRCJhw2CqqiqKo7Gx4YLRI996891TmUJx8ZtvvbF+w9qYGM+ll04aP25CYmJKy9HYsWP7hx/+/fDhw42NgY6dcm+6ccrAgYMA4NixIzNnzjx+/PjEiRMvu+wPqampnPPXX5938mSlw6lyBoTIQnBCSCQSGjZs2JAhw954Y159fb2iqs3r00BwQSlc9sc/+pgo/WapTxAKXCKCc0FQAlMHRAKCgGACuSoxAnLEQJNouW1TL5woESq4AIKB3bsin39Ojx0WoTBr28nx/66M6dMHAUOFhQ2vvwtFhcqooa5x4xzpGQKAG0b4yBF97WqxL4+XlZFGjZtouh3YJgUGD4i/dJKckGhNkFl2LhC1guP6ylWBg3tJVaUc0nWfTx4z1jd2rBIb9wuXuv8cIkVXd1n/LC4uWr161fLly7dt/c7vb0hMSh4zZuwNN9zQtWu3KOXefPONV155Jatd+3lz5+Zk50RXZa5as/Kxh+9/5ZXXevc8p/jx/3Eu+TrGQWWTATh4hEWGDQw7CV232UcF0XQBkpAlExkirTcYu2Vqm5tvalrX9WsguvTt11olbp3K7/drmpacnNzqWwzDqK+rS0pObvliTU3Nk089vnbtmr/8+emJEy8GgECgsb6+vri4eOfOnXl5uw/sP1hbW6fpoQkTxr3yytxW31VRUa46HHGxca1up/kAUVVVpel6YmKCQ3VGn6Oua35/g7VrWsuulN8aViMlY0ZtDTcMKS5BcjqiazdNQzcDITUuFqNL9JqfFA8EjYZ6Fg4hR1QUOTaW+rzRhTYtFt5yQCIATF3nDY1gmsTllL3eJlnplz36nx+RWmlHwUDjtu3bFi1atHr16pKSksysjBsm3/DHK65KiE+wjl+5csX/PPGEy+l8be683NzcKJeef+6p40ePvfq3d0P79obuvy/+ZClFhRgEAA1kjHAFVQI6IhdcAaJwpgecVBs7zjt5siMr6we0/+iiox/MV6OqMZ65UvWsg0MIsXacPeukCyX0rB6n1XBFTb/lrgxHjx557H8eW75y6ZVXXjln9jzTNIUQkiRHr8sw9Hvuue+bJd+YTH/wwQemTLn19PNgtBH2rLfTkiHRf0bPwLmILnT4wXvHH3r37Pva4fdOVTXvK3r2mZyoKntqEdrpW+Fx0TR/dtbFsPhPvP4r7a33kzc/OX1aE6PLj1XV0a5d+7Fjfzd48OC4uPj848cWLVq4adPG1DZtsrNzACA7O/vcc4evX7du8ZKvR44c6fV6refdq+c5ixYucjgc3YefGyopJ3sPUkURBFAAVwAkRWKCCM4loklKRGd6XAqZOsV76y1KXBz82DwS+TFYSeZZA9H3fdxqO/qe05Ezh0icJhNjy+nXlovPDcO49967165f6/N6VdXRp0/fpKQkSqn1iVAoePDgwdmzZ69YsTKihbOyMqZPvycuLr7lIvbofgVnbiPR/EWn7L+l5t78qdMu7/vv/Qffbd6k4if8acWr5ksUcGoi8bTdJFt+pKkfUggQ2KzIn+WwVrxt6QW+78jfLCL9cL5XVFTw1Vdfzp8/v7y8fNKky6dNm2atPK2rq3vyqcdr6+tmvzwn1htr1VGb129892/zZrzyqlRX779jmq+sQKYg6ciRmlSipknRNClUEcJ79I296Ub3wIEAgD9jN+N/Y3DOp0y5cdmylR63xzD19PTUoUOHJicnS5Li9/uPHDm8b9/e2tp6zoWiyE8//ZdJky6393z998Gv//uRWuYVZaWl8z+b/8nHH3k8nvvuv3/smHGWPv7irBl1tXVPPv6U0+m0uDTvrzOSktpM+tM1Fa/Ok996K54AmiAECspREkI3aiUpMvGixCm3Otq0Ec1bZPwnsYgQ8u2ypXfddadpMFV1GIYZiUSsXIlzZk1SGaaWmJh4z/R7r7zyT0Lwf8n+dTZ+5dTuh1M+i59er3fQoMETJlwYiYQ/+PD90rLSnj16ulyu4UOHF58o3rJl0znn9LV2XezYufO6tWs6dO0a1z67dtM6R20DIbJJGFOEZmjB+CRy480Jt0xRYuOsprv/sN+eZI1YdnZOXLx377699fUNQnBZlghFRLSSLofDOer8kU888fiY0eOsbg/bfP+TI9L35XuFhfnvvPd2MBi56YabO3XqBABLlix2OB0jR4wCEIhk+9YtjTVVI8dNLJ39V8f77/pQYZLwG4beuVfsbVPcQ4ZYZ/yPTmYEAB45cmTZsm+2bd9aW1NjMibLclxcfNcu3YYPH96/f39ZVsWvJ1Ta+L9BpGjeEq1ut2zZsnXblmFDh/fufQ4hJC9vZ0yMJyenAwBwxvZv35rToxcvKfZPn5ZUWRmQJX3U6NjbbnOmZ0R3APzPfh4ttDXh9/s1TVdV1ev1Rt3Hv1KetvGvJVKr2ikcDuft2pGUmJyT04EQKC0tTkhIdDhcAKCFI1wI1aGWPvI4X7PKdcO1cVddQ10uYBzpf4v1nFW8PrNR1cZ/KZFaOdRAIOB0Oq2O/dO8rABAqN+2Q6+vSxw9ilgLW/4rrafFFBP8R/zqE5tI/wv28QPcaLFKrmklsv2cbNhE+plsa6KQTSIbNpFs2Pgvga3/2LBhE8mGDZtINmzYRLJhw4ZNJBs2bCLZsGETyYYNm0g2bNiwiWTDhk0kGzZsItmwYRPJhg0bNpFs2LCJZMOGTSQbNmwi2bBhwyaSDRs2kWzYsIlkw4ZNJBs2bNhEsmHDJpINGzaRbNiwiWTDho1fgv8Pcmlv/0fMkroAAAAASUVORK5CYII="""

# ─── DATABASE ────────────────────────────────────────────────────────────────

# ── Logo bytes cache — sirf ek baar decode hota hai startup pe ──────────────
import base64 as _b64
_LOGO_BYTES = _b64.b64decode(LOGO_B64)

def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY,
            username TEXT UNIQUE,
            password TEXT
        );
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            hsn TEXT DEFAULT '',
            sale_rate REAL DEFAULT 0,
            gst_percent REAL DEFAULT 18,
            opening_stock REAL DEFAULT 0,
            barcode TEXT DEFAULT '',
            unit TEXT DEFAULT 'Pcs',
            mrp REAL DEFAULT 0,
            purchase_rate REAL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS parties (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            ptype TEXT DEFAULT 'Customer',
            mobile TEXT DEFAULT '',
            gstin TEXT DEFAULT '',
            address TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bill_no TEXT UNIQUE NOT NULL,
            bill_date TEXT NOT NULL,
            party TEXT DEFAULT '',
            party_gstin TEXT DEFAULT '',
            party_address TEXT DEFAULT '',
            party_mobile TEXT DEFAULT '',
            grand_total REAL DEFAULT 0,
            pay_mode TEXT DEFAULT 'Cash',
            due_date TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS sale_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER NOT NULL,
            product TEXT,
            hsn TEXT DEFAULT '',
            qty REAL DEFAULT 0,
            rate REAL DEFAULT 0,
            taxable REAL DEFAULT 0,
            gst_percent REAL DEFAULT 18,
            gst_amt REAL DEFAULT 0,
            grand REAL DEFAULT 0,
            FOREIGN KEY(sale_id) REFERENCES sales(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS purchases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bill_no TEXT UNIQUE NOT NULL,
            bill_date TEXT NOT NULL,
            party TEXT DEFAULT '',
            grand_total REAL DEFAULT 0,
            pay_mode TEXT DEFAULT 'Credit',
            due_date TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS purchase_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            purchase_id INTEGER NOT NULL,
            product TEXT,
            qty REAL DEFAULT 0,
            rate REAL DEFAULT 0,
            taxable REAL DEFAULT 0,
            gst_percent REAL DEFAULT 0,
            gst_amt REAL DEFAULT 0,
            total REAL DEFAULT 0,
            FOREIGN KEY(purchase_id) REFERENCES purchases(id) ON DELETE CASCADE
        );
    """)
    c.execute("INSERT OR IGNORE INTO users(username,password) VALUES(?,?)", ("admin","empo123"))
    # returns table — sale/purchase returns track karne ke liye
    c.execute("""CREATE TABLE IF NOT EXISTS returns (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        return_no   TEXT UNIQUE NOT NULL,
        return_type TEXT NOT NULL,
        return_date TEXT NOT NULL,
        orig_bill   TEXT NOT NULL,
        party       TEXT NOT NULL,
        product     TEXT NOT NULL,
        qty         REAL DEFAULT 0,
        rate        REAL DEFAULT 0,
        gst_percent REAL DEFAULT 0,
        gst_amt     REAL DEFAULT 0,
        total_amt   REAL DEFAULT 0,
        reason      TEXT DEFAULT '',
        created_at  TEXT DEFAULT CURRENT_TIMESTAMP
    )""")
    # bill_payments table — partial payments track karne ke liye
    c.execute("""CREATE TABLE IF NOT EXISTS bill_payments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        bill_type TEXT NOT NULL,
        bill_no   TEXT NOT NULL,
        party     TEXT NOT NULL,
        pay_date  TEXT NOT NULL,
        amount    REAL NOT NULL,
        pay_mode  TEXT DEFAULT 'Cash',
        note      TEXT DEFAULT '',
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )""")
    # Existing DB mein naye columns add karo (agar pehle se nahi hain)
    for sql in [
        "ALTER TABLE purchases ADD COLUMN pay_mode TEXT DEFAULT 'Credit'",
        "ALTER TABLE purchases ADD COLUMN due_date TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN pay_mode TEXT DEFAULT 'Cash'",
        "ALTER TABLE sales ADD COLUMN due_date TEXT DEFAULT ''",
        "ALTER TABLE products ADD COLUMN low_stock_alert REAL DEFAULT 5",
        "ALTER TABLE parties ADD COLUMN state TEXT DEFAULT 'Uttar Pradesh'",
        "ALTER TABLE parties ADD COLUMN email TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN gst_type TEXT DEFAULT 'CGST_SGST'",
        "ALTER TABLE sales ADD COLUMN bill_mode TEXT DEFAULT 'GST'",
        "ALTER TABLE sales ADD COLUMN party_state TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN irn TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN ack_no TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN ack_date TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN ewb_no TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN ewb_date TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN vehicle_no TEXT DEFAULT ''",
        "ALTER TABLE sales ADD COLUMN transport_mode TEXT DEFAULT 'Road'",
        "ALTER TABLE sales ADD COLUMN distance INTEGER DEFAULT 0",
        "ALTER TABLE sales ADD COLUMN transporter TEXT DEFAULT ''",
        # ── Grocery new columns ──
        "ALTER TABLE products ADD COLUMN barcode TEXT DEFAULT ''",
        "ALTER TABLE products ADD COLUMN unit TEXT DEFAULT 'Pcs'",
        "ALTER TABLE products ADD COLUMN mrp REAL DEFAULT 0",
        "ALTER TABLE products ADD COLUMN purchase_rate REAL DEFAULT 0",
        # ── Unit column for bills ──
        "ALTER TABLE sale_items ADD COLUMN unit TEXT DEFAULT 'Pcs'",
        "ALTER TABLE sale_items ADD COLUMN batch_no TEXT DEFAULT ''",
        "ALTER TABLE sale_items ADD COLUMN mfg_date TEXT DEFAULT ''",
        "ALTER TABLE sale_items ADD COLUMN expiry_date TEXT DEFAULT ''",
        "ALTER TABLE purchase_items ADD COLUMN unit TEXT DEFAULT 'Pcs'",
        "ALTER TABLE purchase_items ADD COLUMN batch_no TEXT DEFAULT ''",
        "ALTER TABLE purchase_items ADD COLUMN mfg_date TEXT DEFAULT ''",
        "ALTER TABLE purchase_items ADD COLUMN expiry_date TEXT DEFAULT ''",
        # ── Expiry tracking flag for products ──
        "ALTER TABLE products ADD COLUMN track_expiry INTEGER DEFAULT 0",
        # ── New Medical fields for Purchase/Sale ──
        "ALTER TABLE expiry_stock ADD COLUMN packing TEXT DEFAULT ''",
        "ALTER TABLE expiry_stock ADD COLUMN mfg_company TEXT DEFAULT ''",
        "ALTER TABLE sale_items ADD COLUMN packing TEXT DEFAULT ''",
        "ALTER TABLE sale_items ADD COLUMN mfg_company TEXT DEFAULT ''",
        "ALTER TABLE sale_items ADD COLUMN mrp REAL DEFAULT 0",
        "ALTER TABLE sale_items ADD COLUMN disc_amt REAL DEFAULT 0",
        "ALTER TABLE purchase_items ADD COLUMN packing TEXT DEFAULT ''",
        "ALTER TABLE purchase_items ADD COLUMN mfg_company TEXT DEFAULT ''",
        "ALTER TABLE purchase_items ADD COLUMN sale_rate REAL DEFAULT 0",
        "ALTER TABLE purchase_items ADD COLUMN free_qty REAL DEFAULT 0",
        "ALTER TABLE products ADD COLUMN packing TEXT DEFAULT ''",
        "ALTER TABLE products ADD COLUMN mfg_company TEXT DEFAULT ''",
    ]:
        try: conn.execute(sql)
        except: pass

    # Custom units table — user-defined units
    c.execute("""CREATE TABLE IF NOT EXISTS custom_units (
        id   INTEGER PRIMARY KEY AUTOINCREMENT,
        unit TEXT UNIQUE NOT NULL
    )""")

    # Stock movements table — manual in/out entries
    c.execute("""CREATE TABLE IF NOT EXISTS stock_movements (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        move_date   TEXT NOT NULL,
        product     TEXT NOT NULL,
        move_type   TEXT NOT NULL,
        qty         REAL NOT NULL,
        ref_no      TEXT DEFAULT '',
        reason      TEXT DEFAULT '',
        created_at  TEXT DEFAULT CURRENT_TIMESTAMP
    )""")
    # ── Grocery: Expiry Stock Batches ────────────────────────────────────────
    c.execute("""CREATE TABLE IF NOT EXISTS expiry_stock (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        product      TEXT NOT NULL,
        batch_no     TEXT DEFAULT '',
        mfg_date     TEXT DEFAULT '',
        expiry_date  TEXT NOT NULL,
        qty          REAL DEFAULT 0,
        purchase_rate REAL DEFAULT 0,
        mrp          REAL DEFAULT 0,
        supplier     TEXT DEFAULT '',
        packing      TEXT DEFAULT '',
        mfg_company  TEXT DEFAULT '',
        created_at   TEXT DEFAULT CURRENT_TIMESTAMP
    )""")
    # ── FIFO Stock Layers ─────────────────────────────────────────────────────
    c.execute("""CREATE TABLE IF NOT EXISTS fifo_layers (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        product       TEXT NOT NULL,
        purchase_date TEXT NOT NULL,
        purchase_bill TEXT DEFAULT '',
        batch_no      TEXT DEFAULT '',
        qty_in        REAL DEFAULT 0,
        qty_remaining REAL DEFAULT 0,
        cost_per_unit REAL DEFAULT 0,
        created_at    TEXT DEFAULT CURRENT_TIMESTAMP
    )""")
    # ── FIFO Consumption Log ──────────────────────────────────────────────────
    c.execute("""CREATE TABLE IF NOT EXISTS fifo_consumption (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        layer_id    INTEGER NOT NULL,
        sale_bill   TEXT DEFAULT '',
        sale_date   TEXT DEFAULT '',
        product     TEXT NOT NULL,
        qty_used    REAL DEFAULT 0,
        cost_per_unit REAL DEFAULT 0,
        total_cost  REAL DEFAULT 0,
        created_at  TEXT DEFAULT CURRENT_TIMESTAMP
    )""")
    # Expenses table
    c.execute("""CREATE TABLE IF NOT EXISTS expenses (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        exp_date     TEXT NOT NULL,
        category     TEXT NOT NULL,
        description  TEXT DEFAULT '',
        amount       REAL NOT NULL,
        pay_mode     TEXT DEFAULT 'Cash',
        ref_no       TEXT DEFAULT '',
        created_at   TEXT DEFAULT CURRENT_TIMESTAMP
    )""")
    # Settings table
    c.execute("""CREATE TABLE IF NOT EXISTS settings (
        key   TEXT PRIMARY KEY,
        value TEXT DEFAULT ''
    )""")
    for k, v in [("einvoice_enabled","0"), ("ewaybill_enabled","0"),
                 ("install_date",""), ("license_expiry",""), ("license_serial",""), ("license_customer",""),
                 ("shop_name","BhugtanEase"), ("shop_address",""), ("shop_city",""),
                 ("shop_state","Uttar Pradesh"), ("shop_gstin",""), ("shop_mobile",""),
                 ("shop_email",""), ("shop_bank",""), ("shop_ifsc",""), ("shop_account",""),
                 ("shop_upi",""), ("shop_print_tnc","")]:
        c.execute("INSERT OR IGNORE INTO settings(key,value) VALUES(?,?)", (k,v))

    # ── Migration: Purane Cash/UPI/Bank/Cheque bills jo bill_payments mein nahi hain unhe fix karo ──
    try:
        for tbl, btype in [("sales","sale"), ("purchases","pur")]:
            existing = {r[0] for r in conn.execute(
                f"SELECT bill_no FROM bill_payments WHERE bill_type=?", (btype,)).fetchall()}
            rows = conn.execute(
                f"SELECT bill_no, bill_date, party, grand_total, pay_mode FROM {tbl} "
                f"WHERE pay_mode != 'Credit'").fetchall()
            for r in rows:
                if r["bill_no"] not in existing:
                    c.execute(
                        "INSERT OR IGNORE INTO bill_payments(bill_type,bill_no,party,pay_date,amount,pay_mode,note) VALUES(?,?,?,?,?,?,?)",
                        (btype, r["bill_no"], r["party"], r["bill_date"], r["grand_total"], r["pay_mode"], "Auto-migrated")
                    )
    except Exception:
        pass

    conn.commit()
    conn.close()

# ─── SHOP SETTINGS HELPER ────────────────────────────────────────────────────
def get_shop():
    """DB se shop ki details fetch karo. Har jagah isi se use karo."""
    try:
        conn = get_db()
        rows = {r[0]: r[1] for r in conn.execute("SELECT key,value FROM settings WHERE key LIKE 'shop_%'").fetchall()}
        conn.close()
    except:
        rows = {}
    return {
        "name":    rows.get("shop_name",    "BhugtanEase"),
        "address": rows.get("shop_address", ""),
        "city":    rows.get("shop_city",    ""),
        "state":   rows.get("shop_state",   "Uttar Pradesh"),
        "gstin":   rows.get("shop_gstin",   ""),
        "mobile":  rows.get("shop_mobile",  ""),
        "email":   rows.get("shop_email",   ""),
        "bank":    rows.get("shop_bank",    ""),
        "ifsc":    rows.get("shop_ifsc",    ""),
        "account": rows.get("shop_account", ""),
        "upi":     rows.get("shop_upi",     ""),
        "tnc":     rows.get("shop_print_tnc",""),
    }

# ─── HELPERS ─────────────────────────────────────────────────────────────────
MON = ["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

def fmt_date(d):
    try:
        y,m,dd = d.split("-")
        return f"{dd}-{MON[int(m)]}-{y}"
    except:
        return d or ""

def fmt_exp_mmyy(d):
    """Expiry date ko MM/YY format mein convert karo (display ke liye).
    Accepts MM/YY, MM/YYYY, YYYY-MM, ya YYYY-MM-DD — sab ko MM/YY mein normalize karta hai."""
    d = (d or "").strip()
    if not d:
        return ""
    # Already MM/YY or MM/YYYY
    if "/" in d:
        try:
            mm, yy = d.split("/")
            mm = mm.zfill(2)
            yy = yy[-2:].zfill(2)
            return f"{mm}/{yy}"
        except:
            return d
    # YYYY-MM or YYYY-MM-DD
    try:
        parts = d.split("-")
        if len(parts) >= 2:
            y, m = parts[0], parts[1]
            return f"{int(m):02d}/{y[2:]}"
    except:
        pass
    return d

def exp_to_storage(d):
    """Expiry date ko storage format (MM/YY) mein normalize karo.
    Accepts MM/YY, MM/YYYY, YYYY-MM, YYYY-MM-DD ya empty string."""
    return fmt_exp_mmyy(d)

def exp_sort_key(d):
    """Expiry date (MM/YY ya legacy formats) se sortable 'YYYY-MM' key banao.
    Sahi date-comparison ke liye use karo (string MM/YY directly compare nahi hota)."""
    d = (d or "").strip()
    if not d:
        return ""
    if "/" in d:
        try:
            mm, yy = d.split("/")
            mm = int(mm)
            yy = yy[-2:]
            full_yr = 2000 + int(yy)
            return f"{full_yr:04d}-{mm:02d}"
        except:
            return ""
    # legacy YYYY-MM / YYYY-MM-DD
    try:
        parts = d.split("-")
        if len(parts) >= 2:
            return f"{int(parts[0]):04d}-{int(parts[1]):02d}"
    except:
        pass
    return ""

# SQL expression: expiry_date column (MM/YY) ko 'YYYY-MM' sortable string mein convert karta hai
EXP_KEY_SQL = "('20' || substr(expiry_date,4,2) || '-' || substr(expiry_date,1,2))"

def exp_key_for_date(date_obj):
    """datetime.date object se 'YYYY-MM' key banao (EXP_KEY_SQL se compare karne ke liye)."""
    return f"{date_obj.year:04d}-{date_obj.month:02d}"


def today_str():
    return datetime.date.today().isoformat()

# ─── CALENDAR PICKER ──────────────────────────────────────────────────────────
def auto_titlecase(var):
    """StringVar pe trace lagao — har word ka pehla letter capital, baaki small.
    Space dene ke baad agla word bhi capital se start hoga."""
    _guard = {"on": False}
    def _fix(*a):
        if _guard["on"]: return
        val = var.get()
        new_val = val.title()
        if new_val != val:
            _guard["on"] = True
            try:
                var.set(new_val)
            finally:
                _guard["on"] = False
    var.trace_add("write", _fix)


class DatePicker:
    """Click karne pe popup calendar open hota hai. YYYY-MM-DD format mein value set karta hai."""

    def __init__(self, parent, textvariable, width=13, **kw):
        self.var = textvariable
        self._frame = tk.Frame(parent, bg=kw.get("bg", "white"))

        self._entry = ttk.Entry(self._frame, textvariable=textvariable, width=width)
        self._entry.pack(side="left")

        self._btn = tk.Button(
            self._frame, text="📅", font=("Segoe UI", 9),
            relief="flat", cursor="hand2", bd=0,
            bg=kw.get("bg","#EBF4FF"), fg="#2B6CB0",
            activebackground="#BEE3F8",
            command=self._open_cal
        )
        self._btn.pack(side="left", padx=(2,0))

        # ── Box pe focus jaate hi calendar auto-open ho jaaye ───────────────
        self._suppress_focus = False
        def _on_focus(ev):
            if self._suppress_focus:
                return
            self._entry.after(10, self._open_cal)
        self._entry.bind("<FocusIn>", _on_focus)

    def pack(self, **kw):        self._frame.pack(**kw)
    def grid(self, **kw):        self._frame.grid(**kw)
    def place(self, **kw):       self._frame.place(**kw)
    def pack_forget(self, **kw): self._frame.pack_forget()
    def grid_forget(self, **kw): self._frame.grid_forget()
    def place_forget(self,**kw): self._frame.place_forget()
    def bind(self, *a, **kw): return self._entry.bind(*a, **kw)
    def focus_set(self): self._entry.focus_set()
    def select_range(self, *a, **kw): return self._entry.select_range(*a, **kw)

    def _open_cal(self):
        # Parse current value
        try:
            cur = datetime.date.fromisoformat(self.var.get())
        except:
            cur = datetime.date.today()

        top = tk.Toplevel()
        top.title("Date Select Karo")
        top.resizable(False, False)
        top.grab_set()
        top.configure(bg="white")

        self._cal_year  = cur.year
        self._cal_month = cur.month
        self._top       = top

        # Pehle calendar build karo taaki actual size mil sake
        self._build_cal(top, cur)

        # Ab actual size ke saath position calculate karo — sabhi 4 sides check
        top.update_idletasks()
        cal_w = top.winfo_reqwidth()  or 230
        cal_h = top.winfo_reqheight() or 260
        scr_w = top.winfo_screenwidth()
        scr_h = top.winfo_screenheight()
        btn_x = self._btn.winfo_rootx()
        btn_y = self._btn.winfo_rooty()
        btn_h = self._btn.winfo_height() or 24

        # X: button ke neeche align, right ya left edge se cut na ho
        pop_x = btn_x - 100
        pop_x = max(4, min(pop_x, scr_w - cal_w - 4))

        # Y: niche try karo, jagah nahi toh upar
        if btn_y + btn_h + 4 + cal_h <= scr_h - 8:
            pop_y = btn_y + btn_h + 4
        else:
            pop_y = btn_y - cal_h - 4
        pop_y = max(4, min(pop_y, scr_h - cal_h - 4))

        top.geometry(f"+{pop_x}+{pop_y}")

    def _build_cal(self, top, selected=None):
        for w in top.winfo_children():
            w.destroy()

        y = self._cal_year
        m = self._cal_month

        # Header
        hdr = tk.Frame(top, bg="#1A365D"); hdr.pack(fill="x")
        tk.Button(hdr, text="◀", bg="#1A365D", fg="white", relief="flat",
                  font=("Segoe UI",11), cursor="hand2",
                  command=lambda: self._shift(-1, top)).pack(side="left", padx=6, pady=4)
        tk.Label(hdr, text=f"{MON[m]}  {y}", font=("Segoe UI",10,"bold"),
                 bg="#1A365D", fg="white").pack(side="left", expand=True)
        tk.Button(hdr, text="▶", bg="#1A365D", fg="white", relief="flat",
                  font=("Segoe UI",11), cursor="hand2",
                  command=lambda: self._shift(1, top)).pack(side="right", padx=6, pady=4)

        # Year jump
        yf = tk.Frame(top, bg="white"); yf.pack(fill="x", pady=2)
        tk.Label(yf, text="Year:", font=("Segoe UI",7), bg="white", fg="#718096").pack(side="left", padx=(8,2))
        yv = tk.StringVar(value=str(y))
        ye = ttk.Entry(yf, textvariable=yv, width=6, font=("Segoe UI",7))
        ye.pack(side="left")
        def jump_year(ev=None):
            try:
                self._cal_year = int(yv.get())
                self._build_cal(top)
            except: pass
        ye.bind("<Return>", jump_year)
        tk.Button(yf, text="Go", font=("Segoe UI",7), relief="flat", cursor="hand2",
                  bg="#EBF4FF", fg="#2B6CB0", command=jump_year).pack(side="left", padx=4)

        # Day names
        gf = tk.Frame(top, bg="white"); gf.pack(padx=6, pady=(0,2))
        days = ["Mo","Tu","We","Th","Fr","Sa","Su"]
        for i, d in enumerate(days):
            clr = "#E53E3E" if i >= 5 else "#4A5568"
            tk.Label(gf, text=d, width=3, font=("Segoe UI",8,"bold"),
                     bg="white", fg=clr).grid(row=0, column=i, padx=1)

        # Calendar grid
        import calendar
        cal = calendar.monthcalendar(y, m)
        today = datetime.date.today()
        try:    sel = datetime.date.fromisoformat(self.var.get())
        except: sel = None

        for r, week in enumerate(cal):
            for c, day in enumerate(week):
                if day == 0:
                    tk.Label(gf, text="", width=3, bg="white").grid(row=r+1, column=c, padx=1, pady=1)
                    continue
                d_obj = datetime.date(y, m, day)
                is_sel   = sel and d_obj == sel
                is_today = d_obj == today
                is_wknd  = c >= 5

                if is_sel:
                    bg, fg = "#2B6CB0", "white"
                elif is_today:
                    bg, fg = "#BEE3F8", "#1A365D"
                elif is_wknd:
                    bg, fg = "white", "#E53E3E"
                else:
                    bg, fg = "white", "#2D3748"

                btn = tk.Button(
                    gf, text=str(day), width=3,
                    font=("Segoe UI", 9, "bold" if is_today or is_sel else "normal"),
                    bg=bg, fg=fg, relief="flat", cursor="hand2", bd=0,
                    activebackground="#EBF4FF",
                    command=lambda dobj=d_obj: self._pick(dobj)
                )
                btn.grid(row=r+1, column=c, padx=1, pady=1)

        # Today button
        bf = tk.Frame(top, bg="white"); bf.pack(pady=(2,6))
        tk.Button(bf, text="Today", font=("Segoe UI",7), relief="flat",
                  bg="#EBF4FF", fg="#2B6CB0", cursor="hand2",
                  command=lambda: self._pick(datetime.date.today())
                  ).pack(side="left", padx=4)
        tk.Button(bf, text="Clear", font=("Segoe UI",7), relief="flat",
                  bg="#FFF5F5", fg=C_RED, cursor="hand2",
                  command=lambda: self._close_cal(clear=True)
                  ).pack(side="left", padx=4)
        tk.Button(bf, text="✕ Close", font=("Segoe UI",7), relief="flat",
                  bg="#F7FAFC", fg="#718096", cursor="hand2",
                  command=self._close_cal).pack(side="left", padx=4)

    def _close_cal(self, clear=False):
        if clear: self.var.set("")
        self._suppress_focus = True
        self._top.destroy()
        self._entry.after(150, lambda: setattr(self, "_suppress_focus", False))

    def _shift(self, delta, top):
        self._cal_month += delta
        if self._cal_month > 12: self._cal_month = 1;  self._cal_year += 1
        if self._cal_month < 1:  self._cal_month = 12; self._cal_year -= 1
        self._build_cal(top)

    def _pick(self, d_obj):
        self.var.set(d_obj.isoformat())
        self._suppress_focus = True
        self._top.destroy()
        self._entry.after(150, lambda: setattr(self, "_suppress_focus", False))


def make_date_entry(parent, textvariable, width=13, bg="white"):
    """DatePicker banao aur return karo — .pack() ya .grid() call karo."""
    return DatePicker(parent, textvariable, width=width, bg=bg)


def make_exp_mmyy_entry(parent, textvariable, width=8, font=("Segoe UI",9)):
    """Expiry (MM/YY) combobox banao. textvariable mein seedha 'MM/YY' format store hota hai.
    Return: ttk.Combobox widget — .pack()/.grid() call karo."""
    import re as _re
    def _key_handler(event):
        val = textvariable.get()
        if len(val) == 2 and val.isdigit() and event.char not in ("", "/", "\b"):
            textvariable.set(val + "/")
            event.widget.icursor(3)
            return "break"

    def _on_change(*a):
        raw = textvariable.get().strip()
        m = _re.match(r"^(\d{1,2})/(\d{2}|\d{4})$", raw)
        if m:
            mm, yy = m.group(1).zfill(2), m.group(2)[-2:].zfill(2)
            norm = f"{mm}/{yy}"
            if norm != raw and 1 <= int(mm) <= 12:
                textvariable.set(norm)

    cur_yr = datetime.date.today().year
    suggestions = [f"{str(m).zfill(2)}/{str(y)[2:]}"
                    for y in range(cur_yr, cur_yr+10)
                    for m in range(1,13)]
    cb = ttk.Combobox(parent, textvariable=textvariable, values=suggestions, font=font, width=width)
    cb.bind("<KeyPress>", _key_handler)
    cb.bind("<FocusOut>", _on_change)
    add_autocomplete(cb, lambda: suggestions)
    return cb



def next_sale_no():
    conn = get_db()
    # Sirf April 2026 (new FY 2026-27) ke baad ke bills count karo
    n = conn.execute(
        "SELECT COUNT(*) FROM sales WHERE bill_date >= '2026-04-01'"
    ).fetchone()[0] + 1
    conn.close()
    return f"INV-{n}"

def next_pur_no():
    conn = get_db()
    # Sirf April 2026 (new FY 2026-27) ke baad ke purchases count karo
    n = conn.execute(
        "SELECT COUNT(*) FROM purchases WHERE bill_date >= '2026-04-01'"
    ).fetchone()[0] + 1
    conn.close()
    return f"PUR-{n}"

def get_stock(product_name):
    conn = get_db()
    row = conn.execute("SELECT opening_stock FROM products WHERE name=?", (product_name,)).fetchone()
    opening = row[0] if row else 0
    sold   = conn.execute("SELECT COALESCE(SUM(qty),0) FROM sale_items WHERE product=?",     (product_name,)).fetchone()[0]
    bought = conn.execute("SELECT COALESCE(SUM(qty+free_qty),0) FROM purchase_items WHERE product=?", (product_name,)).fetchone()[0]
    # Manual stock movements
    man_in  = conn.execute("SELECT COALESCE(SUM(qty),0) FROM stock_movements WHERE product=? AND move_type='IN'",  (product_name,)).fetchone()[0]
    man_out = conn.execute("SELECT COALESCE(SUM(qty),0) FROM stock_movements WHERE product=? AND move_type='OUT'", (product_name,)).fetchone()[0]
    conn.close()
    return opening + bought + man_in - sold - man_out

# ─── FIFO HELPERS ─────────────────────────────────────────────────────────────

def fifo_add_layer(conn, product, purchase_date, purchase_bill, batch_no, qty, cost_per_unit):
    """Purchase hone pe ek naya FIFO layer add karo."""
    conn.execute(
        "INSERT INTO fifo_layers(product,purchase_date,purchase_bill,batch_no,"
        "qty_in,qty_remaining,cost_per_unit) VALUES(?,?,?,?,?,?,?)",
        (product, purchase_date, purchase_bill, batch_no, qty, qty, cost_per_unit)
    )

def fifo_consume(conn, product, qty_needed, sale_bill, sale_date):
    """Sale hone pe FIFO order mein layers consume karo. COGS return karo."""
    layers = conn.execute(
        "SELECT id, qty_remaining, cost_per_unit FROM fifo_layers "
        "WHERE product=? AND qty_remaining>0 ORDER BY purchase_date ASC, id ASC",
        (product,)
    ).fetchall()
    remaining = qty_needed
    total_cost = 0.0
    for layer in layers:
        if remaining <= 0:
            break
        lid, qty_avail, cpu = layer["id"], layer["qty_remaining"], layer["cost_per_unit"]
        use = min(remaining, qty_avail)
        cost = round(use * cpu, 4)
        conn.execute(
            "UPDATE fifo_layers SET qty_remaining=qty_remaining-? WHERE id=?",
            (use, lid)
        )
        conn.execute(
            "INSERT INTO fifo_consumption(layer_id,sale_bill,sale_date,product,"
            "qty_used,cost_per_unit,total_cost) VALUES(?,?,?,?,?,?,?)",
            (lid, sale_bill, sale_date, product, use, cpu, cost)
        )
        total_cost += cost
        remaining  -= use
    # Agar layers khatam ho gayi aur abhi bhi qty baaki hai (opening stock se)
    if remaining > 0:
        fallback_cost = conn.execute(
            "SELECT COALESCE(purchase_rate,0) FROM products WHERE name=?", (product,)
        ).fetchone()
        cpu = fallback_cost[0] if fallback_cost else 0
        cost = round(remaining * cpu, 4)
        conn.execute(
            "INSERT INTO fifo_consumption(layer_id,sale_bill,sale_date,product,"
            "qty_used,cost_per_unit,total_cost) VALUES(?,?,?,?,?,?,?)",
            (-1, sale_bill, sale_date, product, remaining, cpu, cost)
        )
        total_cost += cost
    return round(total_cost, 2)

def fifo_get_stock_value(conn):
    """Remaining FIFO layers ki cost — balance sheet ke liye."""
    row = conn.execute(
        "SELECT COALESCE(SUM(qty_remaining * cost_per_unit), 0) FROM fifo_layers"
    ).fetchone()
    return round(row[0], 2) if row else 0

def fifo_product_layers(conn, product):
    """Ek product ki saari remaining layers — FIFO report ke liye."""
    return [dict(r) for r in conn.execute(
        "SELECT purchase_date, purchase_bill, batch_no, qty_in, qty_remaining, cost_per_unit "
        "FROM fifo_layers WHERE product=? AND qty_remaining>0 ORDER BY purchase_date ASC, id ASC",
        (product,)
    ).fetchall()]

# ─── GROCERY HELPERS ─────────────────────────────────────────────────────────

def get_expiry_alerts():
    """Products jinki expiry 30 din mein hai ya ho gayi."""
    conn = get_db()
    soon_key = exp_key_for_date(datetime.date.today() + datetime.timedelta(days=30))
    rows = conn.execute(
        "SELECT product, batch_no, expiry_date, qty FROM expiry_stock "
        "WHERE qty > 0 AND " + EXP_KEY_SQL + " <= ? ORDER BY " + EXP_KEY_SQL, (soon_key,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_low_stock_alerts():
    """Products jinка current stock low_stock_alert se kam ya barabar ho."""
    conn = get_db()
    products = conn.execute(
        "SELECT name, opening_stock, unit, low_stock_alert FROM products "
        "WHERE low_stock_alert > 0 ORDER BY name"
    ).fetchall()
    result = []
    for pr in products:
        pname = pr["name"]
        bought  = conn.execute("SELECT COALESCE(SUM(qty+free_qty),0) FROM purchase_items WHERE product=?", (pname,)).fetchone()[0]
        sold    = conn.execute("SELECT COALESCE(SUM(qty),0) FROM sale_items    WHERE product=?", (pname,)).fetchone()[0]
        man_in  = conn.execute("SELECT COALESCE(SUM(qty),0) FROM stock_movements WHERE product=? AND move_type='IN'",  (pname,)).fetchone()[0]
        man_out = conn.execute("SELECT COALESCE(SUM(qty),0) FROM stock_movements WHERE product=? AND move_type='OUT'", (pname,)).fetchone()[0]
        current = (pr["opening_stock"] or 0) + bought + man_in - sold - man_out
        alert   = pr["low_stock_alert"] or 5
        if current <= alert:
            result.append({
                "name": pname,
                "stock": round(current, 2),
                "low_stock_alert": alert,
                "unit": pr["unit"] or ""
            })
    conn.close()
    result.sort(key=lambda x: x["stock"])
    return result

def generate_barcode_image(barcode_val, product_name="", mrp=0):
    """
    EAN13 ya Code128 barcode generate karo aur PIL Image return karo.
    barcode library required; fallback to text if not installed.
    """
    if not BARCODE_OK or not PIL_OK:
        return None
    try:
        buf = io.BytesIO()
        # Code128 — grocery ke liye best (alphanumeric support)
        bc = barcode.get('code128', barcode_val, writer=ImageWriter())
        options = dict(module_height=12, text_distance=2, font_size=8,
                       quiet_zone=2, write_text=True)
        bc.write(buf, options=options)
        buf.seek(0)
        img = Image.open(buf).convert("RGB")
        return img
    except Exception as e:
        return None

def scan_barcode_from_camera(callback):
    """
    Webcam se barcode scan karo. cv2 optional; agar nahi hai to manual entry dialog."""
    try:
        import cv2
        from pyzbar import pyzbar
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            raise RuntimeError("Camera nahi mili")

        win = tk.Toplevel()
        win.title("📷 Barcode Scan Karo — Camera")
        win.configure(bg="#1A1A2E")
        win.resizable(False, False)
        lbl_img = tk.Label(win, bg="#1A1A2E")
        lbl_img.pack(padx=10, pady=4)
        status = tk.Label(win, text="Camera ke saamne barcode rakho...",
                          font=("Segoe UI",9), bg="#1A1A2E", fg="#68D391")
        status.pack(pady=(0, 6))
        result = [None]

        def update_frame():
            ret, frame = cap.read()
            if not ret:
                win.after(50, update_frame)
                return
            codes = pyzbar.decode(frame)
            if codes:
                result[0] = codes[0].data.decode("utf-8")
                cap.release()
                cv2.destroyAllWindows()
                win.destroy()
                callback(result[0])
                return
            # Show frame
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            h, w = rgb.shape[:2]
            scale = min(480/w, 320/h)
            rgb = cv2.resize(rgb, (int(w*scale), int(h*scale)))
            photo = ImageTk.PhotoImage(Image.fromarray(rgb))
            lbl_img.config(image=photo)
            lbl_img.image = photo
            win.after(30, update_frame)

        def on_close():
            cap.release()
            win.destroy()
        win.protocol("WM_DELETE_WINDOW", on_close)
        win.after(30, update_frame)
    except Exception:
        # Fallback — manual entry
        _manual_barcode_entry(callback)

def _manual_barcode_entry(callback):
    """Camera nahi hai to manually barcode type karo."""
    dlg = tk.Toplevel()
    dlg.title("Barcode Enter Karo")
    dlg.configure(bg="#1A1A2E")
    dlg.resizable(False, False)
    w, h = 360, 200
    sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
    dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
    dlg.grab_set()

    tk.Label(dlg, text="🔍 Barcode / Code Daalo",
             font=("Segoe UI", 12, "bold"), bg="#1A1A2E", fg="white").pack(pady=(18, 4))
    tk.Label(dlg, text="Barcode scanner se scan karo ya manually type karo:",
             font=("Segoe UI", 9), bg="#1A1A2E", fg="#A0AEC0").pack()
    var = tk.StringVar()
    e = tk.Entry(dlg, textvariable=var, font=("Courier New", 14, "bold"),
                 bg="#2D3748", fg="#68D391", insertbackground="#68D391",
                 relief="flat", width=22, justify="center")
    e.pack(pady=10, ipady=6)
    e.focus_set()

    def confirm(ev=None):
        v = var.get().strip()
        if v:
            dlg.destroy()
            callback(v)
    e.bind("<Return>", confirm)
    tk.Button(dlg, text="✅ OK", font=("Segoe UI", 10, "bold"),
              bg="#276749", fg="white", relief="flat", cursor="hand2",
              command=confirm).pack()



def num_to_words(n):
    n = max(0, int(round(n)))
    ones = ["","One","Two","Three","Four","Five","Six","Seven","Eight","Nine",
            "Ten","Eleven","Twelve","Thirteen","Fourteen","Fifteen","Sixteen",
            "Seventeen","Eighteen","Nineteen"]
    tens = ["","","Twenty","Thirty","Forty","Fifty","Sixty","Seventy","Eighty","Ninety"]
    def w(n):
        if n == 0: return ""
        if n < 20: return ones[n]
        if n < 100: return tens[n//10]+(" "+ones[n%10] if n%10 else "")
        if n < 1000: return ones[n//100]+" Hundred"+(" "+w(n%100) if n%100 else "")
        if n < 100000: return w(n//1000)+" Thousand"+(" "+w(n%1000) if n%1000 else "")
        if n < 10000000: return w(n//100000)+" Lakh"+(" "+w(n%100000) if n%100000 else "")
        return w(n//10000000)+" Crore"+(" "+w(n%10000000) if n%10000000 else "")
    return ("INR "+w(n)+" Only") if n else "INR Zero Only"

# ─── EXCEL EXPORT HELPER ─────────────────────────────────────────────────────
def export_to_excel(headers, rows, default_filename="report"):
    """
    headers : list of column names
    rows    : list of lists (data rows)
    Saves as .xlsx using openpyxl.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        messagebox.showerror("Library Missing",
            "openpyxl install nahi hai.\n\n"
            "CMD mein ye command chalao:\n"
            "pip install openpyxl\n\nPhir dobara try karo.")
        return

    from tkinter import filedialog
    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel File","*.xlsx")],
        initialfile=default_filename,
        title="Excel file kahaan save karein?"
    )
    if not save_path:
        messagebox.showinfo("Cancel", "Save cancel kar diya.")
        return

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = default_filename[:30]

        # Header style
        hdr_fill  = PatternFill("solid", fgColor="1A365D")
        hdr_font  = Font(bold=True, color="FFFFFF", size=10)
        bd_side   = Side(style="thin", color="CCCCCC")
        bd        = Border(left=bd_side, right=bd_side, top=bd_side, bottom=bd_side)

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = bd

        # Data rows
        for ri, row in enumerate(rows, 2):
            fill_color = "F0F4F8" if ri % 2 == 0 else "FFFFFF"
            row_fill = PatternFill("solid", fgColor=fill_color)
            for ci, val in enumerate(row, 1):
                clean = val
                if isinstance(val, str):
                    clean = val.replace("₹","").replace(",","").strip()
                    try:
                        clean = float(clean) if "." in clean else int(clean)
                    except:
                        clean = val
                cell = ws.cell(row=ri, column=ci, value=clean)
                cell.fill   = row_fill
                cell.border = bd
                cell.alignment = Alignment(vertical="center")

        # Auto column width
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except: pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        ws.row_dimensions[1].height = 20
        wb.save(save_path)
        messagebox.showinfo("Excel Ready!", f"File save ho gayi:\n{save_path}")
        try:
            import os, sys
            if sys.platform == "win32": os.startfile(save_path)
        except: pass
    except Exception as e:
        messagebox.showerror("Excel Error", f"Excel nahi ban saki:\n{str(e)}")


# ─── EXCEL IMPORT HELPERS ────────────────────────────────────────────────────

def _download_excel_template(template_type):
    """
    Template Excel file download karo — user ko pata chale kaise bharna hai.
    template_type: 'products' | 'parties' | 'purchase'
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        messagebox.showerror("Library Missing",
            "openpyxl install nahi hai.\n\nCMD mein ye command chalao:\npip install openpyxl")
        return

    from tkinter import filedialog

    TEMPLATES = {
        "products": {
            "filename": "Products_Import_Template",
            "headers": ["Product Name*", "HSN Code", "Sale Rate*", "MRP", "Purchase Rate",
                        "GST %*", "Opening Stock", "Unit", "Barcode"],
            "sample": [
                ["Paracetamol 500mg", "3004", "12.50", "15.00", "8.00", "5", "100", "Strip", ""],
                ["Amoxicillin 250mg", "3004", "45.00", "55.00", "30.00", "12", "50", "Strip", ""],
                ["Vitamin C 500mg",   "3004", "25.00", "30.00", "18.00", "5",  "200", "Tablet", ""],
            ],
            "notes": [
                "INSTRUCTIONS:",
                "1. '*' wale columns zaroori hain",
                "2. GST % mein sirf number likhein: 0, 5, 12, 18, 28",
                "3. Unit examples: Strip, Tablet, Capsule, Pcs, Box, Bottle",
                "4. Row 1 header hai — DELETE MAT KARO",
                "5. Sample data (rows 2-4) hata dein ya upar se likhtey jayein",
            ]
        },
        "parties": {
            "filename": "Parties_Import_Template",
            "headers": ["Party Name*", "Type*", "Mobile", "GSTIN", "Address", "State", "Email"],
            "sample": [
                ["Apollo Pharmacy",    "Supplier", "9876543210", "09AAAAA0000A1Z5", "MG Road, Varanasi", "Uttar Pradesh", ""],
                ["Sharma Medical",     "Customer", "9123456789", "", "Civil Lines, Lucknow",   "Uttar Pradesh", ""],
                ["MedPlus Healthcare", "Both",     "8765432100", "09BBBBB1111B2Z6", "Hazratganj",     "Uttar Pradesh", ""],
            ],
            "notes": [
                "INSTRUCTIONS:",
                "1. Type column mein sirf ye likhein: Customer / Supplier / Both",
                "2. GSTIN 15 digit ka hona chahiye (optional)",
                "3. State ka pura naam likhein (e.g. Uttar Pradesh)",
                "4. Row 1 header hai — DELETE MAT KARO",
            ]
        },
        "purchase": {
            "filename": "Purchase_Import_Template",
            "headers": ["Product Name*", "Qty*", "Purchase Rate*", "GST %", "Batch No",
                        "Mfg Date (YYYY-MM-DD)", "Expiry Date (MM/YY)", "Supplier"],
            "sample": [
                ["Paracetamol 500mg", "100", "8.00",  "5",  "B001", "2024-01-01", "12/26", "Apollo Pharmacy"],
                ["Amoxicillin 250mg", "50",  "30.00", "12", "B002", "2024-03-01", "09/25", "MedPlus"],
            ],
            "notes": [
                "INSTRUCTIONS:",
                "1. Product Name pehle se database mein hona chahiye",
                "2. Date format: YYYY-MM-DD (e.g. 2025-12-31)",
                "3. GST % mein sirf number likhein",
                "4. Row 1 header hai — DELETE MAT KARO",
            ]
        }
    }

    if template_type not in TEMPLATES:
        return

    tmpl = TEMPLATES[template_type]
    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel File", "*.xlsx")],
        initialfile=tmpl["filename"],
        title="Template kahan save karein?"
    )
    if not save_path:
        return

    try:
        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Data"

        # Header row style
        hdr_fill = PatternFill("solid", fgColor="1A365D")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        bd       = Border(
            left=Side(style="thin", color="AAAAAA"),
            right=Side(style="thin", color="AAAAAA"),
            top=Side(style="thin", color="AAAAAA"),
            bottom=Side(style="thin", color="AAAAAA"))

        for ci, h in enumerate(tmpl["headers"], 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = bd

        # Sample rows
        for ri, row in enumerate(tmpl["sample"], 2):
            sample_fill = PatternFill("solid", fgColor="EBF8FF")
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill   = sample_fill
                cell.border = bd

        # Notes sheet
        ws_note = wb.create_sheet("Instructions")
        note_font = Font(size=10)
        for i, note in enumerate(tmpl["notes"], 1):
            ws_note.cell(row=i, column=1, value=note).font = note_font
        ws_note.column_dimensions["A"].width = 60

        # Auto column width
        for col in ws.columns:
            max_len = 0
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except: pass
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        ws.row_dimensions[1].height = 22

        wb.save(save_path)
        messagebox.showinfo("Template Ready!",
            f"Template save ho gayi:\n{save_path}\n\n"
            "Instructions sheet mein pura guide hai.\n"
            "Blue rows mein sample data hai — hata dein ya upar se likhein.")
        try:
            import os, sys
            if sys.platform == "win32":
                os.startfile(save_path)
        except: pass
    except Exception as e:
        messagebox.showerror("Error", f"Template nahi bani:\n{str(e)}")


def import_products_from_excel(refresh_callback=None):
    """
    Excel se products import karo.
    Columns: Product Name | HSN | Sale Rate | MRP | Purchase Rate | GST % | Opening Stock | Unit | Barcode
    """
    try:
        import openpyxl
    except ImportError:
        messagebox.showerror("Library Missing",
            "openpyxl install nahi hai.\n\nCMD mein ye command chalao:\npip install openpyxl")
        return

    from tkinter import filedialog
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel File", "*.xlsx *.xls"), ("All Files", "*.*")],
        title="Products Excel File Select Karo"
    )
    if not file_path:
        return

    try:
        # Try openpyxl first, fallback to pandas for .xls
        if file_path.lower().endswith(".xls"):
            import pandas as pd
            df = pd.read_excel(file_path)
            rows_raw = [list(df.columns)] + df.values.tolist()
        else:
            wb   = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws   = wb.active
            rows_raw = [[cell.value for cell in row] for row in ws.iter_rows()]
            wb.close()
    except Exception as e:
        messagebox.showerror("File Error", f"Excel file nahi khul saki:\n{str(e)}")
        return

    if not rows_raw or len(rows_raw) < 2:
        messagebox.showinfo("Khaali File", "File mein koi data nahi mila!")
        return

    # Preview window
    _show_import_preview(
        rows_raw=rows_raw,
        import_type="products",
        refresh_callback=refresh_callback
    )


def import_parties_from_excel(refresh_callback=None):
    """
    Excel se parties (customers/suppliers) import karo.
    Columns: Party Name | Type | Mobile | GSTIN | Address | State | Email
    """
    try:
        import openpyxl
    except ImportError:
        messagebox.showerror("Library Missing",
            "openpyxl install nahi hai.\n\nCMD mein ye command chalao:\npip install openpyxl")
        return

    from tkinter import filedialog
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel File", "*.xlsx *.xls"), ("All Files", "*.*")],
        title="Parties Excel File Select Karo"
    )
    if not file_path:
        return

    try:
        if file_path.lower().endswith(".xls"):
            import pandas as pd
            df = pd.read_excel(file_path)
            rows_raw = [list(df.columns)] + df.values.tolist()
        else:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            rows_raw = [[cell.value for cell in row] for row in ws.iter_rows()]
            wb.close()
    except Exception as e:
        messagebox.showerror("File Error", f"Excel file nahi khul saki:\n{str(e)}")
        return

    if not rows_raw or len(rows_raw) < 2:
        messagebox.showinfo("Khaali File", "File mein koi data nahi mila!")
        return

    _show_import_preview(
        rows_raw=rows_raw,
        import_type="parties",
        refresh_callback=refresh_callback
    )


def import_purchase_from_excel(refresh_callback=None):
    """
    Excel se purchase/stock import karo.
    Columns: Product | Qty | Rate | GST% | Batch No | Mfg Date | Expiry Date | Supplier
    """
    try:
        import openpyxl
    except ImportError:
        messagebox.showerror("Library Missing",
            "openpyxl install nahi hai.\n\nCMD mein ye command chalao:\npip install openpyxl")
        return

    from tkinter import filedialog
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel File", "*.xlsx *.xls"), ("All Files", "*.*")],
        title="Purchase/Stock Excel File Select Karo"
    )
    if not file_path:
        return

    try:
        if file_path.lower().endswith(".xls"):
            import pandas as pd
            df = pd.read_excel(file_path)
            rows_raw = [list(df.columns)] + df.values.tolist()
        else:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            rows_raw = [[cell.value for cell in row] for row in ws.iter_rows()]
            wb.close()
    except Exception as e:
        messagebox.showerror("File Error", f"Excel file nahi khul saci:\n{str(e)}")
        return

    if not rows_raw or len(rows_raw) < 2:
        messagebox.showinfo("Khaali File", "File mein koi data nahi mila!")
        return

    _show_import_preview(
        rows_raw=rows_raw,
        import_type="purchase",
        refresh_callback=refresh_callback
    )


def _show_import_preview(rows_raw, import_type, refresh_callback=None):
    """
    Import data ka preview dikhao — user confirm kare tab save karo.
    """
    import tkinter as _tk
    from tkinter import ttk as _ttk

    TYPE_CONFIG = {
        "products": {
            "title":   "📦 Products Import Preview",
            "color":   "#276749",
            "cols":    ["Product Name", "HSN", "Sale Rate", "MRP", "Purchase Rate",
                        "GST %", "Opening Stock", "Unit", "Barcode"],
            "widths":  [22, 8, 10, 10, 12, 6, 12, 7, 12],
        },
        "parties": {
            "title":   "👥 Parties Import Preview",
            "color":   "#2B6CB0",
            "cols":    ["Party Name", "Type", "Mobile", "GSTIN", "Address", "State", "Email"],
            "widths":  [22, 10, 13, 18, 22, 16, 20],
        },
        "purchase": {
            "title":   "🛒 Purchase/Stock Import Preview",
            "color":   "#C05621",
            "cols":    ["Product", "Qty", "Rate", "GST%", "Batch No", "Mfg Date", "Expiry Date", "Supplier"],
            "widths":  [22, 7, 10, 6, 10, 13, 13, 16],
        },
    }

    cfg = TYPE_CONFIG.get(import_type, TYPE_CONFIG["products"])

    # Skip header row — row 1 = header
    data_rows = rows_raw[1:]
    # Filter empty rows
    data_rows = [r for r in data_rows
                 if any(str(v).strip() not in ("", "None", "nan") for v in r if v is not None)]

    if not data_rows:
        messagebox.showinfo("Khaali Data", "Excel mein koi valid data nahi mila (header ke baad)!")
        return

    # Window
    win = _tk.Toplevel()
    win.title(cfg["title"])
    win.configure(bg="#FFFFFF")
    win.state("zoomed")
    win.grab_set()

    # Header bar
    hdr = _tk.Frame(win, bg=cfg["color"], pady=4)
    hdr.pack(fill="x")
    _tk.Label(hdr, text=cfg["title"], font=("Segoe UI", 12, "bold"),
              bg=cfg["color"], fg="white").pack(side="left", padx=16)
    _tk.Label(hdr, text=f"{len(data_rows)} rows mili — check karke Import dabao",
              font=("Segoe UI",9), bg=cfg["color"], fg="#CCE5FF").pack(side="left", padx=8)

    # Info bar
    info_f = _tk.Frame(win, bg="#EBF8FF", pady=6)
    info_f.pack(fill="x")
    _tk.Label(info_f,
              text="✅ = Import hoga  |  ⚠️ = Warning (phir bhi import hoga)  |  ❌ = Skip hoga (galat data)",
              font=("Segoe UI", 9), bg="#EBF8FF", fg="#2B6CB0", padx=12).pack(side="left")

    # Stats frame
    stats_f = _tk.Frame(win, bg="#FFFFFF")
    stats_f.pack(fill="x", padx=12, pady=(6, 2))

    stats_lbl = _tk.Label(stats_f, text="Data check ho raha hai...",
                           font=("Segoe UI", 9), bg="#FFFFFF", fg="#4A5568")
    stats_lbl.pack(side="left")

    # Scrollable table
    outer = _tk.Frame(win, bg="#FFFFFF")
    outer.pack(fill="both", expand=True, padx=12, pady=(4, 6))

    canvas = _tk.Canvas(outer, bg="#FFFFFF", highlightthickness=0)
    vsb = _ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
    hsb = _ttk.Scrollbar(outer, orient="horizontal", command=canvas.xview)
    canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    vsb.pack(side="right", fill="y")
    hsb.pack(side="bottom", fill="x")
    canvas.pack(side="left", fill="both", expand=True)

    inner = _tk.Frame(canvas, bg="#FFFFFF")
    wid = canvas.create_window((0, 0), window=inner, anchor="nw")
    inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>",
        lambda ev: canvas.yview_scroll(-1*(ev.delta//120), "units")))
    canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

    # Table header
    STATUS_COL  = "Status"
    REASON_COL  = "Reason"
    all_cols = [STATUS_COL] + cfg["cols"] + [REASON_COL]
    all_widths = [8] + cfg["widths"] + [24]

    THEAD_BG = "#DBEAFE"
    for ci, (col, w) in enumerate(zip(all_cols, all_widths)):
        _tk.Label(inner, text=col, font=("Segoe UI", 9, "bold"),
                  bg=THEAD_BG, fg="#1A365D", width=w, anchor="w",
                  padx=8, pady=5).grid(row=0, column=ci, sticky="nsew", padx=1, pady=1)
        inner.columnconfigure(ci, minsize=w*7)

    # Validate and build rows
    validated = []
    for row in data_rows:
        # Pad row to expected length
        ncols = len(cfg["cols"])
        padded = list(row) + [""] * ncols
        padded = padded[:ncols]
        # Clean values
        cleaned = [str(v).strip() if v is not None else "" for v in padded]
        # Remove "None"/"nan" strings
        cleaned = ["" if v in ("None", "nan", "NaN") else v for v in cleaned]

        status = "✅"
        reason = ""

        if import_type == "products":
            name      = cleaned[0]
            sale_rate = cleaned[2]
            gst_pct   = cleaned[5]
            if not name:
                status = "❌"; reason = "Product naam khaali hai"
            elif not sale_rate:
                status = "❌"; reason = "Sale Rate khaali hai"
            else:
                try: float(sale_rate)
                except: status = "❌"; reason = f"Sale Rate '{sale_rate}' number nahi hai"
                try:
                    gp = float(gst_pct) if gst_pct else 0
                    if gp not in (0, 5, 12, 18, 28):
                        status = "⚠️"; reason = f"GST {gp}% unusual hai (0/5/12/18/28 chahiye)"
                except:
                    if gst_pct: status = "⚠️"; reason = f"GST '{gst_pct}' clear nahi"

        elif import_type == "parties":
            name  = cleaned[0]
            ptype = cleaned[1].strip().title() if cleaned[1] else "Customer"
            if not name:
                status = "❌"; reason = "Party naam khaali hai"
            elif ptype not in ("Customer", "Supplier", "Both"):
                status = "⚠️"; reason = f"Type '{ptype}' unknown — 'Customer' use hoga"

        elif import_type == "purchase":
            prod = cleaned[0]
            qty  = cleaned[1]
            rate = cleaned[2]
            if not prod:
                status = "❌"; reason = "Product naam khaali hai"
            elif not qty or not rate:
                status = "❌"; reason = "Qty ya Rate khaali hai"
            else:
                try: float(qty)
                except: status = "❌"; reason = f"Qty '{qty}' number nahi"
                try: float(rate)
                except: status = "❌"; reason = f"Rate '{rate}' number nahi"

        validated.append((status, cleaned, reason))

    # Fill table
    ok_count   = sum(1 for s,_,_ in validated if s == "✅")
    warn_count = sum(1 for s,_,_ in validated if s == "⚠️")
    skip_count = sum(1 for s,_,_ in validated if s == "❌")

    stats_lbl.config(
        text=f"Total: {len(validated)}  |  ✅ Import: {ok_count+warn_count}  |  ❌ Skip: {skip_count}",
        fg="#276749" if skip_count == 0 else "#C05621"
    )

    STATUS_COLORS = {"✅": "#F0FFF4", "⚠️": "#FFFBEB", "❌": "#FFF5F5"}
    STATUS_FGS    = {"✅": "#276749", "⚠️": "#975A16", "❌": "#9B2C2C"}

    for ri, (status, cleaned, reason) in enumerate(validated, 1):
        bg = STATUS_COLORS.get(status, "#FFFFFF")
        fg = STATUS_FGS.get(status, "#4A5568")
        row_vals = [status] + cleaned + [reason]
        for ci, val in enumerate(row_vals):
            cell_fg = fg if ci in (0, len(row_vals)-1) else "#4A5568"
            _tk.Label(inner, text=str(val)[:60], font=("Segoe UI", 9),
                      bg=bg, fg=cell_fg, anchor="w", padx=6, pady=4
                      ).grid(row=ri, column=ci, sticky="nsew", padx=1, pady=0)

    # Bottom buttons
    bf = _tk.Frame(win, bg="#FFFFFF")
    bf.pack(pady=(4, 12))

    def do_import():
        conn = get_db()
        success = 0; skipped = 0; errors = []

        for status, cleaned, reason in validated:
            if status == "❌":
                skipped += 1
                continue

            try:
                if import_type == "products":
                    name       = cleaned[0].strip()
                    hsn        = cleaned[1].strip()
                    sale_rate  = float(cleaned[2]) if cleaned[2] else 0
                    mrp        = float(cleaned[3]) if cleaned[3] else 0
                    pur_rate   = float(cleaned[4]) if cleaned[4] else 0
                    gst_pct    = float(cleaned[5]) if cleaned[5] else 0
                    opening    = float(cleaned[6]) if cleaned[6] else 0
                    unit       = cleaned[7].strip() if cleaned[7] else "Pcs"
                    barcode    = cleaned[8].strip() if len(cleaned) > 8 else ""
                    # Insert or update
                    existing = conn.execute(
                        "SELECT id FROM products WHERE name=?", (name,)).fetchone()
                    if existing:
                        conn.execute(
                            "UPDATE products SET hsn=?,sale_rate=?,mrp=?,purchase_rate=?,"
                            "gst_percent=?,opening_stock=?,unit=?,barcode=? WHERE name=?",
                            (hsn, sale_rate, mrp, pur_rate, gst_pct, opening, unit, barcode, name))
                    else:
                        conn.execute(
                            "INSERT INTO products(name,hsn,sale_rate,mrp,purchase_rate,"
                            "gst_percent,opening_stock,unit,barcode) VALUES(?,?,?,?,?,?,?,?,?)",
                            (name, hsn, sale_rate, mrp, pur_rate, gst_pct, opening, unit, barcode))
                    success += 1

                elif import_type == "parties":
                    name   = cleaned[0].strip()
                    ptype  = cleaned[1].strip().title() if cleaned[1] else "Customer"
                    if ptype not in ("Customer", "Supplier", "Both"):
                        ptype = "Customer"
                    mobile = cleaned[2].strip() if cleaned[2] else ""
                    gstin  = cleaned[3].strip().upper() if cleaned[3] else ""
                    addr   = cleaned[4].strip() if cleaned[4] else ""
                    state  = cleaned[5].strip() if cleaned[5] else "Uttar Pradesh"
                    email  = cleaned[6].strip() if len(cleaned) > 6 and cleaned[6] else ""
                    existing = conn.execute(
                        "SELECT id FROM parties WHERE name=?", (name,)).fetchone()
                    if existing:
                        conn.execute(
                            "UPDATE parties SET ptype=?,mobile=?,gstin=?,address=?,state=?,email=? WHERE name=?",
                            (ptype, mobile, gstin, addr, state, email, name))
                    else:
                        conn.execute(
                            "INSERT INTO parties(name,ptype,mobile,gstin,address,state,email) VALUES(?,?,?,?,?,?,?)",
                            (name, ptype, mobile, gstin, addr, state, email))
                    success += 1

                elif import_type == "purchase":
                    prod     = cleaned[0].strip()
                    qty      = float(cleaned[1]) if cleaned[1] else 0
                    rate     = float(cleaned[2]) if cleaned[2] else 0
                    gst_pct  = float(cleaned[3]) if cleaned[3] else 0
                    batch_no = cleaned[4].strip() if cleaned[4] else ""
                    mfg_date = cleaned[5].strip() if cleaned[5] else ""
                    exp_date = exp_to_storage(cleaned[6].strip()) if cleaned[6] else ""
                    supplier = cleaned[7].strip() if len(cleaned) > 7 and cleaned[7] else ""

                    taxable = round(qty * rate, 2)
                    gst_amt = round(taxable * gst_pct / 100, 2)
                    total   = round(taxable + gst_amt, 2)

                    import datetime as _dt2
                    bill_no = f"IMP-{_dt2.date.today().strftime('%Y%m%d')}-{success+1:04d}"

                    # Check if purchase bill exists for today's batch import
                    existing_pur = conn.execute(
                        "SELECT id FROM purchases WHERE bill_no=?", (bill_no,)).fetchone()
                    if not existing_pur:
                        conn.execute(
                            "INSERT INTO purchases(bill_no,bill_date,party,grand_total,pay_mode) VALUES(?,?,?,?,?)",
                            (bill_no, _dt2.date.today().isoformat(), supplier, total, "Cash"))

                    pur_id = conn.execute(
                        "SELECT id FROM purchases WHERE bill_no=?", (bill_no,)).fetchone()[0]

                    conn.execute(
                        "INSERT INTO purchase_items(purchase_id,product,qty,rate,taxable,gst_percent,gst_amt,total,"
                        "batch_no,mfg_date,expiry_date) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                        (pur_id, prod, qty, rate, taxable, gst_pct, gst_amt, total,
                         batch_no, mfg_date, exp_date))

                    # FIFO layer add karo
                    fifo_add_layer(conn, prod, _dt2.date.today().isoformat(),
                                   bill_no, batch_no, qty, rate)

                    # Expiry stock bhi add karo agar expiry date diya (MM/YY format)
                    if exp_date:
                        try:
                            exp_norm = exp_to_storage(exp_date)
                            import re as _re_imp2
                            if _re_imp2.match(r"^(0[1-9]|1[0-2])/\d{2}$", exp_norm):
                                conn.execute(
                                    "INSERT INTO expiry_stock(product,batch_no,mfg_date,expiry_date,qty,purchase_rate,supplier) "
                                    "VALUES(?,?,?,?,?,?,?)",
                                    (prod, batch_no, mfg_date, exp_norm, qty, rate, supplier))
                        except: pass

                    success += 1

            except Exception as e:
                errors.append(f"Row {cleaned[0]}: {str(e)[:60]}")

        conn.commit()
        conn.close()

        # Result message
        msg = f"Import complete!\n\n✅ Imported: {success}\n❌ Skipped: {skipped}"
        if errors:
            msg += f"\n\n⚠️ Errors ({len(errors)}):\n" + "\n".join(errors[:5])
            if len(errors) > 5:
                msg += f"\n... aur {len(errors)-5} errors"

        messagebox.showinfo("Import Done!", msg)
        win.destroy()
        if refresh_callback:
            refresh_callback()

    make_btn(bf, f"✅  Import Karo ({ok_count+warn_count} rows)", do_import, bg=cfg["color"]).pack(side="left", padx=8)
    make_btn(bf, "❌  Cancel", win.destroy, bg="#4A5568").pack(side="left", padx=4)


# ─── THEME ───────────────────────────────────────────────────────────────────
C_BG      = "#F0F4F8"
C_WHITE   = "#FFFFFF"
C_SIDEBAR = "#1A365D"
C_NAV_ACT = "#2B6CB0"
C_NAV_FG  = "#A0C4E8"
C_TOP     = "#185FA5"
C_ACCENT  = "#185FA5"
C_BORDER  = "#CBD5E0"
C_GRAY    = "#4A5568"
C_DARK    = "#2D3748"
C_LGRAY   = "#718096"
C_GREEN   = "#276749"
C_RED     = "#9B2C2C"
C_AMBER   = "#975A16"
C_PURPLE  = "#553C9A"
C_LIGHT   = "#EBF4FF"
C_THEAD   = "#DBEAFE"
HDR_BG    = "#F0F0F0"

def style_setup():
    s = ttk.Style()
    s.theme_use("clam")
    s.configure("TEntry", padding=2, relief="flat", fieldbackground=C_WHITE)
    s.configure("TCombobox", padding=2, relief="flat", fieldbackground=C_WHITE)
    s.configure("Vertical.TScrollbar", background=C_BORDER, troughcolor=C_BG, relief="flat")
    s.map("TCombobox", fieldbackground=[("readonly", C_WHITE)])

# ─── REUSABLE WIDGETS ────────────────────────────────────────────────────────
def make_btn(parent, text, cmd, bg=C_ACCENT, fg=C_WHITE, font_size=9, padx=8, pady=2):
    b = tk.Button(parent, text=text, command=cmd, bg=bg, fg=fg,
                  font=("Segoe UI", font_size, "bold"), relief="flat",
                  activebackground="#0C447C", activeforeground=C_WHITE,
                  cursor="hand2", padx=padx, pady=pady, bd=0)
    return b

def make_label(parent, text, font_size=10, bold=False, color=C_GRAY, bg=C_BG, anchor="w"):
    wt = "bold" if bold else "normal"
    return tk.Label(parent, text=text, font=("Segoe UI", font_size, wt),
                    fg=color, bg=bg, anchor=anchor)

def labeled_entry(parent, label, row, col, width=18, default="", colspan=1, bg=C_BG):
    make_label(parent, label, 9, bg=bg).grid(row=row*2, column=col, sticky="w",
                                               padx=(6,2), pady=(2,0), columnspan=colspan)
    var = tk.StringVar(value=default)
    e = ttk.Entry(parent, textvariable=var, width=width, font=("Segoe UI",9))
    e.grid(row=row*2+1, column=col, sticky="ew", padx=(6,2), pady=(0,1), columnspan=colspan)
    return var, e

def labeled_combo(parent, label, row, col, values, width=18, bg=C_BG):
    make_label(parent, label, 9, bg=bg).grid(row=row*2, column=col, sticky="w", padx=(6,2), pady=(2,0))
    var = tk.StringVar()
    cb = ttk.Combobox(parent, textvariable=var, values=values, width=width,
                      font=("Segoe UI",9), state="readonly")
    cb.grid(row=row*2+1, column=col, sticky="ew", padx=(6,2), pady=(0,4))
    return var, cb

def scrollable_frame(parent):
    outer = tk.Frame(parent, bg=C_WHITE, highlightthickness=1, highlightbackground=C_BORDER)

    # Horizontal scrollbar at bottom
    hsb = ttk.Scrollbar(outer, orient="horizontal")
    hsb.pack(side="bottom", fill="x")

    # Vertical scrollbar on right
    vsb = ttk.Scrollbar(outer, orient="vertical")
    vsb.pack(side="right", fill="y")

    canvas = tk.Canvas(outer, bg=C_WHITE, highlightthickness=0,
                       yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    canvas.pack(side="left", fill="both", expand=True)

    vsb.config(command=canvas.yview)
    hsb.config(command=canvas.xview)

    inner = tk.Frame(canvas, bg=C_WHITE)
    win_id = canvas.create_window((0,0), window=inner, anchor="nw")

    def on_inner_resize(e):
        canvas.configure(scrollregion=canvas.bbox("all"))

    def on_enter(e):
        canvas.bind_all("<MouseWheel>",
            lambda ev: canvas.yview_scroll(-1*(ev.delta//120), "units"))
        # Shift+scroll = horizontal
        canvas.bind_all("<Shift-MouseWheel>",
            lambda ev: canvas.xview_scroll(-1*(ev.delta//120), "units"))

    def on_leave(e):
        canvas.unbind_all("<MouseWheel>")
        canvas.unbind_all("<Shift-MouseWheel>")

    inner.bind("<Configure>", on_inner_resize)
    canvas.bind("<Enter>", on_enter)
    canvas.bind("<Leave>", on_leave)

    return outer, inner


def make_table(parent, cols, col_weights=None):
    outer, inner = scrollable_frame(parent)
    outer.pack(fill="both", expand=True, pady=2)
    for i, col in enumerate(cols):
        # col_weights are character widths — convert to pixel minwidth
        w = col_weights[i] if col_weights else 12
        lbl = tk.Label(inner, text=col, font=("Segoe UI", 9, "bold"),
                       bg=C_THEAD, fg=C_ACCENT, width=w, anchor="w",
                       padx=6, pady=2)
        lbl.grid(row=0, column=i, sticky="nsew", padx=1, pady=1)
        inner.columnconfigure(i, minsize=w*7)  # ~7px per char
    return inner

def table_row(inner, vals, row_n, fgs=None, bg=None):
    if bg is None:
        bg = C_WHITE if row_n % 2 == 0 else "#F7FAFC"
    for i, val in enumerate(vals):
        fg = (fgs[i] if fgs and fgs[i] else C_GRAY)
        tk.Label(inner, text=str(val), font=("Segoe UI", 9),
                 bg=bg, fg=fg, anchor="w", padx=6, pady=2
                 ).grid(row=row_n, column=i, sticky="nsew", padx=1, pady=0)

def clear_table_rows(inner):
    for w in inner.winfo_children():
        try:
            info = w.grid_info()
            if info and int(info.get("row", 0)) > 0:
                w.destroy()
        except Exception:
            pass

def add_autocomplete(cb, get_values_fn=None):
    """
    Reliable autocomplete for ttk.Combobox.
    Type karo → filtered list ek popup Listbox me dikhti hai.
    Click ya Enter se select hota hai — original combobox me value set hoti hai.
    """
    _all_vals = []
    _popup    = {"win": None, "lb": None}

    def _get_all():
        if get_values_fn:
            try:
                return [str(v) for v in get_values_fn()]
            except Exception:
                return []
        return [str(v) for v in (cb["values"] or [])]

    def _close_popup():
        try:
            if _popup["win"] and _popup["win"].winfo_exists():
                _popup["win"].destroy()
        except Exception:
            pass
        _popup["win"] = None
        _popup["lb"]  = None

    def _show_popup(filtered):
        _close_popup()
        if not filtered:
            return

        # Position popup just below the combobox
        cb.update_idletasks()
        x = cb.winfo_rootx()
        y = cb.winfo_rooty() + cb.winfo_height()
        w = max(cb.winfo_width(), 200)
        h = min(len(filtered) * 22 + 6, 220)

        win = tk.Toplevel(cb)
        win.wm_overrideredirect(True)
        win.geometry(f"{w}x{h}+{x}+{y}")
        win.attributes("-topmost", True)
        _popup["win"] = win

        frame = tk.Frame(win, bg="#FFFFFF",
                         highlightthickness=1, highlightbackground="#CBD5E0")
        frame.pack(fill="both", expand=True)

        sb = tk.Scrollbar(frame, orient="vertical")
        lb = tk.Listbox(frame, yscrollcommand=sb.set,
                        font=("Segoe UI",9),
                        bg="#FFFFFF", fg="#2D3748",
                        selectbackground="#3182CE",
                        selectforeground="#FFFFFF",
                        activestyle="none",
                        relief="flat", bd=0,
                        highlightthickness=0)
        sb.config(command=lb.yview)
        sb.pack(side="right", fill="y")
        lb.pack(side="left", fill="both", expand=True)
        _popup["lb"] = lb

        for item in filtered:
            lb.insert("end", item)

        def _select(event=None):
            try:
                idx = lb.curselection()
                if idx:
                    val = lb.get(idx[0])
                    cb.set(val)
                    cb.event_generate("<<ComboboxSelected>>")
            except Exception:
                pass
            _close_popup()
            cb.focus_set()

        def _hover(event):
            lb.selection_clear(0, "end")
            idx = lb.nearest(event.y)
            lb.selection_set(idx)

        lb.bind("<ButtonRelease-1>", _select)
        lb.bind("<Return>",          _select)
        lb.bind("<Motion>",          _hover)

        # Close popup if user clicks elsewhere
        win.bind("<FocusOut>", lambda e: cb.after(20, _maybe_close))

    def _maybe_close():
        try:
            fw = cb.focus_get()
            if _popup["lb"] and fw == _popup["lb"]:
                return
            _close_popup()
        except Exception:
            _close_popup()

    def _filter_and_show(*args):
        typed  = cb.get().strip().lower()
        all_v  = _get_all()
        if not typed:
            filtered = all_v
        else:
            filtered = [v for v in all_v if typed in v.lower()]
        _show_popup(filtered)

    def _on_key(event):
        if event.keysym in ("Escape",):
            _close_popup()
            return
        if event.keysym == "Down":
            if _popup["lb"]:
                _popup["lb"].focus_set()
                _popup["lb"].selection_set(0)
            return
        if event.keysym in ("Return", "Tab"):
            _close_popup()
            return
        cb.after(80, _filter_and_show)

    cb.bind("<KeyRelease>",  _on_key)
    cb.bind("<FocusIn>",     lambda e: cb.after(100, _filter_and_show))
    cb.bind("<FocusOut>",    lambda e: cb.after(30, _maybe_close))
    # Disable built-in dropdown to avoid conflict
    cb.configure(state="normal")


def add_entry_autocomplete(entry_widget, get_values_fn):
    """
    ttk.Entry widget pe autocomplete suggestion popup lagao.
    Type karo → filtered list popup mein dikhti hai.
    Click ya Enter/Down se select hota hai.
    get_values_fn: callable jo list of strings return kare.
    """
    _popup = {"win": None, "lb": None}

    def _close_popup():
        try:
            if _popup["win"] and _popup["win"].winfo_exists():
                _popup["win"].destroy()
        except Exception:
            pass
        _popup["win"] = None
        _popup["lb"]  = None

    def _show_popup(filtered):
        _close_popup()
        if not filtered:
            return
        entry_widget.update_idletasks()
        x = entry_widget.winfo_rootx()
        y = entry_widget.winfo_rooty() + entry_widget.winfo_height()
        w = max(entry_widget.winfo_width(), 200)
        h = min(len(filtered) * 22 + 6, 200)

        win = tk.Toplevel(entry_widget)
        win.wm_overrideredirect(True)
        win.geometry(f"{w}x{h}+{x}+{y}")
        win.attributes("-topmost", True)
        _popup["win"] = win

        frame = tk.Frame(win, bg="#FFFFFF",
                         highlightthickness=1, highlightbackground="#CBD5E0")
        frame.pack(fill="both", expand=True)

        sb = tk.Scrollbar(frame, orient="vertical")
        lb = tk.Listbox(frame, yscrollcommand=sb.set,
                        font=("Segoe UI",9),
                        bg="#FFFFFF", fg="#2D3748",
                        selectbackground="#3182CE",
                        selectforeground="#FFFFFF",
                        activestyle="none",
                        relief="flat", bd=0,
                        highlightthickness=0)
        sb.config(command=lb.yview)
        sb.pack(side="right", fill="y")
        lb.pack(side="left", fill="both", expand=True)
        _popup["lb"] = lb

        for item in filtered:
            lb.insert("end", item)

        def _select(event=None):
            try:
                idx = lb.curselection()
                if idx:
                    val = lb.get(idx[0])
                    entry_widget.delete(0, "end")
                    entry_widget.insert(0, val)
            except Exception:
                pass
            _close_popup()
            entry_widget.focus_set()

        def _hover(event):
            lb.selection_clear(0, "end")
            idx = lb.nearest(event.y)
            lb.selection_set(idx)

        lb.bind("<ButtonRelease-1>", _select)
        lb.bind("<Return>",          _select)
        lb.bind("<Motion>",          _hover)
        win.bind("<FocusOut>", lambda e: entry_widget.after(20, _maybe_close))

    def _maybe_close():
        try:
            fw = entry_widget.focus_get()
            if _popup["lb"] and fw == _popup["lb"]:
                return
            _close_popup()
        except Exception:
            _close_popup()

    def _filter_and_show(*args):
        typed = entry_widget.get().strip().lower()
        try:
            all_v = [str(v) for v in get_values_fn() if v]
        except Exception:
            all_v = []
        # Duplicates hata do
        seen = set(); unique = []
        for v in all_v:
            if v not in seen: seen.add(v); unique.append(v)
        if not typed:
            filtered = unique
        else:
            filtered = [v for v in unique if typed in v.lower()]
        _show_popup(filtered)

    def _on_key(event):
        if event.keysym == "Escape":
            _close_popup(); return
        if event.keysym == "Down":
            if _popup["lb"]:
                _popup["lb"].focus_set()
                _popup["lb"].selection_set(0)
            return
        if event.keysym in ("Return", "Tab"):
            _close_popup(); return
        entry_widget.after(80, _filter_and_show)

    entry_widget.bind("<KeyRelease>", _on_key)
    entry_widget.bind("<FocusIn>",    lambda e: entry_widget.after(100, _filter_and_show))
    entry_widget.bind("<FocusOut>",   lambda e: entry_widget.after(30, _maybe_close))


def stat_card(parent, title, value, color=C_ACCENT):
    f = tk.Frame(parent, bg=C_WHITE, highlightthickness=1, highlightbackground=C_BORDER)
    f.pack(side="left", fill="both", expand=True, padx=4)
    tk.Label(f, text=title, font=("Segoe UI", 9), bg=C_WHITE, fg=C_LGRAY).pack(anchor="w", padx=12, pady=(10, 2))
    tk.Label(f, text=value, font=("Segoe UI", 15, "bold"), bg=C_WHITE, fg=color).pack(anchor="w", padx=12, pady=(0, 12))

def section_title(parent, text):
    tk.Label(parent, text=text, font=("Segoe UI", 11, "bold"),
             bg=C_BG, fg="#1A365D").pack(anchor="w", pady=(0,2))

def divider(parent):
    tk.Frame(parent, bg=C_BORDER, height=1).pack(fill="x", pady=3)

# ─── BARCODE WINDOW FUNCTIONS ─────────────────────────────────────────────────

def _show_barcode_window(barcode_val, product_name="", mrp=0):
    """Single product ka barcode dikhao aur print karo."""
    dlg = tk.Toplevel()
    dlg.title(f"🏷 Barcode — {product_name}")
    dlg.configure(bg=C_WHITE)
    dlg.resizable(False, False)
    w, h = 480, 420
    sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
    dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
    dlg.grab_set()

    tk.Label(dlg, text="🏷 Barcode / Label Print",
             font=("Segoe UI", 13, "bold"), bg=C_WHITE, fg="#1A365D").pack(pady=(14, 4))

    # ── Barcode display ───────────────────────────────────────────────────
    canvas_frame = tk.Frame(dlg, bg="#F7FAFC", relief="ridge", bd=1)
    canvas_frame.pack(padx=20, pady=6, fill="x")

    lbl_bc = tk.Label(canvas_frame, bg="#F7FAFC")
    lbl_bc.pack(pady=4)

    lbl_prod_name = tk.Label(canvas_frame,
                              text=product_name, font=("Segoe UI", 11, "bold"),
                              bg="#F7FAFC", fg="#1A365D")
    lbl_prod_name.pack()
    lbl_mrp_show = tk.Label(canvas_frame,
                              text=f"MRP: ₹{mrp:.2f}" if mrp else "",
                              font=("Segoe UI",9), bg="#F7FAFC", fg="#276749")
    lbl_mrp_show.pack(pady=(0, 6))
    lbl_bc_val = tk.Label(canvas_frame, text=barcode_val,
                           font=("Courier New", 9), bg="#F7FAFC", fg="#4A5568")
    lbl_bc_val.pack(pady=(0, 8))

    def render_barcode(val):
        img = generate_barcode_image(val, product_name, mrp)
        if img and PIL_OK:
            img_resized = img.resize((360, 80), Image.LANCZOS)
            photo = ImageTk.PhotoImage(img_resized)
            lbl_bc.config(image=photo)
            lbl_bc.image = photo
        else:
            lbl_bc.config(text=f"[Barcode: {val}]\n(python-barcode install karo for image)",
                          font=("Courier New", 9), fg="#718096")

    # ── Edit section ─────────────────────────────────────────────────────
    ef = tk.Frame(dlg, bg=C_WHITE); ef.pack(fill="x", padx=20, pady=4)
    tk.Label(ef, text="Barcode Value:", font=("Segoe UI", 9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
    v_bc  = tk.StringVar(value=barcode_val)
    v_mrp = tk.StringVar(value=str(mrp) if mrp else "0")
    ttk.Entry(ef, textvariable=v_bc, width=18).pack(side="left", padx=4)
    tk.Label(ef, text="MRP ₹:", font=("Segoe UI", 9), bg=C_WHITE, fg=C_GRAY).pack(side="left", padx=(8,0))
    ttk.Entry(ef, textvariable=v_mrp, width=8).pack(side="left", padx=4)

    nf = tk.Frame(dlg, bg=C_WHITE); nf.pack(fill="x", padx=20, pady=2)
    tk.Label(nf, text="Copies:", font=("Segoe UI", 9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
    v_copies = tk.StringVar(value="1")
    ttk.Spinbox(nf, from_=1, to=100, textvariable=v_copies, width=5).pack(side="left", padx=4)

    def refresh_preview(*args):
        render_barcode(v_bc.get().strip() or barcode_val)
        try:
            lbl_mrp_show.config(text=f"MRP: ₹{float(v_mrp.get()):.2f}")
        except: pass
    v_bc.trace_add("write", refresh_preview)
    v_mrp.trace_add("write", refresh_preview)
    render_barcode(barcode_val)

    def print_labels():
        try:
            copies = int(v_copies.get())
        except: copies = 1
        bc_val  = v_bc.get().strip() or barcode_val
        mrp_val = float(v_mrp.get() or 0)
        if BARCODE_OK and PIL_OK:
            _print_barcode_labels(bc_val, product_name, mrp_val, copies)
        else:
            # Fallback: text file
            import tkinter.filedialog as fd
            path = fd.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text","*.txt")],
                initialfile=f"barcode_{bc_val}.txt")
            if path:
                with open(path, "w", encoding="utf-8") as f:
                    for _ in range(copies):
                        f.write(f"Product: {product_name}\n")
                        f.write(f"Barcode: {bc_val}\n")
                        f.write(f"MRP: Rs.{mrp_val:.2f}\n")
                        f.write("-"*30 + "\n")
                messagebox.showinfo("Saved!", f"Label file saved:\n{path}", parent=dlg)

    def save_barcode_to_db():
        bc_val = v_bc.get().strip()
        if not bc_val or not product_name: return
        conn = get_db()
        conn.execute("UPDATE products SET barcode=? WHERE name=?", (bc_val, product_name))
        conn.commit(); conn.close()
        messagebox.showinfo("Saved!", "Barcode product mein save ho gaya! ✅", parent=dlg)

    bf2 = tk.Frame(dlg, bg=C_WHITE); bf2.pack(pady=4)

    def make_b(t, cmd, bg):
        return tk.Button(bf2, text=t, font=("Segoe UI", 9, "bold"),
                         bg=bg, fg="white", relief="flat", cursor="hand2",
                         padx=10, pady=6, bd=0, command=cmd)

    make_b("🖨 Print Labels", print_labels, "#276749").pack(side="left", padx=4)
    make_b("💾 DB mein Save", save_barcode_to_db, "#2B6CB0").pack(side="left", padx=4)
    make_b("✕ Close", dlg.destroy, C_GRAY).pack(side="left", padx=4)


def _print_barcode_labels(barcode_val, product_name, mrp, copies=1):
    """PIL se barcode labels PNG mein save karo aur open karo."""
    if not PIL_OK:
        messagebox.showerror("Error", "PIL/Pillow install karo:\npip install Pillow")
        return
    try:
        label_w, label_h = 400, 150
        img = Image.new("RGB", (label_w * copies, label_h), "white")
        bc_img_raw = generate_barcode_image(barcode_val, product_name, mrp)

        for idx in range(copies):
            x_offset = idx * label_w
            # Paste barcode
            if bc_img_raw:
                bc_resized = bc_img_raw.resize((360, 80), Image.LANCZOS)
                img.paste(bc_resized, (x_offset + 20, 10))
            # Draw text
            draw = ImageDraw.Draw(img)
            draw.text((x_offset + 20, 96), product_name[:35], fill="#1A365D")
            if mrp:
                draw.text((x_offset + 20, 114), f"MRP: Rs.{mrp:.2f}", fill="#276749")
            draw.text((x_offset + 20, 132), barcode_val, fill="#4A5568")

        import tempfile, subprocess, sys
        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        img.save(tmp.name, "PNG", dpi=(203, 203))
        tmp.close()
        if sys.platform.startswith("win"):
            os.startfile(tmp.name, "print")
        elif sys.platform.startswith("darwin"):
            subprocess.run(["open", tmp.name])
        else:
            subprocess.run(["xdg-open", tmp.name])
        messagebox.showinfo("Label Ready!",
                            f"{copies} label(s) generate ho gaye!\nFile: {tmp.name}\n\n"
                            "File open hogi — wahan se print karo.")
    except Exception as e:
        messagebox.showerror("Error", f"Label generate nahi hua:\n{e}")


def _show_batch_barcode_window(products):
    """Saare products ke barcodes ek saath generate karo."""
    dlg = tk.Toplevel()
    dlg.title("🏷 Batch Barcode Print — Saare Products")
    dlg.configure(bg=C_WHITE)
    dlg.geometry("700x500")
    dlg.grab_set()

    tk.Label(dlg, text="🏷 Batch Barcode Print",
             font=("Segoe UI", 12, "bold"), bg=C_WHITE, fg="#1A365D").pack(pady=(12, 4))
    tk.Label(dlg, text="Products select karo aur barcode labels generate karo:",
             font=("Segoe UI", 9), bg=C_WHITE, fg=C_GRAY).pack()

    # Scrollable list with checkboxes
    frame_outer = tk.Frame(dlg, bg=C_WHITE)
    frame_outer.pack(fill="both", expand=True, padx=16, pady=4)

    canvas = tk.Canvas(frame_outer, bg=C_WHITE, highlightthickness=0)
    scrollbar = ttk.Scrollbar(frame_outer, orient="vertical", command=canvas.yview)
    scroll_frame = tk.Frame(canvas, bg=C_WHITE)
    scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    check_vars = []
    copy_vars  = []
    for i, pr in enumerate(products):
        row = tk.Frame(scroll_frame, bg=C_WHITE if i%2==0 else "#F7FAFC")
        row.pack(fill="x", pady=1)
        var = tk.BooleanVar(value=True)
        check_vars.append(var)
        tk.Checkbutton(row, variable=var, bg=row["bg"]).pack(side="left", padx=4)
        tk.Label(row, text=pr["name"], font=("Segoe UI", 9), bg=row["bg"],
                 fg=C_DARK, width=28, anchor="w").pack(side="left")
        tk.Label(row, text=pr.get("barcode","") or "—",
                 font=("Courier New", 8), bg=row["bg"], fg="#718096", width=14, anchor="w").pack(side="left")
        tk.Label(row, text=f"MRP ₹{pr.get('mrp',0):.2f}",
                 font=("Segoe UI", 9), bg=row["bg"], fg="#276749", width=12, anchor="w").pack(side="left")
        tk.Label(row, text="Copies:", font=("Segoe UI",7), bg=row["bg"], fg=C_GRAY).pack(side="left")
        cv = tk.StringVar(value="1")
        copy_vars.append(cv)
        ttk.Spinbox(row, from_=1, to=50, textvariable=cv, width=4).pack(side="left", padx=4)

    def generate_all():
        selected = [(products[i], int(copy_vars[i].get() or 1))
                    for i, var in enumerate(check_vars) if var.get()]
        if not selected:
            messagebox.showwarning("Koi Select Nahi","Koi product select nahi kiya!", parent=dlg)
            return
        for pr, copies in selected:
            bc = pr.get("barcode","") or pr["name"]
            _print_barcode_labels(bc, pr["name"], pr.get("mrp",0) or 0, copies)
        messagebox.showinfo("Done!", f"{len(selected)} products ke labels generate ho gaye!", parent=dlg)

    def select_all():
        for v in check_vars: v.set(True)
    def deselect_all():
        for v in check_vars: v.set(False)

    bf3 = tk.Frame(dlg, bg=C_WHITE); bf3.pack(pady=4)
    for t, cmd, bg in [("✅ Sab Select", select_all, "#2B6CB0"),
                        ("☐ Deselect All", deselect_all, C_GRAY),
                        ("🖨 Generate Labels", generate_all, "#276749"),
                        ("Close", dlg.destroy, C_GRAY)]:
        tk.Button(bf3, text=t, font=("Segoe UI", 9, "bold"),
                  bg=bg, fg="white", relief="flat", cursor="hand2",
                  padx=10, pady=6, bd=0, command=cmd).pack(side="left", padx=4)


# ─── REUSABLE BARCODE SCAN BUTTON HELPER ──────────────────────────────────────
def add_barcode_scan_btn(parent, prod_var, rate_var=None, gst_var=None,
                         on_found=None, side="left", padx=(2,0)):
    """
    Kisi bhi page pe Product field ke paas 📷 Scan button lagao.

    prod_var  : tk.StringVar — product naam yahan set hoga
    rate_var  : tk.StringVar (optional) — auto-fill rate
    gst_var   : tk.StringVar (optional) — auto-fill GST
    on_found  : callable(product_dict) — extra action after product found
    """
    def do_scan():
        def on_scan(barcode_val):
            conn_ = get_db()
            row = conn_.execute("SELECT * FROM products WHERE barcode=?", (barcode_val,)).fetchone()
            conn_.close()
            if row:
                row = dict(row)
                prod_var.set(row["name"])
                if rate_var:
                    # For sale: use sale_rate; for purchase: use purchase_rate if available
                    rate_var.set(str(row.get("sale_rate", 0)))
                if gst_var:
                    gst_var.set(str(int(row.get("gst_percent", 5))))
                if on_found:
                    on_found(row)
                messagebox.showinfo("✅ Product Mila!",
                                    f"📦 {row['name']}\n"
                                    f"Barcode: {barcode_val}\n"
                                    f"MRP: ₹{row.get('mrp', 0):.2f}")
            else:
                messagebox.showwarning("⚠️ Nahi Mila",
                    f"Barcode '{barcode_val}' kisi product se match nahi hua.\n"
                    "Products page pe barcode save karo pehle.")
        scan_barcode_from_camera(on_scan)

    btn = tk.Button(parent, text="📷 Scan", font=("Segoe UI", 8, "bold"),
                    bg="#2B6CB0", fg="white", relief="flat", cursor="hand2",
                    padx=6, pady=2, bd=0, command=do_scan)
    btn.pack(side=side, padx=padx)
    return btn


# ══════════════════════════════════════════════════════════════════════════════
#  LOGIN
# ══════════════════════════════════════════════════════════════════════════════
class LoginWin:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("BhugtanEase — Login")
        self.root.configure(bg=C_WHITE)

        self.root.resizable(False, False)
        w, h = 400, 520
        sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        _apply_logo(self.root)
        style_setup()
        self._build()
        self.root.mainloop()

    def _build(self):
        top = tk.Frame(self.root, bg=C_WHITE, height=160)
        top.pack(fill="x")
        top.pack_propagate(False)

        # ── Embedded Logo (cached bytes se seedha load — koi decode/disk nahi) ──
        logo_shown = False
        try:
            from PIL import Image, ImageTk
            import io
            _img = Image.open(io.BytesIO(_LOGO_BYTES))
            _img.thumbnail((260, 90), Image.LANCZOS)
            self._logo_img = ImageTk.PhotoImage(_img)
            tk.Label(top, image=self._logo_img, bg=C_WHITE).pack(pady=(22, 4))
            logo_shown = True
        except Exception:
            pass

        if not logo_shown:
            tk.Label(top, text="BhugtanEase", font=("Segoe UI", 28, "bold"),
                     bg=C_WHITE, fg=C_TOP).pack(pady=(35, 4))

        tk.Label(top, text="Billing & Inventory Management",
                 font=("Segoe UI",9), bg=C_WHITE, fg=C_LGRAY).pack(pady=(0, 6))
        tk.Frame(top, bg=C_BORDER, height=1).pack(fill="x", pady=(4, 0))

        frm = tk.Frame(self.root, bg=C_WHITE, padx=50)
        frm.pack(fill="x", pady=28)

        tk.Label(frm, text="Username", font=("Segoe UI",9),
                 bg=C_WHITE, fg=C_GRAY).pack(anchor="w", pady=(0,3))
        self.u = tk.StringVar()
        ue = ttk.Entry(frm, textvariable=self.u, font=("Segoe UI", 12))
        ue.pack(fill="x", ipady=5)
        ue.focus_set()

        tk.Label(frm, text="Password", font=("Segoe UI",9),
                 bg=C_WHITE, fg=C_GRAY).pack(anchor="w", pady=(14,3))
        self.p = tk.StringVar()
        pe = ttk.Entry(frm, textvariable=self.p, show="*", font=("Segoe UI", 12))
        pe.pack(fill="x", ipady=5)
        pe.bind("<Return>", lambda e: self._login())

        self.err = tk.Label(frm, text="", font=("Segoe UI", 9),
                            bg=C_WHITE, fg=C_RED)
        self.err.pack(pady=(8,0))

        make_btn(frm, "  LOGIN  ", self._login).pack(fill="x", pady=(14,0))

        tk.Label(self.root, text="Ashrisha Ecommerce Solution Pvt Ltd",
                 font=("Segoe UI", 9, "bold"), bg=C_WHITE, fg="#A0AEC0").pack(pady=(10, 2))

    def _login(self):
        conn = get_db()
        row = conn.execute("SELECT id FROM users WHERE username=? AND password=?",
                           (self.u.get().strip(), self.p.get())).fetchone()
        conn.close()
        if row:
            root = self.root
            # Sabhi login widgets hatao
            for w in root.winfo_children():
                w.destroy()
            # Window ko app mode mein configure karo
            root.title("BhugtanEase Billing Software")
            root.configure(bg=C_BG)
            root.resizable(True, True)
            root.state("zoomed")
            root.update()
            # AppWin ko isi root pe build karo
            style_setup()
            app = object.__new__(AppWin)
            app.root = root
            try:
                app._build_shell()
                app._show("dashboard")
            except Exception:
                import traceback, logging
                logging.basicConfig(filename="bhugtanease_error.log",
                                    level=logging.ERROR, encoding="utf-8")
                logging.error("App Load Error:\n" + traceback.format_exc())
                messagebox.showerror("Error",
                    "Software load karne mein problem aayi.\n"
                    "Support se sampark karein.\n\n"
                    "Error log: bhugtanease_error.log")
        else:
            self.err.config(text="Galat username ya password!")

# ══════════════════════════════════════════════════════════════════════════════
#  MAIN APP
# ══════════════════════════════════════════════════════════════════════════════
class AppWin:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("BhugtanEase Billing Software")
        self.root.state("zoomed")
        self.root.configure(bg=C_BG)
        _apply_logo(self.root)
        style_setup()
        self._build_shell()
        self._show("dashboard")

        def _on_close():
            try:
                self._do_auto_backup_on_exit()
            except: pass
            self.root.destroy()

        self.root.protocol("WM_DELETE_WINDOW", _on_close)
        self.root.mainloop()

    def _build_shell(self):
        # Topbar
        top = tk.Frame(self.root, bg=C_TOP, height=50)
        top.pack(fill="x")
        top.pack_propagate(False)

        # ── Header Logo ──────────────────────────────────────────────────────
        _header_logo_shown = False
        try:
            from PIL import Image, ImageTk
            import io
            _img = Image.open(io.BytesIO(_LOGO_BYTES))
            _img.thumbnail((160, 36), Image.LANCZOS)
            self._header_logo_img = ImageTk.PhotoImage(_img)
            tk.Label(top, image=self._header_logo_img, bg=C_TOP).pack(side="left", padx=10, pady=7)
            _header_logo_shown = True
        except Exception:
            pass

        _sh = get_shop()
        if not _header_logo_shown:
            self._hdr_name_lbl = tk.Label(top, text=_sh['name'] or "BhugtanEase",
                     font=("Segoe UI", 14, "bold"), bg=C_TOP, fg=C_WHITE)
            self._hdr_name_lbl.pack(side="left", padx=14)
        else:
            self._hdr_name_lbl = None

        _hdr_addr = f"{_sh['address']}, {_sh['city']}".strip(", ")
        if _sh['gstin']: _hdr_addr += f"  |  GSTIN: {_sh['gstin']}"
        self._hdr_addr_lbl = tk.Label(top, text=_hdr_addr or "Shop Settings mein address set karein",
                 font=("Segoe UI", 9), bg=C_TOP, fg="#B3D4F0")
        self._hdr_addr_lbl.pack(side="left", padx=6)
        tk.Label(top, text=datetime.date.today().strftime("%d %b %Y"),
                 font=("Segoe UI",9), bg=C_TOP, fg="#B3D4F0").pack(side="right", padx=12)
        make_btn(top, "Logout", self._logout, bg="#2C5282").pack(side="right", pady=10, padx=8)

        body = tk.Frame(self.root, bg=C_BG)
        body.pack(fill="both", expand=True)

        # ── Sidebar with scroll ──────────────────────────────────────────────
        sb_outer = tk.Frame(body, bg=C_SIDEBAR, width=190)
        sb_outer.pack(side="left", fill="y")
        sb_outer.pack_propagate(False)

        sb_canvas = tk.Canvas(sb_outer, bg=C_SIDEBAR, highlightthickness=0, width=190)
        sb_vsb = ttk.Scrollbar(sb_outer, orient="vertical", command=sb_canvas.yview)
        sb_canvas.configure(yscrollcommand=sb_vsb.set)
        # Scrollbar hamesha right side pe — sidebar scroll karne ke liye
        sb_vsb.pack(side="right", fill="y")
        sb_canvas.pack(side="left", fill="both", expand=True)

        sb = tk.Frame(sb_canvas, bg=C_SIDEBAR, width=190)
        sb_win = sb_canvas.create_window((0, 0), window=sb, anchor="nw")

        def _sb_resize(e):
            sb_canvas.itemconfig(sb_win, width=e.width)
        def _sb_inner(e):
            sb_canvas.configure(scrollregion=sb_canvas.bbox("all"))

        sb_canvas.bind("<Configure>", _sb_resize)
        sb.bind("<Configure>", _sb_inner)
        sb_canvas.bind("<Enter>",
            lambda e: sb_canvas.bind_all("<MouseWheel>",
                lambda ev: sb_canvas.yview_scroll(-1*(ev.delta//120), "units")))
        sb_canvas.bind("<Leave>", lambda e: sb_canvas.unbind_all("<MouseWheel>"))

        self.nav_btns = {}

        # Dashboard — always visible
        dash_btn = tk.Button(sb, text="🏠  Dashboard", font=("Segoe UI", 10, "bold"),
                             bg=C_NAV_ACT, fg=C_WHITE, relief="flat",
                             anchor="w", padx=12, cursor="hand2",
                             activebackground=C_NAV_ACT, activeforeground=C_WHITE,
                             command=lambda: self._show("dashboard"))
        dash_btn.pack(fill="x", ipady=2)
        self.nav_btns["dashboard"] = dash_btn

        tk.Frame(sb, bg="#2A5298", height=1).pack(fill="x")

        # Collapsible sections
        SECTIONS = [
            ("📋  TRANSACTIONS", [
                ("Sale Bill",           "sale"),
                ("Purchase",            "purchase"),
                ("Sale Return",         "salereturn"),
                ("Purchase Return",     "purreturn"),
                ("Exchange",            "exchange"),
            ]),
            ("🕘  HISTORY", [
                ("Sale History",        "salehistory"),
                ("Purchase History",    "purhistory"),
                ("Sale Return Hist.", "salereturnhistory"),
                ("Pur. Return Hist.", "purreturnhistory"),
                ("Exchange History",  "exchangehistory"),
            ]),
            ("📦  INVENTORY", [
                ("Stock",               "stock"),
                ("Stock In/Out",        "stockinout"),
                ("Products",            "products"),
                ("Expiry Manager",      "expiry"),
            ]),
            ("💼  ACCOUNTS", [
                ("Ledger",              "ledger"),
                ("Parties",             "parties"),
                ("Payment Reminders",   "reminders"),
                ("Expenses",            "expenses"),
            ]),
            ("📊  REPORTS", [
                ("P&L Report",          "pl"),
                ("Sale Report",         "salereport"),
                ("Balance Sheet",       "balancesheet"),
                ("GSTR-1",              "gstr1"),
                ("GSTR-3B",             "gstr3b"),
                ("FIFO Stock",          "fiforeport"),
            ]),
            ("🏛  GST COMPLIANCE", [
                ("E-Invoice",           "einvoice"),
                ("E-Way Bill",          "ewaybill"),
            ]),
            ("⚙️  SETTINGS", [
                ("Shop Settings",       "shopsettings"),
                ("Backup & Restore",    "backup"),
            ]),
        ]

        # Store child frames and arrow vars for accordion control
        self._sec_frames = {}
        self._sec_arrows = {}
        self._sec_hdrs   = {}

        def _close_all():
            for sl, cf in self._sec_frames.items():
                cf.pack_forget()
                cur = self._sec_hdrs[sl].cget("text")
                if cur.startswith("▾  "):
                    self._sec_hdrs[sl].config(text=cur[3:], bg="#0F2D5E", fg="#7BAFD4")
                else:
                    self._sec_hdrs[sl].config(bg="#0F2D5E", fg="#7BAFD4")
                # Wrapper ka bg bhi reset
                cf.master.config(bg="#0F2D5E")

        def _toggle_section(sec_lbl):
            cf = self._sec_frames[sec_lbl]
            hdr = self._sec_hdrs[sec_lbl]
            cur_text = hdr.cget("text")
            if cf.winfo_ismapped():
                cf.pack_forget()
                hdr.config(bg="#0F2D5E", fg="#7BAFD4",
                           text=cur_text.replace("▾  ",""))
                # Wrapper ka bg bhi reset
                cf.master.config(bg="#0F2D5E")
            else:
                _close_all()
                cf.pack(fill="x")
                hdr.config(bg=C_NAV_ACT, fg=C_WHITE,
                           text="▾  " + cur_text.replace("▾  ",""))
                # Wrapper ka bg active color
                cf.master.config(bg=C_NAV_ACT)

        for sec_label, items in SECTIONS:
            arrow_var = tk.StringVar(value="  ▸")
            self._sec_arrows[sec_label] = arrow_var

            # ── Wrapper: header + child ek saath ────────────────────────────
            wrapper = tk.Frame(sb, bg="#0F2D5E")
            wrapper.pack(fill="x")

            # Header button — wrapper ke andar
            hdr = tk.Button(
                wrapper,
                text=sec_label,
                font=("Segoe UI", 9, "bold"),
                bg="#0F2D5E", fg="#7BAFD4",
                relief="flat", cursor="hand2",
                anchor="w", padx=10,
                activebackground=C_NAV_ACT,
                activeforeground=C_WHITE,
                command=lambda sl=sec_label: _toggle_section(sl)
            )
            hdr.pack(fill="x", ipady=7)
            self._sec_hdrs[sec_label] = hdr

            # Child frame — usi wrapper ke andar (hidden by default)
            child = tk.Frame(wrapper, bg="#162B50")
            self._sec_frames[sec_label] = child
            # child NOT packed — hidden by default

            # Nav buttons inside child
            for label, key in items:
                btn = tk.Button(child, text="    " + label,
                                font=("Segoe UI",9),
                                bg="#162B50", fg=C_NAV_FG,
                                relief="flat", anchor="w",
                                padx=12, cursor="hand2",
                                activebackground=C_NAV_ACT,
                                activeforeground=C_WHITE,
                                command=lambda k=key: self._show(k))
                btn.pack(fill="x", ipady=6)
                self.nav_btns[key] = btn

            # Separator — wrapper ke baad (ek section ke neeche)
            tk.Frame(sb, bg="#2A5298", height=1).pack(fill="x")

        tk.Label(sb, bg=C_SIDEBAR, height=2).pack(fill="x")

        # ── Content area — always scrollable ────────────────────────────────
        content_outer = tk.Frame(body, bg=C_BG)
        content_outer.pack(fill="both", expand=True)

        self._main_canvas = tk.Canvas(content_outer, bg=C_BG, highlightthickness=0)
        self._main_vsb = ttk.Scrollbar(content_outer, orient="vertical",
                                        command=self._main_canvas.yview)
        self._main_canvas.configure(yscrollcommand=self._main_vsb.set)
        self._main_vsb.pack(side="right", fill="y")
        self._main_canvas.pack(fill="both", expand=True)

        self.content = tk.Frame(self._main_canvas, bg=C_BG)
        self._content_win = self._main_canvas.create_window(
            (0, 0), window=self.content, anchor="nw")

        def _on_canvas_resize(e):
            self._main_canvas.itemconfig(self._content_win, width=e.width)
            # Form pages: set content height = canvas height so no vertical scroll needed
            if getattr(self, "_fullscreen_page", False):
                self._main_canvas.itemconfig(self._content_win, height=e.height)
        def _on_content_resize(e):
            if not getattr(self, "_fullscreen_page", False):
                self._main_canvas.configure(scrollregion=self._main_canvas.bbox("all"))
            else:
                self._main_canvas.configure(scrollregion=(0, 0, e.width, e.height))

        self._main_canvas.bind("<Configure>", _on_canvas_resize)
        self.content.bind("<Configure>", _on_content_resize)

        # Mousewheel on main content (bind when mouse enters canvas)
        self._main_canvas.bind("<Enter>",
            lambda e: self._main_canvas.bind_all("<MouseWheel>",
                lambda ev: self._main_canvas.yview_scroll(-1*(ev.delta//120), "units")))
        self._main_canvas.bind("<Leave>",
            lambda e: self._main_canvas.unbind_all("<MouseWheel>"))

    # Pehle kuch pages full-screen (no-scroll) mode mein the — ab sabhi pages
    # scrollable hain taaki koi bhi button/total/option screen se cut na ho.
    _FULLSCREEN_PAGES = set()

    def _show(self, name):
        for w in self.content.winfo_children():
            w.destroy()
        self._fullscreen_page = name in self._FULLSCREEN_PAGES
        self._main_vsb.pack_forget() if self._fullscreen_page else self._main_vsb.pack(side="right", fill="y")
        self._main_canvas.yview_moveto(0)
        for k, b in self.nav_btns.items():
            b.config(bg=C_SIDEBAR, fg=C_NAV_FG)
        if name in self.nav_btns:
            self.nav_btns[name].config(bg=C_NAV_ACT, fg=C_WHITE)
        try:
            getattr(self, f"_pg_{name}")()
        except Exception:
            import traceback, logging
            logging.basicConfig(filename="bhugtanease_error.log",
                                level=logging.ERROR, encoding="utf-8")
            logging.error(f"Page '{name}' load error:\n" + traceback.format_exc())
            ef = tk.Frame(self.content, bg=C_BG, padx=20, pady=20)
            ef.pack(fill="x")
            tk.Label(ef, text="⚠️  Yeh page load nahi ho saka.", font=("Segoe UI",12,"bold"),
                     bg=C_BG, fg=C_RED).pack(anchor="w")
            tk.Label(ef, text="Kripya software restart karein ya support se sampark karein.",
                     font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(anchor="w", pady=(4,0))

    def _logout(self):
        self.root.destroy()
        LoginWin()

    def _content_scroll(self):
        """Returns a padded frame inside the main scrollable content area."""
        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        return p

    # ══════════════════════════════════════════════════════════════════════════
    #  DASHBOARD
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_dashboard(self):
        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "Dashboard")

        conn = get_db()
        total_sale = conn.execute("SELECT COALESCE(SUM(grand_total),0) FROM sales").fetchone()[0]
        total_pur  = conn.execute("SELECT COALESCE(SUM(grand_total),0) FROM purchases").fetchone()[0]
        gst_amt    = conn.execute("SELECT COALESCE(SUM(gst_amt),0) FROM sale_items").fetchone()[0]
        pur_gst    = conn.execute("SELECT COALESCE(SUM(gst_amt),0) FROM purchase_items").fetchone()[0]

        # ── Sale Returns — revenue reduce karo ───────────────────────────────
        sr_rows    = [dict(r) for r in conn.execute(
            "SELECT * FROM returns WHERE return_type='sale_return'").fetchall()]
        sr_total   = sum(r["total_amt"] for r in sr_rows)          # GST-inclusive return amt
        sr_gst     = sum(r["gst_amt"]   for r in sr_rows)          # GST part of returns
        sr_taxable = sr_total - sr_gst                             # ex-GST return value

        # ── Purchase Returns — purchase cost reduce karo ──────────────────────
        pr_rows    = [dict(r) for r in conn.execute(
            "SELECT * FROM returns WHERE return_type='pur_return'").fetchall()]
        pr_total   = sum(r["total_amt"] for r in pr_rows)

        net_sale   = total_sale - sr_total                         # actual net sale after returns
        taxable    = (total_sale - gst_amt) - sr_taxable           # net sale ex-GST after returns

        # ── Net Profit = Net Sale (ex-GST) − COGS — FIFO-based (P&L ke jaisa) ──
        # Step 1: FIFO COGS — fifo_consumption table se (same as P&L report)
        fifo_rows_dash = [dict(r) for r in conn.execute(
            "SELECT product, SUM(total_cost) as cogs FROM fifo_consumption GROUP BY product"
        ).fetchall()]
        fifo_cogs_by_prod_dash = {fr["product"]: fr["cogs"] for fr in fifo_rows_dash}

        # Step 2: Avg purchase cost map (fallback for products without FIFO data)
        all_pi = [dict(r) for r in conn.execute("SELECT * FROM purchase_items").fetchall()]
        pur_cost_map = {}
        for x in all_pi:
            k = x["product"]
            if k not in pur_cost_map: pur_cost_map[k] = {"cost": 0, "qty": 0}
            ex_gst = x.get("taxable", 0) or max(0, x.get("total", 0) - x.get("gst_amt", 0))
            pur_cost_map[k]["cost"] += ex_gst
            pur_cost_map[k]["qty"]  += x.get("qty", 0)

        # Step 3: COGS calculate karo — FIFO first, avg fallback
        all_si = [dict(r) for r in conn.execute("SELECT * FROM sale_items").fetchall()]
        cogs = 0.0
        fifo_products_used = set()
        for prod, fifo_cogs in fifo_cogs_by_prod_dash.items():
            cogs += fifo_cogs
            fifo_products_used.add(prod)
        for x in all_si:
            if x["product"] not in fifo_products_used:
                d = pur_cost_map.get(x["product"])
                if d and d["qty"] > 0:
                    cogs += (d["cost"] / d["qty"]) * x.get("qty", 0)

        # Step 4: Returned items ka COGS minus karo
        cogs_returned = 0.0
        for r in sr_rows:
            d = pur_cost_map.get(r["product"])
            if d and d["qty"] > 0:
                cogs_returned += (d["cost"] / d["qty"]) * r["qty"]
        cogs_net = cogs - cogs_returned

        # Step 5: Expenses minus karo — Net Profit = Sale - COGS - Expenses
        total_expenses = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM expenses"
        ).fetchone()[0] or 0
        profit = taxable - cogs_net - total_expenses  # Net Profit — P&L ke same formula
        prods      = [dict(r) for r in conn.execute("SELECT * FROM products").fetchall()]
        stock_val  = sum(get_stock(pr["name"]) * pr["sale_rate"] for pr in prods)

        # Customer received — no double count
        originally_cash = conn.execute("""
            SELECT COALESCE(SUM(grand_total),0) FROM sales
            WHERE pay_mode='Cash'
            AND bill_no NOT IN (SELECT DISTINCT bill_no FROM bill_payments WHERE bill_type='sale')
        """).fetchone()[0]
        credit_rcvd = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM bill_payments WHERE bill_type='sale'").fetchone()[0]
        total_received = originally_cash + credit_rcvd
        remain_market  = total_sale - total_received

        # Supplier payments — no double count
        pur_orig_cash = conn.execute("""
            SELECT COALESCE(SUM(grand_total),0) FROM purchases
            WHERE pay_mode='Cash'
            AND bill_no NOT IN (SELECT DISTINCT bill_no FROM bill_payments WHERE bill_type='pur')
        """).fetchone()[0]
        pur_credit_paid = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM bill_payments WHERE bill_type='pur'").fetchone()[0]
        total_pur_paid = pur_orig_cash + pur_credit_paid
        pur_pending    = max(0, total_pur - total_pur_paid)

        conn.close()

        # Row 1
        cf = tk.Frame(p, bg=C_BG); cf.pack(fill="x", pady=(0,3))
        stat_card(cf, "Gross Sale",        f"₹{round(total_sale):,}",  C_ACCENT)
        stat_card(cf, "Sale Returns",      f"₹{round(sr_total):,}",  C_RED)
        stat_card(cf, "Net Sale",          f"₹{round(net_sale):,}",    "#2B6CB0")
        stat_card(cf, "Total Purchase",    f"₹{round(total_pur):,}",   C_AMBER)
        stat_card(cf, "Net Profit",        f"₹{round(profit):,}",      C_GREEN if profit>=0 else C_RED)
        stat_card(cf, "Total Expenses",    f"₹{round(total_expenses):,}", C_AMBER)
        stat_card(cf, "Stock Value",       f"₹{round(stock_val):,}",   C_PURPLE)

        # Row 2: Customer
        tk.Label(p, text="👥 Customer (Receivable)", font=("Segoe UI",9,"bold"),
                 bg=C_BG, fg=C_GRAY).pack(anchor="w", padx=4)
        cf2 = tk.Frame(p, bg=C_BG); cf2.pack(fill="x", pady=(2,6))
        stat_card(cf2, "💰 Total Sale",           f"₹{round(total_sale):,}",     C_ACCENT)
        stat_card(cf2, "✅ Received (Customers)",  f"₹{round(total_received):,}", C_GREEN)
        stat_card(cf2, "⏳ Baki (Market mein)",    f"₹{round(remain_market):,}",
                  C_RED if remain_market > 0 else C_GREEN)

        # Row 3: Supplier
        tk.Label(p, text="🏭 Supplier (Payable)", font=("Segoe UI",9,"bold"),
                 bg=C_BG, fg=C_GRAY).pack(anchor="w", padx=4)
        cf3 = tk.Frame(p, bg=C_BG); cf3.pack(fill="x", pady=(2,14))
        stat_card(cf3, "🛒 Total Purchase",        f"₹{round(total_pur):,}",      C_AMBER)
        stat_card(cf3, "✅ Paid (Suppliers)",       f"₹{round(total_pur_paid):,}", C_GREEN)
        stat_card(cf3, "⏳ Dena Baaki (Suppliers)",f"₹{round(pur_pending):,}",
                  C_RED if pur_pending > 0 else C_GREEN)

        # ── Product-wise stock details ──────────────────────────────────────
        stock_data = []
        conn2 = get_db()
        for pr in prods:
            pname = pr["name"]
            rate  = pr["sale_rate"] or 0
            opening = pr["opening_stock"] or 0

            pur_qty = conn2.execute(
                "SELECT COALESCE(SUM(qty+free_qty),0) FROM purchase_items WHERE product=?", (pname,)).fetchone()[0]
            pur_rate_avg = conn2.execute(
                "SELECT COALESCE(SUM(total)/NULLIF(SUM(qty),0),0) FROM purchase_items WHERE product=?", (pname,)).fetchone()[0]
            sale_qty = conn2.execute(
                "SELECT COALESCE(SUM(qty),0) FROM sale_items WHERE product=?", (pname,)).fetchone()[0]
            sale_taxable = conn2.execute(
                "SELECT COALESCE(SUM(taxable),0) FROM sale_items WHERE product=?", (pname,)).fetchone()[0]

            total_in  = opening + pur_qty
            remain    = total_in - sale_qty
            pur_amt   = pur_qty * pur_rate_avg if pur_rate_avg else pur_qty * rate
            sale_amt  = sale_taxable
            remain_amt= remain * rate

            stock_data.append({
                "name": pname, "rate": rate,
                "pur_qty": pur_qty,   "pur_amt": pur_amt,
                "sale_qty": sale_qty, "sale_amt": sale_amt,
                "remain": remain,     "remain_amt": remain_amt,
            })
        conn2.close()
        conn.close()

        # ── Recent Sales (compact) ──────────────────────────────────────────
        two = tk.Frame(p, bg=C_BG)
        two.pack(fill="x")
        two.columnconfigure(0, weight=1)
        two.columnconfigure(1, weight=1)

        lf = tk.Frame(two, bg=C_BG)
        lf.grid(row=0, column=0, sticky="nsew", padx=(0,8))
        make_label(lf, "Recent Sales", 11, True, color="#1A365D").pack(anchor="w", pady=(0,3))
        tbl_holder = tk.Frame(lf, bg=C_BG, height=220)
        tbl_holder.pack(fill="x"); tbl_holder.pack_propagate(False)
        conn3 = get_db()
        inn = make_table(tbl_holder, ["Bill No","Party","Amount","Date"], [18,22,12,14])
        recent = conn3.execute("SELECT * FROM sales ORDER BY id DESC LIMIT 8").fetchall()
        conn3.close()
        for i, s in enumerate(recent):
            table_row(inn, [s["bill_no"], s["party"][:20],
                            f"₹{round(s['grand_total']):,}", fmt_date(s["bill_date"])], i+1)

        rf = tk.Frame(two, bg=C_BG)
        rf.grid(row=0, column=1, sticky="nsew", padx=(8,0))
        make_label(rf, "Stock Alert", 11, True, color="#1A365D").pack(anchor="w", pady=(0,3))
        tbl_holder2 = tk.Frame(rf, bg=C_BG, height=220)
        tbl_holder2.pack(fill="x"); tbl_holder2.pack_propagate(False)
        inn2 = make_table(tbl_holder2, ["Product","Remain Qty","Status"], [26,10,10])
        for i, pr in enumerate(prods):
            stk = get_stock(pr["name"])
            s   = "⚠ Low" if stk<=2 else "• Med" if stk<=5 else "✓ OK"
            c   = C_RED if stk<=2 else C_AMBER if stk<=5 else C_GREEN
            table_row(inn2, [pr["name"][:24], int(stk), s], i+1, fgs=[None,None,c])

        # ── Product-wise Stock Summary Table ───────────────────────────────
        divider(p)
        make_label(p, "📦  Product-wise Stock Summary", 12, True, color="#1A365D").pack(anchor="w", pady=(6,6))

        # Column headers
        SCOLS = ["Product", "Rate(₹)",
                 "Purchase Qty", "Purchase Amt(₹)",
                 "Sale Qty",     "Sale Amt(₹)",
                 "Remain Qty",   "Remain Amt(₹)", "Status"]
        SWEIGHTS = [20, 8, 12, 14, 10, 13, 11, 14, 8]

        stk_tbl = make_table(p, SCOLS, SWEIGHTS)

        t_pqty=0; t_pamt=0; t_sqty=0; t_samt=0; t_rqty=0; t_ramt=0
        for i, sd in enumerate(stock_data):
            status = "⚠ Low" if sd["remain"]<=2 else "• Med" if sd["remain"]<=5 else "✓ OK"
            sc     = C_RED if sd["remain"]<=2 else C_AMBER if sd["remain"]<=5 else C_GREEN
            t_pqty+=sd["pur_qty"]; t_pamt+=sd["pur_amt"]
            t_sqty+=sd["sale_qty"]; t_samt+=sd["sale_amt"]
            t_rqty+=sd["remain"];   t_ramt+=sd["remain_amt"]
            table_row(stk_tbl, [
                sd["name"][:18],
                f"{sd['rate']:.0f}",
                f"{sd['pur_qty']:.0f}",
                f"₹{sd['pur_amt']:,.0f}",
                f"{sd['sale_qty']:.0f}",
                f"₹{sd['sale_amt']:,.0f}",
                f"{sd['remain']:.0f}",
                f"₹{sd['remain_amt']:,.0f}",
                status,
            ], i+1, fgs=[None,None,None,None,None,None,None,None,sc])

        # Totals row
        bg_tot = "#DBEAFE"
        tot_vals = ["TOTAL","",
                    f"{t_pqty:.0f}", f"₹{t_pamt:,.0f}",
                    f"{t_sqty:.0f}", f"₹{t_samt:,.0f}",
                    f"{t_rqty:.0f}", f"₹{t_ramt:,.0f}", ""]
        for j, val in enumerate(tot_vals):
            tk.Label(stk_tbl, text=str(val), font=("Segoe UI", 9, "bold"),
                     bg=bg_tot, fg=C_ACCENT, anchor="w", padx=6, pady=5
                     ).grid(row=len(stock_data)+1, column=j, sticky="nsew", padx=1)

    # ══════════════════════════════════════════════════════════════════════════
    #  SALE BILL
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_sale(self):
        self._sale_items = []
        p = tk.Frame(self.content, bg=C_BG, padx=10, pady=3)
        p.pack(fill="both", expand=True)
        section_title(p, "New Sale Bill")

        # ── GST / Non-GST Bill Mode Toggle ────────────────────────────────────
        mode_row = tk.Frame(p, bg=C_BG)
        mode_row.pack(fill="x", pady=(0,2))
        self._s_bill_mode = tk.StringVar(value="GST")
        self._sale_mode_btn = tk.Button(
            mode_row, text="✅ GST Bill  —  click to switch to Non-GST",
            font=("Segoe UI",9,"bold"), bg="#276749", fg="white",
            relief="flat", cursor="hand2", bd=0, padx=10, pady=5,
            command=lambda: self._sale_toggle_bill_mode())
        self._sale_mode_btn.pack(side="left")

        conn = get_db()
        all_parties  = [r[0] for r in conn.execute("SELECT name FROM parties ORDER BY name").fetchall()]
        all_products = [dict(r) for r in conn.execute("SELECT * FROM products ORDER BY name").fetchall()]
        conn.close()
        self._prod_dict = {pr["name"]: pr for pr in all_products}

        # ── TOP FORM: Bill No | Date | GST Type | Override (4 equal columns) ──────
        tf = tk.Frame(p, bg=C_BG)
        tf.pack(fill="x", pady=(0,1))
        for i in range(4): tf.columnconfigure(i, weight=1)

        LBL  = {"font":("Segoe UI",8), "bg":C_BG, "fg":C_GRAY, "anchor":"w"}
        PADY = {"pady":(1,0)}

        self._s_billno = tk.StringVar(value=next_sale_no())
        tk.Label(tf,text="Bill No",**LBL).grid(row=0,column=0,sticky="ew",padx=(0,3),**PADY)
        billno_entry = ttk.Entry(tf,textvariable=self._s_billno,font=("Segoe UI",9))
        billno_entry.grid(row=1,column=0,sticky="ew",padx=(0,3),pady=(0,2),ipady=1)

        self._s_date = tk.StringVar(value=today_str())
        tk.Label(tf,text="Date (YYYY-MM-DD)",**LBL).grid(row=0,column=1,sticky="ew",padx=3,**PADY)
        date_picker = make_date_entry(tf,self._s_date,width=14)
        date_picker.grid(row=1,column=1,sticky="ew",padx=3,pady=(0,2))

        # ── GST Type indicator (moved up from bottom) ────────────────────────
        OWN_STATE = "Uttar Pradesh"  # Company ka state
        OWN_GST_CODE = "09"          # UP ka GST state code
        gst_type_var = tk.StringVar(value="CGST+SGST")
        tk.Label(tf,text="GST Type",**LBL).grid(row=0,column=2,sticky="ew",padx=3,**PADY)
        gst_state_lbl = tk.Label(tf, font=("Segoe UI",8,"bold"), bg=C_BG,
                                  text="🟢 CGST + SGST (default)", fg="#276749", anchor="w")
        gst_state_lbl.grid(row=1,column=2,sticky="ew",padx=3,pady=(0,2))

        tk.Label(tf,text="Override",**LBL).grid(row=0,column=3,sticky="ew",padx=3,**PADY)
        gst_override_cb = ttk.Combobox(tf, textvariable=gst_type_var, state="readonly",
                     values=["CGST+SGST","IGST"])
        gst_override_cb.grid(row=1,column=3,sticky="ew",padx=3,pady=(0,2),ipady=1)
        self._sale_gst_override_cb = gst_override_cb
        self._gst_type_var = gst_type_var

        # ── Walk-in Customer: compact 4-column grid ──────────────────────────
        walkin_frame = tk.Frame(p, bg="#FFFBEA",
                                highlightthickness=1, highlightbackground="#F6C90E")
        walkin_frame.pack(fill="x", pady=(1, 1))

        WLG = {"bg":"#FFFBEA"}
        wf_inner = tk.Frame(walkin_frame, **WLG)
        wf_inner.pack(fill="x", padx=6, pady=2)
        # 4 columns: name (under Walk-in label) | mobile | address | gstin
        wf_inner.columnconfigure(0, weight=2)
        wf_inner.columnconfigure(1, weight=1)
        wf_inner.columnconfigure(2, weight=2)
        wf_inner.columnconfigure(3, weight=1)

        tk.Label(wf_inner, text="👤 Walk-in / Name:", font=("Segoe UI",8,"bold"),
                 **WLG, fg="#92400E").grid(row=0, column=0, sticky="w", padx=(4,2))
        tk.Label(wf_inner, text="Mobile", font=("Segoe UI",7), **WLG, fg="#78350F"
                 ).grid(row=0, column=1, sticky="w", padx=(0,2))
        tk.Label(wf_inner, text="Address", font=("Segoe UI",7), **WLG, fg="#78350F"
                 ).grid(row=0, column=2, sticky="w", padx=(0,2))
        tk.Label(wf_inner, text="GSTIN", font=("Segoe UI",7), **WLG, fg="#78350F"
                 ).grid(row=0, column=3, sticky="w", padx=(0,2))

        self._s_walkin_name  = tk.StringVar()
        self._s_walkin_mob   = tk.StringVar()
        self._s_walkin_addr  = tk.StringVar()
        self._s_walkin_gstin = tk.StringVar()
        self._s_party = self._s_walkin_name   # alias — naya alag Party field nahi hai ab

        walkin_name_entry = ttk.Entry(wf_inner, textvariable=self._s_walkin_name, font=("Segoe UI",9))
        walkin_name_entry.grid(row=1, column=0, sticky="ew", padx=(4,4), ipady=2)
        walkin_mob_entry = ttk.Entry(wf_inner, textvariable=self._s_walkin_mob, font=("Segoe UI",9))
        walkin_mob_entry.grid(row=1, column=1, sticky="ew", padx=(0,4), ipady=2)
        walkin_addr_entry = ttk.Entry(wf_inner, textvariable=self._s_walkin_addr, font=("Segoe UI",9))
        walkin_addr_entry.grid(row=1, column=2, sticky="ew", padx=(0,4), ipady=2)
        walkin_gstin_entry = ttk.Entry(wf_inner, textvariable=self._s_walkin_gstin, font=("Segoe UI",9))
        walkin_gstin_entry.grid(row=1, column=3, sticky="ew", padx=(0,4), ipady=2)

        # Walk-in fields autocomplete
        add_entry_autocomplete(walkin_name_entry, lambda: [r[0] for r in get_db().execute("SELECT DISTINCT name FROM parties WHERE ptype IN ('Customer','Both') ORDER BY name").fetchall()])
        add_entry_autocomplete(walkin_mob_entry,  lambda: [r[0] for r in get_db().execute("SELECT DISTINCT mobile FROM parties WHERE mobile!='' ORDER BY mobile").fetchall()])
        add_entry_autocomplete(walkin_addr_entry, lambda: [r[0] for r in get_db().execute("SELECT DISTINCT address FROM parties WHERE address!='' ORDER BY address").fetchall()])
        add_entry_autocomplete(walkin_gstin_entry,lambda: [r[0] for r in get_db().execute("SELECT DISTINCT gstin FROM parties WHERE gstin!='' ORDER BY gstin").fetchall()])

        # ── Save as Regular Customer checkbox — DEFAULT CHECKED ──────────────
        self._s_save_as_customer = tk.BooleanVar(value=True)
        save_cust_chk = tk.Checkbutton(
            wf_inner, text="✅ Save as Regular Customer (one-time customer ho to uncheck karo)",
            variable=self._s_save_as_customer, font=("Segoe UI",7,"bold"),
            bg="#FFFBEA", fg="#92400E", activebackground="#FFFBEA",
            selectcolor="#FFF3C4", anchor="w", cursor="hand2")
        save_cust_chk.grid(row=2, column=0, columnspan=4, sticky="w", padx=(4,0), pady=(1,0))

        def _detect_gst_type(ev=None):
            typed_gstin = self._s_walkin_gstin.get().strip().upper()
            name = self._s_walkin_name.get().strip()
            conn2 = get_db()
            pr = conn2.execute("SELECT state, gstin FROM parties WHERE name=?", (name,)).fetchone()
            conn2.close()
            gstin = typed_gstin or ((pr["gstin"] or "").strip() if pr else "")
            p_state = (pr["state"] or "").strip() if pr else ""
            if gstin and len(gstin) >= 2:
                if gstin[:2] != OWN_GST_CODE:
                    gst_type_var.set("IGST")
                    gst_state_lbl.config(text=f"🔵 IGST (Interstate — GSTIN {gstin[:2]})", fg="#6B46C1")
                else:
                    gst_type_var.set("CGST+SGST")
                    gst_state_lbl.config(text="🟢 CGST + SGST (Intrastate — UP)", fg="#276749")
            elif p_state and p_state != OWN_STATE:
                gst_type_var.set("IGST")
                gst_state_lbl.config(text=f"🔵 IGST (Interstate — {p_state})", fg="#6B46C1")
            else:
                gst_type_var.set("CGST+SGST")
                gst_state_lbl.config(text="🟢 CGST + SGST (default)", fg="#276749")

        # Jab walk-in Name field se hatein (FocusOut) — agar existing party hai to fields auto-fill karo
        def _fill_walkin_from_party(ev=None):
            nm = self._s_walkin_name.get().strip()
            conn2 = get_db()
            row2 = conn2.execute("SELECT name,mobile,address,gstin FROM parties WHERE name=?", (nm,)).fetchone()
            conn2.close()
            if row2:
                if not self._s_walkin_mob.get().strip():
                    self._s_walkin_mob.set(row2["mobile"] or "")
                if not self._s_walkin_addr.get().strip():
                    self._s_walkin_addr.set(row2["address"] or "")
                if not self._s_walkin_gstin.get().strip():
                    self._s_walkin_gstin.set(row2["gstin"] or "")
            _detect_gst_type(ev)

        walkin_name_entry.bind("<FocusOut>", _fill_walkin_from_party)
        self._s_walkin_gstin.trace_add("write", lambda *a: _detect_gst_type())

        # ══ ADD ITEM — Sale Bill ══════════════════════════════════════════════
        # Layout: 8 equal columns, weight=1 each. har box same size.
        # Row 0: labels | Row 1: fields | Row 2: labels | Row 3: fields | Row 4: +Add button
        af = tk.Frame(p, bg=C_LIGHT, highlightthickness=1, highlightbackground=C_BORDER)
        af.pack(fill="x", pady=(0,2))
        AF = {"bg": C_LIGHT}
        NCOLS = 14
        for i in range(NCOLS-1):
            af.columnconfigure(i, weight=1, uniform="sal_af")
        af.columnconfigure(NCOLS-1, weight=1, uniform=None)

        # Uniform label / entry padding
        PL = {"padx":(2,2), "pady":(4,1), "sticky":"w"}
        PE = {"padx":(2,2), "pady":(0,4), "sticky":"ew", "ipady":1}

        # ── SINGLE ROW: Product(2col) | Batch | Expiry | MRP | Qty | +Free | Rate | GST | Packing | HSN | Discount | MfgCo | +Add ──
        tk.Label(af,text="Product / Medicine *",font=("Segoe UI",8,"bold"),**AF,fg="#1A365D"
                 ).grid(row=0,column=0,columnspan=2,**PL)
        self._s_prod = tk.StringVar()
        prod_cb = ttk.Combobox(af,textvariable=self._s_prod,
                               values=[pr["name"] for pr in all_products],font=("Segoe UI",9),width=14)
        prod_cb.grid(row=1,column=0,columnspan=2,**PE)
        self._sale_prod_cb = prod_cb
        add_autocomplete(prod_cb, lambda: [r["name"] for r in get_db().execute(
            "SELECT name FROM products ORDER BY name").fetchall()])

        tk.Label(af,text="Batch No",font=("Segoe UI",7),**AF,fg=C_GRAY).grid(row=0,column=2,**PL)
        self._s_batch_no = tk.StringVar()
        _batch_entry = ttk.Entry(af,textvariable=self._s_batch_no,font=("Segoe UI",9),width=6)
        _batch_entry.grid(row=1,column=2,**PE)
        add_entry_autocomplete(_batch_entry, lambda: [r[0] for r in get_db().execute(
            "SELECT DISTINCT batch_no FROM sale_items WHERE batch_no!='' ORDER BY batch_no").fetchall()])

        tk.Label(af,text="Expiry (MM/YYYY)",font=("Segoe UI",7),**AF,fg="#9B2C2C").grid(row=0,column=3,**PL)
        self._s_expiry_date = tk.StringVar()
        _s_exp_raw = tk.StringVar(value="")

        def _s_parse_exp_raw(raw):
            raw = raw.strip()
            import re as _re
            m = _re.match(r"^(\d{1,2})/(\d{2}|\d{4})$", raw)
            if m:
                mm, yy = m.group(1).zfill(2), m.group(2)
                yy = yy[-2:].zfill(2)
                if 1 <= int(mm) <= 12:
                    return f"{mm}/{yy}"
            return ""

        def _s_on_exp_change(*args):
            self._s_expiry_date.set(_s_parse_exp_raw(_s_exp_raw.get()))

        def _s_exp_key(event):
            w = event.widget
            val = _s_exp_raw.get()
            if len(val) == 2 and val.isdigit() and event.char not in ("", "/", "\b"):
                _s_exp_raw.set(val + "/")
                w.icursor(3)
                return "break"

        _s_exp_raw.trace_add("write", _s_on_exp_change)
        import datetime as _sdt
        _s_cur_yr = _sdt.date.today().year
        _s_exp_suggestions = [f"{str(m).zfill(2)}/{y}"
                               for y in range(_s_cur_yr, _s_cur_yr+10)
                               for m in range(1,13)]
        _exp_entry = ttk.Combobox(af,textvariable=_s_exp_raw,values=_s_exp_suggestions,font=("Segoe UI",9),width=8)
        _exp_entry.grid(row=1,column=3,padx=(2,2),pady=(0,4),sticky="ew")
        _exp_entry.bind("<KeyPress>", _s_exp_key)
        add_autocomplete(_exp_entry, lambda: _s_exp_suggestions)
        self._s_exp_raw = _s_exp_raw

        tk.Label(af,text="MRP (₹)",font=("Segoe UI",7),**AF,fg=C_GRAY).grid(row=0,column=4,**PL)
        self._s_mrp = tk.StringVar()
        _mrp_entry = ttk.Entry(af,textvariable=self._s_mrp,font=("Segoe UI",9),width=6)
        _mrp_entry.grid(row=1,column=4,**PE)

        tk.Label(af,text="Qty",font=("Segoe UI",8,"bold"),**AF,fg="#1A365D").grid(row=0,column=5,**PL)
        self._s_qty = tk.StringVar()
        qty_entry = ttk.Entry(af,textvariable=self._s_qty,font=("Segoe UI",9),width=5)
        qty_entry.grid(row=1,column=5,**PE)
        self._sale_qty_entry = qty_entry

        tk.Label(af,text="+Free",font=("Segoe UI",7),**AF,fg="#C53030").grid(row=0,column=6,**PL)
        self._s_free_qty = tk.StringVar(value="0")
        free_qty_entry = ttk.Entry(af,textvariable=self._s_free_qty,font=("Segoe UI",9),width=3)
        free_qty_entry.grid(row=1,column=6,**PE)

        tk.Label(af,text="Rate (₹) *",font=("Segoe UI",8,"bold"),**AF,fg="#1A365D").grid(row=0,column=7,**PL)
        self._s_rate = tk.StringVar()
        _rate_entry = ttk.Entry(af,textvariable=self._s_rate,font=("Segoe UI",9),width=6)
        _rate_entry.grid(row=1,column=7,**PE)

        tk.Label(af,text="GST %",font=("Segoe UI",7),**AF,fg=C_GRAY).grid(row=0,column=8,**PL)
        self._s_gst = tk.StringVar(value="5")
        _gst_cb = ttk.Combobox(af,textvariable=self._s_gst,values=["0","5","12","18","28"],
                     font=("Segoe UI",9),state="readonly",width=4)
        _gst_cb.grid(row=1,column=8,**PE)
        self._sale_gst_cb = _gst_cb

        tk.Label(af,text="Packing",font=("Segoe UI",7),**AF,fg=C_GRAY).grid(row=0,column=9,**PL)
        self._s_packing = tk.StringVar()
        _packing_entry = ttk.Entry(af,textvariable=self._s_packing,font=("Segoe UI",9),width=6)
        _packing_entry.grid(row=1,column=9,**PE)
        add_entry_autocomplete(_packing_entry, lambda: [r[0] for r in get_db().execute(
            "SELECT DISTINCT packing FROM products WHERE packing!='' ORDER BY packing").fetchall()])

        tk.Label(af,text="HSN/SAC",font=("Segoe UI",7),**AF,fg=C_GRAY).grid(row=0,column=10,**PL)
        self._s_hsn = tk.StringVar()
        _hsn_entry = ttk.Entry(af,textvariable=self._s_hsn,font=("Segoe UI",9),width=6)
        _hsn_entry.grid(row=1,column=10,**PE)
        add_entry_autocomplete(_hsn_entry, lambda: [r[0] for r in get_db().execute(
            "SELECT DISTINCT hsn FROM products WHERE hsn!='' ORDER BY hsn").fetchall()])

        tk.Label(af,text="Discount",font=("Segoe UI",7),**AF,fg="#C53030").grid(row=0,column=11,**PL)
        disc_inner = tk.Frame(af,**AF)
        disc_inner.grid(row=1,column=11,padx=(2,2),pady=(0,4),sticky="ew")
        self._s_disc = tk.StringVar(value="0")
        self._s_disc_type = tk.StringVar(value="%")
        _disc_entry = ttk.Entry(disc_inner,textvariable=self._s_disc,font=("Segoe UI",9),width=4)
        _disc_entry.pack(side="left",ipady=1)
        def _toggle_disc_type():
            self._s_disc_type.set("₹" if self._s_disc_type.get()=="%" else "%")
            _disc_type_btn.config(text=self._s_disc_type.get())
        _disc_type_btn = tk.Button(disc_inner,textvariable=self._s_disc_type,
                                   text="%",font=("Segoe UI",10,"bold"),
                                   bg="#C53030",fg="white",relief="flat",
                                   cursor="hand2",bd=0,width=2,
                                   command=_toggle_disc_type)
        _disc_type_btn.pack(side="left",padx=(2,0),fill="y")

        tk.Label(af,text="Mfg Company",font=("Segoe UI",7),**AF,fg=C_GRAY).grid(row=0,column=12,**PL)
        self._s_mfg_company = tk.StringVar()
        _mfg_co_entry = ttk.Entry(af,textvariable=self._s_mfg_company,font=("Segoe UI",9),width=8)
        _mfg_co_entry.grid(row=1,column=12,**PE)
        add_entry_autocomplete(_mfg_co_entry, lambda: [r[0] for r in get_db().execute(
            "SELECT DISTINCT mfg_company FROM products WHERE mfg_company!='' ORDER BY mfg_company").fetchall()])

        # ── + Add Item button — last column, single row ──────
        make_btn(af, "+Add", self._sale_add_item).grid(row=0,column=13,rowspan=2,padx=(2,4),pady=(4,4),sticky="nsew")

        # ── Enter Key Navigation (Top Section): BillNo→Date→Walkin→GSTIN→GST Override→Product ──
        def _focus_next_top(target_widget, select_all=True):
            real = getattr(target_widget, "_entry", target_widget)
            def _handler(e):
                try:
                    real.focus_set()
                    if select_all and hasattr(real, 'select_range'):
                        real.select_range(0, "end")
                except: pass
                return "break"
            return _handler

        billno_entry.bind("<Return>", _focus_next_top(date_picker))
        getattr(date_picker, "_entry", date_picker).bind("<Return>", _focus_next_top(walkin_name_entry))
        walkin_name_entry.bind("<Return>", _focus_next_top(walkin_mob_entry))
        walkin_mob_entry.bind("<Return>", _focus_next_top(walkin_addr_entry))
        walkin_addr_entry.bind("<Return>", _focus_next_top(walkin_gstin_entry))
        walkin_gstin_entry.bind("<Return>", _focus_next_top(gst_override_cb))
        gst_override_cb.bind("<Return>", _focus_next_top(prod_cb))

        # ── Enter Key Navigation: Product→Qty→Rate→GST→Disc→Batch→Expiry→Add ──
        def _focus_next(target_widget):
            def _handler(e):
                try:
                    target_widget.focus_set()
                    if hasattr(target_widget, 'select_range'):
                        target_widget.select_range(0, "end")
                except: pass
                return "break"
            return _handler

        # Product Enter → Qty pe jaaye (scanner workflow + manual)
        def _prod_enter(e):
            name = self._s_prod.get()
            if name in self._prod_dict:
                pr = self._prod_dict[name]
                self._s_rate.set(str(pr["sale_rate"]))
                if self._s_bill_mode.get() != "NONGST":
                    self._s_gst.set(str(int(pr["gst_percent"])))
                self._s_hsn.set(str(pr.get("hsn","") or ""))
            qty_entry.focus_set()
            qty_entry.select_range(0,"end")
            return "break"
        prod_cb.bind("<Return>", _prod_enter)

        # Qty Enter → Free Qty
        qty_entry.bind("<Return>", _focus_next(free_qty_entry))
        # Free Qty Enter → Rate
        free_qty_entry.bind("<Return>", _focus_next(_rate_entry))
        # Rate Enter → GST
        _rate_entry.bind("<Return>", _focus_next(_gst_cb))
        # GST Enter → Batch
        _gst_cb.bind("<Return>", _focus_next(_batch_entry))
        # Batch Enter → Expiry
        _batch_entry.bind("<Return>", _focus_next(_exp_entry))
        # Expiry Enter → MRP
        _exp_entry.bind("<Return>", _focus_next(_mrp_entry))
        # MRP Enter → Packing (Row 2 shuru)
        _mrp_entry.bind("<Return>", _focus_next(_packing_entry))
        # Packing Enter → HSN
        _packing_entry.bind("<Return>", _focus_next(_hsn_entry))
        # HSN Enter → Discount
        _hsn_entry.bind("<Return>", _focus_next(_disc_entry))
        # Disc Enter → Mfg Company
        _disc_entry.bind("<Return>", _focus_next(_mfg_co_entry))
        # Mfg Company Enter → Add Item
        _mfg_co_entry.bind("<Return>", lambda e: self._sale_add_item())

        def on_prod_select(e=None):
            name = self._s_prod.get()
            if name in self._prod_dict:
                pr = self._prod_dict[name]
                self._s_rate.set(str(pr["sale_rate"]))
                if self._s_bill_mode.get() != "NONGST":
                    self._s_gst.set(str(int(pr["gst_percent"])))
                self._s_hsn.set(str(pr.get("hsn","") or ""))
                try: self._s_mrp.set(str(pr.get("mrp","") or ""))
                except: pass
                # Packing aur Mfg Company bhi products table se
                try: self._s_packing.set(str(pr.get("packing","") or ""))
                except: pass
                try: self._s_mfg_company.set(str(pr.get("mfg_company","") or ""))
                except: pass
            # ── Batch/MFD/Expiry/MRP/Packing/MfgCo auto-fill from expiry_stock (FIFO) ──
            try:
                db = get_db()
                row = db.execute(
                    "SELECT batch_no, mfg_date, expiry_date, mrp, packing, mfg_company "
                    "FROM expiry_stock "
                    "WHERE product=? AND qty>0 ORDER BY " + EXP_KEY_SQL + " ASC, id ASC LIMIT 1",
                    (name,)
                ).fetchone()
                db.close()
                if row:
                    self._s_batch_no.set(row["batch_no"] or "")
                    self._s_expiry_date.set(row["expiry_date"] or "")
                    try:
                        if row["expiry_date"]:
                            self._s_exp_raw.set(fmt_exp_mmyy(row["expiry_date"]))
                        else:
                            self._s_exp_raw.set("")
                    except: pass
                    if row["mrp"]: self._s_mrp.set(str(row["mrp"]))
                    try:
                        if row["packing"]: self._s_packing.set(row["packing"])
                        if row["mfg_company"]: self._s_mfg_company.set(row["mfg_company"])
                    except: pass
                else:
                    self._s_batch_no.set("")
                    self._s_expiry_date.set("")
                    try: self._s_exp_raw.set("")
                    except: pass
            except:
                pass
        prod_cb.bind("<<ComboboxSelected>>", on_prod_select)

        # ── Bottom bar: Totals + Payment + Buttons (table ke UPAR pack hoga) ──
        bottom_bar = tk.Frame(p, bg=C_BG)
        bottom_bar.pack(fill="x", side="bottom")

        # Buttons row
        bf = tk.Frame(bottom_bar, bg=C_BG)
        bf.pack(fill="x", pady=2)
        make_btn(bf,"💾  Save Bill",self._sale_save).pack(side="left",padx=(0,8))
        make_btn(bf,"🖨️  Preview & Print",self._sale_preview,bg=C_PURPLE).pack(side="left",padx=(0,8))
        make_btn(bf,"🗑️  Clear",self._sale_clear,bg=C_GRAY).pack(side="left")

        # Payment Mode row
        spmf = tk.Frame(bottom_bar, bg=C_BG); spmf.pack(fill="x", pady=(2,0))
        tk.Label(spmf, text="Payment Mode:", font=("Segoe UI",10,"bold"), bg=C_BG, fg=C_GRAY
                 ).pack(side="left", padx=(0,8))
        self._s_paymode = tk.StringVar(value="Cash")
        _s_pay_modes = ["Cash", "UPI", "Bank Transfer", "Cheque", "Credit"]
        _s_pay_btns  = []
        for mode in _s_pay_modes:
            rb = tk.Radiobutton(spmf, text=mode, variable=self._s_paymode, value=mode,
                                font=("Segoe UI",9), bg=C_BG, fg=C_GRAY,
                                selectcolor=C_LIGHT, activebackground=C_BG,
                                command=lambda: _s_toggle_due())
            rb.pack(side="left", padx=6)
            _s_pay_btns.append(rb)

        def _s_pay_enter(ev, idx):
            self._s_paymode.set(_s_pay_modes[idx])
            _s_toggle_due()
            if _s_pay_modes[idx] == "Credit":
                self._s_due_entry.focus_set()

        def _s_pay_tab(ev, idx):
            nxt = (idx + 1) % len(_s_pay_btns)
            _s_pay_btns[nxt].focus_set()
            return "break"

        def _s_pay_shift_tab(ev, idx):
            prv = (idx - 1) % len(_s_pay_btns)
            _s_pay_btns[prv].focus_set()
            return "break"

        for i, rb in enumerate(_s_pay_btns):
            rb.bind("<Return>", lambda ev, idx=i: _s_pay_enter(ev, idx))
            rb.bind("<space>",  lambda ev, idx=i: _s_pay_enter(ev, idx))
            rb.bind("<Right>",  lambda ev, idx=i: _s_pay_tab(ev, idx))
            rb.bind("<Left>",   lambda ev, idx=i: _s_pay_shift_tab(ev, idx))

        sduef = tk.Frame(bottom_bar, bg=C_BG); sduef.pack(fill="x", pady=(1,2))
        self._s_due_label = tk.Label(sduef, text="Due Date (YYYY-MM-DD):", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY)
        self._s_due = tk.StringVar()
        self._s_due_entry = make_date_entry(sduef, self._s_due, width=14)
        self._s_due_note = tk.Label(sduef, text="(YYYY-MM-DD format mein bharo)", font=("Segoe UI",7), bg=C_BG, fg="#999")

        def _s_toggle_due():
            mode = self._s_paymode.get()
            if mode == "Credit":
                self._s_due_label.pack(side="left", padx=(0,6))
                self._s_due_entry.pack(side="left", padx=(0,6))
                self._s_due_note.pack(side="left")
            else:
                self._s_due.set("")
                self._s_due_label.pack_forget()
                self._s_due_entry.pack_forget()
                self._s_due_note.pack_forget()
        _s_toggle_due()

        # Totals row (ek hi line mein sab)
        totrow = tk.Frame(bottom_bar, bg=C_WHITE, highlightthickness=1, highlightbackground=C_BORDER)
        totrow.pack(fill="x", pady=2)
        tk.Label(totrow,text="Sub Total:",font=("Segoe UI",9),bg=C_WHITE,fg=C_GRAY
                 ).pack(side="left",padx=8,pady=2)
        self._s_subtotal_var = tk.StringVar(value="₹0")
        tk.Label(totrow,textvariable=self._s_subtotal_var,font=("Segoe UI",11,"bold"),
                 bg=C_WHITE,fg="#4A5568").pack(side="left",padx=(0,16))

        # Item Discount Total
        tk.Label(totrow,text="Item Disc:",font=("Segoe UI",9),bg=C_WHITE,fg="#C53030"
                 ).pack(side="left",padx=(0,4))
        self._s_item_disc_var = tk.StringVar(value="₹0")
        tk.Label(totrow,textvariable=self._s_item_disc_var,font=("Segoe UI",11,"bold"),
                 bg=C_WHITE,fg="#C53030").pack(side="left",padx=(0,16))

        # GST Total
        tk.Label(totrow,text="GST:",font=("Segoe UI",9),bg=C_WHITE,fg=C_GRAY
                 ).pack(side="left",padx=(0,4))
        self._s_gst_total_var = tk.StringVar(value="₹0")
        tk.Label(totrow,textvariable=self._s_gst_total_var,font=("Segoe UI",11,"bold"),
                 bg=C_WHITE,fg="#4A5568").pack(side="left",padx=(0,16))

        # Bill Discount (inline)
        tk.Label(totrow, text="Bill Discount (₹):", font=("Segoe UI",9,"bold"),
                 bg=C_WHITE, fg="#C53030").pack(side="left", padx=(0,4))
        self._s_bill_disc = tk.StringVar(value="0")
        disc_entry = ttk.Entry(totrow, textvariable=self._s_bill_disc, width=8, font=("Segoe UI",9))
        disc_entry.pack(side="left", padx=(0,16))

        # Grand Total (inline)
        tk.Label(totrow,text="Grand Total:",font=("Segoe UI",10,"bold"),bg=C_WHITE,fg=C_GRAY
                 ).pack(side="left",padx=(0,4))
        self._s_total_var = tk.StringVar(value="₹0")
        tk.Label(totrow,textvariable=self._s_total_var,font=("Segoe UI",14,"bold"),
                 bg=C_WHITE,fg=C_ACCENT).pack(side="left")
        self._s_words_var = tk.StringVar()
        tk.Label(totrow,textvariable=self._s_words_var,font=("Segoe UI",8,"italic"),
                 bg=C_WHITE,fg=C_LGRAY).pack(side="left",padx=10)

        def _update_grand(*args):
            try:
                sub = sum(it["grand"] for it in self._sale_items)
                bd  = float(self._s_bill_disc.get() or 0)
                bd  = max(0, min(bd, sub))
                net = sub - bd
            except:
                sub = 0; net = 0
            self._s_subtotal_var.set(f"₹{round(sub):,}")
            self._s_total_var.set(f"₹{round(net):,}")
            self._s_words_var.set(num_to_words(net))
        self._s_bill_disc.trace_add("write", _update_grand)
        self._update_sale_grand = _update_grand

        # ── Items Table — baaki poori jagah le legi ───────────────────────────
        self._s_tbl = make_table(p,
            ["#","Description","HSN/SAC","Batch No","Expiry","Packing","Qty","MRP","Rate","Disc","Disc Amt","Taxable","GST%","GST AMT","Total","Mfg",""],
            [3,13,7,9,9,7,7,8,8,6,8,9,5,8,8,11,4])

    def _sale_add_item(self):
        prod = self._s_prod.get().strip()
        if not prod:
            messagebox.showerror("Error","Product select karo!"); return
        try:
            qty  = float(self._s_qty.get())
            rate = float(self._s_rate.get())
            gst  = float(self._s_gst.get())
            free_qty = float(self._s_free_qty.get() or 0)
        except:
            messagebox.showerror("Error","Qty, Rate aur Free Qty sahi numbers bharo!"); return
        if qty <= 0:
            messagebox.showerror("Error","Qty 0 se zyada honi chahiye!"); return
        if free_qty < 0:
            messagebox.showerror("Error","Free Qty negative nahi ho sakti!"); return

        # ── Expiry check for tracked products ──────────────────────────────
        conn_exp = get_db()
        pr_exp = conn_exp.execute(
            "SELECT track_expiry FROM products WHERE name=?", (prod,)).fetchone()
        if pr_exp and int(pr_exp["track_expiry"] or 0):
            today_key = exp_key_for_date(datetime.date.today())
            soon_key  = exp_key_for_date(datetime.date.today()+datetime.timedelta(days=30))
            expired = conn_exp.execute(
                "SELECT COUNT(*) FROM expiry_stock WHERE product=? AND qty>0 AND " + EXP_KEY_SQL + "<?",
                (prod, today_key)).fetchone()[0]
            soon = conn_exp.execute(
                "SELECT COUNT(*) FROM expiry_stock WHERE product=? AND qty>0 "
                "AND " + EXP_KEY_SQL + ">=? AND " + EXP_KEY_SQL + "<=?",
                (prod, today_key, soon_key)).fetchone()[0]
            total_avail = conn_exp.execute(
                "SELECT COALESCE(SUM(qty),0) FROM expiry_stock WHERE product=? AND qty>0",
                (prod,)).fetchone()[0]
            conn_exp.close()
            if expired > 0:
                ans = messagebox.askyesno("⛔ EXPIRED BATCH WARNING!",
                    f"'{prod}' ki {expired} batch(es) EXPIRE HO GAYI HAI!\n\n"
                    f"Kya aap phir bhi sale bill mein add karna chahte ho?\n"
                    f"(Expiry Manager se expired stock remove karein)")
                if not ans:
                    return
            elif soon > 0:
                messagebox.showwarning("⚠️ Expiry Warning",
                    f"'{prod}' ki {soon} batch(es) 30 din mein expire hone wali hai!\n"
                    f"Dhyan se bechein.")
        else:
            conn_exp.close()

        hsn  = self._s_hsn.get().strip() if hasattr(self,"_s_hsn") else self._prod_dict.get(prod, {}).get("hsn", "")
        unit_val = "Strip"  # Medical: unit fixed
        batch_no_val = getattr(self, "_s_batch_no", None)
        batch_no_val = batch_no_val.get().strip() if batch_no_val else ""
        expiry_date_val = getattr(self, "_s_expiry_date", None)
        expiry_date_val = expiry_date_val.get().strip() if expiry_date_val else ""
        packing_val = getattr(self, "_s_packing", None)
        packing_val = packing_val.get().strip() if packing_val else ""
        mrp_val = getattr(self, "_s_mrp", None)
        try: mrp_val = float(mrp_val.get()) if mrp_val and mrp_val.get().strip() else 0.0
        except: mrp_val = 0.0
        mfg_company_val = getattr(self, "_s_mfg_company", None)
        mfg_company_val = mfg_company_val.get().strip() if mfg_company_val else ""
        try:
            disc_val = float(self._s_disc.get() or 0)
        except:
            disc_val = 0
        disc_type = getattr(self, "_s_disc_type", None)
        disc_type = disc_type.get() if disc_type else "%"
        gross = qty * rate
        if disc_type == "%":
            disc_pct = disc_val
            disc_amt = round(gross * disc_pct / 100, 2)
        else:
            # ₹ mein discount — pehle % nikalo display ke liye
            disc_amt = round(min(disc_val, gross), 2)
            disc_pct = round(disc_amt / gross * 100, 2) if gross else 0
        taxable  = round(gross - disc_amt, 2)
        gst_amt  = round(taxable * gst / 100, 2)
        grand    = round(taxable + gst_amt, 2)
        disc_label = f"{disc_pct}%" if disc_type=="%" else f"₹{disc_amt}"
        self._sale_items.append({
            "product":prod,"hsn":hsn,"unit":unit_val,"qty":qty,"free_qty":free_qty,"rate":rate,
                "batch_no":batch_no_val,"mfg_date":"","expiry_date":expiry_date_val,
                "packing":packing_val,"mrp":mrp_val,"mfg_company":mfg_company_val,
            "disc":disc_pct,"disc_amt":disc_amt,"disc_label":disc_label,
            "taxable":taxable,"gst":gst,"gst_amt":gst_amt,"grand":grand
        })
        self._sale_render()
        self._s_qty.set("")
        self._s_free_qty.set("0")
        self._s_prod.set("")
        self._s_rate.set("")
        self._s_disc.set("0")
        try: self._s_batch_no.set("")
        except: pass
        try: self._s_expiry_date.set("")
        except: pass
        try: self._s_exp_raw.set("")
        except: pass
        try: self._s_mrp.set("")
        except: pass
        try: self._s_packing.set("")
        except: pass
        try: self._s_mfg_company.set("")
        except: pass
        try: self._s_hsn.set("")
        except: pass
        try: self._s_disc_type.set("%"); _disc_type_btn.config(text="%")
        except: pass
        # ── Barcode scanner ke liye: focus wapas product field pe ────────────
        try: self._sale_prod_cb.focus_set()
        except: pass

    def _sale_toggle_bill_mode(self):
        if self._s_bill_mode.get() != "NONGST":
            # ── Switch to Non-GST Bill (Kacha) ──────────────────────────────
            self._s_bill_mode.set("NONGST")
            self._sale_mode_btn.config(
                text="🧾 Non-GST Bill  —  click to switch to GST",
                bg="#C53030")
            self._s_gst.set("0")
            try: self._sale_gst_cb.config(state="disabled")
            except: pass
            try: self._sale_gst_override_cb.config(state="disabled")
            except: pass
            # Cart mein already-added items ka GST bhi 0 kar do
            changed = False
            for it in self._sale_items:
                if it.get("gst", 0) != 0:
                    it["gst"] = 0
                    it["gst_amt"] = 0
                    it["grand"] = it["taxable"]
                    changed = True
            if changed:
                self._sale_render()
        else:
            # ── Switch back to GST Bill (Pakka) ─────────────────────────────
            self._s_bill_mode.set("GST")
            self._sale_mode_btn.config(
                text="✅ GST Bill  —  click to switch to Non-GST",
                bg="#276749")
            self._s_gst.set("5")
            try: self._sale_gst_cb.config(state="readonly")
            except: pass
            try: self._sale_gst_override_cb.config(state="readonly")
            except: pass

    def _sale_render(self):
        clear_table_rows(self._s_tbl)
        total = 0
        for i, it in enumerate(self._sale_items):
            total += it["grand"]
            bg = C_WHITE if i%2==0 else "#F7FAFC"
            disc_pct = it.get("disc", 0)
            disc_amt = it.get("disc_amt", 0)
            disc_show = it.get("disc_label", f"{disc_pct}%" if disc_pct else "-")
            qty_show = it["qty"]
            if it.get("free_qty"):
                qty_show = f"{it['qty']}+{it['free_qty']}F"
            vals = [i+1, it["product"][:14], it["hsn"], it.get("batch_no",""), fmt_exp_mmyy(it.get("expiry_date","")),
                    it.get("packing",""), qty_show,
                    f"₹{it.get('mrp',0):.2f}" if it.get("mrp") else "-",
                    f"₹{it['rate']:.2f}",
                    disc_show if disc_amt else "-",
                    f"₹{disc_amt:.2f}" if disc_amt else "-",
                    f"₹{it['taxable']:.2f}",
                    f"{it['gst']}%", f"₹{it['gst_amt']:.2f}",
                    f"₹{it['grand']:.0f}",
                    it.get("mfg_company","")[:10], "✕"]
            for j, val in enumerate(vals):
                if j == 16:
                    tk.Button(self._s_tbl, text="✕", font=("Segoe UI",9),
                              bg=bg, fg=C_RED, relief="flat", cursor="hand2", bd=0,
                              command=lambda idx=i: self._sale_del(idx)
                              ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=0)
                else:
                    tk.Label(self._s_tbl, text=str(val), font=("Segoe UI",9),
                             bg=bg, fg=C_GRAY, anchor="w", padx=6, pady=4
                             ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=0)
        sub = sum(it["grand"] for it in self._sale_items)
        try:
            bd  = float(self._s_bill_disc.get() or 0)
            bd  = max(0, min(bd, sub))
        except:
            bd = 0
        net = sub - bd
        item_disc_sum = sum(it.get("disc_amt",0) for it in self._sale_items)
        gst_sum = sum(it.get("gst_amt",0) for it in self._sale_items)
        self._s_item_disc_var.set(f"₹{round(item_disc_sum):,}")
        self._s_gst_total_var.set(f"₹{round(gst_sum):,}")
        self._s_subtotal_var.set(f"₹{round(sub):,}")
        self._s_total_var.set(f"₹{round(net):,}")
        self._s_words_var.set(num_to_words(net))

    def _sale_del(self, idx):
        self._sale_items.pop(idx)
        self._sale_render()

    def _sale_save(self):
        if not self._sale_items:
            messagebox.showerror("Error","Koi item nahi add kiya!"); return
        bill_no = self._s_billno.get()
        date    = self._s_date.get()
        party   = self._s_party.get().strip() or "Walk-in"
        sub_total = sum(it["grand"] for it in self._sale_items)
        try:
            bill_disc = float(self._s_bill_disc.get() or 0)
            bill_disc = max(0, min(bill_disc, sub_total))
        except:
            bill_disc = 0
        grand   = round(sub_total - bill_disc, 2)
        gst_type = getattr(self, "_gst_type_var", None)
        gst_type_val = gst_type.get() if gst_type else "CGST+SGST"
        # Walk-in fields
        wk_name  = getattr(self, "_s_walkin_name", None)
        wk_mob   = getattr(self, "_s_walkin_mob",  None)
        wk_addr  = getattr(self, "_s_walkin_addr", None)
        wk_gstin = getattr(self, "_s_walkin_gstin", None)
        walkin_name  = wk_name.get().strip()  if wk_name  else ""
        walkin_mob   = wk_mob.get().strip()   if wk_mob   else ""
        walkin_addr  = wk_addr.get().strip()  if wk_addr  else ""
        walkin_gstin = wk_gstin.get().strip().upper() if wk_gstin else ""
        if walkin_name:
            party = walkin_name
        conn = get_db()
        pr = conn.execute("SELECT * FROM parties WHERE name=?", (self._s_party.get().strip(),)).fetchone()
        p_gstin = walkin_gstin or (pr["gstin"] if pr else "")
        p_addr  = walkin_addr or (pr["address"] if pr else "")
        p_mob   = walkin_mob  or (pr["mobile"]  if pr else "")
        p_state = (pr["state"] if pr and pr["state"] else "") if pr else ""
        try:
            c = conn.cursor()
            s_pm = getattr(self, "_s_paymode", None)
            s_dd = getattr(self, "_s_due", None)
            spm = s_pm.get() if s_pm else "Cash"
            sdd = s_dd.get().strip() if s_dd else ""
            # ── "Save as Regular Customer" checkbox: walk-in ko Parties master mein add karo ──
            save_cust_var = getattr(self, "_s_save_as_customer", None)
            if save_cust_var and save_cust_var.get() and walkin_name:
                existing_cust = c.execute(
                    "SELECT id FROM parties WHERE name=?", (walkin_name,)
                ).fetchone()
                if not existing_cust:
                    c.execute(
                        "INSERT INTO parties(name,ptype,mobile,gstin,address,state,email) "
                        "VALUES(?,?,?,?,?,?,?)",
                        (walkin_name, "Customer", walkin_mob, walkin_gstin, walkin_addr, "Uttar Pradesh", "")
                    )
                else:
                    # Already exists — mobile/address/gstin khali ho to update kar do
                    c.execute(
                        "UPDATE parties SET "
                        "mobile=CASE WHEN (mobile IS NULL OR mobile='') AND ?!='' THEN ? ELSE mobile END, "
                        "address=CASE WHEN (address IS NULL OR address='') AND ?!='' THEN ? ELSE address END, "
                        "gstin=CASE WHEN (gstin IS NULL OR gstin='') AND ?!='' THEN ? ELSE gstin END "
                        "WHERE name=?",
                        (walkin_mob, walkin_mob, walkin_addr, walkin_addr, walkin_gstin, walkin_gstin, walkin_name)
                    )
            c.execute("INSERT INTO sales(bill_no,bill_date,party,party_gstin,party_address,party_mobile,grand_total,pay_mode,due_date,gst_type,party_state,bill_mode) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                      (bill_no,date,party,p_gstin,p_addr,p_mob,grand,spm,sdd,gst_type_val,p_state,
                       getattr(self, "_s_bill_mode", None).get() if getattr(self, "_s_bill_mode", None) else "GST"))
            sid = c.lastrowid
            for it in self._sale_items:
                c.execute("INSERT INTO sale_items(sale_id,product,hsn,unit,qty,rate,taxable,gst_percent,gst_amt,grand,batch_no,mfg_date,expiry_date,disc_amt) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                          (sid,it["product"],it["hsn"],it.get("unit","Pcs"),it["qty"],it["rate"],
                           it["taxable"],it["gst"],it["gst_amt"],it["grand"],
                           it.get("batch_no",""),it.get("mfg_date",""),it.get("expiry_date",""),it.get("disc_amt",0)))
                # ── FIFO consume karo (Qty + Free Qty dono stock se kategi) ──
                fifo_consume(c, it["product"], it["qty"]+it.get("free_qty",0), bill_no, date)
            # ── Auto payment entry — Cash/UPI/Bank/Cheque pe turant paid mark karo ──
            if spm != "Credit":
                c.execute(
                    "INSERT INTO bill_payments(bill_type,bill_no,party,pay_date,amount,pay_mode,note) VALUES(?,?,?,?,?,?,?)",
                    ("sale", bill_no, party, date, grand, spm, "Auto-paid at billing")
                )
            conn.commit()
            messagebox.showinfo("Saved",f"Bill saved!\nBill No: {bill_no}\nTotal: ₹{round(grand):,}")
            self._sale_items = []
            self._show("sale")
        except sqlite3.IntegrityError:
            messagebox.showerror("Error","Bill number already exists! Page refresh karo.")
        finally:
            conn.close()

    def _sale_clear(self):
        self._sale_items = []
        try: self._s_bill_disc.set("0")
        except: pass
        self._sale_render()

    def _sale_preview(self):
        if not self._sale_items:
            messagebox.showerror("Error","Pehle items add karo!"); return
        party_key = self._s_party.get().strip() or "Walk-in"
        conn = get_db()
        pr = conn.execute("SELECT * FROM parties WHERE name=?", (party_key,)).fetchone()
        conn.close()
        pdata = dict(pr) if pr else {"name":party_key,"gstin":"","address":"","mobile":"","state":""}
        wk_name  = getattr(self, "_s_walkin_name", None)
        wk_mob   = getattr(self, "_s_walkin_mob",  None)
        wk_addr  = getattr(self, "_s_walkin_addr", None)
        wk_gstin = getattr(self, "_s_walkin_gstin", None)
        if wk_name and wk_name.get().strip():
            pdata["name"]    = wk_name.get().strip()
        if wk_mob and wk_mob.get().strip():
            pdata["mobile"]  = wk_mob.get().strip()
        if wk_addr and wk_addr.get().strip():
            pdata["address"] = wk_addr.get().strip()
        if wk_gstin and wk_gstin.get().strip():
            pdata["gstin"]   = wk_gstin.get().strip().upper()
        gst_type = getattr(self, "_gst_type_var", None)
        try:
            _bd = float(self._s_bill_disc.get() or 0)
            _bd = max(0, min(_bd, sum(it["grand"] for it in self._sale_items)))
        except:
            _bd = 0
        InvoiceWin(self.root, {
            "bill_no":self._s_billno.get(), "date":self._s_date.get(),
            "party":pdata, "items":self._sale_items,
            "grand":round(sum(it["grand"] for it in self._sale_items) - _bd, 2),
            "bill_disc": _bd,
            "gst_type": gst_type.get() if gst_type else "CGST+SGST",
            "bill_mode": getattr(self, "_s_bill_mode", None).get() if getattr(self, "_s_bill_mode", None) else "GST",
            "label": "BILL (Non-GST)" if (getattr(self, "_s_bill_mode", None) and self._s_bill_mode.get()=="NONGST") else "Tax Invoice"
        })

    def _pur_preview(self):
        if not self._pur_items:
            messagebox.showerror("Error","Pehle items add karo!"); return
        party_key = self._p_party.get().strip() or "Supplier"
        conn = get_db()
        pr = conn.execute("SELECT * FROM parties WHERE name=?", (party_key,)).fetchone()
        conn.close()
        pdata = dict(pr) if pr else {"name":party_key,"gstin":"","address":"","mobile":"","state":""}
        gst_type = getattr(self, "_gst_type_var", None)
        # Convert purchase items to invoice-compatible format
        inv_items = []
        for it in self._pur_items:
            inv_items.append({
                "product":   it["product"],
                "hsn":       it.get("hsn",""),
                "batch_no":  it.get("batch_no",""),
                "expiry_date": it.get("expiry_date",""),
                "packing":   it.get("packing",""),
                "qty":       it["qty"],
                "rate":      it["rate"],
                "mrp":       it.get("exp_mrp",0),
                "disc":      it.get("disc",0),
                "disc_amt":  it.get("disc_amt",0),
                "disc_label":f"{it.get('disc',0)}%",
                "gst":       it["gst"],
                "gst_amt":   it["gst_amt"],
                "taxable":   it["taxable"],
                "grand":     it["total"],
                "mfg_company": it.get("mfg_company",""),
                "free_qty":  0,
            })
        grand = round(sum(it["total"] for it in self._pur_items), 2)
        InvoiceWin(self.root, {
            "bill_no":  self._p_billno.get(),
            "date":     self._p_date.get(),
            "party":    pdata,
            "items":    inv_items,
            "grand":    grand,
            "bill_disc": 0,
            "gst_type": gst_type.get() if gst_type else "CGST+SGST",
            "label":    "Purchase Order"
        })

    # ══════════════════════════════════════════════════════════════════════════
    #  PURCHASE
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_purchase(self):
        self._pur_items = []
        p = tk.Frame(self.content, bg=C_BG, padx=10, pady=3)
        p.pack(fill="both", expand=True)
        section_title(p, "New Purchase")

        conn = get_db()
        all_parties  = [r[0] for r in conn.execute("SELECT name FROM parties ORDER BY name").fetchall()]
        all_products = [dict(r) for r in conn.execute("SELECT * FROM products ORDER BY name").fetchall()]
        conn.close()

        # ── TOP FORM: Purchase No | Date | Supplier | +New Supplier ──────────
        tf = tk.Frame(p, bg=C_BG)
        tf.pack(fill="x", pady=(0,2))
        for i in range(4): tf.columnconfigure(i, weight=1)

        LBL  = {"font":("Segoe UI",8), "bg":C_BG, "fg":C_GRAY, "anchor":"w"}
        PADY = {"pady":(1,0)}

        self._p_billno = tk.StringVar(value=next_pur_no())
        tk.Label(tf,text="Purchase No",**LBL).grid(row=0,column=0,sticky="ew",padx=(6,3),**PADY)
        p_billno_entry = ttk.Entry(tf,textvariable=self._p_billno,font=("Segoe UI",9))
        p_billno_entry.grid(row=1,column=0,sticky="ew",padx=(6,3),pady=(0,2),ipady=1)

        self._p_date = tk.StringVar(value=today_str())
        tk.Label(tf,text="Date (YYYY-MM-DD)",**LBL).grid(row=0,column=1,sticky="ew",padx=3,**PADY)
        p_date_picker = make_date_entry(tf,self._p_date,width=14)
        p_date_picker.grid(row=1,column=1,sticky="ew",padx=3,pady=(0,2))

        self._p_party = tk.StringVar()
        tk.Label(tf,text="Supplier",**LBL).grid(row=0,column=2,sticky="ew",padx=3,**PADY)
        sup_cb = ttk.Combobox(tf,textvariable=self._p_party,values=all_parties,font=("Segoe UI",9))
        sup_cb.grid(row=1,column=2,sticky="ew",padx=3,pady=(0,2),ipady=1)
        add_autocomplete(sup_cb, lambda: [r["name"] for r in get_db().execute("SELECT name FROM parties WHERE ptype IN ('Supplier','Both') ORDER BY name").fetchall()])

        # ── Supplier GST + Contact info display ───────────────────────────────
        sup_info_frame = tk.Frame(tf, bg=C_BG)
        sup_info_frame.grid(row=2, column=2, sticky="w", padx=3, pady=(2,2))
        sup_mobile_lbl = tk.Label(sup_info_frame, text="", font=("Segoe UI",7), bg=C_BG, fg="#276749")
        sup_mobile_lbl.pack(side="left", padx=(0,8))
        sup_gstin_lbl = tk.Label(sup_info_frame, text="", font=("Segoe UI",7), bg=C_BG, fg="#2B6CB0")
        sup_gstin_lbl.pack(side="left")

        def _refresh_sup_info(*args):
            name = self._p_party.get().strip()
            if not name:
                sup_mobile_lbl.config(text="")
                sup_gstin_lbl.config(text="")
                return
            conn2 = get_db()
            row = conn2.execute(
                "SELECT mobile, gstin FROM parties WHERE name=?", (name,)
            ).fetchone()
            conn2.close()
            if row:
                mob  = row["mobile"] or ""
                gst  = row["gstin"]  or ""
                sup_mobile_lbl.config(text=f"📞 {mob}" if mob else "📞 —")
                sup_gstin_lbl.config(text=f"🏷 GST: {gst}" if gst else "🏷 GST: —")
            else:
                sup_mobile_lbl.config(text="")
                sup_gstin_lbl.config(text="")

        self._p_party.trace_add("write", _refresh_sup_info)

        def quick_add_supplier():
            name = self._p_party.get().strip()
            if not name:
                messagebox.showinfo("Quick Add","Pehle Supplier ka naam type karo upar, phir ye button dabaao.")
                return
            conn2 = get_db()
            existing = conn2.execute("SELECT id,mobile,gstin FROM parties WHERE name=?", (name,)).fetchone()

            # ── Dialog for mobile + gstin ─────────────────────────────────────
            dlg = tk.Toplevel()
            dlg.title("Supplier Details")
            dlg.resizable(False, False)
            dlg.configure(bg=C_BG)
            dlg.grab_set()
            W2, H2 = 360, 230
            sw = dlg.winfo_screenwidth(); sh = dlg.winfo_screenheight()
            dlg.geometry(f"{W2}x{H2}+{(sw-W2)//2}+{(sh-H2)//2}")

            tk.Label(dlg, text=f"Supplier: {name}", font=("Segoe UI",10,"bold"),
                     bg=C_BG, fg=C_DARK).pack(pady=(14,8))

            frm = tk.Frame(dlg, bg=C_BG); frm.pack(padx=20, fill="x")

            tk.Label(frm, text="📞 Contact Number:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY
                     ).grid(row=0, column=0, sticky="w", pady=4)
            v_mob = tk.StringVar(value=existing["mobile"] if existing else "")
            ttk.Entry(frm, textvariable=v_mob, width=22).grid(row=0, column=1, padx=8, pady=4)

            tk.Label(frm, text="🏷 GST Number:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY
                     ).grid(row=1, column=0, sticky="w", pady=4)
            v_gst = tk.StringVar(value=existing["gstin"] if existing else "")
            ttk.Entry(frm, textvariable=v_gst, width=22).grid(row=1, column=1, padx=8, pady=4)

            def save_supplier():
                mob2 = v_mob.get().strip()
                gst2 = v_gst.get().strip().upper()
                conn3 = get_db()
                if existing:
                    conn3.execute(
                        "UPDATE parties SET mobile=?, gstin=? WHERE name=?",
                        (mob2, gst2, name))
                else:
                    conn3.execute(
                        "INSERT INTO parties(name,ptype,mobile,gstin,address) VALUES(?,?,?,?,?)",
                        (name, "Supplier", mob2, gst2, ""))
                    all_parties.append(name)
                    sup_cb["values"] = sorted(all_parties)
                conn3.commit(); conn3.close()
                _refresh_sup_info()
                dlg.destroy()
                messagebox.showinfo("Done", f"Supplier '{name}' save ho gaya! ✅")

            bf2 = tk.Frame(dlg, bg=C_BG); bf2.pack(pady=14)
            make_btn(bf2, "✅ Save", save_supplier, bg="#276749").pack(side="left", padx=6)
            make_btn(bf2, "Cancel", dlg.destroy, bg=C_GRAY).pack(side="left", padx=6)

        make_btn(tf, "+ New Supplier", quick_add_supplier, bg="#276749").grid(row=1, column=3, padx=6, sticky="w")


        # ══ ADD ITEM — Purchase ═══════════════════════════════════════════════
        # Row 1: HSN(1) | Batch(1) | Expiry(1) | Product(3) | Packing(1) | Qty(1) | +Add(1) = 9 col
        # Row 2: MRP(1) | PurRate(1) | Disc(1) | GST(1) | SaleRate(1) | MfgCo(3) | [+Add spans] = 9 col
        af = tk.Frame(p, bg=C_LIGHT, highlightthickness=1, highlightbackground=C_BORDER)
        af.pack(fill="x", pady=(0,2))
        AF = {"bg": C_LIGHT}
        NCOLS = 15
        for i in range(NCOLS-1):
            af.columnconfigure(i, weight=1, uniform="pur_af")
        af.columnconfigure(NCOLS-1, weight=1, uniform=None)

        PL = {"padx":(2,2), "pady":(4,1), "sticky":"w"}
        PE = {"padx":(2,2), "pady":(0,4), "sticky":"ew", "ipady":1}

        # ── SINGLE ROW: HSN | Batch | Expiry | Product(2col) | Packing | Qty | MRP | PurRate | Disc | GST | SaleRate | MfgCo | +Add ──
        tk.Label(af,text="HSN/SAC",font=("Segoe UI",7),**AF,fg=C_GRAY).grid(row=0,column=0,**PL)
        self._p_hsn = tk.StringVar()
        _p_hsn_entry = ttk.Entry(af,textvariable=self._p_hsn,font=("Segoe UI",9),width=6)
        _p_hsn_entry.grid(row=1,column=0,**PE)
        add_entry_autocomplete(_p_hsn_entry, lambda: [r[0] for r in get_db().execute(
            "SELECT DISTINCT hsn FROM products WHERE hsn!='' ORDER BY hsn").fetchall()])

        tk.Label(af,text="Batch No",font=("Segoe UI",7),**AF,fg=C_GRAY).grid(row=0,column=1,**PL)
        self._p_batch_no = tk.StringVar()
        _p_batch_entry = ttk.Entry(af,textvariable=self._p_batch_no,font=("Segoe UI",9),width=6)
        _p_batch_entry.grid(row=1,column=1,**PE)
        add_entry_autocomplete(_p_batch_entry, lambda: [r[0] for r in get_db().execute(
            "SELECT DISTINCT batch_no FROM purchase_items WHERE batch_no!='' ORDER BY batch_no").fetchall()])

        tk.Label(af,text="Expiry (MM/YYYY)",font=("Segoe UI",7),**AF,fg="#9B2C2C").grid(row=0,column=2,**PL)
        self._p_expiry_date = tk.StringVar()
        _p_exp_raw = tk.StringVar(value="")

        def _parse_exp_raw(raw):
            raw = raw.strip()
            import re as _re
            m = _re.match(r"^(\d{1,2})/(\d{2}|\d{4})$", raw)
            if m:
                mm, yy = m.group(1).zfill(2), m.group(2)
                yy = yy[-2:].zfill(2)
                if 1 <= int(mm) <= 12:
                    return f"{mm}/{yy}"
            return ""

        def _on_exp_change(*args):
            self._p_expiry_date.set(_parse_exp_raw(_p_exp_raw.get()))

        def _exp_key(event):
            w = event.widget
            val = _p_exp_raw.get()
            if len(val) == 2 and val.isdigit() and event.char not in ("", "/", "\b"):
                _p_exp_raw.set(val + "/")
                w.icursor(3)
                return "break"

        _p_exp_raw.trace_add("write", _on_exp_change)
        import datetime as _dt
        cur_yr = _dt.date.today().year
        exp_suggestions = [f"{str(m).zfill(2)}/{y}"
                           for y in range(cur_yr, cur_yr+10)
                           for m in range(1,13)]
        exp_cb = ttk.Combobox(af,textvariable=_p_exp_raw,values=exp_suggestions,font=("Segoe UI",9),width=6)
        exp_cb.grid(row=1,column=2,**PE)
        exp_cb.bind("<KeyPress>", _exp_key)
        add_autocomplete(exp_cb, lambda: exp_suggestions)
        self._p_exp_raw = _p_exp_raw

        tk.Label(af,text="Product / Medicine *",font=("Segoe UI",8,"bold"),**AF,fg="#1A365D"
                 ).grid(row=0,column=3,columnspan=2,**PL)
        self._p_prod = tk.StringVar()
        auto_titlecase(self._p_prod)
        p_prod_cb = ttk.Combobox(af,textvariable=self._p_prod,
                     values=[pr["name"] for pr in all_products],font=("Segoe UI",9),width=14)
        p_prod_cb.grid(row=1,column=3,columnspan=2,**PE)
        add_autocomplete(p_prod_cb, lambda: [r["name"] for r in get_db().execute(
            "SELECT name FROM products ORDER BY name").fetchall()])

        tk.Label(af,text="Packing",font=("Segoe UI",7),**AF,fg=C_GRAY).grid(row=0,column=5,**PL)
        self._p_packing = tk.StringVar()
        _p_pack_entry = ttk.Entry(af,textvariable=self._p_packing,font=("Segoe UI",9),width=6)
        _p_pack_entry.grid(row=1,column=5,**PE)
        add_entry_autocomplete(_p_pack_entry, lambda: [r[0] for r in get_db().execute(
            "SELECT DISTINCT packing FROM products WHERE packing!='' ORDER BY packing").fetchall()])

        tk.Label(af,text="Qty *",font=("Segoe UI",8,"bold"),**AF,fg="#1A365D").grid(row=0,column=6,**PL)
        self._p_qty = tk.StringVar()
        _p_qty_entry = ttk.Entry(af,textvariable=self._p_qty,font=("Segoe UI",9),width=5)
        _p_qty_entry.grid(row=1,column=6,**PE)

        tk.Label(af,text="+Free",font=("Segoe UI",7),**AF,fg="#C53030").grid(row=0,column=7,**PL)
        self._p_free_qty = tk.StringVar(value="0")
        _p_free_qty_entry = ttk.Entry(af,textvariable=self._p_free_qty,font=("Segoe UI",9),width=3)
        _p_free_qty_entry.grid(row=1,column=7,**PE)

        tk.Label(af,text="MRP (₹) *",font=("Segoe UI",8,"bold"),**AF,fg="#1A365D").grid(row=0,column=8,**PL)
        self._p_exp_mrp = tk.StringVar(value="")
        _p_mrp_entry = ttk.Entry(af,textvariable=self._p_exp_mrp,font=("Segoe UI",9),width=6)
        _p_mrp_entry.grid(row=1,column=8,**PE)

        tk.Label(af,text="Pur. Rate *",font=("Segoe UI",8,"bold"),**AF,fg="#1A365D").grid(row=0,column=9,**PL)
        self._p_rate = tk.StringVar()
        _p_rate_entry = ttk.Entry(af,textvariable=self._p_rate,font=("Segoe UI",9),width=7)
        _p_rate_entry.grid(row=1,column=9,**PE)

        tk.Label(af,text="Discount",font=("Segoe UI",7),**AF,fg="#C53030").grid(row=0,column=10,**PL)
        disc_inner = tk.Frame(af,**AF)
        disc_inner.grid(row=1,column=10,padx=(2,2),pady=(0,4),sticky="ew")
        self._p_disc = tk.StringVar(value="0")
        self._p_disc_type = tk.StringVar(value="%")
        _p_disc_entry = ttk.Entry(disc_inner,textvariable=self._p_disc,font=("Segoe UI",9),width=4)
        _p_disc_entry.pack(side="left",ipady=1)
        def _p_toggle_disc_type():
            self._p_disc_type.set("₹" if self._p_disc_type.get()=="%" else "%")
            _p_disc_type_btn.config(text=self._p_disc_type.get())
        _p_disc_type_btn = tk.Button(disc_inner,textvariable=self._p_disc_type,
                                   text="%",font=("Segoe UI",10,"bold"),
                                   bg="#C53030",fg="white",relief="flat",
                                   cursor="hand2",bd=0,width=2,
                                   command=_p_toggle_disc_type)
        _p_disc_type_btn.pack(side="left",padx=(2,0),fill="y")

        tk.Label(af,text="GST %",font=("Segoe UI",7),**AF,fg=C_GRAY).grid(row=0,column=11,**PL)
        self._p_gst = tk.StringVar(value="5")
        _p_gst_cb = ttk.Combobox(af,textvariable=self._p_gst,values=["0","5","12","18","28"],
                     font=("Segoe UI",9),state="readonly",width=4)
        _p_gst_cb.grid(row=1,column=11,**PE)

        tk.Label(af,text="Sale Rate (₹)",font=("Segoe UI",7),**AF,fg=C_GRAY).grid(row=0,column=12,**PL)
        self._p_sale_rate = tk.StringVar()
        _p_sale_entry = ttk.Entry(af,textvariable=self._p_sale_rate,font=("Segoe UI",9),width=6)
        _p_sale_entry.grid(row=1,column=12,**PE)

        tk.Label(af,text="Mfg Company",font=("Segoe UI",7),**AF,fg=C_GRAY
                 ).grid(row=0,column=13,**PL)
        self._p_mfg_company = tk.StringVar()
        _p_mfg_co_entry = ttk.Entry(af,textvariable=self._p_mfg_company,font=("Segoe UI",9),width=8)
        _p_mfg_co_entry.grid(row=1,column=13,**PE)
        add_entry_autocomplete(_p_mfg_co_entry, lambda: [r[0] for r in get_db().execute(
            "SELECT DISTINCT mfg_company FROM products WHERE mfg_company!='' ORDER BY mfg_company").fetchall()])

        # +Add, single row, last column
        make_btn(af, "+Add", self._pur_add_item).grid(row=0,column=14,rowspan=2,padx=(2,4),pady=(4,4),sticky="nsew")

        self._p_unit = tk.StringVar(value="Strip")

        # ── Enter Navigation (Top Section): BillNo→Date→Supplier→HSN ───────
        def _p_focus_top(target_widget):
            real = getattr(target_widget, "_entry", target_widget)
            def _h(e):
                try:
                    real.focus_set()
                    if hasattr(real,"select_range"): real.select_range(0,"end")
                except: pass
                return "break"
            return _h

        p_billno_entry.bind("<Return>", _p_focus_top(p_date_picker))
        getattr(p_date_picker, "_entry", p_date_picker).bind("<Return>", _p_focus_top(sup_cb))
        sup_cb.bind("<Return>", _p_focus_top(_p_hsn_entry), add="+")

        # ── Enter Navigation: HSN→Batch→Expiry→Product→Packing→Qty→MRP→Rate→Disc→GST→SaleRate→MfgCo→Add ─
        def _p_focus(target):
            def _h(e):
                try:
                    target.focus_set()
                    if hasattr(target,"select_range"): target.select_range(0,"end")
                except: pass
                return "break"
            return _h

        def _p_prod_enter(e):
            name = self._p_prod.get()
            conn2 = get_db()
            pr2 = conn2.execute("SELECT * FROM products WHERE name=?", (name,)).fetchone()
            conn2.close()
            if pr2:
                pr2 = dict(pr2)
                self._p_gst.set(str(int(pr2["gst_percent"])))
                if pr2["purchase_rate"]: self._p_rate.set(str(pr2["purchase_rate"]))
                if pr2.get("mrp"):       self._p_exp_mrp.set(str(pr2["mrp"]))
                if pr2.get("hsn"):       self._p_hsn.set(str(pr2["hsn"]))
                if pr2.get("sale_rate"): self._p_sale_rate.set(str(pr2["sale_rate"]))
            _p_pack_entry.focus_set()
            _p_pack_entry.select_range(0,"end")
            return "break"

        _p_hsn_entry.bind("<Return>",   _p_focus(_p_batch_entry))
        _p_batch_entry.bind("<Return>", _p_focus(exp_cb))
        exp_cb.bind("<Return>",         _p_focus_top(p_prod_cb))
        p_prod_cb.bind("<Return>", _p_prod_enter)
        _p_pack_entry.bind("<Return>",  _p_focus(_p_qty_entry))
        _p_qty_entry.bind("<Return>",   _p_focus(_p_free_qty_entry))
        _p_free_qty_entry.bind("<Return>", _p_focus(_p_mrp_entry))
        _p_mrp_entry.bind("<Return>",   _p_focus(_p_rate_entry))
        _p_rate_entry.bind("<Return>",  _p_focus(_p_disc_entry))
        _p_disc_entry.bind("<Return>",  _p_focus(_p_gst_cb))
        _p_gst_cb.bind("<Return>",      _p_focus(_p_sale_entry))
        _p_sale_entry.bind("<Return>",  _p_focus(_p_mfg_co_entry))
        _p_mfg_co_entry.bind("<Return>", lambda e: self._pur_add_item())

        # Auto-fill from products table on selection
        def on_p_prod_select(e=None):
            name = self._p_prod.get()
            conn2 = get_db()
            pr2 = conn2.execute("SELECT * FROM products WHERE name=?", (name,)).fetchone()
            conn2.close()
            if pr2:
                pr2 = dict(pr2)
                self._p_gst.set(str(int(pr2["gst_percent"])))
                if pr2["purchase_rate"]: self._p_rate.set(str(pr2["purchase_rate"]))
                if pr2.get("mrp"):       self._p_exp_mrp.set(str(pr2["mrp"]))
                if pr2.get("hsn"):       self._p_hsn.set(str(pr2["hsn"]))
                if pr2.get("sale_rate"): self._p_sale_rate.set(str(pr2["sale_rate"]))
        p_prod_cb.bind("<<ComboboxSelected>>", on_p_prod_select)


        # ── Bottom bar: Totals + Payment + Buttons (side="bottom" se pehle pack) ──
        p_bottom = tk.Frame(p, bg=C_BG)
        p_bottom.pack(fill="x", side="bottom")

        # ── 1. Payment Mode (TOP) ────────────────────────────────────────────
        pmf = tk.Frame(p_bottom, bg=C_BG); pmf.pack(fill="x", pady=(4,0))
        tk.Label(pmf, text="Payment Mode:", font=("Segoe UI",10,"bold"), bg=C_BG, fg=C_GRAY
                 ).pack(side="left", padx=(0,8))
        self._p_paymode = tk.StringVar(value="Credit")
        _p_pay_modes = ["Cash", "UPI", "Bank Transfer", "Cheque", "Credit"]
        _p_pay_btns  = []
        for mode in _p_pay_modes:
            rb = tk.Radiobutton(pmf, text=mode, variable=self._p_paymode, value=mode,
                                font=("Segoe UI",9), bg=C_BG, fg=C_GRAY,
                                selectcolor=C_LIGHT, activebackground=C_BG,
                                command=lambda: _toggle_due())
            rb.pack(side="left", padx=6)
            _p_pay_btns.append(rb)

        def _p_pay_enter(ev, idx):
            # Enter/Space se select karo aur next field pe jaao
            _p_pay_modes_local = ["Cash","UPI","Bank Transfer","Cheque","Credit"]
            self._p_paymode.set(_p_pay_modes_local[idx])
            _toggle_due()
            # Credit select hua toh due date pe jaao warna Save button pe
            if _p_pay_modes_local[idx] == "Credit":
                self._due_entry.focus_set()
            else:
                pass  # Tab se aage jayenge

        def _p_pay_tab(ev, idx):
            # Tab/Right Arrow se agla radiobutton
            nxt = (idx + 1) % len(_p_pay_btns)
            _p_pay_btns[nxt].focus_set()
            return "break"

        def _p_pay_shift_tab(ev, idx):
            # Shift+Tab/Left Arrow se pichla radiobutton
            prv = (idx - 1) % len(_p_pay_btns)
            _p_pay_btns[prv].focus_set()
            return "break"

        for i, rb in enumerate(_p_pay_btns):
            rb.bind("<Return>", lambda ev, idx=i: _p_pay_enter(ev, idx))
            rb.bind("<space>",  lambda ev, idx=i: _p_pay_enter(ev, idx))
            rb.bind("<Right>",  lambda ev, idx=i: _p_pay_tab(ev, idx))
            rb.bind("<Left>",   lambda ev, idx=i: _p_pay_shift_tab(ev, idx))

        duef = tk.Frame(p_bottom, bg=C_BG); duef.pack(fill="x", pady=(1,2))
        self._due_label = tk.Label(duef, text="Due Date (YYYY-MM-DD):", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY)
        self._p_due = tk.StringVar()
        self._due_entry = make_date_entry(duef, self._p_due, width=14)
        self._due_note = tk.Label(duef, text="(YYYY-MM-DD format mein bharo)", font=("Segoe UI",7), bg=C_BG, fg="#999")

        def _toggle_due():
            mode = self._p_paymode.get()
            if mode == "Credit":
                self._due_label.pack(side="left", padx=(0,6))
                self._due_entry.pack(side="left", padx=(0,6))
                self._due_note.pack(side="left")
            else:
                self._p_due.set("")
                self._due_label.pack_forget()
                self._due_entry.pack_forget()
                self._due_note.pack_forget()
        _toggle_due()

        # ── 2. Buttons (Save, Preview & Print, Clear) ────────────────────────
        bf = tk.Frame(p_bottom, bg=C_BG); bf.pack(fill="x", pady=2)
        make_btn(bf,"💾  Save Purchase",self._pur_save).pack(side="left",padx=(0,8))
        make_btn(bf,"🗑️  Clear",self._pur_clear,bg=C_GRAY).pack(side="left")

        # Totals Summary (Taxable + Discount + GST + Grand Total)
        totrow = tk.Frame(p_bottom, bg=C_WHITE, highlightthickness=1, highlightbackground=C_BORDER)
        totrow.pack(fill="x", pady=2)

        def _tot_block(label, var, fg=C_GRAY):
            blk = tk.Frame(totrow, bg=C_WHITE)
            blk.pack(side="left", padx=12, pady=4)
            tk.Label(blk,text=label,font=("Segoe UI",8),bg=C_WHITE,fg=C_GRAY).pack(anchor="w")
            tk.Label(blk,textvariable=var,font=("Segoe UI",11,"bold"),bg=C_WHITE,fg=fg).pack(anchor="w")
            return blk

        self._p_total_taxable = tk.StringVar(value="₹0")
        self._p_total_disc    = tk.StringVar(value="₹0")
        self._p_total_gst     = tk.StringVar(value="₹0")
        self._p_total         = tk.StringVar(value="₹0")

        _tot_block("Taxable Amt", self._p_total_taxable)
        _tot_block("Discount", self._p_total_disc, fg="#C53030")
        _tot_block("GST Amt", self._p_total_gst)
        _tot_block("Grand Total (with GST)", self._p_total, fg=C_ACCENT)

        # ── Items Table — baaki poori jagah ───────────────────────────────────
        self._p_tbl = make_table(p,
            ["#","HSN","Batch No","Expiry","Product","Packing","Qty","MRP","Pur Rate","Disc%","GST%","Sale Rate","Taxable","GST Amt","Total","Mfg Co",""],
            [3,7,9,9,15,8,5,8,9,6,5,9,9,8,9,11,4])

    def _pur_add_item(self):
        prod = self._p_prod.get().strip()
        if not prod:
            messagebox.showerror("Error","Product select karo!"); return
        try:
            qty  = float(self._p_qty.get())
            rate = float(self._p_rate.get())
            gst  = float(self._p_gst.get())
            free_qty = float(self._p_free_qty.get() or 0)
        except:
            messagebox.showerror("Error","Qty, Rate aur Free Qty sahi numbers bharo!"); return
        if qty <= 0:
            messagebox.showerror("Error","Qty 0 se zyada honi chahiye!"); return
        if free_qty < 0:
            messagebox.showerror("Error","Free Qty negative nahi ho sakti!"); return

        # Expiry date — optional, but validate format if given (MM/YY)
        exp_str = self._p_expiry_date.get().strip()
        if exp_str:
            import re as _re_chk
            if not _re_chk.match(r"^(0[1-9]|1[0-2])/\d{2}$", exp_str):
                messagebox.showerror("Format Galat!",
                    "Expiry date MM/YY format mein honi chahiye.\nExample: 12/27")
                return

        try:    disc_val = float(self._p_disc.get()) if self._p_disc.get().strip() else 0.0
        except: disc_val = 0.0
        p_disc_type = getattr(self, "_p_disc_type", None)
        p_disc_type = p_disc_type.get() if p_disc_type else "%"

        unit_v   = getattr(self, "_p_unit", None)
        unit_val = unit_v.get() if unit_v else "Strip"
        gross    = round(qty * rate, 2)
        if p_disc_type == "%":
            disc_pct = disc_val
            disc_amt = round(gross * disc_pct / 100, 2)
        else:
            disc_amt = round(min(disc_val, gross), 2)
            disc_pct = round(disc_amt / gross * 100, 2) if gross else 0
        taxable  = round(gross - disc_amt, 2)
        gst_amt  = round(taxable * gst / 100, 2)
        total    = round(taxable + gst_amt, 2)

        batch_str   = self._p_batch_no.get().strip()
        mrp_val     = float(self._p_exp_mrp.get() or 0)
        packing_str = getattr(self,"_p_packing",None)
        packing_str = packing_str.get().strip() if packing_str else ""
        hsn_str     = getattr(self,"_p_hsn",None)
        hsn_str     = hsn_str.get().strip() if hsn_str else ""
        mfg_co_str  = getattr(self,"_p_mfg_company",None)
        mfg_co_str  = mfg_co_str.get().strip() if mfg_co_str else ""
        sale_rate_v = getattr(self,"_p_sale_rate",None)
        try:    sale_rate = float(sale_rate_v.get()) if sale_rate_v and sale_rate_v.get().strip() else 0.0
        except: sale_rate = 0.0

        self._pur_items.append({
            "product"    : prod,
            "unit"       : unit_val,
            "qty"        : qty,
            "free_qty"   : free_qty,
            "rate"       : rate,
            "taxable"    : taxable,
            "gst"        : gst,
            "gst_amt"    : gst_amt,
            "total"      : total,
            "batch_no"   : batch_str,
            "mfg_date"   : "",
            "expiry_date": exp_str,
            "exp_mrp"    : mrp_val,
            "packing"    : packing_str,
            "hsn"        : hsn_str,
            "mfg_company": mfg_co_str,
            "sale_rate"  : sale_rate,
            "disc"       : disc_pct,
            "disc_amt"   : disc_amt,
        })
        self._pur_render()
        self._p_prod.set("")
        self._p_qty.set("")
        self._p_free_qty.set("0")
        self._p_rate.set("")
        self._p_disc.set("0")
        try: self._p_disc_type.set("%"); _p_disc_type_btn.config(text="%")
        except: pass
        self._p_batch_no.set("")
        if hasattr(self, "_p_exp_raw"):  self._p_exp_raw.set("")
        self._p_expiry_date.set("")
        self._p_exp_mrp.set("")
        try: self._p_packing.set("")
        except: pass
        try: self._p_mfg_company.set("")
        except: pass
        try: self._p_sale_rate.set("")
        except: pass

    def _pur_render(self):
        clear_table_rows(self._p_tbl)
        total_sum = 0
        taxable_sum = 0
        disc_sum = 0
        gst_sum = 0
        for i,it in enumerate(self._pur_items):
            total_sum += it["total"]
            taxable_sum += it["taxable"]
            gst_sum += it["gst_amt"]
            gross_i = it["qty"] * it["rate"]
            disc_sum += round(gross_i - it["taxable"], 2)
            bg = C_WHITE if i%2==0 else "#F7FAFC"
            exp_info = ""
            if it.get("expiry_date"):
                exp_info = f"📅{it['expiry_date']}"
                if it.get("batch_no"): exp_info = f"B:{it['batch_no']} " + exp_info
            qty_show = it["qty"]
            if it.get("free_qty"):
                qty_show = f"{it['qty']}+{it['free_qty']}F"
            vals = [i+1, it.get("hsn",""), it.get("batch_no","—"),
                    fmt_exp_mmyy(it.get("expiry_date","")) or "—", it["product"][:14], it.get("packing",""),
                    qty_show,
                    f"₹{it.get('exp_mrp',0):.2f}" if it.get("exp_mrp") else "—",
                    f"₹{it['rate']:.2f}",
                    f"{it.get('disc',0)}%" if it.get("disc") else "—",
                    f"{it['gst']}%",
                    f"₹{it.get('sale_rate',0):.2f}" if it.get("sale_rate") else "—",
                    f"₹{it['taxable']:.2f}",
                    f"₹{it['gst_amt']:.2f}",
                    f"₹{it['total']:.0f}",
                    it.get("mfg_company","")[:10],
                    "✕"]
            for j,val in enumerate(vals):
                if j==16:
                    tk.Button(self._p_tbl,text="✕",font=("Segoe UI",9),
                              bg=bg,fg=C_RED,relief="flat",cursor="hand2",bd=0,
                              command=lambda idx=i: self._pur_del(idx)
                              ).grid(row=i+1,column=j,sticky="nsew",padx=1)
                else:
                    tk.Label(self._p_tbl,text=str(val),font=("Segoe UI",9),
                             bg=bg,fg=C_GRAY,anchor="w",padx=4,pady=4
                             ).grid(row=i+1,column=j,sticky="nsew",padx=1)
        self._p_total_taxable.set(f"₹{round(taxable_sum):,}")
        self._p_total_disc.set(f"₹{round(disc_sum):,}")
        self._p_total_gst.set(f"₹{round(gst_sum):,}")
        self._p_total.set(f"₹{round(total_sum):,}")

    def _pur_del(self,idx):
        self._pur_items.pop(idx); self._pur_render()

    def _pur_save(self):
        if not self._pur_items:
            messagebox.showerror("Error","Koi item nahi!"); return
        bill_no = self._p_billno.get().strip()
        if not bill_no:
            messagebox.showerror("Error","Purchase number bharo!"); return
        date    = self._p_date.get().strip()
        party   = self._p_party.get().strip() or "Supplier"
        grand   = sum(it["total"] for it in self._pur_items)
        conn = get_db()
        try:
            c = conn.cursor()
            pay_mode = getattr(self, "_p_paymode", None)
            due_date = getattr(self, "_p_due", None)
            pm = pay_mode.get() if pay_mode else "Cash"
            dd = due_date.get().strip() if due_date else ""
            c.execute("INSERT INTO purchases(bill_no,bill_date,party,grand_total,pay_mode,due_date) VALUES(?,?,?,?,?,?)",
                      (bill_no,date,party,grand,pm,dd))
            pid = c.lastrowid
            for it in self._pur_items:
                c.execute("""INSERT INTO purchase_items
                    (purchase_id,product,unit,qty,free_qty,rate,taxable,gst_percent,gst_amt,total)
                    VALUES(?,?,?,?,?,?,?,?,?,?)""",
                    (pid, it["product"], it.get("unit","Pcs"), it["qty"], it.get("free_qty",0), it["rate"],
                     it["taxable"], it["gst"], it["gst_amt"], it["total"]))
                # ── Auto-add to products list if not already present ──────────
                existing = c.execute(
                    "SELECT id FROM products WHERE name=?", (it["product"],)
                ).fetchone()
                if not existing:
                    c.execute(
                        "INSERT INTO products(name, hsn, purchase_rate, sale_rate, "
                        "mrp, gst_percent, opening_stock, unit, barcode, track_expiry) "
                        "VALUES(?,?,?,?,?,?,0,?,?,0)",
                        (
                            it["product"],
                            it.get("hsn", ""),
                            it["rate"],
                            it.get("sale_rate", 0) or it["rate"],
                            it.get("exp_mrp", 0),
                            it["gst"],
                            it.get("unit", "Strip"),
                            it.get("barcode", ""),
                        )
                    )
                else:
                    # Product exist karta hai — sab fields update karo
                    c.execute(
                        "UPDATE products SET purchase_rate=?, mrp=?, "
                        "hsn=CASE WHEN hsn='' OR hsn IS NULL THEN ? ELSE hsn END, "
                        "sale_rate=CASE WHEN ? > 0 THEN ? ELSE sale_rate END "
                        "WHERE name=?",
                        (it["rate"], it.get("exp_mrp",0),
                         it.get("hsn",""),
                         it.get("sale_rate",0), it.get("sale_rate",0),
                         it["product"])
                    )
                # ── FIFO Layer add karo (Qty + Free Qty dono stock mein) ──────
                fifo_add_layer(c, it["product"], date, bill_no,
                               it.get("batch_no",""), it["qty"]+it.get("free_qty",0), it["rate"])
                # ── Auto-add to expiry_stock if expiry_date given ──────────────
                if it.get("expiry_date"):
                    # Check karo agar same batch already hai to qty update karo
                    same = c.execute(
                        "SELECT id FROM expiry_stock WHERE product=? AND batch_no=? AND expiry_date=?",
                        (it["product"], it.get("batch_no",""), it["expiry_date"])
                    ).fetchone()
                    if same:
                        c.execute(
                            "UPDATE expiry_stock SET qty=qty+?, purchase_rate=?, mrp=? WHERE id=?",
                            (it["qty"]+it.get("free_qty",0), it["rate"], it.get("exp_mrp",0), same["id"])
                        )
                    else:
                        c.execute(
                            "INSERT INTO expiry_stock(product,batch_no,mfg_date,expiry_date,qty,"
                            "purchase_rate,mrp,supplier,packing,mfg_company) VALUES(?,?,?,?,?,?,?,?,?,?)",
                            (it["product"], it.get("batch_no",""),
                             it.get("mfg_date",""),
                             it["expiry_date"], it["qty"]+it.get("free_qty",0),
                             it["rate"],
                             it.get("exp_mrp", 0),
                             party,
                             it.get("packing",""),
                             it.get("mfg_company",""))
                        )
                    # products table mein bhi packing/mfg_company update karo
                    if it.get("packing") or it.get("mfg_company"):
                        c.execute(
                            "UPDATE products SET "
                            "packing=CASE WHEN ? != '' THEN ? ELSE packing END, "
                            "mfg_company=CASE WHEN ? != '' THEN ? ELSE mfg_company END "
                            "WHERE name=?",
                            (it.get("packing",""), it.get("packing",""),
                             it.get("mfg_company",""), it.get("mfg_company",""),
                             it["product"])
                        )
            # ── Auto payment entry — Cash/UPI/Bank/Cheque pe turant paid mark karo ──
            if pm != "Credit":
                c.execute(
                    "INSERT INTO bill_payments(bill_type,bill_no,party,pay_date,amount,pay_mode,note) VALUES(?,?,?,?,?,?,?)",
                    ("pur", bill_no, party, date, grand, pm, "Auto-paid at billing")
                )
            conn.commit()
            messagebox.showinfo("Saved",
                f"Purchase saved!\n{bill_no}\n"
                f"Taxable: ₹{round(sum(it['taxable'] for it in self._pur_items)):,}\n"
                f"GST: ₹{round(sum(it['gst_amt'] for it in self._pur_items)):,}\n"
                f"Total: ₹{round(grand):,}")
            self._pur_items=[]; self._show("purchase")
        except sqlite3.IntegrityError:
            messagebox.showerror("Error","Bill number already exists! Alag number use karo.")
        finally:
            conn.close()

    def _pur_clear(self):
        self._pur_items=[]; self._pur_render()

    # ══════════════════════════════════════════════════════════════════════════
    #  SALE HISTORY
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_salehistory(self):
        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "Sale History")

        ff = tk.Frame(p, bg=C_BG); ff.pack(fill="x", pady=(0,10))
        tk.Label(ff,text="From:",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY).pack(side="left")
        sh_from = tk.StringVar()
        make_date_entry(ff,sh_from,width=12).pack(side="left",padx=4)
        tk.Label(ff,text="To:",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY).pack(side="left")
        sh_to = tk.StringVar()
        make_date_entry(ff,sh_to,width=12).pack(side="left",padx=4)
        tk.Label(ff,text="Party:",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY).pack(side="left",padx=(10,4))
        sh_party = tk.StringVar()
        conn=get_db()
        parties=[""]+[r[0] for r in conn.execute("SELECT name FROM parties ORDER BY name").fetchall()]
        conn.close()
        ttk.Combobox(ff,textvariable=sh_party,values=parties,width=22,font=("Segoe UI",9)
                     ).pack(side="left",padx=4)

        tbl = make_table(p,
            ["Bill No","Date","Party","Pay Mode","Taxable","Discount","GST","Grand Total","Edit","Delete"],
            [13,11,16,10,11,10,10,12,5,5])

        tot_f = tk.Frame(p,bg=C_WHITE,highlightthickness=1,highlightbackground=C_BORDER)
        tot_f.pack(fill="x",pady=4)
        sh_tot_var = tk.StringVar(value="Total: ₹0.00  |  Bills: 0")
        tk.Label(tot_f,textvariable=sh_tot_var,font=("Segoe UI",11,"bold"),
                 bg=C_WHITE,fg=C_ACCENT).pack(side="left",padx=12,pady=4)

        def load():
            clear_table_rows(tbl)
            frm=sh_from.get(); to=sh_to.get(); party=sh_party.get()
            conn=get_db()
            q=("SELECT s.*, "
               " COALESCE(SUM(si.gst_amt),0) as total_gst,"
               " COALESCE(SUM(si.taxable),0) as total_taxable,"
               " COALESCE(SUM(si.disc_amt),0) as total_disc"
               " FROM sales s LEFT JOIN sale_items si ON s.id=si.sale_id WHERE 1=1")
            params=[]
            if frm: q+=" AND s.bill_date>=?"; params.append(frm)
            if to:  q+=" AND s.bill_date<=?"; params.append(to)
            if party: q+=" AND s.party=?"; params.append(party)
            q+=" GROUP BY s.id ORDER BY s.id DESC"
            rows=conn.execute(q,params).fetchall()
            conn.close()
            grand_sum=0
            for i,r in enumerate(rows):
                grand_sum+=r["grand_total"]
                sid=r["id"]; rr=dict(r)
                bg=C_WHITE if i%2==0 else "#F7FAFC"
                pay_mode = rr.get("pay_mode","—") or "—"
                taxable  = round(rr.get("total_taxable",0) or 0, 0)
                gst_amt  = round(rr.get("total_gst",0) or 0, 0)
                disc_amt = round(rr.get("total_disc",0) or 0, 0)
                vals=[r["bill_no"], fmt_date(r["bill_date"]), r["party"][:14],
                      pay_mode,
                      f"₹{taxable:,.0f}",
                      f"₹{disc_amt:,.0f}" if disc_amt>0 else "—",
                      f"₹{gst_amt:,.0f}",
                      f"₹{r['grand_total']:.0f}",
                      "",""]
                for j,val in enumerate(vals):
                    if j==8:
                        tk.Button(tbl,text="✏",font=("Segoe UI",9),
                                  bg="#2B6CB0",fg=C_WHITE,relief="flat",cursor="hand2",bd=0,pady=3,
                                  command=lambda s=sid: self._edit_sale(s, load)
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=1)
                    elif j==9:
                        tk.Button(tbl,text="🗑",font=("Segoe UI",9),
                                  bg=C_RED,fg=C_WHITE,relief="flat",cursor="hand2",bd=0,pady=3,
                                  command=lambda s=sid,bn=r["bill_no"]: self._delete_sale(s,bn,load)
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=1)
                    else:
                        tk.Label(tbl,text=str(val),font=("Segoe UI",9),
                                 bg=bg,fg=C_GRAY,anchor="w",padx=5,pady=4
                                 ).grid(row=i+1,column=j,sticky="nsew",padx=1)
            sh_tot_var.set(f"Total: ₹{round(grand_sum):,}  |  Bills: {len(rows)}")

        def export_sh():
            try:
                frm=sh_from.get(); to=sh_to.get(); party=sh_party.get()
                conn=get_db()
                q="SELECT s.*, COALESCE(SUM(si.gst_amt),0) as total_gst FROM sales s LEFT JOIN sale_items si ON s.id=si.sale_id WHERE 1=1"
                params=[]
                if frm: q+=" AND s.bill_date>=?"; params.append(frm)
                if to:  q+=" AND s.bill_date<=?"; params.append(to)
                if party: q+=" AND s.party=?"; params.append(party)
                q+=" GROUP BY s.id ORDER BY s.id DESC"
                rows=[dict(r) for r in conn.execute(q,params).fetchall()]; conn.close()
                data=[]
                for r in rows:
                    taxable=r["grand_total"]-r["total_gst"]
                    data.append([r["bill_no"],fmt_date(r["bill_date"]),r["party"],
                                  f"{taxable:.2f}",f"{r['total_gst']:.2f}",f"{r['grand_total']:.0f}",
                                  r.get("pay_mode","") or "",r.get("due_date","") or ""])
                export_to_excel(["Bill No","Date","Party","Taxable","GST Amt","Grand Total","Pay Mode","Due Date"],
                                data, "Sale_History")
            except Exception as e:
                messagebox.showerror("Export Error", str(e))
        make_btn(ff,"🔍  Show",load).pack(side="left",padx=8)
        make_btn(ff,"📥 Excel",export_sh,bg=C_GREEN).pack(side="left",padx=4)
        load()

    def _view_sale(self, sale_id):
        conn = get_db()
        sale  = dict(conn.execute("SELECT * FROM sales WHERE id=?", (sale_id,)).fetchone())
        items = [dict(r) for r in conn.execute("SELECT * FROM sale_items WHERE sale_id=?", (sale_id,)).fetchall()]
        conn.close()
        pdata = {"name":sale["party"],"gstin":sale["party_gstin"],
                 "address":sale["party_address"],"mobile":sale["party_mobile"],
                 "state": sale.get("party_state","") or ""}
        InvoiceWin(self.root, {
            "bill_no":sale["bill_no"], "date":sale["bill_date"],
            "party":pdata, "items":items, "grand":sale["grand_total"],
            "gst_type": sale.get("gst_type","CGST+SGST") or "CGST+SGST"
        })

    def _direct_print_sale(self, sale_id, parent_win=None):
        """Sale bill preview window kholo with Print option."""
        conn = get_db()
        sale  = dict(conn.execute("SELECT * FROM sales WHERE id=?", (sale_id,)).fetchone())
        items = [dict(r) for r in conn.execute("SELECT * FROM sale_items WHERE sale_id=?", (sale_id,)).fetchall()]
        conn.close()
        pdata = {"name": sale["party"], "gstin": sale.get("party_gstin","") or "",
                 "address": sale.get("party_address","") or "",
                 "mobile": sale.get("party_mobile","") or "",
                 "state": sale.get("party_state","") or ""}
        InvoiceWin(parent_win or self.root, {
            "bill_no": sale["bill_no"], "date": sale["bill_date"],
            "party": pdata, "items": items, "grand": sale["grand_total"],
            "bill_disc": sale.get("bill_disc", 0) or 0,
            "gst_type": sale.get("gst_type","CGST+SGST") or "CGST+SGST"
        })

    def _direct_print_purchase(self, pid, parent_win=None):
        """Purchase bill preview window kholo with Print option."""
        conn = get_db()
        pur   = dict(conn.execute("SELECT * FROM purchases WHERE id=?", (pid,)).fetchone())
        items = [dict(r) for r in conn.execute(
            "SELECT * FROM purchase_items WHERE purchase_id=?", (pid,)).fetchall()]
        conn.close()
        pdata = {"name": pur["party"], "gstin": "", "address": "", "mobile": "", "state": ""}
        # convert purchase items to invoice format
        inv_items = []
        for it in items:
            inv_items.append({
                "product":     it["product"],
                "hsn":         it.get("hsn","") or "",
                "batch_no":    it.get("batch_no","") or "",
                "expiry_date": it.get("expiry_date","") or "",
                "packing":     it.get("packing","") or "",
                "qty":         it["qty"],
                "rate":        it["rate"],
                "mrp":         it.get("exp_mrp", 0) or 0,
                "disc":        it.get("disc_pct", 0) or 0,
                "disc_amt":    round(it["qty"]*it["rate"] - it.get("taxable", it["qty"]*it["rate"]), 2),
                "disc_label":  f"{it.get('disc_pct',0)}%",
                "gst":         it.get("gst_percent", 0) or 0,
                "gst_amt":     it.get("gst_amt", 0) or 0,
                "taxable":     it.get("taxable", it["qty"]*it["rate"]),
                "grand":       it.get("total", it["qty"]*it["rate"]),
                "mfg_company": it.get("mfg_company","") or "",
                "free_qty":    0,
            })
        InvoiceWin(parent_win or self.root, {
            "bill_no":   pur["bill_no"],
            "date":      pur["bill_date"],
            "party":     pdata,
            "items":     inv_items,
            "grand":     pur["grand_total"],
            "bill_disc": 0,
            "gst_type":  "CGST+SGST",
            "label":     "Purchase Order"
        })


    def _delete_sale(self, sale_id, bill_no, refresh_cb):
        if not messagebox.askyesno("Delete?",
            f"Sale Bill '{bill_no}' delete karna chahte ho?\nYe permanent hoga!"):
            return
        conn=get_db()
        conn.execute("DELETE FROM sale_items WHERE sale_id=?", (sale_id,))
        conn.execute("DELETE FROM sales WHERE id=?", (sale_id,))
        conn.commit(); conn.close()
        refresh_cb()

    def _edit_sale(self, sale_id, refresh_cb):
        """Edit sale bill — popup with all items editable."""
        conn=get_db()
        sale = dict(conn.execute("SELECT * FROM sales WHERE id=?", (sale_id,)).fetchone())
        items= [dict(r) for r in conn.execute("SELECT * FROM sale_items WHERE sale_id=?", (sale_id,)).fetchall()]
        all_products=[r[0] for r in conn.execute("SELECT name FROM products ORDER BY name").fetchall()]
        all_parties =[r[0] for r in conn.execute("SELECT name FROM parties ORDER BY name").fetchall()]
        prod_dict   ={dict(r)["name"]:dict(r) for r in conn.execute("SELECT * FROM products").fetchall()}
        conn.close()

        win=tk.Toplevel(self.root)
        win.title(f"Edit Sale Bill — {sale['bill_no']}")
        win.state("zoomed"); win.configure(bg=C_BG); win.grab_set()

        # ── Scrollable wrapper so full page is always accessible ──
        _e_canvas = tk.Canvas(win, bg=C_BG, highlightthickness=0)
        _e_vsb = ttk.Scrollbar(win, orient="vertical", command=_e_canvas.yview)
        _e_canvas.configure(yscrollcommand=_e_vsb.set)
        _e_vsb.pack(side="right", fill="y")
        _e_canvas.pack(side="left", fill="both", expand=True)
        _e_inner = tk.Frame(_e_canvas, bg=C_BG)
        _e_win_id = _e_canvas.create_window((0,0), window=_e_inner, anchor="nw")
        def _e_canvas_resize(ev): _e_canvas.itemconfig(_e_win_id, width=ev.width)
        def _e_inner_resize(ev): _e_canvas.configure(scrollregion=_e_canvas.bbox("all"))
        _e_canvas.bind("<Configure>", _e_canvas_resize)
        _e_inner.bind("<Configure>", _e_inner_resize)
        _e_canvas.bind_all("<MouseWheel>",
            lambda ev: _e_canvas.yview_scroll(-1*(ev.delta//120), "units"))

        p=tk.Frame(_e_inner,bg=C_BG,padx=20,pady=14); p.pack(fill="x")
        tk.Label(p,text=f"Sale Bill Edit: {sale['bill_no']}",
                 font=("Segoe UI",14,"bold"),bg=C_BG,fg="#1A365D").pack(anchor="w",pady=(0,10))

        # ── Row 1: Bill No | Date | Party ──
        hf=tk.Frame(p,bg=C_BG); hf.pack(fill="x")
        for i in range(3): hf.columnconfigure(i,weight=1)

        tk.Label(hf,text="Bill No",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY
                 ).grid(row=0,column=0,sticky="w",padx=6,pady=(4,1))
        e_billno=tk.StringVar(value=sale["bill_no"])
        ttk.Entry(hf,textvariable=e_billno,font=("Segoe UI",9)
                  ).grid(row=1,column=0,sticky="ew",padx=6,pady=(0,3))

        tk.Label(hf,text="Date (YYYY-MM-DD)",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY
                 ).grid(row=0,column=1,sticky="w",padx=6,pady=(4,1))
        e_date=tk.StringVar(value=sale["bill_date"])
        make_date_entry(hf,e_date,width=14).grid(row=1,column=1,sticky="w",padx=6,pady=(0,3))

        tk.Label(hf,text="Party",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY
                 ).grid(row=0,column=2,sticky="w",padx=6,pady=(4,1))
        e_party=tk.StringVar(value=sale["party"])
        ttk.Combobox(hf,textvariable=e_party,values=all_parties,font=("Segoe UI",9)
                     ).grid(row=1,column=2,sticky="ew",padx=6,pady=(0,3))

        # ── Row 2: Payment Mode ──
        pmrow=tk.Frame(p,bg=C_BG); pmrow.pack(fill="x",pady=(0,4))
        tk.Label(pmrow,text="Payment Mode:",font=("Segoe UI",10,"bold"),bg=C_BG,fg=C_GRAY
                 ).pack(side="left",padx=(6,10))
        existing_pm=sale.get("pay_mode","Cash") or "Cash"
        e_paymode=tk.StringVar(value=existing_pm)
        for mode in ["Cash","Credit","UPI","Bank Transfer","Cheque"]:
            tk.Radiobutton(pmrow,text=mode,variable=e_paymode,value=mode,
                           bg=C_BG,fg=C_DARK,font=("Segoe UI",9),
                           activebackground=C_BG,
                           command=lambda: _toggle_due()
                           ).pack(side="left",padx=6)

        # ── Row 3: Due Date (Credit pe hi dikhega) ──
        duerow=tk.Frame(p,bg=C_BG); duerow.pack(fill="x",pady=(0,3))
        due_lbl=tk.Label(duerow,text="Due Date (YYYY-MM-DD):",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY)
        e_due=tk.StringVar(value=sale.get("due_date","") or "")
        due_entry=make_date_entry(duerow,e_due,width=16)

        def _toggle_due():
            if e_paymode.get()=="Credit":
                due_lbl.pack(side="left",padx=(6,6))
                due_entry.pack(side="left")
            else:
                due_lbl.pack_forget()
                due_entry.pack_forget()
        _toggle_due()

        # Items editor
        edit_items=[dict(it) for it in items]

        tk.Label(p,text="Items",font=("Segoe UI",10,"bold"),bg=C_BG,fg="#1A365D"
                 ).pack(anchor="w",pady=(6,4))

        tbl_f=tk.Frame(p,bg=C_BG); tbl_f.pack(fill="x")
        item_rows_widgets=[]

        EDIT_UNITS = ["Pcs","Kg","Gm","Litre","ml","Dozen","Box","Packet","Bag","Bottle","Carton","Meter","Feet","Set","Pair"]
        try:
            _cu3 = get_db()
            _cu3_list = [r[0] for r in _cu3.execute("SELECT unit FROM custom_units ORDER BY unit").fetchall()]
            _cu3.close()
            EDIT_UNITS = list(dict.fromkeys(EDIT_UNITS + _cu3_list))
        except Exception:
            pass

        def render_items():
            for w in tbl_f.winfo_children(): w.destroy()
            item_rows_widgets.clear()
            for ci,col in enumerate(["Product","Unit","Qty","Rate(₹)","GST%","Taxable","GST Amt","Total",""]):
                tk.Label(tbl_f,text=col,font=("Segoe UI",8,"bold"),bg="#DBEAFE",
                         fg=C_ACCENT,padx=6,pady=4,anchor="w"
                         ).grid(row=0,column=ci,sticky="ew",padx=1,pady=1)
            for ri,it in enumerate(edit_items):
                row_w={}
                prod_v=tk.StringVar(value=it.get("product",""))
                unit_v=tk.StringVar(value=it.get("unit","Pcs") or "Pcs")
                qty_v =tk.StringVar(value=str(it.get("qty",0)))
                rate_v=tk.StringVar(value=str(it.get("rate",0)))
                gst_v =tk.StringVar(value=str(int(it.get("gst_percent",18))))

                cb=ttk.Combobox(tbl_f,textvariable=prod_v,values=all_products,width=18,font=("Segoe UI",9))
                cb.grid(row=ri+1,column=0,sticky="ew",padx=1,pady=1)
                ttk.Combobox(tbl_f,textvariable=unit_v,values=EDIT_UNITS,width=6,state="readonly",
                             font=("Segoe UI",9)).grid(row=ri+1,column=1,sticky="ew",padx=1,pady=1)
                ttk.Entry(tbl_f,textvariable=qty_v,width=6,font=("Segoe UI",9)
                          ).grid(row=ri+1,column=2,sticky="ew",padx=1,pady=1)
                ttk.Entry(tbl_f,textvariable=rate_v,width=9,font=("Segoe UI",9)
                          ).grid(row=ri+1,column=3,sticky="ew",padx=1,pady=1)
                ttk.Combobox(tbl_f,textvariable=gst_v,values=["0","5","12","18","28"],
                             width=4,state="readonly",font=("Segoe UI",9)
                             ).grid(row=ri+1,column=4,sticky="ew",padx=1,pady=1)

                try:
                    qty=float(qty_v.get()); rate=float(rate_v.get()); gst=float(gst_v.get())
                    taxbl=qty*rate; gst_a=taxbl*gst/100; grnd=taxbl+gst_a
                except: taxbl=gst_a=grnd=0

                tk.Label(tbl_f,text=f"{taxbl}",font=("Segoe UI",9),bg=C_WHITE,padx=4,pady=3
                         ).grid(row=ri+1,column=5,sticky="ew",padx=1,pady=1)
                tk.Label(tbl_f,text=f"{gst_a}",font=("Segoe UI",9),bg=C_WHITE,padx=4,pady=3
                         ).grid(row=ri+1,column=6,sticky="ew",padx=1,pady=1)
                tk.Label(tbl_f,text=f"{grnd}",font=("Segoe UI",9),bg=C_WHITE,padx=4,pady=3
                         ).grid(row=ri+1,column=7,sticky="ew",padx=1,pady=1)
                tk.Button(tbl_f,text="✕",font=("Segoe UI",9),bg=C_RED,fg=C_WHITE,
                          relief="flat",cursor="hand2",bd=0,
                          command=lambda idx=ri: (edit_items.pop(idx), render_items())
                          ).grid(row=ri+1,column=8,sticky="nsew",padx=1,pady=1)
                row_w.update({"prod":prod_v,"unit":unit_v,"qty":qty_v,"rate":rate_v,"gst":gst_v})
                item_rows_widgets.append(row_w)

        render_items()

        # Add new item row
        af=tk.Frame(p,bg=C_LIGHT,highlightthickness=1,highlightbackground=C_BORDER)
        af.pack(fill="x",pady=4)
        for i in range(6): af.columnconfigure(i,weight=1)
        tk.Label(af,text="New Item:",font=("Segoe UI",9,"bold"),bg=C_LIGHT,fg=C_ACCENT
                 ).grid(row=0,column=0,sticky="w",padx=6,pady=(2,0))
        n_prod=tk.StringVar(); n_unit=tk.StringVar(value="Pcs"); n_qty=tk.StringVar(); n_rate=tk.StringVar(); n_gst=tk.StringVar(value="5")
        ttk.Combobox(af,textvariable=n_prod,values=all_products,width=20,font=("Segoe UI",9)
                     ).grid(row=1,column=0,sticky="ew",padx=4,pady=(0,3))
        ttk.Combobox(af,textvariable=n_unit,values=EDIT_UNITS,width=7,state="readonly",font=("Segoe UI",9)
                     ).grid(row=1,column=1,sticky="ew",padx=4,pady=(0,3))
        ttk.Entry(af,textvariable=n_qty,width=6,font=("Segoe UI",9)
                  ).grid(row=1,column=2,sticky="ew",padx=4,pady=(0,3))
        ttk.Entry(af,textvariable=n_rate,width=9,font=("Segoe UI",9)
                  ).grid(row=1,column=3,sticky="ew",padx=4,pady=(0,3))
        ttk.Combobox(af,textvariable=n_gst,values=["0","5","12","18","28"],width=4,state="readonly",
                     font=("Segoe UI",9)).grid(row=1,column=4,sticky="ew",padx=4,pady=(0,3))

        def add_new_item():
            if not n_prod.get(): messagebox.showerror("Error","Product select karo!",parent=win); return
            try: qty=float(n_qty.get()); rate=float(n_rate.get()); gst=float(n_gst.get())
            except: messagebox.showerror("Error","Qty/Rate sahi bharo!",parent=win); return
            taxbl=qty*rate; gst_a=taxbl*gst/100
            edit_items.append({"product":n_prod.get(),"unit":n_unit.get(),"qty":qty,"rate":rate,
                                "gst_percent":gst,"taxable":taxbl,"gst_amt":gst_a,"grand":taxbl+gst_a,
                                "hsn":prod_dict.get(n_prod.get(),{}).get("hsn","")})
            n_prod.set(""); n_qty.set(""); n_rate.set("")
            render_items()

        make_btn(af, "+ Add", add_new_item).grid(row=1,column=5,padx=8,pady=(0,3),sticky="ew")

        def collect_and_save():
            """Read current widget values back into edit_items then save."""
            updated=[]
            for w in item_rows_widgets:
                try:
                    prod=w["prod"].get().strip()
                    unit=w.get("unit",tk.StringVar(value="Pcs")).get()
                    qty=float(w["qty"].get()); rate=float(w["rate"].get()); gst=float(w["gst"].get())
                    taxbl=qty*rate; gst_a=taxbl*gst/100
                    updated.append({"product":prod,"unit":unit,"qty":qty,"rate":rate,
                                    "gst_percent":gst,"taxable":taxbl,"gst_amt":gst_a,"grand":taxbl+gst_a,
                                    "hsn":prod_dict.get(prod,{}).get("hsn","")})
                except: pass
            for it in edit_items:
                if not any(u["product"]==it.get("product") for u in updated):
                    updated.append(it)
            if not updated:
                messagebox.showerror("Error","Koi item nahi!",parent=win); return
            grand=sum(it["grand"] for it in updated)
            conn=get_db()
            try:
                conn.execute("UPDATE sales SET bill_no=?,bill_date=?,party=?,grand_total=?,pay_mode=?,due_date=? WHERE id=?",
                             (e_billno.get().strip(),e_date.get().strip(),e_party.get().strip(),
                              grand,e_paymode.get(),e_due.get().strip(),sale_id))
                conn.execute("DELETE FROM sale_items WHERE sale_id=?", (sale_id,))
                for it in updated:
                    conn.execute("INSERT INTO sale_items(sale_id,product,hsn,unit,qty,rate,taxable,gst_percent,gst_amt,grand,batch_no,mfg_date,expiry_date,disc_amt) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                                 (sale_id,it["product"],it.get("hsn",""),it.get("unit","Pcs"),it["qty"],it["rate"],
                                  it["taxable"],it["gst_percent"],it["gst_amt"],it["grand"],
                                  it.get("batch_no",""),it.get("mfg_date",""),it.get("expiry_date",""),it.get("disc_amt",0)))
                conn.commit()
                messagebox.showinfo("Done","Bill update ho gaya!",parent=win)
                win.destroy(); refresh_cb()
            except Exception as e:
                messagebox.showerror("Error",str(e),parent=win)
            finally: conn.close()

        bf2=tk.Frame(p,bg=C_BG); bf2.pack(fill="x",pady=4)
        make_btn(bf2,"💾 Save Changes",collect_and_save).pack(side="left",padx=(0,8))
        make_btn(bf2,"✕ Cancel",win.destroy,bg=C_GRAY).pack(side="left")

    # ══════════════════════════════════════════════════════════════════════════
    #  PURCHASE HISTORY
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_purhistory(self):
        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "Purchase History")

        ff = tk.Frame(p,bg=C_BG); ff.pack(fill="x",pady=(0,10))
        tk.Label(ff,text="From:",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY).pack(side="left")
        ph_from=tk.StringVar()
        make_date_entry(ff,ph_from,width=12).pack(side="left",padx=4)
        tk.Label(ff,text="To:",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY).pack(side="left")
        ph_to=tk.StringVar()
        make_date_entry(ff,ph_to,width=12).pack(side="left",padx=4)
        tk.Label(ff,text="Supplier:",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY).pack(side="left",padx=(10,4))
        ph_party=tk.StringVar()
        conn=get_db()
        parties=[""]+[r[0] for r in conn.execute("SELECT name FROM parties ORDER BY name").fetchall()]
        conn.close()
        ttk.Combobox(ff,textvariable=ph_party,values=parties,width=22,font=("Segoe UI",9)
                     ).pack(side="left",padx=4)

        tbl = make_table(p,
            ["Bill No","Date","Supplier","Pay Mode","Items","Taxable","GST","Discount","Grand Total","View","Edit","Delete"],
            [13,11,16,10,5,11,10,10,12,5,5,5])

        tot_f=tk.Frame(p,bg=C_WHITE,highlightthickness=1,highlightbackground=C_BORDER)
        tot_f.pack(fill="x",pady=4)
        ph_tot=tk.StringVar(value="Total: ₹0.00  |  Bills: 0")
        tk.Label(tot_f,textvariable=ph_tot,font=("Segoe UI",11,"bold"),
                 bg=C_WHITE,fg=C_ACCENT).pack(side="left",padx=12,pady=4)

        def load():
            clear_table_rows(tbl)
            frm=ph_from.get(); to=ph_to.get(); party=ph_party.get()
            conn=get_db()
            q=("SELECT p.*, COUNT(pi.id) as item_count,"
               " COALESCE(SUM(pi.taxable),0) as tot_taxable,"
               " COALESCE(SUM(pi.gst_amt),0) as tot_gst,"
               " COALESCE(SUM(pi.qty*pi.rate)-SUM(COALESCE(pi.taxable,pi.qty*pi.rate)),0) as tot_disc"
               " FROM purchases p LEFT JOIN purchase_items pi ON p.id=pi.purchase_id WHERE 1=1")
            params=[]
            if frm: q+=" AND p.bill_date>=?"; params.append(frm)
            if to:  q+=" AND p.bill_date<=?"; params.append(to)
            if party: q+=" AND p.party=?"; params.append(party)
            q+=" GROUP BY p.id ORDER BY p.id DESC"
            rows=conn.execute(q,params).fetchall()
            conn.close()
            grand_sum=0
            for i,r in enumerate(rows):
                grand_sum+=r["grand_total"]
                pid=r["id"]; rr=dict(r)
                bg=C_WHITE if i%2==0 else "#F7FAFC"
                pay_mode = rr.get("pay_mode","—") or "—"
                taxable  = round(rr.get("tot_taxable",0) or 0, 0)
                gst_amt  = round(rr.get("tot_gst",0) or 0, 0)
                disc_amt = round(max(rr.get("tot_disc",0) or 0, 0), 0)
                vals=[r["bill_no"], fmt_date(r["bill_date"]), r["party"][:14],
                      pay_mode, r["item_count"],
                      f"₹{taxable:,.0f}", f"₹{gst_amt:,.0f}",
                      f"₹{disc_amt:,.0f}" if disc_amt>0 else "—",
                      f"₹{round(r['grand_total']):,}",
                      "","",""]
                for j,val in enumerate(vals):
                    if j==9:
                        tk.Button(tbl,text="👁",font=("Segoe UI",9),
                                  bg=C_AMBER,fg=C_WHITE,relief="flat",cursor="hand2",bd=0,pady=3,
                                  command=lambda pid=pid,rrd=rr: self._view_purchase(pid, rrd)
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=1)
                    elif j==10:
                        tk.Button(tbl,text="✏",font=("Segoe UI",9),
                                  bg="#2B6CB0",fg=C_WHITE,relief="flat",cursor="hand2",bd=0,pady=3,
                                  command=lambda pid=pid: self._edit_purchase(pid, load)
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=1)
                    elif j==11:
                        tk.Button(tbl,text="🗑",font=("Segoe UI",9),
                                  bg=C_RED,fg=C_WHITE,relief="flat",cursor="hand2",bd=0,pady=3,
                                  command=lambda pid=pid,bn=r["bill_no"]: self._delete_purchase(pid,bn,load)
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=1)
                    else:
                        tk.Label(tbl,text=str(val),font=("Segoe UI",9),
                                 bg=bg,fg=C_GRAY,anchor="w",padx=5,pady=4
                                 ).grid(row=i+1,column=j,sticky="nsew",padx=1)
            ph_tot.set(f"Total: ₹{round(grand_sum):,}  |  Bills: {len(rows)}")

        def export_ph():
            try:
                frm=ph_from.get(); to=ph_to.get(); party=ph_party.get()
                conn=get_db()
                q="SELECT p.*, COUNT(pi.id) as item_count FROM purchases p LEFT JOIN purchase_items pi ON p.id=pi.purchase_id WHERE 1=1"
                params=[]
                if frm: q+=" AND p.bill_date>=?"; params.append(frm)
                if to:  q+=" AND p.bill_date<=?"; params.append(to)
                if party: q+=" AND p.party=?"; params.append(party)
                q+=" GROUP BY p.id ORDER BY p.id DESC"
                rows=[dict(r) for r in conn.execute(q,params).fetchall()]; conn.close()
                data=[[r["bill_no"],fmt_date(r["bill_date"]),r["party"],
                       f"{r['grand_total']:.0f}",r["item_count"],
                       r.get("pay_mode","") or "",r.get("due_date","") or ""] for r in rows]
                export_to_excel(["Bill No","Date","Supplier","Grand Total","Items","Pay Mode","Due Date"],
                                data,"Purchase_History")
            except Exception as e:
                messagebox.showerror("Export Error", str(e))
        make_btn(ff,"🔍  Show",load,bg=C_AMBER).pack(side="left",padx=8)
        make_btn(ff,"📥 Excel Export",export_ph,bg=C_GREEN).pack(side="left",padx=4)
        make_btn(ff,"📤 Stock Import (Excel)",lambda: import_purchase_from_excel(load),bg="#C05621").pack(side="left",padx=4)
        make_btn(ff,"📋 Template",lambda: _download_excel_template("purchase"),bg="#6B46C1").pack(side="left",padx=4)
        load()

    def _view_purchase(self, pid, prow):
        conn=get_db()
        items=[dict(r) for r in conn.execute("SELECT * FROM purchase_items WHERE purchase_id=?",(pid,)).fetchall()]
        conn.close()
        win=tk.Toplevel(self.root)
        win.title(f"Purchase Details - {prow['bill_no']}")
        win.geometry("920x560"); win.configure(bg=C_WHITE)
        _apply_logo(win)

        # ── Header ──────────────────────────────────────────────────────────
        hdr=tk.Frame(win,bg="#EBF8FF",pady=10); hdr.pack(fill="x",padx=0)
        tk.Label(hdr,text=f"🧾  Purchase Bill: {prow['bill_no']}",
                 font=("Segoe UI",13,"bold"),bg="#EBF8FF",fg=C_ACCENT).pack(anchor="w",padx=20)
        info_txt=(f"Date: {fmt_date(prow['bill_date'])}    "
                  f"Supplier: {prow['party']}    "
                  f"Pay Mode: {prow.get('pay_mode','—') or '—'}")
        tk.Label(hdr,text=info_txt,font=("Segoe UI",9),bg="#EBF8FF",fg=C_GRAY).pack(anchor="w",padx=20)
        tk.Frame(win,bg=C_BORDER,height=1).pack(fill="x")

        # ── Scrollable Table ─────────────────────────────────────────────────
        cols    = ["#","Product","Batch No","Expiry","Qty","Rate (₹)","Taxable (₹)",
                   "GST %","CGST (₹)","SGST (₹)","GST Amt (₹)","Total (₹)"]
        widths  = [3,22,10,10,5,10,11,6,9,9,10,11]
        inner   = make_table(win, cols, widths)
        inner.master.master.pack(fill="both",expand=True,padx=12,pady=6)

        tot_taxable=0; tot_gst=0; tot_grand=0
        for i,it in enumerate(items):
            gst_pct  = float(it.get("gst_percent") or it.get("gst_pct") or 0)
            gst_amt  = float(it.get("gst_amt") or 0)
            taxable  = float(it.get("taxable") or 0)
            total    = float(it.get("total") or 0)
            half_gst = round(gst_amt/2,2)
            tot_taxable += taxable; tot_gst += gst_amt; tot_grand += total
            batch   = it.get("batch_no","") or "—"
            expiry  = it.get("expiry_date","") or "—"
            qty_disp = it["qty"]
            if it.get("free_qty"):
                qty_disp = f"{it['qty']}+{it['free_qty']}F"
            row_vals=[
                i+1,
                it["product"],
                batch,
                expiry,
                qty_disp,
                f"{it['rate']:.2f}",
                f"{taxable:.2f}",
                f"{gst_pct:.1f}%",
                f"{half_gst:.2f}",
                f"{half_gst:.2f}",
                f"{gst_amt:.2f}",
                f"{total:.2f}",
            ]
            table_row(inner, row_vals, i+1)

        # ── GST Summary Footer ───────────────────────────────────────────────
        tk.Frame(win,bg=C_BORDER,height=1).pack(fill="x",padx=12)
        sumf=tk.Frame(win,bg="#F0FFF4",pady=6); sumf.pack(fill="x",padx=12,pady=(0,4))

        def _sum_lbl(parent, label, value, fg=C_GRAY):
            f=tk.Frame(parent,bg="#F0FFF4"); f.pack(side="left",padx=14)
            tk.Label(f,text=label,font=("Segoe UI",8),bg="#F0FFF4",fg=C_GRAY).pack(anchor="w")
            tk.Label(f,text=value,font=("Segoe UI",10,"bold"),bg="#F0FFF4",fg=fg).pack(anchor="w")

        _sum_lbl(sumf,"Taxable Amount",  f"₹{tot_taxable:.2f}")
        _sum_lbl(sumf,"CGST",            f"₹{tot_gst/2:.2f}")
        _sum_lbl(sumf,"SGST",            f"₹{tot_gst/2:.2f}")
        _sum_lbl(sumf,"Total GST",       f"₹{tot_gst:.2f}", fg="#C05621")
        _sum_lbl(sumf,"Grand Total",     f"₹{round(prow['grand_total']):,}", fg=C_ACCENT)

        # ── Buttons ──────────────────────────────────────────────────────────
        bf=tk.Frame(win,bg=C_WHITE); bf.pack(pady=(2,10))
        make_btn(bf,"✕  Close",win.destroy,bg=C_GRAY).pack(side="left",padx=6)

    def _delete_purchase(self, pid, bill_no, refresh_cb):
        conn=get_db()
        items = [dict(r) for r in conn.execute(
            "SELECT * FROM purchase_items WHERE purchase_id=?", (pid,)).fetchall()]

        # ── Check: kya is purchase ki stock se koi sale ho chuki hai? ──────────
        consumed_layers = conn.execute(
            "SELECT fl.product, fl.qty_in, fl.qty_remaining FROM fifo_layers fl "
            "WHERE fl.purchase_bill=? AND fl.qty_remaining < fl.qty_in",
            (bill_no,)
        ).fetchall()
        if consumed_layers:
            conn.close()
            messagebox.showerror("Delete nahi ho sakta",
                f"Purchase '{bill_no}' ka stock pehle se sale/use ho chuka hai.\n"
                "Pehle related sales delete karo, fir purchase delete karo.")
            return

        if not messagebox.askyesno("Delete?",
            f"Purchase '{bill_no}' delete karna chahte ho?\n"
            "Stock, expiry aur payment entries bhi revert ho jayengi.\nYe permanent hoga!"):
            conn.close()
            return

        try:
            # ── Stock (expiry_stock) se Qty+Free Qty minus karo ────────────────
            for it in items:
                total_qty = (it.get("qty") or 0) + (it.get("free_qty") or 0)
                if it.get("expiry_date"):
                    target_exp = fmt_exp_mmyy(it["expiry_date"])
                    target_batch = (it.get("batch_no") or "").strip()
                    rows = conn.execute(
                        "SELECT id, qty, expiry_date, batch_no FROM expiry_stock WHERE product=?",
                        (it["product"],)
                    ).fetchall()
                    for row in rows:
                        if (row["batch_no"] or "").strip() != target_batch:
                            continue
                        if fmt_exp_mmyy(row["expiry_date"]) != target_exp:
                            continue
                        new_qty = (row["qty"] or 0) - total_qty
                        if new_qty <= 0:
                            conn.execute("DELETE FROM expiry_stock WHERE id=?", (row["id"],))
                        else:
                            conn.execute("UPDATE expiry_stock SET qty=? WHERE id=?", (new_qty, row["id"]))
                        break

            # ── FIFO layers is purchase bill ke remove karo ─────────────────────
            conn.execute("DELETE FROM fifo_layers WHERE purchase_bill=?", (bill_no,))

            # ── Auto-paid payment entry remove karo ─────────────────────────────
            conn.execute("DELETE FROM bill_payments WHERE bill_type='pur' AND bill_no=?", (bill_no,))

            # ── Purchase items aur purchase record hatao ────────────────────────
            conn.execute("DELETE FROM purchase_items WHERE purchase_id=?", (pid,))
            conn.execute("DELETE FROM purchases WHERE id=?", (pid,))
            conn.commit()
        except Exception as e:
            conn.rollback()
            messagebox.showerror("Error", str(e))
        finally:
            conn.close()
        refresh_cb()

    def _edit_purchase(self, pid, refresh_cb):
        conn=get_db()
        pur  = dict(conn.execute("SELECT * FROM purchases WHERE id=?", (pid,)).fetchone())
        items= [dict(r) for r in conn.execute("SELECT * FROM purchase_items WHERE purchase_id=?",(pid,)).fetchall()]
        all_products=[r[0] for r in conn.execute("SELECT name FROM products ORDER BY name").fetchall()]
        all_parties =[r[0] for r in conn.execute("SELECT name FROM parties ORDER BY name").fetchall()]
        conn.close()

        win=tk.Toplevel(self.root)
        win.title(f"Edit Purchase — {pur['bill_no']}")
        win.state("zoomed"); win.configure(bg=C_BG); win.grab_set()

        p=tk.Frame(win,bg=C_BG,padx=20,pady=14); p.pack(fill="x")
        tk.Label(p,text=f"Purchase Edit: {pur['bill_no']}",
                 font=("Segoe UI",14,"bold"),bg=C_BG,fg="#1A365D").pack(anchor="w",pady=(0,10))

        hf=tk.Frame(p,bg=C_BG); hf.pack(fill="x")
        for i in range(2): hf.columnconfigure(i,weight=1)

        tk.Label(hf,text="Date (YYYY-MM-DD)",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY
                 ).grid(row=0,column=0,sticky="w",padx=6,pady=(4,1))
        e_date=tk.StringVar(value=pur["bill_date"])
        make_date_entry(hf,e_date,width=14).grid(row=1,column=0,sticky="w",padx=6,pady=(0,3))

        tk.Label(hf,text="Supplier",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY
                 ).grid(row=0,column=1,sticky="w",padx=6,pady=(4,1))
        e_party=tk.StringVar(value=pur["party"])
        ttk.Combobox(hf,textvariable=e_party,values=all_parties,font=("Segoe UI",9)
                     ).grid(row=1,column=1,sticky="ew",padx=6,pady=(0,3))

        # Payment Mode row
        pmf=tk.Frame(p,bg=C_BG); pmf.pack(fill="x",pady=(0,2))
        tk.Label(pmf,text="Payment Mode:",font=("Segoe UI",10,"bold"),bg=C_BG,fg=C_GRAY
                 ).pack(side="left",padx=(0,8))
        existing_pm = pur.get("pay_mode","Credit") or "Credit"
        e_paymode=tk.StringVar(value=existing_pm)
        for mode in ["Cash","UPI","Bank Transfer","Cheque","Credit"]:
            tk.Radiobutton(pmf,text=mode,variable=e_paymode,value=mode,
                           font=("Segoe UI",9),bg=C_BG,fg=C_GRAY,
                           selectcolor=C_LIGHT,activebackground=C_BG,
                           command=lambda: _etoggle_due()
                           ).pack(side="left",padx=6)

        duef=tk.Frame(p,bg=C_BG); duef.pack(fill="x",pady=(0,3))
        e_due_label=tk.Label(duef,text="Due Date (YYYY-MM-DD):",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY)
        e_due=tk.StringVar(value=pur.get("due_date","") or "")
        e_due_entry=make_date_entry(duef,e_due,width=14)

        def _etoggle_due():
            if e_paymode.get()=="Credit":
                # Auto fill NAHI — blank rakho, manually bharein
                e_due_label.pack(side="left",padx=(0,6))
                e_due_entry.pack(side="left")
            else:
                e_due_label.pack_forget()
                e_due_entry.pack_forget()
        _etoggle_due()

        edit_items=[dict(it) for it in items]

        tk.Label(p,text="Items",font=("Segoe UI",10,"bold"),bg=C_BG,fg="#1A365D"
                 ).pack(anchor="w",pady=(6,4))

        tbl_f=tk.Frame(p,bg=C_BG); tbl_f.pack(fill="x")
        item_rows_widgets=[]

        def render_items():
            for w in tbl_f.winfo_children(): w.destroy()
            item_rows_widgets.clear()
            for ci,col in enumerate(["Product","Qty","Rate (₹)","Disc %","GST %","Batch No","Expiry Date","GST Amt","Total",""]):
                tk.Label(tbl_f,text=col,font=("Segoe UI",8,"bold"),bg="#DBEAFE",
                         fg=C_ACCENT,padx=6,pady=4,anchor="w"
                         ).grid(row=0,column=ci,sticky="ew",padx=1,pady=1)
            for ri,it in enumerate(edit_items):
                prod_v   = tk.StringVar(value=it.get("product",""))
                qty_v    = tk.StringVar(value=str(it.get("qty",0)))
                rate_v   = tk.StringVar(value=str(it.get("rate",0)))
                gst_v    = tk.StringVar(value=str(it.get("gst_percent",it.get("gst",0))))
                batch_v  = tk.StringVar(value=it.get("batch_no","") or "")
                expiry_v = tk.StringVar(value=fmt_exp_mmyy(it.get("expiry_date","")) or "")
                # Existing discount reverse-calculate karo (qty*rate - taxable)
                try:
                    _qty_f  = float(it.get("qty",0))
                    _rate_f = float(it.get("rate",0))
                    _gross  = _qty_f * _rate_f
                    _taxable_saved = float(it.get("taxable") or _gross)
                    _existing_disc = round(((_gross - _taxable_saved) / _gross * 100), 2) if _gross > 0 else 0.0
                except:
                    _existing_disc = 0.0
                disc_v = tk.StringVar(value=str(_existing_disc) if _existing_disc > 0 else "0")

                # Live recalc labels
                gst_lbl = tk.Label(tbl_f,text="₹0.00",font=("Segoe UI",9),bg=C_WHITE,padx=4,pady=3)
                tot_lbl = tk.Label(tbl_f,text="₹0",   font=("Segoe UI",9),bg=C_WHITE,padx=4,pady=3,fg=C_ACCENT)

                def _recalc(event=None, qv=qty_v, rv=rate_v, dv=disc_v, gv=gst_v, gl=gst_lbl, tl=tot_lbl):
                    try:
                        _q = float(qv.get() or 0)
                        _r = float(rv.get() or 0)
                        _d = float(dv.get() or 0)
                        _g = float(gv.get() or 0)
                        _gross   = _q * _r
                        _taxable = round(_gross * (1 - _d/100), 2)
                        _gamt    = round(_taxable * _g / 100, 2)
                        _total   = _taxable + _gamt
                        gl.config(text=f"₹{_gamt:.2f}")
                        tl.config(text=f"₹{_total:.0f}")
                    except:
                        pass

                qty_e  = ttk.Entry(tbl_f,textvariable=qty_v,width=6,font=("Segoe UI",9))
                rate_e = ttk.Entry(tbl_f,textvariable=rate_v,width=10,font=("Segoe UI",9))
                disc_e = ttk.Entry(tbl_f,textvariable=disc_v,width=6,font=("Segoe UI",9))
                gst_cb = ttk.Combobox(tbl_f,textvariable=gst_v,values=["0","5","12","18","28"],
                                      width=5,font=("Segoe UI",9),state="readonly")

                for _w in [qty_e, rate_e, disc_e]:
                    _w.bind("<KeyRelease>", _recalc)
                gst_cb.bind("<<ComboboxSelected>>", _recalc)

                ttk.Combobox(tbl_f,textvariable=prod_v,values=all_products,width=22,font=("Segoe UI",9)
                             ).grid(row=ri+1,column=0,sticky="ew",padx=1,pady=1)
                qty_e .grid(row=ri+1,column=1,sticky="ew",padx=1,pady=1)
                rate_e.grid(row=ri+1,column=2,sticky="ew",padx=1,pady=1)
                disc_e.grid(row=ri+1,column=3,sticky="ew",padx=1,pady=1)
                gst_cb.grid(row=ri+1,column=4,sticky="ew",padx=1,pady=1)
                ttk.Entry(tbl_f,textvariable=batch_v,width=10,font=("Segoe UI",9)
                          ).grid(row=ri+1,column=5,sticky="ew",padx=1,pady=1)
                make_exp_mmyy_entry(tbl_f,expiry_v,width=8
                                ).grid(row=ri+1,column=6,sticky="ew",padx=1,pady=1)
                gst_lbl.grid(row=ri+1,column=7,sticky="ew",padx=1,pady=1)
                tot_lbl.grid(row=ri+1,column=8,sticky="ew",padx=1,pady=1)
                tk.Button(tbl_f,text="✕",font=("Segoe UI",9),bg=C_RED,fg=C_WHITE,
                          relief="flat",cursor="hand2",bd=0,
                          command=lambda idx=ri: (edit_items.pop(idx), render_items())
                          ).grid(row=ri+1,column=9,sticky="nsew",padx=1,pady=1)
                _recalc()
                item_rows_widgets.append({"prod":prod_v,"qty":qty_v,"rate":rate_v,
                                          "disc":disc_v,"gst":gst_v,
                                          "batch":batch_v,"expiry":expiry_v})

        render_items()

        af=tk.Frame(p,bg=C_LIGHT,highlightthickness=1,highlightbackground=C_BORDER)
        af.pack(fill="x",pady=4)
        for i in range(7): af.columnconfigure(i,weight=1)
        tk.Label(af,text="New Item:",font=("Segoe UI",9,"bold"),bg=C_LIGHT,fg=C_ACCENT
                 ).grid(row=0,column=0,sticky="w",padx=6,pady=(2,0))
        n_prod=tk.StringVar(); n_qty=tk.StringVar(); n_rate=tk.StringVar()
        n_batch=tk.StringVar(); n_expiry=tk.StringVar()

        tk.Label(af,text="Product",font=("Segoe UI",7),bg=C_LIGHT,fg=C_GRAY).grid(row=0,column=0,sticky="w",padx=4)
        ttk.Combobox(af,textvariable=n_prod,values=all_products,width=20,font=("Segoe UI",9)
                     ).grid(row=1,column=0,sticky="ew",padx=4,pady=(0,3))
        tk.Label(af,text="Qty",font=("Segoe UI",7),bg=C_LIGHT,fg=C_GRAY).grid(row=0,column=1,sticky="w",padx=4)
        ttk.Entry(af,textvariable=n_qty,width=7,font=("Segoe UI",9)
                  ).grid(row=1,column=1,sticky="ew",padx=4,pady=(0,3))
        tk.Label(af,text="Rate",font=("Segoe UI",7),bg=C_LIGHT,fg=C_GRAY).grid(row=0,column=2,sticky="w",padx=4)
        ttk.Entry(af,textvariable=n_rate,width=9,font=("Segoe UI",9)
                  ).grid(row=1,column=2,sticky="ew",padx=4,pady=(0,3))
        tk.Label(af,text="Batch No",font=("Segoe UI",7),bg=C_LIGHT,fg=C_GRAY).grid(row=0,column=3,sticky="w",padx=4)
        ttk.Entry(af,textvariable=n_batch,width=10,font=("Segoe UI",9)
                  ).grid(row=1,column=3,sticky="ew",padx=4,pady=(0,3))
        tk.Label(af,text="Expiry (MM/YY)",font=("Segoe UI",7),bg=C_LIGHT,fg=C_GRAY).grid(row=0,column=4,sticky="w",padx=4)
        make_exp_mmyy_entry(af,n_expiry,width=8).grid(row=1,column=4,sticky="ew",padx=4,pady=(0,3))

        def add_new_item():
            if not n_prod.get(): messagebox.showerror("Error","Product select karo!",parent=win); return
            try: qty=float(n_qty.get()); rate=float(n_rate.get())
            except: messagebox.showerror("Error","Qty/Rate sahi bharo!",parent=win); return
            edit_items.append({"product":n_prod.get(),"qty":qty,"rate":rate,"total":qty*rate,
                               "batch_no":n_batch.get().strip(),
                               "mfg_date":"",
                               "expiry_date":n_expiry.get().strip()})
            n_prod.set(""); n_qty.set(""); n_rate.set("")
            n_batch.set(""); n_expiry.set("")
            render_items()

        make_btn(af, "+ Add", add_new_item).grid(row=1,column=5,padx=8,pady=(0,3),sticky="ew")

        def collect_and_save():
            updated=[]
            for w in item_rows_widgets:
                try:
                    prod=w["prod"].get().strip()
                    qty=float(w["qty"].get())
                    rate=float(w["rate"].get())
                    disc=float(w["disc"].get() or 0)
                    gst=float(w["gst"].get())
                    gross=qty*rate
                    taxable=round(gross*(1-disc/100),2)
                    gst_amt=round(taxable*gst/100,2)
                    total=taxable+gst_amt
                    updated.append({"product":prod,"qty":qty,"rate":rate,
                                    "taxable":taxable,"gst":gst,"gst_amt":gst_amt,"total":total,
                                    "disc_pct":disc,
                                    "batch_no":w["batch"].get().strip(),
                                    "mfg_date":"",
                                    "expiry_date":w["expiry"].get().strip()})
                except: pass
            if not updated:
                messagebox.showerror("Error","Koi item nahi!",parent=win); return
            grand=sum(it["total"] for it in updated)
            conn=get_db()
            try:
                conn.execute("UPDATE purchases SET bill_date=?,party=?,grand_total=?,pay_mode=?,due_date=? WHERE id=?",
                             (e_date.get(), e_party.get(), grand, e_paymode.get(), e_due.get().strip(), pid))
                conn.execute("DELETE FROM purchase_items WHERE purchase_id=?", (pid,))
                for it in updated:
                    conn.execute("INSERT INTO purchase_items(purchase_id,product,qty,rate,taxable,gst_percent,gst_amt,total,batch_no,mfg_date,expiry_date) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                                 (pid,it["product"],it["qty"],it["rate"],it["taxable"],it["gst"],it["gst_amt"],it["total"],
                                  it.get("batch_no",""),"",it.get("expiry_date","")))
                    # expiry_stock bhi update karo agar expiry_date diya
                    if it.get("expiry_date"):
                        conn.execute(
                            "INSERT OR REPLACE INTO expiry_stock(product,batch_no,mfg_date,expiry_date,qty,purchase_rate,supplier) "
                            "VALUES(?,?,?,?,?,?,?)",
                            (it["product"],it.get("batch_no",""),"",
                             it["expiry_date"],it["qty"],it["rate"],e_party.get()))
                conn.commit()
                messagebox.showinfo("Done","Purchase update ho gaya!",parent=win)
                win.destroy(); refresh_cb()
            except Exception as e:
                messagebox.showerror("Error",str(e),parent=win)
            finally: conn.close()

        bf2=tk.Frame(p,bg=C_BG); bf2.pack(fill="x",pady=4)
        make_btn(bf2,"💾 Save Changes",collect_and_save).pack(side="left",padx=(0,8))
        make_btn(bf2,"✕ Cancel",win.destroy,bg=C_GRAY).pack(side="left")

    # ══════════════════════════════════════════════════════════════════════════
    #  ALL BILLS (Sale + Purchase tabs)
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_allbills(self):
        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "All Bills")

        tabf = tk.Frame(p, bg=C_BG); tabf.pack(fill="x", pady=(0,3))
        self._bills_tab = tk.StringVar(value="sale")

        sale_btn = make_btn(tabf, "Sale Bills", lambda: switch("sale"))
        sale_btn.pack(side="left", padx=(0,4))
        pur_btn  = make_btn(tabf, "Purchase Bills", lambda: switch("purchase"), bg=C_BORDER)
        pur_btn.config(fg=C_GRAY)
        pur_btn.pack(side="left")

        sf = tk.Frame(p, bg=C_BG); sf.pack(fill="x", pady=(0,3))
        tk.Label(sf,text="Search:",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY).pack(side="left")
        sv = tk.StringVar()
        ttk.Entry(sf,textvariable=sv,width=28,font=("Segoe UI",9)).pack(side="left",padx=6)

        self._ab_tbl_frame = tk.Frame(p, bg=C_BG)
        self._ab_tbl_frame.pack(fill="x")
        self._ab_tbl = make_table(self._ab_tbl_frame,
            ["Bill No","Date","Party","Amount","View","Edit","Delete"],
            [20,12,24,13,7,7,7])

        def switch(t):
            self._bills_tab.set(t)
            if t == "sale":
                sale_btn.config(bg=C_ACCENT, fg=C_WHITE)
                pur_btn.config(bg=C_BORDER, fg=C_GRAY)
            else:
                sale_btn.config(bg=C_BORDER, fg=C_GRAY)
                pur_btn.config(bg=C_ACCENT, fg=C_WHITE)
            refresh()

        def refresh():
            clear_table_rows(self._ab_tbl)
            tab = self._bills_tab.get()
            q = sv.get().lower()
            conn = get_db()
            if tab == "sale":
                rows = conn.execute("SELECT * FROM sales ORDER BY id DESC").fetchall()
            else:
                rows = conn.execute("SELECT * FROM purchases ORDER BY id DESC").fetchall()
            conn.close()
            ri = 0
            for r in rows:
                if q and q not in r["bill_no"].lower() and q not in r["party"].lower():
                    continue
                bg = C_WHITE if ri%2==0 else "#F7FAFC"
                rid = r["id"]; rr = dict(r)
                for j, val in enumerate([r["bill_no"], fmt_date(r["bill_date"]),
                                          r["party"][:22], f"₹{round(r['grand_total']):,}",
                                          "","",""]):
                    if j == 4:
                        # View
                        if tab == "sale":
                            tk.Button(self._ab_tbl, text="👁", font=("Segoe UI",9),
                                      bg=C_ACCENT, fg=C_WHITE, relief="flat", cursor="hand2", bd=0,
                                      command=lambda s=rid: self._view_sale(s)
                                      ).grid(row=ri+1, column=j, sticky="nsew", padx=1, pady=1)
                        else:
                            tk.Button(self._ab_tbl, text="👁", font=("Segoe UI",9),
                                      bg=C_AMBER, fg=C_WHITE, relief="flat", cursor="hand2", bd=0,
                                      command=lambda s=rid, rrd=rr: self._view_purchase(s, rrd)
                                      ).grid(row=ri+1, column=j, sticky="nsew", padx=1, pady=1)
                    elif j == 5:
                        # Edit
                        if tab == "sale":
                            tk.Button(self._ab_tbl, text="✏", font=("Segoe UI",9),
                                      bg="#2B6CB0", fg=C_WHITE, relief="flat", cursor="hand2", bd=0,
                                      command=lambda s=rid: self._edit_sale(s, refresh)
                                      ).grid(row=ri+1, column=j, sticky="nsew", padx=1, pady=1)
                        else:
                            tk.Button(self._ab_tbl, text="✏", font=("Segoe UI",9),
                                      bg="#2B6CB0", fg=C_WHITE, relief="flat", cursor="hand2", bd=0,
                                      command=lambda s=rid: self._edit_purchase(s, refresh)
                                      ).grid(row=ri+1, column=j, sticky="nsew", padx=1, pady=1)
                    elif j == 6:
                        # Delete
                        bn = r["bill_no"]
                        if tab == "sale":
                            tk.Button(self._ab_tbl, text="🗑", font=("Segoe UI",9),
                                      bg=C_RED, fg=C_WHITE, relief="flat", cursor="hand2", bd=0,
                                      command=lambda s=rid, b=bn: self._delete_sale(s, b, refresh)
                                      ).grid(row=ri+1, column=j, sticky="nsew", padx=1, pady=1)
                        else:
                            tk.Button(self._ab_tbl, text="🗑", font=("Segoe UI",9),
                                      bg=C_RED, fg=C_WHITE, relief="flat", cursor="hand2", bd=0,
                                      command=lambda s=rid, b=bn: self._delete_purchase(s, b, refresh)
                                      ).grid(row=ri+1, column=j, sticky="nsew", padx=1, pady=1)
                    else:
                        tk.Label(self._ab_tbl, text=str(val), font=("Segoe UI",9),
                                 bg=bg, fg=C_GRAY, anchor="w", padx=6, pady=4
                                 ).grid(row=ri+1, column=j, sticky="nsew", padx=1)
                ri += 1

        def export_ab():
            try:
                tab=self._bills_tab.get(); q=sv.get().lower()
                conn=get_db()
                if tab=="sale":
                    rows=[dict(r) for r in conn.execute("SELECT * FROM sales ORDER BY id DESC").fetchall()]
                    hdrs=["Bill No","Date","Party","Amount","Pay Mode","Due Date"]
                    data=[[r["bill_no"],fmt_date(r["bill_date"]),r["party"],
                           f"{r['grand_total']:.0f}",r.get("pay_mode","") or "",r.get("due_date","") or ""] for r in rows
                          if not q or q in r["bill_no"].lower() or q in r["party"].lower()]
                    fname="All_Sale_Bills"
                else:
                    rows=[dict(r) for r in conn.execute("SELECT * FROM purchases ORDER BY id DESC").fetchall()]
                    hdrs=["Bill No","Date","Supplier","Amount","Pay Mode","Due Date"]
                    data=[[r["bill_no"],fmt_date(r["bill_date"]),r["party"],
                           f"{r['grand_total']:.0f}",r.get("pay_mode","") or "",r.get("due_date","") or ""] for r in rows
                          if not q or q in r["bill_no"].lower() or q in r["party"].lower()]
                    fname="All_Purchase_Bills"
                conn.close()
                export_to_excel(hdrs, data, fname)
            except Exception as e:
                messagebox.showerror("Export Error", str(e))

        sv.trace_add("write", lambda *a: refresh())
        make_btn(sf,"📥 Excel",export_ab,bg=C_GREEN).pack(side="left",padx=8)
        refresh()

    # ══════════════════════════════════════════════════════════════════════════
    #  STOCK
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_stock(self):
        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "Stock Register")

        sf = tk.Frame(p, bg=C_BG); sf.pack(fill="x", pady=(0,3))
        tk.Label(sf, text="Search:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        sv = tk.StringVar()
        ttk.Entry(sf, textvariable=sv, width=24, font=("Segoe UI",9)).pack(side="left", padx=6)

        tk.Label(sf, text="Filter:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left", padx=(10,0))
        fv = tk.StringVar(value="All")
        ttk.Combobox(sf, textvariable=fv, width=12, state="readonly",
                     values=["All","Low Stock","Out of Stock","OK"]).pack(side="left", padx=4)

        tbl = make_table(p,
            ["Product","HSN","Opening","Purchased","Sold","Manual IN","Manual OUT","Current","Rate","Value","Alert","Status"],
            [20, 7, 8, 10, 8, 10, 11, 9, 10, 12, 7, 8])

        def load_rows():
            conn = get_db()
            prods = [dict(r) for r in conn.execute("SELECT * FROM products ORDER BY name").fetchall()]
            rows = []
            for pr in prods:
                bought  = conn.execute("SELECT COALESCE(SUM(qty+free_qty),0) FROM purchase_items WHERE product=?", (pr["name"],)).fetchone()[0]
                sold    = conn.execute("SELECT COALESCE(SUM(qty),0) FROM sale_items    WHERE product=?", (pr["name"],)).fetchone()[0]
                man_in  = conn.execute("SELECT COALESCE(SUM(qty),0) FROM stock_movements WHERE product=? AND move_type='IN'",  (pr["name"],)).fetchone()[0]
                man_out = conn.execute("SELECT COALESCE(SUM(qty),0) FROM stock_movements WHERE product=? AND move_type='OUT'", (pr["name"],)).fetchone()[0]
                stk     = pr["opening_stock"] + bought + man_in - sold - man_out
                val     = stk * pr["sale_rate"]
                alert   = pr.get("low_stock_alert") or 5
                try:    alert = float(alert)
                except: alert = 5
                if stk <= 0:        st="✗ Out";  clr=C_RED
                elif stk <= alert:  st="⚠ Low";  clr=C_AMBER
                else:               st="✓ OK";   clr=C_GREEN
                rows.append({
                    "name": pr["name"], "hsn": pr["hsn"],
                    "opening": pr["opening_stock"], "bought": int(bought),
                    "sold": int(sold), "man_in": int(man_in), "man_out": int(man_out),
                    "stk": int(stk), "rate": pr["sale_rate"],
                    "val": round(val), "alert": int(alert),
                    "st": st, "clr": clr
                })
            conn.close()
            return rows

        all_rows = []

        def render(q="", filt="All"):
            clear_table_rows(tbl)
            for i, r in enumerate(all_rows):
                if q    and q.lower() not in r["name"].lower(): continue
                if filt == "Low Stock"    and r["st"] != "⚠ Low":  continue
                if filt == "Out of Stock" and r["st"] != "✗ Out":  continue
                if filt == "OK"           and r["st"] != "✓ OK":   continue
                vals = [r["name"], r["hsn"], r["opening"], r["bought"],
                        r["sold"], r["man_in"], r["man_out"],
                        r["stk"], f"₹{r['rate']:.2f}", f"₹{r['val']:,}",
                        r["alert"], r["st"]]
                fgs  = [None]*11 + [r["clr"]]
                table_row(tbl, vals, i+1, fgs=fgs)

        def refresh():
            nonlocal all_rows
            all_rows = load_rows()
            render(sv.get(), fv.get())

        def export_stk():
            try:
                data = [[r["name"], r["hsn"], r["opening"], r["bought"],
                         r["sold"], r["man_in"], r["man_out"],
                         r["stk"], f"₹{r['rate']:.2f}", f"₹{r['val']:,}",
                         r["alert"], r["st"]] for r in all_rows]
                export_to_excel(
                    ["Product","HSN","Opening","Purchased","Sold","Manual IN","Manual OUT",
                     "Current Stock","Rate","Value","Low Alert","Status"],
                    data, "Stock_Register")
            except Exception as e:
                messagebox.showerror("Export Error", str(e))

        sv.trace_add("write", lambda *a: render(sv.get(), fv.get()))
        fv.trace_add("write", lambda *a: render(sv.get(), fv.get()))
        make_btn(sf, "🔄 Refresh",  refresh,    bg="#2B6CB0").pack(side="left", padx=8)
        make_btn(sf, "📥 Excel",    export_stk, bg=C_GREEN  ).pack(side="left", padx=4)
        make_btn(sf, "➕ Stock In/Out", lambda: self._show("stockinout"), bg=C_AMBER).pack(side="left", padx=4)
        refresh()

    # ══════════════════════════════════════════════════════════════════════════
    #  STOCK IN / OUT
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_stockinout(self):
        import datetime as _dt

        # Ensure table exists
        db0 = get_db()
        db0.execute("""CREATE TABLE IF NOT EXISTS stock_movements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            move_date TEXT NOT NULL, product TEXT NOT NULL,
            move_type TEXT NOT NULL, qty REAL NOT NULL,
            ref_no TEXT DEFAULT '', reason TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP)""")
        db0.commit(); db0.close()

        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "📦 Stock In / Out")

        # ── Entry Form ──────────────────────────────────────────────────────
        form = tk.Frame(p, bg=C_WHITE, highlightthickness=1, highlightbackground=C_BORDER)
        form.pack(fill="x", pady=(0,10))
        tk.Label(form, text="Manual Stock Entry", font=("Segoe UI",10,"bold"),
                 bg=C_WHITE, fg="#1A365D").pack(anchor="w", padx=16, pady=(10,6))

        row1 = tk.Frame(form, bg=C_WHITE); row1.pack(fill="x", padx=16, pady=4)

        # Date
        tk.Label(row1, text="Date:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_date = tk.StringVar(value=str(_dt.date.today()))
        make_date_entry(row1, v_date, width=13).pack(side="left", padx=(2,14))

        # Type
        tk.Label(row1, text="Type:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_type = tk.StringVar(value="IN")
        cb_type = ttk.Combobox(row1, textvariable=v_type, width=8, state="readonly",
                               values=["IN","OUT"])
        cb_type.pack(side="left", padx=(2,14))

        # Product
        tk.Label(row1, text="Product:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        conn_p = get_db()
        prod_list = [r[0] for r in conn_p.execute("SELECT name FROM products ORDER BY name").fetchall()]
        conn_p.close()
        v_prod = tk.StringVar()
        cb_prod = ttk.Combobox(row1, textvariable=v_prod, values=prod_list, width=22)
        add_autocomplete(cb_prod, lambda: [r["name"] for r in get_db().execute("SELECT name FROM products ORDER BY name").fetchall()])
        cb_prod.pack(side="left", padx=(2,4))

        # Current stock label
        stk_lbl = tk.Label(row1, text="Current: -", font=("Segoe UI",9,"bold"),
                           bg=C_WHITE, fg="#2B6CB0")
        stk_lbl.pack(side="left", padx=(0,14))

        def on_prod_change(ev=None):
            pr = v_prod.get().strip()
            if pr:
                stk = get_stock(pr)
                clr = C_RED if stk<=0 else C_AMBER if stk<=5 else C_GREEN
                stk_lbl.config(text=f"Current: {int(stk)}", fg=clr)
            else:
                stk_lbl.config(text="Current: -", fg="#2B6CB0")
        cb_prod.bind("<<ComboboxSelected>>", on_prod_change)

        row2 = tk.Frame(form, bg=C_WHITE); row2.pack(fill="x", padx=16, pady=4)

        # Qty
        tk.Label(row2, text="Qty:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_qty = tk.StringVar()
        ttk.Entry(row2, textvariable=v_qty, width=10).pack(side="left", padx=(2,14))

        # Ref No
        tk.Label(row2, text="Ref No:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_ref = tk.StringVar()
        ttk.Entry(row2, textvariable=v_ref, width=14).pack(side="left", padx=(2,14))

        # Reason
        tk.Label(row2, text="Reason:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_reason = tk.StringVar()
        reason_opts = ["","Damage/Wastage","Theft/Loss","Adjustment","Transfer","Opening Entry",
                       "Return to Supplier","Sample/Testing","Other"]
        ttk.Combobox(row2, textvariable=v_reason, values=reason_opts, width=20).pack(side="left", padx=(2,0))

        def do_save():
            prod = v_prod.get().strip()
            if not prod:
                messagebox.showerror("Error","Product select karo!"); return
            try:
                qty = float(v_qty.get())
                assert qty > 0
            except:
                messagebox.showerror("Error","Qty sahi bharo (positive number)!"); return
            mtype = v_type.get()
            if mtype == "OUT":
                cur = get_stock(prod)
                if qty > cur:
                    if not messagebox.askyesno("Warning",
                        f"Current stock {int(cur)} hai lekin {int(qty)} OUT kar rahe ho.\nFir bhi save karo?"):
                        return
            d = get_db()
            d.execute(
                "INSERT INTO stock_movements(move_date,product,move_type,qty,ref_no,reason) VALUES(?,?,?,?,?,?)",
                (v_date.get(), prod, mtype, qty, v_ref.get().strip(), v_reason.get().strip())
            )
            d.commit(); d.close()
            new_stk = get_stock(prod)
            messagebox.showinfo("Saved!",
                f"{'Stock IN' if mtype=='IN' else 'Stock OUT'} saved!\n"
                f"Product: {prod}\nQty: {int(qty)}\nNew Stock: {int(new_stk)}")
            v_prod.set(""); v_qty.set(""); v_ref.set(""); v_reason.set("")
            stk_lbl.config(text="Current: -", fg="#2B6CB0")
            load_history()

        bf = tk.Frame(form, bg=C_WHITE); bf.pack(anchor="w", padx=16, pady=(4,12))
        make_btn(bf, "➕ Save Stock IN",  lambda: [v_type.set("IN"),  do_save()], bg=C_GREEN ).pack(side="left", padx=(0,8))
        make_btn(bf, "➖ Save Stock OUT", lambda: [v_type.set("OUT"), do_save()], bg=C_RED   ).pack(side="left", padx=(0,8))
        make_btn(bf, "Reset", lambda: [v_prod.set(""),v_qty.set(""),v_ref.set(""),v_reason.set(""),
                                        stk_lbl.config(text="Current: -",fg="#2B6CB0")], bg=C_GRAY).pack(side="left")

        # ── Filter bar for history ───────────────────────────────────────────
        hf = tk.Frame(p, bg=C_BG); hf.pack(fill="x", pady=(8,4))
        tk.Label(hf, text="History — Product:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_hprod = tk.StringVar()
        ttk.Combobox(hf, textvariable=v_hprod, values=[""]+prod_list, width=22).pack(side="left", padx=4)

        tk.Label(hf, text="Type:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left", padx=(8,0))
        v_htype = tk.StringVar(value="All")
        ttk.Combobox(hf, textvariable=v_htype, width=8, state="readonly",
                     values=["All","IN","OUT"]).pack(side="left", padx=4)

        tk.Label(hf, text="From:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left", padx=(8,0))
        v_hfrom = tk.StringVar()
        make_date_entry(hf, v_hfrom, width=12).pack(side="left", padx=4)

        tk.Label(hf, text="To:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_hto = tk.StringVar()
        make_date_entry(hf, v_hto, width=12).pack(side="left", padx=4)

        make_btn(hf, "🔍 Filter", lambda: load_history(), bg="#2B6CB0").pack(side="left", padx=6)
        make_btn(hf, "📥 Excel",  lambda: export_history(), bg=C_GREEN).pack(side="left", padx=4)

        # ── History Table ────────────────────────────────────────────────────
        tk.Label(p, text="Movement History", font=("Segoe UI",10,"bold"),
                 bg=C_BG, fg="#1A365D").pack(anchor="w", pady=(4,2))
        htbl = make_table(p,
            ["Date","Type","Product","Qty","Ref No","Reason","Stock After"],
            [11, 7, 22, 7, 13, 20, 11])

        def load_history():
            clear_table_rows(htbl)
            hprod = v_hprod.get().strip()
            htype = v_htype.get()
            hfrom = v_hfrom.get().strip()
            hto   = v_hto.get().strip()

            d = get_db()
            q  = "SELECT * FROM stock_movements WHERE 1=1"
            params = []
            if hprod: q += " AND product=?";    params.append(hprod)
            if htype != "All": q += " AND move_type=?"; params.append(htype)
            if hfrom: q += " AND move_date>=?"; params.append(hfrom)
            if hto:   q += " AND move_date<=?"; params.append(hto)
            q += " ORDER BY move_date DESC, id DESC"
            rows = d.execute(q, params).fetchall()
            d.close()

            # Calculate running stock after each movement (for display)
            for i, r in enumerate(rows):
                stk_after = get_stock(r["product"])
                # Running stock is complex to calc backwards, show current stock for latest
                type_fg = C_GREEN if r["move_type"]=="IN" else C_RED
                qty_txt = ("+" if r["move_type"]=="IN" else "-") + str(int(r["qty"]))
                table_row(htbl,
                    [fmt_date(r["move_date"]), r["move_type"], r["product"],
                     qty_txt, r["ref_no"] or "-", r["reason"] or "-",
                     str(int(stk_after)) if i==0 else "-"],
                    i+1,
                    fgs=[None, type_fg, None, type_fg, None, None, None])

        def export_history():
            try:
                d = get_db()
                rows = d.execute(
                    "SELECT * FROM stock_movements ORDER BY move_date DESC, id DESC"
                ).fetchall()
                d.close()
                data = [[fmt_date(r["move_date"]), r["move_type"], r["product"],
                         int(r["qty"]), r["ref_no"] or "", r["reason"] or ""] for r in rows]
                export_to_excel(
                    ["Date","Type","Product","Qty","Ref No","Reason"],
                    data, "Stock_Movements")
            except Exception as e:
                messagebox.showerror("Export Error", str(e))

        load_history()



    # ══════════════════════════════════════════════════════════════════════════
    #  PRODUCTS
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_products(self):
        p=tk.Frame(self.content,bg=C_BG,padx=20,pady=14)
        p.pack(fill="x")
        section_title(p,"🛒 Products Master — Medical")

        # ── Expiry Alert Banner ──────────────────────────────────────────────
        alerts = get_expiry_alerts()
        if alerts:
            exp_frame = tk.Frame(p, bg="#FFF5F5", relief="flat", bd=1)
            exp_frame.pack(fill="x", pady=(0, 4))
            today_key = exp_key_for_date(datetime.date.today())
            expired_count = sum(1 for a in alerts if exp_sort_key(a["expiry_date"]) < today_key)
            soon_count = len(alerts) - expired_count
            msg = ""
            if expired_count: msg += f"⚠️ {expired_count} batch EXPIRED! "
            if soon_count:    msg += f"🔔 {soon_count} batch 30 din mein expire hogi!"
            tk.Label(exp_frame, text=msg, font=("Segoe UI", 9, "bold"),
                     bg="#FFF5F5", fg="#C53030", anchor="w", padx=10, pady=5).pack(fill="x")
            tk.Button(exp_frame, text="📋 Expiry Details Dekho",
                      font=("Segoe UI",7), bg="#C53030", fg="white",
                      relief="flat", cursor="hand2", padx=8, pady=2,
                      command=lambda: self._expiry_stock_window()).pack(side="right", padx=6, pady=4)

        # ── Low Stock Alert Banner ────────────────────────────────────────────
        low_items = get_low_stock_alerts()
        if low_items:
            low_frame = tk.Frame(p, bg="#FFFBEB", relief="flat", bd=1,
                                 highlightthickness=1, highlightbackground="#F6AD55")
            low_frame.pack(fill="x", pady=(0, 8))
            zero_count = sum(1 for x in low_items if x["stock"] <= 0)
            low_count  = len(low_items) - zero_count
            msg2 = ""
            if zero_count: msg2 += f"🚨 {zero_count} product OUT OF STOCK!  "
            if low_count:  msg2 += f"⚠️ {low_count} product ka stock kam hai!"
            tk.Label(low_frame, text=msg2, font=("Segoe UI", 9, "bold"),
                     bg="#FFFBEB", fg="#744210", anchor="w", padx=10, pady=5).pack(side="left", fill="x", expand=True)
            def _show_low_stock_popup(items=low_items):
                win = tk.Toplevel()
                win.title("⚠️ Low Stock Products")
                win.configure(bg="#FFFBEB")
                win.geometry("520x380")
                win.grab_set()
                tk.Label(win, text="⚠️  Low Stock Alert", font=("Segoe UI", 13, "bold"),
                         bg="#FFFBEB", fg="#744210").pack(pady=(12, 4))
                tk.Label(win, text="Neeche diye products ka stock low_stock_alert se kam ya khatam hai:",
                         font=("Segoe UI", 9), bg="#FFFBEB", fg="#744210").pack()
                import tkinter.ttk as _ttk
                cols = ("Product", "Stock", "Alert Limit", "Unit")
                tree = _ttk.Treeview(win, columns=cols, show="headings", height=12)
                for col, w in zip(cols, (220, 80, 90, 80)):
                    tree.heading(col, text=col); tree.column(col, width=w, anchor="center")
                tree.column("Product", anchor="w")
                for it in items:
                    stk = it["stock"] or 0
                    tag = "zero" if stk <= 0 else "low"
                    tree.insert("", "end", values=(it["name"], stk, it["low_stock_alert"] or 0, it["unit"] or ""), tags=(tag,))
                tree.tag_configure("zero", foreground="#C53030", font=("Segoe UI", 9, "bold"))
                tree.tag_configure("low",  foreground="#744210")
                tree.pack(fill="both", expand=True, padx=12, pady=4)
                tk.Button(win, text="Close", command=win.destroy,
                          font=("Segoe UI", 10, "bold"), bg="#744210", fg="white",
                          relief="flat", padx=20, pady=6).pack(pady=(0, 12))
            tk.Button(low_frame, text="📋 Low Stock Dekho",
                      font=("Segoe UI",7), bg="#D97706", fg="white",
                      relief="flat", cursor="hand2", padx=8, pady=2,
                      command=_show_low_stock_popup).pack(side="right", padx=6, pady=4)

        # ── Form (ROW 1) ─────────────────────────────────────────────────────
        frm = tk.Frame(p, bg=C_BG)
        frm.pack(fill="x", pady=(0, 2))

        # Row 1 + Row 2 — fields 2 rows mein baante hain taaki bina scroll
        # ke sab kuch screen par fit ho jaye
        row1 = tk.Frame(frm, bg=C_BG)
        row1.pack(fill="x", pady=2)

        row2 = tk.Frame(frm, bg=C_BG)
        row2.pack(fill="x", pady=2)

        def lbl(parent, text):
            tk.Label(parent, text=text, font=("Segoe UI", 9),
                     bg=C_BG, fg=C_GRAY).pack(side="left", padx=(8, 2))

        def ent(parent, var, width):
            e = ttk.Entry(parent, textvariable=var, width=width)
            e.pack(side="left", padx=(0, 6))
            return e

        # ── Row 1: Name, HSN, Sale Rate, MRP, Purchase Rate, GST% ──────────
        pr_name = tk.StringVar()
        lbl(row1, "Product Name"); pr_name_entry = ent(row1, pr_name, 18)
        auto_titlecase(pr_name)
        add_entry_autocomplete(pr_name_entry, lambda: [r[0] for r in get_db().execute(
            "SELECT DISTINCT name FROM products WHERE name!='' ORDER BY name").fetchall()])

        pr_hsn = tk.StringVar()
        lbl(row1, "HSN Code"); pr_hsn_entry = ent(row1, pr_hsn, 8)
        add_entry_autocomplete(pr_hsn_entry, lambda: [r[0] for r in get_db().execute(
            "SELECT DISTINCT hsn FROM products WHERE hsn!='' ORDER BY hsn").fetchall()])

        pr_rate = tk.StringVar()
        lbl(row1, "Sale Rate (₹)"); pr_rate_entry = ent(row1, pr_rate, 8)

        pr_mrp = tk.StringVar(value="0")
        lbl(row1, "MRP (₹)"); pr_mrp_entry = ent(row1, pr_mrp, 7)

        pr_pur = tk.StringVar(value="0")
        lbl(row1, "Purchase Rate"); pr_pur_entry = ent(row1, pr_pur, 7)

        pr_gst = tk.StringVar(value="5")
        lbl(row1, "GST %")
        pr_gst_cb = ttk.Combobox(row1, textvariable=pr_gst,
                     values=["0","5","12","18","28"],
                     width=4, state="readonly")
        pr_gst_cb.pack(side="left", padx=(0,6))

        # ── Row 2: Opening Stock, Packing, Mfg Company, Expiry ─────────────
        pr_stk = tk.StringVar(value="0")
        lbl(row2, "Opening Stock"); pr_stk_entry = ent(row2, pr_stk, 8)

        pr_packing = tk.StringVar()
        lbl(row2, "Packing")
        pr_packing_entry = ent(row2, pr_packing, 16)
        add_entry_autocomplete(pr_packing_entry, lambda: [r[0] for r in get_db().execute(
            "SELECT DISTINCT packing FROM products WHERE packing!='' ORDER BY packing").fetchall()])

        pr_mfg_company = tk.StringVar()
        lbl(row2, "Mfg Company")
        pr_mfg_company_entry = ent(row2, pr_mfg_company, 18)
        add_entry_autocomplete(pr_mfg_company_entry, lambda: [r[0] for r in get_db().execute(
            "SELECT DISTINCT mfg_company FROM products WHERE mfg_company!='' ORDER BY mfg_company").fetchall()])

        # Expiry Date — hamesha visible, MM/YY format
        lbl(row2, "Expiry (MM/YY)")
        pr_expiry_date = tk.StringVar()
        pr_expiry_entry = make_exp_mmyy_entry(row2, pr_expiry_date, width=7)
        pr_expiry_entry.pack(side="left", padx=(0,6))

        # Unit/Barcode form se hata diye gaye hain (user request).
        # DB column zaroor bharte hain kyunki Sale/Purchase Bill inhe use karte hain.
        pr_unit = tk.StringVar(value="Pcs")
        pr_barcode = tk.StringVar()
        pr_batch_no = tk.StringVar()
        pr_exp_qty = tk.StringVar(value="0")

        edit_id = [None]

        # ── Table ─────────────────────────────────────────────────────────────
        tbl=make_table(p,["Product","HSN","Packing","Mfg Company","Sale Rate","MRP","Pur Rate","GST%","Stock","Expiry","Edit","Del"],
                       [18,7,10,14,9,9,9,5,6,8,6,5])

        def refresh():
            clear_table_rows(tbl)
            conn=get_db()
            rows=[dict(r) for r in conn.execute("SELECT * FROM products ORDER BY name").fetchall()]
            conn.close()
            today = datetime.date.today().isoformat()
            soon_key = exp_key_for_date(datetime.date.today()+datetime.timedelta(days=30))
            today_key = exp_key_for_date(datetime.date.today())
            for i,pr in enumerate(rows):
                stk=get_stock(pr["name"])
                # Check expiry — saari active batches ka nearest expiry date dhoondo
                conn2=get_db()
                exp_rows = conn2.execute(
                    "SELECT expiry_date FROM expiry_stock WHERE product=? AND qty>0",
                    (pr["name"],)
                ).fetchall()
                conn2.close()
                exp_dates = sorted([r[0] for r in exp_rows if r[0]], key=exp_sort_key)
                nearest_exp = exp_dates[0] if exp_dates else ""
                nearest_key = exp_sort_key(nearest_exp) if nearest_exp else ""
                exp_cnt = sum(1 for d in exp_dates if exp_sort_key(d) and exp_sort_key(d) <= soon_key)
                bg=C_WHITE if i%2==0 else "#F7FAFC"
                mrp_val = float(pr.get("mrp", 0) or 0)
                vals=[pr["name"], pr.get("hsn","") or "",
                      pr.get("packing","") or "", pr.get("mfg_company","") or "",
                      f"₹{pr['sale_rate']:.2f}", f"₹{mrp_val:.2f}",
                      f"₹{pr.get('purchase_rate',0):.2f}",
                      f"{pr['gst_percent']}%", int(stk),"","",""]
                for j,val in enumerate(vals):
                    if j==9:
                        # Expiry column — nearest batch expiry date (MM/YY)
                        if nearest_exp:
                            if nearest_key and nearest_key < today_key:
                                exp_label = f"⚠️ {nearest_exp}"
                                exp_fg    = "#C53030"
                                exp_bg    = "#FFF5F5"
                            elif nearest_key and nearest_key <= soon_key:
                                exp_label = f"🔔 {nearest_exp}"
                                exp_fg    = "#9B2C2C"
                                exp_bg    = "#FFFBEB"
                            else:
                                exp_label = nearest_exp
                                exp_fg    = "#276749"
                                exp_bg    = bg
                        else:
                            exp_label = "—"
                            exp_fg    = C_GRAY
                            exp_bg    = bg
                        tk.Label(tbl, text=exp_label, font=("Segoe UI",8,"bold"),
                                 bg=exp_bg, fg=exp_fg, anchor="w", padx=4, pady=4
                                 ).grid(row=i+1,column=j,sticky="nsew",padx=1)
                    elif j==10:
                        tk.Button(tbl,text="✏",font=("Segoe UI",8,"bold"),
                                  bg="#2B6CB0",fg=C_WHITE,relief="flat",cursor="hand2",bd=0,pady=3,
                                  command=lambda r=pr: load_edit(r)
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=2,pady=1)
                    elif j==11:
                        tk.Button(tbl,text="🗑",font=("Segoe UI",8,"bold"),
                                  bg=C_RED,fg=C_WHITE,relief="flat",cursor="hand2",bd=0,pady=3,
                                  command=lambda r=pr: delete_prod(r)
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=2,pady=1)
                    else:
                        lbl_fg = "#C53030" if j==8 and exp_cnt else C_GRAY
                        tk.Label(tbl,text=str(val),font=("Segoe UI",9),
                                 bg=bg,fg=lbl_fg,anchor="w",padx=4,pady=4
                                 ).grid(row=i+1,column=j,sticky="nsew",padx=1)

        def load_edit(pr):
            edit_id[0] = pr["id"]
            pr_name.set(pr["name"]); pr_hsn.set(pr["hsn"])
            pr_rate.set(str(pr["sale_rate"])); pr_gst.set(str(int(pr["gst_percent"])))
            pr_stk.set(str(int(pr["opening_stock"])))
            pr_mrp.set(str(pr.get("mrp",0) or 0))
            pr_pur.set(str(pr.get("purchase_rate",0) or 0))
            pr_unit.set(pr.get("unit","Pcs") or "Pcs")
            pr_barcode.set(pr.get("barcode","") or "")
            pr_packing.set(pr.get("packing","") or "")
            pr_mfg_company.set(pr.get("mfg_company","") or "")
            # Is product ki saari active batches mein se nearest expiry date dikhao
            conn_e = get_db()
            exp_rows_e = conn_e.execute(
                "SELECT expiry_date FROM expiry_stock WHERE product=? AND qty>0",
                (pr["name"],)).fetchall()
            conn_e.close()
            exp_dates_e = sorted([r[0] for r in exp_rows_e if r[0]], key=exp_sort_key)
            pr_expiry_date.set(exp_dates_e[0] if exp_dates_e else "")
            save_btn.config(text="💾 Update Product", bg=C_AMBER)
            cancel_btn.pack(side="left", padx=4)

        def cancel_edit():
            edit_id[0]=None
            pr_name.set(""); pr_hsn.set(""); pr_rate.set("")
            pr_gst.set("5"); pr_stk.set("0"); pr_mrp.set("0")
            pr_pur.set("0"); pr_unit.set("Pcs"); pr_barcode.set("")
            pr_packing.set(""); pr_mfg_company.set("")
            pr_expiry_date.set("")
            save_btn.config(text="+ Add Product", bg=C_ACCENT)
            cancel_btn.pack_forget()

        def save():
            name=pr_name.get().strip()
            if not name: messagebox.showerror("Error","Product naam bharo!"); return
            try:
                rate=float(pr_rate.get() or 0)
                stk=float(pr_stk.get() or 0)
                mrp=float(pr_mrp.get() or 0)
                pur=float(pr_pur.get() or 0)
            except:
                messagebox.showerror("Error","Rate/Stock/MRP number hona chahiye!"); return

            # Expiry date validation — MM/YY format (agar bhara gaya hai)
            exp_date_str = pr_expiry_date.get().strip()
            track_expiry_val = 0
            if exp_date_str:
                import re as _re_pchk
                if not _re_pchk.match(r"^(0[1-9]|1[0-2])/\d{2}$", exp_date_str):
                    messagebox.showerror("Format Galat!",
                        "Expiry date MM/YY format mein honi chahiye.\nExample: 12/27")
                    return
                track_expiry_val = 1

            bc = pr_barcode.get().strip()
            conn=get_db()
            try:
                if edit_id[0]:
                    conn.execute(
                        "UPDATE products SET name=?,hsn=?,sale_rate=?,gst_percent=?,opening_stock=?,"
                        "barcode=?,unit=?,mrp=?,purchase_rate=?,track_expiry=?,"
                        "packing=?,mfg_company=? WHERE id=?",
                        (name,pr_hsn.get(),rate,float(pr_gst.get()),stk,
                         bc,pr_unit.get(),mrp,pur,track_expiry_val,
                         pr_packing.get().strip(), pr_mfg_company.get().strip(),
                         edit_id[0]))
                    conn.commit()
                    # Naya expiry batch banao agar date diya gaya hai
                    if exp_date_str:
                        conn.execute(
                            "INSERT INTO expiry_stock(product,batch_no,expiry_date,qty,"
                            "purchase_rate,mrp,supplier) VALUES(?,?,?,?,?,?,?)",
                            (name, pr_batch_no.get().strip(), exp_date_str,
                             stk, pur, mrp, ""))
                        conn.commit()
                    messagebox.showinfo("Done","Product update ho gaya! ✅")
                    cancel_edit()
                else:
                    conn.execute(
                        "INSERT INTO products(name,hsn,sale_rate,gst_percent,opening_stock,"
                        "barcode,unit,mrp,purchase_rate,track_expiry,packing,mfg_company) "
                        "VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                        (name,pr_hsn.get(),rate,float(pr_gst.get()),stk,
                         bc,pr_unit.get(),mrp,pur,track_expiry_val,
                         pr_packing.get().strip(), pr_mfg_company.get().strip()))
                    conn.commit()
                    # Auto-insert into expiry_stock if expiry date provided
                    if exp_date_str:
                        exp_qty = stk
                        conn.execute(
                            "INSERT INTO expiry_stock(product,batch_no,expiry_date,qty,"
                            "purchase_rate,mrp,supplier) VALUES(?,?,?,?,?,?,?)",
                            (name, pr_batch_no.get().strip(), exp_date_str,
                             exp_qty, pur, mrp, ""))
                        conn.commit()
                    messagebox.showinfo("Done","Product add ho gaya! ✅")
                    cancel_edit()
                refresh()
            except sqlite3.IntegrityError:
                messagebox.showerror("Error","Product pehle se exist karta hai!")
            finally: conn.close()

        def delete_prod(pr):
            if not messagebox.askyesno("Delete?",f"'{pr['name']}' delete karna chahte ho?\nSale/Purchase history safe rahegi."):
                return
            conn=get_db()
            conn.execute("DELETE FROM products WHERE id=?", (pr["id"],))
            conn.commit(); conn.close()
            refresh()

        def export_prods():
            try:
                conn=get_db()
                rows=[dict(r) for r in conn.execute("SELECT * FROM products ORDER BY name").fetchall()]
                conn.close()
                data=[[pr["name"],pr.get("barcode",""),pr.get("unit","Pcs"),
                       f"{pr['sale_rate']:.2f}",f"{pr.get('mrp',0):.2f}",
                       f"{int(pr['gst_percent'])}%",int(get_stock(pr["name"]))] for pr in rows]
                export_to_excel(["Product","Barcode","Unit","Sale Rate","MRP","GST %","Stock"],data,"Products")
            except Exception as e:
                messagebox.showerror("Export Error", str(e))

        bf=tk.Frame(frm,bg=C_BG)
        bf.pack(fill="x", pady=4)
        save_btn=make_btn(bf,"+ Add Product",save)
        save_btn.pack(side="left", padx=(0,4))
        cancel_btn=make_btn(bf,"✕ Cancel",cancel_edit,bg=C_GRAY)
        make_btn(bf,"📥 Excel Export",export_prods,bg=C_GREEN).pack(side="left",padx=4)
        make_btn(bf,"📤 Excel Import",lambda: import_products_from_excel(refresh),bg="#276749").pack(side="left",padx=4)
        make_btn(bf,"📋 Template Download",lambda: _download_excel_template("products"),bg="#2B6CB0").pack(side="left",padx=4)
        make_btn(bf,"📦 Expiry Stock",lambda: self._expiry_stock_window(),bg="#D69E2E").pack(side="left",padx=4)
        # cancel_btn shown only during edit

        # ── Enter Key Navigation: Name→HSN→Rate→MRP→PurRate→GST%→Stock→
        #    Packing→MfgCo→Expiry→Save ─────────────────────────────────────
        def _pf_focus(w):
            def _go(e):
                target = getattr(w, "_entry", w)
                target.focus_set()
                try: target.select_range(0, "end")
                except: pass
                return "break"
            return _go

        pr_name_entry.bind("<Return>", _pf_focus(pr_hsn_entry), add="+")
        pr_hsn_entry.bind("<Return>", _pf_focus(pr_rate_entry), add="+")
        pr_rate_entry.bind("<Return>", _pf_focus(pr_mrp_entry), add="+")
        pr_mrp_entry.bind("<Return>", _pf_focus(pr_pur_entry), add="+")
        pr_pur_entry.bind("<Return>", _pf_focus(pr_gst_cb), add="+")
        pr_gst_cb.bind("<Return>", _pf_focus(pr_stk_entry), add="+")
        pr_stk_entry.bind("<Return>", _pf_focus(pr_packing_entry), add="+")
        pr_packing_entry.bind("<Return>", _pf_focus(pr_mfg_company_entry), add="+")
        pr_mfg_company_entry.bind("<Return>", _pf_focus(pr_expiry_entry), add="+")
        pr_expiry_entry.bind("<Return>", lambda e: save(), add="+")

        refresh()

    def _expiry_stock_window(self):
        """Expiry date manage karne ki window."""
        dlg = tk.Toplevel()
        dlg.title("📦 Expiry Stock Management — Medical")
        dlg.configure(bg=C_BG)
        dlg.geometry("900x620")
        dlg.grab_set()

        # Top tabs
        notebook = ttk.Notebook(dlg)
        notebook.pack(fill="both", expand=True, padx=10, pady=4)

        # ── Tab 1: Add Expiry Batch ──────────────────────────────────────────
        tab_add = tk.Frame(notebook, bg=C_BG, padx=16, pady=12)
        notebook.add(tab_add, text="➕ Batch Add Karo")

        tk.Label(tab_add, text="Naya Batch — Expiry Date ke saath",
                 font=("Segoe UI",11,"bold"), bg=C_BG, fg=C_DARK).pack(anchor="w", pady=(0,3))

        frm2 = tk.Frame(tab_add, bg=C_BG); frm2.pack(fill="x")
        for i in range(4): frm2.columnconfigure(i, weight=1)

        # Product dropdown
        tk.Label(frm2,text="Product",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY
                 ).grid(row=0,column=0,sticky="w",padx=6,pady=(4,1))
        conn_=get_db()
        prod_names=[r[0] for r in conn_.execute("SELECT name FROM products ORDER BY name").fetchall()]
        conn_.close()
        v_prod = tk.StringVar()
        cb_prod = ttk.Combobox(frm2,textvariable=v_prod,values=prod_names,width=22,state="readonly")
        cb_prod.grid(row=0,column=0,sticky="sw",padx=6,pady=(20,2))

        tk.Label(frm2,text="Batch No",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY
                 ).grid(row=0,column=1,sticky="w",padx=6,pady=(4,1))
        v_batch = tk.StringVar()
        ttk.Entry(frm2,textvariable=v_batch,width=14).grid(row=0,column=1,sticky="sw",padx=6,pady=(20,2))

        tk.Label(frm2,text="Mfg Date (YYYY-MM-DD)",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY
                 ).grid(row=0,column=2,sticky="w",padx=6,pady=(4,1))
        v_mfg = tk.StringVar()
        make_date_entry(frm2, v_mfg, width=13).grid(row=0,column=2,sticky="sw",padx=6,pady=(20,2))

        tk.Label(frm2,text="Expiry Date (MM/YY)",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY
                 ).grid(row=0,column=3,sticky="w",padx=6,pady=(4,1))
        v_exp = tk.StringVar()
        make_exp_mmyy_entry(frm2, v_exp, width=8).grid(row=0,column=3,sticky="sw",padx=6,pady=(20,2))

        tk.Label(frm2,text="Quantity",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY
                 ).grid(row=1,column=0,sticky="w",padx=6,pady=(2,0))
        v_qty = tk.StringVar(value="0")
        ttk.Entry(frm2,textvariable=v_qty,width=10).grid(row=1,column=0,sticky="sw",padx=6,pady=(22,2))

        tk.Label(frm2,text="Purchase Rate",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY
                 ).grid(row=1,column=1,sticky="w",padx=6,pady=(2,0))
        v_pur = tk.StringVar(value="0")
        ttk.Entry(frm2,textvariable=v_pur,width=10).grid(row=1,column=1,sticky="sw",padx=6,pady=(22,2))

        tk.Label(frm2,text="MRP",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY
                 ).grid(row=1,column=2,sticky="w",padx=6,pady=(2,0))
        v_mrp2 = tk.StringVar(value="0")
        ttk.Entry(frm2,textvariable=v_mrp2,width=10).grid(row=1,column=2,sticky="sw",padx=6,pady=(22,2))

        tk.Label(frm2,text="Supplier",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY
                 ).grid(row=1,column=3,sticky="w",padx=6,pady=(2,0))
        v_sup = tk.StringVar()
        ttk.Entry(frm2,textvariable=v_sup,width=18).grid(row=1,column=3,sticky="sw",padx=6,pady=(22,2))

        # ── Tab 2: Expiry Stock List ─────────────────────────────────────────
        tab_list = tk.Frame(notebook, bg=C_BG)
        notebook.add(tab_list, text="📋 Expiry List")

        tbl2 = make_table(tab_list,
                          ["Product","Batch","Mfg","Expiry","Qty","Rate","MRP","Supplier","Status"],
                          [20,10,12,12,6,8,8,14,10])
        tbl2.pack(fill="both", expand=True, padx=8, pady=4)

        def refresh_list():
            clear_table_rows(tbl2)
            conn=get_db()
            rows=[dict(r) for r in conn.execute(
                "SELECT * FROM expiry_stock ORDER BY " + EXP_KEY_SQL).fetchall()]
            conn.close()
            today_key = exp_key_for_date(datetime.date.today())
            soon_key  = exp_key_for_date(datetime.date.today()+datetime.timedelta(days=30))
            for i,r in enumerate(rows):
                bg=C_WHITE if i%2==0 else "#F7FAFC"
                ek = exp_sort_key(r["expiry_date"])
                if ek < today_key:
                    status = "❌ EXPIRED"; row_fg="#C53030"
                elif ek <= soon_key:
                    status = "⚠️ Soon"; row_fg="#D69E2E"
                else:
                    status = "✅ OK"; row_fg="#276749"
                vals=[r["product"],r["batch_no"],r["mfg_date"],fmt_exp_mmyy(r["expiry_date"]),
                      int(r["qty"]),f"₹{r['purchase_rate']:.2f}",f"₹{r['mrp']:.2f}",
                      r["supplier"],status]
                for j,val in enumerate(vals):
                    fg = row_fg if j in [0,8] else C_GRAY
                    tk.Label(tbl2,text=str(val),font=("Segoe UI",9),
                             bg=bg,fg=fg,anchor="w",padx=4,pady=4
                             ).grid(row=i+1,column=j,sticky="nsew",padx=1)

        def add_batch():
            prod = v_prod.get().strip()
            exp  = v_exp.get().strip()
            if not prod: messagebox.showerror("Error","Product select karo!", parent=dlg); return
            if not exp:  messagebox.showerror("Error","Expiry date bharo!", parent=dlg); return
            import re as _re_eb
            if not _re_eb.match(r"^(0[1-9]|1[0-2])/\d{2}$", exp):
                messagebox.showerror("Format Galat!","Expiry date MM/YY format mein honi chahiye.\nExample: 12/27", parent=dlg); return
            try:
                qty=float(v_qty.get() or 0)
                pur=float(v_pur.get() or 0)
                mrp=float(v_mrp2.get() or 0)
            except:
                messagebox.showerror("Error","Qty/Rate number hona chahiye!", parent=dlg); return
            conn=get_db()
            conn.execute(
                "INSERT INTO expiry_stock(product,batch_no,mfg_date,expiry_date,qty,purchase_rate,mrp,supplier)"
                " VALUES(?,?,?,?,?,?,?,?)",
                (prod,v_batch.get(),v_mfg.get(),exp,qty,pur,mrp,v_sup.get()))
            conn.commit(); conn.close()
            messagebox.showinfo("Done","Batch add ho gaya! ✅", parent=dlg)
            v_batch.set(""); v_qty.set("0"); v_mfg.set(""); v_exp.set("")
            notebook.select(tab_list)
            refresh_list()

        make_btn(frm2,"✅ Batch Save Karo",add_batch,bg=C_ACCENT
                 ).grid(row=2,column=0,columnspan=2,sticky="w",padx=6,pady=12)

        refresh_list()

    # ══════════════════════════════════════════════════════════════════════════
    #  PARTIES
    # ══════════════════════════════════════════════════════════════════════════

    def _pg_parties(self):
        p=tk.Frame(self.content,bg=C_BG,padx=20,pady=14)
        p.pack(fill="x")
        section_title(p,"Parties / Customers / Suppliers")

        STATES = ["Andhra Pradesh","Arunachal Pradesh","Assam","Bihar","Chhattisgarh",
                  "Goa","Gujarat","Haryana","Himachal Pradesh","Jharkhand","Karnataka",
                  "Kerala","Madhya Pradesh","Maharashtra","Manipur","Meghalaya","Mizoram",
                  "Nagaland","Odisha","Punjab","Rajasthan","Sikkim","Tamil Nadu","Telangana",
                  "Tripura","Uttar Pradesh","Uttarakhand","West Bengal",
                  "Delhi","Jammu & Kashmir","Ladakh","Chandigarh","Puducherry"]

        # ── Form container ────────────────────────────────────────────────────
        frm = tk.Frame(p, bg=C_BG)
        frm.pack(fill="x", pady=(0,3))

        def lbl(parent, text):
            tk.Label(parent, text=text, font=("Segoe UI",9),
                     bg=C_BG, fg=C_GRAY).pack(side="left", padx=(8,2))

        def ent(parent, var, width):
            e = ttk.Entry(parent, textvariable=var, width=width)
            e.pack(side="left", padx=(0,6))
            return e

        # Row 1: saare fields ek hi row mein — compact widths, readable labels
        row1 = tk.Frame(frm, bg=C_BG); row1.pack(fill="x", pady=(4,2))
        pt_name = tk.StringVar()
        lbl(row1, "Party Name");  pt_name_entry = ent(row1, pt_name, 16)
        pt_type = tk.StringVar(value="Customer")
        lbl(row1, "Type")
        pt_type_cb = ttk.Combobox(row1, textvariable=pt_type,
                     values=["Customer","Supplier","Both"],
                     width=9, state="readonly")
        pt_type_cb.pack(side="left", padx=(0,6))
        pt_mob = tk.StringVar()
        lbl(row1, "Mobile");  pt_mob_entry = ent(row1, pt_mob, 11)
        pt_gst = tk.StringVar()
        lbl(row1, "GSTIN");   pt_gst_entry = ent(row1, pt_gst, 13)
        pt_state = tk.StringVar(value="Uttar Pradesh")
        lbl(row1, "State")
        pt_state_cb = ttk.Combobox(row1, textvariable=pt_state, values=STATES,
                     width=14)
        pt_state_cb.pack(side="left", padx=(0,6))
        pt_email = tk.StringVar()
        lbl(row1, "Email");  pt_email_entry = ent(row1, pt_email, 16)
        pt_addr = tk.StringVar()
        lbl(row1, "Address"); pt_addr_entry = ent(row1, pt_addr, 22)

        edit_id=[None]

        # ── Buttons Row ───────────────────────────────────────────────────────
        bf = tk.Frame(frm, bg=C_BG); bf.pack(fill="x", pady=4)

        tbl=make_table(p,["Party Name","Type","Mobile","GSTIN","State","Address","Edit","Delete"],
                       [18,10,13,18,14,18,7,7])

        def refresh():
            clear_table_rows(tbl)
            conn=get_db()
            for i,pt in enumerate([dict(r) for r in conn.execute("SELECT * FROM parties ORDER BY name").fetchall()]):
                bg=C_WHITE if i%2==0 else "#F7FAFC"
                st = pt.get("state","") or ""
                for j,val in enumerate([pt["name"],pt["ptype"],pt["mobile"],
                                        pt["gstin"],st,pt["address"][:16],"",""]):
                    if j==6:
                        tk.Button(tbl,text="✏ Edit",font=("Segoe UI",8,"bold"),
                                  bg="#2B6CB0",fg=C_WHITE,relief="flat",cursor="hand2",bd=0,pady=3,
                                  command=lambda r=pt: load_edit(r)
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=2,pady=1)
                    elif j==7:
                        tk.Button(tbl,text="🗑 Del",font=("Segoe UI",8,"bold"),
                                  bg=C_RED,fg=C_WHITE,relief="flat",cursor="hand2",bd=0,pady=3,
                                  command=lambda r=pt: delete_party(r)
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=2,pady=1)
                    else:
                        tk.Label(tbl,text=str(val),font=("Segoe UI",9),
                                 bg=bg,fg=C_GRAY,anchor="w",padx=6,pady=4
                                 ).grid(row=i+1,column=j,sticky="nsew",padx=1)

        def load_edit(pt):
            edit_id[0]=pt["id"]
            pt_name.set(pt["name"]); pt_type.set(pt["ptype"])
            pt_mob.set(pt["mobile"]); pt_gst.set(pt["gstin"]); pt_addr.set(pt["address"])
            pt_state.set(pt.get("state","") or "Uttar Pradesh")
            pt_email.set(pt.get("email","") or "")
            save_btn.config(text="💾 Update Party", bg=C_AMBER)
            cancel_btn.pack(side="left",padx=4)

        def cancel_edit():
            edit_id[0]=None
            pt_name.set(""); pt_mob.set(""); pt_gst.set(""); pt_addr.set("")
            pt_type.set("Customer"); pt_state.set("Uttar Pradesh"); pt_email.set("")
            save_btn.config(text="+ Add Party", bg=C_ACCENT)
            cancel_btn.pack_forget()

        def save():
            name=pt_name.get().strip()
            if not name: messagebox.showerror("Error","Party naam bharo!"); return
            conn=get_db()
            try:
                if edit_id[0]:
                    conn.execute("UPDATE parties SET name=?,ptype=?,mobile=?,gstin=?,address=?,state=?,email=? WHERE id=?",
                                 (name,pt_type.get(),pt_mob.get(),pt_gst.get(),pt_addr.get(),
                                  pt_state.get(),pt_email.get().strip(),edit_id[0]))
                    conn.commit()
                    messagebox.showinfo("Done","Party update ho gayi!")
                    cancel_edit()
                else:
                    conn.execute("INSERT INTO parties(name,ptype,mobile,gstin,address,state,email) VALUES(?,?,?,?,?,?,?)",
                                 (name,pt_type.get(),pt_mob.get(),pt_gst.get(),pt_addr.get(),
                                  pt_state.get(),pt_email.get().strip()))
                    conn.commit()
                    messagebox.showinfo("Done","Party add ho gayi!")
                    pt_name.set(""); pt_mob.set(""); pt_gst.set(""); pt_addr.set("")
                    pt_state.set("Uttar Pradesh"); pt_email.set("")
                refresh()
            except sqlite3.IntegrityError:
                messagebox.showerror("Error","Party pehle se exist karti hai!")
            finally: conn.close()

        def delete_party(pt):
            if not messagebox.askyesno("Delete?",f"'{pt['name']}' delete karna chahte ho?"):
                return
            conn=get_db()
            conn.execute("DELETE FROM parties WHERE id=?", (pt["id"],))
            conn.commit(); conn.close()
            refresh()

        def export_parties():
            try:
                conn=get_db()
                rows=[dict(r) for r in conn.execute("SELECT * FROM parties ORDER BY name").fetchall()]
                conn.close()
                data=[[pt["name"],pt["ptype"],pt["mobile"],pt["gstin"],
                       pt.get("state",""),pt.get("email",""),pt["address"]] for pt in rows]
                export_to_excel(["Party Name","Type","Mobile","GSTIN","State","Email","Address"],data,"Parties")
            except Exception as e:
                messagebox.showerror("Export Error", str(e))

        save_btn=make_btn(bf,"+ Add Party",save)
        save_btn.pack(side="left")
        cancel_btn=make_btn(bf,"✕ Cancel",cancel_edit,bg=C_GRAY)
        make_btn(bf,"📥 Excel Export",export_parties,bg=C_GREEN).pack(side="left",padx=4)
        make_btn(bf,"📤 Excel Import",lambda: import_parties_from_excel(refresh),bg="#2B6CB0").pack(side="left",padx=4)
        make_btn(bf,"📋 Template",lambda: _download_excel_template("parties"),bg="#6B46C1").pack(side="left",padx=4)
        refresh()

        # ── Enter Key Navigation: Name→Type→Mobile→GSTIN→State→Email→Address→Save ──
        def _pt_focus(w):
            def _go(e):
                target = getattr(w, "_entry", w)
                target.focus_set()
                try: target.select_range(0, "end")
                except: pass
                return "break"
            return _go
        pt_name_entry.bind("<Return>",  _pt_focus(pt_type_cb),    add="+")
        pt_type_cb.bind("<Return>",     _pt_focus(pt_mob_entry),  add="+")
        pt_mob_entry.bind("<Return>",   _pt_focus(pt_gst_entry),  add="+")
        pt_gst_entry.bind("<Return>",   _pt_focus(pt_state_cb),   add="+")
        pt_state_cb.bind("<Return>",    _pt_focus(pt_email_entry),add="+")
        pt_email_entry.bind("<Return>", _pt_focus(pt_addr_entry), add="+")
        pt_addr_entry.bind("<Return>",  lambda e: save(),         add="+")

    # ══════════════════════════════════════════════════════════════════════════
    #  LEDGER
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_ledger(self):
        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "Ledger")

        nb = ttk.Notebook(p)
        nb.pack(fill="both", expand=True, pady=(0,10))
        tab_cust = tk.Frame(nb, bg=C_BG); nb.add(tab_cust, text="  👤 Customer Ledger  ")
        tab_supp = tk.Frame(nb, bg=C_BG); nb.add(tab_supp, text="  🏭 Supplier Ledger  ")

        def _build_tab(parent, is_cust):
            db_table  = "sales"    if is_cust else "purchases"
            etype     = "sale"     if is_cust else "pur"
            col_lbl   = "Customer" if is_cust else "Supplier"
            type_lbl  = "Sale"     if is_cust else "Purchase"
            pay_lbl   = "Receive kiya (₹):" if is_cust else "Pay kiya (₹):"
            baki_head = "Lena Baaki (Customers se)" if is_cust else "Dena Baaki (Suppliers ko)"

            ff = tk.Frame(parent, bg=C_BG); ff.pack(fill="x", pady=(10,8), padx=8)
            conn = get_db()
            parties = [""]+sorted(set(r[0] for r in conn.execute(
                "SELECT party FROM "+db_table).fetchall()))
            conn.close()
            tk.Label(ff, text=col_lbl+":", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
            l_party = tk.StringVar()
            ttk.Combobox(ff, textvariable=l_party, values=parties, width=22,
                         font=("Segoe UI",9)).pack(side="left", padx=6)
            tk.Label(ff, text="From:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
            l_from = tk.StringVar()
            make_date_entry(ff, l_from, width=12).pack(side="left", padx=4)
            tk.Label(ff, text="To:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
            l_to = tk.StringVar()
            make_date_entry(ff, l_to, width=12).pack(side="left", padx=4)

            tbl = make_table(parent,
                ["Date", type_lbl+" No", col_lbl, "Bill Amt", "Paid", "Baki", "Status", ""],
                [11, 15, 18, 12, 12, 12, 14, 7])

            def open_payment_popup(e):
                import datetime as _dt2
                win2 = tk.Toplevel(parent.winfo_toplevel())
                win2.title("Payment — "+e["bill"])
                win2.geometry("500x530")
                win2.configure(bg=C_BG); win2.grab_set(); win2.resizable(True, True)

                tk.Label(win2,
                    text="Bill: "+e["bill"]+"  |  "+col_lbl+": "+e["party"],
                    font=("Segoe UI",11,"bold"), bg=C_BG, fg="#1A365D"
                    ).pack(pady=(14,2), padx=16, anchor="w")
                tk.Label(win2,
                    text="Bill Amount: ₹{:,.0f}  |  Baki: ₹{:,.0f}".format(e["amt"],e["baki"]),
                    font=("Segoe UI",9), bg=C_BG, fg=C_RED).pack(padx=16, anchor="w")
                tk.Frame(win2, bg=C_BORDER, height=1).pack(fill="x", pady=8, padx=16)

                ef = tk.Frame(win2, bg=C_BG); ef.pack(fill="x", padx=16, pady=4)
                tk.Label(ef, text=pay_lbl, font=("Segoe UI",9), bg=C_BG, fg=C_GRAY
                         ).grid(row=0, column=0, sticky="w", pady=(0,2))
                amt_v = tk.StringVar()
                amt_e = ttk.Entry(ef, textvariable=amt_v, font=("Segoe UI",13), width=14)
                amt_e.grid(row=1, column=0, sticky="w"); amt_e.focus()

                tk.Label(ef, text="Date:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY
                         ).grid(row=0, column=1, sticky="w", padx=(16,0), pady=(0,2))
                date_v = tk.StringVar(value=str(_dt2.date.today()))
                make_date_entry(ef, date_v, width=13).grid(row=1, column=1, padx=(16,0), sticky="w")

                tk.Label(ef, text="Mode:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY
                         ).grid(row=2, column=0, sticky="w", pady=(10,2))
                mode_v = tk.StringVar(value="Cash")
                ttk.Combobox(ef, textvariable=mode_v,
                             values=["Cash","UPI","Bank Transfer","Cheque"],
                             font=("Segoe UI",9), state="readonly", width=16
                             ).grid(row=3, column=0, sticky="w")

                tk.Label(ef, text="Note (optional):", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY
                         ).grid(row=2, column=1, sticky="w", padx=(16,0), pady=(10,2))
                note_v = tk.StringVar()
                ttk.Entry(ef, textvariable=note_v, font=("Segoe UI",9), width=18
                          ).grid(row=3, column=1, padx=(16,0), sticky="w")

                def save_payment():
                    try:
                        amt = float(amt_v.get())
                        if amt <= 0: raise ValueError
                    except:
                        messagebox.showerror("Error","Sahi amount bharo!",parent=win2); return
                    if amt > e["baki"]+0.01:
                        if not messagebox.askyesno("Zyada?",
                            "Baki ₹{:,.0f} hai lekin ₹{:,.0f} enter kiya. Save karo?".format(e["baki"],amt),
                            parent=win2): return
                    conn4 = get_db()
                    conn4.execute(
                        "INSERT INTO bill_payments(bill_type,bill_no,party,pay_date,amount,pay_mode,note) VALUES(?,?,?,?,?,?,?)",
                        (etype, e["bill"], e["party"], date_v.get(), amt, mode_v.get(), note_v.get().strip())
                    )
                    total_paid = conn4.execute(
                        "SELECT COALESCE(SUM(amount),0) FROM bill_payments WHERE bill_type=? AND bill_no=?",
                        (etype, e["bill"])
                    ).fetchone()[0]
                    if round(total_paid,2) >= round(e["amt"],2)-0.01:
                        conn4.execute(
                            "UPDATE "+db_table+" SET pay_mode='Paid' WHERE bill_no=?", (e["bill"],))
                    conn4.commit(); conn4.close()
                    messagebox.showinfo("Saved!","Payment save ho gayi ✓", parent=win2)
                    win2.destroy(); show()

                # Save button HAMESHA neeche fixed
                tk.Frame(win2, bg=C_BORDER, height=1).pack(side="bottom", fill="x", padx=16, pady=(4,0))
                bf = tk.Frame(win2, bg=C_BG); bf.pack(side="bottom", padx=16, anchor="w", pady=(0,10))
                make_btn(bf, "💾 Save Payment", save_payment, bg=C_GREEN).pack(side="left", padx=(0,8))
                make_btn(bf, "✕ Cancel", win2.destroy, bg=C_GRAY).pack(side="left")

                tk.Label(win2, text="Pehle ki payments:", font=("Segoe UI",9,"bold"),
                         bg=C_BG, fg=C_GRAY).pack(padx=16, anchor="w", pady=(10,2))
                ph_tbl = make_table(win2, ["Date","Amount","Mode","Note"], [12,12,14,20])
                ph_tbl.master.master.pack(fill="both", expand=True, padx=16, pady=(0,4))
                conn3 = get_db()
                for ri,pv in enumerate(conn3.execute(
                    "SELECT * FROM bill_payments WHERE bill_type=? AND bill_no=? ORDER BY pay_date",
                    (etype, e["bill"])).fetchall()):
                    table_row(ph_tbl, [fmt_date(pv["pay_date"]),
                        "₹{:,.0f}".format(pv["amount"]),
                        pv["pay_mode"], (pv["note"] or "")[:18]], ri+1)
                conn3.close()

            def show():
                clear_table_rows(tbl)
                party = l_party.get(); frm = l_from.get(); to = l_to.get()
                conn = get_db()
                all_payments = {}
                for bp in conn.execute(
                    "SELECT bill_type,bill_no,COALESCE(SUM(amount),0) as paid FROM bill_payments GROUP BY bill_type,bill_no"
                ).fetchall():
                    all_payments[(bp["bill_type"],bp["bill_no"])] = round(bp["paid"], 2)

                entries = []
                for row in conn.execute("SELECT * FROM "+db_table).fetchall():
                    if party and row["party"] != party: continue
                    if frm   and row["bill_date"] < frm: continue
                    if to    and row["bill_date"] > to:  continue
                    paid = all_payments.get((etype, row["bill_no"]), 0)
                    amt  = round(row["grand_total"], 2)
                    baki = round(max(0, amt - paid), 2)
                    entries.append({
                        "date":  row["bill_date"], "bill":  row["bill_no"],
                        "party": row["party"],     "amt":   amt,
                        "paid":  paid,             "baki":  baki,
                        "due":   row["due_date"] or "", "etype": etype,
                    })
                conn.close()
                entries.sort(key=lambda x: x["date"])

                import datetime as _dt
                today_d = _dt.date.today()
                total_baki = 0
                row_fg = C_GREEN if is_cust else "#C05621"

                for i, e in enumerate(entries):
                    bg   = C_WHITE if i%2==0 else "#F7FAFC"
                    baki = e["baki"]

                    # STATUS — sirf baki pe decide
                    if baki > 0.01:
                        total_baki += baki
                        try: bill_d = _dt.date.fromisoformat(e["date"])
                        except: bill_d = today_d
                        days_since = (today_d - bill_d).days
                        if e["due"]:
                            try: due_d = _dt.date.fromisoformat(e["due"])
                            except: due_d = None
                            if due_d:
                                dl = (due_d - today_d).days
                                if dl < 0:    status=f"OVERDUE {abs(dl)}d"; sfg=C_RED
                                elif dl == 0: status="Due TODAY!";           sfg=C_RED
                                else:         status=f"Due {dl}d baad";      sfg=C_AMBER
                            else: status=f"{days_since}d ho gaye"; sfg=C_AMBER
                        else:     status=f"{days_since}d ho gaye"; sfg=C_AMBER
                    else:
                        status="✔ Paid"; sfg=C_GREEN

                    paid_str = "₹{:,.0f}".format(e["paid"]) if e["paid"] > 0 else "-"
                    baki_str = "₹{:,.0f}".format(baki)      if baki > 0.01    else "✔ Clear"
                    row_vals = [fmt_date(e["date"]), e["bill"],
                                e["party"][:16], "₹{:,.0f}".format(e["amt"]),
                                paid_str, baki_str, status]

                    for j, val in enumerate(row_vals):
                        fg = row_fg if j==1 else (
                             C_GREEN if (j==5 and baki<=0.01) else (
                             sfg     if j==6 else C_GRAY))
                        tk.Label(tbl, text=str(val), font=("Segoe UI",9),
                                 bg=bg, fg=fg, anchor="w", padx=5, pady=4
                                 ).grid(row=i+1, column=j, sticky="nsew", padx=1)

                    if baki > 0.01:
                        tk.Button(tbl, text="💰Pay", font=("Segoe UI",7),
                                  bg="#276749", fg=C_WHITE, relief="flat",
                                  cursor="hand2", bd=0, pady=2,
                                  command=lambda ev=e: open_payment_popup(ev)
                                  ).grid(row=i+1, column=7, sticky="nsew", padx=2, pady=2)
                    else:
                        tk.Label(tbl, text="-", font=("Segoe UI",9),
                                 bg=bg, fg=C_LGRAY, anchor="center"
                                 ).grid(row=i+1, column=7, sticky="nsew", padx=1)

                for w in parent.winfo_children():
                    if getattr(w, "_ledger_summary", False): w.destroy()
                sf = tk.Frame(parent, bg=C_WHITE, highlightthickness=1, highlightbackground=C_BORDER)
                sf._ledger_summary = True
                sf.pack(fill="x", pady=6, padx=8)
                tk.Label(sf, text=baki_head+": ₹{:,.0f}".format(total_baki),
                         font=("Segoe UI",10,"bold"), bg=C_WHITE, fg=C_RED
                         ).pack(side="left", padx=16, pady=6)

            def export_ledger():
                try:
                    party=l_party.get(); frm=l_from.get(); to=l_to.get()
                    conn = get_db()
                    all_payments = {}
                    for bp in conn.execute(
                        "SELECT bill_type,bill_no,COALESCE(SUM(amount),0) as paid FROM bill_payments GROUP BY bill_type,bill_no"
                    ).fetchall():
                        all_payments[(bp["bill_type"],bp["bill_no"])] = round(bp["paid"],2)
                    rows = []
                    for row in [dict(r) for r in conn.execute("SELECT * FROM "+db_table).fetchall()]:
                        if party and row["party"]!=party: continue
                        if frm   and row["bill_date"]<frm: continue
                        if to    and row["bill_date"]>to:  continue
                        paid = all_payments.get((etype, row["bill_no"]), 0)
                        baki = round(max(0, row["grand_total"]-paid), 2)
                        rows.append([fmt_date(row["bill_date"]), row["bill_no"], row["party"],
                                     "{:.0f}".format(row["grand_total"]),
                                     "{:.0f}".format(paid),
                                     "{:.0f}".format(baki),
                                     "Paid" if baki<=0.01 else "Pending"])
                    conn.close()
                    rows.sort(key=lambda x: x[0])
                    export_to_excel(
                        ["Date","Bill No",col_lbl,"Bill Amt","Paid","Baki","Status"],
                        rows, "Ledger_"+col_lbl)
                except Exception as ex:
                    messagebox.showerror("Export Error", str(ex))

            make_btn(ff, "Show", show).pack(side="left", padx=8)
            make_btn(ff, "📥 Excel", export_ledger, bg=C_GREEN).pack(side="left", padx=4)
            show()

        _build_tab(tab_cust, is_cust=True)
        _build_tab(tab_supp, is_cust=False)
    # ══════════════════════════════════════════════════════════════════════════
    #  PAYMENT REMINDERS
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_reminders(self):
        import datetime as _dt
        import webbrowser, urllib.parse

        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "💬 Payment Reminders")

        # ── Filter bar ──────────────────────────────────────────────────────
        ff = tk.Frame(p, bg=C_BG); ff.pack(fill="x", pady=(0,3))

        tk.Label(ff, text="Filter:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_filter = tk.StringVar(value="All Pending")
        ttk.Combobox(ff, textvariable=v_filter, width=16, state="readonly",
                     values=["All Pending","Overdue","Due Today","Due in 7 days","Due in 30 days"]
                     ).pack(side="left", padx=6)

        tk.Label(ff, text="Party:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left", padx=(10,0))
        v_party = tk.StringVar()
        conn0 = get_db()
        all_parties = sorted(set(
            [r[0] for r in conn0.execute("SELECT DISTINCT party FROM sales").fetchall()]
        ))
        conn0.close()
        ttk.Combobox(ff, textvariable=v_party, width=20,
                     values=[""]+all_parties).pack(side="left", padx=4)

        make_btn(ff, "🔍 Refresh", lambda: load_data(), bg="#2B6CB0").pack(side="left", padx=10)
        make_btn(ff, "📲 Send All WhatsApp", lambda: send_all_wa(), bg="#25D366").pack(side="left", padx=4)
        make_btn(ff, "📧 Send All Email", lambda: send_all_email(), bg="#2B6CB0").pack(side="left", padx=4)
        make_btn(ff, "💬 Send All SMS", lambda: send_all_sms(), bg="#7B2D8B").pack(side="left", padx=4)

        # ── Summary bar ─────────────────────────────────────────────────────
        sum_frame = tk.Frame(p, bg=C_WHITE, highlightthickness=1, highlightbackground=C_BORDER)
        sum_frame.pack(fill="x", pady=(0,3))
        lbl_total   = tk.Label(sum_frame, text="Total Pending: ₹0", font=("Segoe UI",10,"bold"),
                               bg=C_WHITE, fg=C_RED, padx=16, pady=6)
        lbl_total.pack(side="left")
        lbl_count   = tk.Label(sum_frame, text="| 0 bills", font=("Segoe UI",9),
                               bg=C_WHITE, fg=C_GRAY, padx=8)
        lbl_count.pack(side="left")
        lbl_overdue = tk.Label(sum_frame, text="| Overdue: 0", font=("Segoe UI",10,"bold"),
                               bg=C_WHITE, fg=C_RED, padx=8)
        lbl_overdue.pack(side="left")

        # ── Table ───────────────────────────────────────────────────────────
        cols   = ["Party","Mobile","Bill No","Bill Amt","Baki","Due Date","Status","WhatsApp","Email","SMS"]
        widths = [16, 13, 14, 11, 11, 12, 14, 9, 8, 7]
        tbl = make_table(p, cols, widths)

        pending_entries = []

        def load_data():
            nonlocal pending_entries
            clear_table_rows(tbl)
            pending_entries.clear()

            filt   = v_filter.get()
            pfilter= v_party.get().strip()
            today  = _dt.date.today()

            conn = get_db()
            all_payments = {}
            for bp in conn.execute(
                "SELECT bill_type,bill_no,COALESCE(SUM(amount),0) as paid "
                "FROM bill_payments GROUP BY bill_type,bill_no"
            ).fetchall():
                all_payments[(bp["bill_type"], bp["bill_no"])] = round(bp["paid"],2)

            entries = []

            # Sales (Customer ko lena hai)
            for row in conn.execute("SELECT * FROM sales").fetchall():
                if pfilter and row["party"] != pfilter: continue
                paid = all_payments.get(("sale", row["bill_no"]), 0)
                baki = round(max(0, row["grand_total"] - paid), 2)
                if baki <= 0.01: continue
                due_str = row["due_date"] or ""
                try:    due_d = _dt.date.fromisoformat(due_str) if due_str else None
                except: due_d = None
                if due_d:
                    dl = (due_d - today).days
                    if dl < 0:    status = f"OVERDUE {abs(dl)}d"; urgency=0
                    elif dl == 0: status = "Due TODAY!";           urgency=1
                    elif dl <= 7: status = f"Due {dl}d baad";      urgency=2
                    else:         status = f"Due {dl}d baad";      urgency=3
                else:
                    days_old = (today - _dt.date.fromisoformat(row["bill_date"])).days
                    status = f"{days_old}d pending"; urgency=3; dl=999

                # Mobile: party_mobile column se try karo, phir parties table se
                try:    mobile = row["party_mobile"] or ""
                except: mobile = ""
                if not mobile:
                    pm = conn.execute("SELECT mobile FROM parties WHERE name=?", (row["party"],)).fetchone()
                    mobile = pm["mobile"] if pm else ""

                entries.append({
                    "type":"Sale","party":row["party"],"mobile":mobile or "",
                    "bill":row["bill_no"],"amt":row["grand_total"],"baki":baki,
                    "due":due_str,"status":status,"urgency":urgency,
                    "dl": dl if due_d else 999
                })

            # Purchases reminders me nahi dikhana — sirf customer (Sale) pending bills hi
            # reminder ka kaam hai, supplier ko reminder bhejne ki zaroorat nahi.
            conn.close()

            # Filter apply
            if filt == "Overdue":
                entries = [e for e in entries if "OVERDUE" in e["status"]]
            elif filt == "Due Today":
                entries = [e for e in entries if "TODAY" in e["status"]]
            elif filt == "Due in 7 days":
                entries = [e for e in entries if e["dl"] <= 7]
            elif filt == "Due in 30 days":
                entries = [e for e in entries if e["dl"] <= 30]

            entries.sort(key=lambda x: x["urgency"])
            pending_entries[:] = entries

            total_baki = sum(e["baki"] for e in entries)
            overdue_cnt= sum(1 for e in entries if "OVERDUE" in e["status"])
            lbl_total.config(  text=f"Total Pending: ₹{total_baki:,.0f}")
            lbl_count.config(  text=f"| {len(entries)} bills")
            lbl_overdue.config(text=f"| Overdue: {overdue_cnt}")

            for i, e in enumerate(entries):
                bg = C_WHITE if i%2==0 else "#F7FAFC"
                if "OVERDUE" in e["status"]: sfg=C_RED
                elif "TODAY" in e["status"]: sfg=C_RED
                else:                        sfg=C_AMBER

                vals = [
                    e["party"][:16], e["mobile"],
                    e["bill"], f"₹{e['amt']:,.0f}", f"₹{e['baki']:,.0f}",
                    fmt_date(e["due"]) if e["due"] else "-", e["status"]
                ]
                col_fgs = [C_GRAY, C_GRAY, "#2B6CB0", C_GRAY, C_RED, C_GRAY, sfg]
                for j,(val,fg) in enumerate(zip(vals, col_fgs)):
                    tk.Label(tbl, text=str(val), font=("Segoe UI",9),
                             bg=bg, fg=fg, anchor="w", padx=4, pady=4
                             ).grid(row=i+1, column=j, sticky="nsew", padx=1)

                # WhatsApp button
                tk.Button(tbl, text="WA", font=("Segoe UI",8,"bold"),
                          bg="#25D366", fg="white", relief="flat", cursor="hand2",
                          command=lambda ev=e: send_wa(ev)
                          ).grid(row=i+1, column=7, sticky="nsew", padx=2, pady=2)

                # Email button
                tk.Button(tbl, text="Mail", font=("Segoe UI",7),
                          bg="#2B6CB0", fg="white", relief="flat", cursor="hand2",
                          command=lambda ev=e: send_email_single(ev)
                          ).grid(row=i+1, column=8, sticky="nsew", padx=2, pady=2)

                # SMS button
                tk.Button(tbl, text="SMS", font=("Segoe UI",8,"bold"),
                          bg="#7B2D8B", fg="white", relief="flat", cursor="hand2",
                          command=lambda ev=e: send_sms(ev)
                          ).grid(row=i+1, column=9, sticky="nsew", padx=2, pady=2)

        # ── WhatsApp single ─────────────────────────────────────────────────
        def send_wa(e):
            dlg = tk.Toplevel(p.winfo_toplevel())
            dlg.title("💬 WhatsApp Reminder")
            dlg.geometry("430x340"); dlg.configure(bg=C_WHITE); dlg.grab_set()

            tk.Label(dlg, text="💬 WhatsApp Payment Reminder", font=("Segoe UI",11,"bold"),
                     bg="#25D366", fg="white", pady=4).pack(fill="x")

            frm2 = tk.Frame(dlg, bg=C_WHITE, padx=16, pady=4); frm2.pack(fill="both", expand=True)

            tk.Label(frm2, text="Mobile:", font=("Segoe UI",9), bg=C_WHITE).pack(anchor="w")
            v_mob = tk.StringVar(value=e["mobile"].replace("+91","").replace(" ",""))
            ttk.Entry(frm2, textvariable=v_mob, width=18).pack(anchor="w", pady=(2,8))

            msg = _build_wa_msg(e)
            tk.Label(frm2, text="Message:", font=("Segoe UI",9), bg=C_WHITE).pack(anchor="w")
            txt = tk.Text(frm2, width=48, height=9, font=("Segoe UI",9), wrap="word")
            txt.pack(fill="x"); txt.insert("1.0", msg)

            def do_send():
                mob = v_mob.get().strip().replace(" ","").replace("-","")
                if not mob:
                    messagebox.showwarning("Mobile?","Mobile number daalo!",parent=dlg); return
                if len(mob)==10 and mob.isdigit(): mob = "91"+mob
                msg2 = txt.get("1.0","end-1c")
                url  = f"https://wa.me/{mob}?text={urllib.parse.quote(msg2)}"
                webbrowser.open(url); dlg.destroy()

            bf2 = tk.Frame(dlg, bg=C_WHITE); bf2.pack(pady=(0,10))
            make_btn(bf2,"💬 WhatsApp Bhejo", do_send, bg="#25D366").pack(side="left",padx=6)
            make_btn(bf2,"Cancel", dlg.destroy, bg=C_GRAY).pack(side="left",padx=6)

        # ── Email single ────────────────────────────────────────────────────
        def send_email_single(e):
            import webbrowser, urllib.parse
            sub  = urllib.parse.quote(f"Payment Reminder — Bill {e['bill']} — BHUGTANEASE")
            body = urllib.parse.quote(_build_email_body(e))
            webbrowser.open(f"mailto:?subject={sub}&body={body}")

        # ── Send All WhatsApp ────────────────────────────────────────────────
        def send_all_wa():
            if not pending_entries:
                messagebox.showinfo("No Data","Pehle Refresh karo!"); return
            no_mobile = [e for e in pending_entries if not e["mobile"].strip()]
            has_mobile= [e for e in pending_entries if e["mobile"].strip()]
            info = f"Total {len(pending_entries)} pending bills.\n"
            if no_mobile:
                info += f"⚠️ {len(no_mobile)} parties ka mobile number nahi hai.\n"
            info += f"\n{len(has_mobile)} parties ko WhatsApp bhejoge?"
            if not messagebox.askyesno("Confirm", info): return
            for e in has_mobile:
                mob = e["mobile"].strip().replace(" ","").replace("-","").replace("+91","")
                if len(mob)==10 and mob.isdigit(): mob = "91"+mob
                msg = _build_wa_msg(e)
                url = f"https://wa.me/{mob}?text={urllib.parse.quote(msg)}"
                webbrowser.open(url)
            messagebox.showinfo("Done!",
                f"{len(has_mobile)} WhatsApp tabs khul gaye!\n"
                + (f"\n{len(no_mobile)} parties skip ki (mobile nahi tha)." if no_mobile else ""))

        # ── Send All Email ───────────────────────────────────────────────────
        def send_all_email():
            if not pending_entries:
                messagebox.showinfo("No Data","Pehle Refresh karo!"); return
            if not messagebox.askyesno("Confirm",
                f"{len(pending_entries)} parties ko email reminder bhejoge?\n"
                "(Ek-ek karke mail client khulega)"): return
            for e in pending_entries:
                import urllib.parse as _up
                sub  = _up.quote(f"Payment Reminder — Bill {e['bill']} — BHUGTANEASE")
                body = _up.quote(_build_email_body(e))
                webbrowser.open(f"mailto:?subject={sub}&body={body}")
            messagebox.showinfo("Done!", f"{len(pending_entries)} email tabs khul gaye!")

        # ── SMS single ──────────────────────────────────────────────────────
        def send_sms(e):
            """SMS bhejo — Windows messaging app ya manual copy."""
            import subprocess, sys, urllib.parse

            dlg = tk.Toplevel(p.winfo_toplevel())
            dlg.title("💬 SMS Reminder")
            dlg.geometry("440x360"); dlg.configure(bg=C_WHITE); dlg.grab_set()

            tk.Label(dlg, text="💬 SMS Payment Reminder", font=("Segoe UI",11,"bold"),
                     bg="#7B2D8B", fg="white", pady=4).pack(fill="x")

            frm2 = tk.Frame(dlg, bg=C_WHITE, padx=16, pady=4)
            frm2.pack(fill="both", expand=True)

            tk.Label(frm2, text="Mobile No:", font=("Segoe UI",9), bg=C_WHITE).pack(anchor="w")
            v_mob = tk.StringVar(value=e["mobile"].replace("+91","").replace(" ",""))
            ttk.Entry(frm2, textvariable=v_mob, width=18).pack(anchor="w", pady=(2,8))

            msg = _build_sms_msg(e)
            tk.Label(frm2, text=f"Message ({len(msg)} chars):", font=("Segoe UI",9), bg=C_WHITE).pack(anchor="w")
            txt = tk.Text(frm2, width=48, height=8, font=("Segoe UI",9), wrap="word")
            txt.pack(fill="x"); txt.insert("1.0", msg)

            # char counter
            char_lbl = tk.Label(frm2, text=f"{len(msg)} chars | {(len(msg)-1)//160+1} SMS",
                                font=("Segoe UI",7), bg=C_WHITE, fg=C_GRAY)
            char_lbl.pack(anchor="e")
            def _update_count(ev=None):
                t = txt.get("1.0","end-1c")
                char_lbl.config(text=f"{len(t)} chars | {(len(t)-1)//160+1} SMS")
            txt.bind("<KeyRelease>", _update_count)

            def do_open_sms():
                mob = v_mob.get().strip().replace(" ","").replace("-","")
                if not mob:
                    messagebox.showwarning("Mobile?","Mobile number daalo!",parent=dlg); return
                if len(mob)==10 and mob.isdigit(): mob = "+91"+mob
                elif not mob.startswith("+"): mob = "+91"+mob
                msg2 = txt.get("1.0","end-1c")
                # sms: URI scheme — Windows Messages / Android
                sms_url = f"sms:{mob}?body={urllib.parse.quote(msg2)}"
                try:
                    import webbrowser
                    webbrowser.open(sms_url)
                except Exception:
                    pass
                # Windows 10/11 — Your Phone / Link to Windows app
                if sys.platform == "win32":
                    try:
                        subprocess.Popen(["start", "", sms_url], shell=True)
                    except Exception:
                        pass
                messagebox.showinfo("SMS",
                    "SMS app kholne ki koshish ki gayi.\n\n"
                    "Agar app nahi khula toh 'Copy Message' se\n"
                    "manually copy karke bhejo.",
                    parent=dlg)

            def do_copy():
                msg2 = txt.get("1.0","end-1c")
                dlg.clipboard_clear()
                dlg.clipboard_append(msg2)
                messagebox.showinfo("Copied!",
                    f"Message copy ho gaya!\n\nMobile: {v_mob.get()}\n\nAb apne SMS app mein paste karo.",
                    parent=dlg)

            bf2 = tk.Frame(dlg, bg=C_WHITE); bf2.pack(pady=(4,10))
            make_btn(bf2, "📱 SMS App Khullo", do_open_sms, bg="#7B2D8B").pack(side="left", padx=4)
            make_btn(bf2, "📋 Copy Message",   do_copy,      bg=C_AMBER ).pack(side="left", padx=4)
            make_btn(bf2, "Cancel",            dlg.destroy,  bg=C_GRAY  ).pack(side="left", padx=4)

        # ── Send All SMS ─────────────────────────────────────────────────────
        def send_all_sms():
            if not pending_entries:
                messagebox.showinfo("No Data","Pehle Refresh karo!"); return
            no_mobile  = [e for e in pending_entries if not e["mobile"].strip()]
            has_mobile = [e for e in pending_entries if e["mobile"].strip()]
            if not has_mobile:
                messagebox.showwarning("Mobile nahi","Kisi bhi party ka mobile nahi hai!"); return

            # Bulk SMS — ek dialog mein sabka message dikhao copy karne ke liye
            dlg = tk.Toplevel(p.winfo_toplevel())
            dlg.title("💬 Send All SMS")
            dlg.geometry("520x500"); dlg.configure(bg=C_WHITE); dlg.grab_set()

            tk.Label(dlg, text="💬 Bulk SMS — Payment Reminders", font=("Segoe UI",11,"bold"),
                     bg="#7B2D8B", fg="white", pady=4).pack(fill="x")

            info_txt = (
                f"Total: {len(pending_entries)} bills  |  "
                f"Mobile available: {len(has_mobile)}  |  "
                f"Skip (no mobile): {len(no_mobile)}"
            )
            tk.Label(dlg, text=info_txt, font=("Segoe UI",9), bg=C_WHITE,
                     fg=C_GRAY, pady=4).pack(fill="x", padx=16)

            tk.Label(dlg, text="Sabke messages neeche hain — copy karke bhejo:",
                     font=("Segoe UI",9,"bold"), bg=C_WHITE, fg="#1A365D"
                     ).pack(anchor="w", padx=16, pady=(4,2))

            txt_all = tk.Text(dlg, font=("Courier",8), wrap="word", height=18)
            sb2 = ttk.Scrollbar(dlg, command=txt_all.yview)
            txt_all.configure(yscrollcommand=sb2.set)
            sb2.pack(side="right", fill="y", padx=(0,4))
            txt_all.pack(fill="both", expand=True, padx=16, pady=4)

            all_msgs = []
            for e in has_mobile:
                mob = e["mobile"].strip()
                all_msgs.append(
                    f"━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                    f"To: {e['party']} | Mobile: {mob}\n"
                    f"━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                    f"{_build_sms_msg(e)}\n\n"
                )
            txt_all.insert("1.0", "".join(all_msgs))
            txt_all.config(state="disabled")

            def copy_all():
                dlg.clipboard_clear()
                dlg.clipboard_append("".join(all_msgs))
                messagebox.showinfo("Copied!","Sabke messages copy ho gaye!\nAb SMS app mein paste karo.", parent=dlg)

            def open_one_by_one():
                import webbrowser, urllib.parse, sys
                for e in has_mobile:
                    mob = e["mobile"].strip().replace(" ","").replace("-","")
                    if len(mob)==10 and mob.isdigit(): mob = "+91"+mob
                    elif not mob.startswith("+"): mob = "+91"+mob
                    sms_url = f"sms:{mob}?body={urllib.parse.quote(_build_sms_msg(e))}"
                    if sys.platform == "win32":
                        import subprocess
                        subprocess.Popen(["start","",sms_url], shell=True)
                    else:
                        webbrowser.open(sms_url)
                messagebox.showinfo("Done!",
                    f"{len(has_mobile)} SMS app tabs kholne ki koshish ki.\n"
                    "Agar app nahi khula toh 'Copy All' use karo.", parent=dlg)

            bf2 = tk.Frame(dlg, bg=C_WHITE); bf2.pack(pady=(0,10))
            make_btn(bf2, "📱 Ek-Ek Karke Khullo", open_one_by_one, bg="#7B2D8B").pack(side="left", padx=6)
            make_btn(bf2, "📋 Copy All Messages",   copy_all,        bg=C_AMBER ).pack(side="left", padx=6)
            make_btn(bf2, "Close",                  dlg.destroy,     bg=C_GRAY  ).pack(side="left", padx=6)

        # ── SMS message builder ──────────────────────────────────────────────
        def _build_sms_msg(e):
            """Short SMS message — 160 chars ke andar rakhne ki koshish."""
            _sm = get_shop()
            due_txt = f" Due:{fmt_date(e['due'])}" if e["due"] else ""
            urgent  = " *OVERDUE*" if "OVERDUE" in e["status"] else (
                      " *Due TODAY*" if "TODAY" in e["status"] else "")
            return (
                f"{_sm['name']}: Priya {e['party']}, "
                f"Bill {e['bill']} ka Rs.{e['baki']:,.0f} baki hai.{due_txt}{urgent} "
                f"Please jald payment karein. "
                + (f"Contact: {_sm['mobile']}" if _sm['mobile'] else "")
            )


        def _build_wa_msg(e):
            import datetime as _dt2
            _wm = get_shop()
            today = _dt2.date.today().strftime("%d-%b-%Y")
            due_txt = f"\nDue Date : {fmt_date(e['due'])}" if e["due"] else ""
            urgency_line = ""
            if "OVERDUE" in e["status"]:
                urgency_line = "\n⚠️ Ye bill *OVERDUE* ho chuka hai. Please turant payment karein."
            elif "TODAY" in e["status"]:
                urgency_line = "\n🔔 Aaj ka due date hai. Please aaj hi payment karein."
            _city = _wm['city'] or _wm['state'] or ""
            return (
                f"🙏 *{_wm['name']} — Payment Reminder*\n"
                f"━━━━━━━━━━━━━━━━━━━\n"
                f"Dear *{e['party']}*,\n\n"
                f"Aapka ek payment pending hai:\n"
                f"Bill No  : *{e['bill']}*\n"
                f"Bill Amt : ₹{e['amt']:,.0f}\n"
                f"Baki     : *₹{e['baki']:,.0f}*"
                f"{due_txt}"
                f"{urgency_line}\n\n"
                f"Payment karne ke liye humse contact karein.\n"
                + (f"📞 {_wm['mobile']}\n\n" if _wm['mobile'] else "\n\n")
                + f"Thank you! 🙏\n"
                f"*{_wm['name']}{', '+_city if _city else ''}*\n"
                f"_(Date: {today})_"
            )

        def _build_email_body(e):
            import datetime as _dt2
            _em = get_shop()
            today = _dt2.date.today().strftime("%d-%b-%Y")
            due_txt = f"\nDue Date   : {fmt_date(e['due'])}" if e["due"] else ""
            urgency_line = ""
            if "OVERDUE" in e["status"]:
                urgency_line = "\n\nNOTE: Ye bill OVERDUE ho chuka hai. Please turant payment karein."
            elif "TODAY" in e["status"]:
                urgency_line = "\n\nNOTE: Aaj ka due date hai. Please aaj hi payment karein."
            return (
                f"Dear {e['party']},\n\n"
                f"Ye ek payment reminder hai {_em['name']} ki taraf se.\n\n"
                f"Bill Details:\n"
                f"Bill No  : {e['bill']}\n"
                f"Bill Amt : Rs. {e['amt']:,.0f}\n"
                f"Baki     : Rs. {e['baki']:,.0f}"
                f"{due_txt}"
                f"{urgency_line}\n\n"
                f"Kripya jald se jald payment karein.\n"
                + (f"Contact: {_em['mobile']}\n\n" if _em['mobile'] else "\n\n")
                + f"Regards,\n"
                f"{_em['name']}{', '+_em['city'] if _em['city'] else ''}\n"
                f"Date: {today}"
            )

        load_data()

    # ══════════════════════════════════════════════════════════════════════════
    #  EXPENSES
    # ══════════════════════════════════════════════════════════════════════════
    # ══════════════════════════════════════════════════════════════════════════
    #  SHOP SETTINGS
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_shopsettings(self):
        pg = tk.Frame(self.content, bg=C_BG)
        pg.pack(fill="both", expand=True)

        tk.Label(pg, text="⚙️  Shop / Firm Settings",
                 font=("Segoe UI", 14, "bold"), bg=C_BG, fg=C_TOP).pack(anchor="w", padx=20, pady=(16,4))
        tk.Label(pg, text="Yahan diye gaye details sabhi bills, reports aur messages pe print honge.",
                 font=("Segoe UI", 9), bg=C_BG, fg=C_LGRAY).pack(anchor="w", padx=20, pady=(0,10))

        # Load current values
        s = get_shop()

        def make_section(parent, title):
            f = tk.LabelFrame(parent, text=f"  {title}  ",
                              font=("Segoe UI", 10, "bold"),
                              bg=C_BG, fg=C_TOP, bd=1, relief="groove",
                              padx=16, pady=12)
            f.pack(fill="x", padx=20, pady=(0,12))
            return f

        def labeled_field(parent, label, default="", width=40, row=None, col=0):
            if row is not None:
                tk.Label(parent, text=label+":", font=("Segoe UI",9,"bold"),
                         bg=C_BG, fg="#444").grid(row=row, column=col*2, sticky="w", padx=(0,8), pady=4)
                var = tk.StringVar(value=default)
                ent_w = tk.Entry(parent, textvariable=var, font=("Segoe UI",9),
                         width=width, relief="solid", bd=1)
                ent_w.grid(row=row, column=col*2+1, sticky="ew", pady=4)
            else:
                tk.Label(parent, text=label+":", font=("Segoe UI",9,"bold"),
                         bg=C_BG, fg="#444").pack(anchor="w", pady=(4,0))
                var = tk.StringVar(value=default)
                ent_w = tk.Entry(parent, textvariable=var, font=("Segoe UI",9),
                         width=width, relief="solid", bd=1)
                ent_w.pack(fill="x", pady=(2,4))
            var._entry = ent_w
            return var

        # ── Section 1: Basic Info ──────────────────────────────────────────
        sec1 = make_section(pg, "🏪 Dukan / Firm Ki Basic Jaankari")
        sec1.columnconfigure(1, weight=1); sec1.columnconfigure(3, weight=1)

        v_name    = labeled_field(sec1, "Firm / Shop Ka Naam",    s["name"],    row=0, col=0, width=28)
        v_mobile  = labeled_field(sec1, "Mobile Number",          s["mobile"],  row=0, col=1, width=18)
        v_gstin   = labeled_field(sec1, "GSTIN",                  s["gstin"],   row=1, col=0, width=22)
        v_email   = labeled_field(sec1, "Email",                  s["email"],   row=1, col=1, width=26)
        v_address = labeled_field(sec1, "Address (Gali/Mohalla)", s["address"], row=2, col=0, width=30)
        v_city    = labeled_field(sec1, "Shahar / City",          s["city"],    row=2, col=1, width=18)

        tk.Label(sec1, text="State:", font=("Segoe UI",9,"bold"),
                 bg=C_BG, fg="#444").grid(row=3, column=0, sticky="w", padx=(0,8), pady=4)
        states = ["Uttar Pradesh","Delhi","Maharashtra","Gujarat","Rajasthan","Madhya Pradesh",
                  "Bihar","West Bengal","Tamil Nadu","Karnataka","Telangana","Andhra Pradesh",
                  "Kerala","Haryana","Punjab","Jharkhand","Chhattisgarh","Odisha","Other"]
        v_state = tk.StringVar(value=s["state"] or "Uttar Pradesh")
        v_state_cb = ttk.Combobox(sec1, textvariable=v_state, values=states,
                     font=("Segoe UI",9), width=22, state="readonly")
        v_state_cb.grid(row=3, column=1, sticky="ew", pady=4)

        # ── Section 2: Bank Details ────────────────────────────────────────
        sec2 = make_section(pg, "🏦 Bank Details (Bills Pe Print Honge)")
        sec2.columnconfigure(1, weight=1); sec2.columnconfigure(3, weight=1)

        v_bank    = labeled_field(sec2, "Bank Ka Naam", s["bank"],    row=0, col=0, width=24)
        v_account = labeled_field(sec2, "Account No.",  s["account"], row=0, col=1, width=20)
        v_ifsc    = labeled_field(sec2, "IFSC Code",    s["ifsc"],    row=1, col=0, width=16)
        v_upi     = labeled_field(sec2, "UPI ID",       s["upi"],     row=1, col=1, width=24)

        # ── Section 3: Bill Footer ─────────────────────────────────────────
        sec3 = make_section(pg, "📄 Bill Ke Neeche Print Hone Wala Note (Optional)")
        v_tnc = tk.Text(sec3, height=3, font=("Segoe UI",9), relief="solid", bd=1, wrap="word")
        v_tnc.insert("1.0", s["tnc"])
        v_tnc.pack(fill="x", pady=4)
        tk.Label(sec3, text="Example: Goods once sold will not be taken back. Subject to Varanasi jurisdiction.",
                 font=("Segoe UI",7), bg=C_BG, fg=C_LGRAY).pack(anchor="w")

        # ── Save Button ────────────────────────────────────────────────────
        btn_frame = tk.Frame(pg, bg=C_BG)
        btn_frame.pack(fill="x", padx=20, pady=16)

        status_lbl = tk.Label(btn_frame, text="", font=("Segoe UI",10,"bold"),
                               bg=C_BG, fg="#276749")
        status_lbl.pack(side="right", padx=20)

        def save_settings():
            pairs = [
                ("shop_name",     v_name.get().strip()),
                ("shop_mobile",   v_mobile.get().strip()),
                ("shop_gstin",    v_gstin.get().strip().upper()),
                ("shop_email",    v_email.get().strip()),
                ("shop_address",  v_address.get().strip()),
                ("shop_city",     v_city.get().strip()),
                ("shop_state",    v_state.get().strip()),
                ("shop_bank",     v_bank.get().strip()),
                ("shop_account",  v_account.get().strip()),
                ("shop_ifsc",     v_ifsc.get().strip().upper()),
                ("shop_upi",      v_upi.get().strip()),
                ("shop_print_tnc",v_tnc.get("1.0","end-1c").strip()),
            ]
            if not pairs[0][1]:
                messagebox.showwarning("Naam Zaruri", "Firm ka naam khali nahi chhod sakte!")
                return
            try:
                conn = get_db()
                for k, v in pairs:
                    conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", (k,v))
                conn.commit()
                conn.close()
                # Verify save by re-reading from DB
                s2 = get_shop()
                if s2["name"] == pairs[0][1]:
                    # ── Header turant update karo ───────────────────────────
                    _new_addr = f"{s2['address']}, {s2['city']}".strip(", ")
                    if s2['gstin']: _new_addr += f"  |  GSTIN: {s2['gstin']}"
                    self._hdr_addr_lbl.config(text=_new_addr or "Shop Settings mein address set karein")
                    if self._hdr_name_lbl:
                        self._hdr_name_lbl.config(text=s2['name'] or "BhugtanEase")
                    messagebox.showinfo("✅ Saved!", f"Shop settings save ho gayi!\n\nFirm: {s2['name']}\nGSTIN: {s2['gstin']}\nMobile: {s2['mobile']}")
                    status_lbl.config(text="✅ Settings save ho gayi! (Bill pe reflect hongi)")
                else:
                    messagebox.showerror("Error", "Save mein problem aayi — dobara try karein!")
            except Exception as e:
                messagebox.showerror("DB Error", f"Settings save nahi hui!\nError: {e}")

        make_btn(btn_frame, "💾  Save Settings", save_settings, bg="#276749").pack(side="left")
        make_btn(btn_frame, "↩  Cancel", lambda: self._show("dashboard"), bg=C_GRAY).pack(side="left", padx=8)

        # ── Enter Key Navigation: Name→Mobile→GSTIN→Email→Address→City→State→
        #    Bank→Account→IFSC→UPI→Save ────────────────────────────────────
        def _stf_focus(var_or_w):
            target = getattr(var_or_w, "_entry", var_or_w)
            def _go(e):
                target.focus_set()
                try: target.select_range(0, "end")
                except: pass
                return "break"
            return _go
        v_name._entry.bind("<Return>",    _stf_focus(v_mobile),  add="+")
        v_mobile._entry.bind("<Return>",  _stf_focus(v_gstin),   add="+")
        v_gstin._entry.bind("<Return>",   _stf_focus(v_email),   add="+")
        v_email._entry.bind("<Return>",   _stf_focus(v_address), add="+")
        v_address._entry.bind("<Return>",_stf_focus(v_city),     add="+")
        v_city._entry.bind("<Return>",    _stf_focus(v_state_cb),add="+")
        v_state_cb.bind("<Return>",       _stf_focus(v_bank),    add="+")
        v_bank._entry.bind("<Return>",    _stf_focus(v_account), add="+")
        v_account._entry.bind("<Return>", _stf_focus(v_ifsc),    add="+")
        v_ifsc._entry.bind("<Return>",    _stf_focus(v_upi),     add="+")
        v_upi._entry.bind("<Return>",     lambda e: save_settings(), add="+")

        # ── Password Change Section ────────────────────────────────────────
        pwd_frame = make_section(pg, "🔑 Login Password Change Karo")

        pwd_grid = tk.Frame(pwd_frame, bg=C_BG)
        pwd_grid.pack(fill="x")

        tk.Label(pwd_grid, text="Current Password:", font=("Segoe UI",9,"bold"),
                 bg=C_BG, fg="#444").grid(row=0, column=0, sticky="w", padx=(0,12), pady=6)
        v_cur_pwd = tk.StringVar()
        pwd_cur_entry = ttk.Entry(pwd_grid, textvariable=v_cur_pwd, show="*", font=("Segoe UI",9),
                  width=22)
        pwd_cur_entry.grid(row=0, column=1, sticky="w", pady=6)

        tk.Label(pwd_grid, text="Naya Password:", font=("Segoe UI",9,"bold"),
                 bg=C_BG, fg="#444").grid(row=1, column=0, sticky="w", padx=(0,12), pady=6)
        v_new_pwd = tk.StringVar()
        pwd_new_entry = ttk.Entry(pwd_grid, textvariable=v_new_pwd, show="*", font=("Segoe UI",9),
                  width=22)
        pwd_new_entry.grid(row=1, column=1, sticky="w", pady=6)

        tk.Label(pwd_grid, text="Password Confirm Karo:", font=("Segoe UI",9,"bold"),
                 bg=C_BG, fg="#444").grid(row=2, column=0, sticky="w", padx=(0,12), pady=6)
        v_conf_pwd = tk.StringVar()
        pwd_conf_entry = ttk.Entry(pwd_grid, textvariable=v_conf_pwd, show="*", font=("Segoe UI",9),
                  width=22)
        pwd_conf_entry.grid(row=2, column=1, sticky="w", pady=6)

        pwd_msg = tk.Label(pwd_frame, text="", font=("Segoe UI",9,"bold"), bg=C_BG)
        pwd_msg.pack(anchor="w", pady=(2,0))

        def change_password():
            cur  = v_cur_pwd.get()
            new  = v_new_pwd.get().strip()
            conf = v_conf_pwd.get().strip()
            if not cur or not new or not conf:
                pwd_msg.config(text="⚠️ Saare fields bharo!", fg="#D69E2E"); return
            if new != conf:
                pwd_msg.config(text="❌ Naya password aur confirm match nahi karte!", fg=C_RED); return
            if len(new) < 4:
                pwd_msg.config(text="⚠️ Password kam se kam 4 characters ka hona chahiye!", fg="#D69E2E"); return
            conn2 = get_db()
            row2 = conn2.execute(
                "SELECT id FROM users WHERE username='admin' AND password=?", (cur,)
            ).fetchone()
            if not row2:
                conn2.close()
                pwd_msg.config(text="❌ Current password galat hai!", fg=C_RED); return
            conn2.execute("UPDATE users SET password=? WHERE username='admin'", (new,))
            conn2.commit(); conn2.close()
            v_cur_pwd.set(""); v_new_pwd.set(""); v_conf_pwd.set("")
            pwd_msg.config(text="✅ Password successfully change ho gaya!", fg="#276749")
            messagebox.showinfo("✅ Done!", "Password change ho gaya!\nAgla login mein naya password use karein.")

        make_btn(pwd_frame, "🔑  Password Change Karo", change_password, bg="#2B6CB0").pack(anchor="w", pady=(8,0))

        def _pwd_focus(w):
            def _go(e):
                w.focus_set()
                try: w.select_range(0, "end")
                except: pass
                return "break"
            return _go
        pwd_cur_entry.bind("<Return>", _pwd_focus(pwd_new_entry), add="+")
        pwd_new_entry.bind("<Return>", _pwd_focus(pwd_conf_entry), add="+")
        pwd_conf_entry.bind("<Return>", lambda e: change_password(), add="+")

        # ── License Info Section ───────────────────────────────────────────
        lic_frame = make_section(pg, "🔐 License Information")
        status, days_left, install_date_str, customer = _get_license_info()

        # Status row
        lic_row1 = tk.Frame(lic_frame, bg=C_BG); lic_row1.pack(fill="x", pady=4)
        if status == 'ok':
            color = "#276749" if days_left > 30 else "#D69E2E"
            status_text = f"✅ Active — {days_left} din baaki"
        elif status == 'expired':
            color = "#C53030"
            status_text = f"❌ Expired — {abs(days_left)} din pehle expire hua"
        else:
            color = "#D69E2E"
            status_text = "⚠️ License set nahi hai"

        tk.Label(lic_row1, text=f"Status: {status_text}",
                 font=("Segoe UI", 10, "bold"), bg=C_BG, fg=color).pack(side="left", padx=8)

        lic_row2 = tk.Frame(lic_frame, bg=C_BG); lic_row2.pack(fill="x", pady=2)
        conn_l = get_db()
        exp_row = conn_l.execute("SELECT value FROM settings WHERE key='license_expiry'").fetchone()
        conn_l.close()
        exp_str = exp_row[0] if exp_row else "—"
        tk.Label(lic_row2, text=f"Customer: {customer or '—'}    |    Expiry: {exp_str}    |    Install: {install_date_str or '—'}",
                 font=("Segoe UI", 9), bg=C_BG, fg=C_GRAY).pack(side="left", padx=8)



    # ══════════════════════════════════════════════════════════════════════════
    #  BACKUP & RESTORE
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_backup(self):
        import datetime as _dt
        import shutil, threading, glob

        cf = self.content
        for w in cf.winfo_children(): w.destroy()

        # ── Page header ──────────────────────────────────────────────────────
        tk.Label(cf, text="💾  Backup & Restore", font=("Segoe UI", 15, "bold"),
                 bg=C_BG, fg=C_DARK).pack(anchor="w", padx=20, pady=(16, 2))
        tk.Label(cf, text="Laptop kharab ho ya data delete ho — backup se sab wapas aayega",
                 font=("Segoe UI", 9), bg=C_BG, fg="#718096").pack(anchor="w", padx=22, pady=(0,12))

        # ── Auto-backup folder path read ─────────────────────────────────────
        conn0 = get_db()
        _auto_folder = conn0.execute(
            "SELECT value FROM settings WHERE key='backup_auto_folder'").fetchone()
        _auto_enabled = conn0.execute(
            "SELECT value FROM settings WHERE key='backup_auto_enabled'").fetchone()
        _auto_keep = conn0.execute(
            "SELECT value FROM settings WHERE key='backup_auto_keep'").fetchone()
        conn0.close()

        auto_folder_var   = tk.StringVar(value=_auto_folder[0]  if _auto_folder  else "")
        auto_enabled_var  = tk.BooleanVar(value=(_auto_enabled[0] == "1") if _auto_enabled else True)
        auto_keep_var     = tk.StringVar(value=_auto_keep[0] if _auto_keep else "10")

        status_var = tk.StringVar(value="")

        def _status(msg, color="#276749"):
            status_var.set(msg)
            status_lbl.config(fg=color)
            cf.after(5000, lambda: status_var.set(""))

        # ── CARD helper ──────────────────────────────────────────────────────
        def _card(parent, title, color):
            outer = tk.Frame(parent, bg=color, bd=0)
            outer.pack(fill="x", padx=20, pady=6)
            tk.Label(outer, text=title, font=("Segoe UI", 10, "bold"),
                     bg=color, fg="#1A365D", pady=6, padx=12).pack(anchor="w")
            inner = tk.Frame(outer, bg="#FFFFFF", padx=14, pady=12)
            inner.pack(fill="x", padx=1, pady=(0,1))
            return inner

        # ════════════════════════════════════════════════════════════════════
        #  CARD 1 — MANUAL BACKUP
        # ════════════════════════════════════════════════════════════════════
        c1 = _card(cf, "📤  Manual Backup — Abhi Backup Lo", "#DBEAFE")

        tk.Label(c1, text="Database ki copy kisi bhi folder mein save karo (USB, Desktop, Google Drive folder, etc.)",
                 font=("Segoe UI", 9), bg="#FFFFFF", fg="#4A5568", wraplength=680, justify="left"
                 ).pack(anchor="w", pady=(0,3))

        def do_manual_backup():
            from tkinter import filedialog
            dest = filedialog.askdirectory(title="Backup kahan save karein?")
            if not dest: return
            ts  = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            out = os.path.join(dest, f"BhugtanEase_Backup_{ts}.db")
            try:
                # Safe copy using SQLite backup API
                import sqlite3 as _sq3
                src_conn = _sq3.connect(DB_FILE)
                dst_conn = _sq3.connect(out)
                src_conn.backup(dst_conn)
                dst_conn.close(); src_conn.close()
                size_kb = os.path.getsize(out) // 1024
                _status(f"✅ Backup saved: {out}  ({size_kb} KB)")
                messagebox.showinfo("Backup Successful! 🎉",
                    f"Backup save ho gayi:\n\n{out}\n\nSize: {size_kb} KB\n\n"
                    "Isko USB ya Google Drive mein copy karein safe rehne ke liye.")
                _refresh_history()
            except Exception as e:
                _status(f"❌ Backup failed: {e}", "#9B2C2C")
                messagebox.showerror("Backup Error", str(e))

        def do_gdrive_backup():
            import sqlite3 as _sq3
            # Google Drive folder common paths (Windows)
            gdrive_paths = []
            home = os.path.expanduser("~")
            # Google Drive for Desktop (new)
            for drive_letter in "DEFGHIJKLMNOPQRSTUVWXYZ":
                gd = f"{drive_letter}:\\My Drive"
                if os.path.isdir(gd):
                    gdrive_paths.append(gd)
            # Google Drive old app / My Drive in user folder
            for candidate in [
                os.path.join(home, "Google Drive"),
                os.path.join(home, "Google Drive", "My Drive"),
                os.path.join(home, "My Drive"),
                os.path.join(home, "GoogleDrive"),
                # Google Drive for Desktop mounted as folder
                os.path.join(home, "AppData", "Local", "Google", "DriveFS"),
            ]:
                if os.path.isdir(candidate):
                    gdrive_paths.append(candidate)
            # DriveFS virtual drive — find actual sync folder
            try:
                import winreg
                key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                    r"SOFTWARE\Google\DriveFS", 0, winreg.KEY_READ)
                val, _ = winreg.QueryValueEx(key, "PerAccountPreferences")
                winreg.CloseKey(key)
            except Exception:
                pass

            if not gdrive_paths:
                messagebox.showwarning("Google Drive Nahi Mila",
                    "Google Drive app install nahi hai ya folder nahi mila.\n\n"
                    "Google Drive install karo:\nhttps://drive.google.com\n\n"
                    "Ya 'Abhi Backup Lo' button se manually folder choose karo.",
                    parent=cf)
                return

            dest = gdrive_paths[0]
            # BhugtanEase backup subfolder banao
            backup_subfolder = os.path.join(dest, "BhugtanEase_Backups")
            os.makedirs(backup_subfolder, exist_ok=True)
            ts  = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            out = os.path.join(backup_subfolder, f"BhugtanEase_Backup_{ts}.db")
            try:
                src_conn = _sq3.connect(DB_FILE)
                dst_conn = _sq3.connect(out)
                src_conn.backup(dst_conn)
                dst_conn.close(); src_conn.close()
                size_kb = os.path.getsize(out) // 1024
                _status(f"✅ Google Drive backup: {out} ({size_kb} KB)")
                messagebox.showinfo("Google Drive Backup! ☁️",
                    f"Backup Google Drive mein save ho gayi!\n\n"
                    f"Folder: {backup_subfolder}\n"
                    f"File: BhugtanEase_Backup_{ts}.db\n"
                    f"Size: {size_kb} KB\n\n"
                    "Google Drive automatically cloud mein sync karega.",
                    parent=cf)
                _refresh_history()
            except Exception as e:
                _status(f"❌ Google Drive backup failed: {e}", "#9B2C2C")
                messagebox.showerror("Backup Error", str(e), parent=cf)

        btn_row = tk.Frame(c1, bg="#FFFFFF"); btn_row.pack(anchor="w")
        make_btn(btn_row, "Abhi Backup Lo", do_manual_backup, bg="#2B6CB0").pack(side="left", padx=(0,8))
        make_btn(btn_row, "Google Drive Backup", do_gdrive_backup, bg="#1a73e8").pack(side="left", padx=(0,8))
        tk.Label(c1, text="Google Drive app install honi chahiye PC pe",
                 font=("Segoe UI", 8, "italic"), bg="#FFFFFF", fg="#718096").pack(anchor="w", pady=(4,0))

        # ════════════════════════════════════════════════════════════════════
        #  CARD 2 — AUTO BACKUP
        # ════════════════════════════════════════════════════════════════════
        c2 = _card(cf, "🔄  Auto Backup — Har Din Apne Aap", "#F0FFF4")

        tk.Label(c2, text="Software band karte waqt automatically backup ho jaayegi — aapko yaad nahi rakhna hoga.",
                 font=("Segoe UI", 9), bg="#FFFFFF", fg="#4A5568", wraplength=680).pack(anchor="w", pady=(0,3))

        # Enable toggle row
        en_f = tk.Frame(c2, bg="#FFFFFF"); en_f.pack(fill="x", pady=(0,3))
        tk.Checkbutton(en_f, text="Auto Backup Enable karo",
                       variable=auto_enabled_var, font=("Segoe UI", 10, "bold"),
                       bg="#FFFFFF", fg="#276749", activebackground="#FFFFFF",
                       selectcolor="#FFFFFF").pack(side="left")

        # Folder row
        fol_f = tk.Frame(c2, bg="#FFFFFF"); fol_f.pack(fill="x", pady=2)
        tk.Label(fol_f, text="Backup Folder:", font=("Segoe UI", 9), bg="#FFFFFF",
                 fg="#4A5568", width=14, anchor="w").pack(side="left")
        fol_entry = tk.Entry(fol_f, textvariable=auto_folder_var, width=52,
                             font=("Segoe UI", 9), relief="solid", bd=1)
        fol_entry.pack(side="left", padx=4)

        def _browse_folder():
            from tkinter import filedialog
            d = filedialog.askdirectory(title="Auto Backup folder choose karo")
            if d: auto_folder_var.set(d)

        make_btn(fol_f, "📁 Browse", _browse_folder, bg="#718096").pack(side="left", padx=4)

        # Keep N copies row
        keep_f = tk.Frame(c2, bg="#FFFFFF"); keep_f.pack(fill="x", pady=4)
        tk.Label(keep_f, text="Purani copies:", font=("Segoe UI", 9), bg="#FFFFFF",
                 fg="#4A5568", width=14, anchor="w").pack(side="left")
        tk.Spinbox(keep_f, from_=1, to=30, textvariable=auto_keep_var, width=6,
                   font=("Segoe UI", 9), relief="solid", bd=1).pack(side="left", padx=4)
        tk.Label(keep_f, text="copies rakho (baaki delete ho jaayengi)",
                 font=("Segoe UI", 9), bg="#FFFFFF", fg="#718096").pack(side="left", padx=4)

        def save_auto_settings():
            folder = auto_folder_var.get().strip()
            enabled = "1" if auto_enabled_var.get() else "0"
            keep    = auto_keep_var.get().strip() or "10"
            if auto_enabled_var.get() and not folder:
                messagebox.showwarning("Folder Missing", "Auto backup ke liye folder select karo!")
                return
            conn2 = get_db()
            for k, v in [("backup_auto_folder", folder),
                          ("backup_auto_enabled", enabled),
                          ("backup_auto_keep",    keep)]:
                conn2.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", (k, v))
            conn2.commit(); conn2.close()
            _status("✅ Auto backup settings save ho gayi!")

        make_btn(c2, "💾 Settings Save Karo", save_auto_settings, bg="#276749").pack(anchor="w", pady=(8, 2))

        # ════════════════════════════════════════════════════════════════════
        #  CARD 3 — RESTORE
        # ════════════════════════════════════════════════════════════════════
        c3 = _card(cf, "♻️  Restore — Purana Data Wapas Lao", "#FFF5F5")

        tk.Label(c3, text="⚠️  Restore karne se CURRENT data REPLACE ho jaayega. Pehle manual backup zaroor lo!",
                 font=("Segoe UI", 9, "bold"), bg="#FFFFFF", fg="#9B2C2C", wraplength=680
                 ).pack(anchor="w", pady=(0,3))

        def do_restore():
            from tkinter import filedialog
            src = filedialog.askopenfilename(
                filetypes=[("BhugtanEase Backup", "*.db"), ("All Files", "*.*")],
                title="Backup file select karo (.db)")
            if not src: return

            confirm = messagebox.askyesno("Confirm Restore",
                f"Kya aap sure hain?\n\n"
                f"File: {os.path.basename(src)}\n\n"
                "Current data DELETE ho jaayega aur backup wala data aayega.\n\n"
                "Aage badhen?", icon="warning")
            if not confirm: return

            # Pehle current ka emergency backup
            ts  = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            emer_path = DB_FILE + f".emergency_{ts}.bak"
            try:
                shutil.copy2(DB_FILE, emer_path)
            except: pass

            try:
                import sqlite3 as _sq3
                src_conn = _sq3.connect(src)
                dst_conn = _sq3.connect(DB_FILE)
                src_conn.backup(dst_conn)
                dst_conn.close(); src_conn.close()
                messagebox.showinfo("Restore Successful! ✅",
                    "Data restore ho gaya!\n\n"
                    f"Emergency backup yahan save hai:\n{emer_path}\n\n"
                    "Software restart karo (band karke dubara kholo).")
                _status("✅ Restore complete — software restart karo!")
            except Exception as e:
                _status(f"❌ Restore failed: {e}", "#9B2C2C")
                messagebox.showerror("Restore Error", str(e))

        make_btn(c3, "♻️  Backup se Restore Karo", do_restore, bg="#9B2C2C").pack(side="left", padx=(0,8))
        tk.Label(c3, text="👆 .db backup file select karo",
                 font=("Segoe UI", 8, "italic"), bg="#FFFFFF", fg="#718096").pack(side="left")

        # ════════════════════════════════════════════════════════════════════
        #  CARD 4 — BACKUP HISTORY
        # ════════════════════════════════════════════════════════════════════
        c4 = _card(cf, "📋  Recent Auto Backups", "#FFFFF0")

        hist_frame = tk.Frame(c4, bg="#FFFFFF"); hist_frame.pack(fill="x")

        def _refresh_history():
            for w in hist_frame.winfo_children(): w.destroy()
            folder = auto_folder_var.get().strip()
            if not folder or not os.path.isdir(folder):
                tk.Label(hist_frame, text="Auto backup folder set nahi hai ya exist nahi karta.",
                         font=("Segoe UI", 9), bg="#FFFFFF", fg="#718096").pack(anchor="w")
                return
            pattern = os.path.join(folder, "BhugtanEase_Backup_*.db")
            files   = sorted(glob.glob(pattern), reverse=True)
            if not files:
                tk.Label(hist_frame, text="Koi backup nahi mili is folder mein abhi tak.",
                         font=("Segoe UI", 9), bg="#FFFFFF", fg="#718096").pack(anchor="w")
                return

            # Header
            hdr = tk.Frame(hist_frame, bg="#DBEAFE"); hdr.pack(fill="x", pady=(0,2))
            for txt, w in [("File Name", 35), ("Size", 10), ("Date & Time", 20)]:
                tk.Label(hdr, text=txt, font=("Segoe UI", 9, "bold"), bg="#DBEAFE",
                         fg="#1A365D", width=w, anchor="w", padx=6, pady=4
                         ).pack(side="left")

            for i, fp in enumerate(files[:15]):
                fname  = os.path.basename(fp)
                sz_kb  = os.path.getsize(fp) // 1024
                try:
                    ts_str = fname.replace("BhugtanEase_Backup_","").replace(".db","")
                    dt_obj = _dt.datetime.strptime(ts_str, "%Y%m%d_%H%M%S")
                    dt_fmt = dt_obj.strftime("%d %b %Y  %I:%M %p")
                except: dt_fmt = "—"
                row_bg = "#F7FAFC" if i%2==0 else "#FFFFFF"
                rw = tk.Frame(hist_frame, bg=row_bg); rw.pack(fill="x")
                for txt, w in [(fname, 35), (f"{sz_kb} KB", 10), (dt_fmt, 20)]:
                    tk.Label(rw, text=txt, font=("Segoe UI", 9), bg=row_bg,
                             fg="#4A5568", width=w, anchor="w", padx=6, pady=3
                             ).pack(side="left")

            if len(files) > 15:
                tk.Label(hist_frame, text=f"... aur {len(files)-15} aur files",
                         font=("Segoe UI",7), bg="#FFFFFF", fg="#718096").pack(anchor="w", pady=2)

        _refresh_history()
        make_btn(c4, "🔄 Refresh", _refresh_history, bg="#718096").pack(anchor="w", pady=(8,0))

        # ── Status bar ───────────────────────────────────────────────────────
        status_lbl = tk.Label(cf, textvariable=status_var,
                              font=("Segoe UI", 9, "bold"), bg=C_BG, fg="#276749")
        status_lbl.pack(anchor="w", padx=22, pady=6)

    # ── Auto-backup trigger — software band karte waqt call ho ──────────────
    def _do_auto_backup_on_exit(self):
        """Main window close hone par auto backup."""
        import datetime as _dt, shutil, glob
        try:
            conn0 = get_db()
            enabled = conn0.execute(
                "SELECT value FROM settings WHERE key='backup_auto_enabled'").fetchone()
            folder  = conn0.execute(
                "SELECT value FROM settings WHERE key='backup_auto_folder'").fetchone()
            keep    = conn0.execute(
                "SELECT value FROM settings WHERE key='backup_auto_keep'").fetchone()
            conn0.close()

            if not enabled or enabled[0] != "1": return
            if not folder  or not folder[0].strip(): return
            dest = folder[0].strip()
            if not os.path.isdir(dest): return

            ts  = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            out = os.path.join(dest, f"BhugtanEase_Backup_{ts}.db")

            import sqlite3 as _sq3
            src_conn = _sq3.connect(DB_FILE)
            dst_conn = _sq3.connect(out)
            src_conn.backup(dst_conn)
            dst_conn.close(); src_conn.close()

            # Purani copies delete karo
            try:
                max_keep = int(keep[0]) if keep else 10
                pattern  = os.path.join(dest, "BhugtanEase_Backup_*.db")
                all_bkp  = sorted(glob.glob(pattern), reverse=True)
                for old in all_bkp[max_keep:]:
                    os.remove(old)
            except: pass
        except: pass

    # ══════════════════════════════════════════════════════════════════════════
    #  EXPENSES
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_expenses(self):
        import datetime as _dt

        # Ensure table
        db0 = get_db()
        db0.execute("""CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            exp_date TEXT NOT NULL, category TEXT NOT NULL,
            description TEXT DEFAULT '', amount REAL NOT NULL,
            pay_mode TEXT DEFAULT 'Cash', ref_no TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP)""")
        db0.commit(); db0.close()

        CATEGORIES = [
            "Rent", "Electricity", "Water", "Salary/Wages",
            "Transport/Freight", "Packaging", "Office Supplies",
            "Marketing/Ads", "Repairs & Maintenance", "Internet/Phone",
            "Bank Charges", "Professional Fees", "Insurance",
            "Fuel/Vehicle", "Misc/Other"
        ]

        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "💸 Expense Management")

        # ── Summary cards ────────────────────────────────────────────────────
        sum_frame = tk.Frame(p, bg=C_BG); sum_frame.pack(fill="x", pady=(0,3))

        lbl_total  = tk.Label(sum_frame, text="This Month: ₹0",
                              font=("Segoe UI",11,"bold"), bg=C_WHITE,
                              fg=C_RED, padx=16, pady=8,
                              highlightthickness=1, highlightbackground=C_BORDER)
        lbl_total.pack(side="left", padx=(0,8))

        lbl_month  = tk.Label(sum_frame, text="This Year: ₹0",
                              font=("Segoe UI",11,"bold"), bg=C_WHITE,
                              fg=C_AMBER, padx=16, pady=8,
                              highlightthickness=1, highlightbackground=C_BORDER)
        lbl_month.pack(side="left", padx=(0,8))

        lbl_count  = tk.Label(sum_frame, text="Total Entries: 0",
                              font=("Segoe UI",9), bg=C_WHITE,
                              fg=C_GRAY, padx=16, pady=8,
                              highlightthickness=1, highlightbackground=C_BORDER)
        lbl_count.pack(side="left")

        # ── Entry Form ──────────────────────────────────────────────────────
        form = tk.Frame(p, bg=C_WHITE, highlightthickness=1, highlightbackground=C_BORDER)
        form.pack(fill="x", pady=(0,10))
        tk.Label(form, text="New Expense Entry", font=("Segoe UI",10,"bold"),
                 bg=C_WHITE, fg="#1A365D").pack(anchor="w", padx=16, pady=(10,6))

        row1 = tk.Frame(form, bg=C_WHITE); row1.pack(fill="x", padx=16, pady=4)

        tk.Label(row1, text="Date:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_date = tk.StringVar(value=str(_dt.date.today()))
        exp_date_entry = make_date_entry(row1, v_date, width=11)
        exp_date_entry.pack(side="left", padx=(2,10))

        tk.Label(row1, text="Category:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_cat = tk.StringVar()
        exp_cat_cb = ttk.Combobox(row1, textvariable=v_cat, values=CATEGORIES,
                     width=16)
        exp_cat_cb.pack(side="left", padx=(2,10))

        tk.Label(row1, text="Amount (₹):", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_amt = tk.StringVar()
        exp_amt_entry = ttk.Entry(row1, textvariable=v_amt, width=9)
        exp_amt_entry.pack(side="left", padx=(2,10))

        tk.Label(row1, text="Pay Mode:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_pay = tk.StringVar(value="Cash")
        exp_pay_cb = ttk.Combobox(row1, textvariable=v_pay, width=11, state="readonly",
                     values=["Cash","UPI","Bank Transfer","Cheque","Credit Card"]
                     )
        exp_pay_cb.pack(side="left", padx=(2,10))

        tk.Label(row1, text="Description:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_desc = tk.StringVar()
        exp_desc_entry = ttk.Entry(row1, textvariable=v_desc, width=22)
        exp_desc_entry.pack(side="left", padx=(2,10))

        tk.Label(row1, text="Ref/Bill No:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_ref = tk.StringVar()
        exp_ref_entry = ttk.Entry(row1, textvariable=v_ref, width=11)
        exp_ref_entry.pack(side="left", padx=(2,0))

        # Edit mode state
        edit_id = [None]

        def reset_form():
            edit_id[0] = None
            v_date.set(str(_dt.date.today()))
            v_cat.set(""); v_amt.set(""); v_desc.set(""); v_ref.set("")
            v_pay.set("Cash")
            save_btn.config(text="💾 Save Expense", bg=C_GREEN)
            cancel_btn.pack_forget()

        def do_save():
            cat = v_cat.get().strip()
            if not cat:
                messagebox.showerror("Error","Category select karo!"); return
            try:
                amt = float(v_amt.get())
                assert amt > 0
            except:
                messagebox.showerror("Error","Amount sahi bharo!"); return

            db = get_db()
            if edit_id[0]:
                db.execute(
                    "UPDATE expenses SET exp_date=?,category=?,description=?,amount=?,pay_mode=?,ref_no=? WHERE id=?",
                    (v_date.get(), cat, v_desc.get().strip(), amt,
                     v_pay.get(), v_ref.get().strip(), edit_id[0]))
                msg = "Expense update ho gaya!"
            else:
                db.execute(
                    "INSERT INTO expenses(exp_date,category,description,amount,pay_mode,ref_no) VALUES(?,?,?,?,?,?)",
                    (v_date.get(), cat, v_desc.get().strip(), amt,
                     v_pay.get(), v_ref.get().strip()))
                msg = "Expense save ho gaya!"
            db.commit(); db.close()
            messagebox.showinfo("✅ Done!", msg)
            reset_form()
            load_expenses()

        def do_delete(eid):
            if not messagebox.askyesno("Delete?","Ye expense delete karo?"): return
            db = get_db()
            db.execute("DELETE FROM expenses WHERE id=?", (eid,))
            db.commit(); db.close()
            load_expenses()

        def do_edit(row_data):
            edit_id[0] = row_data["id"]
            v_date.set(row_data["exp_date"])
            v_cat.set(row_data["category"])
            v_amt.set(str(row_data["amount"]))
            v_desc.set(row_data["description"] or "")
            v_ref.set(row_data["ref_no"] or "")
            v_pay.set(row_data["pay_mode"] or "Cash")
            save_btn.config(text="✏️ Update Expense", bg="#2B6CB0")
            cancel_btn.pack(side="left", padx=4)

        bf = tk.Frame(form, bg=C_WHITE); bf.pack(anchor="w", padx=16, pady=(4,12))
        save_btn = make_btn(bf, "💾 Save Expense", do_save, bg=C_GREEN)
        save_btn.pack(side="left", padx=(0,8))
        cancel_btn = make_btn(bf, "✕ Cancel Edit", reset_form, bg=C_GRAY)
        # cancel_btn hidden by default

        make_btn(bf, "Reset", reset_form, bg=C_GRAY).pack(side="left")

        # ── Enter Key Navigation: Date→Category→Amount→PayMode→Description→Ref→Save ──
        def _exp_focus(w):
            def _go(e):
                target = getattr(w, "_entry", w)
                target.focus_set()
                try: target.select_range(0, "end")
                except: pass
                return "break"
            return _go
        getattr(exp_date_entry, "_entry", exp_date_entry).bind("<Return>", _exp_focus(exp_cat_cb), add="+")
        exp_cat_cb.bind("<Return>",   _exp_focus(exp_amt_entry),  add="+")
        exp_amt_entry.bind("<Return>",_exp_focus(exp_pay_cb),     add="+")
        exp_pay_cb.bind("<Return>",   _exp_focus(exp_desc_entry), add="+")
        exp_desc_entry.bind("<Return>",_exp_focus(exp_ref_entry), add="+")
        exp_ref_entry.bind("<Return>", lambda e: do_save(),       add="+")

        # ── Filter bar ──────────────────────────────────────────────────────
        hf = tk.Frame(p, bg=C_BG); hf.pack(fill="x", pady=(4,4))

        tk.Label(hf, text="From:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_from = tk.StringVar()
        make_date_entry(hf, v_from, width=12).pack(side="left", padx=4)

        tk.Label(hf, text="To:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_to = tk.StringVar()
        make_date_entry(hf, v_to, width=12).pack(side="left", padx=4)

        tk.Label(hf, text="Category:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left", padx=(8,0))
        v_fcat = tk.StringVar()
        ttk.Combobox(hf, textvariable=v_fcat, values=[""]+CATEGORIES,
                     width=18).pack(side="left", padx=4)

        tk.Label(hf, text="Pay Mode:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left", padx=(8,0))
        v_fpay = tk.StringVar()
        ttk.Combobox(hf, textvariable=v_fpay, width=12, state="readonly",
                     values=["","Cash","UPI","Bank Transfer","Cheque","Credit Card"]
                     ).pack(side="left", padx=4)

        make_btn(hf, "🔍 Filter", lambda: load_expenses(), bg="#2B6CB0").pack(side="left", padx=8)
        make_btn(hf, "📥 Excel",  lambda: export_exp(),    bg=C_GREEN  ).pack(side="left", padx=4)

        # ── Category summary ────────────────────────────────────────────────
        tk.Label(p, text="Category-wise Summary", font=("Segoe UI",10,"bold"),
                 bg=C_BG, fg="#1A365D").pack(anchor="w", pady=(8,2))
        cat_tbl = make_table(p,
            ["Category", "Entries", "Total Amount", "% of Total"],
            [24, 9, 16, 12])

        # ── Main expense table ───────────────────────────────────────────────
        tk.Label(p, text="All Expenses", font=("Segoe UI",10,"bold"),
                 bg=C_BG, fg="#1A365D").pack(anchor="w", pady=(8,2))
        tbl = make_table(p,
            ["Date", "Category", "Description", "Amount", "Pay Mode", "Ref No", "Edit", "Del"],
            [11, 18, 22, 12, 12, 12, 6, 6])

        def load_expenses():
            clear_table_rows(tbl)
            clear_table_rows(cat_tbl)

            frm  = v_from.get().strip()
            to   = v_to.get().strip()
            fcat = v_fcat.get().strip()
            fpay = v_fpay.get().strip()

            db = get_db()
            q  = "SELECT * FROM expenses WHERE 1=1"
            params = []
            if frm:  q += " AND exp_date>=?"; params.append(frm)
            if to:   q += " AND exp_date<=?"; params.append(to)
            if fcat: q += " AND category=?";  params.append(fcat)
            if fpay: q += " AND pay_mode=?";  params.append(fpay)
            q += " ORDER BY exp_date DESC, id DESC"
            rows = [dict(r) for r in db.execute(q, params).fetchall()]

            # Summary totals
            today = _dt.date.today()
            month_start = today.replace(day=1).isoformat()
            year_start  = today.replace(month=1, day=1).isoformat()
            month_total = db.execute(
                "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE exp_date>=?", (month_start,)
            ).fetchone()[0]
            year_total  = db.execute(
                "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE exp_date>=?", (year_start,)
            ).fetchone()[0]
            total_all   = db.execute("SELECT COUNT(*) FROM expenses").fetchone()[0]
            db.close()

            lbl_total.config(text=f"This Month: ₹{month_total:,.0f}")
            lbl_month.config(text=f"This Year: ₹{year_total:,.0f}")
            lbl_count.config(text=f"Total Entries: {total_all}")

            # Category-wise summary
            cat_map = {}
            grand_total = sum(r["amount"] for r in rows)
            for r in rows:
                k = r["category"]
                if k not in cat_map: cat_map[k] = {"count":0,"amt":0}
                cat_map[k]["count"] += 1
                cat_map[k]["amt"]   += r["amount"]

            for i,(cat,v) in enumerate(sorted(cat_map.items(), key=lambda x:-x[1]["amt"])):
                pct = f"{v['amt']/grand_total*100:.1f}%" if grand_total>0 else "0%"
                table_row(cat_tbl,
                    [cat, v["count"], f"₹{v['amt']:,.0f}", pct], i+1,
                    fgs=[None, None, C_RED, C_AMBER])

            # Main table
            for i, r in enumerate(rows):
                bg = C_WHITE if i%2==0 else "#F7FAFC"
                vals = [fmt_date(r["exp_date"]), r["category"],
                        (r["description"] or "")[:20],
                        f"₹{r['amount']:,.0f}",
                        r["pay_mode"] or "", r["ref_no"] or ""]
                for j, val in enumerate(vals):
                    fg = C_RED if j==3 else C_GRAY
                    tk.Label(tbl, text=str(val), font=("Segoe UI",9),
                             bg=bg, fg=fg, anchor="w", padx=4, pady=4
                             ).grid(row=i+1, column=j, sticky="nsew", padx=1)

                # Edit button
                tk.Button(tbl, text="✏", font=("Segoe UI",9),
                          bg="#2B6CB0", fg="white", relief="flat", cursor="hand2",
                          command=lambda rd=r: do_edit(rd)
                          ).grid(row=i+1, column=6, sticky="nsew", padx=2, pady=2)

                # Delete button
                tk.Button(tbl, text="🗑", font=("Segoe UI",9),
                          bg=C_RED, fg="white", relief="flat", cursor="hand2",
                          command=lambda eid=r["id"]: do_delete(eid)
                          ).grid(row=i+1, column=7, sticky="nsew", padx=2, pady=2)

        def export_exp():
            db = get_db()
            rows = [dict(r) for r in db.execute(
                "SELECT * FROM expenses ORDER BY exp_date DESC").fetchall()]
            db.close()
            if not rows:
                messagebox.showinfo("No Data","Koi expense nahi hai!"); return
            data = [[fmt_date(r["exp_date"]), r["category"],
                     r["description"] or "", f"₹{r['amount']:,.0f}",
                     r["pay_mode"] or "", r["ref_no"] or ""] for r in rows]
            try:
                export_to_excel(
                    ["Date","Category","Description","Amount","Pay Mode","Ref No"],
                    data, "Expenses")
            except Exception as e:
                messagebox.showerror("Export Error", str(e))

        load_expenses()

    # ══════════════════════════════════════════════════════════════════════════
    #  P&L REPORT
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_pl(self):
        p = self._content_scroll()
        section_title(p, "Profit & Loss Report")

        # ── Filter bar ──
        ff = tk.Frame(p, bg=C_BG); ff.pack(fill="x", pady=(0, 8))
        tk.Label(ff, text="From:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        pl_from = tk.StringVar()
        make_date_entry(ff, pl_from, width=12).pack(side="left", padx=4)
        tk.Label(ff, text="To:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        pl_to = tk.StringVar()
        make_date_entry(ff, pl_to, width=12).pack(side="left", padx=4)
        make_btn(ff, "📊  Generate Report", lambda: gen(), ).pack(side="left", padx=8)

        # ── Summary cards row ──
        summary_frame = tk.Frame(p, bg=C_BG); summary_frame.pack(fill="x", pady=(4, 10))

        # ── Tabs: Product / Bill / Customer / Received-wise ──
        nb = ttk.Notebook(p)
        nb.pack(fill="both", expand=True, pady=(0, 10))

        tab_prod = tk.Frame(nb, bg=C_BG); nb.add(tab_prod, text="  📦 Product-wise  ")
        tab_prodprofit = tk.Frame(nb, bg=C_BG); nb.add(tab_prodprofit, text="  💹 Product Profit Detail  ")
        tab_bill = tk.Frame(nb, bg=C_BG); nb.add(tab_bill, text="  🧾 Bill-wise  ")
        tab_cust = tk.Frame(nb, bg=C_BG); nb.add(tab_cust, text="  👤 Customer-wise  ")
        tab_recv = tk.Frame(nb, bg=C_BG); nb.add(tab_recv, text="  💰 Received-wise P&L  ")
        tab_exp  = tk.Frame(nb, bg=C_BG); nb.add(tab_exp,  text="  💸 Expenses  ")

        # ── Product table ──
        tbl_prod = make_table(tab_prod,
            ["Product", "Qty Sold", "Sale Amt (ex-GST)", "GST Collected", "Purchase Cost", "Gross Profit"],
            [26, 10, 18, 14, 16, 14])

        # ── Product Profit Detail table ──
        tk.Label(tab_prodprofit,
                 text="  Har product ka Sale Price - Purchase Price = Net Profit (unit-wise aur total)",
                 font=("Segoe UI", 9, "italic"), bg=C_BG, fg=C_LGRAY).pack(anchor="w", padx=6, pady=(6,2))
        tbl_prodprofit = make_table(tab_prodprofit,
            ["Product", "Qty Sold", "Avg Sale Rate", "Avg Pur Rate", "Profit/Unit", "Total Profit"],
            [26, 10, 15, 15, 14, 14])

        # ── Bill table ──
        tbl_bill = make_table(tab_bill,
            ["Bill No", "Date", "Customer", "Sale Amt", "GST", "Purchase Cost*", "Profit/Loss"],
            [18, 12, 20, 13, 10, 14, 14])
        tk.Label(tab_bill, text="  * Purchase cost = bill items ke products ka proportional purchase cost",
                 font=("Segoe UI",7), bg=C_BG, fg=C_LGRAY).pack(anchor="w", padx=6)

        # ── Customer table ──
        tbl_cust = make_table(tab_cust,
            ["Customer", "Bills", "Total Billed", "GST", "Purchase Cost", "Gross Profit"],
            [24, 8, 14, 10, 14, 14])

        # ── Received-wise table ──
        tbl_recv = make_table(tab_recv,
            ["Customer", "Total Billed", "Amount Received", "Pending", "Cost (prop.)", "Profit on Received"],
            [22, 14, 16, 12, 14, 18])
        tk.Label(tab_recv, text="  * Profit on Received = sirf received amount ke proportion mein profit calculate kiya gaya hai",
                 font=("Segoe UI",7), bg=C_BG, fg=C_LGRAY).pack(anchor="w", padx=6)

        # ── Expenses tab ──
        tbl_exp = make_table(tab_exp,
            ["Date","Category","Description","Amount","Pay Mode","Ref No"],
            [11, 18, 24, 13, 13, 12])
        exp_total_lbl = tk.Label(tab_exp, text="",
            font=("Segoe UI",10,"bold"), bg=C_BG, fg=C_RED, anchor="e", padx=12)
        exp_total_lbl.pack(fill="x", pady=4)

        def gen():
            # ── Clear all tables ──
            for tbl in [tbl_prod, tbl_prodprofit, tbl_bill, tbl_cust, tbl_recv, tbl_exp]:
                clear_table_rows(tbl)
            for w in summary_frame.winfo_children(): w.destroy()

            frm = pl_from.get().strip(); to = pl_to.get().strip()

            conn = get_db()
            sales   = [dict(r) for r in conn.execute("SELECT * FROM sales ORDER BY bill_date").fetchall()]
            purs    = [dict(r) for r in conn.execute("SELECT * FROM purchases").fetchall()]
            sitems  = [dict(r) for r in conn.execute("SELECT * FROM sale_items").fetchall()]
            pitems  = [dict(r) for r in conn.execute("SELECT * FROM purchase_items").fetchall()]
            prods   = [dict(r) for r in conn.execute("SELECT * FROM products ORDER BY name").fetchall()]
            payments= [dict(r) for r in conn.execute(
                "SELECT * FROM bill_payments WHERE bill_type='sale'").fetchall()]
            # Expenses
            eq = "SELECT * FROM expenses WHERE 1=1"
            ep = []
            if frm: eq += " AND exp_date>=?"; ep.append(frm)
            if to:  eq += " AND exp_date<=?"; ep.append(to)
            eq += " ORDER BY exp_date DESC"
            expenses = [dict(r) for r in conn.execute(eq, ep).fetchall()]
            conn.close()

            # Apply date filter
            if frm:
                sales = [s for s in sales if s["bill_date"] >= frm]
                purs  = [p for p in purs  if p["bill_date"] >= frm]
            if to:
                sales = [s for s in sales if s["bill_date"] <= to]
                purs  = [p for p in purs  if p["bill_date"] <= to]

            sids = {s["id"] for s in sales}
            pids = {p["id"] for p in purs}
            si   = [x for x in sitems if x["sale_id"] in sids]
            pi   = [x for x in pitems if x["purchase_id"] in pids]

            # Totals for summary
            total_sale  = sum(s["grand_total"] for s in sales)
            gst_coll    = sum(x["gst_amt"] for x in si)
            total_exp   = sum(e["amount"] for e in expenses)

            total_recv  = sum(pay["amount"] for pay in payments
                              if any(s["bill_no"] == pay["bill_no"] for s in sales))

            # ── Sale Returns (date-filtered) ──────────────────────────────────
            conn2 = get_db()
            sr_q  = "SELECT * FROM returns WHERE return_type='sale_return'"
            sr_p  = []
            if frm: sr_q += " AND return_date>=?"; sr_p.append(frm)
            if to:  sr_q += " AND return_date<=?"; sr_p.append(to)
            sr_rows2 = [dict(r) for r in conn2.execute(sr_q, sr_p).fetchall()]

            # ── FIFO-based COGS (accurate) ────────────────────────────────────
            conn3 = get_db()
            fifo_q = "SELECT fc.sale_bill, fc.product, SUM(fc.total_cost) as cogs FROM fifo_consumption fc"
            fifo_params = []
            if frm or to:
                fifo_q += " WHERE 1=1"
                if frm: fifo_q += " AND fc.sale_date>=?"; fifo_params.append(frm)
                if to:  fifo_q += " AND fc.sale_date<=?"; fifo_params.append(to)
            fifo_q += " GROUP BY fc.sale_bill, fc.product"
            fifo_rows = [dict(r) for r in conn3.execute(fifo_q, fifo_params).fetchall()]
            conn3.close()
            # Map: product -> total FIFO COGS
            fifo_cogs_by_prod = {}
            fifo_cogs_by_bill = {}
            for fr in fifo_rows:
                fifo_cogs_by_prod[fr["product"]] = fifo_cogs_by_prod.get(fr["product"],0) + fr["cogs"]
                fifo_cogs_by_bill[fr["sale_bill"]] = fifo_cogs_by_bill.get(fr["sale_bill"],0) + fr["cogs"]

            # ── Avg purchase rate map (fallback for products without FIFO layers) ──
            conn2 = get_db()
            sr_q  = "SELECT * FROM returns WHERE return_type='sale_return'"
            sr_p  = []
            if frm: sr_q += " AND return_date>=?"; sr_p.append(frm)
            if to:  sr_q += " AND return_date<=?"; sr_p.append(to)
            sr_rows2 = [dict(r) for r in conn2.execute(sr_q, sr_p).fetchall()]
            all_pi2 = [dict(r) for r in conn2.execute("SELECT * FROM purchase_items").fetchall()]
            conn2.close()

            pur_cost_map2 = {}
            for x in all_pi2:
                k = x["product"]
                if k not in pur_cost_map2: pur_cost_map2[k] = {"cost": 0, "qty": 0}
                ex_gst = x.get("taxable", 0) or max(0, x.get("total", 0) - x.get("gst_amt", 0))
                pur_cost_map2[k]["cost"] += ex_gst
                pur_cost_map2[k]["qty"]  += x.get("qty", 0)

            # Sale Return amounts
            sr_total2   = sum(r["total_amt"] for r in sr_rows2)
            sr_gst2     = sum(r["gst_amt"]   for r in sr_rows2)
            sr_taxable2 = sr_total2 - sr_gst2

            # Net Sale figures
            net_sale2 = total_sale - sr_total2
            taxable   = (total_sale - gst_coll) - sr_taxable2   # ex-GST, after returns

            # COGS: FIFO se (agar available) warna avg purchase cost
            cogs_summary = 0.0
            for x in si:
                prod = x["product"]
                if prod in fifo_cogs_by_prod:
                    # FIFO cogs already by bill; sum by product across filtered bills
                    pass  # handled below
                else:
                    d = pur_cost_map2.get(prod)
                    if d and d["qty"] > 0:
                        cogs_summary += (d["cost"] / d["qty"]) * x.get("qty", 0)
            # Add FIFO COGS for products that have layers
            cogs_fifo_total = sum(v for prod, v in fifo_cogs_by_prod.items())
            cogs_summary += cogs_fifo_total

            # Returned items ka COGS minus karo
            cogs_returned2 = 0.0
            for r in sr_rows2:
                d = pur_cost_map2.get(r["product"])
                if d and d["qty"] > 0:
                    cogs_returned2 += (d["cost"] / d["qty"]) * r["qty"]
            cogs_net2 = cogs_summary - cogs_returned2

            gross_profit = taxable - cogs_net2
            net_profit   = gross_profit - total_exp

            # ── Summary cards ──
            stat_card(summary_frame, "Gross Sale",              f"₹{round(total_sale):,}",    C_ACCENT)
            stat_card(summary_frame, "Sale Returns",            f"₹{round(sr_total2):,}",   C_RED)
            stat_card(summary_frame, "Net Sale",                f"₹{round(net_sale2):,}",     "#2B6CB0")
            stat_card(summary_frame, "GST Collected",           f"₹{round(gst_coll):,}",      C_AMBER)
            stat_card(summary_frame, "Net Sale (ex-GST)",       f"₹{round(taxable):,}",       C_PURPLE)
            stat_card(summary_frame, "Cost of Sold Items",      f"₹{round(cogs_net2):,}",     C_AMBER)
            stat_card(summary_frame, "Total Expenses",          f"₹{round(total_exp):,}",     C_RED)
            stat_card(summary_frame, "Gross Profit",            f"₹{round(gross_profit):,}",
                      C_GREEN if gross_profit >= 0 else C_RED)
            stat_card(summary_frame, "Net Profit",              f"₹{round(net_profit):,}",
                      C_GREEN if net_profit >= 0 else C_RED)
            # pur_cost_map already computed above as pur_cost_map2 — reuse karo
            pur_cost_map = pur_cost_map2

            def avg_pur_rate(prod):
                d = pur_cost_map.get(prod)
                if d and d["qty"] > 0: return d["cost"] / d["qty"]
                return 0

            # ════ TAB 1: PRODUCT-WISE ════
            prod_map = {}
            for x in si:
                k = x["product"]
                if k not in prod_map:
                    prod_map[k] = {"qty": 0, "sale": 0, "gst": 0, "pur": 0}
                prod_map[k]["qty"]  += x["qty"]
                prod_map[k]["sale"] += x.get("taxable", 0)
                prod_map[k]["gst"]  += x.get("gst_amt", 0)
                prod_map[k]["pur"]  += x["qty"] * avg_pur_rate(k)

            row_n = 1
            for k, v in sorted(prod_map.items()):
                pf = v["sale"] - v["pur"]
                table_row(tbl_prod,
                    [k, f'{v["qty"]:.2f}', f'₹{v["sale"]:,.2f}',
                     f'₹{v["gst"]:,.2f}', f'₹{v["pur"]:,.2f}', f'₹{pf:,.2f}'],
                    row_n,
                    fgs=[None, None, None, None, None, C_GREEN if pf >= 0 else C_RED])
                row_n += 1

            # Total row
            t_sale = sum(v["sale"] for v in prod_map.values())
            t_gst  = sum(v["gst"]  for v in prod_map.values())
            t_pur  = sum(v["pur"]  for v in prod_map.values())
            t_pf   = t_sale - t_pur
            table_row(tbl_prod,
                ["TOTAL", "", f'₹{t_sale:,.2f}', f'₹{t_gst:,.2f}',
                 f'₹{t_pur:,.2f}', f'₹{t_pf:,.2f}'],
                row_n, fgs=[C_ACCENT]*6)

            # ════ TAB: PRODUCT PROFIT DETAIL (Sale Rate - Purchase Rate per unit) ════
            row_n = 1
            pp_total_profit = 0
            for k, v in sorted(prod_map.items()):
                qty      = v["qty"] if v["qty"] > 0 else 1
                avg_sale = v["sale"] / qty          # avg sale rate (ex-GST)
                avg_pur  = avg_pur_rate(k)          # avg purchase rate (ex-GST)
                profit_u = avg_sale - avg_pur       # profit per unit
                total_pf = v["sale"] - v["pur"]     # total profit for this product
                pp_total_profit += total_pf

                color = C_GREEN if profit_u >= 0 else C_RED
                table_row(tbl_prodprofit,
                    [k,
                     f'{v["qty"]:.2f}',
                     f'₹{avg_sale:,.2f}',
                     f'₹{avg_pur:,.2f}' if avg_pur > 0 else "N/A (pur nahi)",
                     f'₹{profit_u:,.2f}' if avg_pur > 0 else "—",
                     f'₹{total_pf:,.2f}'],
                    row_n,
                    fgs=[None, None, C_ACCENT, C_AMBER,
                         color if avg_pur > 0 else C_LGRAY,
                         C_GREEN if total_pf >= 0 else C_RED])
                row_n += 1

            # Total row for product profit
            table_row(tbl_prodprofit,
                ["TOTAL", f'{sum(v["qty"] for v in prod_map.values()):.2f}',
                 "", "", "",
                 f'₹{pp_total_profit:,.2f}'],
                row_n,
                fgs=[C_ACCENT, C_ACCENT, None, None, None,
                     C_GREEN if pp_total_profit >= 0 else C_RED])

            # ════ TAB 2: BILL-WISE ════
            row_n = 1
            for s in sales:
                s_items = [x for x in si if x["sale_id"] == s["id"]]
                s_sale  = sum(x.get("taxable", 0) for x in s_items)
                s_gst   = sum(x.get("gst_amt", 0) for x in s_items)
                s_pur   = sum(x["qty"] * avg_pur_rate(x["product"]) for x in s_items)
                s_pf    = s_sale - s_pur
                table_row(tbl_bill,
                    [s["bill_no"], fmt_date(s["bill_date"]), s.get("party", "")[:18],
                     f'₹{s_sale:,.2f}', f'₹{s_gst:,.2f}',
                     f'₹{s_pur:,.2f}', f'₹{s_pf:,.2f}'],
                    row_n,
                    fgs=[None, None, None, None, None, None, C_GREEN if s_pf >= 0 else C_RED])
                row_n += 1

            # ════ TAB 3: CUSTOMER-WISE ════
            cust_map = {}
            for s in sales:
                k = s.get("party", "Unknown") or "Unknown"
                if k not in cust_map:
                    cust_map[k] = {"bills": 0, "sale": 0, "gst": 0, "pur": 0}
                cust_map[k]["bills"] += 1
                s_items = [x for x in si if x["sale_id"] == s["id"]]
                cust_map[k]["sale"] += sum(x.get("taxable", 0) for x in s_items)
                cust_map[k]["gst"]  += sum(x.get("gst_amt", 0) for x in s_items)
                cust_map[k]["pur"]  += sum(x["qty"] * avg_pur_rate(x["product"]) for x in s_items)

            row_n = 1
            for k, v in sorted(cust_map.items()):
                pf = v["sale"] - v["pur"]
                table_row(tbl_cust,
                    [k, v["bills"], f'₹{v["sale"]:,.2f}',
                     f'₹{v["gst"]:,.2f}', f'₹{v["pur"]:,.2f}', f'₹{pf:,.2f}'],
                    row_n,
                    fgs=[None, None, None, None, None, C_GREEN if pf >= 0 else C_RED])
                row_n += 1

            # ════ TAB 4: RECEIVED-WISE ════
            # received amount per bill
            recv_by_bill = {}
            for pay in payments:
                bn = pay["bill_no"]
                recv_by_bill[bn] = recv_by_bill.get(bn, 0) + pay["amount"]

            # Also count cash bills as fully received
            recv_cust = {}
            for s in sales:
                k = s.get("party", "Unknown") or "Unknown"
                if k not in recv_cust:
                    recv_cust[k] = {"billed": 0, "received": 0, "pur": 0}

                s_items  = [x for x in si if x["sale_id"] == s["id"]]
                s_total  = s["grand_total"]
                s_taxbl  = sum(x.get("taxable", 0) for x in s_items)
                s_pur_c  = sum(x["qty"] * avg_pur_rate(x["product"]) for x in s_items)

                # received amount: payments + cash bills full amount
                if s.get("pay_mode", "Cash").lower() == "cash":
                    received = s_total
                else:
                    received = recv_by_bill.get(s["bill_no"], 0)

                recv_cust[k]["billed"]   += s_total
                recv_cust[k]["received"] += received
                recv_cust[k]["pur"]      += s_pur_c

            row_n = 1
            for k, v in sorted(recv_cust.items()):
                billed   = v["billed"]
                received = v["received"]
                pending  = billed - received
                pur_c    = v["pur"]
                # Profit only on the received portion (proportional)
                if billed > 0:
                    recv_ratio  = received / billed
                    taxable_rcv = received / 1.18  # approx ex-GST of received amount
                    pur_prop    = pur_c * recv_ratio
                    pf_recv     = taxable_rcv - pur_prop
                else:
                    pf_recv = 0

                table_row(tbl_recv,
                    [k, f'₹{billed:,.2f}', f'₹{received:,.2f}',
                     f'₹{pending:,.2f}', f'₹{pur_c * (received/billed if billed else 0):,.2f}',
                     f'₹{pf_recv:,.2f}'],
                    row_n,
                    fgs=[None, None, C_GREEN, C_RED if pending > 0 else None,
                         None, C_GREEN if pf_recv >= 0 else C_RED])
                row_n += 1

            # ════ TAB 5: EXPENSES ════
            cat_totals = {}
            for i, e in enumerate(expenses):
                bg = C_WHITE if i%2==0 else "#F7FAFC"
                table_row(tbl_exp,
                    [fmt_date(e["exp_date"]), e["category"],
                     (e["description"] or "")[:22],
                     f"₹{e['amount']:,.0f}",
                     e["pay_mode"] or "", e["ref_no"] or ""],
                    i+1, fgs=[None, None, None, C_RED, None, None])
                cat_totals[e["category"]] = cat_totals.get(e["category"], 0) + e["amount"]

            # Total row
            table_row(tbl_exp,
                ["TOTAL", "", f"{len(expenses)} entries",
                 f"₹{total_exp:,.0f}", "", ""],
                len(expenses)+1, fgs=[C_RED]*6)

            # Category breakdown below table
            exp_total_lbl.config(
                text="  ".join(f"{cat}: ₹{amt:,.0f}" for cat,amt in
                               sorted(cat_totals.items(), key=lambda x:-x[1]))
                + f"  |  TOTAL: ₹{total_exp:,.0f}"
            )
        # ── gen() ends here ──────────────────────────────────────────────────

        def export_pl_product():
            frm=pl_from.get().strip(); to=pl_to.get().strip()
            conn=get_db()
            sales=[dict(r) for r in conn.execute("SELECT * FROM sales").fetchall()]
            si_all=[dict(r) for r in conn.execute("SELECT * FROM sale_items").fetchall()]
            all_pi=[dict(r) for r in conn.execute("SELECT * FROM purchase_items").fetchall()]
            conn.close()
            if frm: sales=[s for s in sales if s["bill_date"]>=frm]
            if to:  sales=[s for s in sales if s["bill_date"]<=to]
            sids={s["id"] for s in sales}; si=[x for x in si_all if x["sale_id"] in sids]
            pcm={}
            for x in all_pi:
                k=x["product"]
                if k not in pcm: pcm[k]={"cost":0,"qty":0}
                pcm[k]["cost"]+=x.get("taxable",0) or max(0,x.get("total",0)-x.get("gst_amt",0))
                pcm[k]["qty"]+=x.get("qty",0)
            def ar(p): d=pcm.get(p); return d["cost"]/d["qty"] if d and d["qty"]>0 else 0
            pm={}
            for x in si:
                k=x["product"]
                if k not in pm: pm[k]={"qty":0,"sale":0,"gst":0,"pur":0}
                pm[k]["qty"]+=x["qty"]; pm[k]["sale"]+=x.get("taxable",0)
                pm[k]["gst"]+=x.get("gst_amt",0); pm[k]["pur"]+=x["qty"]*ar(k)
            data=[[k,f"{v['qty']:.2f}",f"{v['sale']:.2f}",f"{v['gst']:.2f}",
                   f"{v['pur']:.2f}",f"{v['sale']-v['pur']:.2f}"] for k,v in sorted(pm.items())]
            export_to_excel(["Product","Qty Sold","Sale Amt","GST","Purchase Cost","Gross Profit"],data,"PL_Product_Wise")

        def export_pl_bill():
            frm=pl_from.get().strip(); to=pl_to.get().strip()
            conn=get_db()
            sales=[dict(r) for r in conn.execute("SELECT * FROM sales").fetchall()]
            si_all=[dict(r) for r in conn.execute("SELECT * FROM sale_items").fetchall()]
            all_pi=[dict(r) for r in conn.execute("SELECT * FROM purchase_items").fetchall()]
            conn.close()
            if frm: sales=[s for s in sales if s["bill_date"]>=frm]
            if to:  sales=[s for s in sales if s["bill_date"]<=to]
            sids={s["id"] for s in sales}; si=[x for x in si_all if x["sale_id"] in sids]
            pcm={}
            for x in all_pi:
                k=x["product"]
                if k not in pcm: pcm[k]={"cost":0,"qty":0}
                pcm[k]["cost"]+=x.get("taxable",0) or max(0,x.get("total",0)-x.get("gst_amt",0))
                pcm[k]["qty"]+=x.get("qty",0)
            def ar(p): d=pcm.get(p); return d["cost"]/d["qty"] if d and d["qty"]>0 else 0
            data=[]
            for s in sales:
                s_items=[x for x in si if x["sale_id"]==s["id"]]
                s_sale=sum(x.get("taxable",0) for x in s_items); s_gst=sum(x.get("gst_amt",0) for x in s_items)
                s_pur=sum(x["qty"]*ar(x["product"]) for x in s_items)
                data.append([s["bill_no"],fmt_date(s["bill_date"]),s.get("party",""),
                              f"{s_sale:.2f}",f"{s_gst:.2f}",f"{s_pur:.2f}",f"{s_sale-s_pur:.2f}"])
            export_to_excel(["Bill No","Date","Customer","Sale Amt","GST","Purchase Cost","Profit/Loss"],data,"PL_Bill_Wise")

        ef2=tk.Frame(p,bg=C_BG); ef2.pack(fill="x",pady=(0,4))
        tk.Label(ef2,text="📥 Excel Download:",font=("Segoe UI",9,"bold"),bg=C_BG,fg=C_GRAY).pack(side="left",padx=(0,6))
        make_btn(ef2, "Product-wise", export_pl_product, bg=C_GREEN).pack(side="left",padx=4)
        make_btn(ef2, "Bill-wise", export_pl_bill, bg=C_GREEN).pack(side="left",padx=4)

        gen()

    # ══════════════════════════════════════════════════════════════════════════
    #  SALE REPORT
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_salereport(self):
        p=tk.Frame(self.content,bg=C_BG,padx=20,pady=14)
        p.pack(fill="x")
        section_title(p,"Sale Report — Product Wise")

        ff=tk.Frame(p,bg=C_BG); ff.pack(fill="x",pady=(0,10))
        tk.Label(ff,text="From:",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY).pack(side="left")
        sr_from=tk.StringVar()
        make_date_entry(ff,sr_from,width=12).pack(side="left",padx=4)
        tk.Label(ff,text="To:",font=("Segoe UI",9),bg=C_BG,fg=C_GRAY).pack(side="left")
        sr_to=tk.StringVar()
        make_date_entry(ff,sr_to,width=12).pack(side="left",padx=4)

        tbl=make_table(p,["Product","Qty Sold","Taxable Amt","GST Collected","Total Sale Amt"],
                       [26,10,14,16,16])

        def gen():
            clear_table_rows(tbl)
            frm=sr_from.get(); to=sr_to.get()
            conn=get_db()
            sales=[r[0] for r in conn.execute("SELECT id FROM sales WHERE 1=1"
                   +(" AND bill_date>=?" if frm else "")+(" AND bill_date<=?" if to else ""),
                   [x for x in [frm,to] if x]).fetchall()]
            if not sales: conn.close(); return
            items=[dict(r) for r in conn.execute(
                f"SELECT * FROM sale_items WHERE sale_id IN ({','.join('?'*len(sales))})",sales).fetchall()]
            conn.close()
            mp={}
            for it in items:
                k=it["product"]
                if k not in mp: mp[k]={"qty":0,"tax":0,"gst":0}
                mp[k]["qty"]+=it["qty"]; mp[k]["tax"]+=it.get("taxable",0); mp[k]["gst"]+=it["gst_amt"]
            for i,(k,v) in enumerate(sorted(mp.items())):
                table_row(tbl,[k,int(v["qty"]),f"₹{v['tax']:,.2f}",
                               f"₹{v['gst']:,.2f}",f"₹{v['tax']+v['gst']:,.2f}"],i+1)

        def export_sr():
            frm=sr_from.get(); to=sr_to.get()
            conn=get_db()
            sales=[r[0] for r in conn.execute("SELECT id FROM sales WHERE 1=1"
                   +(" AND bill_date>=?" if frm else "")+(" AND bill_date<=?" if to else ""),
                   [x for x in [frm,to] if x]).fetchall()]
            if not sales: conn.close(); messagebox.showinfo("Info","Koi data nahi!"); return
            items=[dict(r) for r in conn.execute(
                f"SELECT * FROM sale_items WHERE sale_id IN ({','.join('?'*len(sales))})",sales).fetchall()]
            conn.close()
            mp={}
            for it in items:
                k=it["product"]
                if k not in mp: mp[k]={"qty":0,"tax":0,"gst":0}
                mp[k]["qty"]+=it["qty"]; mp[k]["tax"]+=it.get("taxable",0); mp[k]["gst"]+=it["gst_amt"]
            data=[[k,int(v["qty"]),f"{v['tax']:.2f}",f"{v['gst']:.2f}",
                   f"{v['tax']+v['gst']:.2f}"] for k,v in sorted(mp.items())]
            export_to_excel(["Product","Qty Sold","Taxable Amt","GST Collected","Total Sale Amt"],
                            data,"Sale_Report")
        make_btn(ff,"📈  Generate",gen).pack(side="left",padx=8)
        make_btn(ff,"📥 Excel",export_sr,bg=C_GREEN).pack(side="left",padx=4)
        gen()

    # ══════════════════════════════════════════════════════════════════════════
    #  BALANCE SHEET
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_balancesheet(self):
        import datetime as _dt

        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "⚖️ Balance Sheet")

        # ── Date filter ─────────────────────────────────────────────────────
        ff = tk.Frame(p, bg=C_BG); ff.pack(fill="x", pady=(0,3))
        tk.Label(ff, text="As on Date:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_ason = tk.StringVar(value=str(_dt.date.today()))
        make_date_entry(ff, v_ason, width=13).pack(side="left", padx=6)
        tk.Label(ff, text="From:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left", padx=(10,0))
        v_from = tk.StringVar()
        make_date_entry(ff, v_from, width=12).pack(side="left", padx=4)
        tk.Label(ff, text="To:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_to = tk.StringVar(value=str(_dt.date.today()))
        make_date_entry(ff, v_to, width=12).pack(side="left", padx=4)
        make_btn(ff, "📊 Generate", lambda: gen(), bg="#1A365D").pack(side="left", padx=10)
        make_btn(ff, "📥 Excel",    lambda: export_bs(), bg=C_GREEN).pack(side="left", padx=4)

        # ── Main container: two columns ──────────────────────────────────────
        body = tk.Frame(p, bg=C_BG); body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)

        left_col  = tk.Frame(body, bg=C_BG); left_col.grid( row=0, column=0, sticky="nsew", padx=(0,6))
        right_col = tk.Frame(body, bg=C_BG); right_col.grid(row=0, column=1, sticky="nsew", padx=(6,0))

        # ── Helper: section box ──────────────────────────────────────────────
        def section_box(parent, title, color="#1A365D"):
            box = tk.Frame(parent, bg=C_WHITE,
                           highlightthickness=1, highlightbackground=C_BORDER)
            box.pack(fill="x", pady=(0,10))
            tk.Label(box, text=title, font=("Segoe UI",10,"bold"),
                     bg=color, fg="white", padx=12, pady=6).pack(fill="x")
            inner = tk.Frame(box, bg=C_WHITE)
            inner.pack(fill="x", padx=4, pady=4)
            return inner

        def bs_row(parent, label, value, bold=False, fg=C_GRAY, indent=0, bg=C_WHITE):
            row = tk.Frame(parent, bg=bg); row.pack(fill="x", pady=1)
            pad = indent * 16
            tk.Label(row, text=" "*indent + label,
                     font=("Segoe UI", 9, "bold" if bold else "normal"),
                     bg=bg, fg="#1A365D" if bold else fg,
                     anchor="w", padx=8+pad, pady=3).pack(side="left", fill="x", expand=True)
            tk.Label(row, text=str(value),
                     font=("Segoe UI", 9, "bold" if bold else "normal"),
                     bg=bg, fg=fg, anchor="e", padx=12).pack(side="right")

        def bs_divider(parent):
            tk.Frame(parent, bg=C_BORDER, height=1).pack(fill="x", padx=8, pady=2)

        def fmt(v): return f"₹{abs(round(v)):,}"

        # Store references to re-render
        self._bs_frames = {"left": left_col, "right": right_col}

        def gen():
            # Clear columns
            for w in left_col.winfo_children():  w.destroy()
            for w in right_col.winfo_children(): w.destroy()

            ason = v_ason.get().strip() or str(_dt.date.today())
            frm  = v_from.get().strip()
            to   = v_to.get().strip() or ason

            conn = get_db()

            # ── Raw data ────────────────────────────────────────────────────
            all_sales    = [dict(r) for r in conn.execute("SELECT * FROM sales").fetchall()]
            all_purs     = [dict(r) for r in conn.execute("SELECT * FROM purchases").fetchall()]
            all_si       = [dict(r) for r in conn.execute("SELECT * FROM sale_items").fetchall()]
            all_pi       = [dict(r) for r in conn.execute("SELECT * FROM purchase_items").fetchall()]
            all_prods    = [dict(r) for r in conn.execute("SELECT * FROM products").fetchall()]
            all_payments = [dict(r) for r in conn.execute("SELECT * FROM bill_payments").fetchall()]
            all_returns  = [dict(r) for r in conn.execute("SELECT * FROM returns").fetchall()]
            all_moves    = [dict(r) for r in conn.execute("SELECT * FROM stock_movements").fetchall()]
            # Expenses (date filtered)
            eq = "SELECT * FROM expenses WHERE 1=1"
            ep = []
            if frm: eq += " AND exp_date>=?"; ep.append(frm)
            if to:  eq += " AND exp_date<=?"; ep.append(to)
            all_expenses = [dict(r) for r in conn.execute(eq, ep).fetchall()]
            conn.close()

            # Date-filtered sales & purchases
            sales = [s for s in all_sales if (not frm or s["bill_date"]>=frm) and s["bill_date"]<=to]
            purs  = [p for p in all_purs  if (not frm or p["bill_date"]>=frm) and p["bill_date"]<=to]
            sids  = {s["id"] for s in sales}
            pids  = {p["id"] for p in purs}
            si    = [x for x in all_si if x["sale_id"]    in sids]
            pi    = [x for x in all_pi if x["purchase_id"] in pids]

            # ── INCOME SIDE ──────────────────────────────────────────────────
            total_sale_gross = sum(s["grand_total"] for s in sales)
            sale_gst         = sum(x.get("gst_amt",0) for x in si)
            total_sale_net   = total_sale_gross - sale_gst   # ex-GST sale

            sale_returns     = sum(r["total_amt"] for r in all_returns
                                   if r["return_type"]=="sale_return"
                                   and (not frm or r["return_date"]>=frm)
                                   and r["return_date"]<=to)
            net_revenue      = total_sale_net - sale_returns

            # Payments received
            recv_sale  = sum(pay["amount"] for pay in all_payments
                             if pay["bill_type"]=="sale"
                             and any(s["bill_no"]==pay["bill_no"] for s in sales))

            # ── EXPENSE SIDE ─────────────────────────────────────────────────
            total_pur_gross  = sum(p["grand_total"] for p in purs)
            pur_gst          = sum(x.get("gst_amt",0) for x in pi)
            total_pur_net    = total_pur_gross - pur_gst   # ex-GST purchase

            pur_returns      = sum(r["total_amt"] for r in all_returns
                                   if r["return_type"]=="pur_return"
                                   and (not frm or r["return_date"]>=frm)
                                   and r["return_date"]<=to)
            net_purchase     = total_pur_net - pur_returns

            # ── RECEIVABLES / PAYABLES ───────────────────────────────────────
            pay_map = {}
            for pay in all_payments:
                k = (pay["bill_type"], pay["bill_no"])
                pay_map[k] = pay_map.get(k, 0) + pay["amount"]

            total_receivable = 0
            for s in all_sales:   # ALL sales, not date-filtered
                paid = pay_map.get(("sale", s["bill_no"]), 0)
                baki = max(0, s["grand_total"] - paid)
                total_receivable += baki

            total_payable = 0
            for pur in all_purs:
                paid = pay_map.get(("pur", pur["bill_no"]), 0)
                baki = max(0, pur["grand_total"] - paid)
                total_payable += baki

            # ── STOCK VALUE ──────────────────────────────────────────────────
            stock_value = sum(get_stock(pr["name"]) * pr["sale_rate"] for pr in all_prods)
            stock_cost  = 0
            pur_cost_map = {}
            for x in all_pi:
                k = x["product"]
                if k not in pur_cost_map: pur_cost_map[k] = {"cost":0,"qty":0}
                ex = x.get("taxable",0) or max(0, x.get("total",0)-x.get("gst_amt",0))
                pur_cost_map[k]["cost"] += ex
                pur_cost_map[k]["qty"]  += x.get("qty",0)
            for pr in all_prods:
                stk = get_stock(pr["name"])
                d   = pur_cost_map.get(pr["name"])
                avg = (d["cost"]/d["qty"]) if d and d["qty"]>0 else 0
                stock_cost += stk * avg

            # ── GST ──────────────────────────────────────────────────────────
            gst_collected = sale_gst
            gst_paid      = pur_gst
            gst_liability = gst_collected - gst_paid   # amount govt ko dena hai

            # ── PROFIT (COGS-based — Dashboard ke jaisa) ───────────────────────
            total_expenses = sum(e["amount"] for e in all_expenses)

            # Average purchase rate per product (all-time, ex-GST)
            pur_cost_map = {}
            for x in all_pi:
                k = x["product"]
                if k not in pur_cost_map: pur_cost_map[k] = {"cost":0,"qty":0}
                ex = x.get("taxable",0) or max(0, x.get("total",0)-x.get("gst_amt",0))
                pur_cost_map[k]["cost"] += ex
                pur_cost_map[k]["qty"]  += x.get("qty",0)

            # COGS for items sold in this period
            cogs = 0.0
            for x in si:
                d = pur_cost_map.get(x["product"])
                if d and d["qty"] > 0:
                    avg_rate = d["cost"] / d["qty"]
                    cogs += avg_rate * x.get("qty", 0)

            # Returned items ka COGS wapas minus karo (period ke sale returns)
            sale_return_rows = [r for r in all_returns
                                 if r["return_type"]=="sale_return"
                                 and (not frm or r["return_date"]>=frm)
                                 and r["return_date"]<=to]
            cogs_returned = 0.0
            for r in sale_return_rows:
                d = pur_cost_map.get(r["product"])
                if d and d["qty"] > 0:
                    avg_rate = d["cost"] / d["qty"]
                    cogs_returned += avg_rate * r["qty"]
            cogs_net = cogs - cogs_returned

            gross_profit  = net_revenue - cogs_net
            net_profit    = gross_profit - total_expenses

            # ══════════════════════════════════════════════
            # LEFT COLUMN — ASSETS
            # ══════════════════════════════════════════════
            tk.Label(left_col, text="ASSETS  (Hamare Paas Kya Hai)",
                     font=("Segoe UI",11,"bold"), bg=C_BG, fg="#1A365D").pack(anchor="w", pady=(0,4))

            # Current Assets
            ca = section_box(left_col, "📦 Current Assets", "#276749")
            bs_row(ca, "Stock Value (Sale Rate pe)",  fmt(stock_value),  indent=1)
            bs_row(ca, "Stock Cost (Purchase Rate pe)", fmt(stock_cost), indent=1, fg="#718096")
            bs_divider(ca)
            bs_row(ca, "Debtors (Customers se Lena)", fmt(total_receivable), indent=1, fg=C_GREEN)
            bs_row(ca, "Cash Received (Sales)",        fmt(recv_sale),   indent=1, fg="#276749")
            bs_divider(ca)
            total_ca = stock_value + total_receivable
            bs_row(ca, "Total Current Assets", fmt(total_ca), bold=True, fg="#276749")

            # Revenue Summary
            rev = section_box(left_col, "💰 Revenue Summary", "#2B6CB0")
            bs_row(rev, "Gross Sale",             fmt(total_sale_gross), indent=1)
            bs_row(rev, "  Less: GST Collected", f"- {fmt(sale_gst)}",  indent=1, fg=C_AMBER)
            bs_row(rev, "  Less: Sale Returns",  f"- {fmt(sale_returns)}", indent=1, fg=C_RED)
            bs_divider(rev)
            bs_row(rev, "Net Revenue (ex-GST)", fmt(net_revenue), bold=True, fg="#2B6CB0")

            # Net Profit box
            profit_col = C_GREEN if net_profit >= 0 else C_RED
            prof = section_box(left_col, "📈 Profit / Loss", "#1A365D")
            bs_row(prof, "Net Revenue",      fmt(net_revenue),  indent=1, fg=C_GREEN)
            bs_row(prof, "Cost of Goods Sold (COGS)", fmt(cogs_net), indent=1, fg=C_RED)
            bs_divider(prof)
            bs_row(prof, "Gross Profit", fmt(gross_profit), indent=1)
            bs_row(prof, "  Less: Expenses", f"- {fmt(total_expenses)}", indent=1, fg=C_AMBER)
            bs_divider(prof)
            bs_row(prof,
                   "Net Profit" if net_profit>=0 else "Net Loss",
                   fmt(net_profit), bold=True, fg=profit_col)
            bs_row(prof, "Gross Margin",
                   f"{(net_profit/net_revenue*100):.1f}%" if net_revenue>0 else "N/A",
                   indent=1, fg="#718096")

            # ══════════════════════════════════════════════
            # RIGHT COLUMN — LIABILITIES & EQUITY
            # ══════════════════════════════════════════════
            tk.Label(right_col, text="LIABILITIES  (Humara Dena Kitna Hai)",
                     font=("Segoe UI",11,"bold"), bg=C_BG, fg="#742A2A").pack(anchor="w", pady=(0,4))

            # Current Liabilities
            cl = section_box(right_col, "📋 Current Liabilities", "#C05621")
            bs_row(cl, "Creditors (Suppliers ko Dena)", fmt(total_payable), indent=1, fg=C_RED)
            bs_row(cl, "GST Payable (to Govt)",
                   fmt(gst_liability) if gst_liability>=0 else f"GST Credit: {fmt(abs(gst_liability))}",
                   indent=1, fg=C_AMBER)
            bs_divider(cl)
            total_cl = total_payable + max(0, gst_liability)
            bs_row(cl, "Total Current Liabilities", fmt(total_cl), bold=True, fg=C_RED)

            # Purchase Summary
            pur_s = section_box(right_col, "🛒 Purchase Summary", "#6B46C1")
            bs_row(pur_s, "Gross Purchase",          fmt(total_pur_gross), indent=1)
            bs_row(pur_s, "  Less: GST Paid",        f"- {fmt(pur_gst)}",  indent=1, fg=C_AMBER)
            bs_row(pur_s, "  Less: Purchase Returns",f"- {fmt(pur_returns)}", indent=1, fg=C_GREEN)
            bs_divider(pur_s)
            bs_row(pur_s, "Net Purchase (ex-GST)", fmt(net_purchase), bold=True, fg="#6B46C1")

            # GST Summary
            gst_s = section_box(right_col, "🏛️ GST Summary", "#744210")
            bs_row(gst_s, "GST Collected (Output)",  fmt(gst_collected), indent=1, fg=C_RED)
            bs_row(gst_s, "GST Paid (Input Credit)", fmt(gst_paid),      indent=1, fg=C_GREEN)
            bs_divider(gst_s)
            net_gst = gst_collected - gst_paid
            bs_row(gst_s,
                   "GST Payable" if net_gst>=0 else "GST Refund Due",
                   fmt(net_gst), bold=True,
                   fg=C_RED if net_gst>=0 else C_GREEN)

            # Balance check
            net_worth = total_ca - total_cl + net_profit
            bal = section_box(right_col, "⚖️ Net Position", "#1A365D")
            bs_row(bal, "Total Assets",      fmt(total_ca),   indent=1, fg=C_GREEN)
            bs_row(bal, "Total Liabilities", fmt(total_cl),   indent=1, fg=C_RED)
            bs_divider(bal)
            bs_row(bal, "Net Worth / Equity", fmt(net_worth), bold=True,
                   fg=C_GREEN if net_worth>=0 else C_RED,
                   bg="#F0FFF4" if net_worth>=0 else "#FFF5F5")

            # Store for export
            self._bs_data = {
                "ason": ason, "frm": frm, "to": to,
                "total_sale_gross": total_sale_gross, "sale_gst": sale_gst,
                "sale_returns": sale_returns, "net_revenue": net_revenue,
                "recv_sale": recv_sale,
                "total_pur_gross": total_pur_gross, "pur_gst": pur_gst,
                "pur_returns": pur_returns, "net_purchase": net_purchase,
                "stock_value": stock_value, "stock_cost": stock_cost,
                "total_receivable": total_receivable, "total_payable": total_payable,
                "gst_collected": gst_collected, "gst_paid": gst_paid,
                "gst_liability": gst_liability, "gross_profit": gross_profit,
                "net_profit": net_profit, "net_worth": net_worth,
                "total_ca": total_ca, "total_cl": total_cl,
                "cogs_net": cogs_net, "total_expenses": total_expenses,
            }

        def export_bs():
            if not hasattr(self, "_bs_data"):
                messagebox.showinfo("Generate Karo", "Pehle Generate button dabao!"); return
            d = self._bs_data
            rows = [
                ["BALANCE SHEET — BHUGTANEASE", "", ""],
                [f"Period: {d['frm'] or 'All'} to {d['to']}", "", ""],
                ["", "", ""],
                ["ASSETS", "", "LIABILITIES"],
                ["Current Assets", "", "Current Liabilities"],
                [f"  Stock Value (Sale Rate)", fmt(d["stock_value"]),
                 f"  Creditors (Suppliers ko Dena)  {fmt(d['total_payable'])}"],
                [f"  Stock Cost (Pur Rate)", fmt(d["stock_cost"]),
                 f"  GST Payable  {fmt(max(0,d['gst_liability']))}"],
                [f"  Debtors (Customers se Lena)", fmt(d["total_receivable"]),
                 f"Total Current Liabilities  {fmt(d['total_cl'])}"],
                [f"Total Current Assets", fmt(d["total_ca"]), ""],
                ["", "", ""],
                ["Revenue Summary", "", "Purchase Summary"],
                [f"  Gross Sale", fmt(d["total_sale_gross"]),
                 f"  Gross Purchase  {fmt(d['total_pur_gross'])}"],
                [f"  Less: GST Collected", fmt(d["sale_gst"]),
                 f"  Less: GST Paid  {fmt(d['pur_gst'])}"],
                [f"  Less: Sale Returns", fmt(d["sale_returns"]),
                 f"  Less: Pur Returns  {fmt(d['pur_returns'])}"],
                [f"Net Revenue", fmt(d["net_revenue"]),
                 f"Net Purchase  {fmt(d['net_purchase'])}"],
                ["", "", ""],
                ["Profit / Loss", "", "GST Summary"],
                [f"  Net Revenue", fmt(d["net_revenue"]),
                 f"  GST Collected  {fmt(d['gst_collected'])}"],
                [f"  Cost of Goods Sold (COGS)", fmt(d["cogs_net"]),
                 f"  GST Paid  {fmt(d['gst_paid'])}"],
                [f"  Gross Profit", fmt(d["gross_profit"]), ""],
                [f"  Less: Expenses", fmt(d["total_expenses"]), ""],
                [f"Net Profit" if d["net_profit"]>=0 else "Net Loss",
                 fmt(d["net_profit"]),
                 f"Net GST Payable  {fmt(d['gst_liability'])}"],
                ["", "", ""],
                ["NET WORTH / EQUITY", fmt(d["net_worth"]), ""],
            ]
            try:
                export_to_excel(
                    ["Description", "Amount (Assets Side)", "Amount (Liabilities Side)"],
                    rows, "Balance_Sheet")
            except Exception as e:
                messagebox.showerror("Export Error", str(e))

        gen()


        gen()


# ══════════════════════════════════════════════════════════════════════════════
    # ══════════════════════════════════════════════════════════════════════════
    #  FIFO STOCK REPORT
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_fiforeport(self):
        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="both", expand=True)
        section_title(p, "📦 FIFO Stock Report")

        # Info bar
        info_f = tk.Frame(p, bg="#EBF8FF", highlightthickness=1, highlightbackground="#90CDF4")
        info_f.pack(fill="x", pady=(0,10))
        tk.Label(info_f, text="FIFO (First In First Out): Jo stock pehle aaya, woh pehle bikta hai. "
                 "Har purchase ki alag layer track hoti hai — sahi profit calculation ke liye.",
                 font=("Segoe UI",9), bg="#EBF8FF", fg="#2B6CB0", pady=6, padx=12).pack(anchor="w")

        conn = get_db()

        # ── Summary cards ─────────────────────────────────────────────────────
        total_fifo_value = fifo_get_stock_value(conn)
        total_cogs = conn.execute(
            "SELECT COALESCE(SUM(total_cost),0) FROM fifo_consumption"
        ).fetchone()[0] or 0
        total_sale_rev = conn.execute(
            "SELECT COALESCE(SUM(grand_total),0) FROM sales"
        ).fetchone()[0] or 0
        fifo_profit = round(total_sale_rev - total_cogs, 2)

        cards_f = tk.Frame(p, bg=C_BG); cards_f.pack(fill="x", pady=(0,10))
        for i,(lbl,val,clr) in enumerate([
            ("Remaining Stock Value\n(FIFO Cost)", f"₹{round(total_fifo_value):,}", "#276749"),
            ("Total COGS\n(Sold Stock Cost)", f"₹{round(total_cogs):,}", "#C53030"),
            ("Total Revenue\n(Sale Bills)", f"₹{round(total_sale_rev):,}", "#2B6CB0"),
            ("Gross Profit\n(Revenue - COGS)", f"₹{round(fifo_profit):,}", "#744210"),
        ]):
            cf = tk.Frame(cards_f, bg=C_WHITE, highlightthickness=1, highlightbackground=clr,
                          padx=14, pady=4); cf.grid(row=0, column=i, sticky="nsew", padx=5)
            cards_f.columnconfigure(i, weight=1)
            tk.Label(cf, text=lbl, font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY, justify="center").pack()
            tk.Label(cf, text=val, font=("Segoe UI",14,"bold"), bg=C_WHITE, fg=clr).pack()

        # ── Tabs: Layers | COGS Log ───────────────────────────────────────────
        tab_f = tk.Frame(p, bg=C_BG); tab_f.pack(fill="x", pady=(0,3))
        v_tab = tk.StringVar(value="layers")
        content_area = tk.Frame(p, bg=C_BG); content_area.pack(fill="both", expand=True)

        def show_tab(tab):
            v_tab.set(tab)
            for w in content_area.winfo_children(): w.destroy()
            if tab == "layers":
                btn_layers.config(bg=C_ACCENT, fg=C_WHITE)
                btn_cogs.config(bg=C_WHITE, fg=C_GRAY)
                show_layers()
            else:
                btn_layers.config(bg=C_WHITE, fg=C_GRAY)
                btn_cogs.config(bg=C_ACCENT, fg=C_WHITE)
                show_cogs()

        btn_layers = tk.Button(tab_f, text="📦 Stock Layers (Remaining)",
                               font=("Segoe UI",10,"bold"), bg=C_ACCENT, fg=C_WHITE,
                               relief="flat", cursor="hand2", padx=14, pady=6,
                               command=lambda: show_tab("layers"))
        btn_layers.pack(side="left", padx=(0,4))
        btn_cogs = tk.Button(tab_f, text="💰 COGS Log (Sold Stock Cost)",
                             font=("Segoe UI",10,"bold"), bg=C_WHITE, fg=C_GRAY,
                             relief="flat", cursor="hand2", padx=14, pady=6,
                             command=lambda: show_tab("cogs"))
        btn_cogs.pack(side="left")

        def show_layers():
            rows = conn.execute(
                "SELECT product, COUNT(*) as layers, SUM(qty_remaining) as total_qty, "
                "SUM(qty_remaining*cost_per_unit) as stock_value "
                "FROM fifo_layers WHERE qty_remaining>0 GROUP BY product ORDER BY product"
            ).fetchall()
            if not rows:
                tk.Label(content_area, text="Koi FIFO layer nahi mili.\nPurchase karo toh layers automatically banangi.",
                         font=("Segoe UI",11), bg=C_BG, fg=C_GRAY).pack(pady=40)
                return
            tbl = make_table(content_area,
                ["Product","Layers","Total Qty","Stock Value (FIFO Cost)","Avg Cost/Unit",""],
                [20,6,8,18,14,8])
            tbl.master.master.pack(fill="x")
            for i, r in enumerate(rows):
                bg = C_WHITE if i%2==0 else "#F7FAFC"
                avg = round(r["stock_value"]/r["total_qty"],2) if r["total_qty"] else 0
                vals = [r["product"], r["layers"], round(r["total_qty"],2),
                        f"₹{round(r['stock_value']):,}", f"₹{avg}", "🔍"]
                for j, val in enumerate(vals):
                    if j == 5:
                        tk.Button(tbl, text="🔍", font=("Segoe UI",9),
                                  bg=bg, fg=C_ACCENT, relief="flat", cursor="hand2", bd=0,
                                  command=lambda pr=r["product"]: show_product_layers(pr)
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=1)
                    else:
                        tk.Label(tbl, text=str(val), font=("Segoe UI",9),
                                 bg=bg, fg=C_GRAY, anchor="w", padx=6, pady=4
                                 ).grid(row=i+1,column=j,sticky="nsew",padx=1)

        def show_product_layers(product):
            win = tk.Toplevel(self.root)
            win.title(f"FIFO Layers — {product}")
            win.geometry("700x420"); win.configure(bg=C_WHITE)
            tk.Label(win, text=f"📦 FIFO Layers: {product}",
                     font=("Segoe UI",12,"bold"), bg=C_WHITE, fg="#1A365D").pack(anchor="w", padx=16, pady=(12,6))
            c2 = get_db()
            layers = c2.execute(
                "SELECT purchase_date, purchase_bill, batch_no, qty_in, qty_remaining, cost_per_unit "
                "FROM fifo_layers WHERE product=? ORDER BY purchase_date ASC, id ASC",
                (product,)
            ).fetchall()
            c2.close()
            tbl2 = make_table(win,
                ["Purchase Date","Bill No","Batch","Qty In","Remaining","Cost/Unit","Layer Value"],
                [12,10,8,8,9,10,11])
            tbl2.master.master.pack(fill="x", padx=16)
            for i, l in enumerate(layers):
                bg = C_WHITE if i%2==0 else "#EBF8FF"
                rem_clr = "#C53030" if l["qty_remaining"]==0 else "#276749"
                for j, val in enumerate([
                    l["purchase_date"], l["purchase_bill"], l["batch_no"] or "-",
                    round(l["qty_in"],2), round(l["qty_remaining"],2),
                    f"₹{l['cost_per_unit']}", f"₹{round(l['qty_remaining']*l['cost_per_unit'],2)}"
                ]):
                    fg = rem_clr if j==4 else C_GRAY
                    tk.Label(tbl2, text=str(val), font=("Segoe UI",9),
                             bg=bg, fg=fg, anchor="w", padx=5, pady=4
                             ).grid(row=i+1,column=j,sticky="nsew",padx=1)
            make_btn(win, "✕ Close", win.destroy, bg=C_GRAY).pack(pady=4)

        def show_cogs():
            rows = conn.execute(
                "SELECT fc.sale_date, fc.sale_bill, fc.product, "
                "SUM(fc.qty_used) as qty, SUM(fc.total_cost) as cogs, "
                "COALESCE(s.grand_total,0) as revenue "
                "FROM fifo_consumption fc "
                "LEFT JOIN sales s ON s.bill_no=fc.sale_bill "
                "GROUP BY fc.sale_bill, fc.product ORDER BY fc.sale_date DESC LIMIT 200"
            ).fetchall()
            if not rows:
                tk.Label(content_area, text="Koi COGS record nahi.\nSale bill save karo toh COGS automatically track hoga.",
                         font=("Segoe UI",11), bg=C_BG, fg=C_GRAY).pack(pady=40)
                return
            tbl = make_table(content_area,
                ["Sale Date","Bill No","Product","Qty Sold","COGS (Cost)","Sale Revenue","Gross Profit"],
                [10,10,18,8,12,13,12])
            tbl.master.master.pack(fill="x")
            for i, r in enumerate(rows):
                bg = C_WHITE if i%2==0 else "#F7FAFC"
                gp = round(r["revenue"] - r["cogs"], 2)
                gp_clr = "#276749" if gp >= 0 else "#C53030"
                for j, val in enumerate([
                    r["sale_date"], r["sale_bill"], r["product"],
                    round(r["qty"],2), f"₹{round(r['cogs'],2):,}",
                    f"₹{round(r['revenue'],2):,}", f"₹{round(gp,2):,}"
                ]):
                    fg = gp_clr if j==6 else C_GRAY
                    tk.Label(tbl, text=str(val), font=("Segoe UI",9),
                             bg=bg, fg=fg, anchor="w", padx=5, pady=4
                             ).grid(row=i+1,column=j,sticky="nsew",padx=1)

        show_tab("layers")
        # conn is used by show_layers/show_cogs via closure — don't close early

    # ══════════════════════════════════════════════════════════════════════════════
    # ══════════════════════════════════════════════════════════════════════════
    #  GSTR-1  (Outward Supplies)
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_gstr1(self):
        import datetime as _dt
        from collections import defaultdict

        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "📋 GSTR-1 — Outward Supplies (Sales)")

        # ── Info bar ────────────────────────────────────────────────────────
        info = tk.Label(p,
            text="GSTR-1 = Aapke sabhi outward (sale) invoices ki summary. B2B, B2C, HSN-wise aur CDN (Credit/Debit Notes) sections.",
            font=("Segoe UI",9), bg="#EBF4FF", fg="#1A365D", anchor="w", padx=12, pady=6,
            wraplength=900, justify="left")
        info.pack(fill="x", pady=(0,3))

        # ── Filter ──────────────────────────────────────────────────────────
        ff = tk.Frame(p, bg=C_BG); ff.pack(fill="x", pady=(0,3))
        tk.Label(ff, text="Month:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        months = []
        today = _dt.date.today()
        for i in range(12):
            d = today.replace(day=1) - _dt.timedelta(days=i*28)
            d = d.replace(day=1)
            months.append(d.strftime("%Y-%m"))
        months = sorted(set(months), reverse=True)
        v_month = tk.StringVar(value=today.strftime("%Y-%m"))
        ttk.Combobox(ff, textvariable=v_month, values=months, width=10,
                     state="readonly").pack(side="left", padx=6)

        tk.Label(ff, text="OR Custom From:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left", padx=(10,0))
        v_from = tk.StringVar()
        make_date_entry(ff, v_from, width=12).pack(side="left", padx=4)
        tk.Label(ff, text="To:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_to = tk.StringVar()
        make_date_entry(ff, v_to, width=12).pack(side="left", padx=4)

        make_btn(ff, "📊 Generate", lambda: gen(), bg="#1A365D").pack(side="left", padx=10)
        make_btn(ff, "📥 Excel", lambda: export_gstr1(), bg=C_GREEN).pack(side="left", padx=4)

        # ── Notebook tabs ────────────────────────────────────────────────────
        nb = ttk.Notebook(p); nb.pack(fill="both", expand=True, pady=(8,0))

        tab_b2b  = tk.Frame(nb, bg=C_BG); nb.add(tab_b2b,  text="  🏢 B2B Invoices  ")
        tab_b2c  = tk.Frame(nb, bg=C_BG); nb.add(tab_b2c,  text="  👤 B2C Invoices  ")
        tab_hsn  = tk.Frame(nb, bg=C_BG); nb.add(tab_hsn,  text="  📦 HSN Summary  ")
        tab_cdn  = tk.Frame(nb, bg=C_BG); nb.add(tab_cdn,  text="  🔄 Credit/Debit Notes  ")
        tab_sum  = tk.Frame(nb, bg=C_BG); nb.add(tab_sum,  text="  📊 Tax Summary  ")

        # B2B table
        tbl_b2b = make_table(tab_b2b,
            ["#","Bill No","Date","Party","GSTIN","Taxable","CGST","SGST","IGST","Total","Type"],
            [4, 14, 11, 18, 18, 11, 9, 9, 9, 11, 9])

        # B2C table
        tbl_b2c = make_table(tab_b2c,
            ["#","Bill No","Date","Party","Taxable","CGST","SGST","IGST","Total GST","Grand"],
            [4, 14, 11, 20, 12, 9, 9, 9, 10, 11])

        # HSN table
        tbl_hsn = make_table(tab_hsn,
            ["HSN","Description","UOM","Total Qty","Taxable Amt","IGST Rate","IGST","CGST Rate","CGST","SGST Rate","SGST","Total Tax"],
            [8, 16, 6, 9, 12, 8, 10, 8, 10, 8, 10, 10])

        # CDN table
        tbl_cdn = make_table(tab_cdn,
            ["#","Return No","Date","Party","GSTIN","Type","Taxable","GST","Total"],
            [4, 14, 11, 18, 18, 10, 12, 10, 11])

        # Summary table
        tbl_sum = make_table(tab_sum,
            ["Nature","Taxable Amt","IGST","CGST","SGST","Total GST","Grand Total"],
            [20, 14, 12, 12, 12, 12, 14])

        # Summary cards
        sum_cards = tk.Frame(tab_sum, bg=C_BG); sum_cards.pack(fill="x", pady=(8,0))

        gstr1_data = {}  # store for export

        def get_date_range():
            frm = v_from.get().strip()
            to  = v_to.get().strip()
            if not frm and not to:
                m = v_month.get()
                y, mo = int(m[:4]), int(m[5:])
                import calendar
                last = calendar.monthrange(y, mo)[1]
                frm = f"{m}-01"
                to  = f"{m}-{last:02d}"
            return frm, to

        def gen():
            for tbl in [tbl_b2b, tbl_b2c, tbl_hsn, tbl_cdn, tbl_sum]:
                clear_table_rows(tbl)
            for w in sum_cards.winfo_children(): w.destroy()

            frm, to = get_date_range()

            conn = get_db()
            sales = [dict(r) for r in conn.execute(
                "SELECT s.*, si.product, si.hsn, si.qty, si.rate, si.taxable, "
                "si.gst_percent, si.gst_amt, si.grand as item_grand "
                "FROM sales s JOIN sale_items si ON s.id=si.sale_id "
                "WHERE s.bill_date>=? AND s.bill_date<=? ORDER BY s.bill_date, s.bill_no",
                (frm, to)).fetchall()]
            returns = [dict(r) for r in conn.execute(
                "SELECT * FROM returns WHERE return_type='sale_return' "
                "AND return_date>=? AND return_date<=?", (frm, to)).fetchall()]
            conn.close()

            # Group by bill
            bills = defaultdict(lambda: {
                "date":"","party":"","gstin":"","gst_type":"CGST+SGST",
                "taxable":0,"cgst":0,"sgst":0,"igst":0,"total_gst":0,"grand":0,
                "items":[]
            })
            for r in sales:
                bn = r["bill_no"]
                bills[bn]["date"]     = r["bill_date"]
                bills[bn]["party"]    = r["party"]
                bills[bn]["gstin"]    = r["party_gstin"] or ""
                bills[bn]["gst_type"] = r.get("gst_type","CGST+SGST") or "CGST+SGST"
                bills[bn]["grand"]    = r["grand_total"]
                bills[bn]["taxable"] += round(r["taxable"],2)
                ga = round(r["gst_amt"],2)
                if bills[bn]["gst_type"] == "IGST":
                    bills[bn]["igst"] += ga
                else:
                    bills[bn]["cgst"] += round(ga/2,2)
                    bills[bn]["sgst"] += round(ga/2,2)
                bills[bn]["total_gst"] += ga
                bills[bn]["items"].append(r)

            b2b = {k:v for k,v in bills.items() if v["gstin"]}
            b2c = {k:v for k,v in bills.items() if not v["gstin"]}

            # ── B2B ──
            b2b_tax=0; b2b_cgst=0; b2b_sgst=0; b2b_igst=0; b2b_grand=0
            for i,(bn,b) in enumerate(sorted(b2b.items(), key=lambda x:x[1]["date"])):
                table_row(tbl_b2b,
                    [i+1, bn, fmt_date(b["date"]), b["party"][:16], b["gstin"],
                     f"₹{b['taxable']:.2f}",
                     f"₹{b['cgst']:.2f}" if b["cgst"] else "-",
                     f"₹{b['sgst']:.2f}" if b["sgst"] else "-",
                     f"₹{b['igst']:.2f}" if b["igst"] else "-",
                     f"₹{b['grand']:.0f}", b["gst_type"]], i+1,
                    fgs=[None,C_ACCENT,None,None,"#6B46C1",None,C_AMBER,C_AMBER,"#6B46C1",C_GREEN,None])
                b2b_tax+=b["taxable"]; b2b_cgst+=b["cgst"]; b2b_sgst+=b["sgst"]
                b2b_igst+=b["igst"]; b2b_grand+=b["grand"]
            table_row(tbl_b2b,
                ["TOTAL","","","","",f"₹{b2b_tax:.2f}",
                 f"₹{b2b_cgst:.2f}",f"₹{b2b_sgst:.2f}",f"₹{b2b_igst:.2f}",f"₹{b2b_grand:.0f}",""],
                len(b2b)+1, fgs=[C_RED]*11)

            # ── B2C ──
            b2c_tax=0; b2c_cgst=0; b2c_sgst=0; b2c_igst=0; b2c_grand=0
            for i,(bn,b) in enumerate(sorted(b2c.items(), key=lambda x:x[1]["date"])):
                table_row(tbl_b2c,
                    [i+1, bn, fmt_date(b["date"]), b["party"][:18],
                     f"₹{b['taxable']:.2f}",
                     f"₹{b['cgst']:.2f}" if b["cgst"] else "-",
                     f"₹{b['sgst']:.2f}" if b["sgst"] else "-",
                     f"₹{b['igst']:.2f}" if b["igst"] else "-",
                     f"₹{b['total_gst']:.2f}", f"₹{b['grand']:.0f}"], i+1,
                    fgs=[None,C_ACCENT,None,None,None,C_AMBER,C_AMBER,"#6B46C1",C_RED,C_GREEN])
                b2c_tax+=b["taxable"]; b2c_cgst+=b["cgst"]; b2c_sgst+=b["sgst"]
                b2c_igst+=b["igst"]; b2c_grand+=b["grand"]
            table_row(tbl_b2c,
                ["TOTAL","","",f"",f"₹{b2c_tax:.2f}",
                 f"₹{b2c_cgst:.2f}",f"₹{b2c_sgst:.2f}",f"₹{b2c_igst:.2f}",
                 f"₹{b2c_cgst+b2c_sgst+b2c_igst:.2f}",f"₹{b2c_grand:.0f}"],
                len(b2c)+1, fgs=[C_RED]*10)

            # ── HSN Summary ──
            hsn_map = defaultdict(lambda:{"desc":"","qty":0,"taxable":0,"igst_rate":0,
                                           "igst":0,"cgst_rate":0,"cgst":0,"sgst_rate":0,"sgst":0})
            for r in sales:
                h = r["hsn"] or "NA"
                gp = r["gst_percent"]
                ga = round(r["gst_amt"],2)
                gt = r.get("gst_type","CGST+SGST") or "CGST+SGST"
                hsn_map[h]["desc"]    = r["product"]
                hsn_map[h]["qty"]    += r["qty"]
                hsn_map[h]["taxable"]+= round(r["taxable"],2)
                if gt == "IGST":
                    hsn_map[h]["igst_rate"] = gp
                    hsn_map[h]["igst"]     += ga
                else:
                    hsn_map[h]["cgst_rate"] = gp/2
                    hsn_map[h]["sgst_rate"] = gp/2
                    hsn_map[h]["cgst"]     += round(ga/2,2)
                    hsn_map[h]["sgst"]     += round(ga/2,2)

            for i,(h,v) in enumerate(sorted(hsn_map.items())):
                total_tax = v["igst"]+v["cgst"]+v["sgst"]
                table_row(tbl_hsn,
                    [h, v["desc"][:14], "NOS", f"{v['qty']:.0f}",
                     f"₹{v['taxable']:.2f}",
                     f"{v['igst_rate']}%" if v["igst"] else "-",
                     f"₹{v['igst']:.2f}"  if v["igst"] else "-",
                     f"{v['cgst_rate']}%" if v["cgst"] else "-",
                     f"₹{v['cgst']:.2f}"  if v["cgst"] else "-",
                     f"{v['sgst_rate']}%" if v["sgst"] else "-",
                     f"₹{v['sgst']:.2f}"  if v["sgst"] else "-",
                     f"₹{total_tax:.2f}"], i+1)

            # ── CDN (Credit Notes = Sale Returns) ──
            for i,r in enumerate(returns):
                conn2 = get_db()
                orig = conn2.execute("SELECT party_gstin FROM sales WHERE bill_no=?",
                                     (r["orig_bill"],)).fetchone()
                conn2.close()
                gstin = orig["party_gstin"] if orig else ""
                table_row(tbl_cdn,
                    [i+1, r["return_no"], fmt_date(r["return_date"]),
                     r["party"][:16], gstin,
                     "Credit Note", f"₹{r.get('taxable',0) or r['total_amt']:.2f}",
                     f"₹{r['gst_amt']:.2f}", f"₹{r['total_amt']:.2f}"], i+1,
                    fgs=[None,C_RED,None,None,"#6B46C1","#6B46C1",None,C_AMBER,C_RED])

            # ── Tax Summary ──
            tot_b2b_tax = b2b_tax; tot_b2c_tax = b2c_tax
            tot_cgst = b2b_cgst+b2c_cgst; tot_sgst = b2b_sgst+b2c_sgst
            tot_igst = b2b_igst+b2c_igst
            ret_tax  = sum(r.get("taxable",0) or r["total_amt"] for r in returns)
            ret_gst  = sum(r["gst_amt"] for r in returns)

            rows_sum = [
                ["B2B Supplies (Registered)",   f"₹{b2b_tax:.2f}", f"₹{b2b_igst:.2f}", f"₹{b2b_cgst:.2f}", f"₹{b2b_sgst:.2f}", f"₹{b2b_igst+b2b_cgst+b2b_sgst:.2f}", f"₹{b2b_grand:.0f}"],
                ["B2C Supplies (Unregistered)", f"₹{b2c_tax:.2f}", f"₹{b2c_igst:.2f}", f"₹{b2c_cgst:.2f}", f"₹{b2c_sgst:.2f}", f"₹{b2c_igst+b2c_cgst+b2c_sgst:.2f}", f"₹{b2c_grand:.0f}"],
                ["Credit Notes (Returns)",      f"- ₹{ret_tax:.2f}","",                 "",                 "",                  f"- ₹{ret_gst:.2f}",                    f"- ₹{sum(r['total_amt'] for r in returns):.0f}"],
                ["NET TOTAL",                   f"₹{b2b_tax+b2c_tax-ret_tax:.2f}",
                                                 f"₹{tot_igst:.2f}", f"₹{tot_cgst:.2f}", f"₹{tot_sgst:.2f}",
                                                 f"₹{tot_igst+tot_cgst+tot_sgst-ret_gst:.2f}",
                                                 f"₹{b2b_grand+b2c_grand-sum(r['total_amt'] for r in returns):.0f}"],
            ]
            for i,(row,bold) in enumerate(zip(rows_sum,[False,False,False,True])):
                fgs = [C_RED if bold else None]*7
                table_row(tbl_sum, row, i+1, fgs=fgs)

            # Summary cards
            stat_card(sum_cards, "Total Outward",      f"₹{b2b_tax+b2c_tax:.0f}",     C_ACCENT)
            stat_card(sum_cards, "B2B (Registered)",   f"₹{b2b_grand:.0f}",           "#276749")
            stat_card(sum_cards, "B2C (Unregistered)", f"₹{b2c_grand:.0f}",           C_PURPLE)
            stat_card(sum_cards, "Total CGST",         f"₹{tot_cgst:.2f}",            C_AMBER)
            stat_card(sum_cards, "Total SGST",         f"₹{tot_sgst:.2f}",            C_AMBER)
            stat_card(sum_cards, "Total IGST",         f"₹{tot_igst:.2f}",            "#6B46C1")
            stat_card(sum_cards, "Credit Notes",       f"- ₹{ret_gst:.2f}",           C_RED)
            stat_card(sum_cards, "Net GST Liability",  f"₹{tot_igst+tot_cgst+tot_sgst-ret_gst:.2f}", C_GREEN)

            gstr1_data["bills"] = bills
            gstr1_data["returns"] = returns
            gstr1_data["frm"] = frm; gstr1_data["to"] = to

        def export_gstr1():
            if not gstr1_data:
                messagebox.showinfo("Generate Karo","Pehle Generate dabao!"); return
            try:
                bills = gstr1_data["bills"]
                rows_b2b, rows_b2c = [], []
                for bn,b in sorted(bills.items(), key=lambda x:x[1]["date"]):
                    row = [bn, fmt_date(b["date"]), b["party"], b["gstin"],
                           b["gst_type"], b["taxable"], b["cgst"], b["sgst"],
                           b["igst"], b["total_gst"], b["grand"]]
                    if b["gstin"]: rows_b2b.append(row)
                    else:          rows_b2c.append(row)
                hdrs = ["Bill No","Date","Party","GSTIN","Type","Taxable",
                        "CGST","SGST","IGST","Total GST","Grand"]
                export_to_excel(hdrs, rows_b2b+[["---"]*11]+rows_b2c, "GSTR1_Export")
            except Exception as e:
                messagebox.showerror("Export Error", str(e))

        gen()

    # ══════════════════════════════════════════════════════════════════════════
    #  GSTR-3B
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_gstr3b(self):
        import datetime as _dt
        from collections import defaultdict

        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "📋 GSTR-3B — Monthly Summary Return")

        info = tk.Label(p,
            text="GSTR-3B = Monthly self-declaration. Outward tax − Input Tax Credit (ITC) = Net GST Payable.",
            font=("Segoe UI",9), bg="#FFF5F5", fg="#742A2A", anchor="w", padx=12, pady=6)
        info.pack(fill="x", pady=(0,3))

        # ── Filter ──────────────────────────────────────────────────────────
        ff = tk.Frame(p, bg=C_BG); ff.pack(fill="x", pady=(0,10))
        tk.Label(ff, text="Month:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        months = []
        today = _dt.date.today()
        for i in range(12):
            d = today.replace(day=1) - _dt.timedelta(days=i*28)
            months.append(d.replace(day=1).strftime("%Y-%m"))
        months = sorted(set(months), reverse=True)
        v_month = tk.StringVar(value=today.strftime("%Y-%m"))
        ttk.Combobox(ff, textvariable=v_month, values=months,
                     width=10, state="readonly").pack(side="left", padx=6)
        make_btn(ff, "📊 Generate", lambda: gen(), bg="#1A365D").pack(side="left", padx=10)
        make_btn(ff, "📥 Excel", lambda: export_3b(), bg=C_GREEN).pack(side="left", padx=4)
        make_btn(ff, "🖨️ Print View", lambda: print_3b(), bg=C_PURPLE).pack(side="left", padx=4)

        # ── Main body ─────────────────────────────────────────────────────
        body = tk.Frame(p, bg=C_BG); body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=1); body.columnconfigure(1, weight=1)

        left  = tk.Frame(body, bg=C_BG); left.grid( row=0, column=0, sticky="nsew", padx=(0,6))
        right = tk.Frame(body, bg=C_BG); right.grid(row=0, column=1, sticky="nsew", padx=(6,0))

        def section_box(parent, title, color="#1A365D"):
            box = tk.Frame(parent, bg=C_WHITE,
                           highlightthickness=1, highlightbackground=C_BORDER)
            box.pack(fill="x", pady=(0,10))
            tk.Label(box, text=title, font=("Segoe UI",10,"bold"),
                     bg=color, fg="white", padx=12, pady=6).pack(fill="x")
            inner = tk.Frame(box, bg=C_WHITE); inner.pack(fill="x", padx=2, pady=2)
            return inner

        def row3b(parent, label, val, bold=False, fg=C_GRAY, bg=C_WHITE, indent=0):
            r = tk.Frame(parent, bg=bg); r.pack(fill="x", pady=1)
            tk.Label(r, text="  "*indent + label,
                     font=("Segoe UI",9,"bold" if bold else "normal"),
                     bg=bg, fg="#1A365D" if bold else C_GRAY,
                     anchor="w", padx=8, pady=3).pack(side="left", fill="x", expand=True)
            tk.Label(r, text=str(val),
                     font=("Segoe UI",9,"bold" if bold else "normal"),
                     bg=bg, fg=fg, anchor="e", padx=12).pack(side="right")

        def divider(parent):
            tk.Frame(parent, bg=C_BORDER, height=1).pack(fill="x", padx=8, pady=2)

        data_store = {}

        def gen():
            for w in left.winfo_children(): w.destroy()
            for w in right.winfo_children(): w.destroy()

            m = v_month.get()
            y, mo = int(m[:4]), int(m[5:])
            import calendar as _cal
            last = _cal.monthrange(y, mo)[1]
            frm = f"{m}-01"; to = f"{m}-{last:02d}"

            conn = get_db()
            sales   = [dict(r) for r in conn.execute(
                "SELECT s.*, si.taxable, si.gst_percent, si.gst_amt "
                "FROM sales s JOIN sale_items si ON s.id=si.sale_id "
                "WHERE s.bill_date>=? AND s.bill_date<=?", (frm,to)).fetchall()]
            purs    = [dict(r) for r in conn.execute(
                "SELECT p.*, pi.taxable, pi.gst_percent, pi.gst_amt "
                "FROM purchases p JOIN purchase_items pi ON p.id=pi.purchase_id "
                "WHERE p.bill_date>=? AND p.bill_date<=?", (frm,to)).fetchall()]
            returns = [dict(r) for r in conn.execute(
                "SELECT * FROM returns WHERE return_date>=? AND return_date<=?",
                (frm,to)).fetchall()]
            expenses= [dict(r) for r in conn.execute(
                "SELECT * FROM expenses WHERE exp_date>=? AND exp_date<=?",
                (frm,to)).fetchall()]
            conn.close()

            # ── 3.1 Outward supplies ──────────────────────────────────────
            def gst_split(rows):
                igst=0; cgst=0; sgst=0; taxable=0
                for r in rows:
                    ga = round(r["gst_amt"],2)
                    tx = round(r["taxable"],2)
                    gt = r.get("gst_type","CGST+SGST") or "CGST+SGST"
                    taxable += tx
                    if gt == "IGST":  igst += ga
                    else: cgst += round(ga/2,2); sgst += round(ga/2,2)
                return taxable, igst, cgst, sgst

            b2b_sales = [r for r in sales if r.get("party_gstin","")]
            b2c_sales = [r for r in sales if not r.get("party_gstin","")]
            sale_ret  = [r for r in returns if r["return_type"]=="sale_return"]
            pur_ret   = [r for r in returns if r["return_type"]=="pur_return"]

            b2b_tx,b2b_igst,b2b_cgst,b2b_sgst = gst_split(b2b_sales)
            b2c_tx,b2c_igst,b2c_cgst,b2c_sgst = gst_split(b2c_sales)
            ret_amt = sum(r["total_amt"] for r in sale_ret)
            ret_gst = sum(r["gst_amt"] for r in sale_ret)

            total_out_tax = b2b_igst+b2b_cgst+b2b_sgst+b2c_igst+b2c_cgst+b2c_sgst - ret_gst
            total_out_taxable = b2b_tx + b2c_tx

            # ── 3.2 ITC (Input Tax Credit from purchases) ─────────────────
            itc_igst=0; itc_cgst=0; itc_sgst=0; itc_taxable=0
            for r in purs:
                ga = round(r["gst_amt"],2)
                itc_taxable += round(r["taxable"],2)
                # Assume all purchase = intrastate (CGST+SGST) unless no GSTIN
                itc_cgst += round(ga/2,2); itc_sgst += round(ga/2,2)
            # Deduct purchase returns from ITC
            pr_gst = sum(r["gst_amt"] for r in pur_ret)
            total_itc = itc_igst + itc_cgst + itc_sgst - pr_gst
            net_total_itc = max(0, total_itc)

            # ── 3.3 Net GST payable ───────────────────────────────────────
            net_igst  = max(0, b2b_igst+b2c_igst - itc_igst)
            net_cgst  = max(0, b2b_cgst+b2c_cgst - itc_cgst)
            net_sgst  = max(0, b2b_sgst+b2c_sgst - itc_sgst)
            net_payable = net_igst + net_cgst + net_sgst - ret_gst
            net_payable = max(0, net_payable)

            total_exp_amt = sum(e["amount"] for e in expenses)

            # Save for export
            data_store.update({
                "month":m,"frm":frm,"to":to,
                "b2b_tx":b2b_tx,"b2b_igst":b2b_igst,"b2b_cgst":b2b_cgst,"b2b_sgst":b2b_sgst,
                "b2c_tx":b2c_tx,"b2c_igst":b2c_igst,"b2c_cgst":b2c_cgst,"b2c_sgst":b2c_sgst,
                "ret_amt":ret_amt,"ret_gst":ret_gst,"total_out_tax":total_out_tax,
                "itc_cgst":itc_cgst,"itc_sgst":itc_sgst,"itc_igst":itc_igst,
                "pr_gst":pr_gst,"total_itc":net_total_itc,
                "net_igst":net_igst,"net_cgst":net_cgst,"net_sgst":net_sgst,
                "net_payable":net_payable,"expenses":total_exp_amt,
            })

            fmt = lambda v: f"₹{abs(round(v)):,.2f}"

            # ══ LEFT COLUMN ══════════════════════════════════════════════
            # 3.1 Outward Supplies
            s31 = section_box(left, "3.1 — Outward Supplies (Taxable)", "#1A365D")
            row3b(s31, "B2B — Registered Buyers",    "", bold=True)
            row3b(s31, "Taxable Value",   fmt(b2b_tx),   indent=1)
            row3b(s31, "IGST",            fmt(b2b_igst), indent=1, fg="⬛" if b2b_igst else C_GRAY)
            row3b(s31, "CGST",            fmt(b2b_cgst), indent=1, fg=C_AMBER)
            row3b(s31, "SGST",            fmt(b2b_sgst), indent=1, fg=C_AMBER)
            divider(s31)
            row3b(s31, "B2C — Unregistered Buyers",  "", bold=True)
            row3b(s31, "Taxable Value",   fmt(b2c_tx),   indent=1)
            row3b(s31, "IGST",            fmt(b2c_igst), indent=1)
            row3b(s31, "CGST",            fmt(b2c_cgst), indent=1, fg=C_AMBER)
            row3b(s31, "SGST",            fmt(b2c_sgst), indent=1, fg=C_AMBER)
            divider(s31)
            row3b(s31, "Less: Credit Notes (Returns)", f"- {fmt(ret_gst)}", fg=C_RED)
            divider(s31)
            row3b(s31, "TOTAL Outward Tax Liability",
                  fmt(total_out_tax), bold=True, fg=C_RED)

            # 3.2 ITC
            s32 = section_box(left, "3.2 — Input Tax Credit (ITC)", "#276749")
            row3b(s32, "ITC from Purchases",   "", bold=True)
            row3b(s32, "IGST Credit",           fmt(itc_igst), indent=1, fg=C_GREEN)
            row3b(s32, "CGST Credit",           fmt(itc_cgst), indent=1, fg=C_GREEN)
            row3b(s32, "SGST Credit",           fmt(itc_sgst), indent=1, fg=C_GREEN)
            row3b(s32, "Less: Purchase Returns",f"- {fmt(pr_gst)}", indent=1, fg=C_RED)
            divider(s32)
            row3b(s32, "Net ITC Available",     fmt(net_total_itc), bold=True, fg=C_GREEN)

            # ══ RIGHT COLUMN ═════════════════════════════════════════════
            # 3.3 Net Payable
            s33 = section_box(right, "3.3 — Net Tax Payable", "#742A2A")
            row3b(s33, "IGST Payable",       fmt(net_igst), fg="#6B46C1")
            row3b(s33, "  Less ITC (IGST)",  fmt(itc_igst), indent=1, fg=C_GREEN)
            divider(s33)
            row3b(s33, "CGST Payable",       fmt(b2b_cgst+b2c_cgst), fg=C_AMBER)
            row3b(s33, "  Less ITC (CGST)",  fmt(itc_cgst), indent=1, fg=C_GREEN)
            divider(s33)
            row3b(s33, "SGST Payable",       fmt(b2b_sgst+b2c_sgst), fg=C_AMBER)
            row3b(s33, "  Less ITC (SGST)",  fmt(itc_sgst), indent=1, fg=C_GREEN)
            divider(s33)
            row3b(s33, "Less: Credit Notes", f"- {fmt(ret_gst)}", fg=C_RED)
            divider(s33)
            net_clr = C_GREEN if net_payable == 0 else C_RED
            row3b(s33, "🔴 NET GST PAYABLE",  fmt(net_payable),
                  bold=True, fg=net_clr,
                  bg="#FFF5F5" if net_payable>0 else "#F0FFF4")

            # Other info
            s_other = section_box(right, "Other Details", "#2B6CB0")
            row3b(s_other, "Total Invoices (Sales)",     len(set(r["bill_no"] for r in sales)))
            row3b(s_other, "Total Purchases",            len(set(r["bill_no"] for r in purs)))
            row3b(s_other, "Sale Returns (CDN)",         len(sale_ret))
            row3b(s_other, "Purchase Returns",           len(pur_ret))
            row3b(s_other, "Total Expenses (this month)",fmt(total_exp_amt))
            divider(s_other)
            row3b(s_other, "Turnover (Gross Sale)",
                  fmt(sum(r["grand_total"] for r in
                          [dict(r2) for r2 in
                           get_db().execute("SELECT DISTINCT bill_no,grand_total FROM sales WHERE bill_date>=? AND bill_date<=?",(frm,to)).fetchall()])),
                  bold=True, fg=C_ACCENT)

        def export_3b():
            if not data_store:
                messagebox.showinfo("Generate Karo","Pehle Generate dabao!"); return
            d = data_store
            rows = [
                ["GSTR-3B", f"Month: {d['month']}", ""],
                ["Section","Description","Amount"],
                ["3.1","B2B Taxable",           f"₹{d['b2b_tx']:.2f}"],
                ["3.1","B2B CGST",              f"₹{d['b2b_cgst']:.2f}"],
                ["3.1","B2B SGST",              f"₹{d['b2b_sgst']:.2f}"],
                ["3.1","B2B IGST",              f"₹{d['b2b_igst']:.2f}"],
                ["3.1","B2C Taxable",           f"₹{d['b2c_tx']:.2f}"],
                ["3.1","B2C CGST",              f"₹{d['b2c_cgst']:.2f}"],
                ["3.1","B2C SGST",              f"₹{d['b2c_sgst']:.2f}"],
                ["3.1","B2C IGST",              f"₹{d['b2c_igst']:.2f}"],
                ["3.1","Less: Credit Notes",    f"- ₹{d['ret_gst']:.2f}"],
                ["3.1","TOTAL Outward Tax",     f"₹{d['total_out_tax']:.2f}"],
                ["---","---","---"],
                ["3.2","ITC CGST",              f"₹{d['itc_cgst']:.2f}"],
                ["3.2","ITC SGST",              f"₹{d['itc_sgst']:.2f}"],
                ["3.2","ITC IGST",              f"₹{d['itc_igst']:.2f}"],
                ["3.2","Less: Purchase Returns",f"- ₹{d['pr_gst']:.2f}"],
                ["3.2","Net ITC Available",     f"₹{d['total_itc']:.2f}"],
                ["---","---","---"],
                ["3.3","Net IGST Payable",      f"₹{d['net_igst']:.2f}"],
                ["3.3","Net CGST Payable",      f"₹{d['net_cgst']:.2f}"],
                ["3.3","Net SGST Payable",      f"₹{d['net_sgst']:.2f}"],
                ["3.3","NET TOTAL GST PAYABLE", f"₹{d['net_payable']:.2f}"],
            ]
            try:
                export_to_excel(["Section","Description","Amount"], rows, f"GSTR3B_{d['month']}")
            except Exception as e:
                messagebox.showerror("Export Error", str(e))

        def print_3b():
            if not data_store:
                messagebox.showinfo("Generate Karo","Pehle Generate dabao!"); return
            d = data_store
            _sgr = get_shop()
            text = f"""
GSTR-3B SUMMARY
Month: {d['month']}  |  {_sgr['name']}  |  GSTIN: {_sgr['gstin'] or '—'}
{'='*55}

3.1 OUTWARD SUPPLIES
  B2B  Taxable: ₹{d['b2b_tx']:,.2f}  CGST: ₹{d['b2b_cgst']:,.2f}  SGST: ₹{d['b2b_sgst']:,.2f}  IGST: ₹{d['b2b_igst']:,.2f}
  B2C  Taxable: ₹{d['b2c_tx']:,.2f}  CGST: ₹{d['b2c_cgst']:,.2f}  SGST: ₹{d['b2c_sgst']:,.2f}  IGST: ₹{d['b2c_igst']:,.2f}
  Less Credit Notes: - ₹{d['ret_gst']:,.2f}
  TOTAL OUTWARD TAX: ₹{d['total_out_tax']:,.2f}

3.2 INPUT TAX CREDIT (ITC)
  CGST Credit: ₹{d['itc_cgst']:,.2f}
  SGST Credit: ₹{d['itc_sgst']:,.2f}
  IGST Credit: ₹{d['itc_igst']:,.2f}
  Less Pur Returns: - ₹{d['pr_gst']:,.2f}
  NET ITC: ₹{d['total_itc']:,.2f}

3.3 NET GST PAYABLE
  IGST: ₹{d['net_igst']:,.2f}
  CGST: ₹{d['net_cgst']:,.2f}
  SGST: ₹{d['net_sgst']:,.2f}
  ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  NET PAYABLE: ₹{d['net_payable']:,.2f}
{'='*55}
"""
            dlg = tk.Toplevel(p.winfo_toplevel())
            dlg.title("GSTR-3B Print View"); dlg.configure(bg=C_WHITE); dlg.geometry("520x480")
            tk.Label(dlg, text="GSTR-3B Print View", font=("Segoe UI",11,"bold"),
                     bg="#1A365D", fg="white", pady=4).pack(fill="x")
            txt = tk.Text(dlg, font=("Courier New",10), wrap="none", padx=12, pady=4)
            txt.pack(fill="both", expand=True, padx=8, pady=4)
            txt.insert("1.0", text); txt.config(state="disabled")
            bf = tk.Frame(dlg, bg=C_WHITE); bf.pack(pady=(0,10))
            make_btn(bf,"📋 Copy", lambda: [dlg.clipboard_clear(), dlg.clipboard_append(text)],
                     bg=C_AMBER).pack(side="left",padx=6)
            make_btn(bf,"Close", dlg.destroy, bg=C_GRAY).pack(side="left",padx=6)

        gen()

    # ══════════════════════════════════════════════════════════════════════════
    #  E-INVOICE REGISTER PAGE
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_einvoice(self):
        import datetime as _dt

        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "🔗 E-Invoice Register")

        info = tk.Label(p,
            text="Yahan sabhi invoices ka IRN (Invoice Reference Number) status dikhta hai. "
                 "Invoice pe click karo → E-Invoice button → JSON copy karo → IRP portal pe generate karo → IRN paste karo.",
            font=("Segoe UI",9), bg="#FFF5EB", fg="#C05621", anchor="w",
            padx=12, pady=6, wraplength=900, justify="left")
        info.pack(fill="x", pady=(0,3))

        # ── Filter ──────────────────────────────────────────────────────────
        ff = tk.Frame(p, bg=C_BG); ff.pack(fill="x", pady=(0,3))
        tk.Label(ff, text="From:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_from = tk.StringVar()
        make_date_entry(ff, v_from, width=12).pack(side="left", padx=4)
        tk.Label(ff, text="To:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_to = tk.StringVar(value=str(_dt.date.today()))
        make_date_entry(ff, v_to, width=12).pack(side="left", padx=4)
        tk.Label(ff, text="Status:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left", padx=(10,0))
        v_status = tk.StringVar(value="All")
        ttk.Combobox(ff, textvariable=v_status, width=14, state="readonly",
                     values=["All","IRN Generated","Pending"]).pack(side="left", padx=4)
        make_btn(ff, "🔍 Refresh", lambda: load(), bg="#2B6CB0").pack(side="left", padx=8)
        make_btn(ff, "📥 Excel",   lambda: export_ei(), bg=C_GREEN).pack(side="left", padx=4)

        # ── Summary cards ────────────────────────────────────────────────────
        sc = tk.Frame(p, bg=C_BG); sc.pack(fill="x", pady=(0,3))
        lbl_total   = tk.Label(sc, text="Total Bills: 0", font=("Segoe UI",10,"bold"),
                               bg=C_WHITE, fg="#1A365D", padx=16, pady=8,
                               highlightthickness=1, highlightbackground=C_BORDER)
        lbl_total.pack(side="left", padx=(0,8))
        lbl_done    = tk.Label(sc, text="IRN Generated: 0", font=("Segoe UI",10,"bold"),
                               bg=C_WHITE, fg=C_GREEN, padx=16, pady=8,
                               highlightthickness=1, highlightbackground=C_BORDER)
        lbl_done.pack(side="left", padx=(0,8))
        lbl_pending = tk.Label(sc, text="Pending: 0", font=("Segoe UI",10,"bold"),
                               bg=C_WHITE, fg=C_RED, padx=16, pady=8,
                               highlightthickness=1, highlightbackground=C_BORDER)
        lbl_pending.pack(side="left")

        # ── Table ────────────────────────────────────────────────────────────
        tbl = make_table(p,
            ["Bill No","Date","Party","GSTIN","Grand","IRN","ACK No","ACK Date","Status","Action"],
            [14, 11, 18, 20, 10, 18, 14, 11, 12, 8])

        all_rows = []

        def load():
            clear_table_rows(tbl)
            all_rows.clear()
            frm = v_from.get().strip()
            to  = v_to.get().strip()
            st  = v_status.get()

            conn = get_db()
            q = "SELECT * FROM sales WHERE 1=1"
            params = []
            if frm: q += " AND bill_date>=?"; params.append(frm)
            if to:  q += " AND bill_date<=?"; params.append(to)
            q += " ORDER BY bill_date DESC"
            rows = [dict(r) for r in conn.execute(q, params).fetchall()]
            conn.close()

            if st == "IRN Generated": rows = [r for r in rows if r.get("irn","")]
            elif st == "Pending":     rows = [r for r in rows if not r.get("irn","")]

            done    = sum(1 for r in rows if r.get("irn",""))
            pending = len(rows) - done
            lbl_total.config(  text=f"Total Bills: {len(rows)}")
            lbl_done.config(   text=f"IRN Generated: {done}")
            lbl_pending.config(text=f"Pending: {pending}")
            all_rows.extend(rows)

            for i, r in enumerate(rows):
                bg    = C_WHITE if i%2==0 else "#F7FAFC"
                irn   = r.get("irn","") or ""
                ack   = r.get("ack_no","") or ""
                adate = r.get("ack_date","") or ""
                has_irn = bool(irn)
                status_txt = "✅ Generated" if has_irn else "⏳ Pending"
                status_clr = C_GREEN        if has_irn else C_RED

                vals = [r["bill_no"], fmt_date(r["bill_date"]), r["party"][:16],
                        (r.get("party_gstin","") or "")[:18],
                        f"₹{r['grand_total']:,.0f}",
                        irn[:16]+"…" if len(irn)>16 else irn,
                        ack, fmt_date(adate) if adate else "-",
                        status_txt]
                fgs = [C_ACCENT, None, None, "#6B46C1", C_GREEN,
                       C_GREEN if has_irn else C_GRAY,
                       None, None, status_clr]

                for j,(val,fg) in enumerate(zip(vals,fgs)):
                    tk.Label(tbl, text=str(val), font=("Segoe UI",9),
                             bg=bg, fg=fg or C_GRAY, anchor="w", padx=4, pady=4
                             ).grid(row=i+1, column=j, sticky="nsew", padx=1)

                # IRN Entry inline button
                tk.Button(tbl, text="IRN", font=("Segoe UI",8,"bold"),
                          bg="#C05621", fg="white", relief="flat", cursor="hand2",
                          command=lambda rd=r: irn_popup(rd)
                          ).grid(row=i+1, column=9, sticky="nsew", padx=2, pady=2)

        def irn_popup(r):
            dlg = tk.Toplevel(p.winfo_toplevel())
            dlg.title(f"IRN — {r['bill_no']}")
            dlg.geometry("480x200"); dlg.configure(bg=C_WHITE); dlg.grab_set()
            tk.Label(dlg, text=f"IRN Entry — {r['bill_no']}",
                     font=("Segoe UI",10,"bold"), bg="#C05621", fg="white", pady=4).pack(fill="x")
            frm2 = tk.Frame(dlg, bg=C_WHITE, padx=16, pady=12); frm2.pack(fill="x")
            tk.Label(frm2, text="IRN:", font=("Segoe UI",9), bg=C_WHITE).pack(anchor="w")
            v_irn = tk.StringVar(value=r.get("irn","") or "")
            ttk.Entry(frm2, textvariable=v_irn, width=54).pack(anchor="w", pady=(2,8))
            rf = tk.Frame(frm2, bg=C_WHITE); rf.pack(fill="x")
            tk.Label(rf, text="ACK No:", font=("Segoe UI",9), bg=C_WHITE).pack(side="left")
            v_ack = tk.StringVar(value=r.get("ack_no","") or "")
            ttk.Entry(rf, textvariable=v_ack, width=18).pack(side="left", padx=(4,14))
            tk.Label(rf, text="ACK Date:", font=("Segoe UI",9), bg=C_WHITE).pack(side="left")
            v_ad = tk.StringVar(value=r.get("ack_date","") or "")
            make_date_entry(rf, v_ad, width=13).pack(side="left", padx=4)
            def save_irn():
                conn2 = get_db()
                conn2.execute("UPDATE sales SET irn=?,ack_no=?,ack_date=? WHERE id=?",
                              (v_irn.get().strip(), v_ack.get().strip(), v_ad.get().strip(), r["id"]))
                conn2.commit(); conn2.close()
                messagebox.showinfo("✅ Saved","IRN saved!", parent=dlg)
                dlg.destroy(); load()
            bf2 = tk.Frame(dlg, bg=C_WHITE); bf2.pack(pady=(0,10))
            make_btn(bf2,"✅ Save IRN", save_irn, bg=C_GREEN).pack(side="left", padx=6)
            make_btn(bf2,"Cancel", dlg.destroy, bg=C_GRAY).pack(side="left", padx=6)

        def export_ei():
            if not all_rows:
                messagebox.showinfo("No Data","Pehle Refresh karo!"); return
            data = [[r["bill_no"], fmt_date(r["bill_date"]), r["party"],
                     r.get("party_gstin","") or "", f"₹{r['grand_total']:,.0f}",
                     r.get("irn","") or "", r.get("ack_no","") or "",
                     r.get("ack_date","") or "",
                     "Generated" if r.get("irn","") else "Pending"]
                    for r in all_rows]
            try:
                export_to_excel(
                    ["Bill No","Date","Party","GSTIN","Grand","IRN","ACK No","ACK Date","Status"],
                    data, "EInvoice_Register")
            except Exception as e:
                messagebox.showerror("Export Error", str(e))

        load()

    # ══════════════════════════════════════════════════════════════════════════
    #  E-WAY BILL REGISTER PAGE
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_ewaybill(self):
        import datetime as _dt

        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "🚛 E-Way Bill Register")

        info = tk.Label(p,
            text="E-Way Bill — Goods transport ke liye mandatory (value > ₹50,000). "
                 "Invoice kholo → E-Way Bill button → Transport details bharo → JSON copy karo → NIC portal pe generate karo → EWB No save karo.",
            font=("Segoe UI",9), bg="#F0FFF4", fg="#276749", anchor="w",
            padx=12, pady=6, wraplength=900, justify="left")
        info.pack(fill="x", pady=(0,3))

        # ── Filter ──────────────────────────────────────────────────────────
        ff = tk.Frame(p, bg=C_BG); ff.pack(fill="x", pady=(0,3))
        tk.Label(ff, text="From:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_from = tk.StringVar()
        make_date_entry(ff, v_from, width=12).pack(side="left", padx=4)
        tk.Label(ff, text="To:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_to = tk.StringVar(value=str(_dt.date.today()))
        make_date_entry(ff, v_to, width=12).pack(side="left", padx=4)
        tk.Label(ff, text="Status:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left", padx=(10,0))
        v_status = tk.StringVar(value="All")
        ttk.Combobox(ff, textvariable=v_status, width=14, state="readonly",
                     values=["All","EWB Generated","Pending","Value > 50K"]).pack(side="left", padx=4)
        make_btn(ff, "🔍 Refresh",  lambda: load(), bg="#2B6CB0").pack(side="left", padx=8)
        make_btn(ff, "📥 Excel",    lambda: export_ewb(), bg=C_GREEN).pack(side="left", padx=4)
        make_btn(ff, "🌐 NIC Portal", lambda: __import__("webbrowser").open("https://ewaybillgst.gov.in"),
                 bg="#276749").pack(side="left", padx=4)

        # ── Summary cards ────────────────────────────────────────────────────
        sc = tk.Frame(p, bg=C_BG); sc.pack(fill="x", pady=(0,3))
        lbl_total   = tk.Label(sc, text="Total Bills: 0",    font=("Segoe UI",10,"bold"),
                               bg=C_WHITE, fg="#1A365D", padx=16, pady=8,
                               highlightthickness=1, highlightbackground=C_BORDER)
        lbl_total.pack(side="left", padx=(0,8))
        lbl_ewb     = tk.Label(sc, text="EWB Generated: 0",  font=("Segoe UI",10,"bold"),
                               bg=C_WHITE, fg=C_GREEN, padx=16, pady=8,
                               highlightthickness=1, highlightbackground=C_BORDER)
        lbl_ewb.pack(side="left", padx=(0,8))
        lbl_pending = tk.Label(sc, text="Pending: 0",        font=("Segoe UI",10,"bold"),
                               bg=C_WHITE, fg=C_RED, padx=16, pady=8,
                               highlightthickness=1, highlightbackground=C_BORDER)
        lbl_pending.pack(side="left", padx=(0,8))
        lbl_above50 = tk.Label(sc, text="> ₹50K (Mandatory): 0", font=("Segoe UI",10,"bold"),
                               bg=C_WHITE, fg=C_AMBER, padx=16, pady=8,
                               highlightthickness=1, highlightbackground=C_BORDER)
        lbl_above50.pack(side="left")

        # ── Table ────────────────────────────────────────────────────────────
        tbl = make_table(p,
            ["Bill No","Date","Party","Grand","EWB No","Vehicle","Mode","Distance","Status","Edit"],
            [13, 11, 18, 10, 16, 12, 8, 9, 13, 6])

        all_rows = []

        def load():
            clear_table_rows(tbl)
            all_rows.clear()
            frm = v_from.get().strip()
            to  = v_to.get().strip()
            st  = v_status.get()

            conn = get_db()
            q = "SELECT * FROM sales WHERE 1=1"
            params = []
            if frm: q += " AND bill_date>=?"; params.append(frm)
            if to:  q += " AND bill_date<=?"; params.append(to)
            q += " ORDER BY bill_date DESC"
            rows = [dict(r) for r in conn.execute(q, params).fetchall()]
            conn.close()

            if st == "EWB Generated": rows = [r for r in rows if r.get("ewb_no","")]
            elif st == "Pending":     rows = [r for r in rows if not r.get("ewb_no","")]
            elif st == "Value > 50K": rows = [r for r in rows if r["grand_total"] >= 50000]

            done    = sum(1 for r in rows if r.get("ewb_no",""))
            pending = len(rows) - done
            above50 = sum(1 for r in rows if r["grand_total"] >= 50000)

            lbl_total.config(  text=f"Total Bills: {len(rows)}")
            lbl_ewb.config(    text=f"EWB Generated: {done}")
            lbl_pending.config(text=f"Pending: {pending}")
            lbl_above50.config(text=f"> ₹50K (Mandatory): {above50}")
            all_rows.extend(rows)

            for i, r in enumerate(rows):
                bg      = C_WHITE if i%2==0 else "#F7FAFC"
                ewb_no  = r.get("ewb_no","")  or ""
                veh     = r.get("vehicle_no","") or "-"
                mode    = r.get("transport_mode","") or "-"
                dist    = r.get("distance",0)  or "-"
                has_ewb = bool(ewb_no)
                above   = r["grand_total"] >= 50000
                status_txt = "✅ EWB Done" if has_ewb else ("⚠️ Mandatory!" if above else "Optional")
                status_clr = C_GREEN if has_ewb else (C_RED if above else C_GRAY)

                vals = [r["bill_no"], fmt_date(r["bill_date"]), r["party"][:16],
                        f"₹{r['grand_total']:,.0f}", ewb_no[:14] if ewb_no else "-",
                        veh, mode, str(dist), status_txt]
                fgs  = [C_ACCENT, None, None, C_GREEN if above else None,
                        C_GREEN if has_ewb else C_GRAY,
                        None, None, None, status_clr]

                for j,(val,fg) in enumerate(zip(vals,fgs)):
                    tk.Label(tbl, text=str(val), font=("Segoe UI",9),
                             bg=bg, fg=fg or C_GRAY, anchor="w", padx=4, pady=4
                             ).grid(row=i+1, column=j, sticky="nsew", padx=1)

                tk.Button(tbl, text="EWB", font=("Segoe UI",8,"bold"),
                          bg="#276749", fg="white", relief="flat", cursor="hand2",
                          command=lambda rd=r: ewb_popup(rd)
                          ).grid(row=i+1, column=9, sticky="nsew", padx=2, pady=2)

        def ewb_popup(r):
            dlg = tk.Toplevel(p.winfo_toplevel())
            dlg.title(f"E-Way Bill — {r['bill_no']}")
            dlg.geometry("520x260"); dlg.configure(bg=C_WHITE); dlg.grab_set()
            tk.Label(dlg, text=f"🚛 E-Way Bill — {r['bill_no']}",
                     font=("Segoe UI",10,"bold"), bg="#276749", fg="white", pady=4).pack(fill="x")
            frm2 = tk.Frame(dlg, bg=C_WHITE, padx=16, pady=4); frm2.pack(fill="x")

            r1 = tk.Frame(frm2, bg=C_WHITE); r1.pack(fill="x", pady=4)
            tk.Label(r1, text="EWB No:", font=("Segoe UI",9), bg=C_WHITE).pack(side="left")
            v_ewb = tk.StringVar(value=r.get("ewb_no","") or "")
            ttk.Entry(r1, textvariable=v_ewb, width=20).pack(side="left", padx=(4,14))
            tk.Label(r1, text="EWB Date:", font=("Segoe UI",9), bg=C_WHITE).pack(side="left")
            v_ed = tk.StringVar(value=r.get("ewb_date","") or "")
            make_date_entry(r1, v_ed, width=13).pack(side="left", padx=4)

            r2 = tk.Frame(frm2, bg=C_WHITE); r2.pack(fill="x", pady=4)
            tk.Label(r2, text="Vehicle No:", font=("Segoe UI",9), bg=C_WHITE).pack(side="left")
            v_veh = tk.StringVar(value=r.get("vehicle_no","") or "")
            ttk.Entry(r2, textvariable=v_veh, width=14).pack(side="left", padx=(4,14))
            tk.Label(r2, text="Mode:", font=("Segoe UI",9), bg=C_WHITE).pack(side="left")
            v_mode = tk.StringVar(value=r.get("transport_mode","Road") or "Road")
            ttk.Combobox(r2, textvariable=v_mode, width=8, state="readonly",
                         values=["Road","Rail","Air","Ship"]).pack(side="left", padx=(4,14))
            tk.Label(r2, text="Distance (km):", font=("Segoe UI",9), bg=C_WHITE).pack(side="left")
            v_dist = tk.StringVar(value=str(r.get("distance","") or ""))
            ttk.Entry(r2, textvariable=v_dist, width=8).pack(side="left", padx=4)

            r3 = tk.Frame(frm2, bg=C_WHITE); r3.pack(fill="x", pady=4)
            tk.Label(r3, text="Transporter:", font=("Segoe UI",9), bg=C_WHITE).pack(side="left")
            v_trans = tk.StringVar(value=r.get("transporter","") or "")
            ttk.Entry(r3, textvariable=v_trans, width=30).pack(side="left", padx=4)

            def save_ewb():
                try: dist_int = int(v_dist.get()) if v_dist.get().strip() else 0
                except: dist_int = 0
                conn2 = get_db()
                conn2.execute(
                    "UPDATE sales SET ewb_no=?,ewb_date=?,vehicle_no=?,transport_mode=?,distance=?,transporter=? WHERE id=?",
                    (v_ewb.get().strip(), v_ed.get().strip(),
                     v_veh.get().strip().upper(), v_mode.get(),
                     dist_int, v_trans.get().strip(), r["id"]))
                conn2.commit(); conn2.close()
                messagebox.showinfo("✅ Saved","E-Way Bill details saved!", parent=dlg)
                dlg.destroy(); load()

            bf2 = tk.Frame(dlg, bg=C_WHITE); bf2.pack(pady=(0,10))
            make_btn(bf2,"✅ Save EWB",  save_ewb,   bg=C_GREEN).pack(side="left", padx=6)
            make_btn(bf2,"Cancel",       dlg.destroy, bg=C_GRAY).pack(side="left", padx=6)

        def export_ewb():
            if not all_rows:
                messagebox.showinfo("No Data","Pehle Refresh karo!"); return
            data = [[r["bill_no"], fmt_date(r["bill_date"]), r["party"],
                     f"₹{r['grand_total']:,.0f}",
                     r.get("ewb_no","") or "", r.get("ewb_date","") or "",
                     r.get("vehicle_no","") or "", r.get("transport_mode","") or "",
                     str(r.get("distance","") or ""),
                     "Generated" if r.get("ewb_no","") else "Pending"]
                    for r in all_rows]
            try:
                export_to_excel(
                    ["Bill No","Date","Party","Grand","EWB No","EWB Date",
                     "Vehicle No","Mode","Distance","Status"],
                    data, "EWayBill_Register")
            except Exception as e:
                messagebox.showerror("Export Error", str(e))

        load()

    # ══════════════════════════════════════════════════════════════════════════
    #  SALE RETURN
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_salereturn(self):
        """Sale Return — Multiple Products + Return Bill Print"""
        import datetime as _dt
        db = get_db()
        db.execute("CREATE TABLE IF NOT EXISTS returns (id INTEGER PRIMARY KEY AUTOINCREMENT, return_no TEXT UNIQUE NOT NULL, return_type TEXT NOT NULL, return_date TEXT NOT NULL, orig_bill TEXT NOT NULL, party TEXT NOT NULL, product TEXT NOT NULL, qty REAL DEFAULT 0, rate REAL DEFAULT 0, gst_percent REAL DEFAULT 0, gst_amt REAL DEFAULT 0, total_amt REAL DEFAULT 0, reason TEXT DEFAULT '', created_at TEXT DEFAULT CURRENT_TIMESTAMP)")
        db.commit(); db.close()

        self._sr_items = []   # in-memory cart for this session

        p = tk.Frame(self.content, bg=C_BG, padx=10, pady=3)
        p.pack(fill="both", expand=True)
        section_title(p, "Sale Return")

        form = tk.Frame(p, bg=C_WHITE, highlightthickness=1, highlightbackground=C_BORDER)
        form.pack(fill="x", pady=(0,10))
        tk.Label(form, text="Customer Return Entry — Ek ya Zyada Products",
                 font=("Segoe UI",10,"bold"), bg=C_WHITE, fg="#1A365D").pack(anchor="w", padx=16, pady=(10,6))

        # ── Row 1: Date / Bill / Customer / Reason ────────────────────────────
        row1 = tk.Frame(form, bg=C_WHITE); row1.pack(fill="x", padx=16, pady=4)
        tk.Label(row1, text="Date:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_date = tk.StringVar(value=str(_dt.date.today()))
        sr_date_entry = make_date_entry(row1, v_date, width=13)
        sr_date_entry.pack(side="left", padx=(2,14))
        tk.Label(row1, text="Original Sale Bill:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_bill = tk.StringVar()
        db2 = get_db()
        bills = [r[0] for r in db2.execute("SELECT bill_no FROM sales ORDER BY bill_date DESC").fetchall()]
        db2.close()
        cb_bill = ttk.Combobox(row1, textvariable=v_bill, values=bills, width=16)
        cb_bill.pack(side="left", padx=(2,14))
        add_autocomplete(cb_bill)
        tk.Label(row1, text="Customer:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_party = tk.StringVar()
        e_party = ttk.Entry(row1, textvariable=v_party, width=20, state="readonly")
        e_party.pack(side="left", padx=(2,14))
        tk.Label(row1, text="Reason:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_reason = tk.StringVar()
        sr_reason_entry = ttk.Entry(row1, textvariable=v_reason, width=22)
        sr_reason_entry.pack(side="left", padx=(2,0))

        # ── Row 2: Product entry ──────────────────────────────────────────────
        row2 = tk.Frame(form, bg=C_WHITE); row2.pack(fill="x", padx=16, pady=4)
        tk.Label(row2, text="Product:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_prod = tk.StringVar()
        cb_prod = ttk.Combobox(row2, textvariable=v_prod, width=22)
        cb_prod.pack(side="left", padx=(2,4))
        add_autocomplete(cb_prod)

        tk.Label(row2, text="Qty:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_qty = tk.StringVar()
        sr_qty_entry = ttk.Entry(row2, textvariable=v_qty, width=7)
        sr_qty_entry.pack(side="left", padx=(2,10))
        tk.Label(row2, text="Rate:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_rate = tk.StringVar()
        sr_rate_entry = ttk.Entry(row2, textvariable=v_rate, width=9)
        sr_rate_entry.pack(side="left", padx=(2,10))
        tk.Label(row2, text="GST%:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_gst = tk.StringVar(value="0")
        sr_gst_cb = ttk.Combobox(row2, textvariable=v_gst, values=["0","5","12","18","28"],
                     width=5, state="readonly")
        sr_gst_cb.pack(side="left", padx=(2,10))
        v_item_total = tk.StringVar(value="Rs.0")
        tk.Label(row2, text="Item Total:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        tk.Label(row2, textvariable=v_item_total, font=("Segoe UI",9,"bold"),
                 bg=C_WHITE, fg=C_ACCENT).pack(side="left", padx=(2,0))

        def recalc(*a):
            try:
                q=float(v_qty.get() or 0); r=float(v_rate.get() or 0); g=float(v_gst.get() or 0)
                t=round(q*r,2); ga=round(t*g/100,2)
                v_item_total.set("Rs."+str(int(t+ga)))
            except: v_item_total.set("Rs.0")
        for vv in (v_qty, v_rate, v_gst): vv.trace_add("write", recalc)

        def on_bill(ev=None):
            bn = v_bill.get().strip()
            if not bn: return
            d = get_db()
            row = d.execute("SELECT party FROM sales WHERE bill_no=?", (bn,)).fetchone()
            prods = [x[0] for x in d.execute(
                "SELECT DISTINCT si.product FROM sale_items si "
                "JOIN sales s ON si.sale_id=s.id WHERE s.bill_no=?", (bn,)).fetchall()]
            d.close()
            if row:
                e_party.config(state="normal"); v_party.set(row["party"]); e_party.config(state="readonly")
            cb_prod["values"] = prods
            if prods: v_prod.set(prods[0]); on_prod()

        def on_prod(ev=None):
            bn = v_bill.get().strip(); pr = v_prod.get().strip()
            if not bn or not pr: return
            d = get_db()
            row = d.execute(
                "SELECT si.rate, si.gst_percent FROM sale_items si "
                "JOIN sales s ON si.sale_id=s.id WHERE s.bill_no=? AND si.product=? LIMIT 1",
                (bn, pr)).fetchone()
            d.close()
            if row: v_rate.set(str(round(row["rate"],2))); v_gst.set(str(int(row["gst_percent"])))

        cb_bill.bind("<<ComboboxSelected>>", on_bill)
        cb_prod.bind("<<ComboboxSelected>>", on_prod)

        # ── Enter Key Navigation: Date→Bill→Reason→Product→Qty→Rate→GST→AddItem ──
        def _sr_focus(w):
            def _go(e):
                target = getattr(w, "_entry", w)
                target.focus_set()
                try: target.select_range(0, "end")
                except: pass
                return "break"
            return _go
        getattr(sr_date_entry, "_entry", sr_date_entry).bind("<Return>", _sr_focus(cb_bill))
        cb_bill.bind("<Return>",       _sr_focus(sr_reason_entry), add="+")
        sr_reason_entry.bind("<Return>", _sr_focus(cb_prod))
        cb_prod.bind("<Return>",       _sr_focus(sr_qty_entry), add="+")
        sr_qty_entry.bind("<Return>",  _sr_focus(sr_rate_entry))
        sr_rate_entry.bind("<Return>", _sr_focus(sr_gst_cb))
        sr_gst_cb.bind("<Return>",     lambda e: add_item())
        tk.Label(form, text="Return Items Cart:", font=("Segoe UI",9,"bold"),
                 bg=C_WHITE, fg="#1A365D").pack(anchor="w", padx=16, pady=(6,2))
        cart_tbl = make_table(form,
            ["#","Product","Qty","Rate","Taxable","GST%","GST Amt","Total","Del"],
            [3, 22, 6, 9, 9, 6, 9, 9, 4])
        cart_tbl.master.master.pack(fill="x", padx=16, pady=(0,2))

        v_cart_total = tk.StringVar(value="Grand Total: Rs.0")
        tk.Label(form, textvariable=v_cart_total, font=("Segoe UI",10,"bold"),
                 bg=C_WHITE, fg=C_ACCENT, anchor="e").pack(fill="x", padx=16, pady=(0,2))

        def render_cart():
            clear_table_rows(cart_tbl)
            grand = 0
            for i, it in enumerate(self._sr_items):
                grand += it["total"]
                bg = C_WHITE if i%2==0 else "#FFF5F5"
                vals = [i+1, it["product"][:20], it["qty"],
                        "Rs."+str(round(it["rate"],2)), "Rs."+str(round(it["taxable"],2)),
                        str(int(it["gst"]))+"%", "Rs."+str(round(it["gst_amt"],2)),
                        "Rs."+str(int(it["total"])), "X"]
                for j, val in enumerate(vals):
                    if j == 8:
                        tk.Button(cart_tbl, text="✕", font=("Segoe UI",9),
                                  bg=bg, fg=C_RED, relief="flat", cursor="hand2", bd=0,
                                  command=lambda idx=i: (self._sr_items.pop(idx), render_cart())
                                  ).grid(row=i+1, column=j, sticky="nsew", padx=1, pady=0)
                    else:
                        tk.Label(cart_tbl, text=str(val), font=("Segoe UI",9),
                                 bg=bg, fg=C_GRAY, anchor="w", padx=4, pady=3
                                 ).grid(row=i+1, column=j, sticky="nsew", padx=1, pady=0)
            v_cart_total.set("Grand Total: Rs." + str(int(grand)))

        def add_item():
            prod = v_prod.get().strip()
            if not prod: messagebox.showerror("Error", "Product select karo!"); return
            if not v_bill.get().strip(): messagebox.showerror("Error", "Pehle Bill No select karo!"); return
            try:
                qty = float(v_qty.get()); rate = float(v_rate.get()); gst_p = float(v_gst.get())
                assert qty > 0
            except: messagebox.showerror("Error", "Qty aur Rate sahi bharo!"); return
            taxable = round(qty*rate, 2); gst_amt = round(taxable*gst_p/100, 2)
            total   = round(taxable+gst_amt, 2)
            hsn = ""
            try:
                c_ = get_db(); r_ = c_.execute("SELECT hsn FROM products WHERE name=?", (prod,)).fetchone(); c_.close()
                if r_: hsn = r_["hsn"] or ""
            except: pass
            self._sr_items.append({
                "product": prod, "hsn": hsn, "qty": qty, "rate": rate,
                "taxable": taxable, "gst": gst_p, "gst_amt": gst_amt, "total": total
            })
            render_cart()
            v_prod.set(""); v_qty.set(""); v_rate.set(""); v_gst.set("0"); v_item_total.set("Rs.0")

        make_btn(row2, "+ Add Item", add_item, bg=C_ACCENT).pack(side="left", padx=(14,8))
        make_btn(row2, "All Clear", lambda: (self._sr_items.clear(), render_cart()), bg=C_GRAY).pack(side="left")

        # ── Save / Print buttons ──────────────────────────────────────────────
        bf = tk.Frame(form, bg=C_WHITE); bf.pack(anchor="w", padx=16, pady=(2,8))

        def do_save(print_after=False):
            if not self._sr_items:
                messagebox.showerror("Error", "Koi item nahi! Pehle + Add Item karo."); return
            bill = v_bill.get().strip(); party = v_party.get().strip()
            if not bill or not party:
                messagebox.showerror("Error", "Bill No aur Customer zaroori hain!"); return
            d = get_db(); saved_nos = []
            for it in self._sr_items:
                n   = d.execute("SELECT COUNT(*) FROM returns WHERE return_type='sale_return'").fetchone()[0]
                rno = "SR-{:03d}".format(n+1)
                d.execute(
                    "INSERT INTO returns(return_no,return_type,return_date,orig_bill,party,"
                    "product,qty,rate,gst_percent,gst_amt,total_amt,reason) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                    (rno, "sale_return", v_date.get(), bill, party,
                     it["product"], it["qty"], it["rate"], it["gst"], it["gst_amt"],
                     it["total"], v_reason.get().strip()))
                d.execute("UPDATE products SET opening_stock=opening_stock+? WHERE name=?",
                          (it["qty"], it["product"]))
                saved_nos.append(rno); d.commit()
            d.close()
            grand_total = sum(it["total"] for it in self._sr_items)
            snap = list(self._sr_items)
            messagebox.showinfo("Saved!",
                "Sale Return saved!\nReturn Nos: " + ", ".join(saved_nos) +
                "\nTotal: Rs." + str(int(grand_total)))
            if print_after:
                _show_return_bill(saved_nos, bill, party, v_date.get(),
                                  v_reason.get().strip(), snap, "Sale Return / Credit Note")
            self._sr_items.clear()
            v_bill.set(""); e_party.config(state="normal"); v_party.set(""); e_party.config(state="readonly")
            v_prod.set(""); v_qty.set(""); v_rate.set(""); v_gst.set("0"); v_reason.set("")
            render_cart(); load_hist()

        def _show_return_bill(return_nos, orig_bill, party_name, ret_date, reason, items, bill_title):
            """Return Bill preview window — InvoiceWin style."""
            win = tk.Toplevel(self.root)
            win.title(bill_title + " — " + ", ".join(return_nos))
            win.state("zoomed"); win.configure(bg=C_WHITE); _apply_logo(win)
            grand_total   = sum(it["total"]   for it in items)
            total_taxable = sum(it["taxable"] for it in items)
            total_gst     = sum(it["gst_amt"] for it in items)

            hdr = tk.Frame(win, bg=C_TOP, height=44); hdr.pack(fill="x"); hdr.pack_propagate(False)
            tk.Label(hdr, text="  Return Bill Preview", font=("Segoe UI",12,"bold"),
                     bg=C_TOP, fg=C_WHITE).pack(side="left", padx=14, pady=4)
            make_btn(hdr, "Print / PDF",
                     lambda: _do_pdf(win, return_nos, orig_bill, party_name, ret_date,
                                     reason, items, grand_total, total_taxable, total_gst, bill_title),
                     bg=C_GREEN).pack(side="right", padx=10, pady=4)
            make_btn(hdr, "Close", win.destroy, bg=C_GRAY).pack(side="right", padx=4, pady=4)

            outer = tk.Frame(win, bg=C_WHITE); outer.pack(fill="both", expand=True)
            canvas = tk.Canvas(outer, bg="#E5E5E5", highlightthickness=0)
            vsb    = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=vsb.set)
            vsb.pack(side="right", fill="y"); canvas.pack(fill="both", expand=True)
            canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>",
                lambda ev: canvas.yview_scroll(-1*(ev.delta//120), "units")))
            canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))
            paper = tk.Frame(canvas, bg=C_WHITE, width=794)
            wid   = canvas.create_window((0, 20), window=paper, anchor="nw")
            canvas.bind("<Configure>", lambda e: (
                canvas.itemconfig(wid, width=min(794, e.width)),
                canvas.configure(scrollregion=canvas.bbox("all"))))
            paper.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

            _sp    = get_shop()
            HDR_BG = "#F2F2F2"
            main   = tk.Frame(paper, bg=C_WHITE, padx=22, pady=16); main.pack(fill="x")

            def cell(par, text, row, col, font_s=9, bold=False, align="w",
                     bg=C_WHITE, fg="#000000", rowspan=1, colspan=1):
                fr = tk.Frame(par, bg=bg, highlightthickness=1, highlightbackground="#999999")
                fr.grid(row=row, column=col, rowspan=rowspan, columnspan=colspan,
                        sticky="nsew", padx=0, pady=0)
                tk.Label(fr, text=str(text),
                         font=("Arial", font_s, "bold" if bold else "normal"),
                         bg=bg, fg=fg, anchor=align, padx=4, pady=3).pack(fill="both", expand=True)

            # Title bar
            title_f = tk.Frame(main, bg=HDR_BG, highlightthickness=1, highlightbackground="#999999")
            title_f.pack(fill="x")
            tk.Label(title_f, text=bill_title, font=("Arial",12,"bold"),
                     bg=HDR_BG, anchor="center", pady=5).pack(fill="x")

            # Shop + Return meta
            info = tk.Frame(main, bg=C_WHITE); info.pack(fill="x")
            info.columnconfigure(0, weight=3); info.columnconfigure(1, weight=2)

            lf = tk.Frame(info, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
            lf.grid(row=0, column=0, sticky="nsew")
            tk.Label(lf, text=_sp["name"], font=("Arial",10,"bold"),
                     bg=C_WHITE, anchor="w", padx=6, pady=3).pack(fill="x")
            tk.Label(lf, text=(_sp["address"] + ", " + _sp["city"]).strip(", "),
                     font=("Arial",8), bg=C_WHITE, anchor="w", padx=6).pack(fill="x")
            if _sp["gstin"]: tk.Label(lf, text="GSTIN: "+_sp["gstin"],
                     font=("Arial",8), bg=C_WHITE, anchor="w", padx=6).pack(fill="x")
            if _sp["mobile"]: tk.Label(lf, text="Mob: "+_sp["mobile"],
                     font=("Arial",8), bg=C_WHITE, anchor="w", padx=6, pady=(0,4)).pack(fill="x")

            rf = tk.Frame(info, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
            rf.grid(row=0, column=1, sticky="nsew")
            for lbl, val in [
                    ("Return No(s):", ", ".join(return_nos)),
                    ("Date:",         fmt_date(ret_date)),
                    ("Orig Bill No:", orig_bill),
                    ("Customer:",     party_name),
                    ("Reason:",       reason or "—")]:
                rf2 = tk.Frame(rf, bg=C_WHITE); rf2.pack(fill="x", padx=6)
                tk.Label(rf2, text=lbl, font=("Arial",8,"bold"), bg=C_WHITE,
                         width=14, anchor="w").pack(side="left")
                tk.Label(rf2, text=val, font=("Arial",8),
                         bg=C_WHITE, anchor="w").pack(side="left")
            tk.Label(rf, text="", bg=C_WHITE, pady=2).pack()

            # Items table
            it_tbl = tk.Frame(main, bg=C_WHITE); it_tbl.pack(fill="x", pady=(8,0))
            COLS  = ["Sl.","Product","HSN","Qty","Rate","Taxable","GST%","GST Amt","Total"]
            WCOLS = [3, 22, 9, 6, 9, 9, 6, 9, 9]
            for i, c in enumerate(COLS): it_tbl.columnconfigure(i, weight=WCOLS[i])
            for i, c in enumerate(COLS):
                hf2 = tk.Frame(it_tbl, bg=HDR_BG, highlightthickness=1, highlightbackground="#999999")
                hf2.grid(row=0, column=i, sticky="nsew")
                tk.Label(hf2, text=c, font=("Arial",8,"bold"), bg=HDR_BG,
                         anchor="center", padx=3, pady=4, wraplength=55, justify="center").pack(fill="both")

            n_rows = max(len(items), 5)
            for r in range(n_rows):
                bg_r = C_WHITE if r%2==0 else "#FFF5F5"
                if r < len(items):
                    it = items[r]
                    row_vals = [r+1, it["product"], it.get("hsn",""), it["qty"],
                                str(round(it["rate"],2)), str(round(it["taxable"],2)),
                                str(int(it["gst"]))+"%",  str(round(it["gst_amt"],2)),
                                str(int(it["total"]))]
                else:
                    row_vals = ["","","","","","","","",""]
                for c, val in enumerate(row_vals):
                    bf2 = tk.Frame(it_tbl, bg=bg_r, highlightthickness=1, highlightbackground="#999999")
                    bf2.grid(row=r+1, column=c, sticky="nsew")
                    anc = "center" if c in [0,3,4,5,6,7,8] else "w"
                    tk.Label(bf2, text=str(val), font=("Arial",9),
                             bg=bg_r, anchor=anc, padx=3, pady=3).pack(fill="both")

            # Total row
            tot_r = n_rows+1
            for c in range(9):
                tf2 = tk.Frame(it_tbl, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
                tf2.grid(row=tot_r, column=c, sticky="nsew")
                if c == 1:
                    tk.Label(tf2, text="TOTAL", font=("Arial",9,"bold"), bg=C_WHITE,
                             anchor="w", padx=4, pady=4).pack(fill="both")
                elif c == 5:
                    tk.Label(tf2, text=str(round(total_taxable,2)), font=("Arial",9,"bold"),
                             bg=C_WHITE, anchor="center", padx=3, pady=4).pack(fill="both")
                elif c == 7:
                    tk.Label(tf2, text=str(round(total_gst,2)), font=("Arial",9,"bold"),
                             bg=C_WHITE, anchor="center", padx=3, pady=4).pack(fill="both")
                elif c == 8:
                    tk.Label(tf2, text=str(int(grand_total)), font=("Arial",9,"bold"),
                             bg=C_WHITE, anchor="center", padx=3, pady=4).pack(fill="both")
                else:
                    tk.Label(tf2, text="", bg=C_WHITE, pady=4).pack()

            # Amount in words + grand total card
            wf = tk.Frame(main, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
            wf.pack(fill="x")
            wf.columnconfigure(0, weight=3); wf.columnconfigure(1, weight=1)
            tk.Label(wf, text="Amount in Words: " + num_to_words(grand_total),
                     font=("Arial",8), bg=C_WHITE, anchor="w", padx=8, pady=5,
                     wraplength=500, justify="left").grid(row=0, column=0, sticky="w")
            gcard = tk.Frame(wf, bg="#F0FFF4", highlightthickness=1, highlightbackground="#999999")
            gcard.grid(row=0, column=1, sticky="nsew")
            tk.Label(gcard, text="Grand Total\nRs." + str(int(grand_total)),
                     font=("Arial",11,"bold"), bg="#F0FFF4", fg=C_GREEN,
                     anchor="center", justify="center", pady=6).pack(fill="both", expand=True)

            # Signatory
            sig_f = tk.Frame(main, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
            sig_f.pack(fill="x")
            sig_f.columnconfigure(0, weight=1); sig_f.columnconfigure(1, weight=1)
            tk.Label(sig_f, text="Note: Yeh Credit Note hai. Original Bill: " + orig_bill,
                     font=("Arial",8,"italic"), bg=C_WHITE,
                     anchor="w", padx=8, pady=4).grid(row=0, column=0, sticky="w")
            sgn = tk.Frame(sig_f, bg=C_WHITE); sgn.grid(row=0, column=1, sticky="nsew")
            tk.Label(sgn, text="For  " + _sp["name"], font=("Arial",9,"bold"),
                     bg=C_WHITE, anchor="e", padx=10, pady=4).pack(fill="x")
            tk.Label(sgn, text="Authorised Signatory",
                     font=("Arial",8), bg=C_WHITE, anchor="e", padx=10, pady=(0,12)).pack(fill="x")

        def _do_pdf(win_par, return_nos, orig_bill, party_name, ret_date, reason,
                    items, grand_total, total_taxable, total_gst, bill_title):
            import os, subprocess, sys
            from tkinter import filedialog
            save_path = filedialog.asksaveasfilename(
                parent=win_par, title="Return Bill PDF Save Karo",
                defaultextension=".pdf",
                initialfile="_".join(return_nos) + ".pdf",
                filetypes=[("PDF File","*.pdf")])
            if not save_path: return
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.units import mm
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import ParagraphStyle
                from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
                _sp    = get_shop()
                PAGE_W = 190*mm; BDR = colors.black; GRY = colors.HexColor("#f2f2f2")
                def ps(name, size=8, bold=False, align=TA_LEFT):
                    return ParagraphStyle(name, fontSize=size,
                        fontName="Helvetica-Bold" if bold else "Helvetica",
                        alignment=align, leading=size+3)
                doc = SimpleDocTemplate(save_path, pagesize=A4,
                    leftMargin=10*mm, rightMargin=10*mm, topMargin=8*mm, bottomMargin=8*mm)
                BASE = [("BOX",(0,0),(-1,-1),0.5,BDR),("INNERGRID",(0,0),(-1,-1),0.3,BDR),
                        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
                        ("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3),
                        ("VALIGN",(0,0),(-1,-1),"TOP")]
                def T(data, cw, style_cmds, repeat=0):
                    t = Table(data, colWidths=cw, style=TableStyle(style_cmds), repeatRows=repeat)
                    t.spaceAfter = 0; t.spaceBefore = 0; return t
                story = []
                story.append(T([[Paragraph(bill_title, ps("t",11,True,TA_CENTER))]],
                    [PAGE_W], BASE+[("BACKGROUND",(0,0),(-1,-1),GRY)]))
                story.append(Spacer(1,1*mm))
                shop_lines = "<b>" + _sp["name"] + "</b><br/>"
                shop_lines += (_sp["address"] + ", " + _sp["city"]).strip(", ")
                if _sp["gstin"]: shop_lines += "<br/>GSTIN: " + _sp["gstin"]
                if _sp["mobile"]: shop_lines += "  Mob: " + _sp["mobile"]
                info_lines  = "<b>Return No(s):</b> " + ", ".join(return_nos)
                info_lines += "<br/><b>Date:</b> " + fmt_date(ret_date)
                info_lines += "<br/><b>Orig Bill:</b> " + orig_bill
                info_lines += "<br/><b>Customer:</b> " + party_name
                info_lines += "<br/><b>Reason:</b> " + (reason or "—")
                story.append(T([[Paragraph(shop_lines, ps("s",8)),
                                  Paragraph(info_lines, ps("i",8))]],
                               [PAGE_W*0.55, PAGE_W*0.45], BASE))
                story.append(Spacer(1,2*mm))
                hdr_row = [Paragraph(c, ps("h",8,True,TA_CENTER))
                           for c in ["Sl.","Product","HSN","Qty","Rate","Taxable","GST%","GST Amt","Total"]]
                col_w = [8*mm, 55*mm, 20*mm, 12*mm, 18*mm, 20*mm, 12*mm, 20*mm, 20*mm]
                data  = [hdr_row]
                for i, it in enumerate(items):
                    data.append([
                        Paragraph(str(i+1),              ps("d",8,False,TA_CENTER)),
                        Paragraph(it["product"],          ps("d",8)),
                        Paragraph(it.get("hsn",""),       ps("d",8,False,TA_CENTER)),
                        Paragraph(str(it["qty"]),         ps("d",8,False,TA_CENTER)),
                        Paragraph(str(round(it["rate"],2)),     ps("d",8,False,TA_CENTER)),
                        Paragraph(str(round(it["taxable"],2)),  ps("d",8,False,TA_CENTER)),
                        Paragraph(str(int(it["gst"]))+"%",      ps("d",8,False,TA_CENTER)),
                        Paragraph(str(round(it["gst_amt"],2)),  ps("d",8,False,TA_CENTER)),
                        Paragraph(str(int(it["total"])),        ps("d",8,False,TA_RIGHT)),
                    ])
                data.append([
                    Paragraph("",""), Paragraph("<b>TOTAL</b>", ps("d",9,True)), "","","",
                    Paragraph("<b>"+str(round(total_taxable,2))+"</b>", ps("d",8,True,TA_CENTER)), "",
                    Paragraph("<b>"+str(round(total_gst,2))+"</b>",    ps("d",8,True,TA_CENTER)),
                    Paragraph("<b>"+str(int(grand_total))+"</b>",      ps("d",9,True,TA_RIGHT))])
                tbl_style = BASE + [
                    ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#1A365D")),
                    ("TEXTCOLOR",(0,0),(-1,0),  colors.white),
                    ("ROWBACKGROUNDS",(0,1),(-1,-2), [colors.white, colors.HexColor("#FFF5F5")]),
                    ("BACKGROUND",(0,-1),(-1,-1), colors.HexColor("#F0FFF4")),
                    ("FONTNAME",(0,-1),(-1,-1),   "Helvetica-Bold"),
                ]
                story.append(T(data, col_w, tbl_style, repeat=1))
                story.append(Spacer(1,2*mm))
                story.append(T([[
                    Paragraph("Amount in Words: <b>" + num_to_words(grand_total) + "</b>", ps("w",8)),
                    Paragraph("<b>Grand Total: Rs." + str(int(grand_total)) + "</b>", ps("gt",10,True,TA_RIGHT))
                ]], [PAGE_W*0.65, PAGE_W*0.35], BASE))
                story.append(Spacer(1,2*mm))
                story.append(T([[
                    Paragraph("Note: Yeh Credit Note hai. Original Bill: " + orig_bill, ps("n",8)),
                    Paragraph("For " + _sp["name"] + "<br/><br/>Authorised Signatory",  ps("sg",8,False,TA_RIGHT))
                ]], [PAGE_W*0.6, PAGE_W*0.4], BASE))
                doc.build(story)
                if sys.platform == "win32":   os.startfile(save_path)
                elif sys.platform == "darwin": subprocess.run(["open", save_path])
                else:                          subprocess.run(["xdg-open", save_path])
                messagebox.showinfo("PDF Ready!", "Return Bill PDF save ho gaya!\n" + save_path)
            except ImportError:
                messagebox.showerror("Library Missing",
                    "reportlab install nahi hai.\nCMD mein chalao:\npip install reportlab")
            except Exception as e:
                messagebox.showerror("PDF Error", str(e))

        make_btn(bf, "Save Return",         lambda: do_save(False), bg=C_GREEN).pack(side="left", padx=(0,8))
        make_btn(bf, "Save + Print Bill",   lambda: do_save(True),  bg="#6B46C1").pack(side="left")

        # ── History ───────────────────────────────────────────────────────────
        tk.Label(p, text="Sale Return History", font=("Segoe UI",10,"bold"),
                 bg=C_BG, fg="#1A365D").pack(anchor="w", pady=(8,4))
        tbl = make_table(p,
            ["Return No","Date","Bill No","Customer","Product","Qty","Total","Reason","Print"],
            [11,11,13,16,18,7,10,14,6])

        def load_hist():
            clear_table_rows(tbl)
            d = get_db()
            rows = [dict(r) for r in d.execute(
                "SELECT * FROM returns WHERE return_type='sale_return' ORDER BY return_date DESC"
            ).fetchall()]
            d.close()
            for i, r in enumerate(rows):
                bg = C_WHITE if i%2==0 else "#F7FAFC"
                vals = [r["return_no"], fmt_date(r["return_date"]), r["orig_bill"],
                        r["party"][:14], r["product"][:16], str(r["qty"]),
                        "Rs."+str(int(r["total_amt"])), (r["reason"] or "")[:14], ""]
                for j, val in enumerate(vals):
                    if j == 8:
                        tk.Button(tbl, text="Print", font=("Segoe UI",7),
                                  bg=C_ACCENT, fg=C_WHITE, relief="flat", cursor="hand2", bd=0,
                                  command=lambda rr=dict(r): _reprint(rr)
                                  ).grid(row=i+1, column=j, sticky="nsew", padx=1, pady=1)
                    else:
                        tk.Label(tbl, text=str(val), font=("Segoe UI",9),
                                 bg=bg, fg=C_GRAY, anchor="w", padx=4, pady=3
                                 ).grid(row=i+1, column=j, sticky="nsew", padx=1)

        def _reprint(rrow):
            items_r = [{
                "product": rrow["product"], "hsn": "",
                "qty": rrow["qty"], "rate": rrow["rate"],
                "taxable": rrow["qty"]*rrow["rate"],
                "gst": rrow["gst_percent"], "gst_amt": rrow["gst_amt"], "total": rrow["total_amt"]
            }]
            _show_return_bill([rrow["return_no"]], rrow["orig_bill"], rrow["party"],
                              rrow["return_date"], rrow["reason"] or "",
                              items_r, "Sale Return / Credit Note")
        load_hist()

    # ══════════════════════════════════════════════════════════════════════════
    #  PURCHASE RETURN
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_purreturn(self):
        """Purchase Return — Multiple Products + Return Bill Print"""
        import datetime as _dt
        db = get_db()
        db.execute("CREATE TABLE IF NOT EXISTS returns (id INTEGER PRIMARY KEY AUTOINCREMENT, return_no TEXT UNIQUE NOT NULL, return_type TEXT NOT NULL, return_date TEXT NOT NULL, orig_bill TEXT NOT NULL, party TEXT NOT NULL, product TEXT NOT NULL, qty REAL DEFAULT 0, rate REAL DEFAULT 0, gst_percent REAL DEFAULT 0, gst_amt REAL DEFAULT 0, total_amt REAL DEFAULT 0, reason TEXT DEFAULT '', created_at TEXT DEFAULT CURRENT_TIMESTAMP)")
        db.commit(); db.close()

        self._pr_items = []

        p = tk.Frame(self.content, bg=C_BG, padx=10, pady=3)
        p.pack(fill="both", expand=True)
        section_title(p, "Purchase Return")

        form = tk.Frame(p, bg=C_WHITE, highlightthickness=1, highlightbackground=C_BORDER)
        form.pack(fill="x", pady=(0,10))
        tk.Label(form, text="Supplier Return Entry — Ek ya Zyada Products",
                 font=("Segoe UI",10,"bold"), bg=C_WHITE, fg="#1A365D").pack(anchor="w", padx=16, pady=(10,6))

        row1 = tk.Frame(form, bg=C_WHITE); row1.pack(fill="x", padx=16, pady=4)
        tk.Label(row1, text="Date:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_date = tk.StringVar(value=str(_dt.date.today()))
        pr_date_entry = make_date_entry(row1, v_date, width=13)
        pr_date_entry.pack(side="left", padx=(2,14))
        tk.Label(row1, text="Purchase Bill:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_bill = tk.StringVar()
        db2 = get_db()
        bills = [r[0] for r in db2.execute("SELECT bill_no FROM purchases ORDER BY bill_date DESC").fetchall()]
        db2.close()
        cb_bill = ttk.Combobox(row1, textvariable=v_bill, values=bills, width=16)
        cb_bill.pack(side="left", padx=(2,14))
        add_autocomplete(cb_bill)
        tk.Label(row1, text="Supplier:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_party = tk.StringVar()
        e_party = ttk.Entry(row1, textvariable=v_party, width=20, state="readonly")
        e_party.pack(side="left", padx=(2,14))
        tk.Label(row1, text="Reason:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_reason = tk.StringVar()
        pr_reason_entry = ttk.Entry(row1, textvariable=v_reason, width=22)
        pr_reason_entry.pack(side="left", padx=(2,0))

        row2 = tk.Frame(form, bg=C_WHITE); row2.pack(fill="x", padx=16, pady=4)
        tk.Label(row2, text="Product:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_prod = tk.StringVar()
        cb_prod = ttk.Combobox(row2, textvariable=v_prod, width=22)
        cb_prod.pack(side="left", padx=(2,4))
        add_autocomplete(cb_prod)

        tk.Label(row2, text="Qty:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_qty = tk.StringVar()
        pr_qty_entry = ttk.Entry(row2, textvariable=v_qty, width=7)
        pr_qty_entry.pack(side="left", padx=(2,10))
        tk.Label(row2, text="Rate:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_rate = tk.StringVar()
        pr_rate_entry = ttk.Entry(row2, textvariable=v_rate, width=9)
        pr_rate_entry.pack(side="left", padx=(2,10))
        tk.Label(row2, text="GST%:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_gst = tk.StringVar(value="0")
        pr_gst_cb = ttk.Combobox(row2, textvariable=v_gst, values=["0","5","12","18","28"],
                     width=5, state="readonly")
        pr_gst_cb.pack(side="left", padx=(2,10))
        v_item_total = tk.StringVar(value="Rs.0")
        tk.Label(row2, text="Item Total:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        tk.Label(row2, textvariable=v_item_total, font=("Segoe UI",9,"bold"),
                 bg=C_WHITE, fg=C_AMBER).pack(side="left", padx=(2,14))
        def recalc(*a):
            try:
                q=float(v_qty.get() or 0); r=float(v_rate.get() or 0); g=float(v_gst.get() or 0)
                t=round(q*r,2); ga=round(t*g/100,2)
                v_item_total.set("Rs."+str(int(t+ga)))
            except: v_item_total.set("Rs.0")
        for vv in (v_qty, v_rate, v_gst): vv.trace_add("write", recalc)

        def on_bill(ev=None):
            bn = v_bill.get().strip()
            if not bn: return
            d = get_db()
            row   = d.execute("SELECT party FROM purchases WHERE bill_no=?", (bn,)).fetchone()
            prods = [x[0] for x in d.execute(
                "SELECT DISTINCT pi.product FROM purchase_items pi "
                "JOIN purchases pu ON pi.purchase_id=pu.id WHERE pu.bill_no=?", (bn,)).fetchall()]
            d.close()
            if row:
                e_party.config(state="normal"); v_party.set(row["party"]); e_party.config(state="readonly")
            cb_prod["values"] = prods
            if prods: v_prod.set(prods[0]); on_prod()

        def on_prod(ev=None):
            bn = v_bill.get().strip(); pr = v_prod.get().strip()
            if not bn or not pr: return
            d = get_db()
            row = d.execute(
                "SELECT pi.rate, pi.gst_percent FROM purchase_items pi "
                "JOIN purchases pu ON pi.purchase_id=pu.id WHERE pu.bill_no=? AND pi.product=? LIMIT 1",
                (bn, pr)).fetchone()
            d.close()
            if row: v_rate.set(str(round(row["rate"],2))); v_gst.set(str(int(row["gst_percent"])))

        cb_bill.bind("<<ComboboxSelected>>", on_bill)
        cb_prod.bind("<<ComboboxSelected>>", on_prod)

        # ── Enter Key Navigation: Date→Bill→Reason→Product→Qty→Rate→GST→AddItem ──
        def _pr_focus(w):
            def _go(e):
                target = getattr(w, "_entry", w)
                target.focus_set()
                try: target.select_range(0, "end")
                except: pass
                return "break"
            return _go
        getattr(pr_date_entry, "_entry", pr_date_entry).bind("<Return>", _pr_focus(cb_bill))
        cb_bill.bind("<Return>",        _pr_focus(pr_reason_entry), add="+")
        pr_reason_entry.bind("<Return>", _pr_focus(cb_prod))
        cb_prod.bind("<Return>",        _pr_focus(pr_qty_entry), add="+")
        pr_qty_entry.bind("<Return>",   _pr_focus(pr_rate_entry))
        pr_rate_entry.bind("<Return>",  _pr_focus(pr_gst_cb))
        pr_gst_cb.bind("<Return>",      lambda e: add_item())

        tk.Label(form, text="Return Items Cart:", font=("Segoe UI",9,"bold"),
                 bg=C_WHITE, fg="#1A365D").pack(anchor="w", padx=16, pady=(6,2))
        cart_tbl = make_table(form,
            ["#","Product","Qty","Rate","Taxable","GST%","GST Amt","Total","Del"],
            [3, 22, 6, 9, 9, 6, 9, 9, 4])
        cart_tbl.master.master.pack(fill="x", padx=16, pady=(0,4))

        v_cart_total = tk.StringVar(value="Grand Total: Rs.0")
        tk.Label(form, textvariable=v_cart_total, font=("Segoe UI",10,"bold"),
                 bg=C_WHITE, fg=C_AMBER, anchor="e").pack(fill="x", padx=16, pady=(0,4))

        def render_cart():
            clear_table_rows(cart_tbl)
            grand = 0
            for i, it in enumerate(self._pr_items):
                grand += it["total"]
                bg = C_WHITE if i%2==0 else "#FFFBF0"
                vals = [i+1, it["product"][:20], it["qty"],
                        "Rs."+str(round(it["rate"],2)), "Rs."+str(round(it["taxable"],2)),
                        str(int(it["gst"]))+"%", "Rs."+str(round(it["gst_amt"],2)),
                        "Rs."+str(int(it["total"])), "X"]
                for j, val in enumerate(vals):
                    if j == 8:
                        tk.Button(cart_tbl, text="X", font=("Segoe UI",9),
                                  bg=bg, fg=C_RED, relief="flat", cursor="hand2", bd=0,
                                  command=lambda idx=i: (self._pr_items.pop(idx), render_cart())
                                  ).grid(row=i+1, column=j, sticky="nsew", padx=1, pady=0)
                    else:
                        tk.Label(cart_tbl, text=str(val), font=("Segoe UI",9),
                                 bg=bg, fg=C_GRAY, anchor="w", padx=4, pady=3
                                 ).grid(row=i+1, column=j, sticky="nsew", padx=1, pady=0)
            v_cart_total.set("Grand Total: Rs." + str(int(grand)))

        def add_item():
            prod = v_prod.get().strip()
            if not prod: messagebox.showerror("Error","Product select karo!"); return
            if not v_bill.get().strip(): messagebox.showerror("Error","Pehle Bill No select karo!"); return
            try:
                qty = float(v_qty.get()); rate = float(v_rate.get()); gst_p = float(v_gst.get())
                assert qty > 0
            except: messagebox.showerror("Error","Qty aur Rate sahi bharo!"); return
            taxable = round(qty*rate,2); gst_amt = round(taxable*gst_p/100,2)
            total   = round(taxable+gst_amt,2)
            hsn = ""
            try:
                c_ = get_db(); r_ = c_.execute("SELECT hsn FROM products WHERE name=?", (prod,)).fetchone(); c_.close()
                if r_: hsn = r_["hsn"] or ""
            except: pass
            self._pr_items.append({
                "product": prod, "hsn": hsn, "qty": qty, "rate": rate,
                "taxable": taxable, "gst": gst_p, "gst_amt": gst_amt, "total": total
            })
            render_cart()
            v_prod.set(""); v_qty.set(""); v_rate.set(""); v_gst.set("0"); v_item_total.set("Rs.0")

        make_btn(row2, "+ Add Item",  add_item,                                      bg=C_AMBER).pack(side="left", padx=(0,8))
        make_btn(row2, "All Clear",  lambda: (self._pr_items.clear(), render_cart()), bg=C_GRAY ).pack(side="left")

        bf = tk.Frame(form, bg=C_WHITE); bf.pack(anchor="w", padx=16, pady=(4,12))

        def do_save(print_after=False):
            if not self._pr_items:
                messagebox.showerror("Error","Koi item nahi! Pehle + Add Item karo."); return
            bill = v_bill.get().strip(); party = v_party.get().strip()
            if not bill or not party:
                messagebox.showerror("Error","Bill No aur Supplier zaroori hain!"); return
            d = get_db(); saved_nos = []
            for it in self._pr_items:
                n   = d.execute("SELECT COUNT(*) FROM returns WHERE return_type='pur_return'").fetchone()[0]
                rno = "PR-{:03d}".format(n+1)
                d.execute(
                    "INSERT INTO returns(return_no,return_type,return_date,orig_bill,party,"
                    "product,qty,rate,gst_percent,gst_amt,total_amt,reason) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                    (rno, "pur_return", v_date.get(), bill, party,
                     it["product"], it["qty"], it["rate"], it["gst"], it["gst_amt"],
                     it["total"], v_reason.get().strip()))
                d.execute("UPDATE products SET opening_stock=opening_stock-? WHERE name=?",
                          (it["qty"], it["product"]))
                saved_nos.append(rno); d.commit()
            d.close()
            grand_total = sum(it["total"] for it in self._pr_items)
            snap = list(self._pr_items)
            messagebox.showinfo("Saved!",
                "Purchase Return saved!\nReturn Nos: " + ", ".join(saved_nos) +
                "\nTotal: Rs." + str(int(grand_total)))
            if print_after:
                _show_return_bill(saved_nos, bill, party, v_date.get(),
                                  v_reason.get().strip(), snap, "Purchase Return / Debit Note")
            self._pr_items.clear()
            v_bill.set(""); e_party.config(state="normal"); v_party.set(""); e_party.config(state="readonly")
            v_prod.set(""); v_qty.set(""); v_rate.set(""); v_gst.set("0"); v_reason.set("")
            render_cart(); load_hist()

        def _show_return_bill(return_nos, orig_bill, party_name, ret_date, reason, items, bill_title):
            win = tk.Toplevel(self.root)
            win.title(bill_title + " — " + ", ".join(return_nos))
            win.state("zoomed"); win.configure(bg=C_WHITE); _apply_logo(win)
            grand_total   = sum(it["total"]   for it in items)
            total_taxable = sum(it["taxable"] for it in items)
            total_gst     = sum(it["gst_amt"] for it in items)
            HDR_BG = "#FFF8E7"
            hdr = tk.Frame(win, bg="#7B6000", height=44); hdr.pack(fill="x"); hdr.pack_propagate(False)
            tk.Label(hdr, text="  " + bill_title, font=("Segoe UI",12,"bold"),
                     bg="#7B6000", fg=C_WHITE).pack(side="left", padx=14, pady=4)
            make_btn(hdr, "Print / PDF",
                     lambda: _do_pdf(win, return_nos, orig_bill, party_name, ret_date,
                                     reason, items, grand_total, total_taxable, total_gst, bill_title),
                     bg=C_GREEN).pack(side="right", padx=10, pady=4)
            make_btn(hdr, "Close", win.destroy, bg=C_GRAY).pack(side="right", padx=4, pady=4)
            outer  = tk.Frame(win, bg=C_WHITE); outer.pack(fill="both", expand=True)
            canvas = tk.Canvas(outer, bg="#E5E5E5", highlightthickness=0)
            vsb    = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=vsb.set)
            vsb.pack(side="right", fill="y"); canvas.pack(fill="both", expand=True)
            canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>",
                lambda ev: canvas.yview_scroll(-1*(ev.delta//120), "units")))
            canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))
            paper = tk.Frame(canvas, bg=C_WHITE, width=794)
            wid   = canvas.create_window((0, 20), window=paper, anchor="nw")
            canvas.bind("<Configure>", lambda e: (
                canvas.itemconfig(wid, width=min(794, e.width)),
                canvas.configure(scrollregion=canvas.bbox("all"))))
            paper.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            _sp   = get_shop()
            main  = tk.Frame(paper, bg=C_WHITE, padx=22, pady=16); main.pack(fill="x")
            title_f = tk.Frame(main, bg=HDR_BG, highlightthickness=1, highlightbackground="#999999"); title_f.pack(fill="x")
            tk.Label(title_f, text=bill_title, font=("Arial",12,"bold"), bg=HDR_BG, anchor="center", pady=5).pack(fill="x")
            info = tk.Frame(main, bg=C_WHITE); info.pack(fill="x")
            info.columnconfigure(0, weight=3); info.columnconfigure(1, weight=2)
            lf = tk.Frame(info, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999"); lf.grid(row=0, column=0, sticky="nsew")
            tk.Label(lf, text=_sp["name"], font=("Arial",10,"bold"), bg=C_WHITE, anchor="w", padx=6, pady=3).pack(fill="x")
            tk.Label(lf, text=(_sp["address"]+", "+_sp["city"]).strip(", "), font=("Arial",8), bg=C_WHITE, anchor="w", padx=6).pack(fill="x")
            if _sp["gstin"]:  tk.Label(lf, text="GSTIN: "+_sp["gstin"],  font=("Arial",8), bg=C_WHITE, anchor="w", padx=6).pack(fill="x")
            if _sp["mobile"]: tk.Label(lf, text="Mob: " +_sp["mobile"],  font=("Arial",8), bg=C_WHITE, anchor="w", padx=6, pady=(0,4)).pack(fill="x")
            rf = tk.Frame(info, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999"); rf.grid(row=0, column=1, sticky="nsew")
            for lbl, val in [("Return No(s):", ", ".join(return_nos)),("Date:", fmt_date(ret_date)),
                              ("Orig Bill:", orig_bill),("Supplier:", party_name),("Reason:", reason or "—")]:
                rf2 = tk.Frame(rf, bg=C_WHITE); rf2.pack(fill="x", padx=6)
                tk.Label(rf2, text=lbl, font=("Arial",8,"bold"), bg=C_WHITE, width=14, anchor="w").pack(side="left")
                tk.Label(rf2, text=val, font=("Arial",8), bg=C_WHITE, anchor="w").pack(side="left")
            tk.Label(rf, text="", bg=C_WHITE, pady=2).pack()
            it_tbl = tk.Frame(main, bg=C_WHITE); it_tbl.pack(fill="x", pady=(8,0))
            COLS  = ["Sl.","Product","HSN","Qty","Rate","Taxable","GST%","GST Amt","Total"]
            WCOLS = [3, 22, 9, 6, 9, 9, 6, 9, 9]
            for i, c in enumerate(COLS): it_tbl.columnconfigure(i, weight=WCOLS[i])
            for i, c in enumerate(COLS):
                hf2 = tk.Frame(it_tbl, bg=HDR_BG, highlightthickness=1, highlightbackground="#999999"); hf2.grid(row=0, column=i, sticky="nsew")
                tk.Label(hf2, text=c, font=("Arial",8,"bold"), bg=HDR_BG, anchor="center", padx=3, pady=4, wraplength=55, justify="center").pack(fill="both")
            n_rows = max(len(items), 5)
            for r in range(n_rows):
                bg_r = C_WHITE if r%2==0 else "#FFFBF0"
                if r < len(items):
                    it = items[r]
                    row_vals = [r+1, it["product"], it.get("hsn",""), it["qty"],
                                str(round(it["rate"],2)), str(round(it["taxable"],2)),
                                str(int(it["gst"]))+"%", str(round(it["gst_amt"],2)), str(int(it["total"]))]
                else:
                    row_vals = ["","","","","","","","",""]
                for c, val in enumerate(row_vals):
                    bf3 = tk.Frame(it_tbl, bg=bg_r, highlightthickness=1, highlightbackground="#999999"); bf3.grid(row=r+1, column=c, sticky="nsew")
                    anc = "center" if c in [0,3,4,5,6,7,8] else "w"
                    tk.Label(bf3, text=str(val), font=("Arial",9), bg=bg_r, anchor=anc, padx=3, pady=3).pack(fill="both")
            tot_r = n_rows+1
            for c in range(9):
                tf2 = tk.Frame(it_tbl, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999"); tf2.grid(row=tot_r, column=c, sticky="nsew")
                if   c==1: tk.Label(tf2, text="TOTAL", font=("Arial",9,"bold"), bg=C_WHITE, anchor="w", padx=4, pady=4).pack(fill="both")
                elif c==5: tk.Label(tf2, text=str(round(total_taxable,2)), font=("Arial",9,"bold"), bg=C_WHITE, anchor="center", padx=3, pady=4).pack(fill="both")
                elif c==7: tk.Label(tf2, text=str(round(total_gst,2)),     font=("Arial",9,"bold"), bg=C_WHITE, anchor="center", padx=3, pady=4).pack(fill="both")
                elif c==8: tk.Label(tf2, text=str(int(grand_total)),       font=("Arial",9,"bold"), bg=C_WHITE, anchor="center", padx=3, pady=4).pack(fill="both")
                else:      tk.Label(tf2, text="", bg=C_WHITE, pady=4).pack()
            wf2 = tk.Frame(main, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999"); wf2.pack(fill="x")
            wf2.columnconfigure(0, weight=3); wf2.columnconfigure(1, weight=1)
            tk.Label(wf2, text="Amount in Words: " + num_to_words(grand_total),
                     font=("Arial",8), bg=C_WHITE, anchor="w", padx=8, pady=5, wraplength=500, justify="left").grid(row=0, column=0, sticky="w")
            gcard2 = tk.Frame(wf2, bg="#FFFDE7", highlightthickness=1, highlightbackground="#999999"); gcard2.grid(row=0, column=1, sticky="nsew")
            tk.Label(gcard2, text="Grand Total\nRs." + str(int(grand_total)),
                     font=("Arial",11,"bold"), bg="#FFFDE7", fg="#7B6000", anchor="center", justify="center", pady=6).pack(fill="both", expand=True)
            sf2 = tk.Frame(main, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999"); sf2.pack(fill="x")
            sf2.columnconfigure(0, weight=1); sf2.columnconfigure(1, weight=1)
            tk.Label(sf2, text="Note: Yeh Debit Note hai. Original Bill: " + orig_bill,
                     font=("Arial",8,"italic"), bg=C_WHITE, anchor="w", padx=8, pady=4).grid(row=0, column=0, sticky="w")
            sgn2 = tk.Frame(sf2, bg=C_WHITE); sgn2.grid(row=0, column=1, sticky="nsew")
            tk.Label(sgn2, text="For  " + _sp["name"], font=("Arial",9,"bold"), bg=C_WHITE, anchor="e", padx=10, pady=4).pack(fill="x")
            tk.Label(sgn2, text="Authorised Signatory", font=("Arial",8), bg=C_WHITE, anchor="e", padx=10, pady=(0,12)).pack(fill="x")

        def _do_pdf(win_par, return_nos, orig_bill, party_name, ret_date, reason,
                    items, grand_total, total_taxable, total_gst, bill_title):
            import os, subprocess, sys
            from tkinter import filedialog
            save_path = filedialog.asksaveasfilename(
                parent=win_par, title="Return Bill PDF Save Karo",
                defaultextension=".pdf",
                initialfile="_".join(return_nos)+".pdf",
                filetypes=[("PDF File","*.pdf")])
            if not save_path: return
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.units import mm
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import ParagraphStyle
                from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
                _sp    = get_shop()
                PAGE_W = 190*mm; BDR = colors.black; GRY = colors.HexColor("#fff8e7")
                def ps(name, size=8, bold=False, align=TA_LEFT):
                    return ParagraphStyle(name, fontSize=size,
                        fontName="Helvetica-Bold" if bold else "Helvetica",
                        alignment=align, leading=size+3)
                doc = SimpleDocTemplate(save_path, pagesize=A4,
                    leftMargin=10*mm, rightMargin=10*mm, topMargin=8*mm, bottomMargin=8*mm)
                BASE = [("BOX",(0,0),(-1,-1),0.5,BDR),("INNERGRID",(0,0),(-1,-1),0.3,BDR),
                        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
                        ("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3),
                        ("VALIGN",(0,0),(-1,-1),"TOP")]
                def T(data, cw, style_cmds, repeat=0):
                    t = Table(data, colWidths=cw, style=TableStyle(style_cmds), repeatRows=repeat)
                    t.spaceAfter = 0; t.spaceBefore = 0; return t
                story = []
                story.append(T([[Paragraph(bill_title, ps("t",11,True,TA_CENTER))]],
                    [PAGE_W], BASE+[("BACKGROUND",(0,0),(-1,-1),GRY)]))
                story.append(Spacer(1,1*mm))
                shop_lines  = "<b>" + _sp["name"] + "</b><br/>"
                shop_lines += (_sp["address"] + ", " + _sp["city"]).strip(", ")
                if _sp["gstin"]:  shop_lines += "<br/>GSTIN: " + _sp["gstin"]
                if _sp["mobile"]: shop_lines += "  Mob: " + _sp["mobile"]
                info_lines  = "<b>Return No(s):</b> " + ", ".join(return_nos)
                info_lines += "<br/><b>Date:</b> " + fmt_date(ret_date)
                info_lines += "<br/><b>Orig Bill:</b> " + orig_bill
                info_lines += "<br/><b>Supplier:</b> " + party_name
                info_lines += "<br/><b>Reason:</b> " + (reason or "—")
                story.append(T([[Paragraph(shop_lines, ps("s",8)),
                                  Paragraph(info_lines, ps("i",8))]],
                               [PAGE_W*0.55, PAGE_W*0.45], BASE))
                story.append(Spacer(1,2*mm))
                hdr_row = [Paragraph(c, ps("h",8,True,TA_CENTER))
                           for c in ["Sl.","Product","HSN","Qty","Rate","Taxable","GST%","GST Amt","Total"]]
                col_w = [8*mm, 55*mm, 20*mm, 12*mm, 18*mm, 20*mm, 12*mm, 20*mm, 20*mm]
                data  = [hdr_row]
                for i, it in enumerate(items):
                    data.append([
                        Paragraph(str(i+1),                    ps("d",8,False,TA_CENTER)),
                        Paragraph(it["product"],                ps("d",8)),
                        Paragraph(it.get("hsn",""),             ps("d",8,False,TA_CENTER)),
                        Paragraph(str(it["qty"]),               ps("d",8,False,TA_CENTER)),
                        Paragraph(str(round(it["rate"],2)),     ps("d",8,False,TA_CENTER)),
                        Paragraph(str(round(it["taxable"],2)),  ps("d",8,False,TA_CENTER)),
                        Paragraph(str(int(it["gst"]))+"%",      ps("d",8,False,TA_CENTER)),
                        Paragraph(str(round(it["gst_amt"],2)),  ps("d",8,False,TA_CENTER)),
                        Paragraph(str(int(it["total"])),        ps("d",8,False,TA_RIGHT)),
                    ])
                data.append([
                    Paragraph("",""), Paragraph("<b>TOTAL</b>", ps("d",9,True)), "","","",
                    Paragraph("<b>"+str(round(total_taxable,2))+"</b>", ps("d",8,True,TA_CENTER)), "",
                    Paragraph("<b>"+str(round(total_gst,2))+"</b>",    ps("d",8,True,TA_CENTER)),
                    Paragraph("<b>"+str(int(grand_total))+"</b>",      ps("d",9,True,TA_RIGHT))])
                tbl_style = BASE + [
                    ("BACKGROUND",(0,0),(-1,0),  colors.HexColor("#7B6000")),
                    ("TEXTCOLOR",(0,0),(-1,0),   colors.white),
                    ("ROWBACKGROUNDS",(0,1),(-1,-2), [colors.white, colors.HexColor("#FFFBF0")]),
                    ("BACKGROUND",(0,-1),(-1,-1), colors.HexColor("#FFFDE7")),
                    ("FONTNAME",(0,-1),(-1,-1),   "Helvetica-Bold"),
                ]
                story.append(T(data, col_w, tbl_style, repeat=1))
                story.append(Spacer(1,2*mm))
                story.append(T([[
                    Paragraph("Amount in Words: <b>" + num_to_words(grand_total) + "</b>", ps("w",8)),
                    Paragraph("<b>Grand Total: Rs." + str(int(grand_total)) + "</b>", ps("gt",10,True,TA_RIGHT))
                ]], [PAGE_W*0.65, PAGE_W*0.35], BASE))
                story.append(Spacer(1,2*mm))
                story.append(T([[
                    Paragraph("Note: Yeh Debit Note hai. Original Bill: " + orig_bill, ps("n",8)),
                    Paragraph("For " + _sp["name"] + "<br/><br/>Authorised Signatory", ps("sg",8,False,TA_RIGHT))
                ]], [PAGE_W*0.6, PAGE_W*0.4], BASE))
                doc.build(story)
                if sys.platform == "win32":    os.startfile(save_path)
                elif sys.platform == "darwin": subprocess.run(["open", save_path])
                else:                          subprocess.run(["xdg-open", save_path])
                messagebox.showinfo("PDF Ready!", "Return Bill PDF save ho gaya!\n" + save_path)
            except ImportError:
                messagebox.showerror("Library Missing",
                    "reportlab install nahi hai.\nCMD mein chalao:\npip install reportlab")
            except Exception as e:
                messagebox.showerror("PDF Error", str(e))

        make_btn(bf, "Save Return",       lambda: do_save(False), bg=C_AMBER ).pack(side="left", padx=(0,8))
        make_btn(bf, "Save + Print Bill", lambda: do_save(True),  bg="#276749").pack(side="left")

        tk.Label(p, text="Purchase Return History", font=("Segoe UI",10,"bold"),
                 bg=C_BG, fg="#1A365D").pack(anchor="w", pady=(8,4))
        tbl = make_table(p,
            ["Return No","Date","Bill No","Supplier","Product","Qty","Total","Reason","Print"],
            [11,11,14,16,18,7,10,14,6])

        def load_hist():
            clear_table_rows(tbl)
            d = get_db()
            rows = [dict(r) for r in d.execute(
                "SELECT * FROM returns WHERE return_type='pur_return' ORDER BY return_date DESC"
            ).fetchall()]
            d.close()
            for i, r in enumerate(rows):
                bg = C_WHITE if i%2==0 else "#F7FAFC"
                vals = [r["return_no"], fmt_date(r["return_date"]), r["orig_bill"],
                        r["party"][:14], r["product"][:16], str(r["qty"]),
                        "Rs."+str(int(r["total_amt"])), (r["reason"] or "")[:14], ""]
                for j, val in enumerate(vals):
                    if j == 8:
                        tk.Button(tbl, text="Print", font=("Segoe UI",7),
                                  bg=C_AMBER, fg=C_WHITE, relief="flat", cursor="hand2", bd=0,
                                  command=lambda rr=dict(r): _reprint(rr)
                                  ).grid(row=i+1, column=j, sticky="nsew", padx=1, pady=1)
                    else:
                        tk.Label(tbl, text=str(val), font=("Segoe UI",9),
                                 bg=bg, fg=C_GRAY, anchor="w", padx=4, pady=3
                                 ).grid(row=i+1, column=j, sticky="nsew", padx=1)

        def _reprint(rrow):
            items_r = [{
                "product": rrow["product"], "hsn": "",
                "qty": rrow["qty"], "rate": rrow["rate"],
                "taxable": rrow["qty"]*rrow["rate"],
                "gst": rrow["gst_percent"], "gst_amt": rrow["gst_amt"], "total": rrow["total_amt"]
            }]
            _show_return_bill([rrow["return_no"]], rrow["orig_bill"], rrow["party"],
                              rrow["return_date"], rrow["reason"] or "",
                              items_r, "Purchase Return / Debit Note")
        load_hist()

    # ══════════════════════════════════════════════════════════════════════════
    #  EXCHANGE — Customer purana item return kare, naya item le
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_exchange(self):
        import datetime as _dt
        db = get_db()
        db.execute("""CREATE TABLE IF NOT EXISTS exchanges (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            exchange_no TEXT UNIQUE NOT NULL,
            exchange_date TEXT NOT NULL,
            orig_bill TEXT NOT NULL,
            party TEXT NOT NULL,
            reason TEXT DEFAULT '',
            return_total REAL DEFAULT 0,
            new_total REAL DEFAULT 0,
            diff_amt REAL DEFAULT 0,
            diff_type TEXT DEFAULT 'customer_pay',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS exchange_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            exchange_id INTEGER NOT NULL,
            item_type TEXT NOT NULL,
            product TEXT NOT NULL,
            hsn TEXT DEFAULT '',
            qty REAL DEFAULT 0,
            rate REAL DEFAULT 0,
            taxable REAL DEFAULT 0,
            gst_percent REAL DEFAULT 0,
            gst_amt REAL DEFAULT 0,
            total REAL DEFAULT 0
        )""")
        db.commit(); db.close()

        self._ex_return_items = []   # items being returned
        self._ex_new_items    = []   # new items being given

        p = tk.Frame(self.content, bg=C_BG, padx=10, pady=3)
        p.pack(fill="both", expand=True)
        section_title(p, "🔄 Exchange")

        # ── Top info bar ──────────────────────────────────────────────────────
        info_f = tk.Frame(p, bg="#EBF8FF", highlightthickness=1, highlightbackground="#90CDF4")
        info_f.pack(fill="x", pady=(0,10))
        tk.Label(info_f, text="Exchange kaise kaam karta hai: Customer purane item return karega → naya item milega → difference ka payment hoga. (Difference hamesha upar wali yellow strip mein dikhega — manually calculate karne ki zaroorat nahi)",
                 font=("Segoe UI",9), bg="#EBF8FF", fg="#2B6CB0", pady=6, padx=12).pack(anchor="w")

        # ── Header fields ─────────────────────────────────────────────────────
        hdr_f = tk.Frame(p, bg=C_WHITE, highlightthickness=1, highlightbackground=C_BORDER)
        hdr_f.pack(fill="x", pady=(0,3))
        tk.Label(hdr_f, text="Exchange Details",
                 font=("Segoe UI",10,"bold"), bg=C_WHITE, fg="#1A365D").pack(anchor="w", padx=14, pady=(10,6))

        hrow = tk.Frame(hdr_f, bg=C_WHITE); hrow.pack(fill="x", padx=14, pady=(0,10))
        tk.Label(hrow, text="Date:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_date = tk.StringVar(value=str(_dt.date.today()))
        ex_date_entry = make_date_entry(hrow, v_date, width=13)
        ex_date_entry.pack(side="left", padx=(2,16))

        tk.Label(hrow, text="Original Sale Bill:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_bill = tk.StringVar()
        db2 = get_db()
        bills = [r[0] for r in db2.execute("SELECT bill_no FROM sales ORDER BY bill_date DESC").fetchall()]
        db2.close()
        cb_bill = ttk.Combobox(hrow, textvariable=v_bill, values=bills, width=16)
        cb_bill.pack(side="left", padx=(2,16))
        add_autocomplete(cb_bill)

        tk.Label(hrow, text="Customer:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_party = tk.StringVar()
        e_party = ttk.Entry(hrow, textvariable=v_party, width=20, state="readonly")
        e_party.pack(side="left", padx=(2,16))

        tk.Label(hrow, text="Reason:", font=("Segoe UI",9), bg=C_WHITE, fg=C_GRAY).pack(side="left")
        v_reason = tk.StringVar()
        ex_reason_entry = ttk.Entry(hrow, textvariable=v_reason, width=22)
        ex_reason_entry.pack(side="left", padx=(2,0))

        def on_bill(ev=None):
            bn = v_bill.get().strip()
            if not bn: return
            d = get_db()
            row = d.execute("SELECT party FROM sales WHERE bill_no=?", (bn,)).fetchone()
            prods = [x[0] for x in d.execute(
                "SELECT DISTINCT si.product FROM sale_items si "
                "JOIN sales s ON si.sale_id=s.id WHERE s.bill_no=?", (bn,)).fetchall()]
            d.close()
            if row:
                e_party.config(state="normal"); v_party.set(row["party"]); e_party.config(state="readonly")
            cb_ret_prod["values"] = prods
            if prods: v_ret_prod.set(prods[0]); on_ret_prod()

        cb_bill.bind("<<ComboboxSelected>>", on_bill)

        # ── DIFFERENCE SUMMARY (always visible — top par, compact strip) ─────
        v_ret_amt_lbl  = tk.StringVar(value="↩  Return Amount:  ₹0")
        v_new_amt_lbl  = tk.StringVar(value="➕  New Items Amount:  ₹0")
        v_diff_lbl     = tk.StringVar(value="Difference:  ₹0")
        v_diff_msg     = tk.StringVar(value="")

        diff_f = tk.Frame(p, bg="#FEFCBF", highlightthickness=2, highlightbackground="#F6E05E")
        diff_f.pack(fill="x", pady=(0,6))
        diff_row = tk.Frame(diff_f, bg="#FEFCBF"); diff_row.pack(fill="x", padx=14, pady=6)

        tk.Label(diff_row, textvariable=v_ret_amt_lbl, font=("Segoe UI",9,"bold"),
                 bg="#FEFCBF", fg="#C53030").pack(side="left", padx=(0,20))
        tk.Label(diff_row, textvariable=v_new_amt_lbl, font=("Segoe UI",9,"bold"),
                 bg="#FEFCBF", fg="#276749").pack(side="left", padx=(0,20))
        tk.Frame(diff_row, bg="#D69E2E", width=1).pack(side="left", fill="y", padx=10)
        tk.Label(diff_row, textvariable=v_diff_lbl, font=("Segoe UI",12,"bold"),
                 bg="#FEFCBF", fg="#744210").pack(side="left", padx=(0,12))
        tk.Label(diff_row, textvariable=v_diff_msg, font=("Segoe UI",9,"italic"),
                 bg="#FEFCBF", fg="#B7791F").pack(side="left")

        # ────────────────────────────────────────────────────────────────────
        # TWO-COLUMN LAYOUT: Return (left) | New Items (right)
        # ────────────────────────────────────────────────────────────────────
        cols_f = tk.Frame(p, bg=C_BG); cols_f.pack(fill="both", expand=True, pady=(0,3))
        cols_f.columnconfigure(0, weight=1); cols_f.columnconfigure(1, weight=1)

        # ── LEFT: RETURN ITEMS ───────────────────────────────────────────────
        ret_f = tk.Frame(cols_f, bg="#FFF5F5", highlightthickness=1, highlightbackground="#FC8181")
        ret_f.grid(row=0, column=0, sticky="nsew", padx=(0,6))
        tk.Label(ret_f, text="↩  Wapas Aane Wale Items (Return)",
                 font=("Segoe UI",10,"bold"), bg="#FFF5F5", fg="#C53030").pack(anchor="w", padx=12, pady=(10,6))

        ret_row = tk.Frame(ret_f, bg="#FFF5F5"); ret_row.pack(fill="x", padx=12, pady=(0,3))
        tk.Label(ret_row, text="Product:", font=("Segoe UI",9), bg="#FFF5F5", fg=C_GRAY).pack(side="left")
        v_ret_prod = tk.StringVar()
        cb_ret_prod = ttk.Combobox(ret_row, textvariable=v_ret_prod, width=18)
        cb_ret_prod.pack(side="left", padx=(2,6))
        add_autocomplete(cb_ret_prod)
        tk.Label(ret_row, text="Qty:", font=("Segoe UI",9), bg="#FFF5F5", fg=C_GRAY).pack(side="left")
        v_ret_qty = tk.StringVar()
        ex_ret_qty_entry = ttk.Entry(ret_row, textvariable=v_ret_qty, width=6)
        ex_ret_qty_entry.pack(side="left", padx=(2,6))
        tk.Label(ret_row, text="Rate:", font=("Segoe UI",9), bg="#FFF5F5", fg=C_GRAY).pack(side="left")
        v_ret_rate = tk.StringVar()
        ex_ret_rate_entry = ttk.Entry(ret_row, textvariable=v_ret_rate, width=8)
        ex_ret_rate_entry.pack(side="left", padx=(2,6))
        tk.Label(ret_row, text="GST%:", font=("Segoe UI",9), bg="#FFF5F5", fg=C_GRAY).pack(side="left")
        v_ret_gst = tk.StringVar(value="0")
        ex_ret_gst_cb = ttk.Combobox(ret_row, textvariable=v_ret_gst, values=["0","5","12","18","28"],
                     width=4, state="readonly")
        ex_ret_gst_cb.pack(side="left", padx=(2,10))
        make_btn(ret_row, "+Add", lambda: add_ret_item(), bg="#C53030").pack(side="left")

        def on_ret_prod(ev=None):
            bn = v_bill.get().strip(); pr = v_ret_prod.get().strip()
            if not bn or not pr: return
            d = get_db()
            row = d.execute(
                "SELECT si.rate, si.gst_percent FROM sale_items si "
                "JOIN sales s ON si.sale_id=s.id WHERE s.bill_no=? AND si.product=? LIMIT 1",
                (bn, pr)).fetchone()
            d.close()
            if row: v_ret_rate.set(str(round(row["rate"],2))); v_ret_gst.set(str(int(row["gst_percent"])))
        cb_ret_prod.bind("<<ComboboxSelected>>", on_ret_prod)

        ret_tbl = make_table(ret_f, ["#","Product","Qty","Rate","Total","Del"], [3,20,5,9,9,4])
        ret_tbl.master.master.pack(fill="x", padx=12, pady=(0,4))

        v_ret_total = tk.StringVar(value="Return Total: ₹0")
        tk.Label(ret_f, textvariable=v_ret_total, font=("Segoe UI",10,"bold"),
                 bg="#FFF5F5", fg="#C53030", anchor="e").pack(fill="x", padx=12, pady=(0,3))

        def update_diff(*args):
            ret_total = sum(it["total"] for it in self._ex_return_items)
            new_total = sum(it["total"] for it in self._ex_new_items)
            diff = round(new_total - ret_total, 2)
            v_ret_amt_lbl.set(f"↩  Return Amount:  ₹{round(ret_total):,}")
            v_new_amt_lbl.set(f"➕  New Items Amount:  ₹{round(new_total):,}")
            if diff > 0:
                v_diff_lbl.set(f"Difference:  ₹{round(diff):,}  →  Customer ko ₹{round(diff):,} aur dene honge")
                v_diff_msg.set("(Naye items mehnge hain — customer extra payment karega)")
            elif diff < 0:
                v_diff_lbl.set(f"Difference:  ₹{round(abs(diff)):,}  →  Customer ko ₹{round(abs(diff)):,} wapas milenge")
                v_diff_msg.set("(Purane items mehnge the — customer ko refund dena hoga)")
            else:
                v_diff_lbl.set("Difference:  ₹0  →  Seedha exchange! Koi extra payment nahi.")
                v_diff_msg.set("(Dono ki value barabar hai)")

        def render_ret():
            clear_table_rows(ret_tbl)
            grand = 0
            for i, it in enumerate(self._ex_return_items):
                grand += it["total"]
                bg = "#FFF5F5" if i%2==0 else "#FED7D7"
                vals = [i+1, it["product"][:18], it["qty"],
                        f"₹{it['rate']:.2f}", f"₹{it['total']:.0f}", "✕"]
                for j, val in enumerate(vals):
                    if j == 5:
                        tk.Button(ret_tbl, text="✕", font=("Segoe UI",9),
                                  bg=bg, fg=C_RED, relief="flat", cursor="hand2", bd=0,
                                  command=lambda idx=i: (self._ex_return_items.pop(idx), render_ret(), update_diff())
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=0)
                    else:
                        tk.Label(ret_tbl, text=str(val), font=("Segoe UI",9),
                                 bg=bg, fg=C_GRAY, anchor="w", padx=4, pady=3
                                 ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=0)
            v_ret_total.set(f"Return Total: ₹{round(grand):,}")
            update_diff()

        def add_ret_item():
            prod = v_ret_prod.get().strip()
            if not prod: messagebox.showerror("Error","Product select karo!"); return
            if not v_bill.get().strip(): messagebox.showerror("Error","Pehle Bill No select karo!"); return
            try:
                qty = float(v_ret_qty.get()); rate = float(v_ret_rate.get()); gst_p = float(v_ret_gst.get())
                assert qty > 0
            except: messagebox.showerror("Error","Qty aur Rate sahi bharo!"); return
            taxable = round(qty*rate,2); gst_amt = round(taxable*gst_p/100,2); total = round(taxable+gst_amt,2)
            hsn = ""
            try:
                c_ = get_db(); r_ = c_.execute("SELECT hsn FROM products WHERE name=?", (prod,)).fetchone(); c_.close()
                if r_: hsn = r_["hsn"] or ""
            except: pass
            self._ex_return_items.append({
                "product":prod,"hsn":hsn,"qty":qty,"rate":rate,
                "taxable":taxable,"gst":gst_p,"gst_amt":gst_amt,"total":total
            })
            render_ret()
            update_diff()
            v_ret_qty.set(""); v_ret_rate.set(""); v_ret_gst.set("0")

        # (+Add button ab upar wale row mein hai)

        # ── RIGHT: NEW ITEMS ─────────────────────────────────────────────────
        new_f = tk.Frame(cols_f, bg="#F0FFF4", highlightthickness=1, highlightbackground="#68D391")
        new_f.grid(row=0, column=1, sticky="nsew", padx=(6,0))
        tk.Label(new_f, text="➕  Diye Jaane Wale Naye Items",
                 font=("Segoe UI",10,"bold"), bg="#F0FFF4", fg="#276749").pack(anchor="w", padx=12, pady=(10,6))

        new_row = tk.Frame(new_f, bg="#F0FFF4"); new_row.pack(fill="x", padx=12, pady=(0,3))
        db3 = get_db()
        all_prods = [r[0] for r in db3.execute("SELECT name FROM products ORDER BY name").fetchall()]
        prod_dict = {dict(r)["name"]:dict(r) for r in db3.execute("SELECT * FROM products").fetchall()}
        db3.close()
        tk.Label(new_row, text="Product:", font=("Segoe UI",9), bg="#F0FFF4", fg=C_GRAY).pack(side="left")
        v_new_prod = tk.StringVar()
        cb_new_prod = ttk.Combobox(new_row, textvariable=v_new_prod, values=all_prods, width=18)
        cb_new_prod.pack(side="left", padx=(2,6))
        add_autocomplete(cb_new_prod)
        tk.Label(new_row, text="Qty:", font=("Segoe UI",9), bg="#F0FFF4", fg=C_GRAY).pack(side="left")
        v_new_qty = tk.StringVar()
        ex_new_qty_entry = ttk.Entry(new_row, textvariable=v_new_qty, width=6)
        ex_new_qty_entry.pack(side="left", padx=(2,6))
        tk.Label(new_row, text="Rate:", font=("Segoe UI",9), bg="#F0FFF4", fg=C_GRAY).pack(side="left")
        v_new_rate = tk.StringVar()
        ex_new_rate_entry = ttk.Entry(new_row, textvariable=v_new_rate, width=8)
        ex_new_rate_entry.pack(side="left", padx=(2,6))
        tk.Label(new_row, text="GST%:", font=("Segoe UI",9), bg="#F0FFF4", fg=C_GRAY).pack(side="left")
        v_new_gst = tk.StringVar(value="5")
        ex_new_gst_cb = ttk.Combobox(new_row, textvariable=v_new_gst, values=["0","5","12","18","28"],
                     width=4, state="readonly")
        ex_new_gst_cb.pack(side="left", padx=(2,10))
        make_btn(new_row, "+Add", lambda: add_new_item(), bg="#276749").pack(side="left")

        def on_new_prod(ev=None):
            pr = v_new_prod.get()
            if pr in prod_dict:
                v_new_rate.set(str(prod_dict[pr]["sale_rate"]))
                v_new_gst.set(str(int(prod_dict[pr]["gst_percent"])))
        cb_new_prod.bind("<<ComboboxSelected>>", on_new_prod)

        new_tbl = make_table(new_f, ["#","Product","Qty","Rate","Total","Del"], [3,20,5,9,9,4])
        new_tbl.master.master.pack(fill="x", padx=12, pady=(0,4))

        v_new_total = tk.StringVar(value="New Total: ₹0")
        tk.Label(new_f, textvariable=v_new_total, font=("Segoe UI",10,"bold"),
                 bg="#F0FFF4", fg="#276749", anchor="e").pack(fill="x", padx=12, pady=(0,3))

        def render_new():
            clear_table_rows(new_tbl)
            grand = 0
            for i, it in enumerate(self._ex_new_items):
                grand += it["total"]
                bg = "#F0FFF4" if i%2==0 else "#C6F6D5"
                vals = [i+1, it["product"][:18], it["qty"],
                        f"₹{it['rate']:.2f}", f"₹{it['total']:.0f}", "✕"]
                for j, val in enumerate(vals):
                    if j == 5:
                        tk.Button(new_tbl, text="✕", font=("Segoe UI",9),
                                  bg=bg, fg=C_RED, relief="flat", cursor="hand2", bd=0,
                                  command=lambda idx=i: (self._ex_new_items.pop(idx), render_new(), update_diff())
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=0)
                    else:
                        tk.Label(new_tbl, text=str(val), font=("Segoe UI",9),
                                 bg=bg, fg=C_GRAY, anchor="w", padx=4, pady=3
                                 ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=0)
            v_new_total.set(f"New Total: ₹{round(grand):,}")
            update_diff()

        def add_new_item():
            prod = v_new_prod.get().strip()
            if not prod: messagebox.showerror("Error","Product select karo!"); return
            try:
                qty = float(v_new_qty.get()); rate = float(v_new_rate.get()); gst_p = float(v_new_gst.get())
                assert qty > 0
            except: messagebox.showerror("Error","Qty aur Rate sahi bharo!"); return
            taxable = round(qty*rate,2); gst_amt = round(taxable*gst_p/100,2); total = round(taxable+gst_amt,2)
            hsn = ""
            try:
                c_ = get_db(); r_ = c_.execute("SELECT hsn FROM products WHERE name=?", (prod,)).fetchone(); c_.close()
                if r_: hsn = r_["hsn"] or ""
            except: pass
            self._ex_new_items.append({
                "product":prod,"hsn":hsn,"qty":qty,"rate":rate,
                "taxable":taxable,"gst":gst_p,"gst_amt":gst_amt,"total":total
            })
            render_new()
            v_new_qty.set(""); v_new_prod.set(""); v_new_rate.set(""); v_new_gst.set("18")

        # (+Add button ab upar wale row mein hai)

        # ── SAVE BUTTON ───────────────────────────────────────────────────────
        bf = tk.Frame(p, bg=C_BG); bf.pack(fill="x", pady=(0,10))

        def do_save():
            if not self._ex_return_items and not self._ex_new_items:
                messagebox.showerror("Error","Koi bhi item nahi add kiya!"); return
            bill = v_bill.get().strip(); party = v_party.get().strip()
            if not bill or not party:
                messagebox.showerror("Error","Sale Bill aur Customer zaroori hain!"); return
            ret_total = sum(it["total"] for it in self._ex_return_items)
            new_total = sum(it["total"] for it in self._ex_new_items)
            diff      = round(new_total - ret_total, 2)
            diff_type = "customer_pay" if diff > 0 else ("refund" if diff < 0 else "even")

            db = get_db()
            try:
                n  = db.execute("SELECT COUNT(*) FROM exchanges").fetchone()[0]
                ex_no = "EX-{:03d}".format(n+1)
                db.execute(
                    "INSERT INTO exchanges(exchange_no,exchange_date,orig_bill,party,reason,"
                    "return_total,new_total,diff_amt,diff_type) VALUES(?,?,?,?,?,?,?,?,?)",
                    (ex_no, v_date.get(), bill, party, v_reason.get().strip(),
                     ret_total, new_total, abs(diff), diff_type))
                ex_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
                for it in self._ex_return_items:
                    db.execute(
                        "INSERT INTO exchange_items(exchange_id,item_type,product,hsn,qty,rate,"
                        "taxable,gst_percent,gst_amt,total) VALUES(?,?,?,?,?,?,?,?,?,?)",
                        (ex_id,"return",it["product"],it["hsn"],it["qty"],it["rate"],
                         it["taxable"],it["gst"],it["gst_amt"],it["total"]))
                    # Stock wapas add karo (returned item)
                    db.execute("UPDATE products SET opening_stock=opening_stock+? WHERE name=?",
                               (it["qty"], it["product"]))
                for it in self._ex_new_items:
                    db.execute(
                        "INSERT INTO exchange_items(exchange_id,item_type,product,hsn,qty,rate,"
                        "taxable,gst_percent,gst_amt,total) VALUES(?,?,?,?,?,?,?,?,?,?)",
                        (ex_id,"new",it["product"],it["hsn"],it["qty"],it["rate"],
                         it["taxable"],it["gst"],it["gst_amt"],it["total"]))
                    # Stock kam karo (naya item diya)
                    db.execute("UPDATE products SET opening_stock=MAX(0,opening_stock-?) WHERE name=?",
                               (it["qty"], it["product"]))
                db.commit()
                snap_ret = list(self._ex_return_items)
                snap_new = list(self._ex_new_items)
                if messagebox.askyesno("Saved!",
                    f"Exchange saved!\nExchange No: {ex_no}\n\n"
                    f"Return Total: ₹{round(ret_total):,}\n"
                    f"New Items Total: ₹{round(new_total):,}\n"
                    f"Difference: {'Customer dega' if diff>0 else ('Refund' if diff<0 else 'Even')} ₹{round(abs(diff)):,}\n\n"
                    "Kya Exchange Slip print karni hai?"):
                    _show_exchange_slip(ex_no, bill, party, v_date.get(),
                                        v_reason.get().strip(), snap_ret, snap_new,
                                        ret_total, new_total, diff)
                self._ex_return_items.clear(); self._ex_new_items.clear()
                render_ret(); render_new(); update_diff()
                v_bill.set(""); e_party.config(state="normal"); v_party.set(""); e_party.config(state="readonly")
                v_reason.set("")
            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                db.close()

        def _show_exchange_slip(ex_no, orig_bill, party_name, ex_date, reason,
                                ret_items, new_items, ret_total, new_total, diff):
            win = tk.Toplevel(self.root)
            win.title(f"Exchange Slip — {ex_no}")
            win.state("zoomed"); win.configure(bg=C_WHITE); _apply_logo(win)

            hdr = tk.Frame(win, bg="#744210", height=44); hdr.pack(fill="x"); hdr.pack_propagate(False)
            tk.Label(hdr, text=f"  🔄 Exchange Slip — {ex_no}", font=("Segoe UI",12,"bold"),
                     bg="#744210", fg="white").pack(side="left", padx=14, pady=4)
            make_btn(hdr, "🖨️ Print", lambda: win.after(100, lambda: [win.state("normal"),
                     __import__('subprocess').Popen(['mspaint', '/pt', '']) if False else None]),
                     bg=C_GREEN).pack(side="right", padx=10, pady=4)
            make_btn(hdr, "✕ Close", win.destroy, bg=C_GRAY).pack(side="right", padx=4, pady=4)

            outer = tk.Frame(win, bg="#E2E8F0"); outer.pack(fill="both", expand=True)
            canvas = tk.Canvas(outer, bg="#E2E8F0", highlightthickness=0)
            vsb    = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=vsb.set)
            vsb.pack(side="right", fill="y"); canvas.pack(fill="both", expand=True)
            canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>",
                lambda ev: canvas.yview_scroll(-1*(ev.delta//120), "units")))
            canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))
            paper = tk.Frame(canvas, bg=C_WHITE, width=600,
                             highlightthickness=1, highlightbackground="#999")
            wid = canvas.create_window((0, 20), window=paper, anchor="nw")
            canvas.bind("<Configure>", lambda e: (
                canvas.itemconfig(wid, width=min(600, e.width)),
                canvas.configure(scrollregion=canvas.bbox("all"))))
            paper.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

            _sp  = get_shop()
            main = tk.Frame(paper, bg=C_WHITE, padx=22, pady=16); main.pack(fill="x")

            # Header
            tk.Label(main, text="🔄  EXCHANGE SLIP", font=("Arial",14,"bold"),
                     bg=C_WHITE, anchor="center").pack(fill="x", pady=(0,4))
            tk.Frame(main, bg="#744210", height=2).pack(fill="x", pady=(0,3))

            info_tbl = tk.Frame(main, bg=C_WHITE); info_tbl.pack(fill="x", pady=(0,3))
            info_tbl.columnconfigure(0, weight=1); info_tbl.columnconfigure(1, weight=1)
            _sp_name = _sp["name"] if _sp else ""
            for r,(lbl,val) in enumerate([
                    ("Shop:", _sp_name),
                    ("Exchange No:", ex_no),
                    ("Date:", fmt_date(ex_date)),
                    ("Original Bill:", orig_bill),
                    ("Customer:", party_name),
                    ("Reason:", reason or "—")]):
                tk.Label(info_tbl, text=lbl, font=("Arial",9,"bold"), bg=C_WHITE,
                         anchor="w", padx=4, pady=2).grid(row=r, column=0, sticky="w")
                tk.Label(info_tbl, text=val, font=("Arial",9), bg=C_WHITE,
                         anchor="w", padx=4, pady=2).grid(row=r, column=1, sticky="w")

            # Return items
            tk.Frame(main, bg="#C53030", height=2).pack(fill="x", pady=(8,4))
            tk.Label(main, text="↩  Wapas Kiye Gaye Items", font=("Arial",10,"bold"),
                     bg=C_WHITE, fg="#C53030").pack(anchor="w")
            HDR_BG = "#FED7D7"
            rt = tk.Frame(main, bg=C_WHITE); rt.pack(fill="x", pady=(2,0))
            for ci, col in enumerate(["#","Product","Qty","Rate","Total"]):
                tk.Frame(rt, bg=HDR_BG, highlightthickness=1, highlightbackground="#999"
                         ).grid(row=0, column=ci, sticky="nsew")
                tk.Label(rt, text=col, font=("Arial",8,"bold"), bg=HDR_BG,
                         padx=4, pady=3, anchor="center").grid(row=0, column=ci, sticky="nsew")
            for ri, it in enumerate(ret_items):
                bg = C_WHITE if ri%2==0 else "#FFF5F5"
                for ci, val in enumerate([ri+1, it["product"], it["qty"],
                                          f"₹{it['rate']:.2f}", f"₹{it['total']:.0f}"]):
                    tk.Label(rt, text=str(val), font=("Arial",9), bg=bg,
                             padx=4, pady=3, anchor="w").grid(row=ri+1, column=ci, sticky="nsew")
            tk.Label(main, text=f"Return Total: ₹{round(ret_total):,}",
                     font=("Arial",10,"bold"), bg=C_WHITE, fg="#C53030", anchor="e").pack(fill="x")

            # New items
            tk.Frame(main, bg="#276749", height=2).pack(fill="x", pady=(8,4))
            tk.Label(main, text="➕  Diye Gaye Naye Items", font=("Arial",10,"bold"),
                     bg=C_WHITE, fg="#276749").pack(anchor="w")
            HDR_BG2 = "#C6F6D5"
            nt = tk.Frame(main, bg=C_WHITE); nt.pack(fill="x", pady=(2,0))
            for ci, col in enumerate(["#","Product","Qty","Rate","Total"]):
                tk.Frame(nt, bg=HDR_BG2, highlightthickness=1, highlightbackground="#999"
                         ).grid(row=0, column=ci, sticky="nsew")
                tk.Label(nt, text=col, font=("Arial",8,"bold"), bg=HDR_BG2,
                         padx=4, pady=3, anchor="center").grid(row=0, column=ci, sticky="nsew")
            for ri, it in enumerate(new_items):
                bg = C_WHITE if ri%2==0 else "#F0FFF4"
                for ci, val in enumerate([ri+1, it["product"], it["qty"],
                                          f"₹{it['rate']:.2f}", f"₹{it['total']:.0f}"]):
                    tk.Label(nt, text=str(val), font=("Arial",9), bg=bg,
                             padx=4, pady=3, anchor="w").grid(row=ri+1, column=ci, sticky="nsew")
            tk.Label(main, text=f"New Total: ₹{round(new_total):,}",
                     font=("Arial",10,"bold"), bg=C_WHITE, fg="#276749", anchor="e").pack(fill="x")

            # Difference
            tk.Frame(main, bg="#D69E2E", height=2).pack(fill="x", pady=(10,4))
            diff_abs = abs(diff)
            if diff > 0:
                diff_txt = f"Customer ko ₹{round(diff_abs):,} aur dene honge"
                diff_clr = "#C53030"
            elif diff < 0:
                diff_txt = f"Customer ko ₹{round(diff_abs):,} wapas milenge (Refund)"
                diff_clr = "#276749"
            else:
                diff_txt = "Seedha Exchange — koi extra payment nahi"
                diff_clr = "#744210"
            tk.Label(main, text=f"💰  {diff_txt}", font=("Arial",11,"bold"),
                     bg="#FEFCBF", fg=diff_clr, anchor="center", pady=4).pack(fill="x")
            tk.Frame(main, bg="#D69E2E", height=2).pack(fill="x")
            tk.Label(main, text="\n_____________________\nCustomer Signature",
                     font=("Arial",9), bg=C_WHITE, anchor="center").pack(pady=(12,0))

        make_btn(bf, "💾  Save Exchange", do_save, bg="#744210").pack(side="left", padx=(0,8))
        make_btn(bf, "🗑️  Clear All", lambda: (
            self._ex_return_items.clear(), self._ex_new_items.clear(),
            render_ret(), render_new(), update_diff()
        ), bg=C_GRAY).pack(side="left")

        update_diff()

        # ── Enter Key Navigation ────────────────────────────────────────────
        # Header: Date→Bill→Reason→Return Product
        # Return row: Product→Qty→Rate→GST%→(Add Return Item, phir wapas Product)
        # New row: Product→Qty→Rate→GST%→(Add New Item, phir wapas Product)
        def _ex_focus(w):
            def _go(e):
                target = getattr(w, "_entry", w)
                target.focus_set()
                try: target.select_range(0, "end")
                except: pass
                return "break"
            return _go

        getattr(ex_date_entry, "_entry", ex_date_entry).bind("<Return>", _ex_focus(cb_bill), add="+")
        cb_bill.bind("<Return>", _ex_focus(ex_reason_entry), add="+")
        ex_reason_entry.bind("<Return>", _ex_focus(cb_ret_prod), add="+")

        cb_ret_prod.bind("<Return>", _ex_focus(ex_ret_qty_entry), add="+")
        ex_ret_qty_entry.bind("<Return>", _ex_focus(ex_ret_rate_entry), add="+")
        ex_ret_rate_entry.bind("<Return>", _ex_focus(ex_ret_gst_cb), add="+")

        def _ex_ret_gst_enter(e):
            add_ret_item()
            cb_ret_prod.focus_set()
            return "break"
        ex_ret_gst_cb.bind("<Return>", _ex_ret_gst_enter, add="+")

        cb_new_prod.bind("<Return>", _ex_focus(ex_new_qty_entry), add="+")
        ex_new_qty_entry.bind("<Return>", _ex_focus(ex_new_rate_entry), add="+")
        ex_new_rate_entry.bind("<Return>", _ex_focus(ex_new_gst_cb), add="+")

        def _ex_new_gst_enter(e):
            add_new_item()
            cb_new_prod.focus_set()
            return "break"
        ex_new_gst_cb.bind("<Return>", _ex_new_gst_enter, add="+")

    # ══════════════════════════════════════════════════════════════════════════
    #  SALE RETURN HISTORY
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_salereturnhistory(self):
        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "↩ Sale Return History")

        # ── Filter bar ────────────────────────────────────────────────────────
        ff = tk.Frame(p, bg=C_BG); ff.pack(fill="x", pady=(0,3))
        tk.Label(ff, text="From:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_from = tk.StringVar()
        make_date_entry(ff, v_from, width=12).pack(side="left", padx=4)
        tk.Label(ff, text="To:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_to = tk.StringVar()
        make_date_entry(ff, v_to, width=12).pack(side="left", padx=4)
        tk.Label(ff, text="Customer:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left", padx=(10,4))
        v_party = tk.StringVar()
        db0 = get_db()
        party_list = [""]+[r[0] for r in db0.execute("SELECT name FROM parties ORDER BY name").fetchall()]
        db0.close()
        ttk.Combobox(ff, textvariable=v_party, values=party_list, width=20).pack(side="left", padx=4)
        make_btn(ff, "🔍 Search", lambda: load(), bg=C_ACCENT).pack(side="left", padx=8)
        make_btn(ff, "↺ Reset",  lambda: [v_from.set(""), v_to.set(""), v_party.set(""), load()],
                 bg=C_GRAY).pack(side="left")

        # ── Table ─────────────────────────────────────────────────────────────
        tbl = make_table(p,
            ["Return No","Date","Bill No","Customer","Product","Qty","Rate","Total","Reason","Print","Del"],
            [11,11,13,15,17,6,9,9,13,6,5])

        tot_f = tk.Frame(p, bg=C_WHITE, highlightthickness=1, highlightbackground=C_BORDER)
        tot_f.pack(fill="x", pady=4)
        v_summary = tk.StringVar(value="Total Returns: 0  |  Total Amt: ₹0")
        tk.Label(tot_f, textvariable=v_summary, font=("Segoe UI",10,"bold"),
                 bg=C_WHITE, fg=C_RED).pack(side="left", padx=12, pady=6)

        def _print_return_pdf(rrow, title):
            import os
            from tkinter import filedialog
            save_path = filedialog.asksaveasfilename(
                parent=p, title="Return Bill PDF Save Karo",
                defaultextension=".pdf",
                initialfile=f"Return_{rrow['return_no'].replace('/','_')}.pdf",
                filetypes=[("PDF File","*.pdf")]
            )
            if not save_path: return
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.units import mm
                from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                                Paragraph, Spacer, HRFlowable)
                from reportlab.lib.styles import ParagraphStyle
                from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
                BDR  = colors.black
                RED  = colors.HexColor("#C53030")
                GRY  = colors.HexColor("#f2f2f2")
                def ps(n,sz=9,bold=False,align=TA_LEFT,col=colors.black):
                    return ParagraphStyle(n,fontSize=sz,
                        fontName="Helvetica-Bold" if bold else "Helvetica",
                        alignment=align,leading=sz+3,textColor=col)
                doc = SimpleDocTemplate(save_path,pagesize=A4,
                    leftMargin=12*mm,rightMargin=12*mm,topMargin=10*mm,bottomMargin=10*mm)
                _sp = get_shop()
                story=[]
                story.append(Paragraph(_sp["name"] if _sp else "", ps("sh",14,True,TA_CENTER)))
                if _sp and _sp.get("address"):
                    story.append(Paragraph(_sp["address"], ps("sa",8,align=TA_CENTER)))
                if _sp and _sp.get("gstin"):
                    story.append(Paragraph(f"GSTIN: {_sp['gstin']}", ps("sg",8,align=TA_CENTER)))
                story.append(Spacer(1,3*mm))
                story.append(HRFlowable(width="100%",thickness=2,color=RED))
                story.append(Paragraph(title.upper(), ps("t",13,True,TA_CENTER,RED)))
                story.append(HRFlowable(width="100%",thickness=2,color=RED))
                story.append(Spacer(1,3*mm))
                cw=[28*mm,52*mm,28*mm,52*mm]
                BASE=[("BOX",(0,0),(-1,-1),.5,BDR),("INNERGRID",(0,0),(-1,-1),.3,BDR),
                      ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
                      ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
                      ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),
                      ("FONTSIZE",(0,0),(-1,-1),8),("BACKGROUND",(0,0),(0,-1),GRY),("BACKGROUND",(2,0),(2,-1),GRY)]
                info=[["Return No:",rrow["return_no"],"Date:",fmt_date(rrow["return_date"])],
                      ["Orig Bill:",rrow["orig_bill"],"Customer:",rrow["party"]],
                      ["Reason:",rrow["reason"] or "—","",""]]
                story.append(Table(info,colWidths=cw,style=TableStyle(BASE)))
                story.append(Spacer(1,4*mm))
                taxable = round(rrow["qty"]*rrow["rate"],2)
                gst_amt = round(rrow.get("gst_amt",0),2)
                items_data=[
                    ["#","Product","HSN","Qty","Rate","Taxable","GST%","GST Amt","Total"],
                    ["1",rrow["product"],rrow.get("hsn",""),str(rrow["qty"]),
                     f"{rrow['rate']:.2f}",f"{taxable:.2f}",
                     f"{rrow.get('gst_percent',0):.0f}%",f"{gst_amt:.2f}",
                     f"{rrow['total_amt']:.0f}"],
                    ["","","","","","","","Total:",f"₹{round(rrow['total_amt']):,}"]
                ]
                i_cw=[8*mm,38*mm,16*mm,12*mm,18*mm,18*mm,12*mm,18*mm,18*mm]
                i_style=[("BOX",(0,0),(-1,-1),.5,BDR),("INNERGRID",(0,0),(-1,-1),.3,BDR),
                         ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#FED7D7")),
                         ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                         ("FONTNAME",(7,-1),(8,-1),"Helvetica-Bold"),
                         ("TEXTCOLOR",(7,-1),(8,-1),RED),
                         ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#FFF5F5")),
                         ("ALIGN",(3,0),(-1,-1),"RIGHT"),
                         ("FONTSIZE",(0,0),(-1,-1),8),
                         ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
                         ("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3)]
                story.append(Table(items_data,colWidths=i_cw,style=TableStyle(i_style)))
                story.append(Spacer(1,14*mm))
                sig=[["_____________________","_____________________"],
                     ["Customer Signature","Authorized Signature"]]
                story.append(Table(sig,colWidths=[85*mm,85*mm],
                    style=TableStyle([("ALIGN",(0,0),(-1,-1),"CENTER"),
                                      ("FONTSIZE",(0,0),(-1,-1),8),
                                      ("TOPPADDING",(0,0),(-1,-1),2),
                                      ("BOTTOMPADDING",(0,0),(-1,-1),2)])))
                doc.build(story)
                import subprocess
                if os.name=="nt": os.startfile(save_path)
                else: subprocess.Popen(["xdg-open",save_path])
                messagebox.showinfo("Print Ready!","PDF ban gayi! Print karo.",parent=p)
            except ImportError:
                messagebox.showerror("Error","pip install reportlab",parent=p)
            except Exception as e:
                messagebox.showerror("Error",str(e),parent=p)

        def _delete_return(rid, rno):
            if not messagebox.askyesno("Delete?",
                    f"Return '{rno}' delete karna chahte hain?", parent=p):
                return
            db = get_db()
            db.execute("DELETE FROM returns WHERE id=?", (rid,))
            db.commit(); db.close()
            load()

        def load():
            clear_table_rows(tbl)
            frm=v_from.get(); to=v_to.get(); party=v_party.get()
            db=get_db()
            q="SELECT * FROM returns WHERE return_type='sale_return'"
            params=[]
            if frm:   q+=" AND return_date>=?"; params.append(frm)
            if to:    q+=" AND return_date<=?"; params.append(to)
            if party: q+=" AND party=?"; params.append(party)
            q+=" ORDER BY id DESC"
            try: rows=[dict(r) for r in db.execute(q,params).fetchall()]
            except: rows=[]
            db.close()
            total_amt=0
            for i,r in enumerate(rows):
                bg=C_WHITE if i%2==0 else "#F7FAFC"
                total_amt+=r.get("total_amt",0)
                vals=[r["return_no"],fmt_date(r["return_date"]),r["orig_bill"],
                      r["party"][:13],r["product"][:15],str(r["qty"]),
                      f"₹{r['rate']:.0f}",f"₹{r['total_amt']:.0f}",
                      (r["reason"] or "")[:12],"",""]
                for j,val in enumerate(vals):
                    if j==9:
                        tk.Button(tbl,text="🖨️",font=("Segoe UI",9),
                                  bg=C_ACCENT,fg=C_WHITE,relief="flat",cursor="hand2",bd=0,pady=3,
                                  command=lambda rr=r: _print_return_pdf(rr,"Sale Return / Credit Note")
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=1)
                    elif j==10:
                        tk.Button(tbl,text="🗑",font=("Segoe UI",9),
                                  bg=C_RED,fg=C_WHITE,relief="flat",cursor="hand2",bd=0,pady=3,
                                  command=lambda rid=r["id"],rno=r["return_no"]: _delete_return(rid,rno)
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=1)
                    else:
                        tk.Label(tbl,text=str(val),font=("Segoe UI",9),
                                 bg=bg,fg=C_GRAY,anchor="w",padx=4,pady=3
                                 ).grid(row=i+1,column=j,sticky="nsew",padx=1)
            v_summary.set(f"Total Returns: {len(rows)}  |  Total Amt: ₹{round(total_amt):,}")

        load()

    # ══════════════════════════════════════════════════════════════════════════
    #  PURCHASE RETURN HISTORY
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_purreturnhistory(self):
        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "↩ Purchase Return History")

        ff = tk.Frame(p, bg=C_BG); ff.pack(fill="x", pady=(0,3))
        tk.Label(ff, text="From:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_from = tk.StringVar()
        make_date_entry(ff, v_from, width=12).pack(side="left", padx=4)
        tk.Label(ff, text="To:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_to = tk.StringVar()
        make_date_entry(ff, v_to, width=12).pack(side="left", padx=4)
        tk.Label(ff, text="Supplier:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left", padx=(10,4))
        v_party = tk.StringVar()
        db0 = get_db()
        party_list = [""]+[r[0] for r in db0.execute("SELECT name FROM parties ORDER BY name").fetchall()]
        db0.close()
        ttk.Combobox(ff, textvariable=v_party, values=party_list, width=20).pack(side="left", padx=4)
        make_btn(ff, "🔍 Search", lambda: load(), bg=C_ACCENT).pack(side="left", padx=8)
        make_btn(ff, "↺ Reset",  lambda: [v_from.set(""), v_to.set(""), v_party.set(""), load()],
                 bg=C_GRAY).pack(side="left")

        tbl = make_table(p,
            ["Return No","Date","Bill No","Supplier","Product","Qty","Rate","Total","Reason","Print","Del"],
            [11,11,13,15,17,6,9,9,13,6,5])

        tot_f = tk.Frame(p, bg=C_WHITE, highlightthickness=1, highlightbackground=C_BORDER)
        tot_f.pack(fill="x", pady=4)
        v_summary = tk.StringVar(value="Total Returns: 0  |  Total Amt: ₹0")
        tk.Label(tot_f, textvariable=v_summary, font=("Segoe UI",10,"bold"),
                 bg=C_WHITE, fg=C_AMBER).pack(side="left", padx=12, pady=6)

        def _print_return_pdf(rrow, title):
            import os
            from tkinter import filedialog
            save_path = filedialog.asksaveasfilename(
                parent=p, title="Return Bill PDF Save Karo",
                defaultextension=".pdf",
                initialfile=f"Return_{rrow['return_no'].replace('/','_')}.pdf",
                filetypes=[("PDF File","*.pdf")]
            )
            if not save_path: return
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.units import mm
                from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                                Paragraph, Spacer, HRFlowable)
                from reportlab.lib.styles import ParagraphStyle
                from reportlab.lib.enums import TA_CENTER, TA_LEFT
                BDR   = colors.black
                AMBER = colors.HexColor("#744210")
                GRY   = colors.HexColor("#f2f2f2")
                def ps(n,sz=9,bold=False,align=TA_LEFT,col=colors.black):
                    return ParagraphStyle(n,fontSize=sz,
                        fontName="Helvetica-Bold" if bold else "Helvetica",
                        alignment=align,leading=sz+3,textColor=col)
                doc = SimpleDocTemplate(save_path,pagesize=A4,
                    leftMargin=12*mm,rightMargin=12*mm,topMargin=10*mm,bottomMargin=10*mm)
                _sp = get_shop()
                story=[]
                story.append(Paragraph(_sp["name"] if _sp else "", ps("sh",14,True,TA_CENTER)))
                if _sp and _sp.get("address"):
                    story.append(Paragraph(_sp["address"], ps("sa",8,align=TA_CENTER)))
                if _sp and _sp.get("gstin"):
                    story.append(Paragraph(f"GSTIN: {_sp['gstin']}", ps("sg",8,align=TA_CENTER)))
                story.append(Spacer(1,3*mm))
                story.append(HRFlowable(width="100%",thickness=2,color=AMBER))
                story.append(Paragraph(title.upper(), ps("t",13,True,TA_CENTER,AMBER)))
                story.append(HRFlowable(width="100%",thickness=2,color=AMBER))
                story.append(Spacer(1,3*mm))
                cw=[28*mm,52*mm,28*mm,52*mm]
                BASE=[("BOX",(0,0),(-1,-1),.5,BDR),("INNERGRID",(0,0),(-1,-1),.3,BDR),
                      ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
                      ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
                      ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),
                      ("FONTSIZE",(0,0),(-1,-1),8),("BACKGROUND",(0,0),(0,-1),GRY),("BACKGROUND",(2,0),(2,-1),GRY)]
                info=[["Return No:",rrow["return_no"],"Date:",fmt_date(rrow["return_date"])],
                      ["Orig Bill:",rrow["orig_bill"],"Supplier:",rrow["party"]],
                      ["Reason:",rrow["reason"] or "—","",""]]
                story.append(Table(info,colWidths=cw,style=TableStyle(BASE)))
                story.append(Spacer(1,4*mm))
                taxable=round(rrow["qty"]*rrow["rate"],2)
                gst_amt=round(rrow.get("gst_amt",0),2)
                items_data=[
                    ["#","Product","HSN","Qty","Rate","Taxable","GST%","GST Amt","Total"],
                    ["1",rrow["product"],rrow.get("hsn",""),str(rrow["qty"]),
                     f"{rrow['rate']:.2f}",f"{taxable:.2f}",
                     f"{rrow.get('gst_percent',0):.0f}%",f"{gst_amt:.2f}",
                     f"{rrow['total_amt']:.0f}"],
                    ["","","","","","","","Total:",f"₹{round(rrow['total_amt']):,}"]
                ]
                i_cw=[8*mm,38*mm,16*mm,12*mm,18*mm,18*mm,12*mm,18*mm,18*mm]
                i_style=[("BOX",(0,0),(-1,-1),.5,BDR),("INNERGRID",(0,0),(-1,-1),.3,BDR),
                         ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#FEFCBF")),
                         ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                         ("FONTNAME",(7,-1),(8,-1),"Helvetica-Bold"),
                         ("TEXTCOLOR",(7,-1),(8,-1),AMBER),
                         ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#FFFFF0")),
                         ("ALIGN",(3,0),(-1,-1),"RIGHT"),
                         ("FONTSIZE",(0,0),(-1,-1),8),
                         ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
                         ("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3)]
                story.append(Table(items_data,colWidths=i_cw,style=TableStyle(i_style)))
                story.append(Spacer(1,14*mm))
                sig=[["_____________________","_____________________"],
                     ["Supplier Signature","Authorized Signature"]]
                story.append(Table(sig,colWidths=[85*mm,85*mm],
                    style=TableStyle([("ALIGN",(0,0),(-1,-1),"CENTER"),
                                      ("FONTSIZE",(0,0),(-1,-1),8),
                                      ("TOPPADDING",(0,0),(-1,-1),2),
                                      ("BOTTOMPADDING",(0,0),(-1,-1),2)])))
                doc.build(story)
                import subprocess
                if os.name=="nt": os.startfile(save_path)
                else: subprocess.Popen(["xdg-open",save_path])
                messagebox.showinfo("Print Ready!","PDF ban gayi! Print karo.",parent=p)
            except ImportError:
                messagebox.showerror("Error","pip install reportlab",parent=p)
            except Exception as e:
                messagebox.showerror("Error",str(e),parent=p)

        def _delete_return(rid, rno):
            if not messagebox.askyesno("Delete?",
                    f"Return '{rno}' delete karna chahte hain?", parent=p):
                return
            db = get_db()
            db.execute("DELETE FROM returns WHERE id=?", (rid,))
            db.commit(); db.close()
            load()

        def load():
            clear_table_rows(tbl)
            frm=v_from.get(); to=v_to.get(); party=v_party.get()
            db=get_db()
            q="SELECT * FROM returns WHERE return_type='pur_return'"
            params=[]
            if frm:   q+=" AND return_date>=?"; params.append(frm)
            if to:    q+=" AND return_date<=?"; params.append(to)
            if party: q+=" AND party=?"; params.append(party)
            q+=" ORDER BY id DESC"
            try: rows=[dict(r) for r in db.execute(q,params).fetchall()]
            except: rows=[]
            db.close()
            total_amt=0
            for i,r in enumerate(rows):
                bg=C_WHITE if i%2==0 else "#F7FAFC"
                total_amt+=r.get("total_amt",0)
                vals=[r["return_no"],fmt_date(r["return_date"]),r["orig_bill"],
                      r["party"][:13],r["product"][:15],str(r["qty"]),
                      f"₹{r['rate']:.0f}",f"₹{r['total_amt']:.0f}",
                      (r["reason"] or "")[:12],"",""]
                for j,val in enumerate(vals):
                    if j==9:
                        tk.Button(tbl,text="🖨️",font=("Segoe UI",9),
                                  bg=C_AMBER,fg=C_WHITE,relief="flat",cursor="hand2",bd=0,pady=3,
                                  command=lambda rr=r: _print_return_pdf(rr,"Purchase Return / Debit Note")
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=1)
                    elif j==10:
                        tk.Button(tbl,text="🗑",font=("Segoe UI",9),
                                  bg=C_RED,fg=C_WHITE,relief="flat",cursor="hand2",bd=0,pady=3,
                                  command=lambda rid=r["id"],rno=r["return_no"]: _delete_return(rid,rno)
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=1)
                    else:
                        tk.Label(tbl,text=str(val),font=("Segoe UI",9),
                                 bg=bg,fg=C_GRAY,anchor="w",padx=4,pady=3
                                 ).grid(row=i+1,column=j,sticky="nsew",padx=1)
            v_summary.set(f"Total Returns: {len(rows)}  |  Total Amt: ₹{round(total_amt):,}")

        load()

    # ══════════════════════════════════════════════════════════════════════════
    #  EXCHANGE HISTORY
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_exchangehistory(self):
        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="x")
        section_title(p, "🔄 Exchange History")

        ff = tk.Frame(p, bg=C_BG); ff.pack(fill="x", pady=(0,3))
        tk.Label(ff, text="From:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_from = tk.StringVar()
        make_date_entry(ff, v_from, width=12).pack(side="left", padx=4)
        tk.Label(ff, text="To:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_to = tk.StringVar()
        make_date_entry(ff, v_to, width=12).pack(side="left", padx=4)
        tk.Label(ff, text="Customer:", font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(side="left", padx=(10,4))
        v_party = tk.StringVar()
        db0 = get_db()
        party_list = [""]+[r[0] for r in db0.execute("SELECT name FROM parties ORDER BY name").fetchall()]
        db0.close()
        ttk.Combobox(ff, textvariable=v_party, values=party_list, width=20).pack(side="left", padx=4)
        make_btn(ff, "🔍 Search", lambda: load(), bg=C_ACCENT).pack(side="left", padx=8)
        make_btn(ff, "↺ Reset", lambda: [v_from.set(""), v_to.set(""), v_party.set(""), load()],
                 bg=C_GRAY).pack(side="left")

        tbl = make_table(p,
            ["Exch No","Date","Orig Bill","Customer","Return ₹","New ₹","Difference ₹","Print","Del"],
            [14,11,14,18,10,10,10,6,5])

        tot_f = tk.Frame(p, bg=C_WHITE, highlightthickness=1, highlightbackground=C_BORDER)
        tot_f.pack(fill="x", pady=4)
        v_summary = tk.StringVar(value="Total Exchanges: 0")
        tk.Label(tot_f, textvariable=v_summary, font=("Segoe UI",10,"bold"),
                 bg=C_WHITE, fg="#744210").pack(side="left", padx=12, pady=6)

        def _reprint_exchange(ex_id):
            db = get_db()
            row = db.execute("SELECT * FROM exchanges WHERE id=?", (ex_id,)).fetchone()
            if not row:
                db.close(); messagebox.showerror("Error","Record nahi mila!"); return
            ret_items=[dict(r) for r in db.execute(
                "SELECT * FROM exchange_items WHERE exchange_id=? AND item_type='return'",(ex_id,)).fetchall()]
            new_items=[dict(r) for r in db.execute(
                "SELECT * FROM exchange_items WHERE exchange_id=? AND item_type='new'",(ex_id,)).fetchall()]
            db.close()
            import os
            from tkinter import filedialog
            save_path = filedialog.asksaveasfilename(
                parent=p, title="Exchange Slip PDF Save Karo",
                defaultextension=".pdf",
                initialfile=f"Exchange_{row['exchange_no'].replace('/','_')}.pdf",
                filetypes=[("PDF File","*.pdf")]
            )
            if not save_path: return
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.units import mm
                from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                                Paragraph, Spacer, HRFlowable)
                from reportlab.lib.styles import ParagraphStyle
                from reportlab.lib.enums import TA_CENTER, TA_LEFT
                BDR  =colors.black; RED=colors.HexColor("#C53030")
                GREEN=colors.HexColor("#276749"); BROWN=colors.HexColor("#744210")
                GRY  =colors.HexColor("#f2f2f2")
                def ps(n,sz=9,bold=False,align=TA_LEFT,col=colors.black):
                    return ParagraphStyle(n,fontSize=sz,
                        fontName="Helvetica-Bold" if bold else "Helvetica",
                        alignment=align,leading=sz+3,textColor=col)
                doc=SimpleDocTemplate(save_path,pagesize=A4,
                    leftMargin=12*mm,rightMargin=12*mm,topMargin=10*mm,bottomMargin=10*mm)
                _sp=get_shop(); story=[]
                story.append(Paragraph(_sp["name"] if _sp else "",ps("sh",14,True,TA_CENTER)))
                if _sp and _sp.get("address"): story.append(Paragraph(_sp["address"],ps("sa",8,align=TA_CENTER)))
                if _sp and _sp.get("gstin"):   story.append(Paragraph(f"GSTIN: {_sp['gstin']}",ps("sg",8,align=TA_CENTER)))
                story.append(Spacer(1,3*mm))
                story.append(HRFlowable(width="100%",thickness=2,color=BROWN))
                story.append(Paragraph("EXCHANGE SLIP",ps("t",13,True,TA_CENTER,BROWN)))
                story.append(HRFlowable(width="100%",thickness=2,color=BROWN))
                story.append(Spacer(1,3*mm))
                cw=[28*mm,52*mm,28*mm,52*mm]
                BASE=[("BOX",(0,0),(-1,-1),.5,BDR),("INNERGRID",(0,0),(-1,-1),.3,BDR),
                      ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
                      ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
                      ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),
                      ("FONTSIZE",(0,0),(-1,-1),8),("BACKGROUND",(0,0),(0,-1),GRY),("BACKGROUND",(2,0),(2,-1),GRY)]
                story.append(Table(
                    [["Exchange No:",row["exchange_no"],"Date:",fmt_date(row["exchange_date"])],
                     ["Original Bill:",row["orig_bill"],"Customer:",row["party"]],
                     ["Reason:",row["reason"] or "—","",""]],
                    colWidths=cw,style=TableStyle(BASE)))
                story.append(Spacer(1,4*mm))
                r_cw=[10*mm,65*mm,18*mm,28*mm,28*mm]
                for items,label,hbg,rbg,tcol in [
                    (ret_items,"↩  Wapas Kiye Gaye Items (Return)",
                     "#FED7D7","#FFF5F5",RED),
                    (new_items,"➕  Diye Gaye Naye Items",
                     "#C6F6D5","#F0FFF4",GREEN)
                ]:
                    story.append(HRFlowable(width="100%",thickness=1.5,color=tcol))
                    story.append(Paragraph(label,ps("lh",10,True,col=tcol)))
                    story.append(Spacer(1,1*mm))
                    rows2=[[str(i+1),it["product"],str(it["qty"]),
                            f"{it['rate']:.2f}",f"{it['total']:.0f}"] for i,it in enumerate(items)]
                    tamt=sum(it["total"] for it in items)
                    data=[["#","Product","Qty","Rate (₹)","Total (₹)"]]+rows2+\
                         [["","","","Total:",f"₹{round(tamt):,}"]]
                    sty=[("BOX",(0,0),(-1,-1),.5,BDR),("INNERGRID",(0,0),(-1,-1),.3,BDR),
                         ("BACKGROUND",(0,0),(-1,0),colors.HexColor(hbg)),
                         ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                         ("FONTNAME",(3,-1),(4,-1),"Helvetica-Bold"),
                         ("TEXTCOLOR",(3,-1),(4,-1),tcol),
                         ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor(rbg)),
                         ("ALIGN",(2,0),(-1,-1),"RIGHT"),
                         ("FONTSIZE",(0,0),(-1,-1),8),
                         ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
                         ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4)]
                    story.append(Table(data,colWidths=r_cw,style=TableStyle(sty)))
                    story.append(Spacer(1,4*mm))
                story.append(HRFlowable(width="100%",thickness=1.5,color=BROWN))
                diff=row["diff_amt"]; diff_abs=abs(diff)
                if diff>0:   dt=f"Customer ko ₹{round(diff_abs):,} aur dene honge"; dc=RED
                elif diff<0: dt=f"Customer ko ₹{round(diff_abs):,} wapas milenge (Refund)"; dc=GREEN
                else:        dt="Seedha Exchange — koi extra payment nahi"; dc=BROWN
                story.append(Spacer(1,2*mm))
                story.append(Paragraph(f"💰  {dt}",ps("d",11,True,TA_CENTER,dc)))
                story.append(HRFlowable(width="100%",thickness=1.5,color=BROWN))
                story.append(Spacer(1,14*mm))
                story.append(Table(
                    [["_____________________","_____________________"],
                     ["Customer Signature","Authorized Signature"]],
                    colWidths=[85*mm,85*mm],
                    style=TableStyle([("ALIGN",(0,0),(-1,-1),"CENTER"),
                                      ("FONTSIZE",(0,0),(-1,-1),8),
                                      ("TOPPADDING",(0,0),(-1,-1),2),
                                      ("BOTTOMPADDING",(0,0),(-1,-1),2)])))
                doc.build(story)
                import subprocess
                if os.name=="nt": os.startfile(save_path)
                else: subprocess.Popen(["xdg-open",save_path])
                messagebox.showinfo("Print Ready!","PDF ban gayi! Print karo.",parent=p)
            except ImportError:
                messagebox.showerror("Error","pip install reportlab",parent=p)
            except Exception as e:
                messagebox.showerror("Error",str(e),parent=p)

        def _delete_exchange(ex_id, ex_no):
            if not messagebox.askyesno("Delete?",
                    f"Exchange '{ex_no}' delete karna chahte hain?", parent=p):
                return
            db=get_db()
            db.execute("DELETE FROM exchange_items WHERE exchange_id=?",(ex_id,))
            db.execute("DELETE FROM exchanges WHERE id=?",(ex_id,))
            db.commit(); db.close(); load()

        def load():
            clear_table_rows(tbl)
            frm=v_from.get(); to=v_to.get(); party=v_party.get()
            db=get_db()
            q="SELECT * FROM exchanges WHERE 1=1"
            params=[]
            if frm:   q+=" AND exchange_date>=?"; params.append(frm)
            if to:    q+=" AND exchange_date<=?"; params.append(to)
            if party: q+=" AND party=?"; params.append(party)
            q+=" ORDER BY id DESC"
            try: rows=db.execute(q,params).fetchall()
            except: rows=[]
            db.close()
            for i,r in enumerate(rows):
                bg=C_WHITE if i%2==0 else "#F7FAFC"
                eid=r["id"]; eno=r["exchange_no"]; diff=r["diff_amt"]
                diff_str=f"₹{round(abs(diff)):,}"+(" ↑" if diff>0 else(" ↓" if diff<0 else" ="))
                vals=[eno,fmt_date(r["exchange_date"]),r["orig_bill"],r["party"][:16],
                      f"₹{round(r['return_total']):,}",f"₹{round(r['new_total']):,}",diff_str,"",""]
                for j,val in enumerate(vals):
                    if j==7:
                        tk.Button(tbl,text="🖨️",font=("Segoe UI",9),
                                  bg="#744210",fg=C_WHITE,relief="flat",cursor="hand2",bd=0,pady=3,
                                  command=lambda eid=eid: _reprint_exchange(eid)
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=1)
                    elif j==8:
                        tk.Button(tbl,text="🗑",font=("Segoe UI",9),
                                  bg=C_RED,fg=C_WHITE,relief="flat",cursor="hand2",bd=0,pady=3,
                                  command=lambda eid=eid,eno=eno: _delete_exchange(eid,eno)
                                  ).grid(row=i+1,column=j,sticky="nsew",padx=1,pady=1)
                    else:
                        tk.Label(tbl,text=str(val),font=("Segoe UI",9),
                                 bg=bg,fg=C_GRAY,anchor="w",padx=5,pady=4
                                 ).grid(row=i+1,column=j,sticky="nsew",padx=1)
            v_summary.set(f"Total Exchanges: {len(rows)}")

        load()

    # ══════════════════════════════════════════════════════════════════════════
    #  EXPIRY MANAGER — Complete Grocery Expiry Date Management
    # ══════════════════════════════════════════════════════════════════════════
    def _pg_expiry(self):
        p = tk.Frame(self.content, bg=C_BG, padx=14, pady=6)
        p.pack(fill="both", expand=True)

        # Header
        hdr = tk.Frame(p, bg=C_BG); hdr.pack(fill="x", pady=(0, 8))
        tk.Label(hdr, text="🗓  Expiry Manager — Medical Batches",
                 font=("Segoe UI", 15, "bold"), bg=C_BG, fg="#1A365D").pack(side="left")
        make_btn(hdr, "➕ Nayi Batch Add Karo", lambda: _open_batch_form(),
                 bg=C_GREEN).pack(side="right", padx=4)
        make_btn(hdr, "📊 Excel Export", lambda: _export_excel(),
                 bg="#2B6CB0").pack(side="right", padx=4)

        # Alert Banner — Expiry
        alerts = get_expiry_alerts()
        if alerts:
            banner = tk.Frame(p, bg="#FFF5F5", highlightthickness=1, highlightbackground="#FC8181")
            banner.pack(fill="x", pady=(0, 4))
            today_key_val = exp_key_for_date(datetime.date.today())
            exp_today = [a for a in alerts if exp_sort_key(a["expiry_date"]) <= today_key_val]
            exp_soon  = [a for a in alerts if exp_sort_key(a["expiry_date"]) > today_key_val]
            msg_parts = []
            if exp_today: msg_parts.append(f"⛔ {len(exp_today)} batch(es) EXPIRE HO GAYI!")
            if exp_soon:  msg_parts.append(f"⚠️ {len(exp_soon)} batch(es) 30 din mein expire hongi")
            tk.Label(banner, text="   " + "   |   ".join(msg_parts),
                     font=("Segoe UI", 10, "bold"), bg="#FFF5F5", fg="#9B2C2C",
                     pady=4).pack(side="left")

        # Alert Banner — Low Stock
        low_items2 = get_low_stock_alerts()
        if low_items2:
            low_banner = tk.Frame(p, bg="#FFFBEB", highlightthickness=1, highlightbackground="#F6AD55")
            low_banner.pack(fill="x", pady=(0, 8))
            zero2 = sum(1 for x in low_items2 if (x["stock"] or 0) <= 0)
            low2  = len(low_items2) - zero2
            parts2 = []
            if zero2: parts2.append(f"🚨 {zero2} product OUT OF STOCK!")
            if low2:  parts2.append(f"⚠️ {low2} product ka stock low hai")
            tk.Label(low_banner, text="   " + "   |   ".join(parts2),
                     font=("Segoe UI", 10, "bold"), bg="#FFFBEB", fg="#744210",
                     pady=4).pack(side="left")

        # Filter Bar
        fbar = tk.Frame(p, bg=C_BG); fbar.pack(fill="x", pady=(0, 6))
        tk.Label(fbar, text="🔍 Product:", font=("Segoe UI", 9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_search = tk.StringVar()
        ttk.Entry(fbar, textvariable=v_search, width=20, font=("Segoe UI",9)).pack(side="left", padx=(4, 12))
        tk.Label(fbar, text="Filter:", font=("Segoe UI", 9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_filter = tk.StringVar(value="Sab Dikhao")
        filter_cb = ttk.Combobox(fbar, textvariable=v_filter, width=18, state="readonly",
                                 values=["Sab Dikhao", "Expire Ho Gayi", "7 Din Mein", "30 Din Mein", "Safe Stock"])
        filter_cb.pack(side="left", padx=(4, 12))
        tk.Label(fbar, text="Supplier:", font=("Segoe UI", 9), bg=C_BG, fg=C_GRAY).pack(side="left")
        v_supp = tk.StringVar(value="Sab")
        _conn_tmp = get_db()
        _supp_list = ["Sab"] + sorted({r[0] for r in _conn_tmp.execute(
            "SELECT DISTINCT supplier FROM expiry_stock WHERE supplier != ''").fetchall()})
        _conn_tmp.close()
        supp_cb = ttk.Combobox(fbar, textvariable=v_supp, width=18, state="readonly", values=_supp_list)
        supp_cb.pack(side="left", padx=(4, 12))
        make_btn(fbar, "🔄 Refresh", lambda: _load_table(), bg=C_ACCENT).pack(side="left", padx=4)

        # Summary Cards
        cards_row = tk.Frame(p, bg=C_BG); cards_row.pack(fill="x", pady=(0, 8))
        def _get_summary():
            conn = get_db()
            today_k = exp_key_for_date(datetime.date.today())
            d7_k    = exp_key_for_date(datetime.date.today() + datetime.timedelta(days=7))
            d30_k   = exp_key_for_date(datetime.date.today() + datetime.timedelta(days=30))
            expired = conn.execute("SELECT COALESCE(SUM(qty),0) FROM expiry_stock WHERE qty>0 AND " + EXP_KEY_SQL + "<=?", (today_k,)).fetchone()[0]
            soon7   = conn.execute("SELECT COALESCE(SUM(qty),0) FROM expiry_stock WHERE qty>0 AND " + EXP_KEY_SQL + ">? AND " + EXP_KEY_SQL + "<=?", (today_k, d7_k)).fetchone()[0]
            soon30  = conn.execute("SELECT COALESCE(SUM(qty),0) FROM expiry_stock WHERE qty>0 AND " + EXP_KEY_SQL + ">? AND " + EXP_KEY_SQL + "<=?", (today_k, d30_k)).fetchone()[0]
            safe    = conn.execute("SELECT COALESCE(SUM(qty),0) FROM expiry_stock WHERE qty>0 AND " + EXP_KEY_SQL + ">?", (d30_k,)).fetchone()[0]
            at_risk = conn.execute("SELECT COALESCE(SUM(qty*mrp),0) FROM expiry_stock WHERE qty>0 AND " + EXP_KEY_SQL + "<=?", (d30_k,)).fetchone()[0]
            conn.close()
            return expired, soon7, soon30, safe, at_risk
        expired_qty, s7, s30, safe_qty, at_risk_val = _get_summary()
        for title, val, color in [
            ("⛔ Expire Ho Gayi",  f"{expired_qty:.0f} Units", C_RED),
            ("⚠️ 7 Din Mein",     f"{s7:.0f} Units",          "#C05621"),
            ("⚠️ 30 Din Mein",    f"{s30:.0f} Units",         C_AMBER),
            ("✅ Safe Stock",      f"{safe_qty:.0f} Units",    C_GREEN),
            ("💰 At-Risk Value",   f"Rs.{at_risk_val:,.0f}",  C_RED if at_risk_val > 0 else C_GREEN),
        ]:
            stat_card(cards_row, title, val, color)

        # Table
        cols = ["#", "Product", "Batch No", "Mfg Date", "Expiry Date",
                "Qty", "Pur Rate", "MRP", "Supplier", "Status", "Actions"]
        col_w = [3, 20, 12, 10, 10, 6, 8, 8, 14, 14, 24]
        tbl_inner = make_table(p, cols, col_w)
        rows_data = []

        def _status_info(expiry_str):
            try:
                import calendar as _cal
                mm, yy = expiry_str.split("/")
                mm = int(mm); yyyy = 2000 + int(yy)
                last_day = _cal.monthrange(yyyy, mm)[1]
                exp   = datetime.date(yyyy, mm, last_day)
                today = datetime.date.today()
                delta = (exp - today).days
                if delta < 0:  return f"⛔ {abs(delta)}d Pehle Expire", "#9B2C2C", "#FFF5F5"
                elif delta == 0: return "⛔ AAJ Expire!", "#9B2C2C", "#FFF5F5"
                elif delta <= 7:  return f"⚠️ {delta}d Bacha", "#744210", "#FFFBEB"
                elif delta <= 30: return f"⚠️ {delta}d Bacha", "#975A16", "#FEFCE8"
                else:             return f"✅ {delta}d Bacha", "#276749", "#F0FFF4"
            except:
                return "?", C_GRAY, C_WHITE

        def _load_table(*_):
            clear_table_rows(tbl_inner)
            rows_data.clear()
            conn = get_db()
            query = "SELECT * FROM expiry_stock WHERE 1=1"
            params = []
            today_k = exp_key_for_date(datetime.date.today())
            d7_k    = exp_key_for_date(datetime.date.today() + datetime.timedelta(days=7))
            d30_k   = exp_key_for_date(datetime.date.today() + datetime.timedelta(days=30))
            flt = v_filter.get()
            if flt == "Expire Ho Gayi":
                query += " AND " + EXP_KEY_SQL + "<=?"; params.append(today_k)
            elif flt == "7 Din Mein":
                query += " AND " + EXP_KEY_SQL + ">? AND " + EXP_KEY_SQL + "<=?"; params += [today_k, d7_k]
            elif flt == "30 Din Mein":
                query += " AND " + EXP_KEY_SQL + ">? AND " + EXP_KEY_SQL + "<=?"; params += [today_k, d30_k]
            elif flt == "Safe Stock":
                query += " AND " + EXP_KEY_SQL + ">?"; params.append(d30_k)
            srch = v_search.get().strip().lower()
            if srch:
                query += " AND LOWER(product) LIKE ?"; params.append(f"%{srch}%")
            supp = v_supp.get().strip()
            if supp and supp != "Sab":
                query += " AND supplier=?"; params.append(supp)
            query += " ORDER BY " + EXP_KEY_SQL + " ASC"
            rows = [dict(r) for r in conn.execute(query, params).fetchall()]
            conn.close()
            if not rows:
                tk.Label(tbl_inner, text="Koi record nahi mila.", font=("Segoe UI",9),
                         bg=C_WHITE, fg=C_GRAY, pady=16).grid(
                         row=1, column=0, columnspan=len(cols), sticky="w", padx=16)
                return
            for i, row in enumerate(rows, 1):
                status_txt, status_fg, row_bg = _status_info(row["expiry_date"])
                rows_data.append(row)
                vals = [i, row["product"], row["batch_no"] or "—",
                        fmt_date(row["mfg_date"]) if row["mfg_date"] else "—",
                        fmt_exp_mmyy(row["expiry_date"]),
                        f"{row['qty']:.0f}", f"Rs.{row['purchase_rate']:.2f}",
                        f"Rs.{row['mrp']:.2f}", row["supplier"] or "—", status_txt]
                fgs = [None]*9 + [status_fg, None]
                for ci, val in enumerate(vals):
                    tk.Label(tbl_inner, text=str(val), font=("Segoe UI", 9),
                             bg=row_bg, fg=fgs[ci] if fgs[ci] else C_GRAY,
                             anchor="w", padx=8, pady=5
                             ).grid(row=i, column=ci, sticky="nsew", padx=1, pady=0)
                act_frame = tk.Frame(tbl_inner, bg=row_bg)
                act_frame.grid(row=i, column=10, sticky="nsew", padx=1, pady=0)
                def _make_actions(r=row, rb=row_bg):
                    tk.Button(act_frame, text="Edit", font=("Segoe UI",7),
                              bg=rb, fg="#2B6CB0", relief="flat", cursor="hand2",
                              padx=4, command=lambda: _open_batch_form(r)).pack(side="left")
                    tk.Button(act_frame, text="Del", font=("Segoe UI",7),
                              bg=rb, fg=C_RED, relief="flat", cursor="hand2",
                              padx=4, command=lambda: _delete_batch(r)).pack(side="left")
                    tk.Button(act_frame, text="Qty", font=("Segoe UI",7),
                              bg=rb, fg=C_GREEN, relief="flat", cursor="hand2",
                              padx=4, command=lambda: _adjust_qty(r)).pack(side="left")
                    tk.Button(act_frame, text="↩ Return", font=("Segoe UI", 8, "bold"),
                              bg="#6B46C1", fg="white", relief="flat", cursor="hand2",
                              padx=6, pady=2, bd=0,
                              activebackground="#553C9A", activeforeground="white",
                              command=lambda: _return_to_supplier(r)).pack(side="left", padx=(4,2), pady=2)
                    tk.Button(act_frame, text="🗑 Write Off", font=("Segoe UI", 8, "bold"),
                              bg="#9B2C2C", fg="white", relief="flat", cursor="hand2",
                              padx=6, pady=2, bd=0,
                              activebackground="#742A2A", activeforeground="white",
                              command=lambda: _write_off(r)).pack(side="left", padx=(2,4), pady=2)
                _make_actions()

        def _delete_batch(row):
            if not messagebox.askyesno("Confirm Delete",
                    f"Batch delete karoge?\nProduct: {row['product']}\n"
                    f"Expiry: {fmt_exp_mmyy(row['expiry_date'])}\nQty: {row['qty']}\n\nYeh undo nahi hoga!"):
                return
            conn = get_db()
            conn.execute("DELETE FROM expiry_stock WHERE id=?", (row["id"],))
            conn.commit(); conn.close()
            _load_table()
            messagebox.showinfo("Deleted!", "Batch delete ho gayi!")

        def _return_to_supplier(row):
            """Expired/near-expiry product supplier ko return karo — stock se ghata do + chalan print."""
            dlg = tk.Toplevel(p.winfo_toplevel())
            dlg.title("Return to Supplier")
            dlg.configure(bg=C_WHITE)
            dlg.resizable(False, False)
            dlg.geometry("500x520"); dlg.grab_set()

            tk.Label(dlg, text="\u21a9\ufe0f  Return to Supplier",
                     font=("Segoe UI", 12, "bold"), bg="#6B46C1", fg="white", pady=4).pack(fill="x")

            body = tk.Frame(dlg, bg=C_WHITE, padx=22, pady=4)
            body.pack(fill="both", expand=True)

            def _lbl(parent, text, bold=False):
                tk.Label(parent, text=text,
                         font=("Segoe UI", 9, "bold" if bold else "normal"),
                         bg=C_WHITE, fg=C_GRAY, anchor="w").pack(fill="x", pady=(2,0))

            info_box = tk.Frame(body, bg="#F3F0FF",
                                highlightthickness=1, highlightbackground="#6B46C1")
            info_box.pack(fill="x", pady=(0,3))
            tk.Label(info_box, text=f"  {row['product']}",
                     font=("Segoe UI", 10, "bold"), bg="#F3F0FF", fg="#44337A",
                     anchor="w", pady=4).pack(fill="x")
            tk.Label(info_box,
                     text=f"  Batch: {row['batch_no'] or '—'}   |   Expiry: {fmt_exp_mmyy(row['expiry_date'])}   |   Stock: {row['qty']:.0f} units",
                     font=("Segoe UI",7), bg="#F3F0FF", fg="#553C9A",
                     anchor="w", pady=2).pack(fill="x")

            tk.Frame(body, bg=C_BORDER, height=1).pack(fill="x", pady=6)

            r1 = tk.Frame(body, bg=C_WHITE); r1.pack(fill="x", pady=2)
            f_qty = tk.Frame(r1, bg=C_WHITE); f_qty.pack(side="left", fill="x", expand=True, padx=(0,10))
            _lbl(f_qty, "Return Qty *", bold=True)
            v_rqty = tk.StringVar(value=str(int(row["qty"])))
            ttk.Entry(f_qty, textvariable=v_rqty, width=12,
                      font=("Segoe UI",9)).pack(fill="x", ipady=2)

            f_date = tk.Frame(r1, bg=C_WHITE); f_date.pack(side="left", fill="x", expand=True)
            _lbl(f_date, "Return Date *", bold=True)
            v_rdate = tk.StringVar(value=str(datetime.date.today()))
            make_date_entry(f_date, v_rdate, width=14, bg=C_WHITE).pack(anchor="w")

            _lbl(body, "Supplier Name")
            conn_s2 = get_db()
            supp_names2 = [""] + sorted({r2[0] for r2 in conn_s2.execute(
                "SELECT DISTINCT party FROM purchases ORDER BY party").fetchall()})
            conn_s2.close()
            v_supp2 = tk.StringVar(value=row["supplier"] or "")
            ttk.Combobox(body, textvariable=v_supp2, values=supp_names2,
                         width=42, font=("Segoe UI",9)).pack(fill="x", ipady=1)

            _lbl(body, "Reason / Note")
            v_note = tk.StringVar(value="Expiry return")
            ttk.Combobox(body, textvariable=v_note, width=42,
                         values=["Expiry return", "Near expiry", "Quality issue",
                                 "Damaged packaging", "Wrong product"]).pack(fill="x", ipady=1)

            _lbl(body, "Chalan / Debit Note No (optional)")
            v_chalan = tk.StringVar()
            ttk.Entry(body, textvariable=v_chalan, width=22,
                      font=("Segoe UI",9)).pack(anchor="w", ipady=1)

            err_lbl = tk.Label(body, text="", font=("Segoe UI", 9, "bold"),
                               bg=C_WHITE, fg=C_RED, anchor="w")
            err_lbl.pack(fill="x", pady=(4,0))

            saved_data = {}

            def do_save_return(print_chalan=False):
                err_lbl.config(text="")
                try:
                    rqty = float(v_rqty.get().strip())
                    if rqty <= 0: err_lbl.config(text="\u274c Qty 0 se zyada honi chahiye!"); return
                    if rqty > row["qty"]:
                        err_lbl.config(text=f"\u274c Batch mein sirf {row['qty']:.0f} units hai!"); return
                except:
                    err_lbl.config(text="\u274c Qty mein sirf number likhein!"); return

                rdate    = v_rdate.get().strip()
                supplier = v_supp2.get().strip()
                note     = v_note.get().strip() or "Expiry return"
                chalan   = v_chalan.get().strip()
                amt      = round(rqty * row["purchase_rate"], 2)
                new_qty  = row["qty"] - rqty

                conn2 = get_db()
                if new_qty <= 0:
                    conn2.execute("DELETE FROM expiry_stock WHERE id=?", (row["id"],))
                else:
                    conn2.execute("UPDATE expiry_stock SET qty=? WHERE id=?",
                                  (new_qty, row["id"]))
                conn2.execute(
                    "INSERT INTO stock_movements(move_date,product,move_type,qty,ref_no,reason)"
                    " VALUES(?,?,?,?,?,?)",
                    (rdate, row["product"], "Return to Supplier", -rqty,
                     chalan or supplier, note))
                conn2.commit(); conn2.close()

                saved_data.update(dict(
                    product=row["product"], batch_no=row["batch_no"] or "",
                    expiry_date=row["expiry_date"], supplier=supplier,
                    qty=rqty, rate=row["purchase_rate"], amt=amt,
                    rdate=rdate, note=note, chalan=chalan,
                ))
                dlg.destroy(); _load_table()
                if print_chalan:
                    _print_return_chalan(saved_data)
                else:
                    messagebox.showinfo("\u2705 Return Save Ho Gayi!",
                        f"Product : {row['product']}\n"
                        f"Qty     : {rqty:.0f} units\n"
                        f"Supplier: {supplier or '—'}\n"
                        f"Amount  : Rs.{amt:,.2f}\n"
                        f"Remaining Batch: {max(0,new_qty):.0f} units")

            tk.Frame(dlg, bg=C_BORDER, height=1).pack(fill="x")
            btn_bar = tk.Frame(dlg, bg="#F8F8F8", pady=4); btn_bar.pack(fill="x", side="bottom")
            make_btn(btn_bar, "\U0001f4be Save Karo",
                     lambda: do_save_return(False), bg="#6B46C1").pack(side="left", padx=(16,6))
            make_btn(btn_bar, "\U0001f5a8\ufe0f Save + Chalan Print",
                     lambda: do_save_return(True),  bg="#276749").pack(side="left", padx=6)
            make_btn(btn_bar, "Cancel",
                     dlg.destroy, bg=C_GRAY).pack(side="left", padx=6)

        def _print_return_chalan(d):
            """Return to Supplier ka Chalan / Debit Note print window."""
            win = tk.Toplevel(p.winfo_toplevel())
            win.title(f"Return Chalan \u2014 {d['product']}")
            win.configure(bg=C_WHITE)
            win.geometry("680x600"); win.grab_set()
            _apply_logo(win)
            _sp = get_shop()
            today_str = fmt_date(d["rdate"])
            chalan_no = d["chalan"] or ("RC-" + d["rdate"].replace("-",""))

            tb = tk.Frame(win, bg="#276749", pady=6); tb.pack(fill="x")
            tk.Label(tb, text="  \U0001f5a8\ufe0f Return Chalan / Debit Note",
                     font=("Segoe UI", 11, "bold"), bg="#276749", fg="white").pack(side="left", padx=8)

            def do_print():
                import tempfile, webbrowser
                html = _build_chalan_html(d, _sp, chalan_no, today_str)
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".html",
                                                  mode="w", encoding="utf-8")
                tmp.write(html); tmp.close()
                webbrowser.open("file://" + tmp.name)

            make_btn(tb, "\U0001f5a8\ufe0f Print / PDF", do_print, bg="#2B6CB0").pack(side="right", padx=8, pady=2)
            make_btn(tb, "Close", win.destroy, bg=C_GRAY).pack(side="right", padx=4, pady=2)

            outer = tk.Frame(win, bg="#E0E0E0"); outer.pack(fill="both", expand=True, padx=10, pady=4)
            canvas = tk.Canvas(outer, bg="#E0E0E0", highlightthickness=0)
            vsb    = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=vsb.set)
            vsb.pack(side="right", fill="y"); canvas.pack(fill="both", expand=True)
            canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>",
                lambda ev: canvas.yview_scroll(-1*(ev.delta//120), "units")))
            canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))
            paper = tk.Frame(canvas, bg=C_WHITE, width=620)
            wid = canvas.create_window((0, 10), window=paper, anchor="nw")
            canvas.bind("<Configure>", lambda e: (
                canvas.itemconfig(wid, width=min(620, e.width-20)),
                canvas.configure(scrollregion=canvas.bbox("all"))))
            paper.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

            pad = tk.Frame(paper, bg=C_WHITE, padx=24, pady=16); pad.pack(fill="x")

            title_f = tk.Frame(pad, bg="#6B46C1"); title_f.pack(fill="x")
            tk.Label(title_f, text="RETURN TO SUPPLIER / DEBIT NOTE",
                     font=("Arial", 12, "bold"), bg="#6B46C1", fg="white",
                     anchor="center", pady=6).pack(fill="x")

            meta = tk.Frame(pad, bg=C_WHITE); meta.pack(fill="x", pady=(6,0))
            meta.columnconfigure(0, weight=3); meta.columnconfigure(1, weight=2)

            lf2 = tk.Frame(meta, bg=C_WHITE, highlightthickness=1, highlightbackground="#AAAAAA")
            lf2.grid(row=0, column=0, sticky="nsew")
            tk.Label(lf2, text=_sp["name"], font=("Arial", 10, "bold"),
                     bg=C_WHITE, anchor="w", padx=6, pady=3).pack(fill="x")
            for line in [(_sp["address"]+", "+_sp["city"]).strip(", "),
                         "GSTIN: "+(_sp["gstin"] or "\u2014"),
                         "Mob: "+(_sp["mobile"] or "\u2014")]:
                if line.strip(" ,\u2014:"):
                    tk.Label(lf2, text=line, font=("Arial", 8),
                             bg=C_WHITE, anchor="w", padx=6).pack(fill="x")
            tk.Label(lf2, text="", bg=C_WHITE, pady=2).pack()

            rf2 = tk.Frame(meta, bg=C_WHITE, highlightthickness=1, highlightbackground="#AAAAAA")
            rf2.grid(row=0, column=1, sticky="nsew")
            for lt, vt in [("Chalan No :", chalan_no),
                           ("Date      :", today_str),
                           ("Supplier  :", d["supplier"] or "\u2014"),
                           ("Reason    :", d["note"])]:
                rf3 = tk.Frame(rf2, bg=C_WHITE); rf3.pack(fill="x", padx=6, pady=2)
                tk.Label(rf3, text=lt, font=("Arial", 8, "bold"),
                         bg=C_WHITE, width=12, anchor="w").pack(side="left")
                tk.Label(rf3, text=vt, font=("Arial", 8),
                         bg=C_WHITE, anchor="w").pack(side="left")
            tk.Label(rf2, text="", bg=C_WHITE, pady=2).pack()

            tbl_f = tk.Frame(pad, bg=C_WHITE); tbl_f.pack(fill="x", pady=(10,0))
            COLS2 = ["Sl.", "Product", "Batch No", "Expiry", "Qty", "Rate", "Amount"]
            WGTS  = [2, 10, 7, 7, 4, 7, 7]
            for ci, w in enumerate(WGTS): tbl_f.columnconfigure(ci, weight=w)
            HDR_BG2 = "#EDE9FE"
            for ci, c in enumerate(COLS2):
                hf3 = tk.Frame(tbl_f, bg=HDR_BG2, highlightthickness=1, highlightbackground="#AAAAAA")
                hf3.grid(row=0, column=ci, sticky="nsew")
                tk.Label(hf3, text=c, font=("Arial", 8, "bold"), bg=HDR_BG2,
                         anchor="center", padx=4, pady=5).pack(fill="both")

            row_bg2 = "#FAFAFA"
            vals2 = ["1", d["product"], d["batch_no"] or "\u2014",
                     fmt_exp_mmyy(d["expiry_date"]),
                     f"{d['qty']:.0f}", f"Rs.{d['rate']:.2f}", f"Rs.{d['amt']:.2f}"]
            aligns = ["w","w","w","w","e","e","e"]
            for ci, (val, al) in enumerate(zip(vals2, aligns)):
                cf = tk.Frame(tbl_f, bg=row_bg2, highlightthickness=1, highlightbackground="#DDDDDD")
                cf.grid(row=1, column=ci, sticky="nsew")
                tk.Label(cf, text=str(val), font=("Arial", 9),
                         bg=row_bg2, anchor=al, padx=6, pady=6).pack(fill="both")

            tot_f = tk.Frame(pad, bg=C_WHITE); tot_f.pack(fill="x", pady=(4,0))
            tf4 = tk.Frame(tot_f, bg="#EDE9FE",
                           highlightthickness=1, highlightbackground="#AAAAAA")
            tf4.pack(side="right")
            for lt2, vt2 in [("Total Qty Returned :", f"{d['qty']:.0f} units"),
                             ("Total Amount       :", f"Rs. {d['amt']:,.2f}")]:
                tf5 = tk.Frame(tf4, bg="#EDE9FE"); tf5.pack(fill="x", padx=10, pady=3)
                tk.Label(tf5, text=lt2, font=("Arial", 9, "bold"),
                         bg="#EDE9FE", width=22, anchor="w").pack(side="left")
                tk.Label(tf5, text=vt2, font=("Arial", 9),
                         bg="#EDE9FE", anchor="e").pack(side="left")

            foot = tk.Frame(pad, bg=C_WHITE); foot.pack(fill="x", pady=(24,4))
            foot.columnconfigure(0, weight=1); foot.columnconfigure(1, weight=1)
            tk.Label(foot, text="Received By\n\n__________________",
                     font=("Arial", 8), bg=C_WHITE, anchor="w").grid(row=0, column=0, sticky="w")
            tk.Label(foot, text="For  "+_sp["name"]+"\n\n__________________",
                     font=("Arial", 8), bg=C_WHITE, anchor="e").grid(row=0, column=1, sticky="e")

            canvas.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _build_chalan_html(d, _sp, chalan_no, today_str):
            shop_addr = (_sp["address"]+", "+_sp["city"]).strip(", ")
            return (
                "<!DOCTYPE html><html><head><meta charset='utf-8'>"
                "<title>Return Chalan</title><style>"
                "body{font-family:Arial,sans-serif;font-size:12px;margin:24px}"
                "h2{text-align:center;background:#6B46C1;color:#fff;padding:8px;margin:0}"
                ".meta{display:flex;border:1px solid #aaa;margin-top:6px}"
                ".shop{flex:3;padding:8px;border-right:1px solid #aaa}"
                ".info{flex:2;padding:8px}"
                "table{width:100%;border-collapse:collapse;margin-top:10px}"
                "th{background:#EDE9FE;border:1px solid #aaa;padding:6px;text-align:center}"
                "td{border:1px solid #ccc;padding:6px}"
                ".total-box{float:right;margin-top:6px;border:1px solid #aaa;"
                "padding:8px 14px;background:#EDE9FE}"
                ".foot{display:flex;justify-content:space-between;margin-top:40px;font-size:11px}"
                "@media print{body{margin:8px}}"
                "</style></head><body>"
                f"<h2>RETURN TO SUPPLIER / DEBIT NOTE</h2>"
                f"<div class='meta'>"
                f"<div class='shop'><b>{_sp['name']}</b><br>{shop_addr}<br>"
                f"GSTIN: {_sp['gstin'] or '—'} &nbsp; Mob: {_sp['mobile'] or '—'}</div>"
                f"<div class='info'>"
                f"<b>Chalan No :</b> {chalan_no}<br>"
                f"<b>Date &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:</b> {today_str}<br>"
                f"<b>Supplier &nbsp;:</b> {d['supplier'] or '—'}<br>"
                f"<b>Reason &nbsp;&nbsp;&nbsp;:</b> {d['note']}"
                f"</div></div>"
                f"<table><tr><th>Sl.</th><th>Product</th><th>Batch No</th>"
                f"<th>Expiry</th><th>Qty</th><th>Rate (Rs.)</th><th>Amount (Rs.)</th></tr>"
                f"<tr><td>1</td><td>{d['product']}</td><td>{d['batch_no'] or '—'}</td>"
                f"<td>{fmt_exp_mmyy(d['expiry_date'])}</td>"
                f"<td style='text-align:right'>{d['qty']:.0f}</td>"
                f"<td style='text-align:right'>{d['rate']:.2f}</td>"
                f"<td style='text-align:right'>{d['amt']:.2f}</td></tr></table>"
                f"<div class='total-box'>"
                f"<b>Total Qty Returned :</b> {d['qty']:.0f} units<br>"
                f"<b>Total Amount &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:</b> <b>Rs. {d['amt']:,.2f}</b>"
                f"</div>"
                f"<div style='clear:both;padding-top:8px;font-size:11px;color:#555'>"
                f"Note: Please issue Credit Note / adjust in next invoice accordingly.</div>"
                f"<div class='foot'>"
                f"<span>Received By<br><br>__________________</span>"
                f"<span style='text-align:right'>For {_sp['name']}<br><br>__________________</span>"
                f"</div>"
                f"<script>window.onload=function(){{window.print()}}</script>"
                f"</body></html>"
            )

        def _write_off(row):
            """Expired product write off karo — stock se hata do, loss mein daalo."""
            dlg = tk.Toplevel(p.winfo_toplevel())
            dlg.title("Write Off \u2014 Expired Stock")
            dlg.configure(bg=C_WHITE)
            dlg.resizable(False, False)
            dlg.geometry("480x460"); dlg.grab_set()

            tk.Label(dlg, text="\U0001f5d1\ufe0f  Write Off (Expired/Damaged Stock)",
                     font=("Segoe UI", 12, "bold"), bg="#9B2C2C", fg="white", pady=4).pack(fill="x")

            body = tk.Frame(dlg, bg=C_WHITE, padx=22, pady=4)
            body.pack(fill="both", expand=True)

            def _lbl2(parent, text, bold=False):
                tk.Label(parent, text=text,
                         font=("Segoe UI", 9, "bold" if bold else "normal"),
                         bg=C_WHITE, fg=C_GRAY, anchor="w").pack(fill="x", pady=(2,0))

            total_loss = round(row["qty"] * row["purchase_rate"], 2)
            info_box2 = tk.Frame(body, bg="#FFF5F5",
                                 highlightthickness=1, highlightbackground="#9B2C2C")
            info_box2.pack(fill="x", pady=(0,3))
            tk.Label(info_box2, text=f"  {row['product']}",
                     font=("Segoe UI", 10, "bold"), bg="#FFF5F5", fg="#9B2C2C",
                     anchor="w", pady=4).pack(fill="x")
            tk.Label(info_box2,
                     text=f"  Batch: {row['batch_no'] or '—'}   |   Expiry: {fmt_exp_mmyy(row['expiry_date'])}   |   Qty: {row['qty']:.0f}   |   Rate: Rs.{row['purchase_rate']:.2f}",
                     font=("Segoe UI",7), bg="#FFF5F5", fg="#742A2A",
                     anchor="w", pady=2).pack(fill="x")
            tk.Label(info_box2, text=f"  Full Batch Loss: Rs.{total_loss:,.2f}",
                     font=("Segoe UI", 9, "bold"), bg="#FFF5F5", fg="#9B2C2C",
                     anchor="w", pady=2).pack(fill="x")

            tk.Frame(body, bg=C_BORDER, height=1).pack(fill="x", pady=6)

            r1w = tk.Frame(body, bg=C_WHITE); r1w.pack(fill="x", pady=2)
            f_wqty = tk.Frame(r1w, bg=C_WHITE); f_wqty.pack(side="left", fill="x", expand=True, padx=(0,10))
            _lbl2(f_wqty, "Write Off Qty *", bold=True)
            v_wqty = tk.StringVar(value=str(int(row["qty"])))
            ttk.Entry(f_wqty, textvariable=v_wqty, width=12,
                      font=("Segoe UI",9)).pack(fill="x", ipady=2)

            f_wdate = tk.Frame(r1w, bg=C_WHITE); f_wdate.pack(side="left", fill="x", expand=True)
            _lbl2(f_wdate, "Write Off Date *", bold=True)
            v_wdate = tk.StringVar(value=str(datetime.date.today()))
            make_date_entry(f_wdate, v_wdate, width=14, bg=C_WHITE).pack(anchor="w")

            _lbl2(body, "Reason / Note")
            v_wnote = tk.StringVar(value="Expiry write off")
            ttk.Combobox(body, textvariable=v_wnote, width=42,
                         values=["Expiry write off", "Damaged goods", "Quality fail",
                                 "Customer complaint", "Storage loss"]).pack(fill="x", ipady=1)

            loss_lbl = tk.Label(body, text="", font=("Segoe UI", 10, "bold"),
                                bg="#FFF5F5", fg="#9B2C2C", anchor="w", padx=10, pady=6)
            loss_lbl.pack(fill="x", pady=(8,0))

            def update_loss(*_):
                try:
                    q = float(v_wqty.get() or 0)
                    loss = round(q * row["purchase_rate"], 2)
                    loss_lbl.config(
                        text=f"\U0001f4b8  Loss: Rs.{loss:,.2f}  ({q:.0f} units \xd7 Rs.{row['purchase_rate']:.2f})")
                except:
                    loss_lbl.config(text="")
            v_wqty.trace_add("write", update_loss); update_loss()

            err_lbl = tk.Label(body, text="", font=("Segoe UI", 9, "bold"),
                               bg=C_WHITE, fg=C_RED, anchor="w")
            err_lbl.pack(fill="x", pady=(4,0))

            def do_writeoff():
                err_lbl.config(text="")
                try:
                    wqty = float(v_wqty.get().strip())
                    if wqty <= 0: err_lbl.config(text="\u274c Qty 0 se zyada honi chahiye!"); return
                    if wqty > row["qty"]:
                        err_lbl.config(text=f"\u274c Batch mein sirf {row['qty']:.0f} units hai!"); return
                except:
                    err_lbl.config(text="\u274c Qty mein sirf number likhein!"); return

                loss_amt = round(wqty * row["purchase_rate"], 2)
                if not messagebox.askyesno("Confirm Write Off",
                        f"Pakka Write Off karoge?\n\n"
                        f"Product : {row['product']}\n"
                        f"Qty     : {wqty:.0f} units\n"
                        f"Loss    : Rs.{loss_amt:,.2f}\n\n"
                        f"Yeh stock permanent hatega!", parent=dlg):
                    return

                new_qty = row["qty"] - wqty
                conn3 = get_db()
                if new_qty <= 0:
                    conn3.execute("DELETE FROM expiry_stock WHERE id=?", (row["id"],))
                else:
                    conn3.execute("UPDATE expiry_stock SET qty=? WHERE id=?", (new_qty, row["id"]))
                conn3.execute(
                    "INSERT INTO stock_movements(move_date,product,move_type,qty,ref_no,reason)"
                    " VALUES(?,?,?,?,?,?)",
                    (v_wdate.get(), row["product"], "Write Off", -wqty,
                     "", v_wnote.get().strip() or "Expiry write off"))
                try:
                    conn3.execute(
                        "INSERT INTO expenses(exp_date,category,description,amount,pay_mode,ref_no)"
                        " VALUES(?,?,?,?,?,?)",
                        (v_wdate.get(), "Stock Write Off",
                         f"{row['product']} \u2014 {wqty:.0f} units expired/damaged write off",
                         loss_amt, "N/A", f"Batch: {row['batch_no'] or '—'}"))
                except Exception:
                    pass
                conn3.commit(); conn3.close()
                dlg.destroy(); _load_table()
                messagebox.showinfo("\u2705 Write Off Ho Gaya!",
                    f"Product : {row['product']}\n"
                    f"Qty     : {wqty:.0f} units write off ho gayi\n"
                    f"Loss    : Rs.{loss_amt:,.2f}\n"
                    f"(Expenses mein bhi record ho gaya)")

            tk.Frame(dlg, bg=C_BORDER, height=1).pack(fill="x")
            btn_bar2 = tk.Frame(dlg, bg="#F8F8F8", pady=4); btn_bar2.pack(fill="x", side="bottom")
            make_btn(btn_bar2, "\U0001f5d1\ufe0f Write Off Karo",
                     do_writeoff, bg="#9B2C2C").pack(side="left", padx=(16,6))
            make_btn(btn_bar2, "Cancel",
                     dlg.destroy, bg=C_GRAY).pack(side="left", padx=6)

        def _adjust_qty(row):
            dlg = tk.Toplevel(p.winfo_toplevel())
            dlg.title("Qty Adjust Karo")
            dlg.configure(bg=C_WHITE)
            dlg.resizable(False, False)
            dlg.geometry("380x280"); dlg.grab_set()
            tk.Label(dlg, text="Qty Adjust Karo", font=("Segoe UI", 12, "bold"),
                     bg="#1A365D", fg="white", pady=4).pack(fill="x")
            frm = tk.Frame(dlg, bg=C_WHITE, padx=14, pady=6); frm.pack(fill="x")
            tk.Label(frm, text=f"Product: {row['product']}", font=("Segoe UI", 10, "bold"),
                     bg=C_WHITE, fg="#1A365D").pack(anchor="w")
            tk.Label(frm, text=f"Batch: {row['batch_no'] or 'N/A'}  |  Expiry: {fmt_exp_mmyy(row['expiry_date'])}",
                     font=("Segoe UI", 9), bg=C_WHITE, fg=C_GRAY).pack(anchor="w", pady=(2, 8))
            tk.Label(frm, text=f"Current Qty: {row['qty']:.0f}", font=("Segoe UI",9),
                     bg=C_WHITE, fg=C_GREEN).pack(anchor="w")
            tk.Label(frm, text="New Qty:", font=("Segoe UI", 9), bg=C_WHITE, fg=C_GRAY).pack(anchor="w", pady=(8, 2))
            v_qty = tk.StringVar(value=str(row["qty"]))
            ttk.Entry(frm, textvariable=v_qty, width=12, font=("Segoe UI", 12)).pack(anchor="w", ipady=2)
            tk.Label(frm, text="Reason (optional):", font=("Segoe UI", 9), bg=C_WHITE, fg=C_GRAY).pack(anchor="w", pady=(8, 2))
            v_reason = tk.StringVar()
            ttk.Entry(frm, textvariable=v_reason, width=30).pack(anchor="w")
            def save_qty():
                try: new_qty = float(v_qty.get().strip())
                except: messagebox.showerror("Error", "Sahi qty daalo!", parent=dlg); return
                if new_qty < 0: messagebox.showerror("Error", "0 se kam nahi!", parent=dlg); return
                conn = get_db()
                conn.execute("UPDATE expiry_stock SET qty=? WHERE id=?", (new_qty, row["id"]))
                diff = new_qty - row["qty"]
                if diff != 0:
                    conn.execute(
                        "INSERT INTO stock_movements(move_date,product,move_type,qty,ref_no,reason) VALUES(?,?,?,?,?,?)",
                        (today_str(), row["product"], "IN" if diff > 0 else "OUT",
                         abs(diff), f"Expiry Batch #{row['id']}", v_reason.get().strip() or "Expiry batch qty adjust"))
                conn.commit(); conn.close()
                dlg.destroy(); _load_table()
                messagebox.showinfo("Updated!", f"Qty update ho gayi! New Qty: {new_qty:.0f}")
            bf = tk.Frame(dlg, bg=C_WHITE); bf.pack(pady=4)
            make_btn(bf, "Save", save_qty, bg=C_GREEN).pack(side="left", padx=4)
            make_btn(bf, "Cancel", dlg.destroy, bg=C_GRAY).pack(side="left", padx=4)

        def _open_batch_form(existing=None):
            dlg = tk.Toplevel(p.winfo_toplevel())
            dlg.title("Batch Add" if not existing else "Batch Edit")
            dlg.configure(bg=C_WHITE)
            dlg.resizable(False, False)
            w_d, h_d = 520, 540
            sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
            dlg.geometry(f"{w_d}x{h_d}+{(sw-w_d)//2}+{(sh-h_d)//2}")
            dlg.grab_set()
            hdr_bg = "#276749" if not existing else "#2B6CB0"
            hdr_txt = "Nayi Batch Add Karo" if not existing else "Batch Edit Karo"
            tk.Label(dlg, text=hdr_txt, font=("Segoe UI", 12, "bold"),
                     bg=hdr_bg, fg="white", pady=4).pack(fill="x")
            frm = tk.Frame(dlg, bg=C_WHITE, padx=24, pady=4); frm.pack(fill="both", expand=True)
            def lbl(text):
                tk.Label(frm, text=text, font=("Segoe UI", 9), bg=C_WHITE, fg=C_GRAY,
                         anchor="w").pack(fill="x", pady=(8, 1))
            lbl("Product Name *")
            conn_p = get_db()
            all_prods = [r[0] for r in conn_p.execute("SELECT name FROM products ORDER BY name").fetchall()]
            conn_p.close()
            v_prod = tk.StringVar(value=existing["product"] if existing else "")
            prod_cb = ttk.Combobox(frm, textvariable=v_prod, values=all_prods, width=40, font=("Segoe UI",9))
            add_autocomplete(prod_cb, lambda: [r["name"] for r in get_db().execute("SELECT name FROM products ORDER BY name").fetchall()])
            prod_cb.pack(fill="x", ipady=2)
            lbl("Batch Number (optional)")
            v_batch = tk.StringVar(value=existing["batch_no"] if existing else "")
            ttk.Entry(frm, textvariable=v_batch, width=24, font=("Segoe UI",9)).pack(fill="x", ipady=2)
            dates_row = tk.Frame(frm, bg=C_WHITE); dates_row.pack(fill="x", pady=(6, 0))
            df1 = tk.Frame(dates_row, bg=C_WHITE); df1.pack(side="left", fill="x", expand=True, padx=(0, 8))
            df2 = tk.Frame(dates_row, bg=C_WHITE); df2.pack(side="left", fill="x", expand=True)
            tk.Label(df1, text="Mfg Date", font=("Segoe UI", 9), bg=C_WHITE, fg=C_GRAY, anchor="w").pack(fill="x")
            v_mfg = tk.StringVar(value=existing["mfg_date"] if existing else "")
            make_date_entry(df1, v_mfg, width=14, bg=C_WHITE).pack(anchor="w", pady=(2, 0))
            tk.Label(df2, text="Expiry Date * (MM/YY)", font=("Segoe UI", 9), bg=C_WHITE, fg="#9B2C2C", anchor="w").pack(fill="x")
            v_exp = tk.StringVar(value=fmt_exp_mmyy(existing["expiry_date"]) if existing else "")
            make_exp_mmyy_entry(df2, v_exp, width=8).pack(anchor="w", pady=(2, 0))
            num_row = tk.Frame(frm, bg=C_WHITE); num_row.pack(fill="x", pady=(6, 0))
            v_qty   = tk.StringVar(value=str(existing["qty"])           if existing else "")
            v_prate = tk.StringVar(value=str(existing["purchase_rate"]) if existing else "")
            v_mrp   = tk.StringVar(value=str(existing["mrp"])           if existing else "")
            for nf_lbl, nf_var in [("Qty *", v_qty), ("Purchase Rate Rs.", v_prate), ("MRP Rs.", v_mrp)]:
                nf2 = tk.Frame(num_row, bg=C_WHITE); nf2.pack(side="left", fill="x", expand=True, padx=(0, 6))
                tk.Label(nf2, text=nf_lbl, font=("Segoe UI", 9), bg=C_WHITE, fg=C_GRAY, anchor="w").pack(fill="x")
                ttk.Entry(nf2, textvariable=nf_var, width=10, font=("Segoe UI",9)).pack(fill="x", ipady=2)
            def _autofill(ev=None):
                pname = v_prod.get().strip()
                if pname:
                    c2 = get_db()
                    pr = c2.execute("SELECT mrp, purchase_rate FROM products WHERE name=?", (pname,)).fetchone()
                    c2.close()
                    if pr:
                        if not v_mrp.get() or v_mrp.get() == "0": v_mrp.set(str(pr["mrp"] or 0))
                        if not v_prate.get() or v_prate.get() == "0": v_prate.set(str(pr["purchase_rate"] or 0))
            prod_cb.bind("<<ComboboxSelected>>", _autofill)
            lbl("Supplier Name (optional)")
            conn_s = get_db()
            supp_names = [""] + sorted({r[0] for r in conn_s.execute(
                "SELECT DISTINCT party FROM purchases ORDER BY party").fetchall()})
            conn_s.close()
            v_supplier = tk.StringVar(value=existing["supplier"] if existing else "")
            ttk.Combobox(frm, textvariable=v_supplier, values=supp_names, width=40,
                         font=("Segoe UI",9)).pack(fill="x", ipady=2)
            prev_lbl = tk.Label(frm, text="", font=("Segoe UI", 9, "bold"), bg=C_WHITE)
            prev_lbl.pack(anchor="w", pady=(6, 0))
            def _update_preview(*_):
                exp_str = v_exp.get().strip()
                if exp_str:
                    st, sfg, _ = _status_info(exp_str)
                    prev_lbl.config(text=f"Status: {st}", fg=sfg)
                else:
                    prev_lbl.config(text="")
            v_exp.trace_add("write", _update_preview)
            _update_preview()
            err_lbl = tk.Label(frm, text="", font=("Segoe UI", 9, "bold"), bg=C_WHITE, fg=C_RED)
            err_lbl.pack(anchor="w", pady=(4, 0))
            def save_batch():
                prod = v_prod.get().strip()
                exp  = v_exp.get().strip()
                if not prod: err_lbl.config(text="Product naam zaroori hai!"); return
                if not exp:  err_lbl.config(text="Expiry date zaroori hai!"); return
                import re as _re_bf
                if not _re_bf.match(r"^(0[1-9]|1[0-2])/\d{2}$", exp):
                    err_lbl.config(text="Expiry date format galat! (MM/YY)"); return
                try: qty_v   = float(v_qty.get().strip() or 0)
                except: err_lbl.config(text="Qty mein sirf number likhein!"); return
                try: prate_v = float(v_prate.get().strip() or 0)
                except: err_lbl.config(text="Purchase rate mein sirf number likhein!"); return
                try: mrp_v   = float(v_mrp.get().strip() or 0)
                except: err_lbl.config(text="MRP mein sirf number likhein!"); return
                conn = get_db()
                if existing:
                    conn.execute(
                        "UPDATE expiry_stock SET product=?,batch_no=?,mfg_date=?,expiry_date=?,"
                        "qty=?,purchase_rate=?,mrp=?,supplier=? WHERE id=?",
                        (prod, v_batch.get().strip(), v_mfg.get().strip(), exp,
                         qty_v, prate_v, mrp_v, v_supplier.get().strip(), existing["id"]))
                else:
                    conn.execute(
                        "INSERT INTO expiry_stock(product,batch_no,mfg_date,expiry_date,"
                        "qty,purchase_rate,mrp,supplier) VALUES(?,?,?,?,?,?,?,?)",
                        (prod, v_batch.get().strip(), v_mfg.get().strip(), exp,
                         qty_v, prate_v, mrp_v, v_supplier.get().strip()))
                conn.commit(); conn.close()
                dlg.destroy(); _load_table()
                messagebox.showinfo("Saved!", f"Batch save ho gayi!\nProduct: {prod}\nExpiry: {fmt_exp_mmyy(exp)}")
            bf = tk.Frame(dlg, bg=C_WHITE); bf.pack(pady=(4, 12))
            make_btn(bf, "Update Karo" if existing else "Batch Save Karo", save_batch, bg=C_GREEN).pack(side="left", padx=6)
            make_btn(bf, "Cancel", dlg.destroy, bg=C_GRAY).pack(side="left", padx=6)

        def _export_excel():
            if not rows_data:
                messagebox.showinfo("Koi Data Nahi", "Pehle table mein data hona chahiye!"); return
            headers = ["Product", "Batch No", "Mfg Date", "Expiry Date",
                       "Qty", "Purchase Rate", "MRP", "Supplier", "Status"]
            export_rows = []
            for row in rows_data:
                st, _, _ = _status_info(row["expiry_date"])
                export_rows.append([row["product"], row["batch_no"] or "",
                    fmt_date(row["mfg_date"]) if row["mfg_date"] else "",
                    fmt_exp_mmyy(row["expiry_date"]), f"{row['qty']:.0f}",
                    f"{row['purchase_rate']:.2f}", f"{row['mrp']:.2f}",
                    row["supplier"] or "", st])
            export_to_excel(headers, export_rows, "Expiry_Report")

        def _import_from_purchase():
            dlg = tk.Toplevel(p.winfo_toplevel())
            dlg.title("Purchase se Import Karo")
            dlg.configure(bg=C_WHITE)
            dlg.geometry("640x500"); dlg.grab_set()
            tk.Label(dlg, text="Purchase se Batch Import Karo",
                     font=("Segoe UI", 12, "bold"), bg="#2B6CB0", fg="white", pady=4).pack(fill="x")
            tk.Label(dlg, text="Purchase history se products chuniye aur batch add karein:",
                     font=("Segoe UI", 9), bg=C_WHITE, fg=C_GRAY, pady=6).pack()
            conn = get_db()
            pur_items = conn.execute("""
                SELECT pi.product, pi.qty, pi.rate, p.bill_date, p.party
                FROM purchase_items pi
                JOIN purchases p ON pi.purchase_id = p.id
                ORDER BY p.bill_date DESC LIMIT 100
            """).fetchall()
            conn.close()
            if not pur_items:
                tk.Label(dlg, text="Koi purchase history nahi mili!", font=("Segoe UI",9),
                         bg=C_WHITE, fg=C_RED, pady=20).pack()
                make_btn(dlg, "Close", dlg.destroy, bg=C_GRAY).pack(); return
            outer_f, inner_f = scrollable_frame(dlg)
            outer_f.pack(fill="both", expand=True, padx=12, pady=4)
            check_vars = []; exp_vars = []; batch_vars = []
            items_list = [dict(r) for r in pur_items]
            for ci, (hd, wd) in enumerate([("?",2),("Product",18),("Qty",5),
                                           ("Bill Date",9),("Supplier",14),
                                           ("Expiry Date *",12),("Batch No",10)]):
                tk.Label(inner_f, text=hd, font=("Segoe UI", 8, "bold"), bg=C_THEAD, fg=C_ACCENT,
                         width=wd, anchor="w", padx=4, pady=5).grid(row=0, column=ci, sticky="nsew", padx=1)
            for ri, item in enumerate(items_list, 1):
                row_bg = C_WHITE if ri % 2 == 0 else "#F7FAFC"
                cv = tk.BooleanVar(value=False); check_vars.append(cv)
                tk.Checkbutton(inner_f, variable=cv, bg=row_bg).grid(row=ri, column=0, padx=4)
                for ci, (val, wd) in enumerate([(item["product"],18),(f"{item['qty']:.0f}",5),
                                                 (fmt_date(item["bill_date"]),9),(item["party"] or "—",14)],1):
                    tk.Label(inner_f, text=str(val), font=("Segoe UI",7), bg=row_bg, fg=C_GRAY,
                             width=wd, anchor="w", padx=4, pady=4).grid(row=ri, column=ci, sticky="nsew", padx=1)
                ev = tk.StringVar(); exp_vars.append(ev)
                make_exp_mmyy_entry(inner_f, ev, width=8).grid(row=ri, column=5, padx=2, pady=2)
                bv = tk.StringVar(); batch_vars.append(bv)
                ttk.Entry(inner_f, textvariable=bv, width=10, font=("Segoe UI",7)).grid(row=ri, column=6, padx=2, pady=2)
            def do_import():
                selected = [(items_list[i], exp_vars[i].get().strip(), batch_vars[i].get().strip())
                            for i, cv in enumerate(check_vars) if cv.get()]
                if not selected: messagebox.showwarning("Koi Select Nahi","Koi row select nahi kiya!",parent=dlg); return
                no_exp = [s for s in selected if not s[1]]
                if no_exp: messagebox.showerror("Expiry Date Zaroori",
                    f"{len(no_exp)} items mein expiry date nahi hai!", parent=dlg); return
                conn = get_db(); count = 0
                import re as _re_imp
                for item, exp_str, batch_str in selected:
                    try:
                        if not _re_imp.match(r"^(0[1-9]|1[0-2])/\d{2}$", exp_str):
                            continue
                        conn.execute(
                            "INSERT INTO expiry_stock(product,batch_no,expiry_date,qty,purchase_rate,mrp,supplier) VALUES(?,?,?,?,?,?,?)",
                            (item["product"], batch_str, exp_str, item["qty"], item["rate"], 0, item["party"] or ""))
                        count += 1
                    except: pass
                conn.commit(); conn.close()
                dlg.destroy(); _load_table()
                messagebox.showinfo("Import Done!", f"{count} batches import ho gayi!")
            bf2 = tk.Frame(dlg, bg=C_WHITE); bf2.pack(pady=4)
            make_btn(bf2, "Import Selected", do_import, bg=C_GREEN).pack(side="left", padx=6)
            make_btn(bf2, "Cancel", dlg.destroy, bg=C_GRAY).pack(side="left", padx=6)

        # Quick action buttons
        import_btn_row = tk.Frame(p, bg=C_BG); import_btn_row.pack(fill="x", pady=(0, 4))
        make_btn(import_btn_row, "Purchase se Import", _import_from_purchase, bg="#6B46C1").pack(side="left", padx=4)
        make_btn(import_btn_row, "Sirf Alerts (30d)", lambda: [v_filter.set("30 Din Mein"), _load_table()], bg=C_AMBER).pack(side="left", padx=4)
        make_btn(import_btn_row, "Sirf Expired", lambda: [v_filter.set("Expire Ho Gayi"), _load_table()], bg=C_RED).pack(side="left", padx=4)
        make_btn(import_btn_row, "Sab Reset Karo", lambda: [v_filter.set("Sab Dikhao"), v_search.set(""), _load_table()], bg=C_GREEN).pack(side="left", padx=4)

        v_search.trace_add("write", _load_table)
        v_filter.trace_add("write", _load_table)
        v_supp.trace_add("write", _load_table)
        _load_table()

        # Alert summary at bottom
        if alerts:
            div_f = tk.Frame(p, bg=C_BG); div_f.pack(fill="x", pady=(12, 0))
            tk.Frame(div_f, bg=C_BORDER, height=1).pack(fill="x")
            tk.Label(div_f, text="Quick Alert Summary (Top 10)",
                     font=("Segoe UI", 11, "bold"), bg=C_BG, fg="#9B2C2C", pady=6).pack(anchor="w")
            alert_tbl = make_table(div_f, ["Product","Batch","Expiry Date","Qty","Status"], [20,12,10,6,20])
            for i, a in enumerate(alerts[:10], 1):
                st, sfg, sbg = _status_info(a["expiry_date"])
                table_row(alert_tbl,
                    [a["product"], a["batch_no"] or "—", fmt_exp_mmyy(a["expiry_date"]), f"{a['qty']:.0f}", st],
                    i, fgs=[None,None,None,None,sfg], bg=sbg)
            if len(alerts) > 10:
                tk.Label(div_f, text=f"... aur {len(alerts)-10} aur items. Upar filter use karo.",
                         font=("Segoe UI", 9), bg=C_BG, fg=C_GRAY, pady=4).pack(anchor="w")


#  INVOICE WINDOW
# ══════════════════════════════════════════════════════════════════════════════
class InvoiceWin:
    def __init__(self, parent, data):
        self.data = data
        self.win = tk.Toplevel(parent)
        self.win.title(f"{data.get('label','Invoice')} — {data['bill_no']}")
        self.win.configure(bg=C_WHITE)
        self.win.state("zoomed")          # full screen maximized
        self.win.grab_set()
        self._build()

    def _build(self):
        # Toolbar
        tb = tk.Frame(self.win, bg="#1A365D", pady=4)
        tb.pack(fill="x")
        make_btn(tb, "🖨️  Print / Save PDF", self._print).pack(side="left", padx=12)
        make_btn(tb, "🧾  Thermal Print", self._thermal_print, bg="#B7791F").pack(side="left", padx=4)
        make_btn(tb, "📧  Email", self._share_email, bg="#2B6CB0").pack(side="left", padx=4)
        make_btn(tb, "💬  WhatsApp", self._share_whatsapp, bg="#25D366").pack(side="left", padx=4)
        make_btn(tb, "📱  SMS", self._share_sms, bg="#7B2D8B").pack(side="left", padx=4)
        make_btn(tb, "🔗  E-Invoice", self._gen_einvoice, bg="#C05621").pack(side="left", padx=4)
        make_btn(tb, "🚛  E-Way Bill", self._gen_ewaybill, bg="#276749").pack(side="left", padx=4)
        make_btn(tb, "✕  Close", self.win.destroy, bg="#4A5568").pack(side="left", padx=8)
        tk.Label(tb, text=f"Invoice: {self.data['bill_no']}",
                 font=("Segoe UI", 11, "bold"), bg="#1A365D", fg="white"
                 ).pack(side="left", padx=20)
        tk.Label(tb, text=f"Items: {len(self.data.get('items', []))}",
                 font=("Segoe UI", 11, "bold"), bg="#1A365D", fg="#9AE6B4"
                 ).pack(side="left", padx=8)

        # Scrollable area — grey background
        outer = tk.Frame(self.win, bg="#CBD5E0")
        outer.pack(fill="both", expand=True)

        canvas = tk.Canvas(outer, bg="#CBD5E0", highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        hsb = ttk.Scrollbar(outer, orient="horizontal", command=canvas.xview)
        canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        canvas.pack(fill="both", expand=True)

        # A4 white paper — fixed 794px wide (A4 at 96dpi)
        paper = tk.Frame(canvas, bg=C_WHITE, width=794,
                         highlightthickness=1, highlightbackground="#999")
        wid = canvas.create_window((0, 0), window=paper, anchor="nw")

        def _on_canvas(e):
            # Center paper horizontally if canvas is wider
            pw = 794
            cx = e.width
            x = max(0, (cx - pw) // 2)
            canvas.coords(wid, x, 20)
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_paper(e):
            canvas.configure(scrollregion=canvas.bbox("all"))

        canvas.bind("<Configure>", _on_canvas)
        paper.bind("<Configure>", _on_paper)
        canvas.bind("<Enter>",
            lambda e: canvas.bind_all("<MouseWheel>",
                lambda ev: canvas.yview_scroll(-1*(ev.delta//120), "units")))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        self._draw(paper)

    def _draw(self, f):
        d    = self.data
        party= d["party"]
        items= d["items"]
        grand= d["grand"]

        total_taxable = sum(it.get("taxable", it.get("rate",0)*it.get("qty",0)) for it in items)
        total_gst     = sum(it.get("gst_amt",0) for it in items)
        total_item_disc = sum(it.get("disc_amt", 0) for it in items)
        bill_disc       = d.get("bill_disc", 0) or 0

        def cell(parent, text, row, col,
                 font_s=9, bold=False, align="w",
                 bg=C_WHITE, fg="#000000",
                 rowspan=1, colspan=1, padx_in=4, pady_in=3,
                 sticky="nsew"):
            fr = tk.Frame(parent, bg=bg, highlightthickness=1, highlightbackground="#999999")
            fr.grid(row=row, column=col, rowspan=rowspan, columnspan=colspan,
                    sticky=sticky, padx=0, pady=0)
            wt = "bold" if bold else "normal"
            tk.Label(fr, text=str(text), font=("Arial", font_s, wt),
                     bg=bg, fg=fg, anchor=align,
                     padx=padx_in, pady=pady_in).pack(fill="both", expand=True)

        main = tk.Frame(f, bg=C_WHITE, padx=22, pady=16)
        main.pack(fill="x", expand=False)

        # Header
        tbl = tk.Frame(main, bg=C_WHITE)
        tbl.pack(fill="x")
        for col, wt in enumerate([3,1,2,1,2]):
            tbl.columnconfigure(col, weight=wt)

        title_f = tk.Frame(tbl, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
        title_f.grid(row=0, column=0, columnspan=5, sticky="nsew")
        _is_nongst = (self.data.get("bill_mode","GST") == "NONGST")
        _title_txt = self.data.get("label") or ("BILL (Non-GST)" if _is_nongst else "Tax Invoice")
        tk.Label(title_f, text=_title_txt, font=("Arial", 12, "bold"),
                 bg=C_WHITE, anchor="center", pady=4).pack(fill="x")

        cell(tbl,get_shop()["name"],1,0,font_s=10,bold=True)
        cell(tbl,"Invoice :",1,1,font_s=9,align="e")
        cell(tbl,d["bill_no"],1,2,font_s=9,bold=True)
        cell(tbl,"Dated :",1,3,font_s=9,align="e")
        cell(tbl,fmt_date(d["date"]),1,4,font_s=9,bold=True)

        _sp = get_shop()
        cell(tbl,f"{_sp['address']}, {_sp['city']}".strip(", "),2,0,font_s=9)
        cell(tbl,"Total Items :",2,1,font_s=9,align="e")
        cell(tbl,str(len(items)),2,2,font_s=9,bold=True)
        cell(tbl,"Mode/Terms of Payment",2,3,colspan=2,font_s=9)

        cell(tbl,f"Mob: {_sp['mobile']}" if _sp['mobile'] else "",3,0,font_s=9)
        cell(tbl,"Reference No. & Date",3,1,colspan=2,font_s=9)
        cell(tbl,"Other References",3,3,colspan=2,font_s=9)

        cell(tbl,f"GSTIN: {_sp['gstin']}" if _sp['gstin'] else "",4,0,font_s=9)
        cell(tbl,"Buyers order No.",4,1,colspan=2,font_s=9)
        cell(tbl,"Dated",4,3,colspan=2,font_s=9)

        cell(tbl,f"State: {_sp['state']}",5,0,font_s=9)
        cell(tbl,"Dispatch Doc No.",5,1,colspan=2,font_s=9)
        cell(tbl,"Delivery Note date",5,3,colspan=2,font_s=9)

        cell(tbl,"",6,0,font_s=9)
        cell(tbl,"Dispatched through",6,1,colspan=2,font_s=9)
        cell(tbl,"Destination",6,3,colspan=2,font_s=9)

        cell(tbl,"",7,0,font_s=9)
        cell(tbl,"Terms of Delivery",7,1,colspan=4,font_s=9)

        # Consignee
        bill_tbl = tk.Frame(main, bg=C_WHITE)
        bill_tbl.pack(fill="x")
        bill_tbl.columnconfigure(0, weight=1)
        bill_tbl.columnconfigure(1, weight=2)

        p_name = party.get("name","")
        p_addr = party.get("address","")
        p_mob  = party.get("mobile","")
        p_gst  = party.get("gstin","")
        bill_to_text = f"{p_name}\n{p_addr}\nMob: {p_mob}\nGSTIN/UIN: {p_gst}"

        lf2 = tk.Frame(bill_tbl, bg=C_WHITE)
        lf2.grid(row=0, column=0, sticky="nsew")
        cell_frame = tk.Frame(lf2, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
        cell_frame.pack(fill="both", expand=True)
        tk.Label(cell_frame, text="Consignee (Bill to)", font=("Arial",9,"bold"),
                 bg=C_WHITE, anchor="w", padx=4, pady=2).pack(fill="x")
        tk.Label(cell_frame, text=bill_to_text, font=("Arial",9),
                 bg=C_WHITE, anchor="w", padx=4, pady=2, justify="left").pack(fill="x")
        tk.Label(cell_frame, text="Consignee (Ship to)", font=("Arial",9,"bold"),
                 bg=C_WHITE, anchor="w", padx=4, pady=2).pack(fill="x", pady=(10,0))
        tk.Label(cell_frame, text=bill_to_text, font=("Arial",9),
                 bg=C_WHITE, anchor="w", padx=4, pady=2, justify="left").pack(fill="x")

        right_cell = tk.Frame(bill_tbl, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
        right_cell.grid(row=0, column=1, sticky="nsew")
        tk.Label(right_cell, text="", bg=C_WHITE, height=8).pack()

        # Items Table
        it_tbl = tk.Frame(main, bg=C_WHITE)
        it_tbl.pack(fill="x")
        COLS = ["Sl No.","Description","HSN/SAC Code","Batch No","Expiry Date","Qty","Rate","Disc%","Total Rate","GST RATE","GST AMT","Total"]
        WCOLS= [3, 14, 7, 8, 8, 5, 7, 5, 8, 7, 7, 6]
        for i,c in enumerate(COLS):
            it_tbl.columnconfigure(i, weight=WCOLS[i])

        for i,c in enumerate(COLS):
            hf = tk.Frame(it_tbl, bg=HDR_BG, highlightthickness=1, highlightbackground="#999999")
            hf.grid(row=0, column=i, sticky="nsew")
            tk.Label(hf, text=c, font=("Arial",8,"bold"), bg=HDR_BG,
                     anchor="center", padx=3, pady=4, wraplength=60, justify="center").pack(fill="both")

        n_rows = max(len(items), 7)
        for r in range(n_rows):
            if r < len(items):
                it = items[r]
                rate  = it.get("rate",0)
                qty   = it.get("qty",0)
                taxbl = it.get("taxable", rate*qty)
                gst_p = it.get("gst_percent", it.get("gst",18))
                gst_a = it.get("gst_amt",0)
                grnd  = it.get("grand", taxbl+gst_a)
                batch_v = it.get("batch_no","") or ""
                exp_v   = fmt_exp_mmyy(it.get("expiry_date","")) or ""
                disc_pct = it.get("disc", 0)
                disc_str = it.get("disc_label", f"{disc_pct}%" if disc_pct else "-")
                row_vals = [r+1, it.get("product",""), it.get("hsn",""),
                            batch_v, exp_v, qty, f"{rate}", disc_str, f"{taxbl}",
                            f"{gst_p}%", f"{gst_a}", f"{grnd}"]
            else:
                row_vals = ["","","","","","","","","","","",""]

            for c, val in enumerate(row_vals):
                bf_inner = tk.Frame(it_tbl, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
                bf_inner.grid(row=r+1, column=c, sticky="nsew")
                anc = "center" if c in [0,2,3,4,5,6,7,8,9,10,11] else "w"
                tk.Label(bf_inner, text=str(val), font=("Arial",9),
                         bg=C_WHITE, anchor=anc, padx=3, pady=3).pack(fill="both")

        tot_row = n_rows + 1
        for c in range(12):
            tf2 = tk.Frame(it_tbl, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
            tf2.grid(row=tot_row, column=c, sticky="nsew")
            sub_total_items = sum(it.get("grand",0) for it in items)
            if c == 1:
                tk.Label(tf2, text="Sub Total", font=("Arial",9,"bold"),
                         bg=C_WHITE, anchor="w", padx=4, pady=4).pack(fill="both")
            elif c == 11:
                tk.Label(tf2, text=f"₹{sub_total_items:.0f}", font=("Arial",9,"bold"),
                         bg=C_WHITE, anchor="center", padx=3, pady=4).pack(fill="both")
            else:
                tk.Label(tf2, text="", bg=C_WHITE, pady=4).pack()

        # Bill-level discount row (only if applicable)
        if bill_disc > 0:
            disc_row = tot_row + 1
            for c in range(12):
                tf2 = tk.Frame(it_tbl, bg="#FFF5F5", highlightthickness=1, highlightbackground="#999999")
                tf2.grid(row=disc_row, column=c, sticky="nsew")
                if c == 1:
                    tk.Label(tf2, text="Bill Discount", font=("Arial",9,"bold"),
                             bg="#FFF5F5", fg="#C53030", anchor="w", padx=4, pady=4).pack(fill="both")
                elif c == 11:
                    tk.Label(tf2, text=f"- ₹{bill_disc:.0f}", font=("Arial",9,"bold"),
                             bg="#FFF5F5", fg="#C53030", anchor="center", padx=3, pady=4).pack(fill="both")
                else:
                    tk.Label(tf2, text="", bg="#FFF5F5", pady=4).pack()

            grand_row = disc_row + 1
            for c in range(12):
                tf2 = tk.Frame(it_tbl, bg="#EBF4FF", highlightthickness=1, highlightbackground="#999999")
                tf2.grid(row=grand_row, column=c, sticky="nsew")
                if c == 1:
                    tk.Label(tf2, text="Grand Total", font=("Arial",9,"bold"),
                             bg="#EBF4FF", fg="#1A365D", anchor="w", padx=4, pady=4).pack(fill="both")
                elif c == 11:
                    tk.Label(tf2, text=f"₹{grand:.0f}", font=("Arial",9,"bold"),
                             bg="#EBF4FF", fg="#1A365D", anchor="center", padx=3, pady=4).pack(fill="both")
                else:
                    tk.Label(tf2, text="", bg="#EBF4FF", pady=4).pack()

        # Amount in words
        words_f = tk.Frame(main, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
        words_f.pack(fill="x")
        words_f.columnconfigure(0, weight=3); words_f.columnconfigure(1, weight=1)
        tk.Label(words_f, text="Amount Chargeable (In Words)",
                 font=("Arial",8), bg=C_WHITE, anchor="w", padx=4, pady=2
                 ).grid(row=0, column=0, sticky="w")
        tk.Label(words_f, text="E. & O.E.",
                 font=("Arial",8,"italic"), bg=C_WHITE, anchor="e", padx=6
                 ).grid(row=0, column=1, sticky="e")
        tk.Label(words_f, text=num_to_words(grand),
                 font=("Arial",9,"bold"), bg=C_WHITE, anchor="w", padx=4, pady=3
                 ).grid(row=1, column=0, columnspan=2, sticky="w")

        # Tax Summary — CGST+SGST ya IGST (sirf GST bill ke liye)
        gst_type = self.data.get("gst_type","CGST+SGST") or "CGST+SGST"
        is_igst  = (gst_type == "IGST")

        tax_f = tk.Frame(main, bg=C_WHITE)
        tax_f.pack(fill="x")

        if _is_nongst:
            note_f = tk.Frame(tax_f, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
            note_f.pack(fill="x")
            tk.Label(note_f, text="GST not charged on this bill (Non-GST Bill)",
                     font=("Arial",9,"italic"), bg=C_WHITE, fg="#9B2C2C",
                     anchor="center", padx=6, pady=6).pack(fill="x")
        elif is_igst:
            TCOLS = ["HSN","Taxable Amt","Rate (IGST)","IGST Amount","Total Tax"]
            tax_f.columnconfigure(0, weight=2)
            for i in range(1,5): tax_f.columnconfigure(i, weight=1)
            for i,c in enumerate(TCOLS):
                hf2 = tk.Frame(tax_f, bg=HDR_BG, highlightthickness=1, highlightbackground="#999999")
                hf2.grid(row=0, column=i, sticky="nsew")
                tk.Label(hf2, text=c, font=("Arial",8,"bold"), bg=HDR_BG,
                         anchor="center", padx=4, pady=4).pack(fill="both")
            # Group by GST %
            gst_groups = {}
            for it in items:
                gp = it.get("gst_percent", it.get("gst",0))
                hsn = it.get("hsn","")
                k   = (hsn, gp)
                if k not in gst_groups:
                    gst_groups[k] = {"taxable":0,"gst_amt":0}
                gst_groups[k]["taxable"] += it.get("taxable", it.get("rate",0)*it.get("qty",0))
                gst_groups[k]["gst_amt"] += it.get("gst_amt",0)
            for ri,(k,v) in enumerate(gst_groups.items()):
                hsn, gp = k
                vals = [hsn, f"₹{v['taxable']:.2f}", f"{gp}%",
                        f"₹{v['gst_amt']:.2f}", f"₹{v['gst_amt']:.2f}"]
                for ci,val in enumerate(vals):
                    tf3 = tk.Frame(tax_f, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
                    tf3.grid(row=ri+1, column=ci, sticky="nsew")
                    tk.Label(tf3, text=val, font=("Arial",9), bg=C_WHITE,
                             anchor="center", padx=4, pady=3).pack(fill="both")
            # Total row
            tr = len(gst_groups)+1
            for ci,val in enumerate(["Total","",f"",f"₹{total_gst:.2f}",f"₹{total_gst:.2f}"]):
                tf3 = tk.Frame(tax_f, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
                tf3.grid(row=tr, column=ci, sticky="nsew")
                tk.Label(tf3, text=val, font=("Arial",9,"bold"), bg=C_WHITE,
                         anchor="center", padx=4, pady=3).pack(fill="both")
        else:
            # CGST + SGST split
            TCOLS = ["HSN","Taxable Amt","Rate (CGST)","CGST Amt","Rate (SGST)","SGST Amt","Total Tax"]
            tax_f.columnconfigure(0, weight=2)
            for i in range(1,7): tax_f.columnconfigure(i, weight=1)
            for i,c in enumerate(TCOLS):
                hf2 = tk.Frame(tax_f, bg=HDR_BG, highlightthickness=1, highlightbackground="#999999")
                hf2.grid(row=0, column=i, sticky="nsew")
                tk.Label(hf2, text=c, font=("Arial",8,"bold"), bg=HDR_BG,
                         anchor="center", padx=4, pady=4).pack(fill="both")
            gst_groups = {}
            for it in items:
                gp  = it.get("gst_percent", it.get("gst",0))
                hsn = it.get("hsn","")
                k   = (hsn, gp)
                if k not in gst_groups:
                    gst_groups[k] = {"taxable":0,"gst_amt":0}
                gst_groups[k]["taxable"] += it.get("taxable", it.get("rate",0)*it.get("qty",0))
                gst_groups[k]["gst_amt"] += it.get("gst_amt",0)
            for ri,(k,v) in enumerate(gst_groups.items()):
                hsn, gp = k
                half = gp/2
                h_amt = v["gst_amt"]/2
                vals = [hsn, f"₹{v['taxable']:.2f}",
                        f"{half}%", f"₹{h_amt:.2f}",
                        f"{half}%", f"₹{h_amt:.2f}",
                        f"₹{v['gst_amt']:.2f}"]
                for ci,val in enumerate(vals):
                    tf3 = tk.Frame(tax_f, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
                    tf3.grid(row=ri+1, column=ci, sticky="nsew")
                    tk.Label(tf3, text=val, font=("Arial",9), bg=C_WHITE,
                             anchor="center", padx=4, pady=3).pack(fill="both")
            tr = len(gst_groups)+1
            for ci,val in enumerate(["Total","","",f"₹{total_gst/2:.2f}","",f"₹{total_gst/2:.2f}",f"₹{total_gst:.2f}"]):
                tf3 = tk.Frame(tax_f, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
                tf3.grid(row=tr, column=ci, sticky="nsew")
                tk.Label(tf3, text=val, font=("Arial",9,"bold"), bg=C_WHITE,
                         anchor="center", padx=4, pady=3).pack(fill="both")

        # Bank details
        bot_f = tk.Frame(main, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
        bot_f.pack(fill="x")
        bot_f.columnconfigure(0, weight=1); bot_f.columnconfigure(1, weight=1)

        left_bot = tk.Frame(bot_f, bg=C_WHITE)
        left_bot.grid(row=0, column=0, sticky="nsew")
        tk.Label(left_bot, text="Tax Amount (In Words): "+("N/A (Non-GST Bill)" if _is_nongst else num_to_words(total_gst)),
                 font=("Arial",8), bg=C_WHITE, anchor="w", padx=6, pady=4,
                 wraplength=340, justify="left").pack(fill="x")

        right_bot = tk.Frame(bot_f, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
        right_bot.grid(row=0, column=1, sticky="nsew")
        tk.Label(right_bot, text="Company's Bank Details",
                 font=("Arial",9,"bold"), bg=C_WHITE, anchor="w", padx=6, pady=4).pack(fill="x", pady=(6,2))
        _spb = get_shop()
        for k,v in [("A/C Holder's Name", f": {_spb['name']}"),
                    ("Bank Name",         f": {_spb['bank']}" if _spb['bank'] else ": —"),
                    ("A/C No.",           f": {_spb['account']}" if _spb['account'] else ": —"),
                    ("Branch & IFS Code", f": {_spb['ifsc']}" if _spb['ifsc'] else ": —")]:
            rf2 = tk.Frame(right_bot, bg=C_WHITE)
            rf2.pack(fill="x", padx=6)
            tk.Label(rf2, text=k, font=("Arial",8), bg=C_WHITE, width=18, anchor="w").pack(side="left")
            tk.Label(rf2, text=v, font=("Arial",8), bg=C_WHITE, anchor="w").pack(side="left")
        tk.Label(right_bot, text="", bg=C_WHITE, pady=4).pack()

        # Declaration + Signatory
        decl_f = tk.Frame(main, bg=C_WHITE, highlightthickness=1, highlightbackground="#999999")
        decl_f.pack(fill="x")
        decl_f.columnconfigure(0, weight=1); decl_f.columnconfigure(1, weight=1)
        tk.Label(decl_f, text="Declaration",
                 font=("Arial",8,"bold"), bg=C_WHITE, anchor="w", padx=6, pady=4
                 ).grid(row=0, column=0, sticky="w")
        tk.Label(decl_f,
                 text="We declare that this invoice shows the actual price of\nthe goods described and that all particulars are true and correct.",
                 font=("Arial",8), bg=C_WHITE, anchor="w", padx=6, pady=3, justify="left"
                 ).grid(row=1, column=0, sticky="w")

        sign_f = tk.Frame(decl_f, bg=C_WHITE)
        sign_f.grid(row=0, column=1, rowspan=2, sticky="nsew")
        tk.Label(sign_f, text=f"For  {_spb['name']}",
                 font=("Arial",9,"bold"), bg=C_WHITE, anchor="e", padx=10, pady=8
                 ).pack(fill="x", pady=(0,30))
        tk.Label(sign_f, text="Authorised Signatory",
                 font=("Arial",8), bg=C_WHITE, anchor="e", padx=10
                 ).pack(fill="x")

        # Footer
        foot_f = tk.Frame(main, bg=HDR_BG, highlightthickness=1, highlightbackground="#999999")
        foot_f.pack(fill="x")
        tk.Label(foot_f, text="SUBJECT TO VARANASI JURISDICTION",
                 font=("Arial",8), bg=HDR_BG, anchor="center", pady=4
                 ).pack(fill="x")

    def _print(self):
        """Generate real PDF invoice and open it."""
        import os, subprocess, sys
        from tkinter import filedialog

        d     = self.data
        party = d["party"]
        items = d["items"]
        grand = round(d["grand"])
        total_taxable = round(sum(it.get("taxable", it.get("rate",0)*it.get("qty",0)) for it in items), 2)
        total_gst     = round(sum(it.get("gst_amt", 0) for it in items), 2)

        # Ask save location
        save_path = filedialog.asksaveasfilename(
            parent=self.win,
            title="Invoice PDF Save Karo",
            defaultextension=".pdf",
            initialfile=d["bill_no"].replace("/", "_") + ".pdf",
            filetypes=[("PDF File", "*.pdf")]
        )
        if not save_path:
            return

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
            from collections import defaultdict

            PAGE_W = 190 * mm  # usable width
            BDR    = colors.black
            GRY    = colors.HexColor("#f2f2f2")

            def ps(name, size=8, bold=False, align=TA_LEFT):
                return ParagraphStyle(
                    name, fontSize=size,
                    fontName="Helvetica-Bold" if bold else "Helvetica",
                    alignment=align, leading=size + 3
                )

            doc = SimpleDocTemplate(
                save_path, pagesize=A4,
                leftMargin=10*mm, rightMargin=10*mm,
                topMargin=8*mm,  bottomMargin=8*mm
            )

            bill_date = fmt_date(d["date"])
            p_name = party.get("name","")
            p_addr = party.get("address","").replace("\n","<br/>")
            p_mob  = party.get("mobile","")
            p_gst  = party.get("gstin","")

            # ── Helper: make table with no spacing ──
            def T(data, cw, style_cmds, repeat=0):
                tbl = Table(data, colWidths=cw,
                            style=TableStyle(style_cmds),
                            repeatRows=repeat)
                tbl.spaceAfter  = 0
                tbl.spaceBefore = 0
                return tbl

            BASE = [
                ("BOX",       (0,0),(-1,-1), 0.5, BDR),
                ("INNERGRID", (0,0),(-1,-1), 0.3, BDR),
                ("TOPPADDING",(0,0),(-1,-1), 3),
                ("BOTTOMPADDING",(0,0),(-1,-1), 3),
                ("LEFTPADDING",(0,0),(-1,-1), 3),
                ("RIGHTPADDING",(0,0),(-1,-1), 3),
                ("VALIGN",    (0,0),(-1,-1), "TOP"),
            ]

            story = []

            # ════════════════════════════════════════
            # 1. TITLE
            # ════════════════════════════════════════
            story.append(T(
                [[Paragraph(d.get("label","Invoice"), ps("t",11,False,TA_CENTER))]],
                [PAGE_W],
                [("BOX",(0,0),(-1,-1),0.5,BDR),
                 ("TOPPADDING",(0,0),(-1,-1),4),
                 ("BOTTOMPADDING",(0,0),(-1,-1),4)]
            ))

            # ════════════════════════════════════════
            # 2. COMPANY + INVOICE DETAILS (one table, no gap)
            # ════════════════════════════════════════
            _sp = get_shop()
            _sp_addr_line = f"{_sp['address']}, {_sp['city']}".strip(", ")
            company_txt = (
                f"<b>{_sp['name']}</b><br/>"
                + (f"{_sp_addr_line}<br/>" if _sp_addr_line else "")
                + (f"Mob: {_sp['mobile']}<br/>" if _sp['mobile'] else "")
                + (f"GSTIN/UIN : {_sp['gstin']}<br/>" if _sp['gstin'] else "")
                + (f"State Name : {_sp['state']}" if _sp['state'] else "")
            )

            # Right side inner table (invoice fields)
            rw = [33*mm, 32*mm, 38*mm, 22*mm]  # total = 125mm
            right_data = [
                [Paragraph("Invoice :",    ps("r",8,True)),  Paragraph(d["bill_no"],  ps("r",8)),
                 Paragraph("Dated :",      ps("r",8,True)),  Paragraph(bill_date,     ps("r",8))],
                [Paragraph("Delivery Note",ps("r",8)),       Paragraph("",ps("r",8)),
                 Paragraph("Mode/Terms of Payment",ps("r",8)),Paragraph("",ps("r",8))],
                [Paragraph("Reference No. & Date",ps("r",8)),Paragraph("",ps("r",8)),
                 Paragraph("Other References",ps("r",8)),    Paragraph("",ps("r",8))],
                [Paragraph("Buyers order No",ps("r",8)),     Paragraph("",ps("r",8)),
                 Paragraph("Dated",ps("r",8)),               Paragraph("",ps("r",8))],
                [Paragraph("Dispatch Doc No",ps("r",8)),     Paragraph("",ps("r",8)),
                 Paragraph("Delivery Note date",ps("r",8)),  Paragraph("",ps("r",8))],
                [Paragraph("Dispatched through",ps("r",8)),  Paragraph("",ps("r",8)),
                 Paragraph("Destination",ps("r",8)),         Paragraph("",ps("r",8))],
                [Paragraph("Terms of Delivery",ps("r",8)),   Paragraph("",ps("r",8)),
                 Paragraph("",ps("r",8)),                    Paragraph("",ps("r",8))],
            ]
            right_tbl = T(right_data, rw,
                BASE + [("SPAN",(0,6),(3,6))])

            # Company + right merged into one row
            hdr_tbl = T(
                [[Paragraph(company_txt, ps("co",8)), right_tbl]],
                [65*mm, 125*mm],
                BASE + [("INNERGRID",(0,0),(-1,-1),0.5,BDR)]
            )
            story.append(hdr_tbl)

            # ════════════════════════════════════════
            # 3. CONSIGNEE — Bill to | Ship to
            #    Same outer width = 190mm, no gap
            # ════════════════════════════════════════
            def consignee_txt(label):
                return (
                    f"<b>{label}</b><br/>"
                    f"{p_name}<br/>{p_addr}<br/>"
                    f"Mob: {p_mob}<br/>GSTIN/UIN-{p_gst}"
                )

            con_tbl = T(
                [[Paragraph(consignee_txt("Consignee (Bill to)"), ps("bt",8)),
                  Paragraph(consignee_txt("Consignee (Ship to)"), ps("st",8))]],
                [95*mm, 95*mm],
                BASE + [("INNERGRID",(0,0),(-1,-1),0.5,BDR)]
            )
            story.append(con_tbl)

            # ════════════════════════════════════════
            # 4. ITEMS TABLE
            # ════════════════════════════════════════
            # Cols: SlNo|Descriptions|HSN/SAC|Unit|Qty|Rate|Total Rate|GST RATE|GST AMT|Total
            CW = [8*mm, 42*mm, 16*mm, 12*mm, 14*mm, 16*mm, 19*mm, 16*mm, 19*mm, 18*mm]  # = 180mm

            it_head = [
                Paragraph("<b>Sl No.</b>",          ps("h",7,True,TA_CENTER)),
                Paragraph("<b>Descriptions</b>",    ps("h",7,True)),
                Paragraph("<b>HSN/SAC Code</b>",    ps("h",7,True,TA_CENTER)),
                Paragraph("<b>Unit</b>",             ps("h",7,True,TA_CENTER)),
                Paragraph("<b>Quantity</b>",         ps("h",7,True,TA_CENTER)),
                Paragraph("<b>Rate</b>",             ps("h",7,True,TA_CENTER)),
                Paragraph("<b>Total Rate</b>",       ps("h",7,True,TA_CENTER)),
                Paragraph("<b>GST RATE</b>",         ps("h",7,True,TA_CENTER)),
                Paragraph("<b>GST AMOUNT</b>",       ps("h",7,True,TA_CENTER)),
                Paragraph("<b>Total</b>",            ps("h",7,True,TA_CENTER)),
            ]
            it_rows = [it_head]

            n_blank = max(0, 6 - len(items))
            for idx, it in enumerate(items):
                rate  = it.get("rate", 0)
                qty   = it.get("qty", 0)
                taxbl = round(it.get("taxable", rate*qty), 2)
                gst_p = it.get("gst_percent", it.get("gst", 18))
                gst_a = round(it.get("gst_amt", 0), 2)
                grnd  = round(it.get("grand", taxbl + gst_a))
                unit_v = it.get("unit","Pcs") or "Pcs"
                it_rows.append([
                    Paragraph(str(idx+1),               ps("r",8,align=TA_CENTER)),
                    Paragraph(str(it.get("product","")),ps("r",8)),
                    Paragraph(str(it.get("hsn","")),    ps("r",8,align=TA_CENTER)),
                    Paragraph(str(unit_v),              ps("r",8,align=TA_CENTER)),
                    Paragraph(str(qty),                 ps("r",8,align=TA_CENTER)),
                    Paragraph(f"{rate:.2f}",            ps("r",8,align=TA_CENTER)),
                    Paragraph(f"{taxbl:.2f}",           ps("r",8,align=TA_CENTER)),
                    Paragraph(f"{gst_a:.2f}",           ps("r",8,align=TA_CENTER)),
                    Paragraph(f"{gst_a:.2f}",           ps("r",8,align=TA_CENTER)),
                    Paragraph(f"{grnd}",                ps("r",8,align=TA_RIGHT)),
                ])
            for _ in range(n_blank):
                it_rows.append(["","","","","","","","","",""])
            it_rows.append([
                "", Paragraph("<b>Total</b>", ps("r",8,True)),
                "","","","","","","",
                Paragraph(f"<b>{grand}</b>", ps("r",8,True,TA_RIGHT)),
            ])

            story.append(T(it_rows, CW,
                BASE + [
                    ("BACKGROUND",(0,0),(-1,0), GRY),
                    ("ALIGN",(0,0),(-1,-1),"CENTER"),
                ],
                repeat=1
            ))

            # ════════════════════════════════════════
            # 5. AMOUNT IN WORDS
            # ════════════════════════════════════════
            story.append(T(
                [[Paragraph("Amount Chargeable (In Words)", ps("aw",8)),
                  Paragraph("E. &amp; O E.", ps("aw",8,align=TA_RIGHT))],
                 [Paragraph(f"<b>{num_to_words(grand)}</b>", ps("aw",8,True)),
                  Paragraph("", ps("aw",8))]],
                [145*mm, 45*mm],
                BASE
            ))

            # ════════════════════════════════════════
            # 5b. CLASS-WISE GST SUMMARY (Pharmacy-style)
            # ════════════════════════════════════════
            cls_groups = defaultdict(lambda: {"sub":0,"disc":0,"taxbl":0,"gst":0})
            for it in items:
                gp = it.get("gst_percent", it.get("gst",18))
                taxbl = round(it.get("taxable", it.get("rate",0)*it.get("qty",0)), 2)
                gst_a = round(it.get("gst_amt",0), 2)
                disc_a = round(it.get("disc_amt",0), 2)
                cls_groups[gp]["sub"]   += taxbl + disc_a
                cls_groups[gp]["disc"]  += disc_a
                cls_groups[gp]["taxbl"] += taxbl
                cls_groups[gp]["gst"]   += gst_a

            cls_head = [
                Paragraph("<b>CLASS</b>",      ps("ch",7,True,TA_CENTER)),
                Paragraph("<b>SUB TOTAL</b>",  ps("ch",7,True,TA_CENTER)),
                Paragraph("<b>DISC</b>",       ps("ch",7,True,TA_CENTER)),
                Paragraph("<b>TAXBL.VL</b>",   ps("ch",7,True,TA_CENTER)),
                Paragraph("<b>SGST</b>",       ps("ch",7,True,TA_CENTER)),
                Paragraph("<b>CGST</b>",       ps("ch",7,True,TA_CENTER)),
                Paragraph("<b>TOTAL GST</b>",  ps("ch",7,True,TA_CENTER)),
                Paragraph("<b>SUB TOTAL</b>",  ps("ch",7,True,TA_CENTER)),
            ]
            cls_rows = [cls_head]
            t_sub=t_disc=t_taxbl=t_sgst=t_cgst=t_gst=t_final=0
            for gp in sorted(cls_groups.keys()):
                v = cls_groups[gp]
                half = round(v["gst"]/2, 2)
                final = round(v["taxbl"]+v["gst"], 2)
                t_sub+=v["sub"]; t_disc+=v["disc"]; t_taxbl+=v["taxbl"]
                t_sgst+=half; t_cgst+=half; t_gst+=v["gst"]; t_final+=final
                cls_rows.append([
                    Paragraph(f"GST {gp:.2f}%", ps("cr",8,align=TA_CENTER)),
                    Paragraph(f"{v['sub']:.2f}",   ps("cr",8,align=TA_CENTER)),
                    Paragraph(f"{v['disc']:.2f}",  ps("cr",8,align=TA_CENTER)),
                    Paragraph(f"{v['taxbl']:.2f}", ps("cr",8,align=TA_CENTER)),
                    Paragraph(f"{half:.2f}",       ps("cr",8,align=TA_CENTER)),
                    Paragraph(f"{half:.2f}",       ps("cr",8,align=TA_CENTER)),
                    Paragraph(f"{v['gst']:.2f}",   ps("cr",8,align=TA_CENTER)),
                    Paragraph(f"{final:.2f}",      ps("cr",8,align=TA_CENTER)),
                ])
            cls_rows.append([
                Paragraph("<b>SUB TOTAL</b>", ps("cr",8,True,TA_CENTER)),
                Paragraph(f"<b>{t_sub:.2f}</b>",   ps("cr",8,True,TA_CENTER)),
                Paragraph(f"<b>{t_disc:.2f}</b>",  ps("cr",8,True,TA_CENTER)),
                Paragraph(f"<b>{t_taxbl:.2f}</b>", ps("cr",8,True,TA_CENTER)),
                Paragraph(f"<b>{t_sgst:.2f}</b>",  ps("cr",8,True,TA_CENTER)),
                Paragraph(f"<b>{t_cgst:.2f}</b>",  ps("cr",8,True,TA_CENTER)),
                Paragraph(f"<b>{t_gst:.2f}</b>",   ps("cr",8,True,TA_CENTER)),
                Paragraph(f"<b>{t_final:.2f}</b>", ps("cr",8,True,TA_CENTER)),
            ])
            story.append(T(cls_rows,
                [22*mm, 24*mm, 20*mm, 24*mm, 20*mm, 20*mm, 24*mm, 26*mm],
                BASE + [("BACKGROUND",(0,0),(-1,0),GRY),
                        ("BACKGROUND",(0,-1),(-1,-1),GRY)]
            ))

            # Adjustment / Net Payable / TCS / CR-DR / Grand Total strip
            adj_amt   = round(d.get("adjustment",0) or 0, 2)
            tcs_pct   = d.get("tcs_percent",0) or 0
            tcs_amt   = round(d.get("tcs_amt",0) or 0, 2)
            crdr_amt  = round(d.get("crdr_amt",0) or 0, 2)
            net_payable = round(t_final - adj_amt, 2)

            misc_rows = [
                [Paragraph("ADJUSTMENT SLIP NO.: " + str(d.get("adj_slip_no","")), ps("mr",8)),
                 Paragraph(f"AMT.: {adj_amt:.2f}", ps("mr",8,align=TA_RIGHT)),
                 Paragraph(f"NET PAYABLE AMT.: {net_payable:.2f}", ps("mr",8,True,TA_RIGHT))],
                [Paragraph(f"CR/DR NOTE: {crdr_amt:.2f}", ps("mr",8)),
                 Paragraph(f"TCS ({tcs_pct:.2f}%): {tcs_amt:.2f}", ps("mr",8,align=TA_CENTER)),
                 Paragraph(f"<b>GRAND TOTAL: {grand}</b>", ps("mr",9,True,TA_RIGHT))],
            ]
            story.append(T(misc_rows,
                [70*mm, 60*mm, 60*mm],
                BASE
            ))

            # ════════════════════════════════════════
            # 6. GST SUMMARY
            # ════════════════════════════════════════
            gst_type_pdf = self.data.get("gst_type","CGST+SGST") or "CGST+SGST"
            is_igst_pdf  = (gst_type_pdf == "IGST")
            is_nongst_pdf = (self.data.get("bill_mode","GST") == "NONGST")

            gst_groups = defaultdict(lambda: {"taxable":0,"gst":0})
            for it in items:
                gp = it.get("gst_percent", it.get("gst",18))
                ga = round(it.get("gst_amt",0), 2)
                gst_groups[gp]["taxable"] += round(it.get("taxable", it.get("rate",0)*it.get("qty",0)), 2)
                gst_groups[gp]["gst"]     += ga

            sum_tax = round(sum(v["gst"] for v in gst_groups.values()), 2)

            if is_nongst_pdf:
                story.append(T(
                    [[Paragraph("<b>GST not charged on this bill (Non-GST Bill)</b>",
                                ps("tr",8,True,TA_CENTER))]],
                    [PAGE_W],
                    BASE
                ))
            elif is_igst_pdf:
                tax_head = [
                    Paragraph("", ps("th",7)),
                    Paragraph("<b>Taxable Amt</b>",  ps("th",7,True,TA_CENTER)),
                    Paragraph("<b>Rate (IGST)</b>",  ps("th",7,True,TA_CENTER)),
                    Paragraph("<b>IGST Amount</b>",  ps("th",7,True,TA_CENTER)),
                    Paragraph("<b>Total Tax</b>",    ps("th",7,True,TA_CENTER)),
                ]
                tax_rows = [tax_head]
                for gp, vals in gst_groups.items():
                    tt = round(vals["gst"], 2)
                    tax_rows.append([
                        Paragraph("", ps("tr",8)),
                        Paragraph(f"{vals['taxable']:.2f}", ps("tr",8,align=TA_CENTER)),
                        Paragraph(f"{gp:.0f}%",             ps("tr",8,align=TA_CENTER)),
                        Paragraph(f"{tt:.2f}",              ps("tr",8,align=TA_CENTER)),
                        Paragraph(f"{tt:.2f}",              ps("tr",8,align=TA_CENTER)),
                    ])
                tax_rows.append([
                    Paragraph("<b>Total</b>", ps("tr",8,True)),
                    Paragraph("", ps("tr",8)),
                    Paragraph("", ps("tr",8)),
                    Paragraph(f"<b>{sum_tax:.2f}</b>", ps("tr",8,True,TA_CENTER)),
                    Paragraph(f"<b>{sum_tax:.2f}</b>", ps("tr",8,True,TA_CENTER)),
                ])
                story.append(T(tax_rows,
                    [20*mm, 45*mm, 35*mm, 45*mm, 45*mm],
                    BASE + [("BACKGROUND",(0,0),(-1,0),GRY),
                            ("BACKGROUND",(0,-1),(-1,-1),GRY)]
                ))
            else:
                # CGST + SGST
                tax_head = [
                    Paragraph("", ps("th",7)),
                    Paragraph("<b>Rate (CGST)</b>",  ps("th",7,True,TA_CENTER)),
                    Paragraph("<b>CGST Amt</b>",     ps("th",7,True,TA_CENTER)),
                    Paragraph("<b>Rate (SGST)</b>",  ps("th",7,True,TA_CENTER)),
                    Paragraph("<b>SGST Amt</b>",     ps("th",7,True,TA_CENTER)),
                    Paragraph("<b>Total Tax</b>",    ps("th",7,True,TA_CENTER)),
                ]
                tax_rows = [tax_head]
                sum_cgst=0; sum_sgst=0
                for gp, vals in gst_groups.items():
                    cv = round(vals["gst"]/2, 2)
                    sv = round(vals["gst"]/2, 2)
                    tt = round(vals["gst"], 2)
                    sum_cgst+=cv; sum_sgst+=sv
                    tax_rows.append([
                        Paragraph("", ps("tr",8)),
                        Paragraph(f"{gp/2:.1f}%", ps("tr",8,align=TA_CENTER)),
                        Paragraph(f"{cv:.2f}",    ps("tr",8,align=TA_CENTER)),
                        Paragraph(f"{gp/2:.1f}%", ps("tr",8,align=TA_CENTER)),
                        Paragraph(f"{sv:.2f}",    ps("tr",8,align=TA_CENTER)),
                        Paragraph(f"{tt:.2f}",    ps("tr",8,align=TA_CENTER)),
                    ])
                tax_rows.append([
                    Paragraph("<b>Total</b>", ps("tr",8,True)),
                    Paragraph("", ps("tr",8)),
                    Paragraph(f"<b>{sum_cgst:.2f}</b>", ps("tr",8,True,TA_CENTER)),
                    Paragraph("", ps("tr",8)),
                    Paragraph(f"<b>{sum_sgst:.2f}</b>", ps("tr",8,True,TA_CENTER)),
                    Paragraph(f"<b>{sum_tax:.2f}</b>",  ps("tr",8,True,TA_CENTER)),
                ])
                story.append(T(tax_rows,
                    [20*mm, 30*mm, 38*mm, 30*mm, 38*mm, 34*mm],
                    BASE + [("BACKGROUND",(0,0),(-1,0),GRY),
                            ("BACKGROUND",(0,-1),(-1,-1),GRY)]
                ))

            # ════════════════════════════════════════
            # 7. TAX WORDS + BANK
            # ════════════════════════════════════════
            _spdf = get_shop()
            bank_txt = (
                "<b>Company's Bank details</b><br/>"
                f"A/C Holder's Name  : {_spdf['name']}<br/>"
                + (f"Bank Name           : {_spdf['bank']}<br/>" if _spdf['bank'] else "")
                + (f"A/C No.             : {_spdf['account']}<br/>" if _spdf['account'] else "")
                + (f"Branch And IFS Code : {_spdf['ifsc']}" if _spdf['ifsc'] else "")
                + (f"<br/>UPI              : {_spdf['upi']}" if _spdf['upi'] else "")
            )
            story.append(T(
                [[Paragraph(f"<b>Tax Amount (In Words) : {'N/A (Non-GST Bill)' if is_nongst_pdf else num_to_words(round(sum_tax))}</b>",
                            ps("tw",8)),
                  Paragraph(bank_txt, ps("bk",8))]],
                [90*mm, 100*mm],
                BASE + [("INNERGRID",(0,0),(-1,-1),0.5,BDR)]
            ))

            # ════════════════════════════════════════
            # 8. DECLARATION + SIGNATURE
            # ════════════════════════════════════════
            story.append(T(
                [[Paragraph(
                    "<b>Declaration</b><br/>"
                    "We declare that this invoice shows the actual price of<br/>"
                    "the goods described and that all particulars are true and correct",
                    ps("dc",8)),
                  Paragraph(
                    f"For  {_spdf['name']}<br/><br/><br/>Authorised Signatory",
                    ps("sg",8,align=TA_RIGHT))]],
                [110*mm, 80*mm],
                BASE + [("INNERGRID",(0,0),(-1,-1),0.5,BDR),
                        ("BOTTOMPADDING",(0,0),(-1,-1),20)]
            ))

            # ════════════════════════════════════════
            # 9. FOOTER
            # ════════════════════════════════════════
            story.append(T(
                [[Paragraph("SUBJECT TO VARANASI JURISDICTION",
                            ps("ft",8,align=TA_CENTER))]],
                [PAGE_W],
                [("BOX",(0,0),(-1,-1),0.5,BDR),
                 ("TOPPADDING",(0,0),(-1,-1),4),
                 ("BOTTOMPADDING",(0,0),(-1,-1),4)]
            ))

            doc.build(story)

            if sys.platform == "win32":
                os.startfile(save_path)
            elif sys.platform == "darwin":
                import subprocess
                subprocess.Popen(["open", save_path])
            else:
                import subprocess
                subprocess.Popen(["xdg-open", save_path])

            messagebox.showinfo("PDF Ready!",
                f"Invoice PDF save ho gayi:\n{save_path}\n\nAbhi khul rahi hai...")

        except ImportError:
            messagebox.showerror("Library Missing",
                "reportlab install nahi hai.\n\n"
                "CMD mein ye command chalao:\n"
                "pip install reportlab")
        except Exception as e:
            messagebox.showerror("Error", f"PDF nahi ban saki:\n{str(e)}")

    # ──────────────────────────────────────────────────────────────────────────
    #  THERMAL PRINT — ESC/POS direct USB print
    # ──────────────────────────────────────────────────────────────────────────
    def _thermal_print(self):
        """Thermal printer pe seedha ESC/POS print karo."""
        import os, sys, tempfile

        d     = self.data
        party = d["party"]
        items = d["items"]
        grand = round(d["grand"])
        _sp   = get_shop()
        total_taxable = round(sum(it.get("taxable", it.get("rate",0)*it.get("qty",0)) for it in items), 2)
        total_gst     = round(sum(it.get("gst_amt", 0) for it in items), 2)
        gst_type_val  = d.get("gst_type", "CGST+SGST") or "CGST+SGST"

        # ── Settings Dialog ───────────────────────────────────────────────────
        dlg = tk.Toplevel(self.win)
        dlg.title("Thermal Print")
        dlg.configure(bg="#FFFFFF")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.geometry(f"340x230+{(dlg.winfo_screenwidth()-340)//2}+{(dlg.winfo_screenheight()-230)//2}")

        tk.Label(dlg, text="🧾 Thermal Print",
                 font=("Segoe UI",12,"bold"), bg="#FFFFFF", fg="#1A365D").pack(pady=(14,6))

        # Paper width
        fw = tk.LabelFrame(dlg, text="Paper Width", bg="#FFFFFF", fg="#4A5568",
                           font=("Segoe UI",9), bd=1)
        fw.pack(fill="x", padx=18, pady=4)
        v_width = tk.StringVar(value="80mm")
        for val, lbl in [("58mm","58 mm  (2 inch — Chhota)"),
                          ("80mm","80 mm  (3 inch — Bada)")]:
            tk.Radiobutton(fw, text=lbl, variable=v_width, value=val,
                           font=("Segoe UI",9), bg="#FFFFFF",
                           activebackground="#FFFFFF").pack(anchor="w", padx=8, pady=1)

        # GST
        fg = tk.LabelFrame(dlg, text="GST", bg="#FFFFFF", fg="#4A5568",
                           font=("Segoe UI",9), bd=1)
        fg.pack(fill="x", padx=18, pady=4)
        v_gst = tk.BooleanVar(value=True)
        tk.Checkbutton(fg, text="HSN / GST detail print karo",
                       variable=v_gst, font=("Segoe UI",9),
                       bg="#FFFFFF", activebackground="#FFFFFF").pack(anchor="w", padx=8, pady=3)

        result = {"action": None}

        def do_preview():
            result["action"] = "preview"
            dlg.destroy()

        def do_print():
            result["action"] = "print"
            dlg.destroy()

        bf = tk.Frame(dlg, bg="#FFFFFF"); bf.pack(pady=12)
        make_btn(bf, "👁 Preview", do_preview, bg="#2B6CB0").pack(side="left", padx=5)
        make_btn(bf, "🖨 Print",   do_print,   bg="#276749").pack(side="left", padx=5)
        make_btn(bf, "✖ Cancel",  dlg.destroy, bg="#4A5568").pack(side="left", padx=5)
        dlg.wait_window()
        if not result["action"]:
            return

        paper_w_mm     = 58 if v_width.get() == "58mm" else 80
        show_gst       = v_gst.get()
        action         = result["action"]
        chars_per_line = 32 if paper_w_mm == 58 else 48

        # ── Build ESC/POS bytes ───────────────────────────────────────────────
        ESC = b'\x1b'
        GS  = b'\x1d'

        def center(txt):
            t = str(txt)[:chars_per_line]
            pad = (chars_per_line - len(t)) // 2
            return b' ' * pad + t.encode('ascii','replace') + b'\n'

        def left(txt):
            return str(txt)[:chars_per_line].encode('ascii','replace') + b'\n'

        def two_col(a, b):
            a, b = str(a), str(b)
            gap = chars_per_line - len(a) - len(b)
            if gap < 1: gap = 1
            return (a + ' '*gap + b).encode('ascii','replace') + b'\n'

        def divider(ch='-'):
            return (ch * chars_per_line).encode('ascii') + b'\n'

        buf = bytearray()
        buf += ESC + b'@'           # init
        buf += ESC + b'a\x01'       # center
        buf += GS  + b'!\x11'       # double size
        buf += ESC + b'E\x01'       # bold on
        buf += center(_sp.get("name","Shop").upper())
        buf += GS  + b'!\x00'       # normal size
        buf += ESC + b'E\x00'       # bold off
        if _sp.get("address"):
            buf += center((_sp["address"] or "") + (", "+_sp["city"] if _sp.get("city") else ""))
        if _sp.get("mobile"):
            buf += center("Mob: " + str(_sp["mobile"]))
        if _sp.get("gstin"):
            buf += center("GSTIN: " + str(_sp["gstin"]))
        buf += ESC + b'a\x00'       # left align
        buf += divider("=")

        buf += ESC + b'a\x01'
        buf += ESC + b'E\x01'
        buf += center("INVOICE / BILL")
        buf += ESC + b'E\x00'
        buf += ESC + b'a\x00'
        date_str = fmt_date(d["date"])
        bill_str = "Bill: " + str(d["bill_no"])
        # Agar dono fit ho jayein to ek line, warna alag alag
        if len(bill_str) + len(date_str) + 1 <= chars_per_line:
            buf += two_col(bill_str, date_str)
        else:
            buf += left(bill_str)
            buf += left("Date: " + date_str)
        buf += divider()

        if party.get("name"):
            buf += ESC + b'E\x01'
            buf += left("Party: " + str(party["name"]))
            buf += ESC + b'E\x00'
        if party.get("mobile"):
            buf += left("Mob: " + str(party["mobile"]))
        if show_gst and party.get("gstin"):
            buf += left("GSTIN: " + str(party["gstin"]))
        buf += divider()

        # Items — column widths
        # 58mm=32 chars: name(14) qty_rate(10) amt(8)
        # 80mm=48 chars: name(20) qty_rate(16) amt(12)
        if paper_w_mm == 58:
            w1, w2, w3 = 14, 10, 8
        else:
            w1, w2, w3 = 20, 16, 12

        def row3(c1, c2, c3):
            s1 = str(c1)[:w1].ljust(w1)
            s2 = str(c2)[:w2].ljust(w2)
            s3 = str(c3)[:w3].rjust(w3)
            return (s1+s2+s3).encode('ascii','replace') + b'\n'

        buf += ESC + b'E\x01'
        buf += row3("Item", "Qty x Rate", "Amt")
        buf += ESC + b'E\x00'
        buf += divider()

        for it in items:
            name    = str(it.get("product",""))
            qty     = it.get("qty", 0)
            rate    = it.get("rate", 0)
            taxable = it.get("taxable", round(float(rate)*float(qty), 2))
            gst_amt = it.get("gst_amt", 0)
            total   = round(taxable + gst_amt, 2)
            gp      = it.get("gst_percent", it.get("gst", 0))
            hsn     = it.get("hsn","")

            # Item name + qty+rate+amount ek hi row mein
            qty_rate = f"{qty}x{rate}"
            buf += row3(name, qty_rate, f"{total:.2f}")
            if show_gst:
                info = f"  HSN:{hsn} GST:{gp}%" if hsn else f"  GST:{gp}%"
                buf += left(info)

        buf += divider("=")
        buf += two_col("Taxable Amt:", f"Rs.{total_taxable:.2f}")
        if show_gst:
            if "IGST" in gst_type_val:
                buf += two_col("IGST:", f"Rs.{total_gst:.2f}")
            else:
                buf += two_col("CGST:", f"Rs.{total_gst/2:.2f}")
                buf += two_col("SGST:", f"Rs.{total_gst/2:.2f}")
        buf += divider("=")
        buf += ESC + b'E\x01'
        buf += two_col("TOTAL:", f"Rs.{grand:.2f}")
        buf += ESC + b'E\x00'
        buf += divider("=")
        if d.get("pay_mode"):
            buf += two_col("Payment:", str(d["pay_mode"]))
        buf += b'\n'
        buf += ESC + b'a\x01'
        buf += center("** Thank You! Phir Padharo! **")
        buf += ESC + b'a\x00'
        buf += ESC + b'd\x04'       # feed 4 lines
        buf += GS  + b'VA\x03'      # cut

        receipt = bytes(buf)

        # ── Action: Preview ya Print ──────────────────────────────────────────
        if action == "preview":
            # ESC/POS bytes parse karo — align command track karo
            lines_out = []
            i = 0
            current_line = b""
            current_align = "L"  # L=left, C=center, R=right
            line_aligns = []

            while i < len(receipt):
                b = receipt[i:i+1]
                if b == b'\x1b':
                    i += 1
                    if i < len(receipt):
                        cmd = receipt[i:i+1]
                        i += 1
                        if cmd == b'a' and i < len(receipt):
                            v = receipt[i]
                            current_align = "C" if v == 1 else ("R" if v == 2 else "L")
                            i += 1
                        elif cmd in (b'E', b'd') and i < len(receipt):
                            i += 1  # skip param
                        # '@' = init
                elif b == b'\x1d':
                    i += 3
                elif b == b'\n':
                    lines_out.append(current_line.decode("ascii", errors="replace"))
                    line_aligns.append(current_align)
                    current_line = b""
                    i += 1
                elif b >= b' ':
                    current_line += b
                    i += 1
                else:
                    i += 1
            if current_line:
                lines_out.append(current_line.decode("ascii", errors="replace"))
                line_aligns.append(current_align)

            # Ab har line ko uske align ke hisaab se format karo
            formatted = []
            for line, align in zip(lines_out, line_aligns):
                stripped = line.strip()
                if align == "C":
                    formatted.append(stripped.center(chars_per_line))
                elif align == "R":
                    formatted.append(stripped.rjust(chars_per_line))
                else:
                    formatted.append(stripped)
            preview_text = "\n".join(formatted)

            # Preview window
            pwin = tk.Toplevel(self.win)
            pwin.title("Receipt Preview")
            pwin.configure(bg="#FFFFFF")
            pwin.resizable(True, True)
            # Courier New char width: size9=7.2px, size10=8.4px
            char_px = 7.2 if paper_w_mm == 58 else 8.4
            pw_px = int(chars_per_line * char_px) + 60  # exact fit + small padding

            tk.Label(pwin, text="🧾 Receipt Preview",
                     font=("Segoe UI",11,"bold"), bg="#FFFFFF", fg="#1A365D").pack(pady=(10,2))
            tk.Label(pwin, text=f"Paper: {paper_w_mm}mm  |  {chars_per_line} chars/line",
                     font=("Segoe UI",7), bg="#FFFFFF", fg="#718096").pack(pady=(0,4))

            # White receipt-like background — no scrollbar, full content dikhao
            frm_t = tk.Frame(pwin, bg="#FFFEF0", bd=2, relief="groove")
            frm_t.pack(fill="both", expand=True, padx=14, pady=(0,3))
            fsize = 9 if paper_w_mm == 58 else 10
            # Text centered using padx — receipt bich mein dikhega
            txt = tk.Text(frm_t, font=("Courier New", fsize), bg="#FFFEF0", fg="#111111",
                          wrap="none", relief="flat", padx=6, pady=10, state="disabled")
            txt.pack(fill="both", expand=True)

            txt.config(state="normal")
            txt.insert("1.0", preview_text)
            txt.config(state="disabled")

            # Window height: line count ke hisaab se auto-set karo
            line_count = preview_text.count("\n") + 2
            char_h = fsize + 4  # approx pixels per line
            content_h = line_count * char_h + 100  # +100 for header+buttons
            max_h = pwin.winfo_screenheight() - 80
            win_h = min(content_h, max_h)
            pwin.geometry(f"{pw_px}x{win_h}+{(pwin.winfo_screenwidth()-pw_px)//2}+{(pwin.winfo_screenheight()-win_h)//2}")

            # Buttons
            bf2 = tk.Frame(pwin, bg="#FFFFFF"); bf2.pack(pady=(0,10))

            def do_print_from_preview():
                pwin.destroy()
                self._send_escpos(receipt)

            make_btn(bf2, "🖨 Print Karo", do_print_from_preview, bg="#276749").pack(side="left", padx=6)
            make_btn(bf2, "✖ Close",       pwin.destroy,           bg="#4A5568").pack(side="left", padx=6)
            return

        # ── Direct Print via Windows default printer (RAW ESC/POS) ───────────
        self._send_escpos(receipt)

    # ──────────────────────────────────────────────────────────────────────────
    def _send_escpos(self, receipt):
        """ESC/POS bytes Windows default printer pe bhejo."""
        import sys
        if sys.platform != "win32":
            messagebox.showerror("Error", "Yeh feature sirf Windows pe kaam karta hai.")
            return

        try:
            import win32print
            printer_name = win32print.GetDefaultPrinter()
            hp = win32print.OpenPrinter(printer_name)
            try:
                win32print.StartDocPrinter(hp, 1, ("Receipt", None, "RAW"))
                try:
                    win32print.StartPagePrinter(hp)
                    win32print.WritePrinter(hp, receipt)
                    win32print.EndPagePrinter(hp)
                finally:
                    win32print.EndDocPrinter(hp)
            finally:
                win32print.ClosePrinter(hp)
            messagebox.showinfo("Print Ho Gaya!",
                f"Receipt print ho gayi!\nPrinter: {printer_name}")

        except ImportError:
            # pywin32 nahi hai — install karne ka message
            ans = messagebox.askyesno(
                "pywin32 Chahiye",
                "Thermal print ke liye pywin32 library chahiye.\n\n"
                "Abhi install karein?\n"
                "(CMD khulega aur pip install chalega)")
            if ans:
                import subprocess
                subprocess.Popen(
                    ["cmd", "/k", "pip install pywin32 && echo DONE! && pause"],
                    creationflags=subprocess.CREATE_NEW_CONSOLE)
                messagebox.showinfo("Install Ho Raha Hai",
                    "CMD mein install ho raha hai.\n"
                    "Complete hone ke baad software restart karein aur phir print karein.")

        except Exception as ex:
            messagebox.showerror("Print Error",
                f"Print nahi ho saka:\n{str(ex)}\n\n"
                "Check karein:\n"
                "1. Thermal printer on hai aur USB laga hai?\n"
                "2. Windows mein default printer set hai?")


    # ──────────────────────────────────────────────────────────────────────────
    #  HELPER: PDF temp folder mein banao, path return karo
    # ──────────────────────────────────────────────────────────────────────────
    def _generate_pdf_temp(self):
        """Invoice PDF temp folder mein banao aur path return karo. Error pe None."""
        import os, tempfile
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
            from collections import defaultdict
        except ImportError:
            messagebox.showerror("Library Missing",
                "reportlab install nahi hai.\n\nCMD mein chalao:\npip install reportlab")
            return None

        d     = self.data
        party = d["party"]
        items = d["items"]
        grand = round(d["grand"])
        total_taxable = round(sum(it.get("taxable", it.get("rate",0)*it.get("qty",0)) for it in items), 2)
        total_gst     = round(sum(it.get("gst_amt", 0) for it in items), 2)

        safe_bill = d["bill_no"].replace("/", "_").replace("\\", "_")
        tmp_path  = os.path.join(tempfile.gettempdir(), f"{safe_bill}.pdf")

        PAGE_W = 190 * mm
        BDR    = colors.black
        GRY    = colors.HexColor("#f2f2f2")

        def ps(name, size=8, bold=False, align=TA_LEFT):
            return ParagraphStyle(name, fontSize=size,
                fontName="Helvetica-Bold" if bold else "Helvetica",
                alignment=align, leading=size+3)

        def T(data, cw, style_cmds, repeat=0):
            tbl = Table(data, colWidths=cw, style=TableStyle(style_cmds), repeatRows=repeat)
            tbl.spaceAfter=0; tbl.spaceBefore=0
            return tbl

        BASE = [
            ("BOX",(0,0),(-1,-1),0.5,BDR),
            ("INNERGRID",(0,0),(-1,-1),0.3,BDR),
            ("TOPPADDING",(0,0),(-1,-1),3),
            ("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("LEFTPADDING",(0,0),(-1,-1),3),
            ("RIGHTPADDING",(0,0),(-1,-1),3),
            ("VALIGN",(0,0),(-1,-1),"TOP"),
        ]

        try:
            doc = SimpleDocTemplate(tmp_path, pagesize=A4,
                leftMargin=10*mm, rightMargin=10*mm,
                topMargin=8*mm, bottomMargin=8*mm)

            bill_date = fmt_date(d["date"])
            p_name = party.get("name","")
            p_addr = party.get("address","").replace("\n","<br/>")
            p_mob  = party.get("mobile","")
            p_gst  = party.get("gstin","")

            _sp2 = get_shop()
            _sp2_addr = f"{_sp2['address']}, {_sp2['city']}".strip(", ")
            company_txt = (
                f"<b>{_sp2['name']}</b><br/>"
                + (f"{_sp2_addr}<br/>" if _sp2_addr else "")
                + (f"Mob: {_sp2['mobile']}<br/>" if _sp2['mobile'] else "")
                + (f"GSTIN/UIN : {_sp2['gstin']}<br/>" if _sp2['gstin'] else "")
                + (f"State Name : {_sp2['state']}" if _sp2['state'] else "")
            )

            rw = [33*mm, 32*mm, 38*mm, 22*mm]
            right_data = [
                [Paragraph("Invoice :",ps("r",8,True)), Paragraph(d["bill_no"],ps("r",8)),
                 Paragraph("Dated :",ps("r",8,True)),   Paragraph(bill_date,ps("r",8))],
                [Paragraph("Delivery Note",ps("r",8)),  Paragraph("",ps("r",8)),
                 Paragraph("Mode/Terms",ps("r",8)),     Paragraph("",ps("r",8))],
                [Paragraph("Reference No.",ps("r",8)),  Paragraph("",ps("r",8)),
                 Paragraph("Other Ref",ps("r",8)),      Paragraph("",ps("r",8))],
                [Paragraph("Buyers Order No",ps("r",8)),Paragraph("",ps("r",8)),
                 Paragraph("Dated",ps("r",8)),          Paragraph("",ps("r",8))],
                [Paragraph("Dispatch Doc No.",ps("r",8)),Paragraph("",ps("r",8)),
                 Paragraph("Delivery Note Date",ps("r",8)),Paragraph("",ps("r",8))],
                [Paragraph("Dispatched through",ps("r",8)),Paragraph("",ps("r",8)),
                 Paragraph("Destination",ps("r",8)),    Paragraph("",ps("r",8))],
                [Paragraph("Terms of Delivery",ps("r",8)),Paragraph("",ps("r",8,False,TA_LEFT)),
                 Paragraph("",ps("r",8)),               Paragraph("",ps("r",8))],
            ]

            story = []
            story.append(T([[Paragraph("Invoice",ps("t",11,False,TA_CENTER))]],[PAGE_W],
                [("BOX",(0,0),(-1,-1),0.5,BDR),
                 ("TOPPADDING",(0,0),(-1,-1),4),
                 ("BOTTOMPADDING",(0,0),(-1,-1),4)]))

            story.append(T(
                [[Paragraph(company_txt,ps("c",8)),
                  Table(right_data,colWidths=rw,style=TableStyle(BASE))]],
                [65*mm, 125*mm], BASE))

            bill_to = f"<b>Consignee (Bill to)</b><br/>{p_name}<br/>{p_addr}<br/>Mob: {p_mob}<br/>GSTIN: {p_gst}"
            story.append(T([[Paragraph(bill_to,ps("bt",8)),Paragraph("",ps("bt",8))]],
                [95*mm,95*mm],BASE))

            # Items table
            hdr = [Paragraph(x,ps("h",8,True,TA_CENTER)) for x in
                   ["Sl","Description","HSN","Unit","Qty","Rate","Taxable","GST%","GST Amt","Total"]]
            rows = [hdr]
            sum_tax=0
            for i,it in enumerate(items,1):
                qty   = it.get("qty",0)
                rate  = it.get("rate",0)
                taxbl = it.get("taxable", round(qty*rate,2))
                gst_p = it.get("gst_percent",0)
                gst_a = it.get("gst_amt",0)
                tot   = it.get("grand", round(taxbl+gst_a,2))
                unit_v = it.get("unit","Pcs") or "Pcs"
                sum_tax += tot
                rows.append([
                    Paragraph(str(i),ps("i",8,align=TA_CENTER)),
                    Paragraph(str(it.get("product","")),ps("i",8)),
                    Paragraph(str(it.get("hsn","")),ps("i",8,align=TA_CENTER)),
                    Paragraph(str(unit_v),ps("i",8,align=TA_CENTER)),
                    Paragraph(str(qty),ps("i",8,align=TA_CENTER)),
                    Paragraph(f"{rate:.2f}",ps("i",8,align=TA_RIGHT)),
                    Paragraph(f"{taxbl:.2f}",ps("i",8,align=TA_RIGHT)),
                    Paragraph(f"{gst_p:.1f}%",ps("i",8,align=TA_CENTER)),
                    Paragraph(f"{gst_a:.2f}",ps("i",8,align=TA_RIGHT)),
                    Paragraph(f"{tot:.2f}",ps("i",8,align=TA_RIGHT)),
                ])
            rows.append([Paragraph("",ps("i",8))]*8 +
                [Paragraph(f"{total_gst:.2f}",ps("i",8,True,TA_RIGHT)),
                 Paragraph(f"{grand}",ps("i",8,True,TA_RIGHT))])

            story.append(T(rows,
                [8*mm,48*mm,14*mm,12*mm,12*mm,16*mm,17*mm,12*mm,18*mm,18*mm],
                BASE+[("BACKGROUND",(0,0),(-1,0),GRY),
                      ("BACKGROUND",(0,-1),(-1,-1),GRY)],repeat=1))

            def num_to_words(n):
                ones=["","One","Two","Three","Four","Five","Six","Seven","Eight","Nine",
                      "Ten","Eleven","Twelve","Thirteen","Fourteen","Fifteen","Sixteen",
                      "Seventeen","Eighteen","Nineteen"]
                tens=["","","Twenty","Thirty","Forty","Fifty","Sixty","Seventy","Eighty","Ninety"]
                if n==0: return "Zero"
                def _h(x):
                    if x==0: return ""
                    elif x<20: return ones[x]+" "
                    elif x<100: return tens[x//10]+" "+(ones[x%10]+" " if x%10 else "")
                    else: return ones[x//100]+" Hundred "+(("and "+_h(x%100)) if x%100 else "")
                n=int(n); r=""
                if n>=10000000: r+=_h(n//10000000)+"Crore "; n%=10000000
                if n>=100000:   r+=_h(n//100000)+"Lakh ";   n%=100000
                if n>=1000:     r+=_h(n//1000)+"Thousand "; n%=1000
                r+=_h(n)
                return r.strip()+" Only"

            _spdf2 = get_shop()
            bank_txt = (
                "<b>Company's Bank details</b><br/>"
                f"A/C Holder : {_spdf2['name']}<br/>"
                + (f"Bank : {_spdf2['bank']}<br/>" if _spdf2['bank'] else "")
                + (f"A/C No. : {_spdf2['account']}<br/>" if _spdf2['account'] else "")
                + (f"IFS Code : {_spdf2['ifsc']}" if _spdf2['ifsc'] else "")
                + (f"<br/>UPI : {_spdf2['upi']}" if _spdf2['upi'] else "")
            )
            story.append(T(
                [[Paragraph(f"<b>Tax Amount (In Words) : {num_to_words(round(sum_tax))}</b>",ps("tw",8)),
                  Paragraph(bank_txt,ps("bk",8))]],
                [90*mm,100*mm],
                BASE+[("INNERGRID",(0,0),(-1,-1),0.5,BDR)]))

            story.append(T(
                [[Paragraph("<b>Declaration</b><br/>We declare that this invoice shows the actual price of the goods described and that all particulars are true and correct.",ps("dc",8)),
                  Paragraph(f"For {_spdf2['name']}<br/><br/><br/>Authorised Signatory",ps("sg",8,align=TA_RIGHT))]],
                [110*mm,80*mm],
                BASE+[("INNERGRID",(0,0),(-1,-1),0.5,BDR),
                      ("BOTTOMPADDING",(0,0),(-1,-1),20)]))

            story.append(T(
                [[Paragraph("SUBJECT TO VARANASI JURISDICTION",ps("ft",8,align=TA_CENTER))]],
                [PAGE_W],
                [("BOX",(0,0),(-1,-1),0.5,BDR),
                 ("TOPPADDING",(0,0),(-1,-1),4),
                 ("BOTTOMPADDING",(0,0),(-1,-1),4)]))

            doc.build(story)
            return tmp_path

        except Exception as e:
            messagebox.showerror("Error", f"PDF generate nahi hui:\n{str(e)}")
            return None

    # ──────────────────────────────────────────────────────────────────────────
    #  EMAIL SHARE
    # ──────────────────────────────────────────────────────────────────────────
    def _share_email(self):
        """Invoice PDF email se share karo (default mail client via mailto)."""
        import os, urllib.parse, webbrowser

        pdf_path = self._generate_pdf_temp()
        if not pdf_path:
            return

        d = self.data
        party = d["party"]
        to_email = party.get("email", "").strip()

        # Email dialog window
        dlg = tk.Toplevel(self.win)
        dlg.title("📧 Email se Invoice Bhejo")
        dlg.configure(bg=C_WHITE)
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.geometry("420x280")

        tk.Label(dlg, text="📧  Invoice Email Karo", font=("Segoe UI",11,"bold"),
                 bg="#2B6CB0", fg="white", pady=4).pack(fill="x")

        frm = tk.Frame(dlg, bg=C_WHITE, padx=14, pady=6)
        frm.pack(fill="both", expand=True)

        tk.Label(frm, text="To (Email):", font=("Segoe UI",9), bg=C_WHITE).grid(row=0,column=0,sticky="w",pady=6)
        v_to = tk.StringVar(value=to_email)
        ttk.Entry(frm, textvariable=v_to, width=32).grid(row=0,column=1,padx=8,pady=6)

        tk.Label(frm, text="Subject:", font=("Segoe UI",9), bg=C_WHITE).grid(row=1,column=0,sticky="w",pady=6)
        v_sub = tk.StringVar(value=f"Invoice {d['bill_no']} from BHUGTANEASE")
        ttk.Entry(frm, textvariable=v_sub, width=32).grid(row=1,column=1,padx=8,pady=6)

        tk.Label(frm, text="Body:", font=("Segoe UI",9), bg=C_WHITE).grid(row=2,column=0,sticky="nw",pady=6)
        body_txt = tk.Text(frm, width=32, height=4, font=("Segoe UI",9))
        body_txt.grid(row=2,column=1,padx=8,pady=6)
        body_txt.insert("1.0",
            f"Dear {party.get('name','Customer')},\n\n"
            f"Please find attached Invoice {d['bill_no']}.\n\nRegards,\nBHUGTANEASE")

        info = tk.Label(frm,
            text=f"📎 PDF ready: {os.path.basename(pdf_path)}",
            font=("Segoe UI",7), bg=C_WHITE, fg="#2B6CB0")
        info.grid(row=3,column=0,columnspan=2,pady=(4,0),sticky="w")

        def open_mail():
            subject = urllib.parse.quote(v_sub.get())
            body    = urllib.parse.quote(body_txt.get("1.0","end-1c"))
            to      = v_to.get().strip()
            mailto  = f"mailto:{to}?subject={subject}&body={body}"
            webbrowser.open(mailto)
            # PDF folder mein open karo taaki attach kar sako
            os.startfile(os.path.dirname(pdf_path)) if os.name=="nt" else None
            messagebox.showinfo("Email",
                f"Mail client khul gaya!\n\nPDF yahaan se attach karo:\n{pdf_path}",
                parent=dlg)
            dlg.destroy()

        bf = tk.Frame(dlg, bg=C_WHITE)
        bf.pack(pady=(0,12))
        make_btn(bf,"✉️  Mail Client Khullo", open_mail, bg="#2B6CB0").pack(side="left",padx=6)
        make_btn(bf,"Cancel", dlg.destroy, bg=C_GRAY).pack(side="left",padx=6)

    # ──────────────────────────────────────────────────────────────────────────
    #  WHATSAPP SHARE
    # ──────────────────────────────────────────────────────────────────────────
    def _share_whatsapp(self):
        """Invoice details WhatsApp pe bhejo (Web WhatsApp browser mein khulunga)."""
        import webbrowser, urllib.parse, os

        d     = self.data
        party = d["party"]
        items = d["items"]
        grand = round(d["grand"])

        # PDF bhi generate karo
        pdf_path = self._generate_pdf_temp()

        # Message banao
        lines_msg = [
            f"*Invoice: {d['bill_no']}*",
            f"Date: {fmt_date(d['date'])}",
            f"Customer: {party.get('name','')}",
            "",
            "*Items:*"
        ]
        for it in items:
            lines_msg.append(
                f"• {it.get('product','')}  Qty:{it.get('qty',0)}  "
                f"Rate:{it.get('rate',0):.2f}  Total:{it.get('grand', round(it.get('taxable',0)+it.get('gst_amt',0),2)):.2f}"
            )
        lines_msg += ["", f"*Grand Total: Rs. {grand}*", "", "From: BHUGTANEASE"]
        message = "\n".join(lines_msg)

        # WhatsApp dialog
        dlg = tk.Toplevel(self.win)
        dlg.title("💬 WhatsApp se Bhejo")
        dlg.configure(bg=C_WHITE)
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.geometry("400x320")

        tk.Label(dlg, text="💬  WhatsApp Invoice Share", font=("Segoe UI",11,"bold"),
                 bg="#25D366", fg="white", pady=4).pack(fill="x")

        frm = tk.Frame(dlg, bg=C_WHITE, padx=20, pady=4)
        frm.pack(fill="both", expand=True)

        tk.Label(frm, text="Mobile No (10 digit):", font=("Segoe UI",9), bg=C_WHITE).pack(anchor="w")
        v_mob = tk.StringVar(value=party.get("mobile","").strip().replace("+91","").replace(" ",""))
        ttk.Entry(frm, textvariable=v_mob, width=20).pack(anchor="w", pady=(2,8))

        tk.Label(frm, text="Message Preview:", font=("Segoe UI",9), bg=C_WHITE).pack(anchor="w")
        txt = tk.Text(frm, width=44, height=8, font=("Courier",8))
        txt.pack(fill="x")
        txt.insert("1.0", message)

        if pdf_path:
            tk.Label(frm, text=f"📎 PDF: {os.path.basename(pdf_path)}",
                     font=("Segoe UI",7), bg=C_WHITE, fg="#25D366").pack(anchor="w", pady=(4,0))

        def send_wa():
            mob  = v_mob.get().strip().replace(" ","").replace("-","")
            msg  = txt.get("1.0","end-1c")
            if not mob:
                messagebox.showwarning("Mobile?", "Mobile number daalo!", parent=dlg); return
            # 10 digit → add 91 prefix
            if len(mob)==10 and mob.isdigit():
                mob = "91" + mob
            encoded = urllib.parse.quote(msg)
            url = f"https://wa.me/{mob}?text={encoded}"
            webbrowser.open(url)
            if pdf_path:
                # PDF folder open karo taaki manually send kar sakein
                if os.name == "nt":
                    os.startfile(os.path.dirname(pdf_path))
                messagebox.showinfo("WhatsApp",
                    f"WhatsApp Web browser mein khul gaya!\n\n"
                    f"PDF manually attach karne ke liye folder khul gaya:\n{pdf_path}",
                    parent=dlg)
            dlg.destroy()

        bf = tk.Frame(dlg, bg=C_WHITE)
        bf.pack(pady=(4,10))
        make_btn(bf,"💬  WhatsApp Bhejo", send_wa, bg="#25D366").pack(side="left",padx=6)
        make_btn(bf,"Cancel", dlg.destroy, bg=C_GRAY).pack(side="left",padx=6)

    # ──────────────────────────────────────────────────────────────────────────
    #  SMS SHARE
    # ──────────────────────────────────────────────────────────────────────────
    def _share_sms(self):
        """Invoice ka SMS reminder bhejo."""
        import urllib.parse, subprocess, sys

        d     = self.data
        party = d["party"]
        items = d["items"]
        grand = round(d["grand"])

        # Short SMS message banao
        item_summary = ", ".join(
            f"{it.get('product','?')}x{it.get('qty',0)}"
            for it in items[:3]
        )
        if len(items) > 3:
            item_summary += f" +{len(items)-3} more"

        _ssms = get_shop()
        default_msg = (
            f"{_ssms['name']}: Dear {party.get('name','Customer')}, "
            f"Invoice {d['bill_no']} dt {fmt_date(d['date'])} "
            f"Rs.{grand:,} generated. "
            f"Items: {item_summary}. "
            + (f"Contact: {_ssms['mobile']}. " if _ssms['mobile'] else "")
            + f"{_ssms['name']} {_ssms['city'] or ''}".strip()
        )

        mobile = party.get("mobile","").strip()

        # Dialog
        dlg = tk.Toplevel(self.win)
        dlg.title("📱 SMS se Invoice Bhejo")
        dlg.configure(bg=C_WHITE)
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.geometry("440x380")

        tk.Label(dlg, text="📱  SMS Invoice Share", font=("Segoe UI",11,"bold"),
                 bg="#7B2D8B", fg="white", pady=4).pack(fill="x")

        frm = tk.Frame(dlg, bg=C_WHITE, padx=14, pady=6)
        frm.pack(fill="both", expand=True)

        tk.Label(frm, text="Mobile No (10 digit):", font=("Segoe UI",9), bg=C_WHITE).pack(anchor="w")
        v_mob = tk.StringVar(value=mobile.replace("+91","").replace(" ",""))
        ttk.Entry(frm, textvariable=v_mob, width=20).pack(anchor="w", pady=(2,10))

        tk.Label(frm, text="Message:", font=("Segoe UI",9), bg=C_WHITE).pack(anchor="w")
        txt = tk.Text(frm, width=48, height=7, font=("Segoe UI",9), wrap="word")
        txt.pack(fill="x")
        txt.insert("1.0", default_msg)

        # Char counter
        char_lbl = tk.Label(frm,
            text=f"{len(default_msg)} chars | {(len(default_msg)-1)//160+1} SMS",
            font=("Segoe UI",7), bg=C_WHITE, fg=C_GRAY)
        char_lbl.pack(anchor="e", pady=(2,0))

        def _update_count(ev=None):
            t = txt.get("1.0","end-1c")
            char_lbl.config(text=f"{len(t)} chars | {(len(t)-1)//160+1} SMS")
        txt.bind("<KeyRelease>", _update_count)

        tk.Label(frm,
            text="💡 160 chars = 1 SMS. Chhota message = kam charge.",
            font=("Segoe UI",7), bg=C_WHITE, fg="#718096").pack(anchor="w", pady=(4,0))

        def do_open_sms():
            mob = v_mob.get().strip().replace(" ","").replace("-","")
            if not mob:
                messagebox.showwarning("Mobile?","Mobile number daalo!",parent=dlg); return
            if len(mob)==10 and mob.isdigit(): mob = "+91"+mob
            elif not mob.startswith("+"): mob = "+91"+mob
            msg2 = txt.get("1.0","end-1c")
            sms_url = f"sms:{mob}?body={urllib.parse.quote(msg2)}"
            if sys.platform == "win32":
                try:
                    subprocess.Popen(["start","",sms_url], shell=True)
                except Exception:
                    pass
            else:
                import webbrowser
                webbrowser.open(sms_url)
            messagebox.showinfo("SMS",
                "SMS app kholne ki koshish ki gayi.\n\n"
                "Agar app nahi khula toh 'Copy Message'\n"
                "se manually copy karke bhejo.",
                parent=dlg)

        def do_copy():
            msg2 = txt.get("1.0","end-1c")
            mob  = v_mob.get().strip()
            dlg.clipboard_clear()
            dlg.clipboard_append(msg2)
            messagebox.showinfo("Copied!",
                f"Message copy ho gaya!\nMobile: {mob}\n\nAb SMS app mein paste karo.",
                parent=dlg)

        bf = tk.Frame(dlg, bg=C_WHITE)
        bf.pack(pady=(4,12))
        make_btn(bf, "📱 SMS App Khullo", do_open_sms, bg="#7B2D8B").pack(side="left", padx=5)
        make_btn(bf, "📋 Copy Message",   do_copy,      bg=C_AMBER ).pack(side="left", padx=5)
        make_btn(bf, "Cancel",            dlg.destroy,  bg=C_GRAY  ).pack(side="left", padx=5)

    # ──────────────────────────────────────────────────────────────────────────
    #  E-INVOICE GENERATOR
    # ──────────────────────────────────────────────────────────────────────────
    def _gen_einvoice(self):
        """E-Invoice JSON generate karo (IRP format) aur IRN save karo."""
        import json, hashlib, datetime as _dt

        d     = self.data
        party = d["party"]
        items = d["items"]
        grand = round(d["grand"], 2)
        gst_type = d.get("gst_type","CGST+SGST")

        # Validate
        if not party.get("gstin","").strip():
            if not messagebox.askyesno("Warning",
                "Buyer ka GSTIN nahi hai — B2C invoice.\n"
                "B2C invoices ke liye e-invoice mandatory nahi hoti.\n\n"
                "Phir bhi JSON generate karo?"):
                return

        # Build e-invoice JSON (GST IRP Schema v1.1)
        total_taxable = round(sum(it.get("taxable", it.get("rate",0)*it.get("qty",0)) for it in items), 2)
        total_gst     = round(sum(it.get("gst_amt",0) for it in items), 2)

        item_list = []
        for i, it in enumerate(items, 1):
            gp    = it.get("gst_percent", it.get("gst", 18))
            taxbl = round(it.get("taxable", it.get("rate",0)*it.get("qty",0)), 2)
            ga    = round(it.get("gst_amt",0), 2)
            item_obj = {
                "SlNo": str(i),
                "PrdDesc": it.get("product",""),
                "IsServc": "N",
                "HsnCd": it.get("hsn",""),
                "Qty": it.get("qty", 1),
                "Unit": "NOS",
                "UnitPrice": round(it.get("rate",0), 2),
                "TotAmt": taxbl,
                "AssAmt": taxbl,
                "GstRt": gp,
                "IgstAmt": round(ga,2) if gst_type=="IGST" else 0,
                "CgstAmt": round(ga/2,2) if gst_type!="IGST" else 0,
                "SgstAmt": round(ga/2,2) if gst_type!="IGST" else 0,
                "CesAmt": 0,
                "TotItemVal": round(taxbl+ga, 2),
            }
            item_list.append(item_obj)

        _sei = get_shop()
        einv = {
            "Version": "1.1",
            "TranDtls": {
                "TaxSch": "GST",
                "SupTyp": "B2B" if party.get("gstin","").strip() else "B2C",
                "RegRev": "N",
                "IgstOnIntra": "N"
            },
            "DocDtls": {
                "Typ": "INV",
                "No":  d["bill_no"],
                "Dt":  d["date"].replace("-","/") if "-" in d["date"] else d["date"],
            },
            "SellerDtls": {
                "Gstin": _sei['gstin'] or "99AAAAA0000A1Z5",
                "LglNm": _sei['name'],
                "Addr1": _sei['address'] or "",
                "Loc":   _sei['city'] or "",
                "Pin":   999999,
                "Stcd":  "09",
            },
            "BuyerDtls": {
                "Gstin": party.get("gstin","") or "URP",
                "LglNm": party.get("name",""),
                "Pos":   "09",  # Place of Supply code (default UP=09)
                "Addr1": party.get("address","") or "",
                "Loc":   party.get("state","") or "Uttar Pradesh",
                "Pin":   999999,
                "Stcd":  "09",
            },
            "ItemList": item_list,
            "ValDtls": {
                "AssVal":  total_taxable,
                "CgstVal": round(total_gst/2,2) if gst_type!="IGST" else 0,
                "SgstVal": round(total_gst/2,2) if gst_type!="IGST" else 0,
                "IgstVal": round(total_gst,2)   if gst_type=="IGST"  else 0,
                "CesVal":  0,
                "TotInvVal": grand,
            },
        }

        # Generate simulated IRN (SHA-256 of key fields — actual IRN from IRP portal)
        irn_raw = f"{get_shop()['gstin'] or 'NGSTIN'}|{d['bill_no']}|{d['date']}|{grand}"
        irn_sim = hashlib.sha256(irn_raw.encode()).hexdigest()

        json_str = json.dumps(einv, indent=2, ensure_ascii=False)

        # ── Dialog ──────────────────────────────────────────────────────────
        dlg = tk.Toplevel(self.win)
        dlg.title("🔗 E-Invoice Generator")
        dlg.geometry("680x620"); dlg.configure(bg=C_WHITE); dlg.grab_set()

        tk.Label(dlg, text="🔗  E-Invoice — IRP Format JSON",
                 font=("Segoe UI",11,"bold"), bg="#C05621", fg="white", pady=4).pack(fill="x")

        # IRN section
        irn_f = tk.Frame(dlg, bg="#FFF5EB", highlightthickness=1, highlightbackground="#C05621")
        irn_f.pack(fill="x", padx=12, pady=(8,4))

        tk.Label(irn_f, text="IRN (Invoice Reference Number)",
                 font=("Segoe UI",9,"bold"), bg="#FFF5EB", fg="#C05621").pack(anchor="w", padx=10, pady=(6,2))

        irn_var = tk.StringVar()
        # Check if IRN already saved in DB
        conn_irn = get_db()
        sale_row = conn_irn.execute("SELECT irn,ack_no,ack_date FROM sales WHERE bill_no=?",
                                    (d["bill_no"],)).fetchone()
        conn_irn.close()
        existing_irn = ""
        if sale_row:
            existing_irn = sale_row["irn"] or ""

        irn_var.set(existing_irn or irn_sim)

        irn_row = tk.Frame(irn_f, bg="#FFF5EB"); irn_row.pack(fill="x", padx=10, pady=(0,4))
        ttk.Entry(irn_row, textvariable=irn_var, width=58,
                  font=("Courier",9)).pack(side="left")

        ack_f = tk.Frame(irn_f, bg="#FFF5EB"); ack_f.pack(fill="x", padx=10, pady=(0,3))
        tk.Label(ack_f, text="ACK No:", font=("Segoe UI",9), bg="#FFF5EB", fg=C_GRAY).pack(side="left")
        ack_var = tk.StringVar(value=sale_row["ack_no"] if sale_row else "")
        ttk.Entry(ack_f, textvariable=ack_var, width=20).pack(side="left", padx=(4,14))
        tk.Label(ack_f, text="ACK Date:", font=("Segoe UI",9), bg="#FFF5EB", fg=C_GRAY).pack(side="left")
        ack_date_var = tk.StringVar(value=sale_row["ack_date"] if sale_row else "")
        make_date_entry(ack_f, ack_date_var, width=13).pack(side="left", padx=4)

        if existing_irn:
            tk.Label(irn_f, text=f"✅ IRN already saved for this bill",
                     font=("Segoe UI",9), bg="#FFF5EB", fg=C_GREEN, padx=10).pack(anchor="w", pady=(0,4))
        else:
            tk.Label(irn_f,
                     text="⚠️ Ye simulated IRN hai. Actual IRN IRP portal (einvoice1.gst.gov.in) se generate karo.\n"
                          "    JSON copy karo → IRP portal pe upload karo → actual IRN paste karo → Save karo.",
                     font=("Segoe UI",7), bg="#FFF5EB", fg=C_AMBER, padx=10,
                     justify="left", wraplength=620).pack(anchor="w", pady=(0,3))

        # JSON viewer
        tk.Label(dlg, text="E-Invoice JSON (IRP Format):",
                 font=("Segoe UI",9,"bold"), bg=C_WHITE, fg="#1A365D").pack(anchor="w", padx=12, pady=(4,2))
        txt_frame = tk.Frame(dlg, bg=C_WHITE); txt_frame.pack(fill="both", expand=True, padx=12, pady=(0,4))
        sb = ttk.Scrollbar(txt_frame); sb.pack(side="right", fill="y")
        txt = tk.Text(txt_frame, font=("Courier New",9), wrap="none",
                      yscrollcommand=sb.set, height=16)
        txt.pack(fill="both", expand=True)
        sb.config(command=txt.yview)
        txt.insert("1.0", json_str)
        txt.config(state="disabled")

        # Status
        status_lbl = tk.Label(dlg, text="", font=("Segoe UI",9), bg=C_WHITE)
        status_lbl.pack(anchor="w", padx=12)

        def save_irn():
            irn  = irn_var.get().strip()
            ack  = ack_var.get().strip()
            adate= ack_date_var.get().strip()
            if not irn:
                messagebox.showwarning("IRN?","IRN field empty hai!", parent=dlg); return
            conn2 = get_db()
            conn2.execute("UPDATE sales SET irn=?,ack_no=?,ack_date=? WHERE bill_no=?",
                          (irn, ack, adate, d["bill_no"]))
            conn2.commit(); conn2.close()
            status_lbl.config(text=f"✅ IRN saved: {irn[:30]}...", fg=C_GREEN)
            messagebox.showinfo("Saved!", "IRN, ACK No aur Date save ho gaye!", parent=dlg)

        def copy_json():
            dlg.clipboard_clear()
            dlg.clipboard_append(json_str)
            messagebox.showinfo("Copied!", "E-Invoice JSON clipboard mein copy ho gaya!\nAb IRP portal pe paste karo.", parent=dlg)

        def open_irp():
            import webbrowser
            webbrowser.open("https://einvoice1.gst.gov.in")

        def save_json():
            import tkinter.filedialog as fd
            path = fd.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON","*.json")],
                initialfile=f"einvoice_{d['bill_no'].replace('/','_')}.json")
            if path:
                with open(path,"w",encoding="utf-8") as f:
                    f.write(json_str)
                messagebox.showinfo("Saved!", f"JSON saved:\n{path}", parent=dlg)

        bf = tk.Frame(dlg, bg=C_WHITE); bf.pack(pady=(0,12))
        make_btn(bf,"📋 Copy JSON",     copy_json,  bg="#C05621" ).pack(side="left", padx=4)
        make_btn(bf,"💾 Save JSON",     save_json,  bg=C_AMBER   ).pack(side="left", padx=4)
        make_btn(bf,"🌐 Open IRP Portal",open_irp,  bg="#2B6CB0" ).pack(side="left", padx=4)
        make_btn(bf,"✅ Save IRN",      save_irn,   bg=C_GREEN   ).pack(side="left", padx=4)
        make_btn(bf,"Close",            dlg.destroy, bg=C_GRAY   ).pack(side="left", padx=4)

    # ──────────────────────────────────────────────────────────────────────────
    #  E-WAY BILL GENERATOR
    # ──────────────────────────────────────────────────────────────────────────
    def _gen_ewaybill(self):
        """E-Way Bill JSON generate karo aur EWB details save karo."""
        import json, datetime as _dt

        d     = self.data
        party = d["party"]
        items = d["items"]
        grand = round(d["grand"], 2)
        gst_type = d.get("gst_type","CGST+SGST")

        total_taxable = round(sum(it.get("taxable", it.get("rate",0)*it.get("qty",0)) for it in items), 2)
        total_gst     = round(sum(it.get("gst_amt",0) for it in items), 2)

        # Load existing EWB data from DB
        conn_e = get_db()
        sale_row = conn_e.execute(
            "SELECT ewb_no,ewb_date,vehicle_no,transport_mode,distance,transporter,irn "
            "FROM sales WHERE bill_no=?", (d["bill_no"],)).fetchone()
        conn_e.close()

        # ── Dialog ──────────────────────────────────────────────────────────
        dlg = tk.Toplevel(self.win)
        dlg.title("🚛 E-Way Bill")
        dlg.geometry("700x640"); dlg.configure(bg=C_WHITE); dlg.grab_set()

        tk.Label(dlg, text="🚛  E-Way Bill Generator",
                 font=("Segoe UI",11,"bold"), bg="#276749", fg="white", pady=4).pack(fill="x")

        # ── Transport details form ───────────────────────────────────────────
        frm_f = tk.Frame(dlg, bg="#F0FFF4", highlightthickness=1, highlightbackground="#276749")
        frm_f.pack(fill="x", padx=12, pady=(8,6))
        tk.Label(frm_f, text="Transport / Dispatch Details",
                 font=("Segoe UI",10,"bold"), bg="#F0FFF4", fg="#276749").pack(anchor="w", padx=10, pady=(6,4))

        r1 = tk.Frame(frm_f, bg="#F0FFF4"); r1.pack(fill="x", padx=10, pady=2)
        tk.Label(r1, text="Transport Mode:", font=("Segoe UI",9), bg="#F0FFF4", fg=C_GRAY).pack(side="left")
        v_mode = tk.StringVar(value=sale_row["transport_mode"] if sale_row and sale_row["transport_mode"] else "Road")
        ttk.Combobox(r1, textvariable=v_mode, width=10, state="readonly",
                     values=["Road","Rail","Air","Ship"]).pack(side="left", padx=(4,14))
        tk.Label(r1, text="Vehicle No:", font=("Segoe UI",9), bg="#F0FFF4", fg=C_GRAY).pack(side="left")
        v_veh = tk.StringVar(value=sale_row["vehicle_no"] if sale_row else "")
        ttk.Entry(r1, textvariable=v_veh, width=14).pack(side="left", padx=(4,14))
        tk.Label(r1, text="Distance (km):", font=("Segoe UI",9), bg="#F0FFF4", fg=C_GRAY).pack(side="left")
        v_dist = tk.StringVar(value=str(sale_row["distance"] or "") if sale_row else "")
        ttk.Entry(r1, textvariable=v_dist, width=8).pack(side="left", padx=4)

        r2 = tk.Frame(frm_f, bg="#F0FFF4"); r2.pack(fill="x", padx=10, pady=2)
        tk.Label(r2, text="Transporter Name:", font=("Segoe UI",9), bg="#F0FFF4", fg=C_GRAY).pack(side="left")
        v_trans = tk.StringVar(value=sale_row["transporter"] if sale_row else "")
        ttk.Entry(r2, textvariable=v_trans, width=22).pack(side="left", padx=(4,14))
        tk.Label(r2, text="EWB No (after portal):", font=("Segoe UI",9), bg="#F0FFF4", fg=C_GRAY).pack(side="left")
        v_ewb = tk.StringVar(value=sale_row["ewb_no"] if sale_row else "")
        ttk.Entry(r2, textvariable=v_ewb, width=16).pack(side="left", padx=4)

        r3 = tk.Frame(frm_f, bg="#F0FFF4"); r3.pack(fill="x", padx=10, pady=(2,8))
        tk.Label(r3, text="EWB Date:", font=("Segoe UI",9), bg="#F0FFF4", fg=C_GRAY).pack(side="left")
        v_ewbdate = tk.StringVar(value=sale_row["ewb_date"] if sale_row else "")
        make_date_entry(r3, v_ewbdate, width=13, bg="#F0FFF4").pack(side="left", padx=4)
        tk.Label(r3, text="Sub Type:", font=("Segoe UI",9), bg="#F0FFF4", fg=C_GRAY).pack(side="left", padx=(14,0))
        v_subtype = tk.StringVar(value="Supply")
        ttk.Combobox(r3, textvariable=v_subtype, width=14, state="readonly",
                     values=["Supply","Import","Export","Job Work","SKD/CKD","Recipient Not Known",
                             "For Own Use","Others"]).pack(side="left", padx=4)

        # ── JSON preview ─────────────────────────────────────────────────────
        tk.Label(dlg, text="E-Way Bill JSON (NIC Portal Format):",
                 font=("Segoe UI",9,"bold"), bg=C_WHITE, fg="#1A365D").pack(anchor="w", padx=12, pady=(4,2))

        txt_frame = tk.Frame(dlg); txt_frame.pack(fill="both", expand=True, padx=12, pady=(0,4))
        sb2 = ttk.Scrollbar(txt_frame); sb2.pack(side="right", fill="y")
        txt = tk.Text(txt_frame, font=("Courier New",9), wrap="none",
                      yscrollcommand=sb2.set, height=14)
        txt.pack(fill="both", expand=True)
        sb2.config(command=txt.yview)

        def build_ewb_json():
            item_list = []
            for i,it in enumerate(items,1):
                gp  = it.get("gst_percent", it.get("gst",18))
                txbl= round(it.get("taxable", it.get("rate",0)*it.get("qty",0)),2)
                ga  = round(it.get("gst_amt",0),2)
                item_list.append({
                    "productName":    it.get("product",""),
                    "hsnCode":        it.get("hsn",""),
                    "productDesc":    it.get("product",""),
                    "quantity":       it.get("qty",1),
                    "qtyUnit":        "NOS",
                    "taxableAmount":  txbl,
                    "sgstRate":       gp/2 if gst_type!="IGST" else 0,
                    "cgstRate":       gp/2 if gst_type!="IGST" else 0,
                    "igstRate":       gp   if gst_type=="IGST"  else 0,
                    "cessRate":       0,
                })
            _sew = get_shop()
            return {
                "supplyType":         "O",  # Outward
                "subSupplyType":      v_subtype.get(),
                "docType":            "INV",
                "docNo":              d["bill_no"],
                "docDate":            d["date"].replace("-","/") if "-" in d["date"] else d["date"],
                "fromGstin":          _sew['gstin'] or "99AAAAA0000A1Z5",
                "fromTrdName":        _sew['name'],
                "fromAddr1":          _sew['address'] or "",
                "fromPlace":          _sew['city'] or "",
                "fromPincode":        999999,
                "fromStateCode":      9,
                "toGstin":            party.get("gstin","") or "URP",
                "toTrdName":          party.get("name",""),
                "toAddr1":            party.get("address","") or "",
                "toPlace":            party.get("state","") or "Uttar Pradesh",
                "toPincode":          999999,
                "toStateCode":        9,
                "totalValue":         total_taxable,
                "cgstValue":          round(total_gst/2,2) if gst_type!="IGST" else 0,
                "sgstValue":          round(total_gst/2,2) if gst_type!="IGST" else 0,
                "igstValue":          round(total_gst,2)   if gst_type=="IGST"  else 0,
                "cessValue":          0,
                "totInvValue":        grand,
                "transDistance":      v_dist.get() or "0",
                "transporterName":    v_trans.get(),
                "transportMode":      {"Road":"1","Rail":"2","Air":"3","Ship":"4"}.get(v_mode.get(),"1"),
                "vehicleNo":          v_veh.get().upper(),
                "vehicleType":        "R",
                "ItemList":           item_list,
            }

        def refresh_json(*args):
            ewb = build_ewb_json()
            txt.config(state="normal")
            txt.delete("1.0","end")
            txt.insert("1.0", json.dumps(ewb, indent=2, ensure_ascii=False))
            txt.config(state="disabled")

        for v in [v_mode,v_veh,v_dist,v_trans,v_subtype]:
            v.trace_add("write", refresh_json)
        refresh_json()

        status_lbl = tk.Label(dlg, text="", font=("Segoe UI",9), bg=C_WHITE)
        status_lbl.pack(anchor="w", padx=12)

        def save_ewb():
            ewb_no = v_ewb.get().strip()
            veh    = v_veh.get().strip().upper()
            dist   = v_dist.get().strip()
            try: dist_int = int(dist) if dist else 0
            except: dist_int = 0
            conn2 = get_db()
            conn2.execute(
                "UPDATE sales SET ewb_no=?,ewb_date=?,vehicle_no=?,transport_mode=?,distance=?,transporter=? WHERE bill_no=?",
                (ewb_no, v_ewbdate.get(), veh, v_mode.get(), dist_int, v_trans.get(), d["bill_no"]))
            conn2.commit(); conn2.close()
            status_lbl.config(text="✅ E-Way Bill details saved!", fg=C_GREEN)
            messagebox.showinfo("Saved!","E-Way Bill details save ho gaye!", parent=dlg)

        def copy_json():
            ewb = build_ewb_json()
            dlg.clipboard_clear()
            dlg.clipboard_append(json.dumps(ewb, indent=2, ensure_ascii=False))
            messagebox.showinfo("Copied!","E-Way Bill JSON copy ho gaya!\nAb NIC portal pe paste karo.", parent=dlg)

        def save_json_file():
            import tkinter.filedialog as fd
            ewb = build_ewb_json()
            path = fd.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON","*.json")],
                initialfile=f"ewaybill_{d['bill_no'].replace('/','_')}.json")
            if path:
                with open(path,"w",encoding="utf-8") as f:
                    json.dump(ewb, f, indent=2, ensure_ascii=False)
                messagebox.showinfo("Saved!", f"JSON saved:\n{path}", parent=dlg)

        def open_nic():
            import webbrowser
            webbrowser.open("https://ewaybillgst.gov.in")

        bf = tk.Frame(dlg, bg=C_WHITE); bf.pack(pady=(0,12))
        make_btn(bf,"🔄 Refresh JSON", refresh_json, bg="#276749").pack(side="left", padx=4)
        make_btn(bf,"📋 Copy JSON",    copy_json,    bg=C_AMBER  ).pack(side="left", padx=4)
        make_btn(bf,"💾 Save JSON",    save_json_file,bg="#2B6CB0").pack(side="left", padx=4)
        make_btn(bf,"🌐 NIC Portal",   open_nic,     bg="#6B46C1" ).pack(side="left", padx=4)
        make_btn(bf,"✅ Save EWB",     save_ewb,     bg=C_GREEN  ).pack(side="left", padx=4)
        make_btn(bf,"Close",           dlg.destroy,  bg=C_GRAY   ).pack(side="left", padx=4)


if __name__ == "__main__":
    import sys
    init_db()

    # ── LICENSE CHECK ─────────────────────────────────────────────────────
    status, days_left, install_date, customer = _get_license_info()

    if status == 'new':
        # Pehli baar — activate window dikhao
        _show_first_install_window()
        # Re-check — agar activate nahi kiya to band karo
        status, _, _, _ = _get_license_info()
        if status != 'ok':
            sys.exit(0)

    elif status == 'expired':
        # Expire ho gayi — renew screen dikhao
        LicenseExpiredWin(days_left, customer)
        # Re-check
        status, _, _, _ = _get_license_info()
        if status != 'ok':
            sys.exit(0)

    elif status != 'ok':
        # Kuch aur galat — band karo
        sys.exit(0)

    # ── LICENSE OK — Normal flow ──────────────────────────────────────────
    LoginWin()
